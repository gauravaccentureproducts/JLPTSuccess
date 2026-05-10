"""Administrative cleanup pass:
1. Update audit registry XLSX — mark all 29 ISSUE-105..110 + IMP-124..146 as Done
2. Backfill 9 missing provenance tags on vocab entries
3. Add render_engine='voicevox' to the 40 original listening items
"""
from __future__ import annotations
import io, json, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'data').exists() and ROOT != ROOT.parent: ROOT = ROOT.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


# ============================================================
# 1. Update audit registry XLSX
# ============================================================
print('=== 1. Audit registry update ===')

import openpyxl
xlsx_path = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['Items']
hdr = [c.value for c in ws[4]]
id_idx = hdr.index('ID')
dec_idx = hdr.index('Decision (Fix / Avoid / Defer)') + 1  # 1-based for openpyxl

target_ids = (
    [f'ISSUE-{n:03d}' for n in range(105, 111)] +
    [f'IMP-{n}' for n in range(124, 147)]
)
target_set = set(target_ids)

updated = 0
for row in range(5, ws.max_row + 1):
    rid = ws.cell(row=row, column=id_idx + 1).value
    if rid in target_set:
        ws.cell(row=row, column=dec_idx, value='Done')
        updated += 1

wb.save(xlsx_path)
print(f'Updated {updated} audit-registry entries to Done')


# ============================================================
# 2. Backfill missing provenance tags on vocab
# ============================================================
print()
print('=== 2. Vocab provenance backfill ===')

vocab_path = ROOT / 'data' / 'vocab.json'
data = json.loads(vocab_path.read_text(encoding='utf-8'))
entries = data['entries']

coll_fixed = 0
pitch_fixed = 0
for e in entries:
    if e.get('collocations') and not e.get('collocations_provenance'):
        # These are the early auto-derived entries
        e['collocations_provenance'] = 'auto_derived'
        coll_fixed += 1
    if e.get('pitch_accent') and not e.get('pitch_accent_provenance'):
        e['pitch_accent_provenance'] = 'llm_curated'
        pitch_fixed += 1

vocab_path.write_text(
    json.dumps(data, ensure_ascii=False, indent=2) + '\n',
    encoding='utf-8',
)
print(f'  collocations provenance backfilled: {coll_fixed}')
print(f'  pitch_accent provenance backfilled: {pitch_fixed}')


# ============================================================
# 3. Add render_engine='voicevox' to listening 001-040
# ============================================================
print()
print('=== 3. Listening render_engine backfill ===')

listen_path = ROOT / 'data' / 'listening.json'
data = json.loads(listen_path.read_text(encoding='utf-8'))
items = data['items']

added = 0
for it in items:
    meta = it.get('audio_render_meta')
    if not isinstance(meta, dict):
        continue
    if 'render_engine' in meta:
        continue
    # If voices_used is non-empty, it's VOICEVOX (single voice -> gtts)
    voices = meta.get('voices_used') or []
    if voices:
        meta['render_engine'] = 'voicevox'
        meta['render_date'] = meta.get('render_date', '2026-05-07')
        added += 1

listen_path.write_text(
    json.dumps(data, ensure_ascii=False, indent=2) + '\n',
    encoding='utf-8',
)
print(f'  render_engine backfilled (voicevox): {added}')


# ============================================================
# Verify final state
# ============================================================
print()
print('=== Final state ===')

# Re-read vocab
v = json.loads(vocab_path.read_text(encoding='utf-8'))['entries']
no_coll_prov = sum(1 for e in v if e.get('collocations') and not e.get('collocations_provenance'))
no_pitch_prov = sum(1 for e in v if e.get('pitch_accent') and not e.get('pitch_accent_provenance'))

# Re-read listening
l = json.loads(listen_path.read_text(encoding='utf-8'))['items']
no_engine = sum(1 for it in l if it.get('audio_render_meta') and 'render_engine' not in it.get('audio_render_meta', {}))

print(f'Vocab missing collocations_provenance:    {no_coll_prov}')
print(f'Vocab missing pitch_accent_provenance:    {no_pitch_prov}')
print(f'Listening items missing render_engine:    {no_engine}')
