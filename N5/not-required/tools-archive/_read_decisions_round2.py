"""Read xlsx - list rows where Decision = Fix and that are NOT yet Done."""
from __future__ import annotations
import io, sys
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

wb = load_workbook(XLSX, data_only=True)

# --- Audit findings sheet ---
ws = wb['Audit findings']
header = next(r for r in range(1, 8) if ws.cell(row=r, column=1).value == 'ID')
dec_col = next(c for c in range(1, ws.max_column + 1)
               if 'Decision' in (ws.cell(row=header, column=c).value or ''))

print(f'Sheets: {wb.sheetnames}')
print(f'Audit-findings header row: {header}, decision col: {dec_col}\n')

print('=== Audit findings - current decisions ===')
fix_rows = []
for r in range(header + 1, ws.max_row + 1):
    rid = ws.cell(row=r, column=1).value
    if not rid: continue
    title = ws.cell(row=r, column=9).value or ''
    dec = (ws.cell(row=r, column=dec_col).value or '').strip()
    print(f'  {rid:12} | {dec or "<empty>":18} | {title[:55]}')
    if dec.lower() == 'fix':
        fix_rows.append({
            'id': rid,
            'severity': ws.cell(row=r, column=3).value,
            'priority': ws.cell(row=r, column=4).value,
            'category': ws.cell(row=r, column=7).value,
            'location': ws.cell(row=r, column=8).value,
            'title': title,
            'current': ws.cell(row=r, column=10).value,
            'why': ws.cell(row=r, column=11).value,
            'direction': ws.cell(row=r, column=12).value,
            'deps': ws.cell(row=r, column=13).value,
        })

# --- Open questions sheet ---
ws2 = wb['Open questions']
header2 = next(r for r in range(1, 6) if ws2.cell(row=r, column=1).value == 'ID')
dec_col2 = next(c for c in range(1, ws2.max_column + 1)
                if 'Decision' in (ws2.cell(row=header2, column=c).value or ''))
print(f'\n=== Open questions - current decisions ===')
q_fix = []
for r in range(header2 + 1, ws2.max_row + 1):
    qid = ws2.cell(row=r, column=1).value
    if not qid: continue
    topic = ws2.cell(row=r, column=2).value or ''
    dec = (ws2.cell(row=r, column=dec_col2).value or '').strip()
    print(f'  {qid:4} | {dec or "<empty>":18} | {topic[:55]}')
    if dec.lower() == 'fix':
        q_fix.append({
            'id': qid,
            'topic': topic,
            'context': ws2.cell(row=r, column=3).value,
            'decision_needed': ws2.cell(row=r, column=4).value,
        })

print(f'\n=== Total Fix rows: {len(fix_rows)} (audit) + {len(q_fix)} (questions) ===')
for f in fix_rows:
    print(f'\n--- {f["id"]} [{f["severity"]} / {f["priority"]}] ---')
    print(f'  Category: {f["category"]}')
    print(f'  Location: {f["location"]}')
    print(f'  Title:    {f["title"]}')
    print(f'  Direction: {f["direction"]}')
    print(f'  Dependencies: {f["deps"]}')
for q in q_fix:
    print(f'\n--- {q["id"]}: {q["topic"]} ---')
    print(f'  Context: {q["context"][:200] if q["context"] else ""}')
    print(f'  Decision needed: {q["decision_needed"]}')
