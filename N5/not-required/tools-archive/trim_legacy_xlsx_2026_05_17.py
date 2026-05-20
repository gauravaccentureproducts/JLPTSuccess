"""Drop redundant 'User Reported Bugs' sheet from N5/feedback/n5-audit-2026-05-04.xlsx.

The legacy xlsx has 3 sheets:
  - Items (316 rows: ISSUE-NNN + IMP-NNN registry) — UNIQUE, KEEP
  - Questions (47 rows) — UNIQUE, KEEP
  - User Reported Bugs (24 rows, BUG-001..024) — REDUNDANT (all 24
    superseded by N5/specifications/test-scenarios-by-specialist-
    perspective.xlsx which has BUG-001..053)

Idempotent: re-running on an already-trimmed file is a no-op (the
sheet is just absent).
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "feedback" / "n5-audit-2026-05-04.xlsx"


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)
    SHEET = "User Reported Bugs"
    if SHEET not in wb.sheetnames:
        print(f"No-op: '{SHEET}' sheet already absent from {XLSX.name}")
        return 0
    # Final confirmation: every BUG ID in this sheet exists in the current xlsx
    legacy_sh = wb[SHEET]
    legacy_bugs = []
    for r in range(4, legacy_sh.max_row + 1):
        v = legacy_sh.cell(row=r, column=1).value
        if not v:
            continue
        if isinstance(v, str) and v.startswith("="):
            legacy_bugs.append(f"BUG-{r-3:03d}")
        else:
            legacy_bugs.append(str(v))

    # Cross-check against current xlsx
    CURRENT = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
    cwb = openpyxl.load_workbook(CURRENT)
    current_sh = cwb["User Reported Bugs"]
    current_bugs = set()
    for r in range(4, current_sh.max_row + 1):
        v = current_sh.cell(row=r, column=1).value
        if not v:
            continue
        if isinstance(v, str) and v.startswith("="):
            current_bugs.add(f"BUG-{r-3:03d}")
        else:
            current_bugs.add(str(v))

    missing = [b for b in legacy_bugs if b not in current_bugs]
    if missing:
        print(f"ABORT: {len(missing)} legacy BUG IDs not found in current xlsx — "
              f"would lose data. Missing: {missing[:10]}")
        return 1
    print(f"Verified: all {len(legacy_bugs)} legacy bugs present in current xlsx")

    # Safe to drop the sheet
    del wb[SHEET]
    wb.save(XLSX)
    print(f"Dropped '{SHEET}' sheet from {XLSX.name}")
    print(f"Remaining sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
