"""IMP-100: flip 'Avoid' -> 'Done' with the 2026-05-09 resolution.

Root cause analysis showed the '12 missing patterns' was a false
positive caused by check_coverage.py doing naive count comparison.
Real coverage is 100%: 186 MD pattern bullets all map to one of
the 178 JSON entries (some MD bullets share a JSON entry because
the JSON consolidates particle-level abstractions).

Updated check_coverage.py with semantic matching now exits 0.
"""
from __future__ import annotations
import io
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'feedback').exists() and ROOT != ROOT.parent:
    ROOT = ROOT.parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'
wb = load_workbook(XLSX)
ws = wb['Items']

rows = list(ws.iter_rows(values_only=True))
hdr_row_idx = next(ri for ri, r in enumerate(rows[:10])
                   if 'ID' in (r or []) and any('decision' in str(c).lower() for c in (r or [])))
hdr = list(rows[hdr_row_idx])
id_idx = hdr.index('ID')
title_idx = hdr.index('Title')
state_idx = next(i for i, h in enumerate(hdr) if 'current state' in str(h).lower())
dec_idx = next(i for i, h in enumerate(hdr) if 'decision' in str(h).lower())

target_row = None
for ri in range(hdr_row_idx + 2, ws.max_row + 1):
    if ws.cell(row=ri, column=id_idx + 1).value == 'IMP-100':
        target_row = ri
        break

if target_row is None:
    print('ERROR: IMP-100 not found')
    sys.exit(1)

new_state = (
    'RESOLVED 2026-05-09. The "12 missing patterns" was a false positive '
    'from check_coverage.py doing naive count comparison (190 MD bullets '
    'vs 178 JSON entries). Root cause: MD bullets and JSON entries '
    'operate at different abstraction levels. MD has sentence-templates '
    '(～は～です), sub-use enumerations under consolidated parents (the '
    '"～の (consolidated)" bullet has 3 sub-uses listed), and register-'
    'metadata annotations - none of which are separate patterns. JSON '
    'has particle-level entries that subsume the MD templates. '
    'check_coverage.py rewritten 2026-05-09 with semantic matching: '
    '186 effective MD pattern bullets all map to one of the 178 JSON '
    'entries. Tool exits 0. No grammar pattern is genuinely missing.'
)

ws.cell(row=target_row, column=state_idx + 1, value=new_state)
ws.cell(row=target_row, column=dec_idx + 1, value='Done')
title = ws.cell(row=target_row, column=title_idx + 1).value
print(f'IMP-100 ({title}):')
print(f'  Decision: Avoid -> Done')
print(f'  Current state updated to terminal-state resolution.')

wb.save(XLSX)
