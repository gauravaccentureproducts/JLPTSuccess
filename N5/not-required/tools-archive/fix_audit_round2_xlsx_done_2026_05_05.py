"""IMP-028 follow-through: stamp 'Done' on the 13 round-2 Fix items in
feedback/n5-audit-2026-05-04.xlsx.

The audit tracker is the source of truth for outstanding work. This run
walks each sheet, locates the row whose Issue ID matches one of the 13
shipped fixes, and sets Decision = 'Done' (mirroring the v1.12.28 close-
out pattern). Idempotent: rows already marked Done are skipped.

Items closed this batch (v1.12.29):
  ISSUE-008, ISSUE-009, ISSUE-010, ISSUE-011, ISSUE-012,
  IMP-018, IMP-020, IMP-021, IMP-022, IMP-023, IMP-025, IMP-028, IMP-029
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

DONE_IDS = {
    'ISSUE-008', 'ISSUE-009', 'ISSUE-010', 'ISSUE-011', 'ISSUE-012',
    'IMP-018', 'IMP-020', 'IMP-021', 'IMP-022', 'IMP-023',
    'IMP-025', 'IMP-028', 'IMP-029',
    # Open Questions sheet: Q1 (localStorage namespace migration target)
    # is closed by IMP-023's `migrate(oldNS, newNS)` helper. Q2-Q7 stay
    # open - they are product decisions, not implementation items.
    'Q1',
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed. Run `pip install openpyxl`.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)
    closed = []
    skipped_already_done = []
    not_found = set(DONE_IDS)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Header row may be on row 1 OR offset by a title + blank pair (row 4
        # on "Audit findings"). Walk the first 6 rows looking for a line
        # whose cells include the literal "ID" and a "Decision..." entry.
        header_row_idx = None
        for r in range(1, min(7, ws.max_row + 1)):
            cells = [str(c.value or '').strip() for c in ws[r]]
            if any(c.lower() == 'id' for c in cells) and any(
                c.lower().startswith('decision') for c in cells
            ):
                header_row_idx = r
                break
        if header_row_idx is None:
            continue

        header = [c.value for c in ws[header_row_idx]]
        try:
            id_col = next(i for i, v in enumerate(header, start=1)
                          if v and str(v).strip().lower() in ('id', 'issue id', 'item id'))
            # Prefer "Decision (Fix / Avoid / Defer)" over "Decision needed";
            # both share the "decision" prefix on the Open questions sheet.
            dec_candidates = [i for i, v in enumerate(header, start=1)
                              if v and str(v).strip().lower().startswith('decision (')]
            if not dec_candidates:
                dec_candidates = [i for i, v in enumerate(header, start=1)
                                  if v and str(v).strip().lower().startswith('decision')]
            dec_col = dec_candidates[0]
        except (StopIteration, IndexError):
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            id_cell = row[id_col - 1]
            dec_cell = row[dec_col - 1]
            id_val = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
            if id_val not in DONE_IDS:
                continue
            cur = (dec_cell.value or '').strip() if isinstance(dec_cell.value, str) else ''
            not_found.discard(id_val)
            if cur == 'Done':
                skipped_already_done.append((sheet, id_val))
                continue
            dec_cell.value = 'Done'
            closed.append((sheet, id_val, cur or '<blank>'))

    wb.save(XLSX)

    print(f'Closed {len(closed)} item(s) (Decision -> Done):')
    for sh, iid, prev in closed:
        print(f'  [{sh}]  {iid}  ({prev} -> Done)')
    if skipped_already_done:
        print(f'\nAlready Done ({len(skipped_already_done)}):')
        for sh, iid in skipped_already_done:
            print(f'  [{sh}]  {iid}')
    if not_found:
        print(f'\nWARNING: not found in any sheet ({len(not_found)}):')
        for iid in sorted(not_found):
            print(f'  {iid}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
