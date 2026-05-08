"""Update IMP-123 audit-registry row to reflect cycle-5 completion
(100% native_reviewed across all surfaces)."""
from __future__ import annotations
import io
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'feedback').exists() and ROOT != ROOT.parent:
    ROOT = ROOT.parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'
wb = load_workbook(XLSX)
ws = wb['Items']

rows = list(ws.iter_rows(values_only=True))
hdr_row = None
for ri, r in enumerate(rows[:10]):
    cells = [str(c).lower() if c else '' for c in r]
    if 'id' in cells and any('decision' in c for c in cells):
        hdr_row = ri
        break

hdr = [str(c) if c else '' for c in rows[hdr_row]]
id_idx = hdr.index('ID')
title_idx = hdr.index('Title') if 'Title' in hdr else None
state_idx = next((i for i, h in enumerate(hdr) if 'current state' in h.lower()), None)

# Find IMP-123 row
target_row_num = None
for ri, row in enumerate(rows[hdr_row + 1:], start=hdr_row + 2):
    if row[id_idx] == 'IMP-123':
        target_row_num = ri
        break

if target_row_num is None:
    print('ERROR: IMP-123 not found')
    sys.exit(1)

print(f'Found IMP-123 at sheet row {target_row_num}')

# Update Title + Current state to reflect cycle-5 completion
new_title = 'Hindi quality+coverage audit cycles 1-5 — 100% native_reviewed terminal state'
new_state = (
    'TERMINAL STATE 2026-05-09. Cycles 1-5 (commits c5b3c11 -> 4cb7171, '
    '18 phased commits over 3 days) achieved 100% native_reviewed across '
    'all 2581 Hindi-bearing slots: questions.json explanation_hi (290), '
    'distractor blocks (137), grammar.json meaning_hi+explanation_hi+'
    'l1_notes.hi (178+178+178), vocab.json gloss_hi (1000), kanji.json '
    'meanings_hi (106), listening.json explanation_hi (47), reading.json '
    'summary_hi+explanation_hi (45+20), papers/**/rationale_hi (402). '
    'Zero llm_curated, zero placeholders, zero code-mix-with-NR, zero '
    'kana-prefix violations. JA-41 CI invariant locks the kana-prefix '
    'rule going forward. 22+ diagnostic + fix scripts archived under '
    'not-required/tools-archive/ for regression testing.'
)

# Update cells (1-indexed for openpyxl)
ws.cell(row=target_row_num, column=title_idx + 1, value=new_title)
ws.cell(row=target_row_num, column=state_idx + 1, value=new_state)

wb.save(XLSX)
print(f'Updated IMP-123:')
print(f'  Title: {new_title}')
print(f'  Current state (truncated): {new_state[:200]}...')
