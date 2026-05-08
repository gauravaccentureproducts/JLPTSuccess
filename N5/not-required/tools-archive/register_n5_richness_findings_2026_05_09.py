"""Register N5 richness audit findings into the audit registry.

Picks up next-free ISSUE-NNN and IMP-NNN IDs from the xlsx and
appends one row per finding.
"""
from __future__ import annotations
import io
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'feedback').exists() and ROOT != ROOT.parent:
    ROOT = ROOT.parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'
wb = load_workbook(XLSX)
ws = wb['Items']

# Find header row + columns
rows = list(ws.iter_rows(values_only=True))
hdr_row_idx = None
for ri, r in enumerate(rows[:10]):
    cells = [str(c).lower() if c else '' for c in r]
    if 'id' in cells and any('decision' in c for c in cells):
        hdr_row_idx = ri
        break
hdr = list(rows[hdr_row_idx])
print(f'Header row: {hdr_row_idx + 1}')

# Find next free IDs
existing_issue = set()
existing_imp = set()
for r in rows[hdr_row_idx + 1:]:
    iid = r[0]
    if isinstance(iid, str):
        if iid.startswith('ISSUE-'):
            try: existing_issue.add(int(iid.split('-')[1]))
            except: pass
        elif iid.startswith('IMP-'):
            try: existing_imp.add(int(iid.split('-')[1]))
            except: pass

next_issue = (max(existing_issue) + 1) if existing_issue else 1
next_imp = (max(existing_imp) + 1) if existing_imp else 1
print(f'Next free ISSUE: {next_issue} (highest existing: {max(existing_issue) if existing_issue else 0})')
print(f'Next free IMP:   {next_imp} (highest existing: {max(existing_imp) if existing_imp else 0})')


# Build new rows. Column order matches header (16 cols).
def make_row(iid, type_, severity, priority, impact, effort,
             category, location, title, current_state, why_matters,
             direction, deps, decision, description, permission):
    return [iid, type_, severity, priority, impact, effort,
            category, location, title, current_state, why_matters,
            direction, deps, decision, description, permission]


# ============================================================================
# ISSUES (existing problems / correctness gaps)
# ============================================================================

issues = []
i = next_issue

# ISSUE: confusable kanji clusters not all linked (4/8)
issues.append(make_row(
    f'ISSUE-{i:03d}', 'Issue', 'MAJOR', 'P3', 'MEDIUM', 'LOW',
    'Kanji depth - linking',
    'data/kanji.json + js/kanji.js',
    'Confusable kanji clusters - 4/8 incomplete cross-link coverage',
    'Cluster all-linked check: 4/8 clusters have all members listing each other in lookalikes/confusable_with. Missing: 大/犬/太, 木/本/末/未, 千/干/王/玉, 千/午 partial.',
    'WaniKani and Tofugu link confusable clusters bidirectionally; learners conflate them without explicit cross-links.',
    'Walk all 8 cluster definitions; for each member, populate `lookalikes` array with the other members.',
    '', 'Fix',
    'Audit found 29/106 kanji with any lookalikes field; cluster-completeness is the metric. Cluster size: 大/犬/太 (3), 木/本/末/未 (4), 人/入/八 (3), 日/目/白 (3), 千/干/王/玉 (4), 上/止/正 (3), 古/占 (2), 千/午 (2).',
    'No permission required'
))
i += 1

# ISSUE: kanji not linked from reading passages (Density-4 = 0%)
issues.append(make_row(
    f'ISSUE-{i:03d}', 'Issue', 'MAJOR', 'P3', 'HIGH', 'MEDIUM',
    'Linking density - kanji ↔ reading',
    'data/reading.json + js/reading.js',
    'Density-4 zero: no kanji entries reference which reading passages contain them',
    'Density-4 metric: 0/106 kanji entries carry a `appears_in_passages` array linking back to reading.json passages where they occur. Reverse-lookup tooling exists at scan time but no surfaced cross-link.',
    'Jisho/WaniKani learners can pivot from a kanji to passages featuring it; this app cannot. Killer feature for niche N3 (all-in-one).',
    'Build offline cross-reference: scan reading passages for each N5 kanji; populate `kanji_appearances` field on kanji entries.',
    '', 'Fix',
    '',
    'No permission required'
))
i += 1

# ISSUE: listening items have no clickable vocab in transcripts (Density-6)
issues.append(make_row(
    f'ISSUE-{i:03d}', 'Issue', 'MAJOR', 'P3', 'HIGH', 'MEDIUM',
    'Linking density - listening transcript',
    'data/listening.json + js/listening-transcript.js',
    'Density-6 zero: listening transcript words not clickable to vocab entries (0/47)',
    '0/47 listening items have a `vocab_glossary` field tagging which words in the transcript link to vocab.json entries. Hover-popovers on transcript words would be the JapanesePod101-paid-tier feature shipped free.',
    'JapanesePod101 hides this behind paid tier; a free version differentiates and addresses learner pain (looking up unknown words mid-listening).',
    'Author `vocab_glossary` per item; render as click/hover popover on listening detail page.',
    '', 'Fix',
    '',
    'No permission required'
))
i += 1

# ISSUE: paper questions don't link back to grammar/vocab/kanji entries (Density-7)
issues.append(make_row(
    f'ISSUE-{i:03d}', 'Issue', 'MAJOR', 'P3', 'HIGH', 'LOW',
    'Linking density - paper review',
    'data/papers/**/*.json + js/papers.js',
    'Density-7 zero: paper questions have no link-back to source content on review (0/402)',
    'After answering a paper question, learners cannot pivot to the related grammar pattern / vocab / kanji entry. 0/402 paper questions have grammarPatternId / vocab_id / kanji_id field populated.',
    'Bunpro and Renshuu link review-result UI to underlying content; this app stops at the score.',
    'Backfill grammarPatternId on each paper question by matching question stems to grammar.json patterns.',
    '', 'Fix',
    '',
    'No permission required'
))
i += 1

# ISSUE: vocab transitivity-pair mostly absent (978/1000 missing)
issues.append(make_row(
    f'ISSUE-{i:03d}', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'LOW',
    'Vocab depth - structural pairs',
    'data/vocab.json',
    'Transitivity-pair links missing on 978 of 1000 verbs/nouns',
    'Only 22/1000 entries have a transitivity_pair / pair_id field. Mandatory N5 pairs (开ける/開く, 閉める/閉まる, 入れる/入る, etc., 12+ canonical pairs) need full bidirectional linking.',
    'Genki I explicitly drills these pairs; learners need adjacency. None of Bunpro/Jisho ship structural pair-links by default.',
    'Author `pair_id` on the 12+ canonical N5 transitivity pairs; surface "see also" link in vocab page.',
    '', 'Fix',
    '',
    'No permission required'
))
i += 1

# ISSUE: vocab verb-class flags incomplete (132/1000 = 13%)
issues.append(make_row(
    f'ISSUE-{i:03d}', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'LOW',
    'Vocab depth - verb-class metadata',
    'data/vocab.json',
    'Verb-class flag (godan/ichidan/irregular) on only 13% of entries',
    '132/1000 entries carry verb_class or group_1_exception. Coverage should be 100% on verbs (~340 verbs in N5 corpus). Group-1-exception verbs (入る/帰る/走る/知る/切る/要る) per X-6.6 must each carry the explicit flag.',
    'Beginners conflate Group-1-exception verbs as Group-2; correctness during conjugation suffers.',
    'For each entry where pos contains "verb", populate verb_class from {godan, ichidan, irregular}; explicit Group-1-exception flag where applicable.',
    '', 'Fix',
    '',
    'No permission required'
))
i += 1

issue_ids = [r[0] for r in issues]


# ============================================================================
# IMPROVEMENTS (richness ideas)
# ============================================================================

imps = []
m = next_imp

# IMP P1: Grammar examples ≥ 7 (Bunpro-beat)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'HIGH',
    'Grammar depth - example count',
    'data/grammar.json',
    'Grammar pattern examples = 7 each (currently 0/178 at target; 26/178 ≥5)',
    '0/178 patterns ship 7 examples (the bar to BEAT Bunpro at 5). 26/178 reach 5. Diversity rule: each example must use a different attachment surface (verb-stem / verb-te / verb-nai / na-adj / i-adj / noun / clause) AND a different topic cluster.',
    'Bunpro ships 5 examples per pattern with native review; beating to 7 with diverse coverage is the single largest grammar-depth lever.',
    'Author 2-4 additional examples per pattern (varying attachment surface + topic), all using only N5 vocab + kanji per JA-13.',
    '', 'Fix',
    'Audit 2026-05-09: 0/178 at ≥7; 26/178 at ≥5; floor 178/178 at ≥3. Bunpro reference: 5 examples / pattern.',
    'No permission required'
))
m += 1

# IMP P1: Kanji 3-mnemonic structure (WaniKani-beat)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'HIGH',
    'Kanji depth - mnemonics',
    'data/kanji.json',
    'WaniKani 3-mnemonic structure: visual + reading mnemonics absent (0/106 each)',
    '106/106 have radical decomposition + radical_story. 0/106 have a visual mnemonic (story tying glyph shape to meaning) or reading mnemonic (pneumonic for the dominant on-yomi). WaniKani ships all three.',
    'Largest UX gap vs WaniKani for the kanji surface. Mnemonics are the differentiator that makes WaniKani worth $9/month.',
    'Author per-kanji visual mnemonic and reading-hint mnemonic. Source: LLM-curated → native-review pass; or manual authoring against WaniKani public-domain references.',
    '', 'Fix',
    'Audit 2026-05-09: WaniKani ships radical-story + visual + reading mnemonic per kanji. This app: radical-story 100%, visual 0%, reading 0%.',
    'No permission required'
))
m += 1

# IMP P1: Authentic-content layer (no incumbent does this)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'HIGH',
    'Authentic content - cross-surface',
    'data/grammar.json + data/vocab.json + data/kanji.json',
    'Authentic-content layer 0% across every surface (largest leverage gap)',
    '0% of grammar / vocab / kanji / reading / listening entries reference authentic Japanese (anime / J-drama / signage / news headline / song lyric). NO incumbent does this systematically at N5.',
    'Single largest "demonstrably richer than competitors" lever. Bunpro/WaniKani/Tofugu all use invented examples. Real Japanese exposure is what makes content stick.',
    'Per surface: ≥20% of entries get an authentic_media reference (anime quote with episode + timestamp, NHK Easy headline, J-pop lyric, restaurant menu, signage photo). Source via OpenSubtitles / kitsunekko / NHK Easy / hand-curation.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P2: Vocab pitch accent
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'MEDIUM',
    'Vocab depth - pitch accent',
    'data/vocab.json',
    'Tokyo pitch accent: 25% coverage (249/1000), target ≥80%',
    '249/1000 entries have pitch_accent populated (per JCE-1 cycle). 751 vocab entries missing. Migaku/Marumori ship near-100%.',
    'Pitch accent is the single highest-signal "serious learner" trust marker. OJAD has the data but is not learner-facing here; embedding it via popover is the lever.',
    'Author pitch_accent on the remaining 751 entries from OJAD/JMdict-pitch-accent dataset (open data); render as inline diacritic or kanji-popover.',
    '', 'Fix',
    'Audit 2026-05-09: 249/1000 with pitch (25%). Target 80%+.',
    'No permission required'
))
m += 1

# IMP P2: Vocab collocations
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'MEDIUM',
    'Vocab depth - collocations',
    'data/vocab.json',
    'Collocations ≥ 5: 0/1000 (target on high-frequency content words)',
    '200/1000 entries have ≥1 collocation; 0/1000 reach 5. For high-frequency content words (~300-400 nouns/verbs), 5 collocations should be the bar (e.g. 雨: 雨が降る, 雨が止む, 雨に濡れる, 雨宿り, 雨の日).',
    'Migaku derives via sentence-mining; this app authoring directly is faster + more curated. Reading-comprehension pivot.',
    'Author 5 collocations per top-300 high-frequency content word; ≥2 for function/adverb/particle words.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P2: Listening timestamped transcripts
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'MEDIUM',
    'Listening depth - timestamped transcripts',
    'data/listening.json + audio/listening/*.mp3 + js/listening-transcript.js',
    'Word-level timestamps on listening transcripts: 0/47',
    '0/47 items have transcript_timestamps. Bar: word-level (or phrase-level) timestamps clickable to seek the audio. JapanesePod101 paid-tier feature; WaniKani / Bunpro do not ship for N5 chokai.',
    'Listening surface alone with this feature would credibly compete with JP101 for N5 listening prep — at zero cost.',
    'Generate via whisper.cpp word-timestamps on existing MP3s; persist to listening.json; render as click-to-seek UI.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P2: Bunpro grammar-path tags (Genki / Minna lesson tags)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'LOW',
    'Grammar pedagogy - textbook paths',
    'data/grammar.json',
    'Genki/Minna lesson tagging absent (no genki_lesson / minna_chapter on any pattern)',
    'No grammar pattern carries genki_lesson or minna_chapter field. Bunpros stickiness is exactly this: "you finished Genki I L4? Here is your N5 progress." Without lesson tags, no path-following affordance.',
    'Bunpro biggest stickiness feature; cheap to add (lesson-by-lesson lookup table); unlocks "study-with-Genki" path-view as next surface.',
    'Map each pattern to genki_lesson_n + minna_chapter_n via reference table. Then build a path-view on home page.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P2: Kanji vocab cross-links (≥5 per kanji)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
    'Kanji depth - vocab cross-links',
    'data/kanji.json',
    'Kanji entries with ≥5 vocab cross-links: 15/106 (14%)',
    '15/106 kanji entries list ≥5 vocab examples. WaniKani lists ~10 per kanji. Currently average 3 per kanji from auto-scan.',
    'Linking density at the kanji surface; learners pivot from kanji to "what vocab uses this".',
    'Auto-scan vocab.json for each kanji glyph; populate examples array with up to 10 highest-frequency vocab using the kanji.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P2: Audio per grammar example
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'MEDIUM',
    'Grammar depth - example audio',
    'data/grammar.json + audio/grammar/*.mp3',
    'Audio per grammar example: 0/178 patterns (0 / ~700+ examples)',
    '0 patterns have audio rendered for any of their examples. No incumbent ships per-example audio for grammar — leadership opportunity.',
    'Listening reinforcement at point-of-grammar-introduction is novel; gtts/VOICEVOX render is cheap; competitive lead at zero cost.',
    'Render TTS audio per example sentence; cache; expose via play button on grammar detail page.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Vocab counter pairing
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'LOW',
    'Vocab depth - counter pairing',
    'data/vocab.json',
    'Counter-pairing on nouns: 9% (88/1000)',
    'Q41-resolved at 87/589 nouns (~15%); audit recount = 88/1000 entries. Target ≥90% on counter-eligible nouns. Mandatory pairs: 本→冊, cars→台, animals→匹, flat→枚, people→人.',
    'No incumbent does this well at N5. Quick lift; high learner-trust signal.',
    'Author counter field on the remaining counter-eligible nouns.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Honorific chains
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
    'Vocab depth - honorific chains',
    'data/vocab.json',
    'Honorific/humble chain links: 9/1000 (1%)',
    'いる⇄いらっしゃる⇄おる, 食べる⇄召し上がる⇄いただく, 見る⇄ご覧になる⇄拝見する, 言う⇄おっしゃる⇄申す, 行く・来る⇄いらっしゃる⇄参る — only 9 entries cross-linked.',
    'No incumbent links these systematically at N5. Bunpro covers in essays; this app codes structurally.',
    'Author honorific_chain bidirectionally on the ~5 canonical N5 chains.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Bunpro JP-keyboard typed-input reviews
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'HIGH', 'HIGH',
    'Question types - typed input',
    'js/test.js + js/drill.js + js/review.js',
    'JP-keyboard typed-input reviews: missing (only MCQ shipped)',
    'All review surfaces are 4-option multiple-choice. Bunpros signature is type-the-answer reviews with romaji-to-kana auto-conversion (wanakana lib) and 50% partial credit.',
    'Production reviews are pedagogically stronger than recognition; Bunpro and WaniKani both ship.',
    'Add wanakana-based input box as alternate review mode; Levenshtein-fuzzy matching for partial credit; toggle on/off in settings.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Bunpro cloze deletion drills
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'HIGH', 'HIGH',
    'Question types - cloze',
    'js/test.js + data/questions.json',
    'Cloze-deletion drill type: missing',
    'No cloze-style question (fill-the-blank with multiple acceptable answers). Bunpro / Renshuu ship.',
    'Cloze forces production not recognition; complements the planned typed-input mode.',
    'Add new question_type "cloze" with acceptedAnswers list (already in some questions); render input field; render acceptable-list on review.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Tofugu-style essay per pattern (selective)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'HIGH',
    'Grammar pedagogy - long-form essay',
    'data/grammar.json + js/learn-grammar.js',
    'Tofugu-style pedagogical essay per pattern: missing (explanation_en is short)',
    'explanation_en field is 1-3 sentences typical. The N5 trickiest patterns (は/が, から/ので, 〜ている progressive vs resultative, あげる/くれる/もらう) deserve essay-length explainers (300-500 words).',
    'Tofugu ships free essay-quality content; this app at the right granularity (essay-on-demand for trickiest 20-30 patterns) closes the differentiation.',
    'Author 300-500 word `essay_en` field on the top-30 trickiest N5 patterns; expose as "Read more" toggle on grammar detail.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: WaniKani production reviews (English → JP)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'HIGH',
    'Question types - production review',
    'js/review.js',
    'Production reviews (English → Japanese): missing',
    'All reviews currently recognition-direction (Japanese → English). WaniKani ships both: recognition (kanji → meaning + reading) AND production (meaning → kanji writing).',
    'Production reviews are 2x harder retention-wise; WaniKani charges for it; this app could ship free.',
    'Add production-direction review mode (e.g., "type the kanji for: large"). Wanakana input + meaning-to-kanji match.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Frequency rank for vocab
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
    'Vocab depth - frequency rank',
    'data/vocab.json',
    'Frequency rank: 0/1000 entries',
    'No vocab entry has frequency_rank from BCCWJ or similar Japanese-corpus frequency list. Learners cannot prioritize "high-frequency first" within N5.',
    'No incumbent exposes per-word frequency for learners; quick lift; lets learners self-prioritize.',
    'Map each entry to BCCWJ-100k frequency rank (open dataset); add `frequency_rank` field; sort vocab catalog by it as default.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Reading paragraph summaries + lit/nat translation
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
    'Reading depth - paragraph + translation layer',
    'data/reading.json',
    'Per-passage paragraph summary + literal/natural translation: 0/45',
    '0/45 passages have paragraph-level summaries or literal-vs-natural English translation toggle. summary_en exists but is 1-passage-level, not paragraph-level.',
    'Marugoto has paragraph summaries; no incumbent ships literal/natural pair. Closes the N5 dokkai-prep depth gap.',
    'Author paragraph_summary list per passage + translation_literal + translation_natural fields.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P3: Listening slow-version audio
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
    'Listening depth - slow audio',
    'audio/listening/*.mp3 + js/listening.js',
    'Slow-version audio variant: 0/47',
    '0/47 listening items have a slow-speed render (0.7×). Real exam pacing is fast for true beginners; slow variant aids first-pass comprehension.',
    'JapanesePod101 paid feature; cheap to render (ffmpeg atempo=0.7); shipping free unlocks beginner usability.',
    'Render slow MP3 alongside normal; add toggle UI in listening player.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P4: Genki/JLPT-Sensei deep-link SEO
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'MEDIUM',
    'Discoverability - SEO + deep linking',
    'index.html + js/*.js + meta tags',
    'Per-content SEO meta + structured data: missing',
    'Per-grammar / per-vocab / per-kanji pages do not have og:title, og:description, structured-data JSON-LD. JLPT-Sensei ranks because of this; this app does not.',
    'JLPT-Sensei free-reference-rank channel; closing it costs little.',
    'Add og:* tags + JSON-LD per detail page; verify with Google Rich Results test.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P4: PDF/Print artifact (Genki workbook equivalent)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'MEDIUM',
    'Print/Offline - PDF cheat sheet',
    'tools/build_pdf_cheatsheet.py + index.html',
    'Genki-workbook-equivalent print/PDF artifact: missing',
    'No printable cheat sheet, vocab list, or kanji-of-the-day handout. Genki ships physical workbook; learners print summaries.',
    'Niche N2 (institutional self-host) wants this for classroom use. Differentiator from web-only competitors.',
    'Build static-site-rendered HTML→PDF for: grammar cheat sheet, vocab Anki-deck export, kanji writing-practice grid.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P4: Renshuu multi-skill drill
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
    'Drill UX - multi-skill mode',
    'js/drill.js + js/review.js',
    'Multi-skill mixed drill: partial (review queue is unified per cycle-phase 2B but no explicit "mix all" toggle)',
    'Phase 2B unified due-queue exists but not surfaced as "do all 4 skills mixed in one 15-min session". Renshuu ships this.',
    'Single-session multi-skill is the daily-drill UX learner expectation; aligns with cross-skill SRS already in code.',
    'Add "Mixed Drill" mode button on home; pull from unified queue per phase 2B helpers.',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P5: WaniKani SRS gating (vocab locked until kanji)
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'HIGH',
    'Pacing - SRS gating',
    'js/storage.js + js/learn-vocab.js',
    'WaniKani-style SRS gating (vocab locked until kanji learned): missing',
    'No gating between kanji learning and vocab using that kanji. WaniKani signature pacing pedagogy — controversial (rigid) but proven.',
    'Pacing pedagogy: forces orderly progression. Optional toggle would suffice.',
    'Add settings.gating_enabled toggle; when on, hide vocab containing un-learned kanji (per knownKanji set).',
    '', 'Fix',
    '',
    'No permission required'
))
m += 1

# IMP P5: JP-101 lesson notes PDF
imps.append(make_row(
    f'IMP-{m:03d}', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'LOW',
    'Print artifact - lesson notes',
    'tools/build_lesson_notes.py',
    'Lesson-notes PDF download: missing',
    'JP101 ships per-lesson PDF for offline study. This app could generate per-grammar / per-listening-item summary PDF.',
    'Offline study artifact; classroom + commuter use.',
    'tools/ build per-pattern + per-passage PDF summaries.',
    '', 'Defer',
    'P5 — defer behind PDF-cheatsheet (IMP-NNN) infrastructure.',
    'No permission required'
))
m += 1


# ============================================================================
# WRITE
# ============================================================================

print(f'\nIssues to register: {len(issues)}')
print(f'Improvements to register: {len(imps)}')

for row in issues + imps:
    ws.append(row)

wb.save(XLSX)
print(f'Saved.')

# Verify
wb2 = load_workbook(XLSX, read_only=True)
ws2 = wb2['Items']
print(f'New row count: {ws2.max_row} (was {hdr_row_idx + 1 + len(rows[hdr_row_idx + 1:])})')
