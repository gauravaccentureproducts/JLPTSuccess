"""Fix BUG-039 + BUG-040 — follow-ons to my recent test-scenarios xlsx
overhaul.

BUG-039 — Overview header / formula mismatch. Headers say
  B: "Perspectives covered (with #)"  → formula counts scenarios
  C: "Scenario count"                  → formula counts P1s
  D: "Priority HIGH count"             → formula uses obsolete "High"
This script applies option (c) from the bug description: full
priority distribution with correct headers.
  B "Scenarios" = COUNTA on ID column
  C "P1"        = COUNTIF for P1
  D "P2"        = COUNTIF for P2
  E "P3"        = COUNTIF for P3
  F "P4 + P5"   = COUNTIF for P4 + COUNTIF for P5

BUG-040 — Empty ISSUE-6 fields. Fill:
  K (Estimated effort): 5-bucket scale (1h / 4h / 1d / 1wk / >1wk)
    Heuristic per scenario type.
  M (Tools/scripts): "n/a — manual evaluation" where empty (216 rows).
    Existing 32 tool-filled rows kept.
  Q (Depends on): "—" everywhere (default "no dependency").
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"

CATEGORY_TABS = [
    "A. Japanese language", "B. JLPT format", "C. Hindi locale", "D. UX design",
    "E. Accessibility", "F. Security", "G. Privacy and legal", "H. Performance",
    "I. Data engineering", "J. Pedagogy", "K. QA testing", "L. Cultural ethical",
    "M. Operations", "N. End-user POV",
]


# ============================================================
# BUG-039: Overview restructure
# ============================================================
def fix_overview(wb: openpyxl.Workbook) -> int:
    """Replace Overview headers + formulas with priority distribution columns."""
    sh = wb["Overview"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Row 5 headers
    new_headers = [
        (1, "Tab"),
        (2, "Scenarios"),
        (3, "P1"),
        (4, "P2"),
        (5, "P3"),
        (6, "P4 + P5"),
    ]
    for col, label in new_headers:
        cell = sh.cell(row=5, column=col)
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Rows 6..19 per-tab formulas
    n = 0
    for i, tab in enumerate(CATEGORY_TABS):
        r = 6 + i
        sh.cell(row=r, column=1).value = tab
        sh.cell(row=r, column=2).value = f"=COUNTA('{tab}'!A5:A1000)"
        sh.cell(row=r, column=3).value = f"=COUNTIF('{tab}'!G5:G1000, \"P1\")"
        sh.cell(row=r, column=4).value = f"=COUNTIF('{tab}'!G5:G1000, \"P2\")"
        sh.cell(row=r, column=5).value = f"=COUNTIF('{tab}'!G5:G1000, \"P3\")"
        sh.cell(row=r, column=6).value = (
            f"=COUNTIF('{tab}'!G5:G1000, \"P4\") + COUNTIF('{tab}'!G5:G1000, \"P5\")"
        )
        n += 5

    # Row 20 TOTAL
    sh.cell(row=20, column=1).value = "TOTAL"
    sh.cell(row=20, column=1).font = header_font
    for col in (2, 3, 4, 5, 6):
        start_letter = chr(64 + col)
        sh.cell(row=20, column=col).value = f"=SUM({start_letter}6:{start_letter}19)"
        sh.cell(row=20, column=col).font = header_font
    n += 5

    # Set column widths
    widths = {"A": 24, "B": 12, "C": 10, "D": 10, "E": 10, "F": 12}
    for col, w in widths.items():
        sh.column_dimensions[col].width = w

    return n


# ============================================================
# BUG-040: Fill empty operational fields
# ============================================================

# Estimated-effort heuristic by category tab + per-row override
# Default per tab (median scenario effort):
DEFAULT_EFFORT_BY_TAB = {
    "A. Japanese language": "1d",     # native-teacher reviews are time-heavy
    "B. JLPT format": "4h",
    "C. Hindi locale": "1d",          # Hindi reviewer time
    "D. UX design": "1d",             # usability tests
    "E. Accessibility": "4h",          # axe-core + manual checks
    "F. Security": "4h",               # security-scan tools
    "G. Privacy and legal": "4h",
    "H. Performance": "4h",            # Lighthouse runs
    "I. Data engineering": "4h",
    "J. Pedagogy": "1d",               # rewritten to Auto, but pedagogy validation takes review
    "K. QA testing": "1d",             # exploratory sessions
    "L. Cultural ethical": "4h",
    "M. Operations": "4h",
    "N. End-user POV": "1wk",          # user-recruitment + multi-participant testing
}

# Per-row overrides (specific scenarios known to be quick or long)
EFFORT_OVERRIDES = {
    # Quick automated checks
    "A-016": "1h", "A-018": "1h", "A-027": "1h",
    "B-002": "1h", "B-005": "1h",
    "D-004": "1h", "E-001": "1h", "E-002": "1h", "E-003": "1h",
    "F-001": "1h", "F-002": "1h", "F-003": "1h", "F-004": "1h",
    "F-005": "1h", "F-006": "1h", "F-007": "1h", "F-008": "1h",
    "G-002": "1h", "H-001": "1h", "H-002": "1h", "H-003": "1h", "H-015": "1h",
    "B-018": "1h",  # stem-punctuation regex
    # Long-running content audits
    "A-001": "1wk",   # 200 examples × 3 min = 10h, plus triage
    "A-020": "1wk",   # 944 pitch-marks
    "A-024": "1wk",   # 200 grammar examples
    "C-001": "1wk",   # 150 Hindi strings
    # Multi-participant UX
    "D-013": "1wk",   # 8 participants × 20-min session + analysis
    "N-003": "1wk",   # 10 participants
    "N-017": "1wk",   # true-beginner path, 3 participants moderated
    "J-016": "1wk",   # competitor benchmark across 50 items
    # Quarterly / annual ops
    "M-010": ">1wk",  # operational review cycle
    "M-001": "1d",    # deploy rollback drill (initial setup; quarterly re-run is shorter)
    # JEES side-by-side
    "B-016": "1d",
    "B-017": "1d",
    # Restore drill
    "I-016": "4h",
    # HI audio
    "A-031": "1d",
    # LLM disclosure timing
    "L-011": "4h",
}

# Manual evaluation placeholder for empty Tools/scripts cells
MANUAL_TOOL_PLACEHOLDER = "n/a — manual evaluation"

# Default "no dependency" marker
NO_DEPENDENCY = "—"


def fill_operational_fields(wb: openpyxl.Workbook) -> dict:
    """BUG-040: Fill Estimated effort, Tools/scripts, Depends on."""
    stats = {"effort_filled": 0, "tools_filled": 0, "deps_filled": 0}
    for tab in CATEGORY_TABS:
        sh = wb[tab]
        default_effort = DEFAULT_EFFORT_BY_TAB.get(tab, "1d")
        for r in range(5, sh.max_row + 1):
            row_id = sh.cell(row=r, column=1).value
            if not row_id:
                continue
            # K: Estimated effort
            if not sh.cell(row=r, column=11).value:
                effort = EFFORT_OVERRIDES.get(row_id, default_effort)
                sh.cell(row=r, column=11).value = effort
                stats["effort_filled"] += 1
            # M: Tools/scripts — fill empty cells with manual placeholder
            if not sh.cell(row=r, column=13).value:
                sh.cell(row=r, column=13).value = MANUAL_TOOL_PLACEHOLDER
                stats["tools_filled"] += 1
            # Q: Depends on
            if not sh.cell(row=r, column=17).value:
                sh.cell(row=r, column=17).value = NO_DEPENDENCY
                stats["deps_filled"] += 1
    return stats


# ============================================================
# Main
# ============================================================
def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)
    print(f"Loaded {XLSX}")

    print("\n--- BUG-039: Overview headers + formulas restructure ---")
    n = fix_overview(wb)
    print(f"  Wrote {n} cells (5 headers + 14 tabs × 5 cols + TOTAL row)")

    print("\n--- BUG-040: Fill operational fields ---")
    stats = fill_operational_fields(wb)
    for k, v in stats.items():
        print(f"  {k}: {v}")

    # Mark BUG-039 and BUG-040 Fixed
    sh = wb["User Reported Bugs"]
    for r in (42, 43):
        bid = f"BUG-{r-3:03d}"
        if sh.cell(row=r, column=8).value not in ("Fixed", "Verified", "Closed"):
            sh.cell(row=r, column=8).value = "Fixed"
            existing = sh.cell(row=r, column=5).value or ""
            stamp = (
                f"\n\n[FIX 2026-05-17]: Applied by "
                f"tools/fix_bugs_039_040_test_scenarios_followup_2026_05_17.py."
            )
            if "[FIX 2026-05-17]" not in existing or stamp.strip() not in existing:
                sh.cell(row=r, column=5).value = existing + stamp
            print(f"  Marked {bid} Fixed.")

    # Update Summary
    sh.cell(row=4, column=12).value = 40
    sh.cell(row=5, column=12).value = 0
    sh.cell(row=7, column=12).value = 40

    wb.save(XLSX)
    print(f"\nSaved {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
