"""Back-fill Fix Commit hash in xlsx tracker for BUG-133/134/135."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_21_backfill_goi_004_006"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

FIX_COMMIT = "c236cd6"  # main repo
SUB_COMMIT = "292546f"  # submodule (procedure manual F.37)
COMBINED = f"{FIX_COMMIT} (+ submodule {SUB_COMMIT})"

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]

# Columns (from earlier audit): col 10 = Fix Commit
for row in [136, 137, 138]:
    prev = ws.cell(row=row, column=10).value
    ws.cell(row=row, column=10, value=COMBINED)
    print(f"  R{row}: {prev} -> {COMBINED}")

wb.save(XLSX)

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
print("\n=== Verified ===")
for row in [136, 137, 138]:
    print(f"  R{row}: status={ws2.cell(row=row, column=8).value} commit={ws2.cell(row=row, column=10).value}")
