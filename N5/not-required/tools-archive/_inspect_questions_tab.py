"""Look at the Questions tab to see what open product decisions remain."""
from __future__ import annotations
import io
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'feedback').exists() and ROOT != ROOT.parent:
    ROOT = ROOT.parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = load_workbook(ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx')
ws = wb['Questions']
rows = list(ws.iter_rows(values_only=True))
print(f'Total rows: {len(rows)}')
for ri, r in enumerate(rows[:50]):
    cells = [c for c in r if c is not None and str(c).strip()]
    if cells:
        print(f'\nrow {ri}: {[str(c)[:120] for c in cells]}')
