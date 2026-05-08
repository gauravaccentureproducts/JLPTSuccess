"""Remove IMP-053 (RTL CSS - deferred to future N1) from the audit registry.

User decision 2026-05-09: this item was deferred to a future locale that
doesn't exist yet (Arabic/Hebrew/Urdu). Per user, it shouldn't sit in the
N5 audit registry at all - it's not pending, not avoided, just irrelevant
to the current scope. Removing the row entirely.
"""
from __future__ import annotations
import io
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'feedback').exists() and ROOT != ROOT.parent:
    ROOT = ROOT.parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'
wb = load_workbook(XLSX)
ws = wb['Items']

rows = list(ws.iter_rows(values_only=True))
hdr_row_idx = None
for ri, r in enumerate(rows[:10]):
    cells = [str(c).lower() if c else '' for c in r]
    if 'id' in cells and any('decision' in c for c in cells):
        hdr_row_idx = ri
        break

hdr = [str(c) if c else '' for c in rows[hdr_row_idx]]
id_idx = hdr.index('ID')

target_row_num = None
for ri in range(hdr_row_idx + 2, ws.max_row + 1):
    if ws.cell(row=ri, column=id_idx + 1).value == 'IMP-053':
        target_row_num = ri
        break

if target_row_num is None:
    print('ERROR: IMP-053 not found')
    sys.exit(1)

print(f'Found IMP-053 at sheet row {target_row_num}, deleting...')
ws.delete_rows(target_row_num, 1)
wb.save(XLSX)
print('Deleted. Remaining row count:', ws.max_row - hdr_row_idx - 1)
