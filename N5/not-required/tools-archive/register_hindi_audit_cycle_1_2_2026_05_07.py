"""Register Hindi-content audit cycle 1+2 as one IMP row in the audit xlsx.

Mirrors the pattern used by tools/register_audit_*.py during the round-N
close-outs.
"""
from __future__ import annotations
import io
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'feedback').exists() and ROOT != ROOT.parent:
    ROOT = ROOT.parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'
wb = load_workbook(XLSX)
ws = wb['Items']

# Detect header row + columns
rows = list(ws.iter_rows(values_only=True))
hdr_row = None
for ri, r in enumerate(rows[:10]):
    cells = [str(c).lower() if c else '' for c in r]
    if 'id' in cells and any('decision' in c for c in cells):
        hdr_row = ri
        break
if hdr_row is None:
    raise SystemExit("could not find header row")

hdr = [str(c) if c else '' for c in rows[hdr_row]]
print('Headers:', hdr)

# Find next free IMP-NNN
existing = set()
id_idx = hdr.index('ID')
for r in rows[hdr_row + 1:]:
    iid = r[id_idx]
    if isinstance(iid, str) and iid.startswith('IMP-'):
        try:
            existing.add(int(iid.split('-')[1]))
        except ValueError:
            pass
next_imp = max(existing) + 1 if existing else 1
new_id = f'IMP-{next_imp:03d}'
print(f'Next free IMP id: {new_id}')

# Build the row in column order matching hdr
def col(name, value):
    """Return value if column exists, else None."""
    return value if name in hdr else None

def set_cell(row_data, name, value):
    if name in hdr:
        row_data[hdr.index(name)] = value

new_row = [None] * len(hdr)
set_cell(new_row, 'ID', new_id)
set_cell(new_row, 'Type', 'Improvement')
set_cell(new_row, 'Severity', 'IMPROVEMENT')
set_cell(new_row, 'Priority', 'P1')
set_cell(new_row, 'Impact', 'HIGH')
set_cell(new_row, 'Effort', 'HIGH')
set_cell(new_row, 'Category', 'Content depth / Hindi pedagogy')
set_cell(new_row, 'Location',
    'data/grammar.json + data/questions.json + data/papers/**/*.json + '
    'data/kanji.json + data/listening.json + data/reading.json + '
    'data/vocab.json + locales/hi.json + tools/check_content_integrity.py')
set_cell(new_row, 'Title',
    'Hindi quality+coverage audit cycle 1+2 - HI-01..HI-19 closed + JA-41 invariant added')
set_cell(new_row, 'Current state',
    'Cycle 1 (8 phases A-H, commits c5b3c11 -> a3de7e4) closed 17 distinct '
    'audit findings: 433 placeholder Hindi values translated, 29 paper files '
    'with zero Hindi keys filled (402 rationale_hi added), 48 kanji arity '
    'mismatches resolved, 67 kana-prefix-convention violations swept '
    '(143 romanization substitutions across EN+HI), 5 single-entry quality '
    'bugs hand-fixed, 4 UI polish strings, 11 native_reviewed-but-codemix '
    'entries hand-fixed for provenance honesty. '
    'Cycle 2 (3 phases) applied 494 more glossary substitutions and added '
    'JA-41 CI invariant locking the kana-prefix convention (R-1.1 from '
    'prompts/LocaleTransitionEnHi.txt). Final state: 50/50 invariants green, '
    '0 placeholders, 0 provenance violations, 0 kanji arity issues, '
    '0 native_reviewed code-mix. ~213 single-word English residuals in '
    'llm_curated content remain as long-tail polish targets.')
set_cell(new_row, 'Why this matters / Best-in-class',
    'Cycle-1 strategic narrowing (IMP-096) committed to native-Hindi-medium '
    'JLPT pedagogy as the primary niche. Audit ensures the Hindi corpus '
    'actually meets the bar that decision implies. Without this cycle, the '
    'Hindi locale was structurally complete but contained 433 placeholder '
    'admissions ("review pending"), 29 fully un-translated paper files, '
    'and 67 kana-convention violations - a learner running the app in HI '
    'locale would have hit "see English version" placeholders and English '
    'explanations on mock-test paper questions.')
set_cell(new_row, 'Suggested direction',
    'Done. Cycle 3 (deferred): per-surface native-speaker review pass to '
    'flip llm_curated -> native_reviewed entries that pass R-1..R-7 rubric '
    'on inspection; also handle the ~213 long-tail English residuals via '
    'either glossary or hand rewrites. See prompts/LocaleTransitionEnHi.txt '
    'cycle-2 phases 3a-3g for the planned per-surface review schedule.')
set_cell(new_row, 'Dependencies', 'IMP-096 (locale transition to en+hi)')
set_cell(new_row, 'Decision (Fix / Avoid / Defer)', 'Done')
set_cell(new_row, 'Description',
    'Documented in feedback/hindi-audit-findings-2026-05-07.md (HI-01..HI-19 '
    'with severity, surface, fix-recipe). 10 fix scripts and 8 diagnostics '
    'live in not-required/tools-archive/ as re-runnable artifacts.')
set_cell(new_row, 'Permission decision',
    'No further permission needed - user authorized native-Hindi-scholar '
    'reviewer persona on 2026-05-06 and explicitly requested cycle execution '
    '("fix all in the order of priority").')

ws.append(new_row)

# Save
wb.save(XLSX)
print(f'Appended {new_id} to feedback/n5-audit-2026-05-04.xlsx Items sheet.')

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2['Items']
last_row = list(ws2.iter_rows(values_only=True))[-1]
print('\nLast row:')
for h, v in zip(hdr, last_row):
    if v is not None:
        print(f'  {h}: {str(v)[:80]}')
