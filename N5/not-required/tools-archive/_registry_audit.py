"""One-shot status audit on the audit registry XLSX."""
from __future__ import annotations
import io
import sys
from collections import Counter
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = load_workbook('feedback/n5-audit-2026-05-04.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    # Detect header row — first row that contains 'status' or 'id' or 'title'
    hdr_row = 0
    for ri, r in enumerate(rows[:5]):
        cells = [str(c).lower() if c else '' for c in r]
        joined = '|'.join(cells)
        if 'status' in joined or ('id' in cells and 'title' in joined):
            hdr_row = ri
            break
    hdr = [str(c) if c else '' for c in rows[hdr_row]]
    data_rows = rows[hdr_row + 1:]
    print(f'\n=== Sheet: {sheet_name} ({len(data_rows)} data rows, header @ row {hdr_row+1}) ===')
    print('Headers:', hdr)
    status_idx = next((i for i, h in enumerate(hdr) if h and ('status' in h.lower() or h.lower().startswith('decision'))), None)
    if status_idx is None:
        print('  no status / decision column')
        continue
    rows = [None] + data_rows  # so the loop's rows[1:] still works below
    statuses = Counter()
    open_titles = []
    id_idx = next((i for i, h in enumerate(hdr) if h and h.lower() in ('id', 'item id', 'issue id', 'item_id')), None)
    title_idx = next((i for i, h in enumerate(hdr) if h and ('title' in h.lower() or 'description' in h.lower() or 'finding' in h.lower())), None)
    for r in rows[1:]:
        s = r[status_idx] or 'BLANK'
        statuses[s] += 1
        if isinstance(s, str) and s.lower().strip() not in ('done', 'closed', 'resolved', 'wontfix', "won't fix", 'wont fix', 'n/a', 'na'):
            iid = r[id_idx] if id_idx is not None else '?'
            title = r[title_idx] if title_idx is not None else '?'
            open_titles.append((iid, s, title))
    for k, v in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f'  {k!r}: {v}')
    if open_titles:
        print(f'  OPEN: {len(open_titles)}')
        # Full row dump for non-Done items
        loc_idx = next((i for i, h in enumerate(hdr) if h.lower() == 'location'), None)
        cur_idx = next((i for i, h in enumerate(hdr) if 'current state' in h.lower()), None)
        desc_idx = next((i for i, h in enumerate(hdr) if h.lower() == 'description'), None)
        for r in data_rows:
            s = r[status_idx] or 'BLANK'
            if isinstance(s, str) and s.lower().strip() not in ('done', 'closed', 'resolved'):
                iid = r[id_idx] if id_idx is not None else '?'
                title = r[title_idx] if title_idx is not None else '?'
                loc = r[loc_idx] if loc_idx is not None else '?'
                cur = r[cur_idx] if cur_idx is not None else '?'
                desc = r[desc_idx] if desc_idx is not None else '?'
                print(f'\n  --- {iid} ({s}) ---')
                print(f'  Title: {title}')
                print(f'  Location: {loc}')
                print(f'  Current state: {str(cur)[:300]}')
                print(f'  Description: {str(desc)[:300]}')
    else:
        print('  All items closed.')

import sys
sys.exit(0)

status_idx = next((i for i, h in enumerate(hdr) if h and 'status' in h.lower()), None)
id_idx = next((i for i, h in enumerate(hdr) if h and h.lower() in ('id', 'item id', 'issue id')), None)
title_idx = next((i for i, h in enumerate(hdr) if h and 'title' in h.lower()), None)

print(f'Columns: status={status_idx}, id={id_idx}, title={title_idx}')

statuses = Counter()
open_items = []
for r in rows[1:]:
    s = (r[status_idx] if status_idx is not None else None) or 'BLANK'
    statuses[s] += 1
    if isinstance(s, str) and s.lower() not in ('done', 'closed', 'resolved', 'wontfix', "won't fix", 'wont fix'):
        iid = r[id_idx] if id_idx is not None else '?'
        title = r[title_idx] if title_idx is not None else '?'
        open_items.append((iid, s, title))

print('\nStatus distribution:')
for k, v in sorted(statuses.items(), key=lambda x: -x[1]):
    print(f'  {k!r}: {v}')

print(f'\nOpen / non-Done items: {len(open_items)}')
for iid, s, t in open_items[:20]:
    print(f'  {iid} | {s} | {str(t)[:80]}')
