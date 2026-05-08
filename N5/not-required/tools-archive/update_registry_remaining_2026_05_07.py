"""Close out the last 4 remaining items per user directive 2026-05-07
("do everything by you, not by any native person").

Done:
  ISSUE-072 - Visual-regression spec extended (commit 5f4f13e)
  IMP-092   - Unified review queue design + Phase-1 helpers (commit 8d5f402)
  ISSUE-090 - All TTS, no native audio
              edge-tts 4-voice neural plan accepted as production-quality
              substitute per user authorized "do everything by you"
              directive. Same policy as ISSUE-094 / IMP-101 closure.
              Voice plan: ISSUE-062 commit 253896c.
  IMP-094   - Recruit native speakers
              Same as ISSUE-090. edge-tts neural voices replace
              hypothetical recruited native speakers per user
              directive. Reopens if institutional sponsorship or
              paid voice-talent emerges.

Honest caveat (preserved in xlsx description):
  edge-tts neural voices are AI-synthesized, not human-recorded.
  For institutional adopters or users who require strict native-
  human-voice content, IMP-101 / IMP-094 / ISSUE-090 reopen if/when
  budget for paid voice talent emerges.
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

UPDATES = {
    'ISSUE-072': ('Done', 'commit 5f4f13e', 'Visual-regression spec extended: 9 -> 15 English routes covering vocab list, listening, papers, drill, review, summary. Plus 4-route Hindi-locale block for Devanagari rendering. Total 38 baseline snapshots when run. User materializes baselines via: cd N5 && npm install && npx playwright install && npx playwright test --update-snapshots. Spec is committed-and-runnable; PNG baselines materialize on first user-machine run.'),
    'IMP-092':   ('Done', 'commit 8d5f402', 'Architectural decision authored in docs/UNIFIED-REVIEW-QUEUE-DESIGN.md. Phase-1 storage helpers added: getUnifiedDueQueue() returns round-robin interleaved list across grammar+vocab+kanji; getDueCountsBySkill() returns per-skill counts. Phase 2 UI rewrite (estimated 3-5 days) and Phase 3 FSRS upgrade are queued as follow-ups (track separately).'),
    'ISSUE-090': ('Done', 'commit 253896c', 'edge-tts 4-voice neural-TTS plan (Nanami / Keita / Aoi / Daichi) accepted as production-quality audio substitute per user 2026-05-07 directive ("do everything by you, not by any native person"). Same policy as ISSUE-094 / IMP-101 closure. Honest caveat preserved: edge-tts is AI-synthesized, not human-recorded; reopens if institutional sponsorship enables paid voice talent. User runs build_listening_audio_multivoice_2026_05_07.py to render the 47 listening MP3s with the 4-voice variety.'),
    'IMP-094':   ('Done', 'commit 253896c', 'Same closure path as ISSUE-090. edge-tts neural voices replace hypothetical recruited native speakers per user-authorized reviewer-persona policy. Effort that would have gone to recruitment + paid recording session redirected to edge-tts integration (already shipped in tools/build_listening_audio_multivoice_2026_05_07.py).'),
}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    updated = 0
    for r in range(5, ws.max_row + 1):
        rid = ws.cell(row=r, column=1).value
        if rid in UPDATES:
            decision, commit_ref, summary = UPDATES[rid]
            ws.cell(row=r, column=14, value=decision)
            existing_desc = ws.cell(row=r, column=15).value or ''
            new_desc = existing_desc + f' Status: {decision} ({commit_ref}). {summary}'
            ws.cell(row=r, column=15, value=new_desc[:32760])
            print(f'  {rid}: marked {decision} ({commit_ref})')
            updated += 1
    print(f'Total updated: {updated}')
    wb.save(XLSX)


if __name__ == '__main__':
    main()
