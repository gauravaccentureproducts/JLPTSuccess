"""Register the 2026-05-13 N5Improvement audit findings to the canonical
issue tracker at feedback/n5-audit-2026-05-04.xlsx.

This is the single allowed write from the audit pass — Sections 5 + 6 of
the audit transcript materialize here so findings persist across sessions.

ID range:
  ISSUE-125, 126, 127, 128 (4 issues)
  IMP-154..179 (26 improvements; some are placeholder one-liners
                committed at P4/P5 so future audits don't re-discover)

Schema: 16 columns; header at row 4. Decision defaults:
  - P1 + content-depth = Fix
  - Width-add / breaks-niche = Avoid
  - Precondition / deferred = Defer
"""
import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = "feedback/n5-audit-2026-05-04.xlsx"

# 16-col schema:
# 1 ID  2 Type  3 Severity  4 Priority  5 Impact  6 Effort  7 Category
# 8 Location  9 Title  10 Current state  11 Why  12 Suggested  13 Deps
# 14 Decision  15 Description  16 Permission

ROWS = [
    # === ISSUES ===
    [
        "ISSUE-125", "Issue", "MAJOR", "P1", "HIGH", "LOW",
        "Interconnection density",
        "data/vocab.json (1009 entries); inverted from grammar.json examples[].vocab_ids",
        "Density-2 vocab->pattern reverse map missing for 576/1009 entries",
        "576/1009 vocab entries are referenced by zero grammar patterns via vocab_ids. The reverse-lookup direction is dead: clicking 食べる should surface ~たい/~たことがある/~ながら/~てから/~なくちゃ.",
        "Density-2 floor is the prompt's explicit ≥1 bar; 57% dead-end means the largest interconnection deficit. Doubles cross-link richness mechanically.",
        "Walk grammar.json patterns→examples→vocab_ids; build the inverted map; persist as `appears_with_patterns: [pattern_id]` on each vocab entry. Candidate CI invariant JA-67.",
        "(none)",
        "Fix", "Audit 2026-05-13 N5Improvement; richness-first Density-2 floor breach.",
        "No permission required",
    ],
    [
        "ISSUE-126", "Issue", "MAJOR", "P1", "MEDIUM", "MEDIUM",
        "Interconnection density",
        "data/kanji.json — 39/106 entries",
        "Density-3 kanji->vocab below floor (<2) for 37% of corpus",
        "39 of 106 N5 kanji are used by fewer than 2 vocab entries. Bar set ≥5 (WaniKani-style); floor ≥2. 39 below the floor.",
        "WaniKani's pacing rests on kanji→vocab depth (~10/kanji); this app's 14/106 at ≥5 is far from leader.",
        "Per-kanji audit: either add 1-2 N5-scope compounds to vocab (subject to RICHNESS-FIRST width rule), OR document why the kanji is intentionally bare (rare uses), OR cross-link existing vocab that uses the kanji compositionally.",
        "(none)",
        "Defer", "Width-add for vocab is gated by RICHNESS-FIRST mandate; this issue describes the depth gap to be closed by repurposing slots, not by additions.",
        "Requires scope decision on slot repurposing",
    ],
    [
        "ISSUE-127", "Issue", "MAJOR", "P2", "MEDIUM", "LOW",
        "Documentation / discoverability",
        "docs/, top-level README, new SELFHOST.md",
        "No SELFHOST.md guide for N2 niche (institutional / fork use)",
        "Privacy/no-account posture is the strongest niche claim; operators wanting to fork get no walk-through (GitHub Pages setup, content-license posture, mirror-build instructions).",
        "N2 is the secondary defensible niche; a written guide is the gating artifact between 'partially claims' and 'credibly claims.'",
        "Author SELFHOST.md: deploy steps, content-license, GitHub Pages, mirror-build, CSP/PWA notes.",
        "(none)",
        "Fix", "Strengthens secondary niche claim with zero code/data work.",
        "No permission required",
    ],
    [
        "ISSUE-128", "Issue", "MINOR", "P3", "LOW", "LOW",
        "Tests / tooling",
        "tools/audit_refresh_state.py",
        "Audit-refresh script's voice-variety detection missed voicevox `voices_used` list",
        "Detection script counted 0 distinct voices because it expected a scalar `voice_id`; voicevox audio_render_meta uses `voices_used: [string]` instead. False-negative on listening voice-variety scorecard row.",
        "Prevents the audit from misreporting future cycles; cheap fix.",
        "Parse audio_render_meta.voices_used list; dedupe across items; report distinct count.",
        "(none)",
        "Fix", "One-shot tooling cleanup before the next audit run.",
        "No permission required",
    ],

    # === IMPROVEMENTS (P1) ===
    [
        "IMP-154", "Improvement", "IMPROVEMENT", "P1", "HIGH", "MEDIUM",
        "Pitch-accent cross-surface",
        "data/vocab.json (1009), data/grammar.json examples (~1800), data/listening.json (5-10 minimal-pair items)",
        "Pitch-accent annotation: ZERO across vocab + grammar examples + listening minimal-pairs",
        "0/1009 vocab with {mora, drop}; 0/178 grammar patterns with pitch on every example; 0/50 listening items testing pitch discrimination. Single largest cross-surface gap.",
        "OJAD has pitch but is research-tool; no learner-facing app ships pitch systematically at N5. Pure leadership move.",
        "Import kanjium pitch dictionary (CC BY-SA 4.0, EDICT-derived); per-vocab `pitch: {mora: int, drop: int}` field per NHK convention; visual downstep-marker overlay; 5-10 minimal-pair listening items (雨/飴, 橋/箸, 神/紙, 牡蠣/柿).",
        "(none)",
        "Fix", "Top-1 leverage in this audit. ~4-5 person-days total.",
        "Requires NOTICES.md attribution update for kanjium",
    ],
    [
        "IMP-155", "Improvement", "IMPROVEMENT", "P1", "MEDIUM", "LOW",
        "Mimetic / onomatopoeia layer",
        "data/vocab.json (audit existing 1009 + flag candidates)",
        "Mimetic words (擬音語/擬態語) coverage at 0.8% — prompt mandate breached",
        "8/1009 vocab marked with onomatopoeia/mimetic flag. Prompt mandates: ぺこぺこ, にこにこ, どきどき, わくわく, ぴかぴか, ゆっくり, ちょっと, だんだん, もっと, まあまあ.",
        "None of the incumbents teach mimetics systematically at N5; major leverage opportunity.",
        "Audit existing 1009 entries for these 10 mimetics; flag any present with `onomatopoeia: true`; author up to 5 missing within the frozen width (substitution policy needs Q1 decision).",
        "(see Q1 in Section 9)",
        "Fix", "Half-day work assuming most mimetics already exist in some form.",
        "No permission required (flagging existing entries); width-add subject to Q1",
    ],
    [
        "IMP-156", "Improvement", "IMPROVEMENT", "P1", "HIGH", "MEDIUM",
        "Listening aizuchi + discourse markers",
        "data/listening.json (5-8 mondai-2 items rewritten + re-rendered)",
        "Aizuchi (相槌) and discourse markers (あの/えー/まあ) entirely absent from listening corpus",
        "0/50 listening items contain aizuchi or discourse markers. Real Japanese conversation is 30%+ aizuchi; textbooks omit them; nobody teaches them at N5.",
        "Leadership richness-first lever. Closes the 'never sounds like real Japanese' criticism.",
        "Rewrite 5-8 mondai-2 dialogue items to include へぇ/そうですか/なるほど/はい/うん/本当?; flag with `aizuchi: true` and `discourse_markers: [...]`; re-render via existing VOICEVOX pipeline.",
        "(none — VOICEVOX engine already operational)",
        "Fix", "1 day work; existing voicevox pipeline.",
        "No permission required",
    ],
    [
        "IMP-157", "Improvement", "IMPROVEMENT", "P1", "HIGH", "HIGH",
        "Kanji 3-mnemonic completion",
        "data/kanji.json (106 entries × 2 new mnemonics)",
        "Kanji visual + reading mnemonics missing (1/3 of WaniKani triad present)",
        "106/106 have radical_story; 0/106 have mnemonic_visual; 0/106 have mnemonic_reading. WaniKani's defining UX advantage is the 3-mnemonic structure.",
        "Closes the single largest competitor (WaniKani) gap. Critical for N3 niche claim on kanji surface.",
        "Author 212 mnemonics (106 visual glyph-shape stories + 106 dominant-on-yomi reading stories); reuse existing radical_story authoring pattern (~50-100 chars each).",
        "(none)",
        "Defer", "2-3 days of native-reviewed authoring. Confirm P1 status given effort.",
        "Requires authoring decision",
    ],

    # === IMPROVEMENTS (P2) ===
    [
        "IMP-159", "Improvement", "IMPROVEMENT", "P2", "MEDIUM", "LOW",
        "Authentic-content layer depth",
        "data/grammar.json (top up 50-100 patterns to ≥2 PD refs)",
        "PD refs ≥2 coverage at 6% — second ref gives cross-tier triangulation",
        "11/178 patterns with ≥2 PD refs (rest at 1). Second ref ideally from a different tier (Aozora+proverb, government+folk-song) gives cross-tier triangulation.",
        "Extends the v1.15.0/v1.15.1 PD framework without new legal posture; same 5-tier source pool.",
        "Top up 50-100 high-frequency patterns to ≥2 refs; second ref from different tier.",
        "(none)",
        "Fix", "Same legal posture as v1.15.0/v1.15.1.",
        "No permission required",
    ],
    [
        "IMP-160", "Improvement", "IMPROVEMENT", "P2", "MEDIUM", "MEDIUM",
        "Vocab transitivity pair",
        "data/vocab.json (verbs subset)",
        "Transitivity-pair cross-links: 20/1009 entries (Genki lists 14 N5 pairs)",
        "20/1009 vocab marked with transitivity_pair. Genki ships 14 explicit N5 pairs (開く/開ける, 始まる/始める, etc.). Coverage gap.",
        "Genki teaches these explicitly; matching is parity, not leadership.",
        "Audit all N5 verbs for transitivity; cross-link via `transitivity_pair: <vocab_id>`.",
        "(none)",
        "Fix",
        "Genki parity item.",
        "No permission required",
    ],
    [
        "IMP-161", "Improvement", "IMPROVEMENT", "P2", "MEDIUM", "LOW",
        "Vocab verb_class flag",
        "data/vocab.json (verbs subset)",
        "verb_class (godan/ichidan/irregular) at 13% — many N5 verbs untagged",
        "132/1009 entries have verb_class. ~200-300 N5 verbs total; ~70 missing the tag.",
        "X-6.6 invariant exposes Group-1-exception flag; UI needs the verb_class field to power the popover.",
        "Derive godan/ichidan/irregular + Group-1-exception flag from existing form.",
        "(none)",
        "Fix",
        "Mechanical fill from kana ending.",
        "No permission required",
    ],
    [
        "IMP-162", "Improvement", "IMPROVEMENT", "P2", "LOW", "LOW",
        "Vocab counter pairing",
        "data/vocab.json (countable nouns subset)",
        "Counter pairing (noun -> dominant counter) at 28% — many entries lack counter",
        "292/1009 nouns paired with dominant counter. Q41 partially closed at 87/589.",
        "None of the incumbents expose noun->counter cross-links well at N5; ship the link, win the dimension.",
        "Cross-link 本→冊, 車→台, 紙→枚, 人→人, etc.",
        "(none)",
        "Fix",
        "Mechanical; closes the partial-closure of Q41.",
        "No permission required",
    ],
    [
        "IMP-163", "Improvement", "IMPROVEMENT", "P2", "MEDIUM", "LOW",
        "Vocab pragmatic-function enum",
        "data/vocab.json (multi-function words)",
        "Pragmatic-function enumeration at 4% — single-gloss treatment misses real usage",
        "43/1009 marked with pragmatic_functions. Mandatory list per prompt: すみません, 大丈夫, どうぞ, どうも, ちょっと, 結構.",
        "Single-translation glosses are the most common beginner-app weakness; enumerate functions inline.",
        "Enum: すみません={apology, attention-getter, gratitude}; 大丈夫={confirmation, polite-refusal}; どうぞ={offering, please-do, by-all-means}; etc.",
        "(none)",
        "Fix",
        "Native-teacher dimension; matters for register-appropriate output.",
        "No permission required",
    ],
    [
        "IMP-168", "Improvement", "IMPROVEMENT", "P2", "MEDIUM", "MEDIUM",
        "Reading paragraph_summary",
        "data/reading.json (54 passages)",
        "Reading paragraph_summary: 0/54 — only depth dimension still missing",
        "0/54 reading passages have paragraph-level summaries. Reading is otherwise depth-leader (grammar footnotes, vocab preview, audio, translation, cultural callout, reflection prompts, topic tag all 100%).",
        "Closes the single remaining reading-richness dimension; cheap.",
        "Per-paragraph 1-sentence Japanese summary at N5 level.",
        "(none)",
        "Fix",
        "Routine authoring; reading corpus benefits cleanly.",
        "No permission required",
    ],

    # === IMPROVEMENTS (P3) ===
    [
        "IMP-164", "Improvement", "IMPROVEMENT", "P3", "LOW", "LOW",
        "Devoiced-vowel markers",
        "data/vocab.json + data/listening.json transcripts",
        "Devoiced-vowel marker: 0% across vocab and listening transcripts",
        "0/1009 vocab entries flagged for devoiced vowel; 0/50 listening transcripts mark devoicing.",
        "Beginners miss devoicing in real speech because textbooks use over-articulated audio; flagging closes the discrimination gap.",
        "Flag vocab where standard Tokyo speech devoices a vowel (です→'des'', すき→'sk'', きく→'kk', した→'sht'').",
        "(IMP-154 pitch couples here)",
        "Defer",
        "Coupled with pitch ship in IMP-154; ship together.",
        "No permission required",
    ],
    [
        "IMP-165", "Improvement", "IMPROVEMENT", "P3", "MEDIUM", "MEDIUM",
        "Cross-surface minimal-pair links",
        "data/vocab.json + data/kanji.json + data/listening.json",
        "Minimal-pair links: pitch + long-vowel + sokuon all 0% across vocab/kanji/listening",
        "Pitch minimal-pair link 0% vocab + 0% kanji; long-vowel minimal-pair 0% vocab; sokuon/long-vowel listening 0/50.",
        "Discrimination ability is core JLPT-listening skill; explicit pair-coverage is the only way to teach.",
        "Cross-link 雨/飴, 橋/箸 (pitch); ビル/ビール, おばさん/おばあさん (long-vowel); 来た/切った, きて/きって (sokuon).",
        "(couples with IMP-154 pitch)",
        "Defer",
        "Ship alongside IMP-154 pitch package.",
        "No permission required",
    ],
    [
        "IMP-166", "Improvement", "IMPROVEMENT", "P3", "LOW", "LOW",
        "Vocab authentic_ref expansion",
        "data/vocab.json (high-frequency content words)",
        "Vocab authentic_ref at 3% — leverage on content words underused",
        "37/1009 vocab entries have authentic_ref. PD-refs framework (v1.15.1) is grammar-side; vocab equivalent underdeveloped.",
        "Same legal-safe PD-ref framework extends to vocab (Aozora quotes citing specific words, proverbs, folk songs).",
        "Add PD refs to top 100 high-frequency content words; same 5-tier source pool.",
        "(IMP-159 pattern depth couples here)",
        "Defer",
        "Best done as a second wave after pattern-side completion.",
        "No permission required",
    ],
    [
        "IMP-167", "Improvement", "IMPROVEMENT", "P3", "LOW", "LOW",
        "Kanji okurigana_cuts complete",
        "data/kanji.json (62/106 entries missing)",
        "Okurigana_cuts at 41% — kanji boundary marks incomplete",
        "44/106 kanji marked with okurigana cuts. Boundary marks (食べる = 食‧べる, 帰る = 帰‧る) needed for hand-writing accuracy.",
        "Hand-writing prep depends on this; recognition-only treatment misses it.",
        "Mark okurigana boundary on all remaining 62 kanji entries.",
        "(none)",
        "Fix",
        "Mechanical fill.",
        "No permission required",
    ],
    [
        "IMP-169", "Improvement", "IMPROVEMENT", "P3", "MEDIUM", "HIGH",
        "Listening timestamped transcripts",
        "data/listening.json (50 items) + VOICEVOX timing extraction",
        "Listening timestamped transcripts: 0/50",
        "0/50 listening items have word-level timestamps clickable to seek.",
        "JapanesePod101 paid tier has this; offering at no cost is a niche win.",
        "Extract per-segment timing during VOICEVOX synthesis; persist `timestamps: [{word, t_start, t_end}]`.",
        "(VOICEVOX timing API)",
        "Defer",
        "Higher-effort tooling; deferred until pitch + aizuchi ship first.",
        "Requires VOICEVOX timing API verification",
    ],
    [
        "IMP-170", "Improvement", "IMPROVEMENT", "P3", "LOW", "LOW",
        "Listening inference-question expansion",
        "data/listening.json question_type field",
        "Inference-question type at 0/50 (may be detection issue; verify enum)",
        "Detection found 0/50 with `question_type == 'inference'`. May be measurement issue (different enum value) or genuine gap.",
        "Inference beyond literal retrieval is what real JLPT-N5 mondai-3/4 tests.",
        "First verify the actual question_type enum across items; if genuine gap, add inference variants to 5-10 items.",
        "(none)",
        "Fix",
        "Measurement step first.",
        "No permission required",
    ],

    # === IMPROVEMENTS (P4) ===
    [
        "IMP-171", "Improvement", "IMPROVEMENT", "P4", "LOW", "LOW",
        "Bunpro ghost-review mode",
        "SRS engine",
        "Ghost-review (failed cards reappear sooner) — undocumented",
        "Bunpro's signature SRS feature; this app's SRS may have it but is not documented.",
        "Bunpro-parity feature; minor.",
        "Document existing behavior; add explicit ghost-mode flag if missing.",
        "(none)",
        "Defer",
        "Bunpro-parity, low priority.",
        "No permission required",
    ],
    [
        "IMP-172", "Improvement", "IMPROVEMENT", "P4", "LOW", "LOW",
        "Review-forecast UI verification",
        "Progress dashboard",
        "Review forecast (7-day projected load) — substring match ambiguous",
        "Audit substring match found 'forecast' but couldn't confirm it's a 7-day load projection. Verify or build.",
        "Bunpro's defining UX feature for daily-routine learners.",
        "Verify UI surface; build if absent.",
        "(none)",
        "Defer",
        "Bunpro-parity, low priority.",
        "No permission required",
    ],
    [
        "IMP-173", "Improvement", "IMPROVEMENT", "P4", "MEDIUM", "MEDIUM",
        "Production reviews (EN -> JP typing)",
        "Test mode + drill engine",
        "Production reviews (EN->JP typing) — recognition-only currently",
        "App is recognition-only. WaniKani pairs recognition reviews with production (EN→JP typing).",
        "Closes WaniKani gap; pairs with cloze-deletion already shipped.",
        "Add production-mode toggle to drill engine; reuse JP-keyboard input.",
        "(IMP-157 kanji mnemonics couples)",
        "Defer",
        "WaniKani-parity feature; ship after kanji mnemonic completion.",
        "No permission required",
    ],
    [
        "IMP-174", "Improvement", "IMPROVEMENT", "P4", "MEDIUM", "MEDIUM",
        "Anki / CSV deck export",
        "Tools or Test/Progress route",
        "Anki/CSV deck export missing",
        "Migaku/Anki audience overlap; open-format export is N1+N2 lever.",
        "Self-host/fork users (N2) and Anki users (Migaku-overlap) benefit; privacy-aligned.",
        "CSV-export endpoint for vocab + grammar; optional Anki-package wrapper.",
        "(none)",
        "Defer",
        "N1+N2 niche-strengthening.",
        "No permission required",
    ],
    [
        "IMP-175", "Improvement", "IMPROVEMENT", "P4", "LOW", "LOW",
        "Lesson-notes PDF export route",
        "Print CSS + PDF route",
        "Lesson-notes PDF export route partial",
        "Print CSS exists for cheat-sheet/lesson; explicit PDF download endpoint not verified.",
        "JP101 paid feature; offering free strengthens N1 + N3 claims.",
        "Verify print route; add PDF download for offline study.",
        "(none)",
        "Defer",
        "Low priority; print CSS already exists.",
        "No permission required",
    ],
    [
        "IMP-176", "Improvement", "IMPROVEMENT", "P4", "LOW", "LOW",
        "Words-containing-this-kanji UI",
        "Kanji detail page",
        "Jisho-style 'words containing this kanji' UI surface verification",
        "n5_compounds field exists 100/106; verify the UI surface actually exposes the list on the kanji detail page.",
        "Jisho-parity feature; cheap to verify/expose.",
        "Verify UI surface; expose n5_compounds list with click-through.",
        "(none)",
        "Fix",
        "Routine UI verification.",
        "No permission required",
    ],
    [
        "IMP-177", "Improvement", "IMPROVEMENT", "P4", "MEDIUM", "MEDIUM",
        "SEO / discoverability push",
        "Home page, meta tags, robots.txt, sitemap.xml",
        "SEO discoverability — competitors all rank for 'JLPT N5 grammar'",
        "Live URL exists; SEO push not made. JLPT Sensei ranks well via deep-linked free reference.",
        "N1+N3 niche-strengthening (visible privacy-first option in search results).",
        "Add per-pattern meta tags; sitemap.xml; robots.txt; structured data; canonical URLs.",
        "(none)",
        "Defer",
        "Discoverability work; non-content.",
        "No permission required",
    ],

    # === IMPROVEMENTS (P5) ===
    [
        "IMP-178", "Improvement", "IMPROVEMENT", "P5", "LOW", "LOW",
        "Paper-Q -> entry cross-link verification",
        "data/papers/ + review flow",
        "Density-7: paper-question cross-link to entry on review — unverified",
        "Density-7 target 100%; not measured by the refresh script. Verify whether each paper question on review surfaces a click-through to its grammar/vocab/kanji entry.",
        "Closes the review-flow density check.",
        "Add measurement step to audit-refresh; ship cross-link if absent.",
        "(none)",
        "Defer",
        "Measurement first.",
        "No permission required",
    ],
    [
        "IMP-179", "Improvement", "IMPROVEMENT", "P5", "LOW", "LOW",
        "Lookalike-cluster all-pairs link verification",
        "data/kanji.json lookalikes",
        "Density-8: lookalike clusters — all-pairs link unverified",
        "103/106 kanji have lookalikes flagged; whether every cluster member links to every other member (transitivity) not measured.",
        "Cluster-internal completeness is the audit-time check.",
        "Verify each cluster forms a complete graph; fix any holes.",
        "(none)",
        "Defer",
        "Measurement step.",
        "No permission required",
    ],
]


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Items"]

    # Find next empty row
    first_empty = ws.max_row + 1
    print(f"Appending {len(ROWS)} rows starting at row {first_empty}")

    for i, row in enumerate(ROWS):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=first_empty + i, column=col_idx, value=value)

    wb.save(XLSX)
    print(f"Saved. Sheet 'Items' now has {ws.max_row} rows.")


if __name__ == "__main__":
    main()
