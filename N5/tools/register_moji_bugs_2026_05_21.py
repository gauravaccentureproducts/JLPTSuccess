#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Register MOJI-001..007 bugs (moji paper review, 2026-05-21).

Adds 7 rows to the User Reported Bugs tab of
specifications/test-scenarios-by-specialist-perspective.xlsx.

Idempotent: skips bugs whose title prefix (MOJI-NNN) already exists.

Bug source: maintainer-supplied moji paper-1..paper-7 content review.
Same registration pattern as the MOB-001..019 batch on 2026-05-19.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")

REPORTER = "Content audit (moji paper review, 2026-05-21) — paper-1.json through paper-7.json, Q1-Q100"

BUGS = [
    (
        "MOJI-001 — Stem markup convention split: Mondai 1 uses HTML <u>…</u> (50/50), Mondai 2 uses markdown __…__ (50/50); paper-4 mixes both within one file; renderer-fragility + undocumented contract",
        (
            "FILE: paper-1.json through paper-7.json (Moji papers, Q1-Q100).\n\n"
            "PROBLEM:\n"
            "Across all 100 moji questions:\n"
            "  Mondai 1 / <u>…</u>: 50 questions (Papers 1-3 + paper-4 Q46-Q50)\n"
            "  Mondai 2 / __…__:    50 questions (paper-4 Q51-Q60 + Papers 5-7)\n"
            "Convention is consistent per-mondai but paper-4 mixes both within "
            "one file: Q1-Q5 use <u>古い</u>/<u>高い</u>/etc., Q6-Q15 switch to "
            "__がくせい__/__せんせい__/etc. at the Mondai 1→2 boundary.\n\n"
            "TWO DISTINCT PROBLEMS:\n"
            "(a) Rendering fragility: an HTML-only renderer displays `__がくせい__` "
            "literally as underscores around text. A markdown-only renderer "
            "leaves `<u>古い</u>` as raw HTML tags in the displayed stem. Either "
            "way the underline-emphasis fails on half the corpus.\n"
            "(b) Undocumented convention: per-mondai split appears intentional "
            "(Mondai 1 underlines the kanji whose reading is asked; Mondai 2 "
            "marks the kana that needs kanji-conversion) but no schema doc or "
            "_meta block declares 'Mondai 1 ⇒ HTML <u>, Mondai 2 ⇒ markdown __'. "
            "Future authors / renderers / migrations can't tell whether the "
            "split is by design or accident.\n\n"
            "FIX DIRECTION: pick one convention and apply corpus-wide. Recommend "
            "HTML `<u>…</u>` for both (since stem_html is already an HTML-shaped "
            "field name + rendering target is a web UI).\n"
            "  Option A — standardize to HTML <u>:\n"
            "    Mondai 2 stems become: 'うちには <u>こども</u> が ふたり います。'\n"
            "  Option B — standardize to markdown __ (riskier; field name is "
            "stem_html, implying HTML rendering):\n"
            "    Mondai 1 stems become: 'あの 人は __学生__ です。'\n\n"
            "BUILD INVARIANT after fix: stem_html for moji questions must "
            "contain exactly one of {<u>…</u>, __…__} as the emphasis wrapper, "
            "AND the chosen wrapper must match the corpus-wide convention. CI "
            "fails on mismatch."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOJI-002 — 28 moji questions carry spurious grammarPatternId values from auto_inferred; moji tests orthography, not grammar; same anti-pattern class as PAPER-001 n5-013 over-misuse",
        (
            "FILE: paper-1.json through paper-7.json (all moji papers).\n\n"
            "PROBLEM:\n"
            "Among 100 moji questions, 72 correctly have grammarPatternId=null "
            "with provenance='not_applicable_orthography'. The remaining 28 "
            "carry pattern IDs from auto-inference:\n"
            "  moji-1.6 (Q6) → n5-108     |  moji-4.10 (Q55) → n5-041\n"
            "  moji-1.7 (Q7) → n5-108     |  moji-4.14 (Q59) → n5-108\n"
            "  moji-2.3 (Q18) → n5-117    |  moji-5.2 (Q62) → n5-013\n"
            "  moji-2.7 (Q22) → n5-116    |  moji-6.3 (Q78) → n5-188\n"
            "  moji-2.11 (Q26) → n5-017   |  moji-6.7 (Q82) → n5-058\n"
            "  moji-2.12 (Q27) → n5-017   |  moji-6.8 (Q83) → n5-147\n"
            "  moji-2.14 (Q29) → n5-067   |  moji-6.13 (Q88) → n5-118\n"
            "  moji-3.2 (Q32) → n5-108    |  moji-6.14 (Q89) → n5-076\n"
            "  moji-3.5 (Q35) → n5-067    |  moji-6.15 (Q90) → n5-118\n"
            "  moji-3.6 (Q36) → n5-078    |  moji-7.2 (Q92) → n5-039\n"
            "  moji-4.1 (Q46) → n5-078    |  moji-7.6 (Q96) → n5-078\n"
            "  moji-4.2 (Q47) → n5-080    |  moji-7.8 (Q98) → n5-078\n"
            "  moji-4.3 (Q48) → n5-078\n"
            "  moji-4.4 (Q49) → n5-078\n"
            "  moji-4.5 (Q50) → n5-078\n"
            "  moji-4.7 (Q52) → n5-028\n\n"
            "Pattern matches are surface artifacts of auto_inferred picking up "
            "incidental token similarities, not real grammar tests:\n"
            "  - moji-5.2 'うちには __こども__ が ふたり います' → n5-013 because the kana "
            "も appears in 子ども. But the question tests recognition of 子+ども as a "
            "kanji compound, NOT も particle usage. Same class as the n5-013 "
            "over-misuse fixed at PAPER-001 in the bunpou paper sweep.\n"
            "  - moji-3.5 'まちの <u>北</u> に あります' → n5-067 because に appears. "
            "But the question tests reading of 北, not locative-に.\n"
            "  - moji-2.3 '<u>今日</u> は とても いい てんきです' → n5-117 because は "
            "appears. But the question tests jukujikun reading of 今日.\n\n"
            "FIX DIRECTION: scrub all 28 to null + not_applicable_orthography "
            "in one pass. If any individual question is later determined to "
            "legitimately test a pattern, restore individually with manual "
            "review.\n\n"
            "BUILD INVARIANT after fix: moji-category questions with "
            "grammarPatternId non-null must have grammarPatternId_provenance "
            "!= 'auto_inferred' (forcing manual sign-off). CI greps for the "
            "combination and fails."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOJI-003 — moji-3.5 (Q35) is a meaning-discrimination question disguised as a reading question: 3 of 4 distractors are NOT readings of 北 (they're readings of 東/南/西); only instance in 60 Mondai 1 questions",
        (
            "FILE: paper-3.json (moji Mondai 1).\n\n"
            "moji-3.5 (Q35):\n"
            "  stem:    '私の いえは まちの <u>北</u> に あります。'\n"
            "  choices: ['ひがし', 'きた', 'みなみ', 'にし']\n"
            "  correct: 'きた'\n\n"
            "PROBLEM:\n"
            "JLPT N5 Moji Mondai 1 (漢字読み) tests phonological reading: given "
            "a kanji, pick its correct reading. Distractors should be plausible "
            "misreadings of the SAME kanji (e.g., for 北: きた / きだ / ほく / ぼく — "
            "kun-yomi, voicing variant, on-yomi, voiced on-yomi).\n\n"
            "Current Q35 instead presents the four cardinal directions:\n"
            "  ひがし = 東 (east, not 北)\n"
            "  きた  = 北 (north, correct)\n"
            "  みなみ = 南 (south, not 北)\n"
            "  にし  = 西 (west, not 北)\n"
            "A student who knows 北 means 'north' but cannot read it phonologically "
            "still scores correctly by elimination — they just need to know which "
            "direction 北 represents and discard the readings of the other three "
            "directions. This tests direction-vocabulary recognition (Mondai 3 "
            "goi-context territory) rather than kanji-reading discrimination "
            "(Mondai 1).\n\n"
            "This is the ONLY case of antonymic-distractor problem across all 60 "
            "Mondai 1 questions in the corpus. Other apparent 'no-shared-initial-"
            "mora' flags (moji-1.10 day-of-week, moji-2.1 number on/kun, "
            "moji-2.10 voicing variants) are legitimate per-kanji reading tests — "
            "all test real readings of the keyed kanji.\n\n"
            "FIX DIRECTION:\n"
            "Replace three distractors with plausible misreadings of 北:\n"
            "  choices: ['きた', 'きだ', 'ほく', 'ぼく']  (or similar)\n"
            "  きた = correct kun-yomi\n"
            "  きだ = voicing-variant trap (rendaku where it doesn't apply)\n"
            "  ほく = on-yomi (correct in 北西/北部 compounds; trap standalone)\n"
            "  ぼく = voiced on-yomi (no-such-reading trap)\n"
            "Update rationale: '北 = きた (kun-yomi, used standalone or as "
            "direction). On-yomi ホク appears in compounds 北西 (ほくせい) / "
            "北部 (ほくぶ) but not standalone.'"
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOJI-004 — moji-5.2 (Q62) marks the legitimate Japanese spelling 子供 as WRONG; rationale acknowledges 子供 is standard yet the UI still penalizes the learner",
        (
            "FILE: paper-5.json.\n\n"
            "moji-5.2 (Q62):\n"
            "  stem:    'うちには __こども__ が ふたり います。'\n"
            "  choices: ['子ども', '字ども', '小ども', '子供']\n"
            "  correct: '子ども' (option 0)\n"
            "  Marked wrong: '子供' (option 3)\n\n"
            "PROBLEM:\n"
            "The rationale honestly acknowledges:\n"
            "  '子ども is selected here because it follows this corpus's "
            "N5-only-kanji policy (供 is N4). Both 子供 and 子ども are standard "
            "in modern Japanese, and on the actual JLPT both forms appear; the "
            "choice between them is a corpus-internal scope rule, not a "
            "correctness rule.'\n\n"
            "Rationale honesty is excellent — but the question UI flow is:\n"
            "  Learner picks 子供 → marked WRONG → reads rationale → learns the "
            "marking was a corpus-policy decision, not a Japanese-correctness "
            "fact.\n"
            "This creates a confidence-eroding moment where a student who picks "
            "a real spelling sees it labeled wrong. Same anti-pattern class as "
            "REG-001 'だれ vs どなた marked Incorrect when both are legitimate "
            "register variants'.\n\n"
            "FIX DIRECTION:\n"
            "Replace the 子供 distractor with one that's unambiguously wrong "
            "orthographically. Candidates that share 子 shape but produce "
            "non-words:\n"
            "  '子問'  (子 + 問 — both N5; 子問 is not a word)\n"
            "  '子文'  (子 + 文 — both N5; 子文 is not a word)\n"
            "  '子分'  (子 + 分 — both N5; this IS a word meaning 'underling' but "
            "rare enough at N5 to be a clean distractor)\n\n"
            "After replacement, rationale simplifies to:\n"
            "  '子ども (こども - child). Note: 子供 is also a valid spelling but "
            "uses the kanji 供 which is N4 and outside the N5 whitelist.'\n"
            "The follow-up sentence keeps the pedagogical observation without "
            "making the test-taking experience marked 'wrong' for a valid "
            "spelling."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOJI-005 — moji-2.1 and moji-2.2 rationale_hi has over-literal 'के पास है पढ़ते हुए' translation of 'has reading'; word-by-word Hindi instead of natural 'X का पठन Y है'",
        (
            "FILE: paper-2.json.\n\n"
            "Two consecutive questions carry the same translation artifact:\n"
            "  moji-2.1 (Q16) rationale (EN): '七 has reading シチ in 七月.'\n"
            "  moji-2.1 (Q16) rationale_hi:   '七 के पास है पढ़ते हुए シチ में 七月।'\n"
            "  moji-2.2 (Q17) rationale (EN): '四 has reading シ in 四月.'\n"
            "  moji-2.2 (Q17) rationale_hi:   '四 के पास है पढ़ते हुए シ में 四月।'\n\n"
            "PROBLEM:\n"
            "Hindi phrase 'के पास है पढ़ते हुए' is a word-by-word rendering of "
            "'has reading'. Literal mapping:\n"
            "  七 (X) + के पास (near/has) + है (is) + पढ़ते हुए (while reading) "
            "+ シチ + में (in) + 七月 → grammatically broken Hindi.\n\n"
            "NATURAL HINDI:\n"
            "  moji-2.1 rationale_hi: '七 का पठन 七月 में シチ है।'\n"
            "                         (The reading of 七 in 七月 is シチ.)\n"
            "  moji-2.2 rationale_hi: '四 का पठन 四月 में シ है।'\n"
            "                         (The reading of 四 in 四月 is シ.)\n\n"
            "Same class as DOKKAI-002 ('एक महीना ago'), DOKKAI-004 ('आना-जाना by "
            "ट्रेन'), and the older PAPER-004 Hindi-quality sweep. Translation "
            "pass missed two moji-corpus entries where the English source uses "
            "a verb construction ('has reading') that has no direct Hindi "
            "cognate.\n\n"
            "FIX DIRECTION:\n"
            "Rewrite both rationale_hi as natural Hindi using 'X का पठन Y है' or "
            "'X की Y के रूप में पढ़ाई होती है' construction. Update provenance "
            "to native_reviewed_2026_05_21.\n\n"
            "BUILD INVARIANT after fix: corpus-wide signature for this defect "
            "class is 'के पास है' anywhere in rationale_hi. CI greps for it."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOJI-006 — moji-7.2 rationale_hi truncated relative to EN: drops the 'for N5 the 立 form is the only correct match' conclusion sentence; HI learner left ambiguous",
        (
            "FILE: paper-7.json.\n\n"
            "moji-7.2 (Q92):\n"
            "  EN rationale (343 chars): ends with '…Broader-exposure students "
            "should not be misled by the polysemy; for N5 the 立 form is the "
            "only correct match for \"students stand up when the teacher "
            "enters\".'\n"
            "  HI rationale (250 chars): ends with '…व्यापक-अनुभव वाले छात्र इन "
            "वैकल्पिक कान्जी से परिचित हो सकते हैं।' — translates to 'students "
            "with broader exposure may be familiar with these alternative "
            "kanji.'\n\n"
            "PROBLEM:\n"
            "HI cuts off after acknowledging that experienced students might "
            "know the alternative kanji but DROPS the crucial conclusion: 'for "
            "N5 the 立 form is the only correct match.' Without this sentence, "
            "a Hindi-speaking learner who recognized 起ちます / 経ちます / 建ちます "
            "is left wondering whether one of THOSE could also be the correct "
            "answer — EN reader gets a definitive 'no, 立 is the only N5-scope "
            "answer'; HI reader doesn't.\n"
            "This is a content-coverage mismatch (rationales don't say the same "
            "thing in two languages), not a translation-quality issue per se.\n\n"
            "FIX DIRECTION:\n"
            "Extend HI rationale to match EN's conclusion:\n"
            "  rationale_hi: '立ちます (खड़ा होना — たつ का रोज़मर्रा N5 अर्थ)। अन्य "
            "रूप 起ちます / 経ちます / 建ちます भी असली जापानी क्रियाएँ हैं जो "
            "たちます पढ़ी जाती हैं (उठना / समय बीतना / इमारत खड़ी होना), पर वे N3+ "
            "दायरे में हैं। व्यापक-अनुभव वाले छात्र इन वैकल्पिक कान्जी से परिचित हो "
            "सकते हैं; पर N5 स्तर पर \"शिक्षक के आने पर छात्र खड़े होते हैं\" के लिए "
            "立 ही एकमात्र सही उत्तर है।'\n"
            "Update provenance: native_reviewed_2026_05_21.\n\n"
            "BUILD INVARIANT: rationale and rationale_hi should be within ~30% of "
            "each other in character count (after accounting for Hindi's ~1.3× "
            "expansion ratio vs English). CI flags entries where one is "
            "significantly truncated relative to the other for manual review."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOJI-007 — moji-7.8 (Q98) distractor 永い is also a valid reading of ながい; rationale is one-token ('長い.') and misses a teachable polysemy moment that moji-7.2 handles excellently",
        (
            "FILE: paper-7.json.\n\n"
            "moji-7.8 (Q98):\n"
            "  stem:    'この かわは とても __ながい__ です。'\n"
            "  choices: ['永い', '長い', '強い', '高い']\n"
            "  correct: '長い'\n"
            "  rationale (EN): '長い.'\n"
            "  rationale (HI): '長い।'\n\n"
            "PROBLEM:\n"
            "Both rationales are one-token. They miss a teachable moment that "
            "parallel moji-7.2 handles well:\n"
            "  moji-7.2 rationale: '立ちます (stand up - the everyday N5 sense of "
            "たつ). The other forms 起ちます / 経ちます / 建ちます are real Japanese "
            "verbs also read たちます …'\n\n"
            "The 永い distractor is exactly the same shape of polysemy:\n"
            "  ながい = 長い (long in space/time, N5 everyday sense)\n"
            "  ながい = 永い (long-lasting, eternal — N3+ literary/poetic use, "
            "e.g., 永い眠り 'eternal sleep', 永い愛 'everlasting love')\n"
            "For 'この かわは とても ながい' (this river is very long), only 長い is "
            "natural (a river has length, not eternity). 永い is contextually "
            "wrong here but lexically valid as a reading of ながい.\n"
            "A broader-exposure student who recognizes 永い might second-guess "
            "and worry whether 永い could also be correct. Current rationale "
            "doesn't help — just asserts '長い.' parallel to moji-7.2's much more "
            "helpful treatment of 起ちます/経ちます/建ちます.\n\n"
            "FIX DIRECTION:\n"
            "Extend rationale to flag the 永い polysemy:\n"
            "  rationale: '長い (long - the everyday N5 sense of ながい for "
            "physical or temporal length). The distractor 永い is also a real "
            "reading of ながい meaning \"eternal / everlasting\" (N3+ scope; used "
            "in literary contexts like 永い眠り \"eternal sleep\"). For a river, "
            "only 長い is natural.'\n"
            "  rationale_hi: '長い (लंबा — ながい का रोज़मर्रा N5 अर्थ, भौतिक या "
            "कालिक लम्बाई के लिए)। विकर्षक 永い भी ながい का असली पठन है जिसका अर्थ "
            "\"शाश्वत / सनातन\" है (N3+ दायरा; साहित्यिक संदर्भों में 永い眠り "
            "\"शाश्वत निद्रा\" जैसा प्रयोग)। नदी के लिए केवल 長い स्वाभाविक है।'\n"
            "Provenance: native_reviewed_2026_05_21."
        ),
        "Minor", "P3", "Open"
    ),
]


def main() -> None:
    today = date.today().strftime("%Y_%m_%d")
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{today}_moji_bugs")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)
    ws = wb["User Reported Bugs"]
    print(f"Sheet has {ws.max_row} rows currently.")

    # Existing prefixes (for idempotency)
    existing = set()
    for r in range(4, ws.max_row + 1):
        t = ws.cell(row=r, column=4).value
        if t:
            existing.add(str(t).split("—")[0].strip())

    written, skipped = 0, 0
    today_str = date.today().isoformat()
    for title, description, severity, priority, status in BUGS:
        prefix = title.split("—")[0].strip()
        if prefix in existing:
            print(f"  SKIP (already exists): {prefix}")
            skipped += 1
            continue
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=f'="BUG-"&TEXT(ROW()-3,"000")')
        ws.cell(row=new_row, column=2, value=today_str)
        ws.cell(row=new_row, column=3, value=REPORTER)
        ws.cell(row=new_row, column=4, value=title)
        ws.cell(row=new_row, column=5, value=description)
        ws.cell(row=new_row, column=6, value=severity)
        ws.cell(row=new_row, column=7, value=priority)
        ws.cell(row=new_row, column=8, value=status)
        ws.cell(row=new_row, column=9, value="")
        ws.cell(row=new_row, column=10, value="")
        for c in range(1, 11):
            ws.cell(row=new_row, column=c).alignment = WRAP
        max_len = max(len(str(v) or "") for v in [title, description])
        ws.row_dimensions[new_row].height = max(45, min(320, max_len // 5))
        print(f"  WROTE row {new_row}: {prefix}")
        written += 1

    wb.save(WORKBOOK)
    print()
    print(f"Result: {written} new bug rows written, {skipped} skipped")


if __name__ == "__main__":
    main()
