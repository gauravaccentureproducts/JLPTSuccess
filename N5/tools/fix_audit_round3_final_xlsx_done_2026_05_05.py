"""Stamp Done on the round-3 items shipped in v1.12.31 (the second
round-3 close-out, which lands the previously-deferred 20).

Items closed this batch:
  Data:     ISSUE-013, IMP-005
  Storage:  IMP-008, IMP-031, IMP-033, IMP-036
  Routes:   IMP-044, IMP-037, ISSUE-020, IMP-032
  Audio:    IMP-007, IMP-010, IMP-038
  A11y:     IMP-006, IMP-012
  Content:  IMP-019 (scaffold), IMP-042 (workflow doc)
  i18n:     ISSUE-022, IMP-034, IMP-041 (key extraction scaffold)

Total: 20 of 20 deferred items resolved (some as full implementations,
some as scaffolds + documentation per the close-out commit's CHANGELOG
"Caveats" subsection).
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

DONE_IDS = {
    # Phase A
    'IMP-005', 'ISSUE-013',
    # Phase B+C
    'IMP-008', 'IMP-031', 'IMP-033', 'IMP-036',
    'IMP-044', 'IMP-037', 'ISSUE-020', 'IMP-032',
    # Phase D
    'IMP-007', 'IMP-010', 'IMP-038',
    # Phase E
    'IMP-006', 'IMP-012',
    # Phase F
    'IMP-019', 'IMP-042',
    # Phase G
    'ISSUE-022', 'IMP-034', 'IMP-041',
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)
    closed = []
    skipped = []
    not_found = set(DONE_IDS)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_row_idx = None
        for r in range(1, min(7, ws.max_row + 1)):
            cells = [str(c.value or '').strip() for c in ws[r]]
            if any(c.lower() == 'id' for c in cells) and any(
                c.lower().startswith('decision') for c in cells
            ):
                header_row_idx = r
                break
        if header_row_idx is None:
            continue
        header = [c.value for c in ws[header_row_idx]]
        try:
            id_col = next(i for i, v in enumerate(header, start=1)
                          if v and str(v).strip().lower() in ('id', 'issue id', 'item id'))
            dec_candidates = [i for i, v in enumerate(header, start=1)
                              if v and str(v).strip().lower().startswith('decision (')]
            if not dec_candidates:
                dec_candidates = [i for i, v in enumerate(header, start=1)
                                  if v and str(v).strip().lower().startswith('decision')]
            dec_col = dec_candidates[0]
        except (StopIteration, IndexError):
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            id_cell = row[id_col - 1]
            dec_cell = row[dec_col - 1]
            id_val = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
            if id_val not in DONE_IDS:
                continue
            cur = (dec_cell.value or '').strip() if isinstance(dec_cell.value, str) else ''
            not_found.discard(id_val)
            if cur == 'Done':
                skipped.append((sheet, id_val))
                continue
            dec_cell.value = 'Done'
            closed.append((sheet, id_val, cur or '<blank>'))

    wb.save(XLSX)
    print(f'Closed {len(closed)} item(s):')
    for sh, iid, prev in closed:
        print(f'  [{sh}]  {iid}  ({prev} -> Done)')
    if skipped:
        print(f'\nAlready Done ({len(skipped)}): {", ".join(s[1] for s in skipped)}')
    if not_found:
        print(f'\nWARNING: not found ({len(not_found)}): {sorted(not_found)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
