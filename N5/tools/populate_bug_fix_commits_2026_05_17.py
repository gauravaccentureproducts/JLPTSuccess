"""Populate "Fix Commit" + "Fix Date" columns on User Reported Bugs sheet.

INV-9 promotion (Cross-Artifact Sync Protocol) — Partial → Wired:
"Closed bug links to fix commit + regression test (or 'no test —
reason: ...')". Pre-fix, the xlsx had 8 columns (Bug ID / Date /
Reported By / Title / Description / Severity / Priority / Status) —
no fix-commit link. This script:

  1. Adds two columns: Fix Commit (col 9) + Fix Date (col 10)
  2. Walks every Fixed-status row and scans git log for the commit
     that closed it (heuristic: commit message subject mentions the
     BUG-NNN ID)
  3. Populates the cells; falls back to "(unattributable)" when no
     commit references the bug ID

Idempotent: re-running on an already-populated xlsx is a no-op.

Run from N5/:
    python tools/populate_bug_fix_commits_2026_05_17.py
"""
from __future__ import annotations

import io
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"


def git_log_for_bug(bug_id: str) -> tuple[str, str] | None:
    """Find the commit whose message references this BUG ID.

    Searches for commits whose subject mentions:
      - the literal BUG-NNN
      - or a "close BUG-NNN" / "fix BUG-NNN" pattern
      - or "BUG-NNN through BUG-MMM" range covering this ID

    Returns (commit_sha_short, ISO_date_string) of the most recent
    matching commit, or None.
    """
    # Try direct subject-match first
    try:
        result = subprocess.run(
            ["git", "log", "--all", "--oneline", "--format=%h|%ai|%s",
             f"--grep={bug_id}", "-i"],
            cwd=ROOT.parent,  # repo root
            capture_output=True, text=True, encoding="utf-8", timeout=15,
        )
    except subprocess.TimeoutExpired:
        return None
    except Exception:
        return None
    if result.returncode != 0:
        return None
    lines = [ln for ln in result.stdout.splitlines() if ln.strip()]
    if not lines:
        # Try range patterns: "BUG-NNN through BUG-MMM" or "BUG-NNN..MMM"
        # Extract the numeric suffix and search for ranges that cover it
        m = re.match(r"BUG-(\d+)", bug_id)
        if not m:
            return None
        bug_n = int(m.group(1))
        # Search for commit subjects with "BUG-..." that might be a range
        try:
            r2 = subprocess.run(
                ["git", "log", "--all", "--oneline",
                 "--format=%h|%ai|%s", "--grep=BUG-", "-i"],
                cwd=ROOT.parent, capture_output=True, text=True,
                encoding="utf-8", timeout=15,
            )
        except Exception:
            return None
        for ln in (r2.stdout or "").splitlines():
            if not ln.strip():
                continue
            parts = ln.split("|", 2)
            if len(parts) != 3:
                continue
            sha, date, subject = parts
            # Look for "BUG-XXX through BUG-YYY" or "BUG-XXX..YYY"
            range_m = re.search(
                r"BUG-(\d+)\s*(?:through|\.\.|to)\s*BUG-(\d+)",
                subject, re.IGNORECASE,
            )
            if range_m:
                lo, hi = int(range_m.group(1)), int(range_m.group(2))
                if lo <= bug_n <= hi:
                    return (sha, date.split("T")[0].split(" ")[0])
            # Also "BUG-NNN..MMM" without word "through"
            range_m2 = re.search(
                r"BUG-(\d+)\.\.(\d+)", subject,
            )
            if range_m2:
                lo, hi = int(range_m2.group(1)), int(range_m2.group(2))
                if lo <= bug_n <= hi:
                    return (sha, date.split("T")[0].split(" ")[0])
        return None
    # Take the most recent (first line from git log --oneline)
    parts = lines[0].split("|", 2)
    if len(parts) != 3:
        return None
    sha, date, _ = parts
    return (sha, date.split("T")[0].split(" ")[0])


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)
    if "User Reported Bugs" not in wb.sheetnames:
        print("ERROR: User Reported Bugs sheet missing")
        return 1
    b = wb["User Reported Bugs"]

    # Locate / ensure Fix Commit + Fix Date columns exist
    headers = {}
    for c in range(1, b.max_column + 1):
        v = b.cell(row=3, column=c).value
        if isinstance(v, str):
            headers[v.strip()] = c

    fix_commit_col = headers.get("Fix Commit")
    fix_date_col = headers.get("Fix Date")

    next_col = b.max_column + 1
    if fix_commit_col is None:
        fix_commit_col = next_col
        b.cell(row=3, column=fix_commit_col).value = "Fix Commit"
        # Style — match other headers
        cell = b.cell(row=3, column=fix_commit_col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="14452A", end_color="14452A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        next_col += 1
        print(f"Added 'Fix Commit' column at index {fix_commit_col}")

    if fix_date_col is None:
        fix_date_col = next_col
        b.cell(row=3, column=fix_date_col).value = "Fix Date"
        cell = b.cell(row=3, column=fix_date_col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="14452A", end_color="14452A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        print(f"Added 'Fix Date' column at index {fix_date_col}")

    # Walk every Fixed-status row; look up git log; populate
    status_col = headers.get("Status", 8)
    bug_id_col = 1
    populated = 0
    skipped_existing = 0
    unattributable = 0
    for r in range(4, b.max_row + 1):
        bid = b.cell(row=r, column=bug_id_col).value
        if not bid:
            continue
        status = b.cell(row=r, column=status_col).value
        if status not in ("Fixed", "Closed"):
            continue
        # Skip if already populated
        existing = b.cell(row=r, column=fix_commit_col).value
        if existing and isinstance(existing, str) and existing.strip():
            skipped_existing += 1
            continue
        # Resolve bug ID
        if isinstance(bid, str) and bid.startswith("="):
            bid_resolved = f"BUG-{r-3:03d}"
        else:
            bid_resolved = str(bid)
        # Look up git
        info = git_log_for_bug(bid_resolved)
        if info:
            sha, date = info
            b.cell(row=r, column=fix_commit_col).value = sha
            # Convert date to datetime so the cell formats as date
            try:
                b.cell(row=r, column=fix_date_col).value = datetime.strptime(date, "%Y-%m-%d")
            except ValueError:
                b.cell(row=r, column=fix_date_col).value = date
            populated += 1
        else:
            b.cell(row=r, column=fix_commit_col).value = "(unattributable)"
            unattributable += 1
            print(f"  WARN: no commit found for {bid_resolved}")

    wb.save(XLSX)
    print()
    print(f"Populated {populated} rows from git log.")
    print(f"Skipped {skipped_existing} already-populated rows.")
    print(f"Marked {unattributable} as (unattributable).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
