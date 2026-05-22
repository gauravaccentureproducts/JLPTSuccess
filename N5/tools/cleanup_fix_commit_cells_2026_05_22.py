"""One-time cleanup of the xlsx Fix Commit column.

Survey (2026-05-22):
 - 4 rows have valid commit hashes (recent close-outs)
 - 142 rows have a date in BOTH col 9 (Fix Date) and col 10 (Fix Commit)
   — col 10 is redundant
 - 8 rows have a date in col 10 only, with col 9 empty — wrong column

Cleanup:
 - Duplicates (142): replace col 10 with the literal sentinel
     "<no-hash-archived; see Fix Date col 9>"
   (preserves the audit trail: "we didn't record the hash at the time;
    the date is in the proper column 9")
 - Mislocated (8): copy the date from col 10 to col 9 (if col 9 empty),
   then sentinelize col 10
 - Already-hash (4): leave alone

Idempotent: re-running this script on already-cleaned data is a no-op.

After this lands, the new JA-146 invariant accepts either:
  (a) a commit-hash-shaped string (7-40 hex, optional "(+ submodule HASH)")
  (b) a "<...>"-shaped sentinel acknowledging hash absence
and rejects datetime objects + date strings + free text.
"""
import sys, io, os, shutil, re, datetime as _dt
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_fix_commit_cleanup"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]

SENTINEL = "<no-hash-archived; see Fix Date col 9>"
HASH_RE = re.compile(r"^[a-f0-9]{7,40}(\s+\(\+\s+submodule\s+[a-f0-9]{7,40}\))?$")

stats = {"hash_kept": 0, "duplicate_sentinelized": 0, "mislocated_moved": 0, "already_sentinel": 0, "untouched_other": 0}
problems_remaining = []

for r in range(4, ws.max_row + 1):
    title = ws.cell(row=r, column=4).value
    status = ws.cell(row=r, column=8).value
    fd = ws.cell(row=r, column=9).value   # Fix Date
    fc = ws.cell(row=r, column=10).value  # Fix Commit
    if not title or status != "Fixed":
        continue

    fc_is_datetime = isinstance(fc, (_dt.datetime, _dt.date))
    fc_is_date_str = isinstance(fc, str) and len(fc) == 10 and fc[4:5] == "-" and fc[7:8] == "-" and re.match(r"^\d{4}-\d{2}-\d{2}$", fc)
    fc_is_hash = isinstance(fc, str) and HASH_RE.match(fc.strip())
    fc_is_sentinel = isinstance(fc, str) and re.match(r"^<[^>]+>$", fc.strip())
    fd_filled = fd is not None and str(fd).strip() != ""

    if fc_is_hash:
        stats["hash_kept"] += 1
        continue

    if fc_is_sentinel:
        stats["already_sentinel"] += 1
        continue

    if fc_is_datetime or fc_is_date_str:
        if fd_filled:
            # Duplicate — col 9 already has the date; col 10 redundant
            ws.cell(row=r, column=10, value=SENTINEL)
            stats["duplicate_sentinelized"] += 1
        else:
            # Mislocated — move date to col 9, sentinelize col 10
            ws.cell(row=r, column=9, value=fc)
            ws.cell(row=r, column=10, value=SENTINEL)
            stats["mislocated_moved"] += 1
        continue

    # Free text or other shape — leave for human review
    stats["untouched_other"] += 1
    problems_remaining.append((r, fc, (title or "")[:60]))

wb.save(XLSX)

print("=== Cleanup stats ===")
for k, v in stats.items():
    print(f"  {k}: {v}")

if problems_remaining:
    print()
    print(f"Untouched rows needing human review ({len(problems_remaining)}):")
    for r, fc, t in problems_remaining[:10]:
        print(f"  R{r} fc={fc!r} title={t}")

print()
print(f"Saved to {XLSX}; backup at {BAK}")
