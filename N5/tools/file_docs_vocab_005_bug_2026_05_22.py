"""File DOCS-VOCAB-005 (BUG-136) in the xlsx bug tracker.

Honest scope after verification of the original bug-report:
The 28 paper-file `source_file` strings are NOT broken file-path references.
They are explicit prose annotations:
  "(authored in-place; was KnowledgeBank/<x>_questions_n5.md before KnowledgeBank/ merge into data/ + docs/N5-syllabus-methodology.md on 2026-05-14)"
that already document the merge.

The bug-report's proposed fix (replace with docs/N5-syllabus-methodology.md#bunpou-questions)
was rejected because:
  1. Those anchor IDs don't exist in the methodology doc (it has Part C/D/E/F headings).
  2. Pointing source_file at the methodology doc would falsely imply the doc
     contains the questions (it describes authoring methodology, not question content).
  3. Replacing accurate "authored in-place" prose with a non-existent pointer
     would remove correct provenance.

Honest fix (per user choice 2026-05-22): trim the prose to just
"(authored in-place)" — the historical KnowledgeBank breadcrumb is preserved
in CHANGELOG + n5_vocab_whitelist_README.md + git history; doesn't need to
live in every paper-file's metadata.
"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_docs_vocab_005"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
print(f"Current max_row: {ws.max_row}")

title = (
    "DOCS-VOCAB-005 — 28 paper-file source_file fields carry now-stale "
    "KnowledgeBank/ historical breadcrumb in prose annotation"
)
description = """FILES: data/papers/{bunpou,dokkai,goi,moji}/paper-{1..7}.json (28 files).

Bug-report claim (autonomous-loop): "every paper file's source_file field points to the deleted KnowledgeBank/ directory; the audit chain question → source of truth is broken; anyone tracing a question hits a 404."

Verified actual state: source_file values are NOT broken file-path references. They are explicit parenthesized prose annotations of the shape:

  "(authored in-place; was KnowledgeBank/<x>_questions_n5.md before KnowledgeBank/ merge into data/ + docs/N5-syllabus-methodology.md on 2026-05-14)"

A JSON consumer that reads this value as a file path would fail at the leading parenthesis, not 404. The string explicitly says "authored in-place" — i.e., no upstream source file.

Bug-report proposed fix (rejected as factually inaccurate):
- Replace with docs/N5-syllabus-methodology.md#bunpou-questions (and category-parallel anchors). The methodology doc has no #bunpou-questions / #dokkai-questions / #goi-questions / #moji-questions headings; it has ## Part C/D/E/F covering vocab/kanji/grammar/question-authoring conventions.
- Pointing source_file at the methodology doc would falsely imply the doc contains the question content; it describes authoring methodology, not question content.
- This would degrade correctness, not fix it.

Honest fix (per user choice 2026-05-22): trim source_file prose from "(authored in-place; was KnowledgeBank/<x>_questions_n5.md before ...)" to just "(authored in-place)". Historical breadcrumb is preserved in CHANGELOG + n5_vocab_whitelist_README.md + git history (commit 136abc4 of 2026-05-14); doesn't need to live in every paper-file's metadata.

CI invariant JA-145 wires the canonical sentinel: source_file in every data/papers/<cat>/paper-N.json must be either (a) a path that resolves to an existing repo file, OR (b) the literal string "(authored in-place)". Other values fail.

Lineage: same defect class as DOCS-VOCAB-001..004 + DOCS-KANJI-001..004 (governance-doc stale-content) closed by commit a979874 on 2026-05-21 with JA-144. This is DOCS-VOCAB-005.

Proposed CI invariant JA-145: per the canonical sentinel rule above. Catches future drift if source_file gets repopulated with a stale path."""

# Append at row 139 (BUG-136)
r = 139
ws.cell(row=r, column=1, value='="BUG-"&TEXT(ROW()-3,"000")')
ws.cell(row=r, column=2, value="2026-05-22")
ws.cell(row=r, column=3, value="Content audit (DOCS-VOCAB sweep iteration 2)")
ws.cell(row=r, column=4, value=title)
ws.cell(row=r, column=5, value=description)
ws.cell(row=r, column=6, value="Low")
ws.cell(row=r, column=7, value="P4")
ws.cell(row=r, column=8, value="Open")
print(f"  Row {r}: BUG-{r-3:03d} = {title[:70]}...")

wb.save(XLSX)

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
print(f"  Verify R{r}: status={ws2.cell(row=r, column=8).value} sev={ws2.cell(row=r, column=6).value}/{ws2.cell(row=r, column=7).value}")
print("Saved.")
