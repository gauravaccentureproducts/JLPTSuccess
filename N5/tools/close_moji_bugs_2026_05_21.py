#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Mark MOJI-001..007 as Fixed in the User Reported Bugs tab.

Status: 'Open' → 'Fixed'
Fix Commit: (left blank for now; back-filled by populate_bug_fix_commits
            script post-commit, OR set manually if commit hash known)
Fix Date: today (2026-05-21)

Idempotent: skips bugs already at 'Fixed'.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()

# Bug prefixes to close
TARGETS = {f"MOJI-{i:03d}" for i in range(1, 8)}


def main() -> None:
    today_str = date.today().strftime("%Y_%m_%d")
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{today_str}_close_moji")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)
    ws = wb["User Reported Bugs"]
    closed, skipped = 0, 0
    for r in range(4, ws.max_row + 1):
        title = ws.cell(row=r, column=4).value or ""
        prefix = str(title).split("—")[0].strip()
        if prefix not in TARGETS:
            continue
        current_status = (ws.cell(row=r, column=8).value or "").strip()
        if current_status.lower() == "fixed":
            print(f"  SKIP R{r} {prefix}: already Fixed")
            skipped += 1
            continue
        ws.cell(row=r, column=8, value="Fixed")
        # Fix Commit (col 9): leave blank; will be back-filled by
        # populate_bug_fix_commits after the commit lands. Or set here
        # if the commit hash were already known.
        ws.cell(row=r, column=10, value=TODAY)
        ws.cell(row=r, column=8).alignment = WRAP
        ws.cell(row=r, column=10).alignment = WRAP
        print(f"  CLOSED R{r} {prefix}: status Fixed, fix_date {TODAY}")
        closed += 1

    wb.save(WORKBOOK)
    print()
    print(f"Result: {closed} closed, {skipped} skipped")


if __name__ == "__main__":
    main()
