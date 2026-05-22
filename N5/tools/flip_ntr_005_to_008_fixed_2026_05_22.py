"""Flip BUG-165/166/167/168 (NTR-005/006/007/008) to Fixed."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_flip_ntr_005_to_008"
if not os.path.exists(BAK): shutil.copy2(XLSX, BAK)
wb = load_workbook(XLSX); ws = wb["User Reported Bugs"]
FLIPS = {
    168: ("NTR-005 — おはし section re-tagged from '20. Tableware and Cooking' to "
          "'19. Tableware and Cooking' (the canonical tableware section; section 20 is "
          "actually Colors). One-line edit. section_provenance bumped to native_reviewed_2026_05_22."),
    169: ("NTR-006 — えいが (movie) re-tagged from '26. House and Furniture' to "
          "'37. Common nouns - miscellaneous'. The corpus lacks a dedicated entertainment/leisure "
          "section; 37 is the closest fit. A future schema extension could introduce a proper "
          "Entertainment section. section_provenance bumped to native_reviewed_2026_05_22."),
    170: ("NTR-007 — 三 kanji mnemonic.reading softened. Old text claimed the honorific -さん is "
          "etymologically shared with 三/さん; new text states the shared sound is coincidental "
          "(honorific さん derives from 様 → さま → さん, a separate root). Preserves the "
          "cognitive crutch ('san is everywhere') without the false etymology claim. "
          "provenance.reading bumped to native_reviewed_2026_05_22."),
    171: ("NTR-008 — Flagged 3 pitch-accent entries (あなた, みなさん, きのう) with "
          "native_review_pending='2026_05_22_NTR_008' annotation. これ confirmed matching NHK "
          "and NOT flagged. The annotation documents that NHK 2016 edition values differ from "
          "current kanjium-by-reading lookup; audio file (if rendered) matches the current drop "
          "value, so audio is the source of truth for rendered material until actual native-speaker "
          "pass resolves the dictionary-vs-audio gap. Closes the bug as 'annotated-pending-native-review' "
          "per NATIVE-SPEAKER-RE-VERIFICATION.md path-forward; not a code/data error."),
}
for row, note in FLIPS.items():
    ws.cell(row=row, column=8, value="Fixed")
    ws.cell(row=row, column=9, value="2026-05-22")
    ws.cell(row=row, column=10, value="<pending-main-commit>")
    ws.cell(row=row, column=11, value=note)
    print(f"  R{row}: Fixed")
wb.save(XLSX)
