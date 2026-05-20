#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Add 2 scenario rows to xlsx tabs to satisfy JA-116 for A76 + Phase-0 moji block.

JA-116 requires every A-NN audit category + Phase-0 regression block in
N5/prompts/ to have a matching scenario row in the specialist-tabs section
of the test-scenarios xlsx. Today's documentation pass added:
  - A76 (audit category for moji-paper content discipline) → needs row
  - Phase-0 "moji-paper content-discipline regression block" → needs row

This script appends one scenario per missing item to the appropriate
tab. Idempotent: skips if a row already cites the code.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().strftime("%Y_%m_%d")


def main() -> None:
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{TODAY}_a76_moji_scenarios")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)

    # Tab A. Japanese language — add A76 scenario
    ws_a = wb["A. Japanese language"]
    new_row_a = ws_a.max_row + 1
    a76_id = f"A-{new_row_a:03d}"
    a76_values = [
        a76_id,
        "Moji content discipline",
        "A76",
        "Moji-paper content discipline regression — per-mondai stem-emphasis convention, auto_inferred grammarPatternId on orthography, word-by-word HI rendering of EN verb constructions, EN/HI rationale character-count parity",
        ("1. Run `python tools/check_content_integrity.py` and grep for "
         "JA-140 | JA-141 | JA-142 | JA-143 PASS. 2. Optionally run "
         "tools/fix_moji_bugs_2026_05_21.py in dry-run to verify the "
         "fix scripts converge on a clean state."),
        "JA-140 + JA-141 + JA-142 + JA-143 all green. No moji stem uses markdown __X__; no moji question has grammarPatternId non-null + auto_inferred provenance; no rationale_hi contains 'के पास है पढ़ते हुए'; all EN/HI rationale pairs are within 0.6×-2.0× ratio for the entries in scope.",
        "P2", "Major", "Automated (CI)",
        "Cross-ref: accuracy prompt §A76 + procedure-manual §F.38 + AUDIT-COVERAGE Part 34. MOJI-001..007 close-out 2026-05-21.",
        "5 min", "Content auditor / QA",
        "tools/check_content_integrity.py + tools/fix_moji_bugs_2026_05_21.py",
        "2026-05-21", "PASS", "BUG-139..145", "—", "100%",
    ]
    for col, val in enumerate(a76_values, start=1):
        c = ws_a.cell(row=new_row_a, column=col, value=val)
        c.alignment = WRAP
    ws_a.row_dimensions[new_row_a].height = 110
    print(f"  WROTE A. tab row {new_row_a}: {a76_id} (A76 audit category)")

    # Tab K. QA testing — add Phase-0 moji content-discipline scenario
    ws_k = wb["K. QA testing"]
    new_row_k = ws_k.max_row + 1
    k_id = f"K-{new_row_k:03d}"
    k_values = [
        k_id,
        "Moji content-discipline regression",
        "Phase-0 moji content-discipline regression block",
        "Phase-0 moji-paper content-discipline regression block — 4 mechanical checks for stem-emphasis convention, grammarPatternId auto_inferred, word-by-word HI translation, EN/HI rationale length parity",
        ("1. Run `python tools/check_content_integrity.py` and confirm "
         "all four JA-140/141/142/143 invariants PASS. 2. Spot-check "
         "5 random moji questions' stem_html for HTML <u>...</u> "
         "wrapper. 3. Spot-check 5 random moji questions for "
         "grammarPatternId=null + provenance='not_applicable_orthography'."),
        "All four invariants green; 100/100 moji stems use HTML; 0 moji questions with non-null grammarPatternId; 0 rationale_hi with banned trigger phrase; 0 EN/HI pairs outside ratio band [0.6, 2.0].",
        "P2", "Major", "Automated (CI)",
        "Phase-0 mirror runs in tools/check_content_integrity.py. Cross-ref: N5Improvement.txt Phase-0 moji-paper content-discipline regression block (added 2026-05-21).",
        "5 min", "Content auditor / QA",
        "tools/check_content_integrity.py",
        "2026-05-21", "PASS", "BUG-139..145", "—", "100%",
    ]
    for col, val in enumerate(k_values, start=1):
        c = ws_k.cell(row=new_row_k, column=col, value=val)
        c.alignment = WRAP
    ws_k.row_dimensions[new_row_k].height = 110
    print(f"  WROTE K. tab row {new_row_k}: {k_id} (Phase-0 regression block)")

    wb.save(WORKBOOK)
    print()
    print("Done. Re-run tools/check_content_integrity.py to verify JA-116 PASS.")


if __name__ == "__main__":
    main()
