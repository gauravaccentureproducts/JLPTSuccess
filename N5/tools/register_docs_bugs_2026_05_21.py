#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Register 8 documentation-stale-content bugs to User Reported Bugs.

4 for `data/n5_kanji_whitelist.exceptions.md` (DOCS-KANJI-001..004)
4 for `data/n5_vocab_whitelist_README.md`     (DOCS-VOCAB-001..004)

Idempotent: skips bugs whose title prefix (DOCS-KANJI-NNN /
DOCS-VOCAB-NNN) already exists.

Source: maintainer-supplied content audit, 2026-05-21.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()
REPORTER = "Content audit (governance docs review, 2026-05-21) — data/n5_kanji_whitelist.exceptions.md + data/n5_vocab_whitelist_README.md"

BUGS = [
    # ============== n5_kanji_whitelist.exceptions.md (4) ==============
    (
        'DOCS-KANJI-001 — "canonically 103 kanji per JLPT.jp" is factually inaccurate; JLPT.jp explicitly does NOT publish official kanji lists post-2010 reform',
        (
            "FILE: data/n5_kanji_whitelist.exceptions.md (top paragraph).\n\n"
            "CURRENT TEXT:\n"
            "  'Each line below documents a kanji that is in the project "
            "whitelist (n5_kanji_whitelist.json) but not in the official "
            "JLPT N5 kanji scope (canonically 103 kanji per JLPT.jp).'\n\n"
            "AUTHORITY-CLAIM PROBLEM:\n"
            "JLPT.jp's own FAQ "
            "(https://www.jlpt.jp/sp/e/faq/) explicitly states:\n"
            "  'Therefore, we decided that publishing \"Test Content "
            "Specifications\" containing a list of vocabulary, kanji and "
            "grammar items was not necessarily appropriate.'\n\n"
            "Post-2010 reform, JLPT replaced kanji/vocab/grammar lists with "
            "competency-based descriptions. The '103 kanji' figure circulates "
            "widely in third-party study materials (WaniKani, Tofugu, Genki, "
            "'first 103 kanji' flashcard decks) and matches the pre-2010 "
            "旧4級 (Old level 4) published list — but it is NOT a current "
            "canonical figure from JLPT.jp.\n\n"
            "Citing 103 'canonically per JLPT.jp' is misleading. The actual "
            "scope boundary for N5 today is a consensus reconstruction from "
            "textbooks and past papers, not an authoritative published list.\n\n"
            "FIX DIRECTION: rewrite the top paragraph to reflect the actual "
            "sourcing. A defensible version:\n"
            "  'Each line below documents a kanji that is in the project "
            "whitelist (n5_kanji_whitelist.json) but not in the project's "
            "reference scope (data/n5_official_kanji_scope.json — a consensus "
            "reconstruction of N5 kanji scope, ~100-103 characters, derived "
            "from major textbooks (Genki, Minna no Nihongo) and analysis of "
            "pre-2010 旧4級 lists and post-2010 official sample materials, "
            "since JLPT.jp no longer publishes specific kanji lists).'\n\n"
            "Three effects: (a) drops the false JLPT.jp authority claim, "
            "(b) acknowledges 103 as a range estimate not a fixed number, "
            "(c) makes the project's actual scope source-of-truth explicit."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "DOCS-KANJI-002 — Empty Exceptions section has no template / commented-out example; future contributors can't tell the line format",
        (
            "FILE: data/n5_kanji_whitelist.exceptions.md (Exceptions section).\n\n"
            "PROBLEM:\n"
            "The top of the file lists what each exception line requires:\n"
            "  - The kanji glyph\n"
            "  - WHY: a one-sentence reason\n"
            "  - REVIEW_DATE: optional date for re-evaluation\n\n"
            "But the Exceptions section is empty:\n"
            "  '(none currently documented - JA-25 is in bootstrapping mode…)'\n\n"
            "A first-time contributor adding a kanji exception has no concrete "
            "template to copy. The three required fields could be on one "
            "line, three lines, in a table, as a YAML/JSON block — the "
            "format is underspecified.\n\n"
            "FIX DIRECTION: add a commented-out example showing the exact "
            "format the tooling expects. For example:\n\n"
            "  ## Exceptions\n\n"
            "  (none currently documented - JA-25 is in bootstrapping mode…)\n\n"
            "  <!--\n"
            "  Format (one block per exception):\n\n"
            "  ### 妹\n"
            "  - WHY: appears as recognition-only distractor in moji-4.12 (Q57) per\n"
            "    moji-corpus kanji-scope exception (Mondai 2 distractors may use\n"
            "    non-whitelist kanji for family-relation tests).\n"
            "  - REVIEW_DATE: 2026-08-01\n\n"
            "  ### 供\n"
            "  - WHY: appears in 子供 spelling variant of こども; documented in\n"
            "    moji-5.2 rationale as N5-policy-excluded alternative spelling.\n"
            "  - REVIEW_DATE: 2026-07-15\n"
            "  -->\n\n"
            "These examples correspond to actual moji corpus cases (moji-4.12's "
            "妹 distractor and moji-5.2's 子供 discussion) so they double as "
            "self-documenting evidence of where the policy gets applied."
        ),
        "Minor", "P4", "Open"
    ),
    (
        "DOCS-KANJI-003 — 'bootstrapping mode' has no documented exit criterion; data/n5_official_kanji_scope.json target/owner/contents are all undefined",
        (
            "FILE: data/n5_kanji_whitelist.exceptions.md (Notes section).\n\n"
            "PROBLEM:\n"
            "The file repeats twice that JA-25 is currently in 'bootstrapping "
            "mode until the official-scope reference file is added at "
            "data/n5_official_kanji_scope.json', but doesn't say:\n"
            "  - Who creates that file (project owner? CI auto-generation? "
            "manual transcription from a reference?)\n"
            "  - When (target version? prerequisite for what release?)\n"
            "  - What its expected contents are (the 103 kanji from 旧4級? "
            "A consensus list? A merged superset of multiple references?)\n\n"
            "Without these, 'bootstrapping mode' is an indefinite holding "
            "state and the invariant JA-25 is permanently inactive. The risk "
            "is that the whitelist drifts from anyone's understanding of "
            "N5 scope and there's nothing to detect it.\n\n"
            "FIX DIRECTION: add a 'Bootstrapping exit criteria' section to "
            "the Notes:\n\n"
            "  ## Bootstrapping exit criteria\n\n"
            "  JA-25 leaves bootstrapping mode when ALL of:\n"
            "  1. data/n5_official_kanji_scope.json is committed, containing "
            "the consensus N5 kanji list (target: ~100-103 entries; sources: "
            "intersection of Genki I-II vocab kanji + Minna no Nihongo Shokyu "
            "1-2 kanji + pre-2010 旧4級 list).\n"
            "  2. Owner reviews the n5_kanji_whitelist.json delta vs the "
            "scope file and adds WHY entries here for every superset kanji.\n"
            "  3. CI begins running JA-25 in enforce mode (currently no-op).\n\n"
            "  Target: v1.16.0 (or whenever the next major content review "
            "lands).\n"
            "  Owner: [TODO — assign]."
        ),
        "Minor", "P4", "Open"
    ),
    (
        "DOCS-KANJI-004 — REVIEW_DATE format unspecified (ISO 8601? free-form? strict pattern?) — downstream parsing risk",
        (
            "FILE: data/n5_kanji_whitelist.exceptions.md.\n\n"
            "PROBLEM:\n"
            "The REVIEW_DATE field is described as 'optional date for "
            "re-evaluation' but no format. Will the tooling parse "
            "'2026-08', 'August 2026', '2026-08-01', 'Q3 2026', "
            "'next minor release'? Without a constraint, any date "
            "written here is allowed and downstream parsing will fail "
            "inconsistently.\n\n"
            "FIX DIRECTION: specify ISO 8601 (YYYY-MM-DD) as the "
            "canonical format. Add a build invariant: REVIEW_DATE lines "
            "must match `^REVIEW_DATE: \\d{4}-\\d{2}-\\d{2}$` or be "
            "absent. CI greps and fails on malformed dates.\n\n"
            "Related to DOCS-KANJI-002 — once the file template is "
            "concrete and CI parses the format, REVIEW_DATE parsing "
            "must be deterministic so re-evaluation cadence isn't lost "
            "to format ambiguity."
        ),
        "Minor", "P4", "Open"
    ),

    # ============== n5_vocab_whitelist_README.md (4) ==============
    (
        "DOCS-VOCAB-001 — Stale vocab.json entry count: README claims 1041 entries 'as of 2026-05-04 (after v1.12.8)' but actual corpus at v1.15.5+ is 995 entries",
        (
            "FILE: data/n5_vocab_whitelist_README.md (\"Relationship to "
            "data/vocab.json\" section).\n\n"
            "CURRENT TEXT:\n"
            "  'As of 2026-05-04 (after v1.12.8), the whitelist and "
            "vocab.json are fully aligned: every one of the 969 whitelist "
            "tokens matches a form or reading value in some vocab.json "
            "entry.\n"
            "  - vocab.json:  1041 structured catalog entries…'\n\n"
            "ACTUAL STATE: corpus state recorded in prior session work "
            "(v1.15.5 era) shows vocab.json at 995 entries — a 46-entry "
            "difference from the README's claim. This is consistent with "
            "the VOCAB-001..006 cleanup pass that removed nonsense/dupe "
            "entries.\n\n"
            "Three possible explanations:\n"
            "  (a) Between v1.12.8 (1041 entries) and v1.15.5 (995 entries), "
            "46 vocab.json entries were removed (e.g., VOCAB-001 "
            "nonsense-entry cleanup removed 19 entries; over-duplicated-"
            "entry merges removed others). README simply wasn't updated.\n"
            "  (b) The '1041' was a snapshot count from v1.12.8 and the "
            "actual current count is different in either direction — "
            "README is generally stale.\n"
            "  (c) The README's vocab.json is a different file from "
            "data/vocab.json (unlikely but possible if there are multiple "
            "vocab artifacts).\n\n"
            "Most likely (a) — the cleanup work documented in session "
            "summaries (VOCAB-001..006) removed entries, and the README "
            "wasn't refreshed.\n\n"
            "FIX DIRECTION:\n"
            "  1. Recount the current vocab.json (`jq '. | length' "
            "data/vocab.json`) and update the README's entry count.\n"
            "  2. Recount the current n5_vocab_whitelist.json unique tokens "
            "and update the whitelist count.\n"
            "  3. Re-verify the drift = 0 claim by running the alignment "
            "check against the current files.\n"
            "  4. Add a 'Last verified' date to the README (e.g., 'Last "
            "verified against vocab.json and whitelist: 2026-05-21 / "
            "v1.15.5') so future drift is easier to spot.\n"
            "  5. Update the math in the 'size difference' paragraph: "
            "currently '(1041 entries vs 969 unique tokens) reflects…' — "
            "the new numbers and the enumerated explanations need to be "
            "re-reconciled (see also DOCS-VOCAB-004).\n\n"
            "Related to the cross-doc convention proposed at the end of "
            "the audit (Last-verified-against / corpus-version header)."
        ),
        "Major", "P2", "Open"
    ),
    (
        "DOCS-VOCAB-002 — README references data/questions.json as a lint target; that file doesn't exist in current corpus (questions live in per-paper files)",
        (
            "FILE: data/n5_vocab_whitelist_README.md.\n\n"
            "TWO STALE REFERENCES:\n"
            "  'tools/lint_content.py reads whitelist as an allowlist for "
            "the vocab-scope check across grammar.json + questions.json.'\n\n"
            "  'The whitelist serves as the recognition allowlist consumed "
            "by tools/lint_content.py when checking that no out-of-N5-scope "
            "vocabulary appears in data/grammar.json or data/questions.json.'\n\n"
            "ACTUAL CORPUS STRUCTURE (per audited file set):\n"
            "  data/grammar.json        ✓ exists\n"
            "  data/vocab.json          ✓ exists\n"
            "  data/kanji.json          ✓ exists\n"
            "  data/reading.json        ✓ exists\n"
            "  data/listening.json      ✓ exists\n"
            "  data/version.json        ✓ exists\n"
            "  data/paper-{1..7}.json   ✓ exists (per-paper Q-batches, "
            "cycled across bunpou/dokkai/goi/moji categories)\n"
            "  data/questions.json      ✗ NOT in the audited file set\n\n"
            "TWO INTERPRETATIONS:\n"
            "  (a) 'questions.json' is an obsolete name; the corpus was "
            "migrated to per-paper files at some point and the README "
            "wasn't updated. The lint tool now reads paper-{1..7}.json "
            "instead.\n"
            "  (b) 'questions.json' still exists as an aggregator file but "
            "wasn't audited in this session because the audits focus on "
            "per-paper files.\n\n"
            "If (a): the README is stale. The lint target should be "
            "enumerated as 'paper-{1..7}.json across all four categories "
            "(bunpou/dokkai/goi/moji) covering Q1-Q100/102 per category'.\n\n"
            "If (b): the README is correct in concept but the audit "
            "pipeline never touched questions.json, which raises the "
            "question of whether anyone is verifying its content matches "
            "the per-paper files. Cross-file drift risk between "
            "questions.json and paper-{1..7}.json is the same shape as "
            "DOKKAI-001 (passage_text duplicated between question-level "
            "and passage-level).\n\n"
            "FIX DIRECTION: confirm which is true (does data/questions.json "
            "exist?) and update the README to match. If (a), enumerate the "
            "correct lint targets. If (b), document the relationship "
            "between questions.json and paper-{1..7}.json (which is "
            "canonical? when do they diverge?)."
        ),
        "Major", "P2", "Open"
    ),
    (
        "DOCS-VOCAB-003 — README claims KnowledgeBank deleted 2026-05-14, but 28 paper files still carry source_file: KnowledgeBank/*.md references",
        (
            "FILE: data/n5_vocab_whitelist_README.md (first paragraph) + "
            "data/papers/*/paper-{1..7}.json source_file fields.\n\n"
            "README CLAIM:\n"
            "  'The file was originally generated from "
            "KnowledgeBank/vocabulary_n5.md (the now-deleted scope-"
            "reference directory; merged into "
            "docs/N5-syllabus-methodology.md + data/ on 2026-05-14 as a "
            "single source of truth).'\n\n"
            "ACTUAL PAPER FILE CONTENT:\n"
            "Every paper file currently carries:\n"
            "  'source_file': 'KnowledgeBank/grammar_questions_n5.md'   (bunpou)\n"
            "  'source_file': 'KnowledgeBank/dokkai_questions_n5.md'   (dokkai)\n"
            "  'source_file': 'KnowledgeBank/goi_questions_n5.md'      (goi)\n"
            "  'source_file': 'KnowledgeBank/moji_questions_n5.md'     (moji)\n\n"
            "If KnowledgeBank/ was actually deleted on 2026-05-14, these "
            "source_file references are now pointing to nonexistent paths. "
            "Either:\n\n"
            "  (a) KnowledgeBank/ is partially deleted — vocabulary_n5.md "
            "was merged but the other four files (*_questions_n5.md) "
            "still exist. The README's 'now-deleted scope-reference "
            "directory' overstates the deletion.\n"
            "  (b) KnowledgeBank/ is fully deleted and the source_file "
            "references on all 28 paper files (4 categories × 7 papers) "
            "are now stale, pointing to nothing.\n\n"
            "If (a): the README is misleading. The directory wasn't "
            "deleted; only one file inside was merged. Reword:\n"
            "  'The file was originally generated from "
            "KnowledgeBank/vocabulary_n5.md (which was merged into "
            "docs/N5-syllabus-methodology.md + data/ on 2026-05-14; "
            "other files under KnowledgeBank/, such as the "
            "*_questions_n5.md source files for the per-category paper "
            "sets, remain in place).'\n\n"
            "If (b): the README is correct but the paper files are wrong "
            "— their source_file fields are now broken references. Either "
            "restore the KnowledgeBank/ directory or update all 28 paper "
            "files to point to the new source-of-truth location.\n\n"
            "FIX DIRECTION: confirm which is true, then either reword the "
            "README (case a) or update the 28 source_file references on "
            "the paper files (case b). Same drift-detection class as "
            "DOKKAI-001 / JA-117 cross-file-reference resolution."
        ),
        "Major", "P2", "Open"
    ),
    (
        "DOCS-VOCAB-004 — Math doesn't reconcile: 1041 entries minus 969 tokens = 72 surplus, but enumerated explanations sum to only 17; other 55 entries hand-waved as 'cross-referenced by section'",
        (
            "FILE: data/n5_vocab_whitelist_README.md ('size difference' "
            "paragraph).\n\n"
            "CURRENT TEXT:\n"
            "  'The size difference (1041 entries vs 969 unique tokens) "
            "reflects:\n"
            "   - Eight existing multi-form entries…  (=  8 surplus)\n"
            "   - Nine new multi-form entries shipped in v1.12.8… (=  9 surplus)\n"
            "   - vocab.json entries can be cross-referenced by section "
            "(e.g., 人 appears as both pronoun in section 1 and counter "
            "にん in section 9 — same form but different entries; "
            "whitelist counts the form once).'\n\n"
            "MATH GAP:\n"
            "The first two bullets explain 17 surplus entries. The third "
            "bullet is qualitative ('can be cross-referenced') with one "
            "example, not a number. The actual surplus is 72 entries, "
            "leaving 55 entries unexplained.\n\n"
            "For a README that's careful enough to claim 'drift = 0' and "
            "report specific counts, leaving 55/72 of the gap unenumerated "
            "is a credibility hole. A reader can't audit the alignment "
            "claim because the math doesn't add up.\n\n"
            "FIX DIRECTION: when refreshing the counts per DOCS-VOCAB-001, "
            "also itemize all sources of the entry-vs-token gap with actual "
            "numbers. For example:\n\n"
            "  'The size difference (NEW_ENTRIES vs NEW_TOKENS unique "
            "tokens, gap of NEW_GAP) reflects:\n"
            "   - N1 multi-reading entries (one form, multiple readings as "
            "separate tokens): {list with counts, e.g., 何 contributes "
            "なに+なん = 2 tokens for 1 entry, +1 surplus}\n"
            "   - N2 cross-referenced section entries (one form, multiple "
            "part-of-speech roles as separate entries): {list with counts}\n"
            "   - N3 … etc.\n"
            "   - Sum: N1 + N2 + … = NEW_GAP ✓'\n\n"
            "This makes the alignment claim auditable: anyone can run the "
            "entry/token counts and verify the surplus enumeration adds up."
        ),
        "Minor", "P3", "Open"
    ),
]


def main() -> None:
    today_str = date.today().strftime("%Y_%m_%d")
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{today_str}_docs_bugs")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)
    ws = wb["User Reported Bugs"]
    existing = set()
    for r in range(4, ws.max_row + 1):
        t = ws.cell(row=r, column=4).value or ""
        if t:
            existing.add(str(t).split("—")[0].strip())

    written, skipped = 0, 0
    today_str_iso = date.today().isoformat()
    for title, description, severity, priority, status in BUGS:
        prefix = title.split("—")[0].strip()
        if prefix in existing:
            print(f"  SKIP (already exists): {prefix}")
            skipped += 1
            continue
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=f'="BUG-"&TEXT(ROW()-3,"000")')
        ws.cell(row=new_row, column=2, value=today_str_iso)
        ws.cell(row=new_row, column=3, value=REPORTER)
        ws.cell(row=new_row, column=4, value=title)
        ws.cell(row=new_row, column=5, value=description)
        ws.cell(row=new_row, column=6, value=severity)
        ws.cell(row=new_row, column=7, value=priority)
        ws.cell(row=new_row, column=8, value=status)
        ws.cell(row=new_row, column=9, value="")
        ws.cell(row=new_row, column=10, value="")
        for c in range(1, 11):
            ws.cell(row=new_row, column=c).alignment = WRAP
        max_len = max(len(str(v) or "") for v in [title, description])
        ws.row_dimensions[new_row].height = max(45, min(320, max_len // 5))
        print(f"  WROTE row {new_row}: {prefix}")
        written += 1

    wb.save(WORKBOOK)
    print()
    print(f"Result: {written} new bug rows written, {skipped} skipped")


if __name__ == "__main__":
    main()
