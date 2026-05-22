"""Back-fill Fix Commit hashes for NTR-001..013 (BUG-161..173)."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_backfill_ntr"
if not os.path.exists(BAK): shutil.copy2(XLSX, BAK)

# Map row → commit hash of the fix
COMMITS = {
    164: "578b56f",  # NTR-001 — vocab examples + JA-150
    165: "eda9441",  # NTR-002 — n5-045 deprecated
    166: "eda9441",  # NTR-003 — gloss flip
    167: "eda9441",  # NTR-004 — あなた usage warning
    168: "9e0925c",  # NTR-005 — おはし section
    169: "9e0925c",  # NTR-006 — えいが section
    170: "9e0925c",  # NTR-007 — 三 mnemonic
    171: "9e0925c",  # NTR-008 — pitch-accent flags
    172: "eda9441",  # NTR-009 — Hindi sync (paired with NTR-003)
    173: "adac786",  # NTR-010 — q-0226 note
    174: "adac786",  # NTR-011 — collocations rename
    175: "adac786",  # NTR-012 — 七 reading_rule
    176: "adac786",  # NTR-013 — counter cleanup
}

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
for row, h in COMMITS.items():
    # Final commit ce1c195 added the Rule 4/5 docs; the actual data fix is the per-row commit above
    ws.cell(row=row, column=10, value=f"{h} (Rule 4/5 propagation: ce1c195 + submodule 1cef80b)")
    print(f"  R{row}: commit = {h}")
wb.save(XLSX)
