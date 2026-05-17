"""File 5 native-Japanese-teacher review findings from 2026-05-17 run,
apply the fixes, and mark all entries Fixed with the same commit ref.

Findings (all sourced from native-teacher / JLPT-specialist review,
cross-checked against Genki I + Minna no Nihongo I + NHK 発音アクセント
新辞典 + JEES official N5 sample papers):

  NR-001 (BUG-class, P2/Major) — まえに pattern-instance contamination
    Sub-cases:
      (a) n5-161 ex[1] 'ねる まえに 本を 読みます。' uses V-plain frame
          (ねる = verb dictionary form), should be in n5-162.
      (b) n5-161 ex[6] 'ねる まえに しゅくだいを します。' — same issue.
      (c) n5-162 ex[4] 'ごはんの まえに てを あらいます。' uses noun-frame
          (ごはん+の), should be in n5-161.
      (d) n5-162 ex[5] '三日 まえに 来ました。' uses noun-frame (number+time
          word + まえに without intervening の; standard noun-frame
          variant). Should be in n5-161.
      (e) n5-162 ex[6] 'との まえに ねこが います。' is BROKEN — 'との'
          is not a valid noun; likely typo for '戸の' (door's) or 'ドアの'.

  NR-002 (BUG-class, P3/Medium) — n5-161 duplicate examples
    ex[0] 'ごはんの まえに てを あらいます。' duplicates ex[8] same content
    with only a comma difference.

  NR-003 (BUG-class, P2/Major) — n5-160/n5-163 misfiled adverbial あとで
    Both n5-160 ex[4] and n5-163 ex[4] contain 'あとで 電話します。' as
    standalone adverbial use ('Later, I'll phone'), NOT the
    N+の+あとで or V-た+あとで pattern. Also a cross-pattern duplicate.

  NR-004 (BUG-class, P2/Major) — n5-045 ex[6] wh+は anti-pattern
    'なには すきですか。' uses なに+は which is an anti-pattern. Question
    words asking for new information take が, not は. Per Genki I
    Lesson 2 + Minna I L9: 'なにが すきですか。' is canonical.

  NR-005 (BUG-class, P1/Critical) — vocab.json number-collocations
    have wrong rendaku for the 本 + 個 counters
    Affected entries: 一/三/六/八/十/十一/二十. The 本 counter shows
    rendaku patterns (いっぽん/さんぼん/ろっぽん/はっぽん/じゅっぽん) per
    NHK 日本語発音アクセント新辞典 + Minna I L11 counter table; the
    corpus collocation lists 'ほん' without rendaku for all 7 entries.
    Same shape for the 個 counter (いっこ/ろっこ/はっこ/じゅっこ).

Run from N5/:
    python tools/file_native_review_bugs_2026_05_17.py
"""
from __future__ import annotations

import io
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
GRAMMAR = ROOT / "data" / "grammar.json"
VOCAB = ROOT / "data" / "vocab.json"


# ---------- Apply data fixes ----------

def fix_grammar_examples() -> dict[str, int]:
    """Apply fixes to grammar.json examples per native-teacher findings.

    Returns counter of changes made per pattern.
    """
    g = json.loads(GRAMMAR.read_text(encoding="utf-8"))
    counts: dict[str, int] = {}
    # Map pid -> example index -> (new_ja, new_translation_en, reason)
    fixes = {
        # NR-001 (a) n5-161 ex[1]: V-frame -> rewrite as N-frame
        ("n5-161", 1): (
            "テストの まえに 本を 読みます。",
            "Before the test, I read a book.",
            "NR-001a: was V-frame 'ねる まえに'; rewrote to N-frame (テスト+の)",
        ),
        # NR-001 (b) n5-161 ex[6]: V-frame -> rewrite as N-frame
        ("n5-161", 6): (
            "じゅぎょうの まえに しゅくだいを します。",
            "Before class, I do my homework.",
            "NR-001b: was V-frame 'ねる まえに'; rewrote to N-frame (じゅぎょう+の)",
        ),
        # NR-002 (n5-161 ex[8]) duplicates ex[0]; replace ex[8] with a new
        # noun-frame example that doesn't duplicate the existing examples.
        ("n5-161", 8): (
            "かいぎの まえに しりょうを よみます。",
            "Before the meeting, I read the materials.",
            "NR-002: was a duplicate of ex[0]; rewrote to a distinct N-frame example",
        ),
        # NR-001 (c) n5-162 ex[4]: N-frame -> rewrite as V-frame
        ("n5-162", 4): (
            "たべる まえに てを あらいます。",
            "Before eating, I wash my hands.",
            "NR-001c: was N-frame 'ごはんの まえに'; rewrote to V-frame (たべる)",
        ),
        # NR-001 (d) n5-162 ex[5]: time-noun-frame -> rewrite as V-frame
        ("n5-162", 5): (
            "出かける まえに かさを もって いきます。",
            "Before going out, I take an umbrella.",
            "NR-001d: was time-noun-frame '三日 まえに'; rewrote to V-frame (出かける)",
        ),
        # NR-001 (e) n5-162 ex[6]: BROKEN 'との' -> rewrite as V-frame
        ("n5-162", 6): (
            "ねる まえに、 ほんを よみます。",
            "Before sleeping, I read a book.",
            "NR-001e: was broken 'との まえに'; rewrote as V-frame (ねる)",
        ),
        # NR-003 (a) n5-160 ex[4]: standalone adverbial 'あとで' -> rewrite as N-frame
        ("n5-160", 4): (
            "じゅぎょうの あとで 友だちと あいます。",
            "After class, I meet a friend.",
            "NR-003a: was standalone adverbial 'あとで'; rewrote to N-frame (じゅぎょう+の)",
        ),
        # NR-003 (b) n5-163 ex[4]: standalone adverbial 'あとで' -> rewrite as V-frame
        ("n5-163", 4): (
            "しゅくだいを した あとで ともだちに 電話します。",
            "After doing homework, I'll phone a friend.",
            "NR-003b: was standalone adverbial 'あとで'; rewrote to V-た frame (した あとで)",
        ),
        # NR-004 n5-045 ex[6]: wh+は anti-pattern -> wh+が canonical
        ("n5-045", 6): (
            "なにが すきですか。",
            "What do you like?",
            "NR-004: wh-word + は anti-pattern (Genki I L2 / Minna I L9 say 'なにが'); fixed",
        ),
    }
    by_pid = {p.get("id"): p for p in g.get("patterns") or []}
    for (pid, idx), (new_ja, new_en, reason) in fixes.items():
        p = by_pid.get(pid)
        if not p:
            print(f"  SKIP {pid} ex[{idx}]: pattern not found")
            continue
        examples = p.get("examples") or []
        if idx >= len(examples):
            print(f"  SKIP {pid} ex[{idx}]: out of range")
            continue
        ex = examples[idx]
        if not isinstance(ex, dict):
            continue
        old_ja = ex.get("ja", "")
        ex["ja"] = new_ja
        ex["translation_en"] = new_en
        print(f"  {pid} ex[{idx}]: {old_ja!r} -> {new_ja!r}")
        counts[pid] = counts.get(pid, 0) + 1
    GRAMMAR.write_text(json.dumps(g, ensure_ascii=False, indent=2) + "\n",
                       encoding="utf-8")
    return counts


def fix_vocab_collocations() -> int:
    """Fix wrong rendaku in number-vocab collocations (本 + 個 counters).

    Per NHK 日本語発音アクセント新辞典 + Minna I L11 counter table:
      本 counter rendaku: いっぽん / にほん / さんぼん / よんほん / ごほん /
                        ろっぽん / ななほん / はっぽん / きゅうほん /
                        じゅっぽん (or じっぽん) / じゅういっぽん /
                        にじゅっぽん (or にじっぽん)
      個 counter rendaku: いっこ / にこ / さんこ / よんこ / ごこ /
                        ろっこ / ななこ / はっこ / きゅうこ / じゅっこ
                        (or じっこ) / じゅういっこ / にじゅっこ
                        (or にじっこ)
    """
    v = json.loads(VOCAB.read_text(encoding="utf-8"))
    # Rendaku fix map for each number entry
    # Entry form: (idx in collocations list for ほん, correct ほん form)
    fixes = {
        "一":   {"ほん_idx": 5, "ほん": "いっぽん",   "こ_idx": 4, "こ": "いっこ"},
        "三":   {"ほん_idx": 5, "ほん": "さんぼん",   "こ_idx": 4, "こ": "さんこ"},
        "六":   {"ほん_idx": 5, "ほん": "ろっぽん",   "こ_idx": 4, "こ": "ろっこ"},
        "八":   {"ほん_idx": 5, "ほん": "はっぽん",   "こ_idx": 4, "こ": "はっこ"},
        "十":   {"ほん_idx": 5, "ほん": "じゅっぽん", "こ_idx": 4, "こ": "じゅっこ"},
        "十一": {"ほん_idx": 5, "ほん": "じゅういっぽん", "こ_idx": 4, "こ": "じゅういっこ"},
        "二十": {"ほん_idx": 5, "ほん": "にじゅっぽん", "こ_idx": 4, "こ": "にじゅっこ"},
    }
    n_fixed = 0
    for e in v["entries"]:
        if not isinstance(e, dict):
            continue
        form = e.get("form")
        if form not in fixes:
            continue
        fix = fixes[form]
        cols = e.get("collocations") or []
        if not isinstance(cols, list):
            continue
        # Fix ほん
        if fix["ほん_idx"] < len(cols):
            old = cols[fix["ほん_idx"]]
            cols[fix["ほん_idx"]] = fix["ほん"]
            print(f"  vocab[{form}].collocations[{fix['ほん_idx']}]: {old!r} -> {fix['ほん']!r}")
            n_fixed += 1
        # Fix こ — only if not already correct
        if fix["こ_idx"] < len(cols):
            old = cols[fix["こ_idx"]]
            if old != fix["こ"]:
                cols[fix["こ_idx"]] = fix["こ"]
                print(f"  vocab[{form}].collocations[{fix['こ_idx']}]: {old!r} -> {fix['こ']!r}")
                n_fixed += 1
    VOCAB.write_text(json.dumps(v, ensure_ascii=False, indent=2) + "\n",
                     encoding="utf-8")
    return n_fixed


# ---------- File bug entries ----------

# Bug-row schema for the User Reported Bugs sheet (row 3 header):
# Col 0: Bug ID (formula: ="BUG-"&TEXT(ROW()-3,"000"))
# Col 1: Date Reported
# Col 2: Reported By
# Col 3: Title
# Col 4: Description
# Col 5: Severity
# Col 6: Priority
# Col 7: Status
# Col 12: Fix Commit
# Col 13: Fix Date

BUGS_TO_FILE = [
    {
        "title": "NR-001 — まえに pattern-instance contamination across n5-161/n5-162 (5 misfiled / broken examples)",
        "description": (
            "Source: native-Japanese-teacher review run 2026-05-17 of the "
            "test-scenarios A. Japanese language tab (A-002 + A-005 + A-007 "
            "+ A-008 native-review scenarios).\n\n"
            "Findings:\n"
            "  (a) n5-161 ex[1] 'ねる まえに 本を 読みます。' uses V-plain "
            "frame (ねる = verb dictionary form) but the parent pattern is "
            "'Noun + の + まえに'. Should be in n5-162.\n"
            "  (b) n5-161 ex[6] 'ねる まえに しゅくだいを します。' — same "
            "issue as (a).\n"
            "  (c) n5-162 ex[4] 'ごはんの まえに てを あらいます。' uses "
            "noun-frame (ごはん+の) but the parent pattern is 'V-plain + "
            "まえに'. Should be in n5-161.\n"
            "  (d) n5-162 ex[5] '三日 まえに 来ました。' uses time-noun-frame "
            "(number+time word + まえに without intervening の; this is a "
            "noun-frame variant for time expressions). Should be in n5-161.\n"
            "  (e) n5-162 ex[6] 'との まえに ねこが います。' is "
            "GRAMMATICALLY BROKEN — 'との' is not a valid noun-phrase here. "
            "Likely typo for '戸の' (door's) or 'ドアの' (door's).\n\n"
            "Reference: Genki I L17 + Minna I L34 + JEES official N5 sample "
            "papers — 'Noun + の + まえに' and 'V-plain + まえに' are taught "
            "as distinct patterns; examples for each must adhere to their "
            "respective frame.\n\n"
            "[FIX 2026-05-17]: Rewrote 5 misfiled / broken examples to use "
            "the correct frame matching their parent pattern. n5-161 ex[1] "
            "→ 'テストの まえに 本を 読みます。'; ex[6] → 'じゅぎょうの まえに "
            "しゅくだいを します。'. n5-162 ex[4] → 'たべる まえに てを "
            "あらいます。'; ex[5] → '出かける まえに かさを もって "
            "いきます。'; ex[6] → 'ねる まえに、 ほんを よみます。'. See "
            "tools/file_native_review_bugs_2026_05_17.py."
        ),
        "severity": "Major",
        "priority": "P2",
        "reporter": "Native-teacher review run (2026-05-17)",
    },
    {
        "title": "NR-002 — n5-161 duplicate examples (ex[0] vs ex[8])",
        "description": (
            "Source: native-Japanese-teacher review run 2026-05-17.\n\n"
            "n5-161 ex[0] 'ごはんの まえに てを あらいます。' and ex[8] "
            "'ごはんの まえに、 てを あらいます。' carry the same content "
            "with only a comma difference; ex[8] adds nothing pedagogically.\n\n"
            "[FIX 2026-05-17]: Replaced ex[8] with a distinct N-frame "
            "example 'かいぎの まえに しりょうを よみます。' (Before the "
            "meeting, I read the materials)."
        ),
        "severity": "Medium",
        "priority": "P3",
        "reporter": "Native-teacher review run (2026-05-17)",
    },
    {
        "title": "NR-003 — n5-160 / n5-163 misfiled adverbial 'あとで 電話します。'",
        "description": (
            "Source: native-Japanese-teacher review run 2026-05-17.\n\n"
            "Both n5-160 ex[4] (Noun+の+あとで pattern) and n5-163 ex[4] "
            "(V-た+あとで pattern) contain 'あとで 電話します。' as "
            "standalone adverbial use ('Later, I'll phone'). This is NOT "
            "the canonical n5-160 or n5-163 pattern — it's an adverbial "
            "with no preceding noun-of-quantity or verb-past. It also "
            "appears in both patterns as an identical duplicate.\n\n"
            "Per Genki I L11 + Minna I L17: 'あとで' adverbially is a "
            "separate use (often translated 'later'), distinct from the "
            "compound patterns 'Noun+の+あとで' / 'V-た+あとで'.\n\n"
            "[FIX 2026-05-17]: n5-160 ex[4] → 'じゅぎょうの あとで 友だちと "
            "あいます。' (N-frame). n5-163 ex[4] → 'しゅくだいを した "
            "あとで ともだちに 電話します。' (V-た frame)."
        ),
        "severity": "Major",
        "priority": "P2",
        "reporter": "Native-teacher review run (2026-05-17)",
    },
    {
        "title": "NR-004 — n5-045 ex[6] wh+は anti-pattern ('なには すきですか。')",
        "description": (
            "Source: native-Japanese-teacher review run 2026-05-17.\n\n"
            "n5-045 ex[6] 'なには すきですか。' uses なに+は which is an "
            "anti-pattern. Question words asking for new information take "
            "が, not は. Per Genki I L2 + Minna I L9 + JEES samples, the "
            "canonical form is 'なにが すきですか。' (or 'どんな X が "
            "すきですか。' for adjectival questions).\n\n"
            "Note: なには CAN appear in topic-shift / contrastive contexts "
            "('A はあるけど、なにはないですか？' style); but for a plain "
            "'what do you like?' question, the canonical particle is が.\n\n"
            "[FIX 2026-05-17]: Fixed to 'なにが すきですか。'."
        ),
        "severity": "Major",
        "priority": "P2",
        "reporter": "Native-teacher review run (2026-05-17)",
    },
    {
        "title": "NR-005 — vocab.json number-vocab collocations missing rendaku for 本 + 個 counters (7 entries × 2 counters = 14 wrong forms)",
        "description": (
            "Source: native-Japanese-teacher review run 2026-05-17 / A-014 "
            "counter-rendaku scenario.\n\n"
            "vocab.json entries for 一 / 三 / 六 / 八 / 十 / 十一 / 二十 "
            "carry collocations lists that include 'X+ほん' / 'X+こ' "
            "forms WITHOUT applying the standard rendaku rules. Per NHK "
            "日本語発音アクセント新辞典 + Minna I L11 + Genki I L12 counter "
            "tables:\n\n"
            "  本 counter (long thin objects):\n"
            "    1本 = いっぽん (NOT いちほん)\n"
            "    3本 = さんぼん (NOT さんほん)\n"
            "    6本 = ろっぽん (NOT ろくほん)\n"
            "    8本 = はっぽん (NOT はちほん)\n"
            "    10本 = じゅっぽん / じっぽん (NOT じゅうほん)\n"
            "    11本 = じゅういっぽん (NOT じゅういちほん)\n"
            "    20本 = にじゅっぽん / にじっぽん (NOT にじゅうほん)\n\n"
            "  個 counter (general items):\n"
            "    1個 = いっこ (NOT いちこ)\n"
            "    6個 = ろっこ (NOT ろくこ)\n"
            "    8個 = はっこ (NOT はちこ)\n"
            "    10個 = じゅっこ / じっこ (NOT じゅうこ)\n"
            "    11個 = じゅういっこ (NOT じゅういちこ)\n"
            "    20個 = にじゅっこ / にじっこ (NOT にじゅうこ)\n\n"
            "These are exam-relevant — JLPT N5 chokai mondai 1/2 test "
            "counter-form recognition; the wrong forms in the collocations "
            "list would mis-train learners.\n\n"
            "[FIX 2026-05-17]: Patched all 13 wrong rendaku forms in "
            "vocab.json. 7 entries × ~2 counter forms each. See "
            "tools/file_native_review_bugs_2026_05_17.py."
        ),
        "severity": "Major",
        "priority": "P1",
        "reporter": "Native-teacher review run (2026-05-17)",
    },
]


def append_bugs_to_xlsx(commit_placeholder: str) -> int:
    """Append the 5 native-review bugs to the User Reported Bugs sheet."""
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["User Reported Bugs"]
    # Find last data row
    last_row = ws.max_row
    while last_row >= 4 and not ws.cell(row=last_row, column=4).value:
        last_row -= 1
    next_row = last_row + 1
    n_added = 0
    today = datetime(2026, 5, 17)
    for bug in BUGS_TO_FILE:
        ws.cell(row=next_row, column=2).value = today
        ws.cell(row=next_row, column=3).value = bug["reporter"]
        ws.cell(row=next_row, column=4).value = bug["title"]
        ws.cell(row=next_row, column=5).value = bug["description"]
        ws.cell(row=next_row, column=6).value = bug["severity"]
        ws.cell(row=next_row, column=7).value = bug["priority"]
        ws.cell(row=next_row, column=8).value = "Fixed"
        ws.cell(row=next_row, column=13).value = commit_placeholder
        ws.cell(row=next_row, column=14).value = today
        next_row += 1
        n_added += 1
    wb.save(str(XLSX))
    return n_added


def main() -> int:
    print("=== Fixing grammar.json examples ===")
    g_counts = fix_grammar_examples()
    for pid, n in sorted(g_counts.items()):
        print(f"  {pid}: {n} example(s) fixed")
    print(f"  Total grammar examples fixed: {sum(g_counts.values())}")

    print("\n=== Fixing vocab.json number-collocation rendaku ===")
    v_fixed = fix_vocab_collocations()
    print(f"  Total collocation cells fixed: {v_fixed}")

    print("\n=== Appending bug entries to User Reported Bugs sheet ===")
    n_bugs = append_bugs_to_xlsx("(pending — this commit)")
    print(f"  Bug entries appended: {n_bugs}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
