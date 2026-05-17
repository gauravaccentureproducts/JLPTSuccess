"""Update test-scenarios xlsx with new scenarios + Notes linkage based
on the 40 bugs found / fixed across BUG-001..BUG-040.

Adds 12 new scenarios mapping to the JA-NN invariants added during
the BUG-001..BUG-024 / BUG-038 cycle. Each new scenario references
both the originating bug AND the CI invariant that now guards
against regression — so the test plan stays in sync with what the
script actually enforces.

Adds Notes-column linkage on 10 existing scenarios that partially
overlap with the bug classes (e.g., A-008 verb-class accuracy is
adjacent to BUG-016 transitivity coverage).

Idempotent — re-running on already-updated rows is a no-op.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"


# ============================================================
# New scenarios — one per bug class with a wired or planned CI invariant.
# ============================================================
NEW_SCENARIOS = [
    # ---- A tab additions ----
    {
        "tab": "A. Japanese language", "id": "A-032",
        "subcat": "Cross-pattern contamination",
        "persp": "1",
        "scenario": "Cross-pattern explanation_en uniqueness (Levenshtein) — catches BUG-003-class contamination",
        "steps": (
            "1. For each pattern P, compute Levenshtein ratio of P.explanation_en against every OTHER pattern's explanation_en.\n"
            "2. Flag pairs with ratio ≥0.85.\n"
            "3. Manual triage on flagged pairs — verify each is intentional similarity, not contamination.\n"
            "4. Reference: BUG-003 (n5-098 had n5-099's explanation verbatim) + the F.17.1 lesson in procedure manual."
        ),
        "expected": "0 unintentional explanation-similarity hits ≥0.85 across the 178 patterns. Flagged pairs documented as intentional contrast pairs OR rewritten.",
        "priority": "P3", "severity": "Major", "test_type": "Auto+Manual",
        "tools": "Custom Python (Levenshtein cross-similarity scan)",
        "owner": "Native Japanese teacher + SLA pedagogy researcher",
        "effort": "4h",
        "notes": "Planned CI invariant JA-91 (currently as Phase-0 regression in prompts/N5Improvement.txt). Promotion to hard gate queued. Covers BUG-003.",
    },
    {
        "tab": "A. Japanese language", "id": "A-033",
        "subcat": "Cross-pattern contamination",
        "persp": "1",
        "scenario": "Per-example JA-EN content-word overlap — catches BUG-005-class translation_en stuck on one string",
        "steps": (
            "1. For each example, extract content words from JA (kanji + katakana tokens) and from EN (lowercased multi-char tokens).\n"
            "2. Compute overlap; flag patterns with ≥3 zero-overlap examples.\n"
            "3. LLM-as-judge pass on flagged examples (cheap): 'Does this EN render this JA?' YES/NO.\n"
            "4. Reference: BUG-005 (n5-166 ex[5] EN = n5-096 ex[2]'s text) + F.17.1."
        ),
        "expected": "0 patterns with ≥3 zero-overlap examples post-LLM-judge.",
        "priority": "P3", "severity": "Major", "test_type": "Auto+Manual",
        "tools": "Custom Python (content-word overlap heuristic) + LLM-judge pass",
        "owner": "Native Japanese teacher + LLM judge",
        "effort": "1d",
        "notes": "Planned CI invariant JA-92. Covers BUG-005.",
    },
    {
        "tab": "A. Japanese language", "id": "A-034",
        "subcat": "Pattern-instance contamination",
        "persp": "1",
        "scenario": "Every grammar example contains ≥1 pattern-defining surface marker (planned JA-94)",
        "steps": (
            "1. Maintain `data/pattern_markers.json` — regex-friendly marker list per pattern.\n"
            "2. For each example filed under pattern P, assert ≥1 marker from P appears in the JA text (after stripping bunsetsu spaces).\n"
            "3. Late-Nx patterns (late_n5, deferred_to_n4) audit first — those are the highest-risk slots.\n"
            "4. Reference: BUG-006 (10 examples filed under wrong pattern) + F.17.3."
        ),
        "expected": "0 examples lacking a pattern-defining marker.",
        "priority": "P3", "severity": "Major", "test_type": "Auto",
        "tools": "Custom Python (`data/pattern_markers.json` + regex scan)",
        "owner": "Data engineer + native Japanese teacher (for marker authoring)",
        "effort": "1d",
        "notes": "Planned CI invariant JA-94 (Phase-0 regression block in prompts/N5Improvement.txt today). Promotion gated on pattern_markers.json authoring. Covers BUG-006.",
    },
    {
        "tab": "A. Japanese language", "id": "A-035",
        "subcat": "Folk-linguistic terminology",
        "persp": "1",
        "scenario": "common_mistakes WHY rationales use canonical grammar terminology (Genki / Tobira / A Dictionary of Basic Japanese Grammar) — not folk shortcuts",
        "steps": (
            "1. Regex-scan common_mistakes.why for folk-linguistic terms: 'intransitive' applied to encounter verbs (会う/挨拶/出会う); 'subject vs topic' used as synonyms; 'auxiliary' or 'passive' applied to constructions that aren't.\n"
            "2. Manual review of each hit against canonical reference.\n"
            "3. Reference: BUG-008 (n5-004 cm[0] folk-linguistic 'intransitive' claim) + F.17.5."
        ),
        "expected": "0 folk-linguistic terminology hits in common_mistakes WHY fields.",
        "priority": "P3", "severity": "Major", "test_type": "Auto+Manual",
        "tools": "Custom Python (regex scan) + manual review",
        "owner": "Native Japanese teacher",
        "effort": "4h",
        "notes": "Phase-0 regression block A53 in prompts/N5Improvement.txt. Covers BUG-008.",
    },
    {
        "tab": "A. Japanese language", "id": "A-036",
        "subcat": "Particle-pattern alignment",
        "persp": "1",
        "scenario": "Particle-introducing patterns: every example JA contains the canonical particle (planned JA-95)",
        "steps": (
            "1. For patterns whose `pattern` field names a specific particle (が, を, に, で, と, へ, から, まで), require ≥1 occurrence of that particle in each example JA.\n"
            "2. For obligatory-が contexts (interrogative-subject, identificational, existential), additionally require an interrogative head OR existential predicate.\n"
            "3. Reference: BUG-009 (n5-003 ex[6] used は in a が-pattern) + F.17.6."
        ),
        "expected": "0 examples missing their pattern's canonical particle.",
        "priority": "P3", "severity": "Major", "test_type": "Auto",
        "tools": "Custom Python (per-pattern particle-presence assertion)",
        "owner": "Data engineer + native Japanese teacher",
        "effort": "4h",
        "notes": "Planned CI invariant JA-95 (Phase-0 regression block today). Covers BUG-009.",
    },
    {
        "tab": "A. Japanese language", "id": "A-037",
        "subcat": "Register-variant schema",
        "persp": "1",
        "scenario": "common_mistakes entries marked `kind: \"register_variant\"` use form_a/form_b keys, not wrong/right",
        "steps": (
            "1. JA-64 already enforces required fields by `kind`.\n"
            "2. Audit any entry whose WHY mentions 'both correct' / 'register choice' / 'casual' / 'formal' — must carry `kind: register_variant`.\n"
            "3. Phase-0 regression block A57 (in prompts/N5Improvement.txt) runs the strict-key + WHY-signal check.\n"
            "4. Reference: BUG-007/011/013 → final schema captured in F.19.2."
        ),
        "expected": "0 entries where WHY signals register-variant but kind is absent. 0 entries with kind=register_variant carrying stale wrong/right keys.",
        "priority": "P2", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-64 + Phase-0 A57 in N5Improvement.txt)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "CI invariant JA-64 enforces required fields by kind. Phase-0 A57 catches stale keys + unflagged variants. Covers BUG-007/011/013.",
    },
    {
        "tab": "A. Japanese language", "id": "A-038",
        "subcat": "Template-nonsense vocab examples",
        "persp": "1",
        "scenario": "No vocab example uses bare `<form>が あります/います` template (JA-96)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-96.\n"
            "2. Pass: 0 vocab examples matching the bare-template regex.\n"
            "3. Reference: BUG-014 (19 such examples on time/abstract/location nouns) + F.21.1 authoring rule (classify noun type before applying template)."
        ),
        "expected": "JA-96 PASS — 0 bare-template examples across 998 vocab entries.",
        "priority": "P2", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-96)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by CI invariant JA-96 (added 2026-05-16 from BUG-014). Strict regex match — no false positives expected.",
    },
    {
        "tab": "A. Japanese language", "id": "A-039",
        "subcat": "Verb transitivity coverage",
        "persp": "1",
        "scenario": "Every vocab verb (verb-1/verb-2/verb-3) has `transitivity` in {transitive, intransitive, contact} (JA-98)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-98.\n"
            "2. Pass: 132/132 verbs declared with closed-enum value.\n"
            "3. Spot-check 10 random verbs against JMdict vt/vi tags for accuracy.\n"
            "4. Reference: BUG-016 + F.21.3 (field-coverage gap class)."
        ),
        "expected": "JA-98 PASS — 132/132 verbs classified; spot-check shows ≥9/10 agreement with JMdict.",
        "priority": "P2", "severity": "Major", "test_type": "Auto+Manual",
        "tools": "tools/check_content_integrity.py (JA-98) + manual JMdict cross-check",
        "owner": "Data engineer + native Japanese teacher",
        "effort": "4h",
        "notes": "Locked by JA-98. 'contact' value reserved for encounter verbs per BUG-008 lesson.",
    },
    {
        "tab": "A. Japanese language", "id": "A-040",
        "subcat": "Scope discipline",
        "persp": "2",
        "scenario": "All kanji in vocab.form fields are in N5 whitelist OR explicit exception list (JA-99)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-99.\n"
            "2. Pass: 0 OOS kanji in any vocab.form across 995 entries.\n"
            "3. Reference: BUG-017 (3 OOS kanji 倍/籍/末) + BUG-023 (the inverse — vocab kana-only when kanji IS in scope) + F.22.1."
        ),
        "expected": "JA-99 PASS — 0 OOS kanji in vocab.form.",
        "priority": "P2", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-99)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-99. Complements JA-13 (grammar example kanji-scope) for the vocab surface.",
    },
    {
        "tab": "A. Japanese language", "id": "A-041",
        "subcat": "primary_reading correctness",
        "persp": "1",
        "scenario": "Standalone-kun primary_reading lock for 6 kanji (人/中/外/東/車/国) — kun-yomi at standalone use (JA-102)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-102.\n"
            "2. Pass: 人=ひと, 中=なか, 外=そと, 東=ひがし, 車=くるま, 国=くに.\n"
            "3. Reference: BUG-021 (these 6 had on-yomi as primary despite kun being the standalone form) + F.22.2."
        ),
        "expected": "JA-102 PASS — 6/6 kanji on canonical kun primary_reading.",
        "priority": "P3", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-102)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-102. Defensible on-yomi-primary entries (時/社/駅) kept unchanged.",
    },

    # ---- I tab additions ----
    {
        "tab": "I. Data engineering", "id": "I-017",
        "subcat": "Schema canonicalization",
        "persp": "1",
        "scenario": "vocab.json counter field is canonical (null OR {kanji, reading} dict); counter_register null (JA-97)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-97.\n"
            "2. Pass: 0 entries with non-canonical counter shape.\n"
            "3. Reference: BUG-015 (3 inconsistent types pre-fix) + F.21.2."
        ),
        "expected": "JA-97 PASS — counter shape uniform across 995 vocab entries.",
        "priority": "P2", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-97)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-97. The 16 counter-word metadata entries (一つ/二つ/etc) use the isolated `counter_word_metadata` field, not counter_register.",
    },
    {
        "tab": "I. Data engineering", "id": "I-018",
        "subcat": "Provenance disambiguation",
        "persp": "2",
        "scenario": "review_status from closed enum {ai_quality_reviewed, llm_curated, auto_generated, human_native_reviewed}; provenance field present (JA-35)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-35.\n"
            "2. Pass: 0 entries on retired 'native_reviewed' value (post-BUG-012 migration); every item has a value from the closed enum.\n"
            "3. Audit: every item with review_status carries review_status_provenance.\n"
            "4. Reference: BUG-012 (rename + provenance) + F.20."
        ),
        "expected": "JA-35 PASS; 100% of items also have review_status_provenance.",
        "priority": "P2", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-35) + Phase-0 A58 (retired-value scan)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-35 (extended enum). Phase-0 A58 in N5Improvement.txt catches retired-value regression. Reserved future value `human_native_reviewed` for actual native-human review pass.",
    },
    {
        "tab": "I. Data engineering", "id": "I-019",
        "subcat": "Cross-section dedup",
        "persp": "1",
        "scenario": "Automated subset-gloss duplicate detection in vocab.json — same (form, reading) + same POS + gloss-B ⊂ gloss-A",
        "steps": (
            "1. Group vocab.json entries by (form, reading).\n"
            "2. Skip groups with different POS (legitimate polysemy like いくつ question-word vs counter).\n"
            "3. Flag pairs where one gloss is a strict substring of another.\n"
            "4. Manual triage on hits.\n"
            "5. Reference: BUG-018 (10 hand-curated cases) + BUG-019 (3 missed by the hand-list) + F.21.5 automated detector script."
        ),
        "expected": "0 flagged subset-gloss duplicate pairs (manual triage may reject false positives).",
        "priority": "P3", "severity": "Minor", "test_type": "Auto+Manual",
        "tools": "Custom Python (subset-detector script in F.21.5 of procedure manual)",
        "owner": "Data engineer",
        "effort": "4h",
        "notes": "Phase-0 regression script. The BUG-018 → BUG-019 round-trip taught: hand-curated case lists miss the full population; automated detector catches all cases. Run before every release.",
    },
    {
        "tab": "I. Data engineering", "id": "I-020",
        "subcat": "Cross-file integrity",
        "persp": "3",
        "scenario": "kanji.json compound/example.form == linked vocab.form (strict; JA-100)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-100.\n"
            "2. Pass: for every kanji.json n5_compounds + examples row with a vocab_id, form == linked vocab_form exactly.\n"
            "3. Reference: BUG-020 (OOS-kanji drift; vocab right) + BUG-023 (in-scope kanji; kanji.json right) + F.22.1 'TIGHTENING' lesson (default to strict CI gates)."
        ),
        "expected": "JA-100 PASS — strict form-equality between kanji.json compounds and linked vocab entries.",
        "priority": "P2", "severity": "Major", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-100)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-100 (tightened 2026-05-17 from narrow OOS-only to strict equality after BUG-023 round-trip). Catches both directions of drift.",
    },
    {
        "tab": "I. Data engineering", "id": "I-021",
        "subcat": "Field naming canonicalization",
        "persp": "2",
        "scenario": "kanji.json `examples[]` use `form` field name only, not legacy `lemma` (JA-101)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-101.\n"
            "2. Pass: every example object has `form`, no example has `lemma`.\n"
            "3. Reference: BUG-022 (374 form + 20 lemma + 14 dual pre-fix) + F.22.3."
        ),
        "expected": "JA-101 PASS.",
        "priority": "P3", "severity": "Minor", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-101)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-101 (added 2026-05-17 from BUG-022). Same field-name-divergence class as BUG-015 counter schema.",
    },
    {
        "tab": "I. Data engineering", "id": "I-022",
        "subcat": "Within-corpus uniqueness",
        "persp": "1",
        "scenario": "kanji.json n5_compounds: (form, reading) tuple unique within each kanji entry (JA-103)",
        "steps": (
            "1. Run tools/check_content_integrity.py JA-103.\n"
            "2. Pass: 0 duplicate (form, reading) tuples within any kanji entry.\n"
            "3. Different readings PASS (legitimate polysemy — e.g., 一日 ついたち vs いちにち on the 一 entry).\n"
            "4. Reference: BUG-024 + F.22.5 auto-derived-data freshness lesson."
        ),
        "expected": "JA-103 PASS — 0 duplicate compound rows.",
        "priority": "P3", "severity": "Minor", "test_type": "Auto",
        "tools": "tools/check_content_integrity.py (JA-103)",
        "owner": "Data engineer",
        "effort": "1h",
        "notes": "Locked by JA-103 (added 2026-05-17 from BUG-024). Auto-derived from vocab.json — re-run derivation after any vocab dedup pass.",
    },
    {
        "tab": "I. Data engineering", "id": "I-023",
        "subcat": "Auto-derivation freshness",
        "persp": "2",
        "scenario": "Every auto-derived array (kanji.n5_compounds, kanji.examples) is fresh vs its source (vocab.json last_modified)",
        "steps": (
            "1. (Pending) Add `last_derivation_run_at` timestamp on every auto-derived array.\n"
            "2. Compare against vocab.json `last_modified`.\n"
            "3. Pass: every derivation timestamp ≥ source timestamp.\n"
            "4. Reference: BUG-024 lesson — kanji.json auto-derived data inherited vocab.json's pre-dedup state because the pipeline was never re-run. F.22.5."
        ),
        "expected": "All auto-derived arrays' freshness timestamps ≥ vocab.json last_modified.",
        "priority": "P4", "severity": "Minor", "test_type": "Auto",
        "tools": "Custom Python (timestamp comparison; pending implementation)",
        "owner": "Data engineer",
        "effort": "1d",
        "notes": "Queued / not yet wired. Operational rule today: re-run auto-derivation after every vocab dedup pass. CI enforcement is the future iteration per F.22.5.",
    },
]


# ============================================================
# Notes linkage on existing scenarios (idempotent — append if absent)
# ============================================================
NOTES_LINKAGE = {
    # A tab — partial coverage of bug classes
    ("A. Japanese language", "A-001"): " | Related: BUG-014 (template-nonsense examples — JA-96 + A-038 new); BUG-003/005 cross-pattern contamination — A-032/033 new.",
    ("A. Japanese language", "A-005"): " | Related: BUG-009 (が-pattern using は) — A-036 new.",
    ("A. Japanese language", "A-007"): " | Related: BUG-009 (particle alignment) — A-036 new.",
    ("A. Japanese language", "A-008"): " | Related: BUG-016 (transitivity coverage gap) — A-039 new; BUG-008 (folk-linguistic 'intransitive') — A-035 new.",
    ("A. Japanese language", "A-017"): " | Related: BUG-021 (primary_reading kun for 6 standalone kanji) — A-041 new.",
    ("A. Japanese language", "A-020"): " | Related: BUG-004 (pitch_marks.mora 911 wrong values). Planned JA-93 (count_mora algorithm equality) — Phase-0 regression in N5Improvement.txt today.",
    ("A. Japanese language", "A-024"): " | Related: BUG-003 (n5-098 all 10 translation_en stuck on 'I like cats'); BUG-005 (n5-166 ex[5] cross-pattern). A-032/033 new cover the regression-detection lens.",
    ("A. Japanese language", "A-028"): " | Related: BUG-017 (3 OOS kanji 倍/籍/末 in vocab.form) — JA-99 + A-040 new.",
    # I tab — partial coverage
    ("I. Data engineering", "I-001"): " | Related: BUG-015 counter schema — JA-97 + I-017 new; BUG-022 form/lemma split — JA-101 + I-021 new.",
    ("I. Data engineering", "I-002"): " | Related: BUG-016 (transitivity field-coverage gap) — JA-98 + A-039 new; BUG-012 (review_status provenance) — JA-35 + I-018 new.",
    ("I. Data engineering", "I-003"): " | Related: BUG-020/023 (kanji.json ↔ vocab.json form drift) — JA-100 + I-020 new.",
}


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)
    print(f"Loaded {XLSX}")

    # Add new scenarios (idempotent: skip if ID exists)
    print("\n--- Adding 17 new scenarios mapped to bug classes / CI invariants ---")
    added = 0
    skipped = 0
    by_tab: dict[str, list[dict]] = {}
    for spec in NEW_SCENARIOS:
        by_tab.setdefault(spec["tab"], []).append(spec)
    for tab_name, specs in by_tab.items():
        sh = wb[tab_name]
        existing_ids = {sh.cell(row=r, column=1).value
                        for r in range(5, sh.max_row + 1)
                        if sh.cell(row=r, column=1).value}
        target_row = sh.max_row + 1
        for spec in specs:
            if spec["id"] in existing_ids:
                skipped += 1
                continue
            sh.cell(row=target_row, column=1).value = spec["id"]
            sh.cell(row=target_row, column=2).value = spec["subcat"]
            sh.cell(row=target_row, column=3).value = spec["persp"]
            sh.cell(row=target_row, column=4).value = spec["scenario"]
            sh.cell(row=target_row, column=5).value = spec["steps"]
            sh.cell(row=target_row, column=6).value = spec["expected"]
            sh.cell(row=target_row, column=7).value = spec["priority"]
            sh.cell(row=target_row, column=8).value = spec["severity"]
            sh.cell(row=target_row, column=9).value = spec["test_type"]
            sh.cell(row=target_row, column=10).value = spec["notes"]
            sh.cell(row=target_row, column=11).value = spec["effort"]
            sh.cell(row=target_row, column=12).value = spec["owner"]
            sh.cell(row=target_row, column=13).value = spec["tools"]
            sh.cell(row=target_row, column=14).value = None  # Last run date
            sh.cell(row=target_row, column=15).value = "Not Yet Run"
            sh.cell(row=target_row, column=16).value = None  # Tracker link
            sh.cell(row=target_row, column=17).value = "—"
            sh.cell(row=target_row, column=18).value = "0%"
            for c in range(1, 19):
                sh.cell(row=target_row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            target_row += 1
            added += 1
    print(f"  Added {added} new scenarios, skipped {skipped} already present")

    # Append Notes-linkage to existing scenarios
    print("\n--- Adding Notes linkage to existing scenarios ---")
    n = 0
    for (tab, row_id), suffix in NOTES_LINKAGE.items():
        sh = wb[tab]
        for r in range(5, sh.max_row + 1):
            if sh.cell(row=r, column=1).value == row_id:
                existing = sh.cell(row=r, column=10).value or ""
                signature = suffix.split("|")[1].strip()  # the "Related: ..." portion
                if signature not in existing:
                    sh.cell(row=r, column=10).value = existing + suffix
                    n += 1
                break
    print(f"  Updated {n} existing scenarios with bug-class linkage")

    wb.save(XLSX)
    print(f"\nSaved {XLSX}")

    # Print final stats
    total = 0
    for tab in by_tab.keys():
        sh = wb[tab]
        c = sum(1 for r in range(5, sh.max_row + 1) if sh.cell(row=r, column=1).value)
        print(f"  {tab}: {c} scenarios total")
        total += c
    # Get full total
    grand_total = 0
    for sname in wb.sheetnames:
        if sname in ("Overview", "User Reported Bugs"):
            continue
        s = wb[sname]
        grand_total += sum(1 for r in range(5, s.max_row + 1) if s.cell(row=r, column=1).value)
    print(f"\nGrand total scenarios across all 14 category tabs: {grand_total}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
