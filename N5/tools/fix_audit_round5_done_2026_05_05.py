"""Stamp Done on the round-5 items shipped in v1.12.33.

Done in this release (14):
  ISSUE-035 — version.json invariants live count
  ISSUE-036 — CONTRIBUTING.md
  ISSUE-037 — Free pill in trust band (paired with IMP-058)
  ISSUE-038 — robots.txt + sitemap.xml
  ISSUE-039 — Mock + Missed in primary-nav (paired with IMP-061)
  ISSUE-040 — LICENSE href fix (GitHub blob URL)
  ISSUE-041 — _meta schema (paired with IMP-059)
  ISSUE-044 — manifest-path trap doc note in SELF-HOST.md
  ISSUE-046 — auto-locale toast i18n
  ISSUE-047 — README docs links
  IMP-058   — Free pill (paired)
  IMP-059   — _meta schema (paired)
  IMP-060   — Dependabot config
  IMP-061   — primary-nav (paired)
  IMP-062   — unified npm run build
  IMP-063   — manifest share_target

Skipped per user direction (1): IMP-057 (CODE_OF_CONDUCT)
Skip-on-error (3): ISSUE-043, ISSUE-045, IMP-065, IMP-067
Skip-needs-decision (4): ISSUE-042, IMP-064, IMP-066, IMP-068
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

DONE_IDS = {
    'ISSUE-035', 'ISSUE-036', 'ISSUE-037', 'ISSUE-038', 'ISSUE-039',
    'ISSUE-040', 'ISSUE-041', 'ISSUE-044', 'ISSUE-046', 'ISSUE-047',
    'IMP-058', 'IMP-059', 'IMP-060', 'IMP-061', 'IMP-062', 'IMP-063',
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1
    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)
    closed = []
    not_found = set(DONE_IDS)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_row_idx = None
        for r in range(1, min(7, ws.max_row + 1)):
            cells = [str(c.value or '').strip() for c in ws[r]]
            if any(c.lower() == 'id' for c in cells) and any(
                c.lower().startswith('decision') for c in cells
            ):
                header_row_idx = r
                break
        if header_row_idx is None:
            continue
        header = [c.value for c in ws[header_row_idx]]
        try:
            id_col = next(i for i, v in enumerate(header, start=1)
                          if v and str(v).strip().lower() in ('id', 'issue id', 'item id'))
            dec_candidates = [i for i, v in enumerate(header, start=1)
                              if v and str(v).strip().lower().startswith('decision (')]
            if not dec_candidates:
                dec_candidates = [i for i, v in enumerate(header, start=1)
                                  if v and str(v).strip().lower().startswith('decision')]
            dec_col = dec_candidates[0]
        except (StopIteration, IndexError):
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            id_cell = row[id_col - 1]
            dec_cell = row[dec_col - 1]
            id_val = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
            if id_val not in DONE_IDS:
                continue
            cur = (dec_cell.value or '').strip() if isinstance(dec_cell.value, str) else ''
            not_found.discard(id_val)
            if cur == 'Done':
                continue
            dec_cell.value = 'Done'
            closed.append((sheet, id_val, cur or '<blank>'))

    wb.save(XLSX)
    print(f'Closed {len(closed)} item(s):')
    for sh, iid, prev in closed:
        print(f'  [{sh}]  {iid}  ({prev} -> Done)')
    if not_found:
        print(f'\nWARNING: not found: {sorted(not_found)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
