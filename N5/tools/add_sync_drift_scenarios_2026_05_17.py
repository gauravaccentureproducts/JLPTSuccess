"""Add Cross-Artifact Sync Protocol test scenarios to xlsx.

Adds 3 rows to "K. QA testing" tab — one per wired INV-N (INV-4 /
INV-5 / INV-10) — and refreshes the User Reported Bugs Summary to
reflect that the protocol-install batch is not a user-reported bug
(no row added to the Bug Tracker sheet).

Run from N5/:
    python tools/add_sync_drift_scenarios_2026_05_17.py
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent  # N5/
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"


# Three new rows — one per wired INV-N. IDs continue from K-015.
NEW_K_ROWS = [
    {
        "id": "K-016",
        "sub_category": "Cross-artifact sync",
        "persp": "65",
        "scenario": "Version.json count drift detection (INV-4 / JA-107)",
        "test_steps": (
            "1. Run `python tools/check_content_integrity.py`\n"
            "2. Inspect JA-107 status line\n"
            "3. To regress-test: temporarily edit "
            "data/version.json.counts.vocab to a wrong number; re-run\n"
            "4. Expect FAIL with explicit drift message; revert"
        ),
        "expected_result": (
            "JA-107 PASSes on the current corpus snapshot. The check "
            "fails LOUDLY (with specific drift message) when "
            "version.json declares a count that diverges from "
            "data/<corpus>.json actual length. The 2026-05-17 protocol-"
            "install commit demonstrated this against vocab count 1009→995."
        ),
        "priority": "P2",
        "severity": "Major",
        "test_type": "Auto",
        "notes": (
            "Cross-Artifact Sync Protocol INV-4 hard CI gate. JA-107 "
            "added 2026-05-17 alongside Rule 5 install. Companion to "
            "JA-47 (CONTENT-LICENSE.md side). Together they enforce "
            "the version-stamp ↔ live-data symmetry from both public "
            "surfaces."
        ),
        "effort": "1h",
        "owner": "QA / build engineer",
        "tools": "tools/check_content_integrity.py, tools/cross_artifact_sync_report.py",
        "last_run_date": "2026-05-17",
        "last_run_result": "PASS",
        "tracker": "",
        "depends_on": "JA-47, JA-107",
        "coverage": "100%",
    },
    {
        "id": "K-017",
        "sub_category": "Cross-artifact sync",
        "persp": "65",
        "scenario": "Locale-parity drift detection (INV-5 / JA-108)",
        "test_steps": (
            "1. Run `python tools/check_content_integrity.py`\n"
            "2. Inspect JA-108 status line\n"
            "3. To regress-test: temporarily add a key to "
            "locales/en.json that's absent from hi.json; re-run\n"
            "4. Expect FAIL listing the asymmetric key; revert"
        ),
        "expected_result": (
            "JA-108 PASSes on the current corpus snapshot. Strict "
            "key-set equality across every file in locales/*.json — "
            "any asymmetric key (in en but not hi, or vice versa) "
            "fails. The 2026-05-17 install batch surfaced 9 such "
            "asymmetries (3 _meta.* in hi, 6 chokai_detail.* in en); "
            "all closed by mirroring."
        ),
        "priority": "P2",
        "severity": "Major",
        "test_type": "Auto",
        "notes": (
            "Cross-Artifact Sync Protocol INV-5 hard CI gate. JA-108 "
            "added 2026-05-17. Includes _meta block (per Rule-5-install "
            "decision to mirror metadata rather than exempt it). The "
            "chokai_detail keys intentionally carry Japanese text on "
            "both locales — that's an in-app pedagogy convention, not "
            "a translation gap."
        ),
        "effort": "1h",
        "owner": "QA / locale maintainer",
        "tools": "tools/check_content_integrity.py, tools/cross_artifact_sync_report.py",
        "last_run_date": "2026-05-17",
        "last_run_result": "PASS",
        "tracker": "",
        "depends_on": "JA-108",
        "coverage": "100%",
    },
    {
        "id": "K-018",
        "sub_category": "Cross-artifact sync",
        "persp": "65",
        "scenario": "Script-reference drift detection (INV-10 / JA-109)",
        "test_steps": (
            "1. Run `python tools/check_content_integrity.py`\n"
            "2. Inspect JA-109 status line\n"
            "3. To regress-test: rename a script that's referenced in "
            "prompts/N5Improvement.txt without updating the reference; "
            "re-run\n"
            "4. Expect FAIL with the doc + path pair; revert"
        ),
        "expected_result": (
            "JA-109 PASSes on the current corpus snapshot. Every "
            "backtick-quoted `tools/<name>.py` reference inside "
            "prompts/Japanese language Accuracy check.txt, "
            "prompts/N5Improvement.txt, and docs/AUDIT-COVERAGE-*.md "
            "resolves to an actual file on disk. Scope decision: the "
            "cross-level procedure manual is excluded (its refs are "
            "abstract Nx-builder targets, by design)."
        ),
        "priority": "P3",
        "severity": "Minor",
        "test_type": "Auto",
        "notes": (
            "Cross-Artifact Sync Protocol INV-10 hard CI gate. JA-109 "
            "added 2026-05-17. The 2026-05-17 install batch surfaced 1 "
            "drift: N5Improvement.txt referenced "
            "tools/register_dev_issue_list_deferrals_2026_05_05.py "
            "which had been removed in a tools-cleanup pass; closed by "
            "retargeting to tools/register_audit_2026_05_12.py."
        ),
        "effort": "30m",
        "owner": "QA / docs maintainer",
        "tools": "tools/check_content_integrity.py, tools/cross_artifact_sync_report.py",
        "last_run_date": "2026-05-17",
        "last_run_result": "PASS",
        "tracker": "",
        "depends_on": "JA-109",
        "coverage": "100%",
    },
]


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1

    wb = openpyxl.load_workbook(XLSX)
    k = wb["K. QA testing"]

    # Locate insertion point — first empty row after row 4 (header).
    insert_at = k.max_row + 1
    # Defensive: if max_row sits on a trailing-empty cell, scan back.
    while insert_at > 5 and not k.cell(row=insert_at - 1, column=1).value:
        insert_at -= 1

    # Check for idempotency — skip rows whose ID already exists.
    existing_ids = set()
    for r in range(5, k.max_row + 1):
        rid = k.cell(row=r, column=1).value
        if rid:
            existing_ids.add(rid)

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    added = 0
    for spec in NEW_K_ROWS:
        if spec["id"] in existing_ids:
            print(f"  skip {spec['id']} (already present)")
            continue
        row = insert_at + added
        k.cell(row=row, column=1).value = spec["id"]
        k.cell(row=row, column=2).value = spec["sub_category"]
        k.cell(row=row, column=3).value = spec["persp"]
        k.cell(row=row, column=4).value = spec["scenario"]
        k.cell(row=row, column=5).value = spec["test_steps"]
        k.cell(row=row, column=6).value = spec["expected_result"]
        k.cell(row=row, column=7).value = spec["priority"]
        k.cell(row=row, column=8).value = spec["severity"]
        k.cell(row=row, column=9).value = spec["test_type"]
        k.cell(row=row, column=10).value = spec["notes"]
        k.cell(row=row, column=11).value = spec["effort"]
        k.cell(row=row, column=12).value = spec["owner"]
        k.cell(row=row, column=13).value = spec["tools"]
        k.cell(row=row, column=14).value = spec["last_run_date"]
        k.cell(row=row, column=15).value = spec["last_run_result"]
        k.cell(row=row, column=16).value = spec["tracker"]
        k.cell(row=row, column=17).value = spec["depends_on"]
        k.cell(row=row, column=18).value = spec["coverage"]
        for c in range(1, 19):
            cell = k.cell(row=row, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
        print(f"  added {spec['id']} at row {row}: {spec['scenario']}")
        added += 1

    if not added:
        print("No new rows to add (all 3 already present).")
        return 0

    wb.save(XLSX)
    print(f"Saved {added} new scenarios to K. QA testing sheet.")
    print()
    print("Note: User Reported Bugs sheet is NOT touched — this batch is a")
    print("governance/tooling install, not a user-reported bug close-out.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
