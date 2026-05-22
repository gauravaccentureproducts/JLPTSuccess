"""Sync test-scenarios xlsx with N5/prompts/ + N5/feedback/ folders.

User directive 2026-05-17: every info in prompts/ + feedback/ should be
present in test scenarios.

Scope (per user choice, Option 1 — "structured + audit-doc summaries"):

  Prompts/  (structured items):
    - 60 A-NN audit categories in accuracy prompt → tab A scenarios
    - 18 Phase-0 regression blocks in N5Improvement → tab K (and cross-
      refs to A/B/I/C where corpus-specific)
    - 15 FP-NN false-positive class entries → tab K (auditor scope-
      exclusion guidance; not bugs)
    - LegalVetting.txt → tab F + tab G summary scenarios
    - LocaleTransitionEnHi.txt → tab C summary scenario

  Feedback/ (one summary scenario per audit doc, distributed by topic):
    - accuracy-audit-* → tab A
    - audit-drift-findings → tab A
    - audit-round9 → tab A + tab J
    - hindi-audit-findings → tab C
    - legal-vetting-audit-* (×3) → tab F + tab G
    - locale-transition-inventory → tab C
    - MASTER-TASK-LIST → tab M (Operations) as the cross-cutting catch-all
    - n5-richness-audit → tab J
    - native-teacher-* (×4) → tab A
    - ui-testing-plan → tab D + tab E + tab K (cross-link)
    - voicevox-integration-notes → tab I + tab M

  Feedback/closed/ (one summary scenario per closed audit doc):
    - jlpt-n5-content-* / native-teacher-review-* → tab A
    - jlpt-n5-data-* / jlpt-n5-reference-* / coverage-comparison → tab I
    - jlpt-n5-paper-files-audit → tab B
    - jlpt-n5-goi-audit / jlpt-n5-moji-and-source-audit → tab A
    - jlpt-n5-reading-feedback → tab A
    - jlpt-n5-homepage-update / jlpt-n5-tutor-ux-* / jlpt-n5-ui-design → tab D
    - jlpt-n5-infrastructure-* / jlpt-n5-tutor-developer-* → tab M
    - llm-audit-validation-report → tab K
    - f15-23-* / f17-9-* decisions / teacher-audit → tab A
    - jlpt-n5-knowledgebank-md-audit / native-teacher-review-request → tab A

Idempotency: every new row has a unique ID; if a row with that ID
already exists in the target tab, the script skips it.

Run: python tools/sync_test_scenarios_with_prompts_feedback_2026_05_17.py
"""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"

# ---------------------------------------------------------------------------
# Item catalog. Each entry → one scenario row.
#   (tab_letter, ref_code, scenario_title, test_steps, expected_result,
#    priority, severity, test_type, notes, tools, owner, depends_on)
# ---------------------------------------------------------------------------

# --- 60 A-NN audit categories (accuracy prompt) → tab A ---
A_NN_CATEGORIES = [
    ("A1", "Incorrect grammar", "Manual review of grammar.json against native-teacher pass."),
    ("A2", "Unnatural collocations", "Random-sample 200 vocab examples; native ear required."),
    ("A3", "Wrong particle", "Per-pattern check that canonical particle is used correctly."),
    ("A4", "Wrong reading", "Sample 50 vocab/kanji readings against kanjium / Genki."),
    ("A5", "Wrong kanji usage", "JA-13 enforces N5-whitelist; manual sample for usage correctness."),
    ("A6", "Onyomi / kunyomi misclassification", "JA-24 + manual sample of kanji.json on/kun tagging."),
    ("A7", "Vocabulary meaning errors", "Cross-check gloss against external dictionaries."),
    ("A8", "Mistranslation of Japanese → English", "Sample 100 examples; native-bilingual review."),
    ("A9", "Half-applied ウ音便 in keigo politeness_ladder.humble", "Detect a+u→ō / i+u→yū missed contractions."),
    ("A10", "Sokuon / rendaku as separate readings", "JA-22 dedup within entry; manual sample for cross-entry."),
    ("A11", "Mixed orthography for the same morpheme", "わたし vs 私 audit; BUG-046 lineage."),
    ("A12", "meaning_ja CROSS-CONTAMINATION (closed by JA-75)", "JA-71 + JA-75 lock; manual sample of new patterns."),
    ("A13", "PITCH-ACCENT MORA COUNT ERRORS", "JA-70 + JA-90 lock; sample re-validation against kanjium."),
    ("A14", "PLACEHOLDER / SELF-INCRIMINATING TEXT IN SHIPPED DATA", "JA-77 grep across all data/*.json."),
    ("A15", "OFF-PATTERN BOILERPLATE AT EXAMPLE INDICES [4][5][6]", "JA-81 boilerplate-leak guard at threshold 10."),
    ("A16", "FORM-TAG VALUE MISMATCHED WITH EXAMPLE CONTENT", "Sample 50 form-tagged examples for tag↔content agreement."),
    ("A17", "WITHIN-PATTERN DUPLICATE EXAMPLES", "Per-pattern dedup check on ja strings."),
    ("A18", "VOCAB EXAMPLE TEMPLATE-LEAK: 'Xを 見ました。'", "JA-83 locks; manual sample for residue."),
    ("A19", "VOCAB EXAMPLE TEMPLATE-LEAK: 'あの Xは どこですか。'", "JA-83 locks."),
    ("A20", "VOCAB EXAMPLE TEMPLATE-LEAK: 'これは Xです。' / 'あれは Xです。'", "JA-83 locks for non-demonstrative entries."),
    ("A21", "VOCAB EXAMPLE TEMPLATE-LEAK: '「X」と あいさつしました。' for non-greetings", "JA-83 locks."),
    ("A22", "VOCAB CROSS-ENTRY BOILERPLATE (JA-81 analogue for vocab)", "Sentence reuse threshold check across vocab.json examples."),
    ("A23", "VOCAB HEADWORD-MISSING-FROM-EXAMPLE", "Per-entry check that ja text contains the headword form or stem."),
    ("A24", "KANJI.JSON SENTENCE TRANSLATIONS EMPTY", "JA-84 locks every populated ja with non-empty translation_en."),
    ("A25", "KANJI.JSON STROKE-COUNT MISMATCH BETWEEN INT AND TEXT", "JA-84 catches stroke_count int vs why_it_matters prose drift."),
    ("A26", "KANJI.JSON DUPLICATE-FORM EXAMPLES WITH SAME READING", "JA-103 locks (form, reading) uniqueness per kanji."),
    ("A27", "DOKKAI MONDAI RANGE IS 4-6 (NOT 1-4 LIKE LISTENING)", "Per-corpus mondai range validation."),
    ("A28", "READING.JSON FORMAT_ROLE + VOCAB_USED SCHEMA GAPS", "JA-19 + JA-104..106 locks; manual sample for gaps."),
    ("A29", "DOKKAI / READING EXPLANATION_HI EMPTY (LOCALE PARITY)", "JA-85 locks."),
    ("A30", "PAPERS RATIONALE-ANSWER MISALIGNMENT", "JA-32 round-trip; manual review of paper questions."),
    ("A31", "AUTHENTIC.JSON CONTEXT_HI EMPTY (LOCALE PARITY)", "JA-86 locks."),
    ("A32", "QUESTIONS.JSON DISTRACTOR-EXPLANATIONS GAP", "JA-86 + sample."),
    ("A33", "LISTENING MONDAI 3 vs 4 CHOICE-COUNT RULES", "JA-33 closed-enum; sample for choice-count distinction."),
    ("A34", "SCHEMA-TYPE DISCRIMINATION ACROSS QUESTIONS-LIKE FILES", "JA-29 + sample of paraphrase vs kanji_writing."),
    ("A35", "CROSS-CORPUS READING/GLOSS CONSISTENCY", "JA-87 wave-1; JA-100/105 cross-corpus refs."),
    ("A36", "PARTICLE-PRECISION L2-ERROR PATTERNS", "JA-88 wave-2 scan."),
    ("A37", "COUNTER-NOUN SEMANTIC PAIRING (本+つ for books, etc.)", "JA-89 lock1."),
    ("A38", "ANIMACY VIOLATIONS: ANIMATE NOUN + が あります", "JA-89 lock2."),
    ("A39", "BARE-ARTICLE EN: 'There is X.' for singular common nouns", "JA-89 lock3."),
    ("A40", "WRONG-HEADWORD EXAMPLES (ja text doesn't contain headword)", "JA-89 lock4."),
    ("A41", "UNNATURAL ADVERB-VERB PAIRINGS (ぜひ/ただ + bare declarative)", "Sample 50 + FP-15."),
    ("A42", "POTENTIAL/EXISTENTIAL VERB + VOLITIONAL つもり MISMATCH", "Sample grammar patterns using つもり."),
    ("A43", "VOCAB VERB-TEMPLATE LEAKS: '毎日 X ことが できます'", "JA-83 lock; manual sample."),
    ("A44", "MULTI-TRANSLATION EN: 'X / Y' pick-one", "Sample 50 EN translations for slash-separated alternatives."),
    ("A45", "TEMPLATE 'あの Xは どこですか。' IN GRAMMAR.JSON", "JA-83 also applies to grammar.json examples."),
    ("A46", "PITCH ACCENT VALIDATED AGAINST EXTERNAL REFERENCE", "JA-90 vs kanjium; native auditory review for unverified 43."),
    ("A47", "AUDIO RENDERER NORMALIZATION GAP — bunsetsu-space particle drop", "F.12 lesson; tested via Phase-0 audio re-render diff."),
    ("A48", "VERB-CLASS PARTICLE DISAMBIGUATION IN wcp / common_mistakes", "JA-95 (planned); BUG-002 lineage."),
    ("A49", "CROSS-PATTERN EXPLANATION / TRANSLATION CONTAMINATION", "JA-91 (planned); BUG-003/005 lineage."),
    ("A50", "MORA-COUNT SYSTEMATIC ERROR IN pitch_marks", "JA-93 (planned); BUG-004 lineage."),
    ("A51", "PATTERN-INSTANCE CONTAMINATION", "Manual review; BUG-006 lineage."),
    ("A52", "RIGHT/WRONG FRAMING FOR GRAMMATICALLY-VALID ALTERNATIVES", "Manual review; BUG-007 lineage; register-variant framing."),
    ("A53", "FOLK-LINGUISTIC GRAMMAR TERMINOLOGY IN common_mistakes", "Manual review; BUG-008 lineage."),
    ("A54", "PATTERN-PARTICLE MISMATCH IN CANONICAL EXAMPLES", "JA-95 (planned); BUG-009 lineage."),
    ("A55", "MORA-COUNT, PATTERN-MARKER, AND PARTICLE-ALIGNMENT CI INVARIANTS", "JA-93/94/95 trio (planned wiring)."),
    ("A56", "STATIC-MIRROR COVERAGE (CRAWLABILITY) ACROSS ALL CONTENT SURFACES", "Phase-0 static-mirror block; BUG-010 close-out."),
    ("A57", "SCHEMA-LEVEL REGISTER-VARIANT DETECTION IN common_mistakes", "JA-64 (tightened); BUG-011/013 lineage."),
    ("A58", "PROVENANCE-LABEL DISAMBIGUATION (HUMAN vs AI REVIEW)", "JA-35 closed-enum; BUG-012 lineage."),
    ("A59", "READING.JSON BATCH-DRIFT — 6 FIELD DIVERGENCES (BUG-041..046)", "JA-104/105/106 lock; F.23 lessons."),
    ("A60", "LISTENING.JSON MIGRATION DRIFT — 7 FIELD DIVERGENCES (BUG-047..053)", "JA-110/111 lock; F.24 lessons."),
    ("A63", "JA-91 + JA-94 FINAL UNBLOCK — baseline-allowlist pattern for invariants that can't reach 0 false-positives mechanically", "JA-91 (43-pair cross-pattern explanation similarity baseline) + JA-94 (14-example BUG-006-CANDIDATE baseline against the 178-pattern structural-markers catalog). Methodology: snapshot legitimate-case set into data/_jaNN_baseline.json with classification notes; CI trips on NEW drift only. Same shape as JA-67 Density-3 below-floor lock; A63 documents the broader pattern for future reserved-invariant promotions. **2026-05-17 status update:** both JA-91 and JA-94 baselines RESOLVED to empty arrays via Phase A (14 example replacements) + Phase B (33 explanation_en rewrites). Procedure-manual F.25 + F.26."),
    ("A64", "AUDIO PHASE-2 CLOSE-OUT — full VOICEVOX from-source re-render at unified speed_scale supersedes stacked post-processing", "Methodology lesson from the 2026-05-17 audio Phase-2 close-out (commit cdd0e6d). 50 listening items re-rendered from source at speed_scale=1.00 via VOICEVOX 0.25.2; replaced 2026-05-12 render at 0.95 + Phase-1 atempo (47d1edc) + Phase-1.5 rubberband (c79c02e) post-processing layers with a single coherent from-source render + minimal per-item post-processing (16 direct, 29 single-pass atempo, 5 single-pass rubberband). Provenance value of 'one render, one filter per item' justifies the wall-clock cost when the TTS engine is locally available and the baseline speed_scale was suboptimal. Procedure-manual F.27."),
    ("A65", "MULTI-ROLE SPECIALIST-REVIEW-BY-TAB methodology + bounded-honest stamping vocabulary + brutal-honesty re-audit pattern", "Methodology lesson from the 2026-05-17 multi-role specialist-review sweep across all 720+ scenarios in the 15-tab xlsx. Bug-ID naming NR-{ROLE}-NNN; bounded-honest stamping vocabulary (PASS / PASS-limited / PASS-architectural / PASS-spot-check / Manual-deferred / Skipped-external / Blocked). Brutal-honesty re-audit pattern: re-execute deep scans (full corpus, not 30-sample) and re-classify previously-PASS stamps with stricter qualifiers. 16 NR-* bugs surfaced across 5 batches (Native Japanese / Native Hindi+JLPT / Security+Privacy+Data / Brutal-honesty / UI). Procedure-manual F.28."),
    ("A66", "SELENIUM UI TEST CLASS — E2E coverage of every functional surface in spec §5 + static-mirror + a11y + security + Service Worker + audio + i18n + console health", "Test class added 2026-05-17 (commit 5635425). 55 Selenium 4 scenarios via tools/ui_test_suite_2026_05_17.py covering: spec §5.1-5.16 routes, 14 static-mirror routes, sitemap.xml + robots.txt, accessibility landmarks, security headers, Service Worker, audio reachability, locale parity, 0-SEVERE console errors. Auto-driver via Selenium Manager. Critical NR-UI-001 lesson: CSP frame-ancestors + X-Frame-Options are HTTP-header-only — IGNORED via <meta>; always verify runtime effectiveness, not just source presence. Procedure-manual F.29."),
    ("A67", "PAPER-QUESTION grammarPatternId MIS-TAG SCAN (PAPER-001)", "Added 2026-05-18 from PAPER-001 close-out. For every Mondai 1 paper question whose correctIndex resolves to a single Japanese particle, the grammarPatternId field MUST reference the canonical n5-NNN pattern for that particle. Mismatches indicate an early auto-tag pass that assigned default-tag (e.g., n5-013 = も) regardless of the actual answer particle. CI invariant JA-120 (paper bunpou Mondai-1 grammarPatternId matches canonical particle pattern, PAPER-001 drift guard). Canonical map (21 entries): は→n5-002, が→n5-003, を→n5-004, に→n5-005, へ→n5-006, で→n5-007, と→n5-008, から→n5-009, まで→n5-010, や→n5-011, も→n5-013, か→n5-023, ね→n5-025, よ→n5-026, の→n5-028, だけ→n5-033, ぐらい/くらい→n5-035, ごろ→n5-036, など→n5-037, ずつ→n5-038, より→n5-095. Procedure-manual F.30."),
    ("A68", "PAPER-RATIONALE META-FIX-HISTORY SCAN (PAPER-003)", "Added 2026-05-18 from PAPER-003 close-out. For every paper question, the learner-facing rationale and rationale_hi fields MUST NOT contain commit-message-style meta-fix history. Substring-scans 12 trigger phrases: auto_inferred, previously tagged, prior version was, Stem now anchored, Stem now includes, replaces colloquial, replaces ので per, the original option, was dropped because, replaced with, patched to, fix:. CI invariant JA-121. Distractor-analysis content preserved by design — no trigger overlap. Procedure-manual F.30."),
    ("A69", "PAPER-RATIONALE_HI NATURAL-HINDI SCAN (PAPER-004)", "Added 2026-05-18 from PAPER-004 close-out. For every paper question, the learner-facing rationale_hi field MUST be natural Hindi — not word-by-word literal translation of rationale_en. Substring-scans 17 trigger fragments covering apostrophe-s possessive ('s घर, दोस्त's, माता's), English contractions (मैं'm, मैं'll), mojibake artifacts (यहाँre, वहाँre, o'घड़ी), untranslated English filler (yet, lot, have जाना). CI invariant JA-122. CRITICAL anti-pattern: do NOT translate from broken rationale_hi back into clean Hindi — always source from rationale_en + verify against actual stem. Procedure-manual F.30."),
    ("A70", "LLM / SEARCH-CRAWLER ACCESSIBILITY — 8-surface canonical set (LLM-001..005)", "Added 2026-05-18 from LLM-001..005 close-out. For a hash-routed SPA on static hosting, per-entity content is invisible to LLM web-fetch tools and search crawlers (fragments don't travel in HTTP requests). The 8-surface canonical set documented in procedure manual F.31 makes the corpus discoverable: (1) per-entity static mirrors at /Nx/learn/<module>/<id>/index.html, (2) per-module index landing, (3) 7 thin summary pages /Nx/<slug>.html, (4) sitemap.xml with ≥1000 URLs, (5) data/index.json corpus discovery JSON, (6) llms.txt at root + /Nx/, (7) robots.txt, (8) noscript fallback (no hash routes). CI invariants JA-123/124/125/126. Build script: tools/build_llm_surfaces_2026_05_18.py (8 stages). Procedure-manual F.31."),
    ("A71", "REGISTER-VARIANT vs GRAMMAR-ERROR DISTINCTION — 6 defect classes (D1..D6) per REG-001", "Added 2026-05-18 from REG-001 close-out. A grammatical register CHOICE between two correct forms must NOT be framed as WRONG/RIGHT in wrong_corrected_pair. Both forms are grammatical; the choice is register, not correctness. Use kind: register_variant nested in common_mistakes with form_a/form_b/label_a/label_b. Six defect classes: D1 (WRONG/RIGHT framing on register choice), D2 (conflated semantics in alternatives — e.g., identity vs description), D3 (formality vs elevation conflation — sentence-formality ≠ referent-elevation 尊敬), D4 (out-of-Nx-scope as canonical without scope_note), D5 (kana of whitelist kanji), D6 (self-contradicting annotation — ✗ line marked appropriate to a register). CI invariant JA-127 catches D6 marker phrases. REG-001 close-out fixed 6 entries (n5-046 + n5-097/102/127/173/179). 84 SWEEP-1 candidates documented in docs/REG-001-SWEEP-1-candidates_2026_05_18.md for native-speaker triage as REG-002..NN. Procedure-manual F.32."),
    ("A72", "PAPER-QUESTION SCHEMA-DISCIPLINE — 3 durable classes (DOKKAI-001..003)", "Added 2026-05-18 from DOKKAI-001..003 close-out. Three schema-discipline classes for every paper-bank file across all categories (bunpou/goi/moji/dokkai). Class A (single source of truth for passages, CI JA-128): paper questions MUST NOT carry passage_text; canonical text lives in passages[label].text via passage_label foreign key. Class B (English-fragment temporal markers in rationale_hi, CI JA-129): no untranslated 'ago' / 'yet' / 'lot' (+ punctuated variants); rewrite with target-language idiom ('पहले' for ago/before). Class C (schema-shape always-a-key, CI JA-130): every paper question has grammarPatternId as a key; null value requires not_applicable_<reason> provenance (not_applicable_comprehension/vocab/orthography). Same always-a-key-sometimes-null pattern as VOCAB-002 counter-field. DOKKAI-001..003 + horizontal sweep migrated 102 dokkai + 10 bunpou Mondai-3 + 2 rationale-hi rewrites + 107 schema-shape fills. Procedure-manual F.33."),
    ("A73", "MOBILE-UI COMPLIANCE — 5 durable classes (MOB-001..019 + DOKKAI-004)", "Added 2026-05-19 from MOB-001..019 + DOKKAI-004 close-out. Five mobile-UI defect classes documented in procedure manual F.34, CI-enforced via JA-131..134. Class A (touch-target HIG, CI JA-132): canonical CSS class set (.btn-action, .study-order-link, .home-up-link a, .back-link, .toc-expand-all/collapse-all, .brand-link, .skip-link, .btn-tiny, authentic ref-chips, examday/weakareas inline links) must meet min-height: 44px. Class B (iOS Safari auto-zoom, CI JA-133): input/textarea/select font-size must be ≥16px (canonical: max(1rem, 16px)). Class C (dead-end hash routes, CI JA-134): every internal hash href must point to an actual router entry; js/home.js uses ../ for home-up, js/listening-story.js uses canonical #/listeningstory form. Class D (locale-parity for hard-coded UI strings, CI JA-131): no hard-coded UI strings in JS template literals; route through t('key'); MOB-007 fix added nav.all_levels to en+hi locales. Class E (test-infrastructure / scenario-design gaps): MOB-018/019 are not app bugs — Selenium mobile-emulation scrollTo no-op (needs Appium fallback) + audio-UI scenarios target index pages instead of items (needs scenario rewrites). DOKKAI-004 ' by ' carry-over added to JA-129 trigger set. Procedure-manual F.34."),
    ("A74", "RATIONALE CONTENT-DISCIPLINE — copy-paste mismatch + meta-content (GOI-001..003)", "Added 2026-05-19 from GOI-001..003 close-out. Two rationale-content defect classes complement F.30/F.33/F.34.4. Class A (copy-paste content-mismatch, CI JA-136): no rationale_hi value (>30 chars) may be shared verbatim by 2+ questions within the same paper. Catches the GOI-001 hard learner-facing breakage where goi-6.11 rationale_hi was a byte-identical copy of goi-6.12 (about a different topic). Rejected stricter token-overlap proposal (~100 false positives from dictionary-form ↔ polite-form variation). Class B (meta-content in learner-facing rationale, CI JA-121 extended): no commit-trail / meta-doc-pointer phrases — extends JA-121 trigger set with 'Hence the rewording', 'rewording from a prior', 'from a prior version', 'documented at vocabulary_n5.md', 'documented at', 'does not bear on', 'test point this question'. Combined with existing JA-121/122/129/130, 5-invariant family on paper rationale fields. Procedure-manual F.35."),
    ("A75", "MIXED-SCRIPT MOJIBAKE + OFF-BY-ONE rationale_hi SHIFT + EXTENDED FIX-HISTORY (GOI-004..006)", "Added 2026-05-21 from GOI-004..006 close-out + horizontal mojibake sweep. Three durable rationale-content defect classes from the second goi sweep iteration. Class A (mixed-script Devanagari-inside-kana mojibake, CI JA-139): regex `[ぁ-ゖァ-ヺ一-鿿][ऀ-ॣ०-ॿ]` excluding danda U+0964/U+0965 and hyphen-separated cross-script terms. GOI-006 sample (goi-7.4 'あमारी ありません' → 'あまく ありません'); horizontal-deployment finding surfaced 2 more in dokkai-2.11 (ぐらि → ぐらい) + dokkai-3.4 (あमारी → あまり). Operational rule: corpus-wide CI pass runs BEFORE declaring class closed. Class B (narrow off-by-one rationale_hi shift, CI JA-137): 0 own-stem token overlap AND ≥2 next-Q token overlap. False-positive rate <1% vs ~21% for broad detector. GOI-004 sample (goi-7.6+goi-7.7 carried next-Q's Hindi). Class C (extended fix-history phrases, CI JA-121 trigger set extension): 11 new phrases. GOI-005 stripped 7 fields across 5 paper files. No new JA-NN for Class C (existing JA-121 intent unchanged). 7-invariant family on paper rationale fields when combined with A74. Procedure-manual F.37."),
    ("A76", "MOJI-PAPER CONTENT DISCIPLINE — 4 durable classes (MOJI-001..007 + JA-143 same-class follow-ups)", "Added 2026-05-21 from MOJI-001..007 close-out. Four durable defect classes from the moji-paper content review. Combines with A67/A72/A74/A75 into an 11-invariant family across paper-question fields. CI-enforced via JA-140 (moji stem_html uniform <u> emphasis wrapper) + JA-141 (moji grammarPatternId non-null bans auto_inferred provenance) + JA-142 (no over-literal Hindi 'के पास है पढ़ते हुए') + JA-143 (rationale/rationale_hi character-count parity 0.6×–2.0×). Source: AUDIT-COVERAGE Part 34."),
    ("A77", "GOVERNANCE-DOC STALE-CONTENT DISCIPLINE + ORPHANED-CI CLASS (DOCS-KANJI-001..004 + DOCS-VOCAB-001..004)", "Added 2026-05-21 from DOCS-KANJI/DOCS-VOCAB close-out + orphaned-workflow close-out. Six durable classes from the 2026-05-21 governance-doc + CI-hardening batch. Combines with A75 + A76 into a 13-invariant content-governance family. CI-enforced via JA-144 (REVIEW_DATE ISO 8601) + cross-doc convention checks. Closes 8 governance-doc stale-content bugs (kanji + vocab READMEs claiming stale counts, missing templates, undocumented bootstrapping exit criteria, ambiguous date formats). Source: AUDIT-COVERAGE Part 35."),
    ("A78", "CI-RECOVERY TRIAGE — test-side and product-side discoveries when Playwright suite first ran green (2026-05-21)", "Added 2026-05-21 from CI-recovery triage batch. Six durable classes from triaging 65 pre-existing failures masked by Playwright timeouts. When the suite first ran green (after parallelisation + video-off + 15-min timeout), cancellations had been hiding failures across 6 distinct classes: visual-regression OS-baseline mismatch (-win32.png vs -linux.png on CI), stale Playwright assertions against removed/restructured UI surfaces, color-contrast WCAG AA violations, recommender priority drift, etc. Generalises the discoveries so patterns are detectable by future audits. Source: AUDIT-COVERAGE Part 36."),
    ("A79", "PAPER-FILE source_file CANONICAL SENTINEL (DOCS-VOCAB-005)", "Added 2026-05-22 from DOCS-VOCAB-005 close-out. One durable defect class: a JSON metadata field that once held a real path keeps its historical breadcrumb prose long after the upstream file is deleted, embedding 'KnowledgeBank/<x>_questions_n5.md before merge on YYYY-MM-DD' style strings in 28 paper files. Canonical sentinel pattern (CI JA-145): source_file must be either (a) a path that resolves to an existing repo file, OR (b) the literal sentinel '(authored in-place)'. Other values fail. Historical breadcrumbs belong in CHANGELOG + README + git history, not in 28-files of data metadata. Closes the unaddressed-half of DOCS-VOCAB-003 (which closed prematurely as a README-only fix without touching the 28 paper files). Procedure-manual F.41. Source: AUDIT-COVERAGE Part 37."),
    ("A80", "COVERAGE-GAP CLOSURE — type-confusion in string fields + doc-state-vs-code-state drift (meta-audit batch)", "Added 2026-05-22 from meta-audit gap-closure batch. Two coverage classes the existing 147 invariants weren't designed to catch: GAP-A type-confusion in string fields (column typed as string but meant to hold specific shape like commit hash; schema accepts any string so type-checkers don't fire) and GAP-B doc-state vs code-state drift (TASKS.md [ ] items for features already shipped). N5 proof: 150/155 Fixed rows had dates in Fix Commit column instead of hashes; SVA-1.1 + SVA-1.4 were [ ] despite being shipped weeks earlier. Closed via JA-146 (strict) + tools/audit_tasks_md_against_codebase_2026_05_22.py (advisory). Reactive→proactive lesson: every JA-NN has been added AFTER a bug surfaced the pattern; no invariant authored proactively. Procedure-manual F.42. Source: AUDIT-COVERAGE Part 38."),
    ("A81", "GOVERNANCE-DOC CONSISTENCY DEFECTS — 5 durable classes from review-packet meta-audit (DOCS-VOCAB-006 + DOCS-CORE-001 + DOCS-BRAND-001 + DOCS-Q-001 + DOCS-DKE-001)", "Added 2026-05-22 from 5-bug DOCS-* close-out. Five durable defect classes generalized from the BUG-156..160 batch surfaced by review-packet meta-audit pipeline. Class A (README-vs-computed mismatch drift, CI JA-147): README documents 'Known mismatches (N)' between two data files; one side updated without the other; documented count and bulleted token set diverge from actual computed set. Class B (multi-source classification one source updated, CI JA-148): taxonomy lives in authoritative file (n5_core_pattern_ids.json) but consumers read per-entry catalog (grammar.json); reclassification not mirrored into per-entry catalog. Class C (review-packet strip undocumented in packet README): build script strips fields for privacy/anonymity; packet's own README doesn't enumerate all strip categories; reviewers misread blanked fields as bugs. Class D ('bank source' terminology overclaim): README implies source-derived relationship between corpora; actual ID overlap = 0; corpora are independent. Class E (placeholder boilerplate at-rest, CI JA-149): pre-formalization authoring leaves 'Pre-formalization (rationale not individually recorded)' placeholders that never get backfilled; placeholder ratio drifts to 28% over time. 3 of 5 classes shipped a CI invariant; 2 remain conventions documented in procedure-manual F.43. Plus the bug-spec-vs-reality verification reinforcement that rejected DOCS-VOCAB-005 as stale-snapshot artifact (audit ran on packet pre-dating b7f5787 fix). Procedure-manual F.43. Source: AUDIT-COVERAGE Part 39."),
]

# --- 18 Phase-0 regression blocks (N5Improvement) → tab K ---
PHASE0_BLOCKS = [
    ("P0-grammar", "Phase-0 Grammar (run mechanically over all 178 patterns)"),
    ("P0-vocab", "Phase-0 Vocab (1009 entries)"),
    ("P0-kanji", "Phase-0 Kanji (106 entries)"),
    ("P0-reading", "Phase-0 Reading (54 passages)"),
    ("P0-listening", "Phase-0 Listening (50 items)"),
    ("P0-density", "Phase-0 Density (interconnection — Sections 4 / Density-1..10)"),
    ("P0-kanji-gaps", "Phase-0 Kanji content gaps (added 2026-05-15)"),
    ("P0-dokkai-locale", "Phase-0 Dokkai / reading locale parity (added 2026-05-15)"),
    ("P0-auth-locale", "Phase-0 Authentic + Questions locale (added 2026-05-15)"),
    ("P0-xcorpus", "Phase-0 Cross-corpus consistency (added 2026-05-15 wave-1)"),
    ("P0-particle", "Phase-0 Particle-precision L2-error scan (added 2026-05-15 wave-2)"),
    ("P0-writing-discipline", "Phase-0 Audit-doc writing-discipline scan (added 2026-05-15)"),
    ("P0-native-teacher", "Phase-0 Native-teacher bug-class regression block (added 2026-05-16)"),
    ("P0-provenance", "Phase-0 Provenance-label disambiguation regression block (added 2026-05-16)"),
    ("P0-register-variant", "Phase-0 Register-variant kind-flag regression block (added 2026-05-16)"),
    ("P0-static-mirror", "Phase-0 Static-mirror coverage regression block (added 2026-05-16)"),
    ("P0-reading-batch-drift", "Phase-0 Reading.json batch-drift regression block (added 2026-05-17)"),
    ("P0-listening-migration", "Phase-0 Listening.json migration-drift regression block (added 2026-05-17)"),
    ("P0-ja91-ja94-baseline", "Phase-0 JA-91 + JA-94 baseline-stability regression block (added 2026-05-17)"),
    ("P0-selenium-ui-tests", "Phase-0 Selenium UI test regression block (added 2026-05-17)"),
    ("P0-multi-role-specialist", "Phase-0 multi-role specialist-review regression block (added 2026-05-17)"),
    ("P0-paper-question", "Phase-0 Paper-question grammarPatternId/rationale regression block (added 2026-05-18)"),
    ("P0-llm-surfaces", "Phase-0 LLM-surfaces + register-variant regression block (added 2026-05-18)"),
    ("P0-dokkai-schema", "Phase-0 dokkai-schema regression block (added 2026-05-18)"),
    ("P0-mobile-ui", "Phase-0 mobile-UI regression block (added 2026-05-19)"),
    ("P0-rationale-content", "Phase-0 rationale-content regression block (added 2026-05-19)"),
    ("P0-mixed-script-mojibake", "Phase-0 mixed-script + off-by-one + extended-fix-history regression block (added 2026-05-21)"),
    ("P0-ci-recovery-triage", "Phase-0 CI-recovery triage regression block (added 2026-05-21)"),
    ("P0-source-file-sentinel", "Phase-0 paper-file source_file canonical-sentinel regression block (added 2026-05-22)"),
    ("P0-governance-doc-consistency", "Phase-0 governance-doc consistency regression block (added 2026-05-22)"),
    ("P0-fix-commit-shape", "Phase-0 Fix Commit shape + TASKS-vs-code coverage-gap block (added 2026-05-22)"),
]

# --- 15 FP-NN false-positive class entries → tab K ---
FP_ENTRIES = [
    ("FP-1", "ウ音便 contraction where o+u → ō has no visible spelling change"),
    ("FP-2", "Past-marker mid-sentence in relative clauses"),
    ("FP-3", "Mixed kanji/kana orthography within one pattern"),
    ("FP-4", "Cross-pattern reuse of canonical sentences below threshold-10"),
    ("FP-5", "Counter rendaku in headword-presence checks"),
    ("FP-6", "Compound -くる verbs in headword-presence checks"),
    ("FP-7", "Demonstrative-pronoun entries using 'これは Xです。'"),
    ("FP-8", "English translations with trailing parenthetical annotations"),
    ("FP-9", "authentic.json `reading` field is the FULL phrase read-aloud"),
    ("FP-10", "English/short-technical rationales for paper-question answers"),
    ("FP-11", "ひとりでに ('by oneself') is a fixed adverb, not a particle stack"),
    ("FP-12", "〜いです IS the correct polite-affirmative form of an i-adj"),
    ("FP-13", "倍 (ばい) is a multiplier suffix, not a noun-counter"),
    ("FP-14", "i-adjective inflected forms still demonstrate the headword"),
    ("FP-15", "ぜひ / ただ + neutral declarative — semantically thin, not a grammar bug"),
]

# --- Feedback / closed audit docs → one summary scenario per doc ---
# Tuple: (tab, doc_path_rel, summary_title, summary_body)
AUDIT_DOCS = [
    # === feedback/ (current) ===
    ("A", "not-required/_n5_richness_audit_20260509.txt", "n5 richness audit 2026-05-09 — depth-over-width assessment (superseded; moved to not-required/ 2026-05-17)",
     "Cross-cutting richness audit (2026-05-09). Assesses whether N5 corpus depth (collocations, pitch-accent, examples, authentic-content, contrasts, common_mistakes) meets the Tofugu / Bunpro / Genki bar on existing entries before any width additions are considered. NOTE 2026-05-17: file moved to not-required/ — explicitly superseded by n5-richness-audit-2026-05-12.md which states 'the corpus has already absorbed the 2026-05-09 audit fully'. Scenario kept as historical pointer."),
    ("A", "feedback/accuracy-audit-run4-2026-05-13.md", "Accuracy audit run-4 2026-05-13 — bootstrap-with-wrong-state class",
     "4th accuracy audit cycle (2026-05-13). Caught n5-166 set-greetings meaning_ja contamination that 3 native-teacher + 3 accuracy audits missed. Lesson: install scripts that bootstrap from data inherit any contamination already present. JA-71/75 lock; JA-80 attempted then retired."),
    ("A", "feedback/audit-drift-findings-2026-05-12.md", "Audit-drift findings 2026-05-12 — drift between prompt CURRENT STATE and live data",
     "Captures every audit-refresh script encoded in this prompt that risks using stale counts. Background for the WRITING DISCIPLINE rewrite + the cross-artifact sync protocol (Rule 5) install."),
    ("J", "feedback/audit-round9-2026-05-06.md", "Audit round-9 2026-05-06 — VOICEVOX migration plan + voice variety",
     "Pre-VOICEVOX round-9 plan. Inventory of ISSUE-062 / 089 / 090 (voice variety) + per-item voice_planned IDs. Foundation for the 2026-05-12 render that BUG-047..053 later audited."),
    ("C", "feedback/hindi-audit-findings-2026-05-07.md", "Hindi audit findings 2026-05-07 — HI-NN classes + R-1.1 invariant",
     "Native-Hindi audit findings (HI-01..HI-23+). Documented the R-1.1 invariant (Japanese particles attached to Hindi terms must be kana, not Devanagari) locked by JA-41."),
    ("F", "feedback/legal-vetting-audit-2026-05-10.md", "Legal-vetting audit 2026-05-10 (F-1..F-13) — clean-room hygiene + license alignment",
     "First legal-vetting cycle. Findings F-1..F-13 covered: clean-room hygiene (no past-paper provenance signatures); CONTENT-LICENSE.md count parity (JA-47); KanjiVG CC-BY-SA attribution (JA-48); localStorage key vs PRIVACY.md (JA-37); trademark disclaimer; public-domain refs legal status (JA-69)."),
    ("F", "feedback/legal-vetting-audit-2026-05-13.md", "Legal-vetting audit 2026-05-13 — second cycle + cache-version sync",
     "Second legal-vetting cycle. Cache-version 3-place sync invariant (JA-68); locale-narrowing impact on legal posture; public-domain ref author_death_year buffer."),
    ("F", "feedback/legal-vetting-f6-spot-check-2026-05-11.md", "Legal-vetting F-6 spot-check 2026-05-11 — KanjiVG attribution detail",
     "Targeted F-6 follow-up. Verified the CC-BY-SA 3.0 attribution headers are preserved verbatim on every kanji SVG; documented in NOTICES.md."),
    ("C", "feedback/locale-transition-inventory.md", "Locale transition inventory 2026-05-06 — en→hi migration scope",
     "Round-8 locale transition inventory. Maps every UI/data surface that needs Hindi translation. Drove the locales/hi.json key set and JA-39 locale-set guard."),
    ("M", "feedback/MASTER-TASK-LIST.md", "MASTER-TASK-LIST.md — project-wide task inventory",
     "Cross-cutting master task list. Tracks every queued audit / fix / improvement across the project. Operations-tab catch-all for items not yet broken into specific scenarios."),
    ("J", "feedback/n5-richness-audit-2026-05-12.md", "n5 richness audit 2026-05-12 — Tofugu/Bunpro/Genki bar assessment",
     "Pedagogy depth audit. Per-pattern / per-vocab / per-kanji richness scoring against best-in-class incumbents. Drove the Section-2 richness scorecard structure in N5Improvement."),
    ("A", "feedback/native-teacher-audit-2026-05-08.md", "Native-teacher audit 2026-05-08 — pre-VOICEVOX baseline",
     "Pre-round-9 native-teacher audit. Caught 50+ accuracy issues before the VOICEVOX render."),
    ("A", "feedback/native-teacher-review-2026-05-13.md", "Native-teacher review 2026-05-13 (run 1) — F.17 bug classes",
     "Run-1 of the 2026-05-13 native-teacher pass. Surfaced BUG-003..009 bug classes (cross-pattern explanation contamination, mora-count systematic error, pattern-instance contamination, RIGHT/WRONG framing, folk-linguistic terminology, pattern-particle mismatch). Source for procedure manual F.17."),
    ("A", "feedback/native-teacher-review-2026-05-13-run2.md", "Native-teacher review 2026-05-13 run-2 — same-day re-audit",
     "Run-2 native-teacher pass on same day. Caught items run-1 missed; documented ad-hoc-sampling anti-pattern."),
    ("A", "feedback/native-teacher-review-2026-05-13-run3.md", "Native-teacher review 2026-05-13 run-3 — form-field coverage",
     "Run-3 native-teacher pass. Caught F-7 (520 missing form values). Documented in accuracy prompt's AD-HOC-SAMPLING ANTI-PATTERN section."),
    ("D", "feedback/ui-testing-plan.md", "UI testing plan §17.1 — P0 smoke + axe-core a11y",
     "Comprehensive UI testing plan: Playwright P0 smoke tests + axe-core accessibility checks. Includes BrowserStack matrix and visual-regression coverage. Cross-link: tab E (Accessibility) for a11y rules + tab K (QA testing) for execution."),
    ("I", "feedback/voicevox-integration-notes.md", "VOICEVOX integration notes — engine versioning + render workflow",
     "Integration notes for the local VOICEVOX engine. Documents speaker IDs, audio_query/synthesis endpoints, speed_scale parameter (key for BUG-049 fix). Cross-link: tab M (Operations) for re-render runbook."),
    # === feedback/closed/ ===
    ("I", "feedback/closed/coverage-comparison.md", "External-corpus coverage comparison — vocab/kanji/grammar parity",
     "Coverage analysis comparing our 178/1009/106/54/50 against external N5 reference corpora (kanjium, Genki, Bunpro, JLPT.jp legacy)."),
    ("A", "feedback/closed/f15-23-n5-167-nodesu-native-teacher-decision.md", "F-15.23 n5-167 nodesu native-teacher decision — kept-as-is",
     "Native-teacher decision on n5-167 (nodesu pattern). Reviewed and kept-as-is after consideration; documented rationale for future audit cycles."),
    ("A", "feedback/closed/f17-9-vocab-pos-tags-decision.md", "F-17.9 vocab POS tags decision — schema rationale",
     "Decision rationale for vocab.json POS tag schema (locked by JA-31). Documents why per-WORD tagging beats per-section default."),
    ("J", "feedback/closed/jlpt-n5-consolidated-audit-2026-05-01.md", "Consolidated N5 audit 2026-05-01 — initial broad review",
     "First broad-spectrum audit (2026-05-01) covering grammar / vocab / kanji / reading / listening / questions. Baseline for all subsequent targeted audits."),
    ("A", "feedback/closed/jlpt-n5-content-correction-brief.md", "N5 content correction brief — Japanese-language fixes",
     "Aggregated content corrections covering Japanese-language accuracy issues across all corpora."),
    ("I", "feedback/closed/jlpt-n5-content-files-audit-2026-05-03.md", "N5 content files audit 2026-05-03 — KB markdown review",
     "Knowledge-bank markdown audit (KnowledgeBank/ folder, before 2026-05-14 consolidation into data/* + docs/N5-syllabus-methodology.md)."),
    ("I", "feedback/closed/jlpt-n5-data-correction-brief.md", "N5 data correction brief — schema + count corrections",
     "Aggregated data-engineering corrections covering schema drift, count mismatches, and field-coverage gaps."),
    ("I", "feedback/closed/jlpt-n5-data-files-audit-2026-05-02.md", "N5 data files audit 2026-05-02 — initial schema review",
     "First data/*.json schema audit. Baseline for subsequent JA-* invariant introductions."),
    ("J", "feedback/closed/jlpt-n5-dossier-followup-audit-2026-05-02.md", "N5 dossier follow-up audit 2026-05-02 — pedagogy depth",
     "Follow-up review of the initial consolidated audit, focused on pedagogy depth gaps."),
    ("A", "feedback/closed/jlpt-n5-goi-audit-2026-05-04.md", "N5 goi audit 2026-05-04 — vocabulary accuracy",
     "Goi (vocabulary) focused audit. Surfaced BUG-014..018 lineage (template-leak nonsense, schema gaps, OOS kanji, dedup issues)."),
    ("D", "feedback/closed/jlpt-n5-homepage-update.md", "N5 homepage update — UX brief",
     "Homepage UI/UX redesign brief. Drove the index.html card-grid + recommender surface."),
    ("M", "feedback/closed/jlpt-n5-infrastructure-audit-2026-05-03.md", "N5 infrastructure audit 2026-05-03 — Ops review",
     "Deploy / CI / cache / SW infrastructure audit. Drove JA-68 (cache-version sync) and the build pipeline lock."),
    ("A", "feedback/closed/jlpt-n5-knowledgebank-md-audit-2026-05-01.md", "N5 KnowledgeBank md audit 2026-05-01 — content baseline",
     "KnowledgeBank/*.md baseline audit (before the 2026-05-14 consolidation). Drove the Japanese-content quality bar."),
    ("A", "feedback/closed/jlpt-n5-moji-and-source-audit-2026-05-03.md", "N5 moji and source audit 2026-05-03 — kanji + source provenance",
     "Moji (kanji) corpus audit + source-attribution review. Drove the n5_kanji_whitelist + the F-3 legal-vetting findings."),
    ("B", "feedback/closed/jlpt-n5-paper-files-audit-2026-05-03.md", "N5 paper files audit 2026-05-03 — JLPT mock-paper structure",
     "Mock-paper structure audit. Drove the papers/manifest.json schema and JA-32 (round-trip rationales) + JA-33 (mondai range)."),
    ("A", "feedback/closed/jlpt-n5-reading-feedback.md", "N5 reading feedback — dokkai accuracy + Hindi context",
     "Reading-corpus (dokkai) feedback. Drove JA-18/20/27/85 and the format_role schema."),
    ("I", "feedback/closed/jlpt-n5-reference-markdowns-audit-2026-05-04.md", "N5 reference markdowns audit 2026-05-04 — references review",
     "Reference markdown (vocabulary_n5.md / grammar_n5.md / etc.) audit. Drove JA-31 (vocab PoS parity) and the reference-vs-data sync."),
    ("M", "feedback/closed/jlpt-n5-tutor-developer-brief.md", "N5 tutor developer brief (EN) — Ops handoff",
     "Developer brief for ops/build/deploy. Covers the build pipeline (tools/build_*.py, npm run build:*, CI workflow). Cross-link: tab K (QA testing) for test scripts."),
    ("M", "feedback/closed/jlpt-n5-tutor-developer-brief.ja.md", "N5 tutor developer brief (JA) — operational handoff in Japanese",
     "Japanese-language developer brief. Mirror of the EN brief for native-Japanese contributors."),
    ("D", "feedback/closed/jlpt-n5-tutor-ux-developer-brief2.md", "N5 tutor UX developer brief v2 — UX hand-off",
     "Second-revision UX developer brief. Drove the responsive grid + the recommender card."),
    ("D", "feedback/closed/jlpt-n5-ui-design-brief.md", "N5 UI design brief — design-system + tokens",
     "UI design system brief. Locked the green palette (#14452a) + the design tokens used in css/design-tokens.css."),
    ("K", "feedback/closed/llm-audit-validation-report.md", "LLM audit validation report — false-positive characterization",
     "LLM-driven audit validation. Characterized false-positive classes that became FP-1..FP-15 in the accuracy prompt."),
    ("A", "feedback/closed/native-teacher-review-request.md", "Native-teacher review request — pre-2026-05-08 baseline",
     "Original native-teacher review request packet. Baseline for the native-teacher audit cycles that followed."),
    ("A", "feedback/closed/teacher-audit-2026-05-02.md", "Teacher audit 2026-05-02 — early Japanese-language review",
     "Early teacher-led Japanese-language audit. Baseline for the F.17 native-teacher bug classes."),
]

# --- Prompt-file summary scenarios for the 2 not-already-mapped prompts ---
PROMPT_SUMMARY_SCENARIOS = [
    ("F", "LegalVetting", "LegalVetting.txt — legal-posture audit prompt",
     "The legal-vetting audit prompt. Drives the F-1..F-13 finding classes covering clean-room hygiene, license alignment, attribution, trademark disclaimer, public-domain refs, and the 6 wired CI invariants (JA-37 / 47 / 48 / 68 / 69 — JA-30 for clean-room). Cross-link: tab G (Privacy and legal)."),
    ("G", "LegalVetting", "LegalVetting.txt — privacy/legal slant",
     "Same prompt as F-tab cross-link. From the privacy lens: localStorage namespace vs PRIVACY.md (JA-37), no remote-fetch code paths (JA-60), no gamification (JA-59), no per-content discussion (JA-61). Cross-link: tab F."),
    ("C", "LocaleTransitionEnHi", "LocaleTransitionEnHi.txt — en→hi locale transition prompt",
     "The locale-transition audit prompt. Drives the HI-NN findings + the R-1.1 invariant (Japanese particles attached to Hindi terms must be kana). Used during round-8 locale transition (2026-05-06)."),
]


def find_max_id_num(sh, prefix: str) -> int:
    """Return the max numeric suffix for ID with given prefix in sheet col 1."""
    n = 0
    for r in range(5, sh.max_row + 1):
        v = sh.cell(row=r, column=1).value
        if not v or not isinstance(v, str):
            continue
        if v.startswith(prefix + "-"):
            try:
                num = int(v.split("-")[-1])
                n = max(n, num)
            except ValueError:
                continue
    return n


def existing_ids_set(sh) -> set[str]:
    s = set()
    for r in range(5, sh.max_row + 1):
        v = sh.cell(row=r, column=1).value
        if v and isinstance(v, str):
            s.add(v)
    return s


def existing_refs_in_tab(sh) -> set[str]:
    """Pull out every A-NN / P0-* / FP-NN / doc-name token from existing rows."""
    refs = set()
    for r in range(5, sh.max_row + 1):
        for c in [4, 5, 6, 10]:  # Scenario, Test steps, Expected, Notes
            v = sh.cell(row=r, column=c).value
            if isinstance(v, str):
                refs.add(v)
    return refs


def append_scenario(
    sh,
    sid: str,
    sub_category: str,
    persp: str,
    scenario: str,
    test_steps: str,
    expected: str,
    priority: str,
    severity: str,
    test_type: str,
    notes: str,
    effort: str,
    owner: str,
    tools: str,
    last_run_date: str,
    last_run_result: str,
    tracker: str,
    depends_on: str,
    coverage: str,
) -> int:
    """Append a scenario row at sh.max_row + 1; return the row written."""
    row = sh.max_row + 1
    # Defensive: skip blank trailing rows
    while row > 5 and not sh.cell(row=row - 1, column=1).value:
        row -= 1
    values = [sid, sub_category, persp, scenario, test_steps, expected,
              priority, severity, test_type, notes, effort, owner, tools,
              last_run_date, last_run_result, tracker, depends_on, coverage]
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for c, v in enumerate(values, start=1):
        cell = sh.cell(row=row, column=c, value=v)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
    return row


TAB_NAMES = {
    "A": "A. Japanese language",
    "B": "B. JLPT format",
    "C": "C. Hindi locale",
    "D": "D. UX design",
    "E": "E. Accessibility",
    "F": "F. Security",
    "G": "G. Privacy and legal",
    "H": "H. Performance",
    "I": "I. Data engineering",
    "J": "J. Pedagogy",
    "K": "K. QA testing",
    "L": "L. Cultural ethical",
    "M": "M. Operations",
    "N": "N. End-user POV",
}


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1

    wb = openpyxl.load_workbook(XLSX)
    print(f"Loaded {XLSX.name}")
    added_total = 0

    # -------------------------------------------------------------------
    # PASS 1 — A-NN audit categories → tab A (60 items)
    # -------------------------------------------------------------------
    sh_a = wb[TAB_NAMES["A"]]
    max_a = find_max_id_num(sh_a, "A")
    a_existing = existing_ids_set(sh_a)
    a_refs = existing_refs_in_tab(sh_a)
    next_a = max_a + 1
    added = 0
    for code, title, hint in A_NN_CATEGORIES:
        # Build the row
        sid = f"A-{next_a:03d}"
        # Skip if any existing row references this code in Scenario or Notes
        # Word-boundary match — avoid substring false-positives where
        # e.g. "A55" causes "A5" to be skipped. (This bug was caught
        # 2026-05-17 by JA-116; A5 was the only victim — fixed in
        # commit that introduced JA-116.)
        already = any(re.search(rf"\b{re.escape(code)}\b", ref) for ref in a_refs if isinstance(ref, str))
        if already:
            continue
        scenario = f"Accuracy prompt audit category {code}: {title}"
        test_steps = (
            f"1. Reference: prompts/Japanese language Accuracy check.txt {code}\n"
            f"2. Run the category's manual sample / CI check\n"
            f"3. Cross-reference any JA-NN invariant noted below\n"
            f"4. Report findings as candidate bug rows"
        )
        expected = (
            "0 findings against current corpus snapshot OR documented "
            "false-positive class (see FP-NN catalog in same prompt)."
        )
        notes = (
            f"Cross-Artifact Sync 2026-05-17: per Rule 5 prompt → test-scenario "
            f"sync. Source: prompts/Japanese language Accuracy check.txt {code}. "
            f"Hint: {hint}"
        )
        append_scenario(
            sh_a, sid, "Accuracy prompt category", "1", scenario, test_steps, expected,
            "P3", "Major", "Manual",
            notes, "30m", "Native-teacher reviewer", "prompts/Japanese language Accuracy check.txt",
            "", "Not Yet Run", "", code, "0%",
        )
        next_a += 1
        added += 1
    print(f"  Tab A: appended {added} A-NN audit-category scenarios")
    added_total += added

    # -------------------------------------------------------------------
    # PASS 2 — Phase-0 regression blocks → tab K (18 items)
    # -------------------------------------------------------------------
    sh_k = wb[TAB_NAMES["K"]]
    max_k = find_max_id_num(sh_k, "K")
    k_refs = existing_refs_in_tab(sh_k)
    next_k = max_k + 1
    added = 0
    for code, title in PHASE0_BLOCKS:
        already = any(title.lower()[:30] in (ref or "").lower() for ref in k_refs if isinstance(ref, str))
        if already:
            continue
        sid = f"K-{next_k:03d}"
        scenario = f"N5Improvement Phase-0 regression block: {title}"
        test_steps = (
            "1. Reference: prompts/N5Improvement.txt §'" + title + "'\n"
            "2. Run the Python regression block inline at the start of the audit\n"
            "3. Expect 0/0/0/... across every check\n"
            "4. Surface any non-zero count as a candidate bug row"
        )
        expected = "Every check in the block reports 0 against current corpus snapshot."
        notes = (
            f"Cross-Artifact Sync 2026-05-17: per Rule 5 prompt → test-scenario "
            f"sync. Source: prompts/N5Improvement.txt {title!r} block. "
            f"Phase-0 blocks are the standing regression checks; CI wires them "
            f"as JA-NN invariants where applicable."
        )
        append_scenario(
            sh_k, sid, "Phase-0 regression block", "65", scenario, test_steps, expected,
            "P2", "Major", "Auto",
            notes, "10m", "Build engineer", "prompts/N5Improvement.txt; tools/check_content_integrity.py",
            "", "Not Yet Run", "", code, "100%",
        )
        next_k += 1
        added += 1
    print(f"  Tab K: appended {added} Phase-0 regression-block scenarios")
    added_total += added

    # -------------------------------------------------------------------
    # PASS 3 — FP-NN false-positive class entries → tab K (15 items)
    # -------------------------------------------------------------------
    k_refs = existing_refs_in_tab(sh_k)  # refresh
    added = 0
    for code, title in FP_ENTRIES:
        already = any(code in (ref or "") for ref in k_refs if isinstance(ref, str))
        if already:
            continue
        sid = f"K-{next_k:03d}"
        scenario = f"Accuracy prompt false-positive class {code}: {title[:80]}"
        test_steps = (
            f"1. Reference: prompts/Japanese language Accuracy check.txt {code}\n"
            f"2. During audit, when a candidate finding matches this FP shape, "
            f"DO NOT FILE\n"
            f"3. If a finding DOES file under this shape, flag the FP entry as "
            f"needing refinement (true-positive escape)"
        )
        expected = (
            f"0 false-positive findings filed under shape {code}. If the audit "
            f"misses this FP shape, refine the FP catalog or the audit prompt."
        )
        notes = (
            f"Cross-Artifact Sync 2026-05-17: per Rule 5 prompt → test-scenario "
            f"sync. Source: prompts/Japanese language Accuracy check.txt {code} "
            f"(false-positive class catalog). These are documented NON-bugs that "
            f"auditors must learn to recognize and skip."
        )
        append_scenario(
            sh_k, sid, "False-positive class", "65", scenario, test_steps, expected,
            "P4", "Minor", "Manual",
            notes, "15m", "Audit reviewer", "prompts/Japanese language Accuracy check.txt",
            "", "Not Yet Run", "", code, "0%",
        )
        next_k += 1
        added += 1
    print(f"  Tab K: appended {added} FP-NN false-positive scenarios")
    added_total += added

    # -------------------------------------------------------------------
    # PASS 4 — feedback / closed audit docs → distributed by tab
    # -------------------------------------------------------------------
    # Group per tab
    per_tab: dict[str, list] = {letter: [] for letter in TAB_NAMES.keys()}
    for tab, doc_rel, title, body in AUDIT_DOCS:
        per_tab[tab].append((doc_rel, title, body))
    for tab, doc_rel, title, body in PROMPT_SUMMARY_SCENARIOS:
        per_tab[tab].append((doc_rel, title, body))

    for tab_letter, items in per_tab.items():
        if not items:
            continue
        sh = wb[TAB_NAMES[tab_letter]]
        max_n = find_max_id_num(sh, tab_letter)
        refs = existing_refs_in_tab(sh)
        next_n = max_n + 1
        added = 0
        for doc_rel, title, body in items:
            # Check if any existing row mentions this doc path
            doc_key = Path(doc_rel).name
            already = any(doc_key.lower() in (ref or "").lower() for ref in refs if isinstance(ref, str))
            if already:
                continue
            sid = f"{tab_letter}-{next_n:03d}"
            scenario = title
            test_steps = (
                f"1. Read N5/{doc_rel}\n"
                f"2. Confirm every finding-class / check / recommendation in the doc "
                f"is either (a) addressed in a downstream fix script, (b) wired as a "
                f"CI invariant, or (c) documented as an open / deferred action\n"
                f"3. Surface any unaddressed item as a candidate bug or improvement"
            )
            expected = (
                "Every actionable item in the doc maps to a downstream artifact "
                "(fix commit / CI invariant / open tracker entry). 0 unmapped items."
            )
            notes = (
                f"Cross-Artifact Sync 2026-05-17: per Rule 5 feedback/prompts → "
                f"test-scenario sync. Source: N5/{doc_rel}. {body}"
            )
            append_scenario(
                sh, sid, "Audit-doc summary", "—", scenario, test_steps, expected,
                "P4", "Minor", "Manual",
                notes, "1h", "Audit reviewer", f"N5/{doc_rel}",
                "", "Not Yet Run", "", "—", "0%",
            )
            next_n += 1
            added += 1
        if added:
            print(f"  Tab {tab_letter}: appended {added} audit-doc summary scenarios")
        added_total += added

    print()
    print(f"TOTAL added: {added_total} scenarios across all tabs")
    wb.save(XLSX)
    print(f"Saved {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
