#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Register MOB-020 — narrow primary-nav links (40px width, below HIG 44 min).

Discovered by re-running tools/run_mobile_ui_tests.py on 2026-05-21
against the post-MOB-001..016-fix state. The fix correctly resolved
display:none on `test` + `summary` (all 9 nav links now visible) and
added `min-height: 44px`, but the 9 links share 360px viewport width
equally — yielding 40px width per item, 91% of HIG 44px minimum on the
width dimension.

Status: Open (per maintainer convention: register, don't fix).
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()
REPORTER = "Mobile UI Selenium re-run (2026-05-21) — tools/run_mobile_ui_tests.py post-MOB-001..016 fix verification"

TITLE = (
    "MOB-020 — Primary nav links render at 40x44 px width on D-360 "
    "(9 links sharing 360px viewport equally; height now compliant per "
    "MOB-001..016 fix but width is 91% of HIG 44 minimum)"
)
DESCRIPTION = (
    "ROUTE: # (every page header) on D-360 (360x800 mobile-emulated).\n\n"
    "EVIDENCE: getBoundingClientRect for all 9 a[data-route='*'] in "
    ".primary-nav:\n"
    "  learn/grammar  40x44\n"
    "  learn/vocab    40x44\n"
    "  kanji          40x44\n"
    "  reading        40x44\n"
    "  listening      40x44\n"
    "  test           40x44  (newly visible — MOB-001 fix)\n"
    "  sitting        40x44\n"
    "  missed         40x44\n"
    "  summary        40x44  (newly visible — MOB-001 fix)\n\n"
    "STATE CHANGE FROM MOB-001 (registered 2026-05-19):\n"
    "  Before fix: test + summary had display:none (0x0); other 7 at "
    "65x44 (passing HIG).\n"
    "  After MOB-001..016 fix batch (commits between 2026-05-19 and "
    "2026-05-21 by external workflow): all 9 visible, display:flex; "
    "min-height:44px applied (HIG-compliant on height). But width "
    "now 40px because 9 items share 360px viewport equally "
    "(360 / 9 = 40).\n\n"
    "ASSESSMENT:\n"
    "  - Height: 44px — HIG-compliant (matches NFR-M1 floor)\n"
    "  - Width:  40px — 91% of HIG 44 minimum; just 4px short\n"
    "  - Material Design min is 48; width is 83% of that\n"
    "  - Visually: each nav link is a 40px square cell; touch-target "
    "    feel is acceptable but not generous\n\n"
    "POSSIBLE FIXES (do NOT apply per maintainer instruction):\n"
    "  Option A: Add a hamburger/overflow menu at <= 480px breakpoint, "
    "            keep 4-5 primary items in the main nav, each at "
    "            ~72px width.\n"
    "  Option B: Horizontal-scroll the .primary-nav at <=480px so each "
    "            item gets its full natural width (~65-80px).\n"
    "  Option C: Stack nav across 2 rows at <=480px (each row 4-5 "
    "            items, each ~72px wide).\n"
    "  Option D: Accept 40px width as documented exception. HIG 44 "
    "            is a guideline; 40px is 91% of it and the touch "
    "            target is still operable. Document this as a "
    "            scope-bounded gap rather than fix.\n\n"
    "REPRO: GET / at D-360; querySelectorAll('a[data-route]'); "
    "every visible nav link has rect.width == 40 (regardless of "
    "min-width / padding values).\n\n"
    "RELATED:\n"
    "  - MOB-001 (R113) — original 'test + summary display:none' bug; "
    "    closed by the MOB-001..016 fix batch\n"
    "  - JA-132 (CI invariant added 2026-05-19 for MOB-002..016) — "
    "    enforces min-height: 44px on .study-order-link, .btn-action, "
    "    etc.; does NOT cover the per-nav-link width dimension at "
    "    narrow viewports.\n\n"
    "SCENARIO IT FAILS: O-X-021 in `O. Mobile UI testing` workbook tab.\n\n"
    "PRIORITY rationale: Minor / P3. Touch targets are still operable "
    "at 40px (vs 44px HIG min). The original MOB-001 problem (2 nav "
    "items completely missing on mobile) is resolved; this is the "
    "secondary follow-on — narrower-than-ideal but visible-and-tappable."
)


def main() -> None:
    today_str = date.today().strftime("%Y_%m_%d")
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{today_str}_mob_020")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)
    ws = wb["User Reported Bugs"]

    # Idempotency
    for r in range(4, ws.max_row + 1):
        t = ws.cell(row=r, column=4).value or ""
        if t.startswith("MOB-020"):
            print(f"  SKIP: MOB-020 already exists at R{r}")
            return

    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=f'="BUG-"&TEXT(ROW()-3,"000")')
    ws.cell(row=new_row, column=2, value=TODAY)
    ws.cell(row=new_row, column=3, value=REPORTER)
    ws.cell(row=new_row, column=4, value=TITLE)
    ws.cell(row=new_row, column=5, value=DESCRIPTION)
    ws.cell(row=new_row, column=6, value="Minor")
    ws.cell(row=new_row, column=7, value="P3")
    ws.cell(row=new_row, column=8, value="Open")
    ws.cell(row=new_row, column=9, value="")
    ws.cell(row=new_row, column=10, value="")
    for c in range(1, 11):
        ws.cell(row=new_row, column=c).alignment = WRAP
    ws.row_dimensions[new_row].height = 280
    wb.save(WORKBOOK)
    print(f"  WROTE R{new_row}: MOB-020")


if __name__ == "__main__":
    main()
