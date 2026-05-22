"""Flip BUG-170/171/172/173 (NTR-010/011/012/013) to Fixed."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_flip_ntr_010_to_013"
if not os.path.exists(BAK): shutil.copy2(XLSX, BAK)
wb = load_workbook(XLSX); ws = wb["User Reported Bugs"]
FLIPS = {
    173: ("NTR-010 — Appended particle-nuance note to q-0226 explanation_en: 'は after a "
          "time noun introduces contrast (yesterday in particular — implying other days are "
          "different). Fine here, but typical neutral past-negative would drop the particle "
          "(きのう ばんごはんを 食べませんでした) or use に.' Protects learners who notice "
          "the wrinkle. explanation_provenance bumped to native_reviewed_2026_05_22."),
    174: ("NTR-011 — Renamed `collocations` → `particle_examples` on 12 pronoun entries "
          "(section '1. People - Pronouns and Self'). More honest terminology: the field "
          "contains particle-template-illustrative substitutions (わたしの ともだち / わたしは "
          "がくせい / わたしも いく / etc.), not real corpus-linguistics collocations. Real "
          "collocations can override the field where they exist. particle_examples_provenance "
          "documents the rename rationale. UI consumers reading the old `collocations` key "
          "will need updating; small JS follow-up."),
    175: ("NTR-012 — Appended なな-usage note to 七 reading_rule: 'なな is heard in "
          "ordering/listing contexts (especially to disambiguate from いち over the phone); "
          "しち in fixed time/date compounds (七時 しちじ, 七月 しちがつ). NHK uses なな when "
          "reading numerals aloud. 七人 takes either reading; 七つ is exclusively ななつ. For "
          "phone/list disambiguation, prefer なな.' Preserves the prescriptive primary "
          "(しち) while documenting the colloquial-counting reality. reading_rule_provenance "
          "bumped to native_reviewed_2026_05_22."),
    176: ("NTR-013 — Two fixes: (a) みなさん.counter.reading typo '人' (kanji) → 'にん' "
          "(kana), matching the other 8 pronoun entries; (b) collective pronouns (私たち, "
          "みなさん) gain counter.applies_to='noun_of_reference' + a note that the counter "
          "applies to the people the pronoun refers to, not to the pronoun itself. "
          "Individual-referring pronouns (私, あなた, かれ, かのじょ, かた, 人) retain counter "
          "as-is. UI may choose to suppress counter display on collective pronouns based on "
          "the applies_to flag."),
}
for row, note in FLIPS.items():
    ws.cell(row=row, column=8, value="Fixed")
    ws.cell(row=row, column=9, value="2026-05-22")
    ws.cell(row=row, column=10, value="<pending-main-commit>")
    ws.cell(row=row, column=11, value=note)
    print(f"  R{row}: Fixed")
wb.save(XLSX)
