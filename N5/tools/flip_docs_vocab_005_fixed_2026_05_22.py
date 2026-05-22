"""Flip BUG-136 (DOCS-VOCAB-005) to Fixed + backfill DOCS-VOCAB-003
follow-up cross-reference in Fix Notes."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_flip_docs_vocab_005"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]

# Columns 8=Status, 9=Fix Date, 10=Fix Commit, 11=Fix Notes

# BUG-136 (DOCS-VOCAB-005) at row 139 → Fixed
ws.cell(row=139, column=8, value="Fixed")
ws.cell(row=139, column=9, value="2026-05-22")
ws.cell(row=139, column=10, value="<pending-main-commit>")
ws.cell(row=139, column=11, value=(
    "Trimmed source_file prose to literal canonical sentinel "
    "'(authored in-place)' across all 28 paper files in "
    "data/papers/{bunpou,dokkai,goi,moji}/paper-{1..7}.json. "
    "Bug-spec verification rejected the original proposed fix "
    "(replace with docs/N5-syllabus-methodology.md#bunpou-questions "
    "and parallel anchors) because those fragment IDs don't exist "
    "in the methodology doc (which has Part C/D/E/F headings, not "
    "per-category question content) and because pointing source_file "
    "at the doc would falsely imply it contains the questions. User "
    "confirmed Option 1 (canonical-sentinel trim). Wired JA-145: "
    "source_file must be either the literal sentinel OR a resolvable "
    "repo path. Bumped version.json v1.15.5 → v1.15.6 + cacheVersion. "
    "Closes the unaddressed half of DOCS-VOCAB-003 (case (b)) which "
    "had closed prematurely as a README-only fix (case (a) only). "
    "Tools: tools/fix_docs_vocab_005_2026_05_22.py + "
    "tools/file_docs_vocab_005_bug_2026_05_22.py + "
    "tools/audit_docs_vocab_005_2026_05_22.py + "
    "tools/audit_paper_kb_refs_2026_05_22.py. Procedure manual F.41 "
    "(canonical-sentinel pattern + multi-case-bug close-out discipline)."
))

# Backfill DOCS-VOCAB-003 (row 153, BUG-150) — Fix Notes was empty.
# Now we know what actually happened: only case (a) was addressed,
# case (b) was filed-as-follow-up under DOCS-VOCAB-005.
prior_notes_153 = ws.cell(row=153, column=11).value or ""
ws.cell(row=153, column=11, value=(
    "Multi-case bug: case (a) - README reworded to firmly state "
    "KB was deleted 2026-05-14; case (b) - update 28 paper-file "
    "source_file references - NOT addressed in this fix. "
    "Case (b) re-filed-as-follow-up under DOCS-VOCAB-005 "
    "(BUG-136) on 2026-05-22 and closed there with the "
    "canonical-sentinel pattern. Operational lesson "
    "(procedure manual F.41.2): multi-case bug close-outs MUST "
    "explicitly state which case(s) were addressed and what "
    "happened to the rest. This row's prior empty Fix Notes "
    "violated that discipline; back-filled 2026-05-22."
    + ("\n\nPrior Fix Notes: " + prior_notes_153 if prior_notes_153 else "")
))

# Also fix the wrong "2026-05-21" in Fix Commit cell (it's a date string
# where a commit hash should be). Set to <unknown> with note. Don't
# fabricate a hash.
prior_commit_153 = ws.cell(row=153, column=10).value
if prior_commit_153 == "2026-05-21":
    ws.cell(row=153, column=10, value="<unknown-from-2026-05-21-batch; case (b) follow-up commit on DOCS-VOCAB-005>")
    print(f"  Backfilled DOCS-VOCAB-003 Fix Commit cell (was date string, not hash)")

# Also backfill Fix Date if empty
if not ws.cell(row=153, column=9).value:
    ws.cell(row=153, column=9, value="2026-05-21")
    print(f"  Backfilled DOCS-VOCAB-003 Fix Date (was empty)")

wb.save(XLSX)

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
print()
print("=== Verified state ===")
for r in [139, 153]:
    print(f"  R{r}: status={ws2.cell(row=r, column=8).value} date={ws2.cell(row=r, column=9).value} commit={str(ws2.cell(row=r, column=10).value)[:50]}")
print("Saved.")
