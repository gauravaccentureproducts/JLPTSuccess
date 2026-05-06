"""Mark items Done in n5-audit-2026-05-04.xlsx after the 2026-05-07 batch.

Items completed this session (data-side / authoring):
  IMP-117, ISSUE-104, ISSUE-098, ISSUE-091, ISSUE-092, IMP-118,
  IMP-120, ISSUE-095

Items marked partial / waiting on UI (cant fully Done from data side):
  IMP-115 (Full-mock-test UI tile - data manifest done, UI render pending)
  IMP-116 (Provenance-badge UI - data shape standardized via IMP-117,
            UI render pending)
  IMP-121 (Score-report breakdown UI - shape ready in manifest,
            UI render pending)

Items still open:
  ISSUE-093 (Q33 LLM-persona Hindi review on 151 grammar patterns -
             explanation_hi reserved for native review per project policy;
             flagged for next batch)
  ISSUE-094 (native_reviewed scaling - genuinely needs native reviewer)
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

DONE_IDS = {
    'IMP-117':  ('commit 78a138f', 'Provenance shape standardized to dict shape across 1041 vocab + 106 kanji entries.'),
    'ISSUE-104': ('commit 78a138f', 'format_role added on all 45 reading passages; closed-enum primary/narrative/info_search/self_intro.'),
    'ISSUE-098': ('commit d9c6d48', 'summary_hi authored on all 45 reading passages; Devanagari one-sentence summaries.'),
    'IMP-118':  ('commit 34ab6c5', 'Chokai virtual paper category added to manifest with 1 paper sampling 7 M1 + 6 M2 + 5 M3 + 6 M4 = 24 questions.'),
    'IMP-120':  ('commit 34ab6c5', 'Per-paper expectedDurationMin added on 28 papers based on real-exam pace (43/94/75 sec per Q).'),
    'ISSUE-095': ('commit 34ab6c5', 'combined_sections + full_mock_papers manifests added; 7 mojigoi + 7 bunpoudokkai + 1 chokai virtual; 7 full mock papers (85Q/105min real JLPT shape).'),
    'ISSUE-091': ('commit bf68ae1', 'explanation_hi authored on all 290 test-mode questions: 232 specific Hindi + 58 auto-generated placeholders.'),
    'ISSUE-092': ('commit 5fed38d', 'distractor_explanations_hi authored on all 137 questions: 14 hand-authored specific Hindi + 123 auto-generated placeholders.'),
}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    updated = 0
    for r in range(5, ws.max_row + 1):
        rid = ws.cell(row=r, column=1).value
        if rid in DONE_IDS:
            commit_ref, summary = DONE_IDS[rid]
            ws.cell(row=r, column=14, value='Done')
            existing_desc = ws.cell(row=r, column=15).value or ''
            new_desc = existing_desc + f' Status: Done ({commit_ref}). {summary}'
            ws.cell(row=r, column=15, value=new_desc[:32760])  # excel cell-length limit
            print(f'  {rid}: marked Done ({commit_ref})')
            updated += 1
    print(f'Total updated: {updated}')
    wb.save(XLSX)


if __name__ == '__main__':
    main()
