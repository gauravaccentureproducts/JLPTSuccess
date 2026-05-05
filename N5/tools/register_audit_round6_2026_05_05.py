"""Round-6 audit registrar — append the 6 new issues + 6 new improvements
+ 3 new open questions to feedback/n5-audit-2026-05-04.xlsx.

Round-6 ran the same SALEABILITY / NICHE-FIT prompt against the
post-v1.12.36 state. State delta from round-5:
  - v1.12.32 → v1.12.36 (4 incremental releases)
  - IMP-066 closed (GitHub repo metadata + v1.12.34 release)
  - IMP-045/046/047 partial: kanji 100%, vocab 12%, grammar 0%
  - Locale chips now visibly translate home + primary-nav (hotfix)
  - JS bundle minified -57%; WebP icons; reading prev/next nav

The audit identifies 6 fresh gaps, mostly around the partial-localization
UX (post-hotfix realization that Settings/Test/Drill/etc. still hardcode
English while home translates), and a few smaller polish items.

Idempotent. Uses the round-3-style Plain-English column (already on
the Items sheet from the round-5 restructure) and the Permission-
decision dropdown.
"""
from __future__ import annotations
import io, sys
from copy import copy
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# 14-column rows + Plain English (col 15) + Permission decision (col 16).
ISSUES = [
    ('ISSUE-048', 'Issue', 'MAJOR', 'P1', 'HIGH', 'MEDIUM',
     'i18n / UX continuity',
     'js/settings.js + js/test.js + js/drill.js + js/review.js + js/summary.js + js/diagnostic.js',
     '[N1] Settings/Test/Drill/Review/Summary/Diagnostic still hardcode English — partial-localization regression',
     'Round-5 hotfix wired t() into home + nav. These 6 modules still have ~80 hardcoded EN strings.',
     'Vietnamese visitor sees home in Vietnamese, clicks Settings, page reverts to English. Partial localization implies brokenness, not coverage.',
     'Same pattern as the round-5 hotfix — import t() and replace literal strings. About 25 already-defined keys cover most cases; ~30 more keys to add.',
     '—', '',
     'Settings page + test setup + drill + review + summary + diagnostic pages still show English text even after switching to Vietnamese/Indonesian/etc. Home page translates correctly but these other pages do not. Result: confusing partial-translation UX.',
     ''),
    ('ISSUE-049', 'Issue', 'MAJOR', 'P2', 'HIGH', 'HIGH',
     'Content / i18n',
     'data/vocab.json — 128/1041 entries carry gloss_<lc>',
     '[N1] Vocab translation at 12% looks broken — most words show EN even when in non-EN locale',
     'Round-5 IMP-046 translated top 128 most-common vocab. Remaining 913 fall back to English, creating a sprinkled "some translated" pattern.',
     'Q21 badge launch policy waits for ≥10% native_reviewed; vocab at 12% machine_translated still feels broken. UX worse-than-zero until coverage feels deliberate.',
     'Either extend translation to ≥50% (push harder), OR hide partial translations behind a "Show partial translations" Settings toggle.',
     'IMP-046, Q20', '',
     'Open the vocabulary list while on Vietnamese/Indonesian: about 1 in 8 words shows in your language, the other 7 show in English. Looks like a bug to a learner. Either translate ~520+ entries to feel "done" OR hide the few translations until coverage is high.',
     ''),
    ('ISSUE-050', 'Issue', 'MINOR', 'P2', 'MEDIUM', 'LOW',
     'i18n / discoverability',
     'js/i18n.js _flashAutoLocaleToast',
     '[N1] Auto-locale toast fires once; no other in-app discovery of locale chips',
     'After the one-shot toast on first init, a non-EN visitor must notice the small CAPS chip group in the header to switch language.',
     'Niche N1 hinges on discoverability. The 30px chip group is easy to miss on mobile.',
     'Add a permanent "Switch language" entry in the footer linking to the chip group. Re-fire toast on first visit per session.',
     '—', '',
     'When you first visit the app in Indonesian, a small toast appears once saying "App language: Bahasa Indonesia". After that toast disappears, there is no obvious place to change language. The 5 small chips (EN/VI/ID/NE/ZH) at the top are easy to miss.',
     ''),
    ('ISSUE-051', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'OSS hygiene / institutional',
     '.github/FUNDING.yml (missing)',
     '[N3] No .github/FUNDING.yml — GitHub Sponsors button missing',
     'GitHub looks for this file at .github/ root to render a "Sponsor" button on the repo page.',
     'Niche N3 institutional adopters scan for sponsorship signals as project sustainability indicator. Low-cost trust signal.',
     'Add .github/FUNDING.yml; either configure a Sponsors profile or ship a comment-only placeholder.',
     '—', '',
     'GitHub looks for a special FUNDING.yml file to show a "Sponsor this project" button on the repo. We do not have one. This is a tiny file but signals project health to institutional adopters.',
     ''),
    ('ISSUE-052', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'Build / debug',
     'tools/build_min_js.py esbuild invocation',
     '[none] JS minify produces no source maps — production debugging difficult',
     'esbuild invoked without --sourcemap. Stack traces from the minified bundle show line numbers from minified files only.',
     'Niche-N3 self-hosters debugging an issue cannot easily map back to source. Sourcemaps do not ship to clients automatically (devtools fetches on demand).',
     'Add --sourcemap=external to the esbuild invocation; precache the .map files alongside .js.',
     '—', '',
     'When the minified production app crashes, the error message points to a hard-to-read line number in the minified file. Adding source maps lets developers see the original code line in their browser developer tools. The map files do not load for normal users.',
     ''),
    ('ISSUE-053', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'Trust-signal infra',
     'no module yet',
     '[none] Q21 badge UI policy has no code skeleton — when kanji crosses 10% threshold, no infra exists',
     'Q21 documented "≥10% native_reviewed per corpus before badge UI ships." But no JS reads review_status today.',
     'When the first reviewer pass lands and kanji crosses 11+ native_reviewed entries, the project-side response is "now write the badge UI from scratch." Better pre-wired and feature-flagged.',
     'Add a stub js/provenance-badge.js reading review_status; render a pill behind a settings flag defaulting to false. Flip the flag when threshold crosses.',
     'Q21', '',
     'We have a policy: when ≥10% of any content type is reviewed by a native speaker, show a "Native-reviewed" badge. But no code currently reads the review_status field. When that day comes, we would have to write the badge UI from scratch. Better to write a feature-flagged version now.',
     ''),
]

IMPS = [
    ('IMP-069', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Discoverability / contributor',
     'js/settings.js + index.html footer-nav',
     '[N1] "Help translate" link to surface translator recruitment',
     'docs/TRANSLATING.md Q20 recruitment is buried in a docs file. Native speakers do not see the CTA unless browsing docs.',
     'Best-in-class: Wikipedia surfaces "Help translate" in locale banners; Mozilla shows it in feedback flows. JLPTSuccess can match cheaply.',
     'Add "Help translate" link to (a) footer-nav, (b) Settings → Display. Both link to docs/TRANSLATING.md on GitHub.',
     'ISSUE-048', '',
     'A small "Help translate" link in the footer + Settings page would surface the open-source translator recruitment so native speakers actually see it. Currently buried in a docs file nobody reads.',
     ''),
    ('IMP-070', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'HIGH',
     'Listening UX',
     'js/listening.js',
     '[N4] Listening transcript-aligned playback (highlight current line)',
     'Listening items play as single audio track. No alignment between audible position and script text.',
     'Best-in-class: NHK Easy News, Migaku, BookBeat. Current-line highlighting is meaningful study aid.',
     'Add per-line timestamps to data/listening.json + a JS overlay that highlights via setInterval polling audio.currentTime.',
     'IMP-042 (native audio recording with timestamps)', '',
     'When listening to a Japanese audio clip, learners benefit from seeing the line of text being spoken highlighted in real-time. Apps like NHK Easy News do this. Requires per-sentence timestamps (which we do not have yet, but could add when we record native audio).',
     ''),
    ('IMP-071', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'i18n / discoverability',
     'js/learn-vocab.js list view',
     '[N1] Section-level translation coverage indicators in vocab list',
     'Vocab list shows 1041 entries with no indication which are translated for the active locale.',
     'Best-in-class: Crowdin / Weblate dashboards. Mozilla L10n project shows per-section coverage so volunteers focus their time.',
     'Add a small per-section "X% translated" badge in vocab list section headers when on a non-EN locale.',
     'ISSUE-048', '',
     'When a Vietnamese reviewer browses the vocab list, they should see "Numbers: 100% translated, Verbs: 0% translated, Family: 60%" so they know where to focus. Currently no such indicator.',
     ''),
    ('IMP-072', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'Build / debug',
     'tools/build_min_js.py',
     '[none] Sourcemap support in JS minifier — paired with ISSUE-052',
     'No sourcemap output from esbuild today.',
     'Best-in-class: every JS toolchain ships sourcemaps in 2025. Free, lazy-loaded by devtools.',
     'Add --sourcemap=external. SW precache handles the .map files.',
     'ISSUE-052', '',
     'See ISSUE-052. Source maps make production stack traces readable in browser developer tools without bloating the user-facing bundle.',
     ''),
    ('IMP-073', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'Discoverability / SEO',
     'README.md',
     '[N3] Add badges row (license / version / stars / last-commit / JLPT-level)',
     'README has no shields.io badges. Niche-N3 institutional adopters scan badge rows for project health.',
     'Best-in-class: every popular OSS repo. Shields.io free.',
     'Add 4-5 badges at the top of README.md.',
     'ISSUE-051', '',
     'Add a row of small status badges at the top of README.md showing license, version, GitHub stars, last commit. Common for open-source projects; signals project health at a glance.',
     ''),
    ('IMP-074', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'i18n',
     'js/i18n.js#initI18n',
     '[N1] Language detection from referrer (privacy-preserving)',
     'Locale picked from navigator.language only; document.referrer not consulted.',
     'Best-in-class: Wikipedia uses referrer + Accept-Language for L1 detection. Privacy-preserving (referrer is per-request).',
     'In initI18n, fold referrer-domain hints into the locale-pick heuristic (.vn → boost VI; .id → boost ID).',
     '—', '',
     'If a Vietnamese visitor arrives from a .vn website, we should default the app to Vietnamese even if their browser is set to English. Privacy-friendly because the referring URL is not stored.',
     ''),
]

QUESTIONS = [
    ('Q24', 'Settings/test/drill localization scope',
     'ISSUE-048 wires t() into 6 hardcoded-EN modules. Three paths: (a) full pass on all 6 modules (hundreds more keys), (b) Settings only (highest-traffic, fastest visible win), (c) wait until a native reviewer asks for it.',
     'Pick a/b/c.',
     '',
     'Three options for fixing ISSUE-048: (a) translate every page (slowest), (b) translate just the Settings page (most visited), (c) wait until someone asks.'),
    ('Q25', 'Vocab translation completion target',
     'ISSUE-049 says vocab at 12% looks broken. Either push higher or hide. Pick a target: 50% / 75% / 100%, OR commit to "hide until 50% reached."',
     'Pick a coverage target or hide-until threshold.',
     '',
     'Vocab is 12% translated. Pick: should I keep translating to 50%, 75%, or 100%? OR should I hide the partial translations until they reach a threshold?'),
    ('Q26', 'Source maps in production',
     'ISSUE-052: ship .js.map files alongside the minified JS (debug-friendly, devtools-only download), or omit (cleaner repo, harder debugging)?',
     'Yes/no on shipping source maps.',
     '',
     'Source maps help developers debug production code. They do not load for normal users. Should we ship them?'),
]


def _copy_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.number_format = src.number_format


def main() -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.cell_range import CellRange
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1
    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)

    # --- Items sheet ---
    ws = wb['Items']
    rows = list(ws.iter_rows())
    hdr_idx = next(i for i, r in enumerate(rows[:6])
                   if r and any(str(c.value or '').strip().lower() == 'id' for c in r))
    hdr = [str(c.value or '').strip() for c in rows[hdr_idx]]
    plain_col = next((i for i, v in enumerate(hdr) if v.lower() == 'plain english'), None)
    perm_col = next((i for i, v in enumerate(hdr) if v.lower() == 'permission decision'), None)

    existing = {str(r[0].value).strip() for r in rows[hdr_idx+1:]
                if r and r[0].value}
    new_rows = [r for r in ISSUES + IMPS if r[0] not in existing]
    next_row = ws.max_row + 1
    ref = hdr_idx + 2  # canonical formatting row (first data row)
    for r in new_rows:
        # Schema layout:
        # cols 1..14 = original schema (ID..Decision)
        # col 15      = Plain English
        # col 16      = Permission decision
        # The tuple has 16 entries: 14 original + plain + perm.
        for col_offset, val in enumerate(r):
            ws.cell(row=next_row, column=col_offset + 1, value=val)
        for c in range(1, ws.max_column + 1):
            _copy_style(ws.cell(row=ref, column=c),
                        ws.cell(row=next_row, column=c))
        next_row += 1
    last = ws.max_row
    # Extend dropdowns
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for cr in list(dv.sqref.ranges):
            cr_str = str(cr)
            if cr_str.startswith('N'):
                dv.sqref.add(CellRange(f'N{hdr_idx+2}:N{last}'))
            if perm_col is not None and cr_str.startswith(chr(ord('A') + perm_col)):
                dv.sqref.add(CellRange(f'{chr(ord("A")+perm_col)}{hdr_idx+2}:{chr(ord("A")+perm_col)}{last}'))

    # --- Questions sheet ---
    wsq = wb['Questions']
    qrows = list(wsq.iter_rows())
    qhdr_idx = next(i for i, r in enumerate(qrows[:6])
                    if r and any(str(c.value or '').strip().lower() == 'id' for c in r))
    qhdr = [str(c.value or '').strip() for c in qrows[qhdr_idx]]
    qplain_col = next((i for i, v in enumerate(qhdr) if v.lower() == 'plain english'), None)
    existing_q = {str(r[0].value).strip() for r in qrows[qhdr_idx+1:]
                  if r and r[0].value}
    new_qs = [r for r in QUESTIONS if r[0] not in existing_q]
    next_q = wsq.max_row + 1
    refq = qhdr_idx + 2
    for r in new_qs:
        # Tuple: (id, topic, context, decision-needed-text, decision, plain-english)
        for col_offset, val in enumerate(r):
            wsq.cell(row=next_q, column=col_offset + 1, value=val)
        for c in range(1, wsq.max_column + 1):
            _copy_style(wsq.cell(row=refq, column=c),
                        wsq.cell(row=next_q, column=c))
        next_q += 1
    last_q = wsq.max_row
    for dv in wsq.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for cr in list(dv.sqref.ranges):
            if str(cr).startswith('E'):
                dv.sqref.add(CellRange(f'E{qhdr_idx+2}:E{last_q}'))

    wb.save(XLSX)
    print(f'Items:     appended {len(new_rows)} rows '
          f'({sum(1 for r in new_rows if r[0].startswith("ISSUE-"))} issues + '
          f'{sum(1 for r in new_rows if r[0].startswith("IMP-"))} improvements)')
    print(f'Questions: appended {len(new_qs)} rows')
    if existing & {r[0] for r in ISSUES + IMPS}:
        print(f'  skipped (Items): {", ".join(sorted(existing & {r[0] for r in ISSUES + IMPS}))}')
    if existing_q & {r[0] for r in QUESTIONS}:
        print(f'  skipped (Questions): {", ".join(sorted(existing_q & {r[0] for r in QUESTIONS}))}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
