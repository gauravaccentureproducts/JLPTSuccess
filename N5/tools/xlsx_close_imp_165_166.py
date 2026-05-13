"""Close IMP-165 + IMP-166 in the audit tracker.

After this commit lands, all 29 audit findings from the 2026-05-13
N5Improvement cycle have a Done / Defer / Avoid decision. No "Fix"
items remain.
"""
import openpyxl
from pathlib import Path
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = Path("feedback/n5-audit-2026-05-04.xlsx")

UPDATES = {
    "IMP-165": (
        "Done",
        " | SHIPPED 2026-05-13: minimal-pair cross-links added across "
        "vocab and kanji. PITCH PAIRS (3): 雨/あめ (rain/candy), 花/はな "
        "(flower/nose), 入れる/いれる (put-in/make-tea) — same kana, "
        "different pitch drops. LONG-VOWEL PAIRS (6): ビル/ビール, "
        "おばさん/おばあさん, おじさん/おじいさん, ここ/高校, え/ええ, "
        "いつ/五つ. KANJI PITCH NOTES (2): 雨 (rain ↔ 飴 candy), "
        "花 (flower ↔ 鼻 nose) — partners are above N5, but the pitch "
        "contrast is annotated for completeness. Sokuon pairs limited "
        "by the corpus's frozen width (来た/切った partner not a "
        "vocab entry; documented as structural gap). Total: 9 vocab "
        "cross-link pairs (18 bidirectional entries) + 2 kanji notes.",
    ),
    "IMP-166": (
        "Done",
        " | SHIPPED 2026-05-13: vocab authentic_refs coverage 37 -> 85 "
        "(2.3× improvement, +48 entries). Phase 1 reverse-mapped "
        "existing authentic.json vocab_refs (no new links — already "
        "consistent). Phase 2 matched 49 unlinked authentic.json items "
        "to vocab entries by form/reading. Final 85/1009 is below the "
        "original audit's '≥100' target — the gap is constrained by "
        "the authentic.json corpus size (188 items). Further expansion "
        "would require new authentic.json entries (out of "
        "RICHNESS-FIRST scope) or curated PD-quote linkages (more "
        "labor-intensive native review work). Provenance: auto_derived.",
    ),
}


def main():
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["Items"]
    updated = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if not row[0].value:
            continue
        item_id = str(row[0].value).strip()
        if item_id in UPDATES:
            new_decision, suffix = UPDATES[item_id]
            old = row[13].value
            row[13].value = new_decision
            row[14].value = (row[14].value or "") + suffix
            print(f"  {item_id}: {old!r} -> {new_decision!r}")
            updated += 1
    wb.save(str(XLSX))
    print(f"\nUpdated {updated} rows.")


if __name__ == "__main__":
    main()
