"""Build a derived "Unit Tests" sheet in test-scenarios xlsx.

The new sheet is a snapshot index of every scenario whose Test type
is `Auto` or `Auto+Manual` — i.e., every CI-runnable / scriptable
check. It's a single-page operational view of "everything that runs
without human review."

Columns:
  A  Index  (1..N)
  B  ID     (e.g., A-038)
  C  Tab    (single letter for compactness)
  D  Sub-category
  E  Scenario (summary)
  F  Tools / scripts
  G  CI invariant (JA-NN extracted from Notes when present)
  H  Owner / role
  I  Priority (P1..P5)
  J  Severity
  K  Test type (Auto or Auto+Manual)
  L  Estimated effort
  M  Coverage %
  N  Last run result
  O  Last run date
  P  Source-row reference (e.g., "A. Japanese language!row 5")

Sort order: Priority (P1 → P5) then Tab letter then ID.

This is a SNAPSHOT view — re-run this script after any update to
the source tabs to refresh the index.
"""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"

CATEGORY_TABS = [
    "A. Japanese language", "B. JLPT format", "C. Hindi locale", "D. UX design",
    "E. Accessibility", "F. Security", "G. Privacy and legal", "H. Performance",
    "I. Data engineering", "J. Pedagogy", "K. QA testing", "L. Cultural ethical",
    "M. Operations", "N. End-user POV",
]

# Sort priority numerically; non-P1..P5 values sort last
PRIORITY_ORDER = {"P1": 1, "P2": 2, "P3": 3, "P4": 4, "P5": 5}


def extract_ci_invariant(notes: str) -> str:
    """Pull the canonical JA-NN reference from the Notes cell, or '' if none."""
    if not notes:
        return ""
    # Match standalone JA-NN tokens — prefer the FIRST one mentioned
    m = re.search(r"\bJA-\d+\b", notes)
    if m:
        return m.group(0)
    return ""


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)

    # Collect all auto rows
    rows: list[dict] = []
    for tab in CATEGORY_TABS:
        sh = wb[tab]
        tab_letter = tab[0]
        for r in range(5, sh.max_row + 1):
            row_id = sh.cell(row=r, column=1).value
            if not row_id:
                continue
            test_type = sh.cell(row=r, column=9).value
            if test_type not in ("Auto", "Auto+Manual"):
                continue
            notes = sh.cell(row=r, column=10).value or ""
            rows.append({
                "id": row_id,
                "tab": tab,
                "tab_letter": tab_letter,
                "subcat": sh.cell(row=r, column=2).value or "",
                "scenario": sh.cell(row=r, column=4).value or "",
                "tools": sh.cell(row=r, column=13).value or "",
                "ci_invariant": extract_ci_invariant(notes),
                "owner": sh.cell(row=r, column=12).value or "",
                "priority": sh.cell(row=r, column=7).value or "",
                "severity": sh.cell(row=r, column=8).value or "",
                "test_type": test_type,
                "effort": sh.cell(row=r, column=11).value or "",
                "coverage": sh.cell(row=r, column=18).value or "0%",
                "last_run_result": sh.cell(row=r, column=15).value or "Not Yet Run",
                "last_run_date": sh.cell(row=r, column=14).value or "",
                "source_ref": f"{tab}!row {r}",
            })

    # Sort: priority asc (P1..P5), then tab letter, then ID
    rows.sort(key=lambda r: (
        PRIORITY_ORDER.get(r["priority"], 99),
        r["tab_letter"],
        r["id"],
    ))

    print(f"Collected {len(rows)} auto-runnable scenarios")

    # Create / replace the Unit Tests sheet
    SHEET_NAME = "Unit Tests (Auto-runnable)"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    sh = wb.create_sheet(SHEET_NAME)
    # Place the new sheet near the front, after Overview
    wb.move_sheet(SHEET_NAME, offset=-(len(wb.sheetnames) - 2))

    # Title rows
    sh["A1"] = "JLPT N5 Tutor — Unit Tests (Auto-runnable scenarios)"
    sh["A1"].font = Font(bold=True, size=14)
    sh.merge_cells("A1:P1")

    sh["A2"] = (
        f"Snapshot view of {len(rows)} auto-runnable test scenarios across the 14 "
        f"specialist-perspective tabs. Includes Test type = Auto + Auto+Manual. "
        f"Sorted by Priority (P1 → P5) then Tab + ID. Each row's CI invariant "
        f"column links to the JA-NN check in tools/check_content_integrity.py. "
        f"Re-run tools/build_unit_tests_sheet_2026_05_17.py after any source-tab edit."
    )
    sh["A2"].font = Font(italic=True, size=10)
    sh["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sh.merge_cells("A2:P2")
    sh.row_dimensions[2].height = 45

    sh["A3"] = ""  # spacer

    # Header row
    headers = [
        ("A", "#", 5),
        ("B", "ID", 9),
        ("C", "Tab", 5),
        ("D", "Sub-category", 22),
        ("E", "Scenario", 60),
        ("F", "Tools / scripts", 38),
        ("G", "CI invariant", 12),
        ("H", "Owner / role", 24),
        ("I", "Priority", 9),
        ("J", "Severity", 11),
        ("K", "Test type", 13),
        ("L", "Effort", 9),
        ("M", "Coverage %", 11),
        ("N", "Last run result", 14),
        ("O", "Last run date", 14),
        ("P", "Source ref", 32),
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="14452A", end_color="14452A", fill_type="solid")
    for col_letter, label, width in headers:
        cell = sh[f"{col_letter}4"]
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sh.column_dimensions[col_letter].width = width

    # Body rows
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    priority_fills = {
        "P1": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "P2": PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),
        "P3": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "P4": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "P5": PatternFill(start_color="D0E0E3", end_color="D0E0E3", fill_type="solid"),
    }
    for i, r in enumerate(rows, start=1):
        row = 4 + i
        sh.cell(row=row, column=1).value = i
        sh.cell(row=row, column=2).value = r["id"]
        sh.cell(row=row, column=3).value = r["tab_letter"]
        sh.cell(row=row, column=4).value = r["subcat"]
        sh.cell(row=row, column=5).value = r["scenario"]
        sh.cell(row=row, column=6).value = r["tools"]
        sh.cell(row=row, column=7).value = r["ci_invariant"]
        sh.cell(row=row, column=8).value = r["owner"]
        sh.cell(row=row, column=9).value = r["priority"]
        sh.cell(row=row, column=10).value = r["severity"]
        sh.cell(row=row, column=11).value = r["test_type"]
        sh.cell(row=row, column=12).value = r["effort"]
        sh.cell(row=row, column=13).value = r["coverage"]
        sh.cell(row=row, column=14).value = r["last_run_result"]
        sh.cell(row=row, column=15).value = r["last_run_date"]
        sh.cell(row=row, column=16).value = r["source_ref"]
        # Priority-tier fill on the row
        fill = priority_fills.get(r["priority"])
        for c in range(1, 17):
            cell = sh.cell(row=row, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # Freeze panes: header row + first 3 columns
    sh.freeze_panes = "D5"

    # Auto-filter
    sh.auto_filter.ref = f"A4:P{4 + len(rows)}"

    # Summary footer
    footer_row = 4 + len(rows) + 2
    sh.cell(row=footer_row, column=1).value = "Summary by priority:"
    sh.cell(row=footer_row, column=1).font = Font(bold=True)
    for i, p in enumerate(("P1", "P2", "P3", "P4", "P5")):
        count = sum(1 for r in rows if r["priority"] == p)
        sh.cell(row=footer_row + 1 + i, column=1).value = p
        sh.cell(row=footer_row + 1 + i, column=2).value = count
        sh.cell(row=footer_row + 1 + i, column=3).value = f"{100 * count / len(rows):.1f}%" if rows else "0%"
    # CI invariant coverage
    base = footer_row + 7
    sh.cell(row=base, column=1).value = "With CI invariant linkage:"
    sh.cell(row=base, column=1).font = Font(bold=True)
    n_with = sum(1 for r in rows if r["ci_invariant"])
    sh.cell(row=base, column=2).value = f"{n_with} / {len(rows)} ({100*n_with/len(rows):.1f}%)"

    wb.save(XLSX)
    print(f"Saved Unit Tests sheet with {len(rows)} rows.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
