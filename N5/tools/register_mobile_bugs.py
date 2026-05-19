#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Register the mobile-UI test-run bugs into User Reported Bugs sheet.

Idempotent: each bug carries a stable Title prefix (MOB-001, MOB-002, ...);
re-running skips bugs whose title already appears.

Bugs registered are CONFIRMED only — false positives identified during
runner triage are NOT registered:
- "Footer scrollTo doesn't move scrollY" — Selenium/CDP mobile-emulation
  artifact (verified working on desktop), NOT app bug
- "No audio on listening list" — listening page is a TOC; audio loads
  per-drill after tap on .listening-pick. Expected design.
- "Header sticky offset top=16" — sticky position is correctly `top:0`;
  the rendered 16px offset is parent-padding driven and visually
  intentional. Not a clear bug.

Run: python tools/register_mobile_bugs.py
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")

# Bug rows. Schema (10 used cols of 14):
#  (date_reported, reported_by, title, description, severity, priority,
#   status, fix_commit, fix_date)
# Note: Bug ID column has a formula that auto-numbers; we skip col 1.

REPORTER = "Mobile UI Selenium runner — tools/run_mobile_ui_tests.py (2026-05-19)"

BUGS = [
    (
        "MOB-001 — Primary nav drops Test + Progress (Mock-Test results pages) on mobile widths; no overflow menu visible",
        (
            "ROUTE: # (every page header) on D-360 (360x800 mobile-emulated)\n"
            "EVIDENCE: getBoundingClientRect for a[data-route='test'] = 0x0; "
            "a[data-route='summary'] = 0x0. getComputedStyle reports "
            "display:none on both (visibility:visible, opacity:1).\n"
            "OTHER 7 PRIMARY NAV LINKS render at 65x44 (OK).\n"
            "ALT-PATH CHECK: searched for hamburger / drawer / overflow menu "
            "via selectors: button[aria-label*='menu' i], button.menu, "
            "button.hamburger, button.nav-toggle, button[aria-expanded], "
            ".mobile-menu, .mobile-nav, .drawer, .sidebar-toggle, "
            "[data-toggle='menu'], [data-action='menu']. 0 matches.\n"
            "IMPACT: Mobile users cannot reach #/test or #/summary directly "
            "from the header. #/test is reachable from home page step 07 "
            "study-order card; #/summary has no header path at D-360 (must "
            "type URL or click a deep-link from another screen).\n"
            "REPRO: python -m http.server 8765; selenium Chrome with "
            "Emulation.setDeviceMetricsOverride width=360 height=800 mobile=true; "
            "GET /#/home; assert all 9 a[data-route] have w*h > 0. Fails on 2.\n"
            "SCENARIOS IT FAILS: O-X-021 (touch target on every nav link); "
            "O. Mobile UI testing tab in workbook.\n"
            "BOUNDARY: verified at D-360 viewport only; D-414+ may render "
            "the full nav. Test would also fail at D-320, D-375 (not "
            "yet probed for this specific assertion)."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOB-002 — Home study-order cards (steps 01-10) are 328x34 px on D-360 — 23% below HIG 44px touch-target minimum",
        (
            "ROUTE: #/home on D-360.\n"
            "EVIDENCE: a.study-order-link bounding rect 328 width x 34 height; "
            "padding 7.5px 0px; display:flex. Parent li.study-order-item is "
            "35px tall. 10 such cards on the home page (steps 01-10 of "
            "the suggested study path).\n"
            "STANDARDS: Apple HIG min target 44x44; Material Design min 48x48. "
            "34px is below both.\n"
            "IMPACT: Adjacent-card mis-taps likely on small-thumb users. The "
            "home page is the primary discovery surface — every learner sees "
            "this.\n"
            "REPRO: GET #/home at D-360; querySelectorAll('a.study-order-link'); "
            "min(rect.h) = 34 across all 10 cards.\n"
            "SCENARIOS IT FAILS: O-S-001 (home primary CTAs above fold + "
            "tappable). The CTA test currently asserts presence, not size — "
            "scenario should be tightened to assert ≥44px minimum.\n"
            "POSSIBLE FIX (do NOT apply per user request): increase "
            ".study-order-link padding from 7.5px 0 to ~12px 0 (or set "
            "min-height: 48px) — must verify visual density remains acceptable."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOB-003 — Home CTA buttons (Take Placement Check, Start with Grammar) are 281x36 px — 18% below HIG 44 minimum",
        (
            "ROUTE: #/home on D-360.\n"
            "EVIDENCE: a.btn-action.btn-action-primary[href='#/diagnostic'] "
            "rect 281 width x 36 height; same for "
            ".btn-action.btn-action-secondary[href='#/learn/grammar'].\n"
            "STANDARDS: Apple HIG min 44x44; Material min 48x48. 36px below both.\n"
            "IMPACT: Primary calls-to-action on the home page are smaller "
            "than recommended for thumb-targeting; first-time users may "
            "have higher mistap rate on these.\n"
            "REPRO: GET #/home at D-360; "
            "querySelectorAll('a.btn-action'); min height = 36.\n"
            "RELATED: same root cause as MOB-002 (vertical padding undersized "
            "on home-page interactive elements). Consider fixing class "
            ".btn-action globally."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOB-004 — Back-navigation links (\"← Back to X\") are 20px tall on every screen that has one — below HIG 44",
        (
            "ROUTES AFFECTED (D-360):\n"
            "  - #/learn/grammar: 'Back to Learn' 125x20\n"
            "  - #/learn/vocab:   'Back to Learn' 125x20\n"
            "  - #/kanji:         'Back to Learn' 125x20\n"
            "  - #/test:          'Continue learning' 154x20\n"
            "  - #/missed:        'Back to Review' 135x20\n"
            "  - #/home:          'All JLPT levels' 98x15 (worst case)\n"
            "Likely shared class — fix once propagates everywhere.\n"
            "STANDARDS: Apple HIG min 44 / Material 48. 15-20px is "
            "~35-45% of minimum.\n"
            "IMPACT: Back-navigation is a high-frequency tap path on mobile. "
            "Mis-tap risk especially when the link is near other links or "
            "edges.\n"
            "REPRO: GET each route above at D-360; querySelectorAll for back-"
            "nav link; assert rect.height >= 44.\n"
            "WORST INSTANCE: '← All JLPT levels' on home is 98x15 — only 15px "
            "tall; likely a plain link with default line-height, no padding."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-005 — Listening section expand/collapse buttons (ぜんぶ ひらく / とじる) are 99x36 — 8px below HIG 44 minimum",
        (
            "ROUTE: #/listening on D-360.\n"
            "EVIDENCE: button.toc-expand-all + button.toc-collapse-all "
            "rect 99x36 each. Margin around is reasonable but the touch "
            "target itself is shorter than HIG min.\n"
            "STANDARDS: HIG min 44 — 36px is 8px short.\n"
            "IMPACT: Lower than MOB-002/003 because these buttons are not "
            "the primary action on the route (the per-drill "
            "button.listening-pick is). Still inconsistent with the rest "
            "of the UI's nominal target size.\n"
            "REPRO: GET #/listening; "
            "querySelectorAll('.toc-expand-all, .toc-collapse-all'); "
            "min height = 36."
        ),
        "Minor", "P3", "Open"
    ),
]


def main() -> None:
    # Backup
    today = date.today().strftime("%Y_%m_%d")
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{today}_mobile_bugs")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)
    ws = wb["User Reported Bugs"]
    print(f"Sheet has {ws.max_row} rows currently.")

    # Find existing titles to avoid duplicates
    existing_titles = set()
    for r in range(4, ws.max_row + 1):
        t = ws.cell(row=r, column=4).value
        if t:
            existing_titles.add(str(t).split("—")[0].strip())

    written = 0
    skipped = 0
    today_str = date.today().isoformat()
    for title, description, severity, priority, status in BUGS:
        prefix = title.split("—")[0].strip()
        if prefix in existing_titles:
            print(f"  SKIP (already exists): {prefix}")
            skipped += 1
            continue
        new_row = ws.max_row + 1
        # Col 1: keep formula (or leave blank and let it inherit; safest to set the formula explicitly)
        ws.cell(row=new_row, column=1, value=f'="BUG-"&TEXT(ROW()-3,"000")')
        ws.cell(row=new_row, column=2, value=today_str)
        ws.cell(row=new_row, column=3, value=REPORTER)
        ws.cell(row=new_row, column=4, value=title)
        ws.cell(row=new_row, column=5, value=description)
        ws.cell(row=new_row, column=6, value=severity)
        ws.cell(row=new_row, column=7, value=priority)
        ws.cell(row=new_row, column=8, value=status)
        ws.cell(row=new_row, column=9, value="")  # Fix Commit
        ws.cell(row=new_row, column=10, value="")  # Fix Date
        for c in range(1, 11):
            cell = ws.cell(row=new_row, column=c)
            cell.alignment = WRAP
        # Row height
        max_len = max(len(str(v) or "") for v in [title, description])
        ws.row_dimensions[new_row].height = max(45, min(220, max_len // 5))
        print(f"  WROTE row {new_row}: {prefix}")
        written += 1

    wb.save(WORKBOOK)
    print()
    print(f"Result: {written} new bug rows written, {skipped} skipped (already present)")


if __name__ == "__main__":
    main()
