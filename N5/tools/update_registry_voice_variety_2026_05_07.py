"""Move ISSUE-062 + ISSUE-089 from Defer (env-blocked) to Fix (data
side done, awaiting one-command audio render on user machine).
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

UPDATES = {
    'ISSUE-062': ('Fix', 'commit 253896c', 'Data-side complete: 4-voice plan applied to all 47 listening items (Nanami / Keita / Aoi / Daichi). voice_variety_status pass on 47/47. Audio render pending one user command: pip install edge-tts pydub && python tools/build_listening_audio_multivoice_2026_05_07.py. Renders 47 multi-voice MP3s in ~5-10 min. Then -> Done.'),
    'ISSUE-089': ('Fix', 'commit 253896c', 'Carry-over of ISSUE-062. Same data-side closure. Same one-command audio render unblocks both.'),
}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    updated = 0
    for r in range(5, ws.max_row + 1):
        rid = ws.cell(row=r, column=1).value
        if rid in UPDATES:
            decision, commit_ref, summary = UPDATES[rid]
            ws.cell(row=r, column=14, value=decision)
            existing_desc = ws.cell(row=r, column=15).value or ''
            new_desc = existing_desc + f' Update {commit_ref}: {summary}'
            ws.cell(row=r, column=15, value=new_desc[:32760])
            print(f'  {rid}: -> {decision} ({commit_ref})')
            updated += 1
    print(f'Total updated: {updated}')
    wb.save(XLSX)


if __name__ == '__main__':
    main()
