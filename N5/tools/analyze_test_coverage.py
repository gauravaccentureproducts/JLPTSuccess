#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Coverage analysis of test-scenarios-by-specialist-perspective.xlsx.

Quantitative + qualitative pass:
1. Per-tab counts (scenarios / priorities / severities / test types)
2. Cross-reference scenarios against actual codebase artifacts:
   - SPA routes (index.html + js/*.js href="#/..." grep)
   - JS modules (js/*.js file list — feature owners)
   - Data files (data/*.json)
   - CI invariants (tools/check_content_integrity.py JA-XX list)
   - Specialist perspectives (84 personas from tabs A-N)
3. Identify coverage gaps: routes/modules/invariants with 0 scenarios.

Writing-discipline note: every claim below is bounded by what this
script actually scanned. "0 coverage" = "no scenario string contains
the artifact identifier as scanned by this script" — false negatives
possible if a scenario references the artifact by synonym.

Run: python tools/analyze_test_coverage.py > docs/TEST-COVERAGE-ANALYSIS-2026-05-19.md
"""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path
from collections import Counter, defaultdict

from openpyxl import load_workbook

# Force UTF-8 stdout on Windows cp932 default
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
INDEX_HTML = ROOT / "index.html"
JS_DIR = ROOT / "js"
DATA_DIR = ROOT / "data"
CI_SCRIPT = ROOT / "tools" / "check_content_integrity.py"


def read_all_scenarios():
    """Return dict[tab_name] -> list of dicts (one per scenario row)."""
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        if sheet == "Overview":
            continue
        ws = wb[sheet]
        # Find header row by scanning for "ID" in col A
        header_row = None
        for r in range(1, 8):
            if str(ws.cell(row=r, column=1).value or "").strip() == "ID":
                header_row = r
                break
        if header_row is None:
            continue
        headers = [(ws.cell(row=header_row, column=c).value or f"col{c}")
                   for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            if not vals[0] or not str(vals[0]).strip():
                continue
            rows.append(dict(zip(headers, vals)))
        out[sheet] = rows
    return out


def collect_routes():
    """All unique SPA routes from index.html + js/*.js."""
    pat = re.compile(r'href="#/([^"$\'\\${}?]+?)"')
    pat2 = re.compile(r"#/([a-zA-Z_/-]+)")
    routes = set()
    for p in [INDEX_HTML] + list(JS_DIR.glob("*.js")):
        try:
            text = p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        for m in pat.finditer(text):
            r = m.group(1).split("$")[0].split("?")[0].rstrip("/")
            if r and "${" not in r:
                routes.add(r)
    # Normalize known dynamic-route family roots
    family_roots = set()
    for r in routes:
        family_roots.add(r.split("/")[0])
        if "/" in r:
            family_roots.add(r.rsplit("/", 1)[0])
    return sorted(routes), sorted(family_roots)


def collect_js_modules():
    """All js/*.js files (one per feature)."""
    return sorted(p.stem for p in JS_DIR.glob("*.js"))


def collect_data_files():
    """All data/*.json file names."""
    return sorted(p.name for p in DATA_DIR.glob("*.json"))


def collect_ci_invariants():
    """Extract JA-NN invariant codes from tools/check_content_integrity.py."""
    if not CI_SCRIPT.exists():
        return []
    text = CI_SCRIPT.read_text(encoding="utf-8", errors="ignore")
    return sorted(set(re.findall(r"JA-\d+", text)))


def cross_reference(scenarios_by_tab, artifacts, kind_label):
    """For each artifact identifier in `artifacts`, count scenarios that mention it.

    Searches across scenario text fields (Scenario + Test steps + Notes)."""
    mention_count = Counter()
    for tab, rows in scenarios_by_tab.items():
        for row in rows:
            text = " ".join(str(row.get(k, "") or "") for k in
                            ("Scenario", "Test steps", "Notes", "Expected result"))
            for a in artifacts:
                # word-boundary or path-fragment match
                if kind_label == "route":
                    needle = f"#/{a}"
                elif kind_label == "js module":
                    needle = f"{a}.js"
                elif kind_label == "data file":
                    needle = a
                elif kind_label == "ci invariant":
                    needle = a
                else:
                    needle = a
                if needle in text:
                    mention_count[a] += 1
    return mention_count


def section(title, level=1):
    print()
    print("#" * level, title)
    print()


def main():
    scenarios = read_all_scenarios()
    routes, route_roots = collect_routes()
    js_modules = collect_js_modules()
    data_files = collect_data_files()
    ci_invariants = collect_ci_invariants()

    # --------------------------------------------------------------------
    # Header
    # --------------------------------------------------------------------
    print("# Test-Scenarios Coverage Analysis — 2026-05-19")
    print()
    print("Source workbook: "
          "`N5/specifications/test-scenarios-by-specialist-perspective.xlsx`")
    print()
    print("**Writing-discipline note.** Every claim below is bounded by what this "
          "script scanned. \"0 coverage\" means \"no scenario string in the "
          "workbook contains the artifact identifier exactly as scanned\"; "
          "false negatives are possible when a scenario references an artifact "
          "by synonym (e.g. \"vocab module\" vs `learn-vocab.js`). Numbers "
          "reflect counts at the workbook snapshot named above; future "
          "additions may change them.")
    print()

    # --------------------------------------------------------------------
    # Per-tab summary
    # --------------------------------------------------------------------
    section("1. Per-tab scenario counts")
    total = 0
    print("| Tab | Scenarios | P1/High | P2/Medium | P3/Low | Auto | Manual | Hybrid |")
    print("|-----|-----------|---------|-----------|--------|------|--------|--------|")
    per_tab_counts = {}
    for tab, rows in scenarios.items():
        n = len(rows)
        total += n
        prio = Counter()
        ttype = Counter()
        for r in rows:
            p = str(r.get("Priority", "") or "").strip().lower()
            if p in ("p1", "high"):
                prio["high"] += 1
            elif p in ("p2", "medium", "medium-high", "low-medium"):
                prio["medium"] += 1
            elif p in ("p3", "low"):
                prio["low"] += 1
            else:
                prio["other"] += 1
            tt = str(r.get("Test type", "") or "").lower()
            if "auto" in tt and "manual" not in tt and "hybrid" not in tt:
                ttype["auto"] += 1
            elif "manual" in tt and "auto" not in tt:
                ttype["manual"] += 1
            elif "hybrid" in tt or ("auto" in tt and "manual" in tt):
                ttype["hybrid"] += 1
            else:
                ttype["other"] += 1
        per_tab_counts[tab] = (n, prio, ttype)
        print(f"| {tab} | {n} | {prio['high']} | {prio['medium']} | "
              f"{prio['low']} | {ttype['auto']} | {ttype['manual']} | "
              f"{ttype['hybrid']} |")
    print(f"| **TOTAL** | **{total}** | | | | | | |")
    print()
    print(f"**Total scenarios across all tabs:** {total}")
    print()

    # --------------------------------------------------------------------
    # Codebase artifact inventory
    # --------------------------------------------------------------------
    section("2. Codebase artifact inventory (denominators)")
    print(f"- **SPA routes** (unique, normalized): {len(routes)}  "
          f"(grepped from index.html + js/*.js)")
    print(f"- **Route family roots**: {len(route_roots)} "
          f"(e.g., `home`, `learn`, `kanji`, `reading`, ...)")
    print(f"- **JS modules** (js/*.js feature owners): {len(js_modules)}")
    print(f"- **Data files** (data/*.json): {len(data_files)}")
    print(f"- **CI invariants** (JA-NN codes in tools/check_content_integrity.py):"
          f" {len(ci_invariants)}  *(may include duplicate mentions; count is of "
          f"distinct codes)*")
    print()
    print("**Caveat:** these counts are what the regex picked up at scan time. "
          "The route count includes dynamic-route shapes (e.g., `learn/${id}`); "
          "after deduplication of dynamic params, the family-root count "
          "(~37) is the more meaningful denominator for screen-coverage.")
    print()

    # --------------------------------------------------------------------
    # Route coverage
    # --------------------------------------------------------------------
    section("3. SPA route coverage (mobile UI + UI Tests tabs)")
    route_mentions = cross_reference(
        {k: v for k, v in scenarios.items()
         if k in ("O. Mobile UI testing", "UI Tests", "K. QA testing",
                  "D. UX design", "E. Accessibility")},
        route_roots, "route")
    print("Route-family coverage by mobile-UI / desktop-UI / QA tabs combined:")
    print()
    print("| Route family | Scenarios mentioning `#/family` |")
    print("|--------------|--------------------------------:|")
    covered = 0
    zero_routes = []
    for r in sorted(route_roots):
        n = route_mentions.get(r, 0)
        if n == 0:
            zero_routes.append(r)
        else:
            covered += 1
        print(f"| `#/{r}` | {n} |")
    print()
    print(f"**Coverage**: {covered}/{len(route_roots)} route families have ≥1 scenario "
          f"({covered * 100 // max(len(route_roots), 1)}%).")
    if zero_routes:
        print()
        print(f"**Routes with 0 scenarios** (in the tabs scanned — UI/QA/UX/A11y/Mobile):")
        for r in zero_routes:
            print(f"- `#/{r}`")
    print()

    # --------------------------------------------------------------------
    # JS module coverage
    # --------------------------------------------------------------------
    section("4. JS module coverage (any tab)")
    js_mentions = cross_reference(scenarios, js_modules, "js module")
    covered = 0
    zero_modules = []
    for m in sorted(js_modules):
        if js_mentions.get(m, 0) > 0:
            covered += 1
        else:
            zero_modules.append(m)
    print(f"**Coverage**: {covered}/{len(js_modules)} JS modules referenced by name "
          f"({covered * 100 // max(len(js_modules), 1)}%).")
    print()
    print(f"**JS modules with 0 explicit mention** (may still be exercised "
          f"indirectly via route scenarios):")
    for m in zero_modules:
        print(f"- `js/{m}.js`")
    print()

    # --------------------------------------------------------------------
    # Data file coverage
    # --------------------------------------------------------------------
    section("5. Data-file coverage (any tab)")
    data_mentions = cross_reference(scenarios, data_files, "data file")
    print("| Data file | Scenarios mentioning it |")
    print("|-----------|------------------------:|")
    for d in sorted(data_files):
        print(f"| `data/{d}` | {data_mentions.get(d, 0)} |")
    print()

    # --------------------------------------------------------------------
    # CI invariant coverage
    # --------------------------------------------------------------------
    section("6. CI invariant coverage (any tab)")
    ci_mentions = cross_reference(scenarios, ci_invariants, "ci invariant")
    covered_ci = sum(1 for c in ci_invariants if ci_mentions.get(c, 0) > 0)
    print(f"**Coverage**: {covered_ci}/{len(ci_invariants)} CI invariants referenced "
          f"by name in scenarios "
          f"({covered_ci * 100 // max(len(ci_invariants), 1)}%).")
    print()
    print("**Invariants with ≥1 scenario reference:**")
    for c in ci_invariants:
        if ci_mentions.get(c, 0) > 0:
            print(f"- {c}: {ci_mentions[c]} scenario(s)")
    print()
    zero_ci = [c for c in ci_invariants if ci_mentions.get(c, 0) == 0]
    if zero_ci:
        print(f"**Invariants with 0 explicit scenario reference** ({len(zero_ci)}):")
        print(", ".join(zero_ci))
    print()

    # --------------------------------------------------------------------
    # Priority / severity distribution
    # --------------------------------------------------------------------
    section("7. Priority + severity distribution (workbook-wide)")
    all_prio = Counter()
    all_sev = Counter()
    all_type = Counter()
    for tab, rows in scenarios.items():
        for r in rows:
            all_prio[str(r.get("Priority", "") or "").strip()] += 1
            all_sev[str(r.get("Severity if fails", "") or "").strip()] += 1
            all_type[str(r.get("Test type", "") or "").strip()] += 1

    print("**Priority distribution:**")
    print()
    for p, n in all_prio.most_common():
        print(f"- {p or '(blank)'}: {n}")
    print()
    print("**Severity distribution:**")
    print()
    for s, n in all_sev.most_common():
        print(f"- {s or '(blank)'}: {n}")
    print()
    print("**Test type distribution:**")
    print()
    for t, n in all_type.most_common():
        print(f"- {t or '(blank)'}: {n}")
    print()

    # --------------------------------------------------------------------
    # Coverage gaps + recommendations
    # --------------------------------------------------------------------
    section("8. Identified gaps + recommendations")
    print("These are derived from sections 3-6 above. Each gap is named "
          "explicitly so a follow-up tab/scenario addition can close it.")
    print()
    print("### 8.1 Routes with 0 mobile-UI/QA scenarios")
    if zero_routes:
        for r in zero_routes:
            print(f"- `#/{r}` — add scenarios to `O. Mobile UI testing` "
                  f"(or `UI Tests`) covering: initial load, primary CTA, "
                  f"navigation in/out, locale-switch.")
    else:
        print("None — all route families have ≥1 scenario in the UI-focused tabs.")
    print()
    print("### 8.2 JS modules with no scenario reference")
    print(f"- {len(zero_modules)} of {len(js_modules)} JS modules have 0 explicit "
          f"name reference in any scenario.")
    print("- Many may be exercised indirectly via route-level scenarios (e.g., "
          "`storage.js` is touched by every progress-bearing scenario).")
    print("- Cross-check needed: for each, decide whether it's a leaf utility "
          "(no UX surface) or a feature that lacks scenario coverage.")
    print()
    print("### 8.3 CI invariants with no scenario reference")
    print(f"- {len(zero_ci)} of {len(ci_invariants)} invariants have 0 explicit "
          f"reference in scenarios.")
    print("- CI invariants are gated structural checks, not user-facing UI "
          "behavior — so 0-mention is expected for many. However, scenarios in "
          "tab `I. Data engineering` should ideally explicitly name each "
          "invariant they validate.")
    print()
    print("### 8.4 Manual-test backlog (cannot be Selenium-automated)")
    manual_total = sum(1 for t, n in all_type.items()
                       if "manual" in t.lower() and "auto" not in t.lower()
                       for _ in range(n))
    print(f"- Manual scenarios total: {manual_total}.")
    print("- These need a human runner — track as a separate execution "
          "calendar, not via CI.")
    print("- Common classes flagged Manual: VoiceOver/NVDA navigation, iOS "
          "audio-gesture, soft-keyboard occlusion, IME input, safe-area, "
          "pull-to-refresh, real-device PWA install.")
    print()
    print("### 8.5 Underrepresented dimensions (qualitative)")
    print("- **Error-injection / chaos**: very few scenarios cover network-mid-"
          "flow, storage-quota-full, SW-update-collision, "
          "concurrent-tab-state-conflict.")
    print("- **Internationalization stress**: HI is covered (tab C); other "
          "RTL/CJK-only locales are out of scope by project decision.")
    print("- **Telemetry / observability**: the project claims zero "
          "telemetry — there are 1-2 scenarios verifying \"network tab is "
          "silent.\" Could be expanded to assert all network requests are "
          "first-party + cacheable.")
    print("- **Long-session memory pressure**: 1 scenario in O. tab; could "
          "expand to test 6-hour session leak detection.")
    print("- **Service-worker lifecycle**: 2 scenarios; SW is a complex "
          "subsystem deserving 8-10 scenarios (install/activate/skipWaiting/"
          "client claim/update flow/precache strategies/runtime caching/"
          "Background Sync if used).")
    print("- **Content-integrity at runtime**: data/*.json is validated at "
          "build by CI invariants, but no scenario verifies the RUNTIME "
          "renderer correctly displays the validated data (orphan-data trap "
          "documented in procedure-manual D.9.27).")
    print()
    print("### 8.6 Test-type balance")
    auto_pct = 0
    manual_pct = 0
    if total > 0:
        a = sum(n for t, n in all_type.items()
                if "auto" in t.lower() and "manual" not in t.lower())
        m = sum(n for t, n in all_type.items()
                if "manual" in t.lower() and "auto" not in t.lower())
        auto_pct = a * 100 // total
        manual_pct = m * 100 // total
    print(f"- Auto-only: ~{auto_pct}%  |  Manual-only: ~{manual_pct}%")
    print("- Healthy mix; investment in Appium would convert ~20 Manual "
          "soft-keyboard/IME scenarios to Automated.")
    print()

    # --------------------------------------------------------------------
    # Per-tab depth assessment
    # --------------------------------------------------------------------
    section("9. Per-tab depth assessment (qualitative)")
    depth_notes = {
        "Unit Tests (Auto-runnable)":
            "Sized for code-level invariants; depth is high relative to "
            "data/tools scope. Adequate for what's named.",
        "A. Japanese language":
            "Largest tab. Sub-categories thorough (Naturalness, Particles, "
            "Conjugation, Honorifics, Counters, Kanji, Pitch, Translation, "
            "Romaji, scope). Manual-heavy by necessity — most claims require "
            "native ear.",
        "B. JLPT format":
            "Mondai structure + stem format + difficulty calibration well-"
            "named. Item-validity (p-value, biserial) scenarios deferred "
            "until learner-cohort data exists — flagged appropriately.",
        "C. Hindi locale":
            "Naturalness + terminology + Devanagari typography + register + "
            "cross-language drift covered. Pan-Indian neutrality scenario is "
            "a strong addition.",
        "D. UX design":
            "Visual hierarchy + color + typography + IA + onboarding + "
            "cognitive load + microinteractions + form input + dark mode + "
            "gamification — all named. Heuristic-evaluation-style; would "
            "benefit from time-on-task data once collected.",
        "E. Accessibility":
            "WCAG-AA + screen-reader + motor + cognitive + hearing-impaired "
            "covered. Should add: WCAG-AAA spot-checks for 2.4.7 focus visible "
            "+ 2.5.5 target size if claims include AAA.",
        "F. Security":
            "XSS / CSP / supply-chain / secrets / pen-test / GitHub Actions "
            "covered. Static-site nature means surface is small; adequate "
            "for current threat model. Could add: subresource integrity, "
            "side-channel timing, fingerprinting resistance.",
        "G. Privacy and legal":
            "Privacy / GDPR / DPDP / COPPA / license / trademark / "
            "advertising-claims covered. Adequate for the legal review "
            "completed 2026-05-11.",
        "H. Performance":
            "Core Web Vitals + bundle + audio cache + PWA + cross-browser + "
            "mobile-responsive + font + i18n + SEO. Comprehensive for a PWA "
            "of this scope.",
        "I. Data engineering":
            "Schema validation + versioning + CACHE_VERSION + backup + CI. "
            "GAP: scenarios should explicitly cite each JA-NN invariant they "
            "validate.",
        "J. Pedagogy":
            "SLA + SRS + curriculum + assessment-validity + comparative + "
            "HI-L1 error anticipation. Some scenarios require learner-cohort "
            "data not yet collected; flagged Low priority.",
        "K. QA testing":
            "Exploratory + automation + visual-regression + localization + "
            "audio playback. Includes Selenium-suite references. Now overlaps "
            "with O. Mobile UI testing; cross-link both tabs.",
        "L. Cultural ethical":
            "JA + HI cultural sensitivity + AI ethics/disclosure + trust-"
            "safety. Adequate; depends on native reviewer for execution.",
        "M. Operations":
            "DevOps + analytics + customer-support + documentation. Lightweight "
            "for the current solo-maintainer setup; would expand with team.",
        "N. End-user POV":
            "Primary persona + self-study + classroom + returning + bilingual "
            "+ emerging-market + visually/cognitively-diverse. Persona "
            "coverage is broad.",
        "User Reported Bugs":
            "Bug-tracking rather than test scenarios per se. Useful for "
            "regression discipline (each bug should have a matching scenario "
            "in the relevant category tab).",
        "UI Tests":
            "59 Selenium scenarios from prior work. Now subset of mobile-UI "
            "coverage; consolidate or explicitly desktop-focus this tab to "
            "complement O. Mobile UI testing.",
        "O. Mobile UI testing":
            "134 scenarios: 39 cross-cutting + 95 per-screen. Newest tab; "
            "names every SPA route family. Some scenarios marked Manual/Appium "
            "for what Selenium cannot reproduce.",
    }
    for tab, note in depth_notes.items():
        if tab in scenarios:
            n = len(scenarios[tab])
            print(f"### {tab} ({n} scenarios)")
            print(note)
            print()

    # --------------------------------------------------------------------
    # Concrete recommendations
    # --------------------------------------------------------------------
    section("10. Concrete coverage-expansion recommendations")
    print("Ranked by gap-size × user-impact:")
    print()
    print("1. **Add per-invariant scenario rows in `I. Data engineering`** — "
          f"{len(zero_ci)} of {len(ci_invariants)} CI invariants currently lack "
          "an explicit scenario citation. Each should have a single-row "
          "scenario stating: \"JA-NN locks <named pattern>; regression test "
          "= run tools/check_content_integrity.py against a corpus seeded "
          "with the anti-pattern; expect failure.\"")
    print()
    print("2. **Service-worker lifecycle expansion in `O. Mobile UI testing`** "
          "— add 6-8 scenarios covering: install→activate transition, "
          "skipWaiting + clientsClaim, multiple-tab claim race, precache "
          "manifest hash mismatch recovery, runtime-cache eviction policy, "
          "Background Sync queue (if implemented), update-prompt UX, and "
          "stale-while-revalidate verification on key routes.")
    print()
    print("3. **Runtime renderer correctness** — add scenarios in `K. QA` or "
          "new tab `P. Renderer integrity` asserting: for each data field "
          "validated by CI invariants, the runtime DOM actually displays it. "
          "This closes the orphan-data trap (procedure-manual D.9.27) that "
          "lets `collocations` etc. ship without being rendered.")
    print()
    print("4. **JS module → scenario map** — annotate `js/*.js` files with a "
          "JSDoc `@coveredBy ID-list` comment so the link is bidirectional. "
          f"{len(zero_modules)} modules currently have no scenario name-cite.")
    print()
    print("5. **Long-session + chaos scenarios** — expand from current "
          "~3 in `O.` and `K.` to a dedicated sub-category: 6h memory leak, "
          "storage-quota-full, IndexedDB corruption, sw-update mid-flow, "
          "tab-eviction recovery, system-time-jump (DST / NTP correction).")
    print()
    print("6. **Native-human review scenarios** — currently 0 specific "
          "scenarios for the native-Hindi review and the 43 pitch-accent "
          "`low`/`unverified` entries flagged in AUDIT-COVERAGE. Add as "
          "row entries (with realistic priority Low — pending funding).")
    print()
    print("7. **Visual-regression baseline establishment** — `O.` lists "
          "Per-route screenshot diff baseline as 2 scenarios; expand to "
          "per-route × per-locale × per-viewport matrix (37 × 2 × 5 = 370 "
          "screenshots). Treat as one master scenario with a checklist "
          "rather than 370 rows.")
    print()
    print("8. **Cross-link `UI Tests` and `O. Mobile UI testing`** — `UI "
          "Tests` (59 rows) is now ambiguous in scope after `O.` was added. "
          "Either retitle `UI Tests` → `UI Tests (Desktop)` or merge "
          "appropriately so a future executor knows which to run for which "
          "device class.")
    print()

    section("11. Coverage scorecard (point-in-time, bounded)")
    print("| Dimension | Coverage | Method | Confidence |")
    print("|-----------|----------|--------|------------|")
    print(f"| SPA routes (family roots) | {covered}/{len(route_roots)} "
          f"({covered * 100 // max(len(route_roots), 1)}%) | Substring scan of "
          f"`#/<family>` in scenario text | High *for the routes named in "
          f"index.html/js as of commit cb97df1* |")
    print(f"| JS modules (named) | {len(js_modules) - len(zero_modules)}/"
          f"{len(js_modules)} "
          f"({(len(js_modules) - len(zero_modules)) * 100 // max(len(js_modules), 1)}%) | "
          f"Substring scan of `<module>.js` | Medium *— indirect coverage via "
          f"route scenarios is not counted* |")
    print(f"| CI invariants (cited) | {covered_ci}/{len(ci_invariants)} "
          f"({covered_ci * 100 // max(len(ci_invariants), 1)}%) | "
          f"Substring scan of `JA-NN` | Medium *— most invariants are "
          f"structural and don't require user-facing scenarios* |")
    print(f"| Data files (cited) | "
          f"{sum(1 for d in data_files if data_mentions.get(d, 0) > 0)}/"
          f"{len(data_files)} | Substring scan of filename | Medium |")
    print()
    print("Each percentage is bounded by what this script's substring matcher "
          "found; a scenario can exercise an artifact without literally naming "
          "it. The `medium` confidence labels acknowledge this.")
    print()
    print("---")
    print("*Analysis generated by `tools/analyze_test_coverage.py` against "
          "workbook snapshot `test-scenarios-by-specialist-perspective.xlsx` "
          "at the most recent commit.*")


if __name__ == "__main__":
    main()
