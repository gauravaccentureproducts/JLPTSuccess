"""Register the 7 items surfaced by the 2026-05-06 unit-test pass +
Hindi depth-fill cycle into feedback/n5-audit-2026-05-04.xlsx.

Items registered:

  Fixed this cycle (Decision = Fix; status Done in description):
    ISSUE-075  Stale _translation_status metadata in grammar.json
    IMP-097    Vocab gloss_hi Hindi seed (925 entries)
    IMP-098    Grammar meaning_hi Hindi seed (178 patterns)
    IMP-099    Grammar l1_notes Hindi-specific framework (21 patterns)

  Pending (Decision = Defer or Avoid):
    IMP-100    12 KB grammar patterns not in data/grammar.json (Avoid)
    ISSUE-076  29 design-system CSS violations (Defer)
    IMP-101    Native Hindi review pass for 1124 llm_curated entries (Defer)

Idempotent: skips IDs already in the sheet. Run once.
"""
from __future__ import annotations
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# (ID, Type, Severity, Priority, Impact, Effort, Category, Location, Title,
#  Current state, Why-it-matters, Suggested direction, Dependencies,
#  Decision, Description, Permission)
NEW_ROWS = [
    (
        'ISSUE-075', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
        'Documentation / metadata parity',
        'N5/data/grammar.json _translation_status block',
        'Stale _translation_status metadata referenced deprecated locales after Phase 4',
        'Phase 4 of the locale transition (commit 2dc8da6) pruned per-pattern locale fields (explanation_vi/id/ne/zh) but missed the top-level _translation_status metadata block, which still listed translatedPerLocale: {vi:0, id:0, ne:0, zh:0} and fields_per_pattern enumerating the deprecated locales.',
        'Internal-only doc-truth field but it is the schema-fields source for any future schema-extension work; leaving it stale would mislead.',
        'Update _translation_status to reflect en+hi reality: translatedPerLocale: {hi: 0}, fields_per_pattern lists explanation_en + explanation_hi, note references IMP-096 narrowing.',
        'IMP-096 (locale narrowing)',
        'Fix',
        'Surfaced by the 2026-05-06 unit-test pass as finding F4. Fixed in commit 6037e87 ("fix(data): clean stale vi/id/ne/zh metadata in grammar.json _translation_status"). Verification: walked grammar.json for residual vi/ne/zh keys = 0 hits. JA-39 invariant green. Status: Done.',
        'No permission required.'
    ),
    (
        'IMP-097', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'HIGH',
        'Hindi locale depth-fill — vocab',
        'N5/data/vocab.json gloss_hi field on 1041 entries',
        'Hindi gloss seed for 925 vocab entries (11.1% coverage to 100%)',
        'Post-locale-transition Phase 4 baseline: 116/1041 (11.1%) of vocab entries had gloss_hi populated. The remaining 925 entries had no Hindi gloss, leaving the niche-N1 (native-Hindi-medium JLPT prep) value proposition structurally incomplete.',
        'gloss_hi is the dominant niche-N1 content lever. Unlike vocab gloss in English (which is a structural-N5 invariant), gloss_hi is the user-facing Hindi-medium learning surface. Without it, the EN | HI chip switcher renders English fallback for vocab — undermining the multilingual claim.',
        'Author Devanagari Hindi glosses for all 925 missing entries via tools/seed_hindi_vocab_glosses_2026_05_06.py (LLM-curated; ≤30-char style matching the 116 prior entries; homograph disambiguation by gloss-substring hint).',
        'IMP-096 (locale narrowing must be done first)',
        'Fix',
        'Surfaced by the 2026-05-06 unit-test pass as finding F1. Fixed in commit e3892a1 ("data(hi): seed Hindi depth — vocab gloss_hi 100%, grammar meaning_hi 100%, l1_notes Hindi 21 patterns"). Final coverage: 1041/1041 (100%). All values in Devanagari script. All marked gloss_provenance.hi = llm_curated. Native-review pass tracked separately as IMP-101. Status: Done.',
        'No permission required for the LLM-curated seed; native-review pass requires budget decision.'
    ),
    (
        'IMP-098', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'MEDIUM',
        'Hindi locale depth-fill — grammar',
        'N5/data/grammar.json meaning_hi field on 178 patterns',
        'Hindi meaning_hi seed for all 178 grammar patterns (0% coverage to 100%)',
        'Post-locale-transition Phase 4 baseline: 0/178 patterns had meaning_hi populated. Hindi grammar-pattern meanings did not exist; the EN | HI chip switcher rendered English fallback on every grammar surface.',
        'Short pattern meanings are the user-facing summary on grammar list pages and pattern detail headers. Missing Hindi meaning_hi means a Hindi-medium learner sees English-language pattern names — direct inversion of the niche-N1 promise.',
        'Author short Devanagari Hindi meanings for all 178 patterns via tools/seed_hindi_grammar_2026_05_06.py (the SHORT field; OK to LLM-curate per project convention). Long explanation_hi NOT seeded — that field is reserved for native Hindi reviewer authoring per the _translation_status policy.',
        'IMP-096 (locale narrowing must be done first)',
        'Fix',
        'Surfaced by the 2026-05-06 unit-test pass as finding F2. Fixed in commit e3892a1 (same as IMP-097). Final coverage: 178/178 (100%). All meaning_hi values in Devanagari. _translation_status updated with meaning_hi_seeded: 178 counter. The long explanation_hi field remains 0/178 — that is intentional and tracked under IMP-101 (native review). Status: Done.',
        'No permission required for the LLM-curated seed.'
    ),
    (
        'IMP-099', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'MEDIUM',
        'Hindi locale depth-fill — L1-interference notes',
        'N5/data/grammar.json l1_notes field on patterns where Hindi-L1 contrasts apply',
        'Hindi-specific L1-interference notes seeded on 21 grammar patterns (0/178 to 21/178; sparse-by-design)',
        'Post-locale-transition Phase 4 baseline: 0/178 patterns had populated l1_notes (the field existed as empty {} on every pattern). The Hindi-specific 9-area L1-interference framework from N5/prompts/N5Improvement.txt and LocaleTransitionEnHi.txt was unimplemented.',
        'L1-interference notes are described in the audit prompt as "the strongest single content lever for niche-N1 market positioning" — they distinguish a generic English-medium app translated to Hindi from a true Hindi-pedagogically-aware app. Absent or generic notes is a major finding per the prompt.',
        'Author Hindi-L1 notes per the 9 mandatory contrast areas (postposition→particle mapping से→から/で, verb agreement, tense over-marking, politeness mismatch, negative placement, question particle position, plural marking, counter system overlap, SOV shared advantage) on patterns where the contrast actually applies. Sparse coverage is correct: not every pattern triggers a Hindi-L1 contrast.',
        'IMP-096 (locale narrowing must be done first)',
        'Fix',
        'Surfaced by the 2026-05-06 unit-test pass as finding F3. Fixed in commit e3892a1 (same as IMP-097). Final coverage: 21/178 patterns populated, covering n5-002 (は), n5-003 (が), n5-004 (を), n5-005 (に), n5-007 (で), n5-008 (と), n5-009 (から), n5-010 (まで), n5-058 (ます), n5-059 (ません), n5-065 (plain dict), n5-066 (ない), n5-067 (た), n5-072 (ています), n5-023 (か), n5-108 (counters), n5-079 (i-adj), n5-085 (na-adj), n5-104 (たい), n5-164 (さん), n5-169 (たことがある). The remaining 157 patterns either lack a strong Hindi-L1 contrast or have borderline cases not yet authored — those are polish items, not regressions. Status: Done (sparse coverage by design).',
        'No permission required for the LLM-curated seed.'
    ),
    (
        'IMP-100', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'HIGH',
        'Content scope / KB-JSON parity',
        'N5/KnowledgeBank/grammar_n5.md vs N5/data/grammar.json',
        '12 grammar patterns in KnowledgeBank not authored in data/grammar.json',
        'tools/check_coverage.py reports: KnowledgeBank/grammar_n5.md has 190 pattern entries; data/grammar.json has 178; 12 patterns are in the KB but not yet authored as runtime entries with examples / common_mistakes / contrasts.',
        'Width-additions to N5 grammar are explicitly anti-items in the depth-first audit cycle (per the DEPTH-FIRST ENRICHMENT MANDATE in N5/prompts/N5Improvement.txt). The 178 patterns already exceed Bunpro N5 deck width (~140); the 12 in KB are likely lower-frequency or marginal-N5 items that did not survive curation. Going from 178 to 190 dilutes per-pattern depth investment without meaningful coverage gain.',
        'Reserve until a future width-cycle audit (post-depth-fill). When that cycle starts: review each of the 12 KB patterns against JLPT.jp official scope + Bunpro/Genki/Minna; promote to grammar.json only those that are canonically N5; document the rest as KB-only reference material with a "not in runtime" flag.',
        'Width-cycle authorization (must come after depth-first cycle completes).',
        'Avoid',
        'Surfaced by the 2026-05-06 unit-test pass as finding F5. Pre-existing content gap (predates the locale transition). Marked Avoid for this cycle per the depth-first directive in N5/prompts/N5Improvement.txt anti-items list. Will be re-evaluated in a future width-cycle audit. Note: pattern n5-188 (the highest-numbered active id) suggests gaps were left intentionally during prior curation passes — confirm with the original curator before re-promoting.',
        'Requires explicit "override depth-first for these 12 patterns" permission OR a width-cycle audit pass.'
    ),
    (
        'ISSUE-076', 'Issue', 'MINOR', 'P3', 'LOW', 'MEDIUM',
        'Design-system / CSS tokens',
        'N5/css/main.css (29 violations across rules D-3, D-4, D-5, D-6, D-7)',
        '29 design-system rule violations in css/main.css',
        'tools/check_design_system.py reports 29 violations: D-3 (8 box-shadow uses including focus-visible outline-fade), D-4 (4 transform inside :hover for button-lift effects), D-5 (8 legacy brand-accent #14452a hardcoded fallbacks alongside var(--c-primary-dark, #14452a)), D-6 (4 non-token border-radius values 8px/12px), D-7 (1 text-transform: capitalize).',
        'Design-system rule consistency is a niche-N3 (institutional self-host) trust signal — adopters inspect CSS hygiene as a maintainability proxy. But each violation likely has functional intent: D-3 box-shadows are for focus-visible outline-fade animations (a11y-relevant); D-4 transforms are intentional button-lift micro-interactions per the round-3 UX brief; D-5 hardcoded fallbacks are CSS-var fallback-value belt-and-suspenders for older browsers; D-6 8px/12px radius are for specific component scales not in the 2/4/6/999 token set.',
        'Audit each violation with the original-intent context. Either (a) update the design-system spec to admit the functional exceptions (e.g., D-3 may permit focus-visible animations), or (b) refactor to use tokens (e.g., add 8px / 12px to D-6 token list), or (c) suppress per-violation with a documented reason. Do NOT bulk-remove without per-violation reasoning — risk of regression on a11y / micro-interaction polish.',
        'Design-spec update (which is a separate decision; the spec must move first OR per-violation review).',
        'Defer',
        'Surfaced by the 2026-05-06 unit-test pass as finding F6. Pre-existing CSS state (predates the locale transition). Each violation likely has a functional reason that the design-system rules do not yet account for. Bulk-fix without context = visual regression risk on focus-visible a11y, button hover-lift, and component micro-interactions. Recommend: design review pass before any code edits. Tracked separately from locale work.',
        'Requires design review and per-violation authorization.'
    ),
    (
        'IMP-101', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'HIGH',
        'Native Hindi reviewer pass',
        'N5/data/vocab.json (gloss_hi), N5/data/kanji.json (meanings_hi), N5/data/grammar.json (meaning_hi + l1_notes), N5/locales/hi.json',
        'Native Hindi reviewer pass for 1124+ LLM-curated Hindi entries',
        'Post-Hindi-seed-cycle baseline: 1041 vocab gloss_hi + 178 grammar meaning_hi + 21 grammar l1_notes + 106 kanji meanings_hi (already llm_curated in Phase 4) + ~50 hi.json UI strings = ~1396 Hindi entries, all marked review_status: llm_curated. The current credibility ceiling is bounded — a Hindi-medium learner who is also a native Hindi speaker will catch nuance errors (register, idiom, gendered concord, Sanskritized vs colloquial choice) that the LLM seed pass cannot.',
        'Niche-N1 (native-Hindi-medium JLPT prep) credibility requires native review. Without it, the app is "LLM-translated to Hindi" — a degraded version of "natively pedagogically-aware in Hindi." The IMP-097/098/099 seed pass made the app FUNCTIONAL in Hindi today; this cycle makes it CREDIBLE.',
        'Recruit one or two native Hindi-speaking Japanese-language teachers (Indian university Japanese programs are the obvious source — Yoisho Academy in particular has a Hindi-instruction lineage). Provide them the seeded JSON files plus a review-protocol document specifying: register correctness, gendered-noun agreement, false-friend Hindi colloquialisms, Devanagari-vs-Romanized conventions, and whether to revise low-frequency translations. After review, flip the entry-level review_status to native_reviewed for revised entries.',
        'Budget decision (paid reviewer vs volunteer); recruitment access; review-protocol authoring.',
        'Defer',
        'Open question Q-NN in Section 6 of N5/prompts/N5Improvement.txt — pending product-owner decision on native-Hindi-review budget. The depth-first audit prompt explicitly notes: "the credibility ceiling without native review is bounded." This row records the decision IS pending; no autonomous action possible.',
        'Requires user decision on (a) reviewer recruitment channel (paid vs volunteer vs Indian-university partnership) and (b) review-protocol scope (every entry vs sampled vs targeted-corrections-only).'
    ),
]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    existing_ids = set()
    for r in range(5, ws.max_row + 1):
        v = ws[r][0].value
        if v:
            existing_ids.add(v)

    appended = 0
    skipped = 0
    next_row = ws.max_row + 1
    for row in NEW_ROWS:
        if row[0] in existing_ids:
            print(f'  skip {row[0]} (already in sheet)')
            skipped += 1
            continue
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)
        appended += 1
        next_row += 1

    if appended:
        wb.save(XLSX)
        print(f'Appended {appended} rows. Skipped {skipped}.')
    else:
        print(f'Nothing to append. Skipped {skipped}.')


if __name__ == '__main__':
    main()
