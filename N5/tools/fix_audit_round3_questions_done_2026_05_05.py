"""Update Open questions sheet decisions for the questions resolved in
v1.12.31 close-out.

Q9 — vocab + kanji SRS scope: DECIDED. IMP-033 shipped the data plumbing
  without a special vocab+kanji daily cap; existing dailyReviewCap
  applies uniformly across all 3 history sources.

Q10 — mock-paper sitting timing: DECIDED. ISSUE-020/IMP-032 shipped
  real-time enforced timing per section (auto-submit at zero) with a
  60-second between-section break that the user can skip.

Q8 (i18n commit-vs-remove), Q11 (native audio budget), Q12 (round-2 Q2-Q7
unanswered) stay open — they need actual product/budget decisions, not
implementation calls.
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# QID -> (decision, rationale-suffix-appended-to-context)
RESOLVED = {
    'Q2':  ('Done', 'DECIDED 2026-05-04 (v1.12.28, IMP-001): untimed-with-toggle. Test setup ships an exam-mode checkbox that is OFF by default; checking it activates a ~60s/Q countdown that auto-submits at zero. Untimed-default keeps first-time learners unblocked; opt-in timer matches JLPT-real-exam pacing for those preparing for the sitting.'),
    'Q3':  ('Done', 'DECIDED 2026-05-05 (v1.12.31, IMP-006): single opt-in toggle ("Auto-furigana (experimental)") that applies ruby ONLY on a 19-kanji safe-single-reading whitelist (numerals, days, fixed compounds). Off by default. Risk-accepted version of the Pass-13-removed feature; the broader auto-ruby that produced wrong context-dependent readings stays disabled.'),
    'Q5':  ('Done', 'DECIDED 2026-05-04 (v1.12.28+IMP-011 marked Avoid): content-protect.js stays as-is. Round-1 audit accepted that the static-PWA threat model does not warrant DRM-style content protection; the existing CSP + same-origin lock is sufficient.'),
    'Q7':  ('Done', 'DECIDED 2026-05-05 (v1.12.30, IMP-024): default daily review goal = 20 reviews/day. User-tunable in Settings -> Practice -> "Daily review goal" (range 1-200). Round number, matches the round-3 audit recommendation, easily adjusted as learner feedback comes in.'),
    'Q9':  ('Done', 'DECIDED 2026-05-05 (v1.12.31): no special vocab+kanji daily cap. The existing dailyReviewCap setting applies uniformly across grammar + vocab + kanji history sources via storage.getDueCount(). Users who want a tighter cap on a particular surface can lower the global cap; a per-surface cap is YAGNI until learner feedback shows the unified queue is overwhelming.'),
    'Q10': ('Done', 'DECIDED 2026-05-05 (v1.12.31): real-time enforced per-section timer with auto-submit at zero, plus a 60-second between-section break with a "Skip break, start now" button. Strict-real-time was the right default for exam realism; the skippable break preserves casual-practice usability without splitting the implementation.'),
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    wb = load_workbook(XLSX)
    ws = wb['Open questions']

    # Find header row.
    header_row_idx = None
    for r in range(1, min(7, ws.max_row + 1)):
        cells = [str(c.value or '').strip() for c in ws[r]]
        if any(c.lower() == 'id' for c in cells) and any(
            c.lower().startswith('decision') for c in cells
        ):
            header_row_idx = r
            break
    if header_row_idx is None:
        print('ERROR: header row not found in Open questions')
        return 1

    header = [c.value for c in ws[header_row_idx]]
    id_col = next(i for i, v in enumerate(header, start=1)
                  if v and str(v).strip().lower() in ('id', 'issue id', 'item id'))
    ctx_col = None
    for i, v in enumerate(header, start=1):
        if v and 'context' in str(v).strip().lower():
            ctx_col = i; break
    dec_candidates = [i for i, v in enumerate(header, start=1)
                      if v and str(v).strip().lower().startswith('decision (')]
    if not dec_candidates:
        dec_candidates = [i for i, v in enumerate(header, start=1)
                          if v and str(v).strip().lower().startswith('decision')]
    dec_col = dec_candidates[0]

    closed = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        id_cell = row[id_col - 1]
        ctx_cell = row[ctx_col - 1] if ctx_col else None
        dec_cell = row[dec_col - 1]
        id_val = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
        if id_val not in RESOLVED:
            continue
        dec, rationale = RESOLVED[id_val]
        cur = (dec_cell.value or '').strip() if isinstance(dec_cell.value, str) else ''
        if cur == dec:
            continue
        dec_cell.value = dec
        # Append rationale to the existing context cell so the audit
        # trail captures why the question was resolved.
        if ctx_cell:
            existing = (ctx_cell.value or '').rstrip() if isinstance(ctx_cell.value, str) else ''
            ctx_cell.value = (existing + ('\n\n' if existing else '') + rationale).strip()
        closed.append((id_val, cur or '<blank>', dec))

    wb.save(XLSX)
    print(f'Resolved {len(closed)} open question(s):')
    for qid, prev, new in closed:
        print(f'  {qid}: {prev} -> {new}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
