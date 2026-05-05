"""Stamp Done on the round-4 items shipped in v1.12.32 close-out.

Items closed this batch (16 of 22):
  ISSUE-025 — LICENSE (MIT) at repo root
  ISSUE-026 — locales/{vi,id,ne,zh}.json expanded 33 -> 75+ keys (machine-translated; needs native review)
  ISSUE-027 — privacy trust band on home
  ISSUE-028 — header locale-chip switcher
  ISSUE-029 — Accept-Language toast
  ISSUE-030 — review_status provenance scaffold (1405/1405 items)
  ISSUE-031 — docs/SELF-HOST.md
  ISSUE-032 — og: tags + JSON-LD in index.html
  ISSUE-033 — n5_core_pattern_ids.json + JA-34 invariant
  ISSUE-034 — install link in trust band + Firefox/iOS fallback toast
  IMP-048   — trust band visual / ISSUE-027 paired
  IMP-049   — LICENSE + dual-license note + SELF-HOST guide / ISSUE-025+031 paired
  IMP-051   — og: + JSON-LD / ISSUE-032 paired
  IMP-052   — runtime theme-overrides loader (data/theme-overrides.json)
  IMP-055   — round3-features.spec.js Playwright spec
  IMP-056   — docs/TRANSLATING.md

Deferred to next cycle (6 of 22) — content-authoring or product-decision blocked:
  IMP-045 — translate grammar.json explanations to vi/id/ne/zh (HIGH content effort, blocked on Q14)
  IMP-046 — translate vocab.json glosses to vi/id/ne/zh (same, ~1041 entries × 4)
  IMP-047 — translate kanji.json meanings to vi/id/ne/zh (same)
  IMP-050 — kanji radical decomposition + mnemonic (KanjiDic2 import + curated mnemonics)
  IMP-053 — RTL CSS via logical properties (blocked on real RTL locale being authored)
  IMP-054 — Trusted Web Activity / Capacitor wrappers (blocked on Q17 distribution decision)
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

DONE_IDS = {
    'ISSUE-025', 'ISSUE-026', 'ISSUE-027', 'ISSUE-028', 'ISSUE-029',
    'ISSUE-030', 'ISSUE-031', 'ISSUE-032', 'ISSUE-033', 'ISSUE-034',
    'IMP-048', 'IMP-049', 'IMP-051', 'IMP-052', 'IMP-055', 'IMP-056',
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)
    closed = []
    skipped = []
    not_found = set(DONE_IDS)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_row_idx = None
        for r in range(1, min(7, ws.max_row + 1)):
            cells = [str(c.value or '').strip() for c in ws[r]]
            if any(c.lower() == 'id' for c in cells) and any(
                c.lower().startswith('decision') for c in cells
            ):
                header_row_idx = r
                break
        if header_row_idx is None:
            continue
        header = [c.value for c in ws[header_row_idx]]
        try:
            id_col = next(i for i, v in enumerate(header, start=1)
                          if v and str(v).strip().lower() in ('id', 'issue id', 'item id'))
            dec_candidates = [i for i, v in enumerate(header, start=1)
                              if v and str(v).strip().lower().startswith('decision (')]
            if not dec_candidates:
                dec_candidates = [i for i, v in enumerate(header, start=1)
                                  if v and str(v).strip().lower().startswith('decision')]
            dec_col = dec_candidates[0]
        except (StopIteration, IndexError):
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            id_cell = row[id_col - 1]
            dec_cell = row[dec_col - 1]
            id_val = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
            if id_val not in DONE_IDS:
                continue
            cur = (dec_cell.value or '').strip() if isinstance(dec_cell.value, str) else ''
            not_found.discard(id_val)
            if cur == 'Done':
                skipped.append((sheet, id_val))
                continue
            dec_cell.value = 'Done'
            closed.append((sheet, id_val, cur or '<blank>'))

    wb.save(XLSX)
    print(f'Closed {len(closed)} item(s):')
    for sh, iid, prev in closed:
        print(f'  [{sh}]  {iid}  ({prev} -> Done)')
    if skipped:
        print(f'\nAlready Done: {", ".join(s[1] for s in skipped)}')
    if not_found:
        print(f'\nWARNING: not found: {sorted(not_found)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
