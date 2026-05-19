#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Register ALL mobile-UI findings from probe (MOB-006 .. MOB-019).

This is the comprehensive registration pass per maintainer request
'register every issue', including findings previously dropped during
triage (Selenium-emulation artifacts, scenario-coverage gaps, possibly-
intentional design decisions). Every distinct observation that came
out of the probe is recorded; the maintainer decides triage downstream.

Run: python tools/register_mobile_bugs_extended.py
"""
from __future__ import annotations

import shutil
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")

REPORTER = "Mobile UI Selenium runner (extended probe) — tools/probe_more_mobile_issues.py (2026-05-19)"

BUGS = [
    (
        "MOB-006 — Feedback form inputs (4 controls) render at 14.06px font-size — triggers iOS Safari auto-zoom on focus",
        (
            "ROUTE: #/feedback on D-360 (iOS Safari emulation).\n"
            "EVIDENCE: getComputedStyle(input).fontSize values:\n"
            "  <SELECT id='fb-category'>  fontSize=14.0625px\n"
            "  <INPUT  id='fb-title'>     fontSize=14.0625px\n"
            "  <INPUT  id='fb-from'>      fontSize=14.0625px\n"
            "  <TEXTAREA id='fb-body'>    fontSize=14.0625px\n"
            "iOS Safari triggers viewport auto-zoom when a focused "
            "form input has font-size < 16px; this leaves the page "
            "zoomed-in until the user pinches back out. The header "
            "#search-input correctly uses fontSize=16px (no auto-zoom).\n"
            "IMPACT: Every user on iOS who opens the in-app feedback "
            "form will get a jarring zoom on first tap into any of "
            "these 4 fields. The Selenium scenario O-X-015 only "
            "checks #search-input; it does not cover form inputs on "
            "interior routes.\n"
            "REPRO: GET /#/feedback at D-360, "
            "document.querySelectorAll('input, textarea, select') for "
            "any with getComputedStyle(el).fontSize < 16; returns 4.\n"
            "POSSIBLE FIX (do NOT apply per user request): set "
            "font-size:16px on form controls site-wide, e.g. "
            "input,textarea,select{font-size:16px} OR mark the "
            "viewport meta with maximum-scale=1 (worse for a11y)."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOB-007 — '← All JLPT levels' home-up link is not localized to Hindi when locale=hi",
        (
            "ROUTE: #/home on D-360 with locale=hi.\n"
            "EVIDENCE: After toggling locale (lang attribute correctly "
            "flips from 'en' to 'hi' and all primary nav links translate "
            "to व्याकरण / शब्दावली / कान्जी / पठन / श्रवण / मॉक / "
            "छूटे प्रश्न), the home-up link inside .home-up-link still "
            "reads 'All JLPT levels' (English) verbatim. The arrow "
            "character ← is preserved but the label text is not "
            "translated.\n"
            "DOM: <a class='' href='#/levels'>← All JLPT levels</a>; "
            "no data-i18n-key attribute; not wired into the i18n "
            "lookup.\n"
            "IMPACT: Hindi-locale users see one untranslated English "
            "string on the home page. Trust-signal regression — the "
            "claim 'full Hindi locale' becomes weakly false for this "
            "string.\n"
            "REPRO: set localStorage['n5.locale']='hi'; reload #/home; "
            "verify document.querySelector('.home-up-link').textContent."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-008 — Route #/listening/story silently redirects to #/listening (dead-end navigation)",
        (
            "EVIDENCE: GET /#/listening/story; after 2.0s settle "
            "location.hash === '#/listening' (route silently rewrites "
            "the URL). Grep of js/*.js + index.html for "
            "'href=\"#/listening/story\"' shows the href IS referenced "
            "(e.g., template literal `#/listening/story/${esc(c)}` "
            "in js/listening-story.js). The bare #/listening/story "
            "without a chapter ID redirects to listening index; the "
            "homepage CTA 'Try story-mode listening' actually links "
            "to '#/listeningstory' (no slash) which works.\n"
            "INCONSISTENCY: the slash-form '#/listening/story' is a "
            "dead route while the no-slash form '#/listeningstory' is "
            "live. If both are intentional aliases, the slash-form "
            "should land on a story-list page, not silently redirect "
            "to the listening index.\n"
            "REPRO: chromedriver.get('http://127.0.0.1:8765/#/listening/story'); "
            "time.sleep(2); driver.execute_script('return location.hash') "
            "returns '#/listening' (not '#/listening/story').\n"
            "IMPACT: A bookmarked '#/listening/story' link does not "
            "land where the URL suggests. Possibly dead code in "
            "js/listening.js that should be removed if the no-slash "
            "form is the canonical entry."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-009 — Route #/levels silently redirects to #/diagnostic; no Levels page rendered",
        (
            "EVIDENCE: GET /#/levels; after 2.0s settle location.hash "
            "= '#/diagnostic' and #app content shows the Diagnostic "
            "page (10-question placement test). The home-up link "
            "'← All JLPT levels' targets href='#/levels' but lands "
            "the user on the Diagnostic.\n"
            "PROBABLE INTENT: '#/levels' is meant to be a multi-level "
            "(N5/N4/...) selection page or external 'JLPTSuccess root' "
            "page; redirect to #/diagnostic looks unintended.\n"
            "REPRO: get /#/levels; time.sleep(2); location.hash === "
            "'#/diagnostic'.\n"
            "IMPACT: Mid- to low-severity navigation surprise. The "
            "home-up link MOB-007 above lands users in placement-test "
            "mode unexpectedly — both bugs compound."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-010 — Sticky header renders at rect.top=16px on content routes despite CSS top:0 (parent-padding-driven offset)",
        (
            "ROUTE: #/learn/grammar (and likely every content route) on D-360.\n"
            "EVIDENCE: CSS computed style on the sticky header reports "
            "position:'sticky', top:'0px'. Yet after window.scrollTo "
            "(0, 5000), the header's getBoundingClientRect().top = 16, "
            "not 0. Diagnosis: sticky positioning is relative to the "
            "nearest scrolling ancestor and its top-padding. The "
            "header element sits inside a container with ~16px top-"
            "padding, so the 'top:0' rule pins the header to "
            "container_top + container_padding_top = +16px.\n"
            "IMPACT: 16px of extra space appears above the header "
            "during scroll on every content route. Likely intentional "
            "(visual breathing room) but the assertion 'sticky pinned "
            "to viewport top' is false. Either document the offset or "
            "set the parent's padding-top to 0 with the header pinned "
            "tight.\n"
            "REPRO: GET /#/learn/grammar; window.scrollTo(0, 5000); "
            "querySelector('.top-bar,header').getBoundingClientRect().top "
            "= 16.\n"
            "STATUS BAR: borderline — possibly by-design; flagging as "
            "P5 for design review."
        ),
        "Minor", "P5", "Open"
    ),
    (
        "MOB-011 — Authentic-items page has 449 undersized interactive elements (12x17 kanji-refs, 12-48x15 vocab-refs, 46x15 grammar-refs, 104x36 Pronounce buttons)",
        (
            "ROUTE: #/authentic on D-360.\n"
            "EVIDENCE: Per-route touch-target audit:\n"
            "  - <a> kanji-ref chips (parent='authentic-card-kanji-refs'): "
            "12x17 — kanji glyph link to detail page. ~150 instances.\n"
            "  - <a> vocab-ref chips (parent='authentic-card-vocab-refs'): "
            "12-108x15 — vocab form link. ~150 instances.\n"
            "  - <a> grammar-ref chips (parent='authentic-card-grammar-refs'): "
            "46x15 — text 'n5-XXX'. ~60 instances.\n"
            "  - <button class='btn-secondary btn-tiny'> 'Pronounce' "
            "buttons: 104x36 — audio play action. ~85 instances.\n"
            "STANDARDS: HIG min 44x44; Material min 48x48. 12x17 is "
            "≈25% of the HIG min target area.\n"
            "IMPACT: 449 elements below touch-target minimum on a "
            "single page. Density-rich page intentionally designed "
            "with tight chips, but the kanji-refs in particular "
            "(12x17 = ~5% of HIG area) are physically near-impossible "
            "to tap accurately on mobile. The Pronounce buttons are "
            "the worst class because they're the page's primary "
            "interaction.\n"
            "REPRO: GET /#/authentic at D-360; "
            "querySelectorAll('button, a[href]') for any rect with "
            "min(w,h) < 44; returns 449.\n"
            "DESIGN NOTE: this is likely a content-density tradeoff "
            "vs touch-friendliness. A consistent fix is to convert "
            "ref chips into long-press popovers OR to give each "
            "Pronounce button a 44+px tap target."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOB-012 — Header brand-link 'N5' (logo + wordmark) renders at 54x16 on every route",
        (
            "ROUTE: every route at D-360 (header is global).\n"
            "EVIDENCE: <a.brand-link href='../'> rect 54 width x 16 "
            "height. Functions as the home/level-up link.\n"
            "STANDARDS: HIG 44x44, Material 48x48. 16px is 36% of HIG "
            "min height.\n"
            "IMPACT: Logo doubles as a brand-up link to '../' (JLPT "
            "level index). Hard to tap; likely seen as decoration "
            "rather than interactive — discoverability + tappability "
            "both degraded.\n"
            "REPRO: any route, D-360; "
            "document.querySelector('a.brand-link').getBoundingClientRect() "
            "= {width:54, height:16}."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-013 — Skip-link 'Skip to main content' is 187x41 — 3px below HIG 44 touch-target minimum",
        (
            "ROUTE: every route at D-360 (skip-link is in global "
            "skiplink slot).\n"
            "EVIDENCE: <a.skip-link href='#app'> rect 187x41. Below "
            "HIG 44 by 3 px; below Material 48 by 7 px.\n"
            "STANDARDS: Skip-link is a critical a11y feature — must "
            "be easy to activate for keyboard + screen-reader users. "
            "While it's typically visually hidden until focused, the "
            "tap target when revealed should meet HIG.\n"
            "IMPACT: a11y degradation for keyboard / SR users who "
            "rely on skip-link to bypass header chrome.\n"
            "REPRO: GET /#/home; document.querySelector('a.skip-link')."
            "getBoundingClientRect()."
        ),
        "Minor", "P4", "Open"
    ),
    (
        "MOB-014 — Changelog 'docs/CHANGELOG-archive.md' link is 209x17 — below HIG 44",
        (
            "ROUTE: #/changelog on D-360.\n"
            "EVIDENCE: <a href='#'> with text 'docs/CHANGELOG-archive.md' "
            "renders 209x17. Likely a plain inline link with default "
            "line-height + no padding.\n"
            "IMPACT: Low. Changelog is a low-traffic route; archive "
            "link is a secondary action.\n"
            "REPRO: GET /#/changelog; look for the archive link in "
            "the page body."
        ),
        "Minor", "P4", "Open"
    ),
    (
        "MOB-015 — Examday + Weakareas 'See full ... bank →' links render at 139-167x15 (below HIG 44)",
        (
            "EVIDENCE:\n"
            "  - #/examday: 'See full strategy bank →' 139x15\n"
            "  - #/weakareas: 'See full test-strategy bank →' 167x15\n"
            "Both inside <span class='muted small'> parent — small-"
            "text styling makes the link tap area 15px tall.\n"
            "IMPACT: Cross-link between strategy-related pages is the "
            "intended navigation pattern. Tiny tap area degrades the "
            "discovery path for the Strategy Bank.\n"
            "REPRO: GET /#/examday; querySelectorAll('a') with the "
            "matching text; check bounding rect."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-016 — Feedback page action buttons (Open email to send, Cancel) are 159x36 and 88x36 — below HIG 44",
        (
            "ROUTE: #/feedback on D-360.\n"
            "EVIDENCE:\n"
            "  - <button.btn-action.btn-action-primary> 'Open email to send' 159x36\n"
            "  - <a.btn-action.btn-action-secondary href='#/home'> 'Cancel' 88x36\n"
            "Same .btn-action class root cause as MOB-003 (home CTAs). "
            "Fixing .btn-action min-height fixes both pages.\n"
            "IMPACT: Feedback form is a high-friction surface; primary "
            "submit-CTA being below HIG is a poor UX especially "
            "alongside the input-zoom MOB-006."
        ),
        "Major", "P2", "Open"
    ),
    (
        "MOB-017 — Reading list page (#/reading) has no detectable a[href*='reading/'] deep-link pattern — items use button-onclick navigation",
        (
            "ROUTE: #/reading on D-360.\n"
            "EVIDENCE: document.querySelector(\"a[href*='reading/']\") "
            "returns null on /#/reading. Reading list items appear "
            "to use <button> or non-href click handlers to enter the "
            "passage detail view. Compare to #/learn/grammar which "
            "has 178 <a href='#/learn/n5-XXX'> deep-link items.\n"
            "IMPACT (a): mobile-UI test scenarios cannot deep-link "
            "into a passage via simple selector — must enumerate "
            "buttons OR rely on shared 'listening-pick'-style "
            "<button data-id> pattern.\n"
            "IMPACT (b): SEO + crawlability degrade if individual "
            "passages have no anchor href in the list view (search "
            "engines and link-share tools won't surface them).\n"
            "IMPACT (c): bookmark / right-click 'Copy Link Address' "
            "from list view does not produce a passage URL.\n"
            "REPRO: GET /#/reading; assert "
            "querySelectorAll('a[href*=\"reading/\"]').length >= 1."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-018 — Mobile-emulation footer-reachability not verifiable: window.scrollTo no-ops with Emulation.setDeviceMetricsOverride mobile=true (test-infrastructure gap)",
        (
            "INFRASTRUCTURE FINDING (not an app bug per se).\n"
            "EVIDENCE: in Selenium 4.41 + Chrome 148.0.7778 headless "
            "with Emulation.setDeviceMetricsOverride mobile=true, "
            "window.scrollTo(0, 5000) returns scrollY=0 after the "
            "call. document.documentElement.scrollTop=5000 also "
            "no-ops. .scrollIntoView() does nothing. Tested on D-360. "
            "Without mobile-emulation (window-size 1200x800), the "
            "same scrollTo correctly yields scrollY=1023.\n"
            "IMPACT: scenarios O-S-f-* (footer reachability per route) "
            "cannot be validated by Selenium alone. Need (a) Appium / "
            "real device, OR (b) drop mobile-emulation and rely on "
            "window-size-only emulation (loses touch-emulation), OR "
            "(c) directly simulate touch-scroll via "
            "Input.dispatchTouchEvent CDP commands.\n"
            "RECOMMENDATION: split footer-reachability scenarios in "
            "O. Mobile UI testing into 'Auto (window-size only, no "
            "touch emulation)' + 'Manual (Appium / real device with "
            "touch)' variants.\n"
            "STATUS: blocking 5 scenarios in O. Mobile UI testing "
            "(O-S-f-home, O-S-f-learn_grammar, O-S-f-kanji, "
            "O-S-f-reading, O-S-f-settings). 2 of 5 already pass via "
            "manual-style assertion; 3 fail spuriously."
        ),
        "Minor", "P3", "Open"
    ),
    (
        "MOB-019 — Scenario coverage gap: 'Audio UI on listening/reading list pages' looks at TOC surface; audio lives on per-item routes",
        (
            "SCENARIO COVERAGE FINDING (test-design gap).\n"
            "AFFECTED SCENARIOS in workbook tab 'O. Mobile UI testing': "
            "O-S-a-listening, O-S-a-listening_story, O-S-a-reading.\n"
            "EVIDENCE: these scenarios assert >=1 <audio> element or "
            "play-button on #/listening, #/listening/story, #/reading "
            "respectively. All 3 return 0 because these pages are "
            "TOC / index views — audio is loaded after tap on a "
            "specific item (#/listening/n5.listen.NNN, "
            "#/reading/n5.read.NNN). Verified per-item routes load "
            "audio correctly: GET /#/listening/n5.listen.001 returns "
            "1 audio element + 1 play button.\n"
            "FIX (scenario, NOT app): re-target these 3 scenarios to "
            "navigate INTO a representative item before asserting "
            "audio UI. Example: GET /#/listening; click first "
            "button.listening-pick; THEN assert audio UI.\n"
            "STATUS: 3 scenarios in O. Mobile UI testing produce "
            "false-fail until rewritten."
        ),
        "Minor", "P4", "Open"
    ),
]


def main() -> None:
    today = date.today().strftime("%Y_%m_%d")
    bak = WORKBOOK.parent / (WORKBOOK.name + f".bak_{today}_mobile_bugs_extended")
    if not bak.exists():
        shutil.copy2(WORKBOOK, bak)
        print(f"Backup: {bak.name}")
    else:
        i = 2
        while True:
            alt = WORKBOOK.parent / (bak.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)
    ws = wb["User Reported Bugs"]

    # Find existing prefixes
    existing = set()
    for r in range(4, ws.max_row + 1):
        t = ws.cell(row=r, column=4).value
        if t:
            existing.add(str(t).split("—")[0].strip())

    written, skipped = 0, 0
    today_str = date.today().isoformat()
    for title, description, severity, priority, status in BUGS:
        prefix = title.split("—")[0].strip()
        if prefix in existing:
            print(f"  SKIP (already exists): {prefix}")
            skipped += 1
            continue
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=f'="BUG-"&TEXT(ROW()-3,"000")')
        ws.cell(row=new_row, column=2, value=today_str)
        ws.cell(row=new_row, column=3, value=REPORTER)
        ws.cell(row=new_row, column=4, value=title)
        ws.cell(row=new_row, column=5, value=description)
        ws.cell(row=new_row, column=6, value=severity)
        ws.cell(row=new_row, column=7, value=priority)
        ws.cell(row=new_row, column=8, value=status)
        ws.cell(row=new_row, column=9, value="")
        ws.cell(row=new_row, column=10, value="")
        for c in range(1, 11):
            ws.cell(row=new_row, column=c).alignment = WRAP
        max_len = max(len(str(v) or "") for v in [title, description])
        ws.row_dimensions[new_row].height = max(45, min(280, max_len // 5))
        print(f"  WROTE row {new_row}: {prefix}")
        written += 1

    wb.save(WORKBOOK)
    print()
    print(f"Result: {written} new bug rows written, {skipped} skipped")


if __name__ == "__main__":
    main()
