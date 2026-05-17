"""Mark BUG-047..053 status in the User Reported Bugs sheet.

Per the 2026-05-17 listening.json fix batch:
  BUG-047 → Fixed (voice_planned dropped; audio_render_meta canonical)
  BUG-048 → Fixed (audit-status fields refreshed for items 41-50)
  BUG-049 → Open (surface-only this batch — needs audio re-render)
  BUG-050 → Fixed (already-fixed in cdef185 Rule-5 install commit)
  BUG-051 → Fixed (format dropped; format_type canonical)
  BUG-052 → Fixed (_meta.voice_variety_plan rewritten as past-tense)
  BUG-053 → Fixed (voicevox_speaker_catalog character names corrected)

Updates Status column (col 8) and adds Fix Commit + Fix Date.
"""
from __future__ import annotations

import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"

FIXES = {
    "BUG-047": {
        "status": "Fixed",
        "note": (
            "Fixed 2026-05-17. voice_planned field dropped from all 50 "
            "items; audio_render_meta is now the canonical source for "
            "engine + speaker info. JS UI (listening.js voice-attribution) "
            "updated to read from audio_render_meta.voice_provider + "
            "audio_render_meta.voice_planned_for_engine.{F,M}.character. "
            "Locked by JA-110."
        ),
    },
    "BUG-048": {
        "status": "Fixed",
        "note": (
            "Fixed 2026-05-17. 7 items (041-047) had pacing_status="
            "'no_audio' refreshed → 'unmeasured' (audio IS rendered, "
            "just not pacing-measured). 3 items (048-050) had "
            "voice_variety_status=None refreshed → 'rendered'."
        ),
    },
    "BUG-049": {
        "status": "Open",
        "note": (
            "Surfaced 2026-05-17. Needs audio re-render at speed_scale "
            "~1.3 (currently 1.0). VOICEVOX install required (~30 min "
            "budget on maintainer's machine). Documented in "
            "listening.json _meta.pacing_fix_status; tracker stays Open."
        ),
    },
    "BUG-050": {
        "status": "Fixed",
        "note": (
            "Already-fixed 2026-05-17 in commit cdef185 (Cross-Artifact "
            "Sync Protocol install — Rule 5). version.json.counts."
            "listening was bumped to 50 alongside the vocab 1009→995 "
            "drift fix. JA-107 (INV-4 of the protocol) now locks the "
            "count parity."
        ),
    },
    "BUG-051": {
        "status": "Fixed",
        "note": (
            "Fixed 2026-05-17. format field dropped from all 50 items "
            "(was 1:1 redundant with format_type per the bug report). "
            "format_type is canonical with closed enum "
            "{task_understanding, point_understanding, "
            "utterance_expression, immediate_response}. JS consumers "
            "(listening.js byFormat grouping + search.js haystack) "
            "updated to read format_type. Locked by JA-111."
        ),
    },
    "BUG-052": {
        "status": "Fixed",
        "note": (
            "Fixed 2026-05-17. _meta.voice_variety_plan rewritten as "
            "past-tense completion record (status='completed_2026_05_12'). "
            "Bundled with BUG-053's catalog correction. The legacy "
            "voice_variety_plan_2026_05_07 block is marked superseded."
        ),
    },
    "BUG-053": {
        "status": "Fixed",
        "note": (
            "Fixed 2026-05-17 (bundled with BUG-052). "
            "voicevox_speaker_catalog rewritten with correct character→ID "
            "mappings: ID 8=春日部つむぎ (not 'hau-tsumugi' which is ID 10 "
            "雨晴はう); ID 11=玄野武宏 (not 'shirakami-kotaro'); ID 13="
            "青山龍星 (was wrongly filed under '12'). Observed-distribution "
            "block added. Unmet-target note records the 6-of-8 diversity "
            "gap (IDs 14 冥鳴ひまり + 53 ナースロボ＿タイプＴ were planned "
            "but never rendered)."
        ),
    },
}


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)
    if "User Reported Bugs" not in wb.sheetnames:
        print("ERROR: User Reported Bugs sheet missing")
        return 1
    b = wb["User Reported Bugs"]

    # Locate column indices by header row
    headers = {}
    for c in range(1, b.max_column + 1):
        v = b.cell(row=3, column=c).value
        if v:
            headers[v.strip()] = c
    status_col = headers.get("Status")
    if not status_col:
        print(f"ERROR: 'Status' column not found in header. Headers: {list(headers.keys())}")
        return 1
    fix_commit_col = headers.get("Fix Commit") or headers.get("Resolved By Commit")
    fix_date_col = headers.get("Fix Date") or headers.get("Resolution Date")
    notes_col = headers.get("Fix Notes") or headers.get("Notes")
    print(f"Status column: {status_col}; Fix Commit col: {fix_commit_col}; "
          f"Fix Date col: {fix_date_col}; Notes col: {notes_col}")

    today = datetime(2026, 5, 17)
    updated = 0
    for r in range(4, b.max_row + 1):
        bid_cell = b.cell(row=r, column=1).value
        if not bid_cell:
            continue
        # Formula bid → resolve via row position
        if isinstance(bid_cell, str) and bid_cell.startswith("="):
            bid_resolved = f"BUG-{r-3:03d}"
        else:
            bid_resolved = str(bid_cell)
        if bid_resolved not in FIXES:
            continue
        spec = FIXES[bid_resolved]
        b.cell(row=r, column=status_col).value = spec["status"]
        if spec["status"] == "Fixed":
            if fix_date_col:
                b.cell(row=r, column=fix_date_col).value = today
            if notes_col:
                cur = b.cell(row=r, column=notes_col).value or ""
                # Append rather than overwrite (preserve any prior context)
                if spec["note"] not in cur:
                    b.cell(row=r, column=notes_col).value = (cur + "\n\n" + spec["note"]).strip()
        else:
            # Open — just stamp the note
            if notes_col:
                cur = b.cell(row=r, column=notes_col).value or ""
                if spec["note"] not in cur:
                    b.cell(row=r, column=notes_col).value = (cur + "\n\n" + spec["note"]).strip()
        print(f"  row {r}: {bid_resolved} → {spec['status']}")
        updated += 1

    # Update Summary block
    # Look for "Summary" row label in column A
    for r in range(b.max_row, max(1, b.max_row - 30), -1):
        v = b.cell(row=r, column=1).value
        if not v:
            continue
        if isinstance(v, str) and "Total" in v and "bugs" in v.lower():
            # Found a Total bugs line; update count? Maybe.
            print(f"  found Summary row {r}: {v}")
            break

    wb.save(XLSX)
    print(f"\nUpdated {updated} bug rows.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
