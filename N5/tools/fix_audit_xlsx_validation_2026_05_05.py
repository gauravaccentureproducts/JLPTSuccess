"""Repair formatting on the round-3 rows appended to
feedback/n5-audit-2026-05-04.xlsx.

When openpyxl appends rows it inherits NEITHER data validations NOR
cell-level styling. After the round-3 register pass the new rows have:
  - no Fix/Avoid/Defer/Needs decision/Done dropdown on the Decision col
  - no row border (the existing rows have a thin top hairline)
  - no wrap-text alignment (existing rows wrap their long content)

This script fixes both:
  1. Extends the Decision-column data-validation sqref to cover every
     data row.
  2. Copies the canonical row's font / alignment / border to every new
     row, column-by-column.

The Decision cell's green-tinted fill is value-dependent (Done = green
FFC8E6C8) and is NOT auto-applied here — the user picks Fix/Avoid/Defer
from the dropdown first, then a future fix_audit_round*.py close-out
script can stamp the fill on Done rows the same way round-1+2 did.

Idempotent: re-running the script produces no further changes.
"""
from __future__ import annotations
import io, sys
from copy import copy
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# (sheet name, decision column letter, header row index, reference row to copy
#  styling from — picked because it has the canonical font/border/alignment)
TARGETS = [
    ('Audit findings', 'N', 4, 5),
    ('Open questions', 'E', 3, 4),
]


def _copy_style(src_cell, dst_cell, copy_fill: bool = False) -> None:
    """Mirror src_cell's font / alignment / border (and optionally fill) to
    dst_cell. Excel style objects are immutable references so we have to
    rebuild via copy.copy(); a bare assignment shares state and breaks
    when the workbook is saved."""
    dst_cell.font = copy(src_cell.font)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format
    if copy_fill:
        dst_cell.fill = copy(src_cell.fill)


def main() -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    wb = load_workbook(XLSX)
    changes = 0

    for sheet, col, header_row, ref_row in TARGETS:
        ws = wb[sheet]
        first_data_row = header_row + 1
        last_data_row = ws.max_row
        wanted = f'{col}{first_data_row}:{col}{last_data_row}'

        # ---- (a) cell-level styling (font/alignment/border) ----------
        # Walk every column on every row beyond ref_row. If the row's
        # first cell has no top border (proxy for "untouched by author"),
        # mirror ref_row's style across the whole row.
        styled_rows = 0
        for r in range(ref_row + 1, last_data_row + 1):
            probe = ws.cell(row=r, column=1)
            if probe.border and probe.border.top and probe.border.top.style:
                continue  # already styled
            for c in range(1, ws.max_column + 1):
                src = ws.cell(row=ref_row, column=c)
                dst = ws.cell(row=r, column=c)
                _copy_style(src, dst, copy_fill=False)
            styled_rows += 1
        if styled_rows:
            print(f'[{sheet}] styled {styled_rows} unformatted row(s) '
                  f'(font/alignment/border copied from row {ref_row})')
            changes += 1
        else:
            print(f'[{sheet}] all rows already styled; no style change')

        # ---- (b) Decision-column data validation (dropdown) ----------
        # Locate the existing list-validation that points at this column.
        target_dv = None
        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list':
                continue
            sqref_str = str(dv.sqref)
            if f'{col}{first_data_row}' in sqref_str or f'{col}5' in sqref_str or f'{col}4' in sqref_str:
                target_dv = dv
                break

        if target_dv is None:
            # Create a fresh one with the canonical formula.
            print(f'[{sheet}] no existing validation found, creating new')
            target_dv = DataValidation(
                type='list',
                formula1='"Fix,Avoid,Defer,Needs decision,Done"',
                allow_blank=True,
            )
            ws.add_data_validation(target_dv)
            target_dv.add(wanted)
            changes += 1
            continue

        old_sqref = str(target_dv.sqref)
        # Replace any existing range on the target column with the full new range.
        # Other columns in the sqref (Open questions also has D4:D10 for "Decision
        # needed" descriptive text) are preserved.
        from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
        new_ranges = MultiCellRange()
        replaced = False
        for cr in target_dv.sqref.ranges:
            cr_str = str(cr)
            if cr_str.startswith(col):
                # Replace this range — it's the one we want to grow.
                if not replaced:
                    new_ranges.add(CellRange(wanted))
                    replaced = True
            else:
                new_ranges.add(cr)
        if not replaced:
            new_ranges.add(CellRange(wanted))
        target_dv.sqref = new_ranges

        new_sqref = str(target_dv.sqref)
        if old_sqref != new_sqref:
            print(f'[{sheet}] {col} range: {old_sqref} -> {new_sqref}')
            changes += 1
        else:
            print(f'[{sheet}] {col} range already covers row {last_data_row}; no change')

    if changes:
        wb.save(XLSX)
        print(f'\nSaved. {changes} change(s) applied.')
    else:
        print('\nNo changes needed.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
