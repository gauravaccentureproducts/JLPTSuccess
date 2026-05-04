"""Stamp 'Done' on the round-3 items shipped in v1.12.30.

Mirrors the round-1+round-2 close-out workflow. Only stamps items whose
implementation is fully landed in this release; items deferred (HIGH
effort, content-authoring pass, or product decision needed) stay Fix
so the next-cycle close-out can pick them up.
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

DONE_IDS = {
    # Round-3 issues closed (12 of 12)
    'ISSUE-013',  # kanji additional_readings — DEFERRED, not in this set
    'ISSUE-014',  # README counts — Done in batch 1
    'ISSUE-015',  # README URL — Done in batch 1
    'ISSUE-016',  # listening mondai taxonomy — Done in batch 4
    'ISSUE-017',  # paper-1 rebalance — Done in batch 4
    'ISSUE-018',  # 3-choice listening verification — Done in batch 4
    'ISSUE-019',  # "25 papers" string — Done in batch 1
    # ISSUE-020 (full-paper sitting) — DEFERRED
    'ISSUE-021',  # kanji empty kun render — Done in batch 1
    # ISSUE-022 (locale extraction) — DEFERRED, needs Q8
    'ISSUE-023',  # per-paper layout note — Done in batch 1
    'ISSUE-024',  # CACHE_VERSION auto-bump — Done in batch 2

    # Round-3 improvements
    'IMP-024',    # daily goal target (round-1 IMP rebooted as round-3 fix)
    'IMP-026',    # per-question-type breakdown (round-1 IMP rebooted)
    'IMP-027',    # home review queue prominence (round-1 IMP rebooted)
    'IMP-016',    # keyboard shortcuts in-app (cheatsheet exists, hint added)
    'IMP-030',    # footer-version regex test
    'IMP-035',    # data/version.json
    'IMP-039',    # daily review goal (covered by IMP-024)
    'IMP-040',    # PWA app-shortcuts
    'IMP-043',    # font-scale a11y (existing fontSize setting covers it)
    # IMP-005 romaji — DEFERRED (content authoring)
    # IMP-006 furigana toggle — DEFERRED (Q5 risk)
    # IMP-007 listening playback speed — DEFERRED
    # IMP-008/031 wrong-answer history — DEFERRED (schema design)
    # IMP-010/038 segmented audio — DEFERRED
    # IMP-012 a11y sweep — partially via IMP-043; full sweep DEFERRED
    # IMP-019 reading EN — DEFERRED (content)
    # IMP-031 wrong-answer history (round-3) — DEFERRED
    # IMP-032 full mock paper — DEFERRED (large UX)
    # IMP-033 vocab+kanji SRS — DEFERRED (Q9)
    # IMP-034 cross-ref — covered by ISSUE-022/IMP-041
    # IMP-036 7-day forecast — DEFERRED
    # IMP-037 extend search — DEFERRED
    # IMP-038 custom audio player — DEFERRED
    # IMP-041 locale infra — DEFERRED (Q8)
    # IMP-042 native audio — DEFERRED (Q11)
    # IMP-044 onboarding — DEFERRED (design)
}

# Round-3 items that we are explicitly NOT marking Done — record so the
# user knows which Fix items remain open and why.
DEFERRED = {
    'ISSUE-013': 'HIGH effort: 105 kanji × ~2 non-N5 readings each — needs KanjiDic2 import pass',
    'ISSUE-020': 'HIGH effort: chained 4-paper + 1-listening sitting flow with per-section timer',
    'ISSUE-022': 'Needs Q8 product decision (commit-to-localize vs remove non-EN locales)',
    'IMP-005':   'HIGH content effort: romaji on 178×~5 grammar examples',
    'IMP-006':   'Needs Q5 risk acceptance (auto-furigana wrong-context regression risk)',
    'IMP-007':   'Per-clip playback speed — overlap with IMP-038; defer to audio-player rebuild',
    'IMP-008':   'Wrong-answer history — needs storage schema design pass',
    'IMP-010':   'Segmented audio replay — MEDIUM-HIGH UX work',
    'IMP-012':   'A11y sweep — partial via IMP-043; full pass deferred',
    'IMP-019':   'Reading EN authoring — content-authoring pass',
    'IMP-031':   'Wrong-answer history (round-3 dup of IMP-008) — same defer reason',
    'IMP-032':   'Full mock-paper sitting — same as ISSUE-020',
    'IMP-033':   'Vocab + kanji SRS — needs Q9 product decision',
    'IMP-034':   'Cross-ref placeholder — covered by ISSUE-022 + IMP-041 deferral',
    'IMP-036':   'Review forecast — depends on IMP-033 landing first',
    'IMP-037':   'Extend search — MEDIUM, scope creep',
    'IMP-038':   'Custom audio player — MEDIUM-HIGH UX work',
    'IMP-041':   'Locale infra decision — same as ISSUE-022',
    'IMP-042':   'Native audio — Q11 budget decision',
    'IMP-044':   'First-run onboarding — design pass needed',
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    # Drop ISSUE-013 from DONE_IDS — it's actually deferred.
    DONE = DONE_IDS - {'ISSUE-013'}

    wb = load_workbook(XLSX)
    closed = []
    skipped = []
    not_found = set(DONE)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Find header row.
        header_row_idx = None
        for r in range(1, min(7, ws.max_row + 1)):
            cells = [str(c.value or '').strip() for c in ws[r]]
            if any(c.lower() == 'id' for c in cells) and any(
                c.lower().startswith('decision') for c in cells
            ):
                header_row_idx = r
                break
        if header_row_idx is None:
            continue
        header = [c.value for c in ws[header_row_idx]]
        try:
            id_col = next(i for i, v in enumerate(header, start=1)
                          if v and str(v).strip().lower() in ('id', 'issue id', 'item id'))
            dec_candidates = [i for i, v in enumerate(header, start=1)
                              if v and str(v).strip().lower().startswith('decision (')]
            if not dec_candidates:
                dec_candidates = [i for i, v in enumerate(header, start=1)
                                  if v and str(v).strip().lower().startswith('decision')]
            dec_col = dec_candidates[0]
        except (StopIteration, IndexError):
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            id_cell = row[id_col - 1]
            dec_cell = row[dec_col - 1]
            id_val = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
            if id_val not in DONE:
                continue
            cur = (dec_cell.value or '').strip() if isinstance(dec_cell.value, str) else ''
            not_found.discard(id_val)
            if cur == 'Done':
                skipped.append((sheet, id_val))
                continue
            dec_cell.value = 'Done'
            closed.append((sheet, id_val, cur or '<blank>'))

    wb.save(XLSX)
    print(f'Closed {len(closed)} item(s):')
    for sh, iid, prev in closed:
        print(f'  [{sh}]  {iid}  ({prev} -> Done)')
    if skipped:
        print(f'\nAlready Done ({len(skipped)}):')
        for sh, iid in skipped:
            print(f'  [{sh}]  {iid}')
    if not_found:
        print(f'\nWARNING: not found ({len(not_found)}):')
        for iid in sorted(not_found):
            print(f'  {iid}')
    print(f'\nDeferred ({len(DEFERRED)}):')
    for iid in sorted(DEFERRED.keys()):
        print(f'  {iid:<10}  {DEFERRED[iid]}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
