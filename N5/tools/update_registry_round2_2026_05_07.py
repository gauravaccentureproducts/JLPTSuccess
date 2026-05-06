"""Mark final-batch items in registry after the 2026-05-07 round-2 session.

Done:
  ISSUE-093 — explanation_hi on 151 grammar patterns (commit f88a207)
  IMP-115   — Full Mock Test tile polished + 85Q/105min real-shape labelling
  IMP-116   — Provenance-badge UI on kanji surfaces (vocab side already shipped)
  IMP-121   — Score-report breakdown with per-section minimums
  (all three UI items in commit 31a064d)

Avoid:
  ISSUE-094 — native_reviewed scaling on vocab + kanji
              Genuinely requires a native Hindi-speaking Japanese teacher /
              native Japanese reviewer. LLM cannot legitimately upgrade
              review_status from llm_curated to native_reviewed without
              actual native review. Same blocker class as IMP-101.
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

UPDATES = {
    'ISSUE-093': ('Done', 'commit f88a207', 'explanation_hi authored on all 151 remaining grammar patterns. 178/178 (100%) coverage. Per-pattern provenance: explanation_provenance.hi = llm_curated. Q33 LLM-persona scaling honored; native review remains the target end-state.'),
    'IMP-115':   ('Done', 'commit 31a064d', 'Full Mock Test tile polished on #/test page. Reads full_mock_papers from manifest; surfaces 85Q/105min real-JLPT shape with section-by-section breakdown (30Q/25min + 31Q/50min + 24Q/30min). Links to existing #/sitting flow.'),
    'IMP-116':   ('Done', 'commit 31a064d', 'Provenance-badge UI render call site added on kanji.js (vocab side already shipped). Both surfaces gated by feature flag + 10% native-reviewed threshold; will auto-appear when reviewer pass crosses threshold.'),
    'IMP-121':   ('Done', 'commit 31a064d', 'Score-report breakdown matches real JLPT N5 shape: per-section minimum-pass column (~63%/61%/79% raw approximation), 60% study target vs 80/180 official scaled mark, per-section pass/fail status, footnote disclosing approximation vs JLPT scaled equating.'),
    'ISSUE-094': ('Avoid', 'no commit', 'Genuinely requires native Hindi-speaking Japanese teacher / native Japanese reviewer to upgrade review_status from llm_curated to native_reviewed. Same blocker class as IMP-101 (no monetization plan; no budget). Decision: Avoid until institutional sponsorship or paid tier emerges.'),
}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    updated = 0
    for r in range(5, ws.max_row + 1):
        rid = ws.cell(row=r, column=1).value
        if rid in UPDATES:
            decision, commit_ref, summary = UPDATES[rid]
            ws.cell(row=r, column=14, value=decision)
            existing_desc = ws.cell(row=r, column=15).value or ''
            new_desc = existing_desc + f' Status: {decision} ({commit_ref}). {summary}'
            ws.cell(row=r, column=15, value=new_desc[:32760])
            print(f'  {rid}: marked {decision} ({commit_ref})')
            updated += 1
    print(f'Total updated: {updated}')
    wb.save(XLSX)


if __name__ == '__main__':
    main()
