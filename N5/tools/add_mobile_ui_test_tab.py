"""Add 'O. Mobile UI testing' tab to the test-scenarios workbook.

Idempotent: re-running replaces the tab in-place; other tabs untouched.

Scenarios cover every SPA route in the app (37 screens) + cross-cutting
mobile concerns (viewports, orientation, safe-area, soft-keyboard,
audio, network, locale). Each scenario is Selenium-ready: includes
CSS selector hints, viewport profile, action sequence, and a
deterministic expected state.

Writing-discipline note: every scenario phrasing is bounded. "Every
screen" = "every route currently listed in index.html / js/app.js
route table as of commit cb97df1"; new routes added later are not
in scope until the tab is regenerated.

Run: python tools/add_mobile_ui_test_tab.py
"""
from __future__ import annotations

from pathlib import Path
import shutil
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WORKBOOK = Path(__file__).resolve().parent.parent / "specifications" / \
    "test-scenarios-by-specialist-perspective.xlsx"

TAB_NAME = "O. Mobile UI testing"

HEADER_FILL = PatternFill("solid", fgColor="14452A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13, color="14452A")
WRAP = Alignment(wrap_text=True, vertical="top")

# 18 columns matching A-N category tabs schema
COLUMNS = [
    ("ID", 9),
    ("Sub-category", 22),
    ("Persp #", 8),
    ("Scenario", 50),
    ("Test steps", 58),
    ("Expected result", 44),
    ("Priority", 9),
    ("Severity if fails", 12),
    ("Test type", 18),
    ("Notes", 32),
    ("Estimated effort", 14),
    ("Owner / role", 22),
    ("Tools / scripts required", 30),
    ("Last run date", 13),
    ("Last run result", 13),
    ("Tracker link", 16),
    ("Depends on", 14),
    ("Coverage %", 10),
]

PERSPECTIVES = "Mobile UI engineer / Mobile UX / Selenium-Appium E2E tester"

# ---------------------------------------------------------------------------
# Standard device profiles referenced in steps. Selenium / Playwright sets
# these via window-size or mobile-emulation.
# ---------------------------------------------------------------------------
DEVICE_PROFILES = """\
Device profiles referenced below (set via driver.set_window_size() OR
Chrome mobile-emulation deviceMetrics):
- D-320:  320x568   (Galaxy Fold folded / iPhone 5/SE 1st-gen)
- D-360:  360x800   (Pixel 4a / common Android baseline)
- D-375:  375x667   (iPhone SE 2nd-gen)
- D-390:  390x844   (iPhone 12/13/14)
- D-414:  414x896   (iPhone Plus / Max)
- D-768:  768x1024  (iPad portrait)
- D-360L: 800x360   (Pixel 4a landscape)
- D-414N: 414x896 + safe-area 47px top, 34px bottom (notched)
All device profiles use touch-emulation + DPR=2 unless noted.
"""

# ---------------------------------------------------------------------------
# Scenarios. Tuple shape:
#   (sub_category, scenario, steps, expected, priority, severity,
#    test_type, notes, effort, owner, tools)
# ---------------------------------------------------------------------------

CROSS_CUTTING = [
    ("Viewport sizing", "App renders without horizontal scroll at 320px width",
     "1. Launch app at D-320\n2. Wait for #app skeleton to disappear\n3. JS: document.documentElement.scrollWidth <= window.innerWidth",
     "scrollWidth === innerWidth (no horizontal scroll)",
     "P1", "Major", "Automated (Selenium)", "Galaxy Fold folded is real-world baseline",
     "15 min", "Mobile UI engineer", "selenium 4.x, chromedriver, ui_mobile_suite.py"),
    ("Viewport sizing", "App renders without horizontal scroll at 360px width",
     "1. D-360\n2. Same scrollWidth assertion per route",
     "scrollWidth === innerWidth on every route",
     "P1", "Major", "Automated (Selenium)", "Loop across 37 routes",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Viewport sizing", "App renders without horizontal scroll at 375px (iPhone SE)",
     "1. D-375\n2. scrollWidth assertion across all routes",
     "scrollWidth === innerWidth on every route",
     "P1", "Major", "Automated (Selenium)", "",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Viewport sizing", "App renders without horizontal scroll at 414px",
     "1. D-414\n2. scrollWidth assertion across all routes",
     "scrollWidth === innerWidth",
     "P2", "Minor", "Automated (Selenium)", "",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Viewport sizing", "App renders without horizontal scroll at 768px (tablet)",
     "1. D-768\n2. scrollWidth assertion across all routes",
     "scrollWidth === innerWidth",
     "P2", "Minor", "Automated (Selenium)", "Tablet is a target form factor per PWA",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Orientation", "Portrait-to-landscape transition preserves route + scroll position",
     "1. D-390 portrait, load #/learn/grammar\n2. Scroll halfway\n3. Rotate to D-360L\n4. Verify hash == #/learn/grammar; scrollY > 0",
     "Route + approximate scroll preserved",
     "P1", "Major", "Automated (Selenium)", "Test on Chrome mobile-emulation rotate",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Orientation", "Landscape rendering does not break sticky header",
     "1. D-360L\n2. Scroll content\n3. Verify .top-bar still position:sticky and visible",
     "Header sticks in landscape",
     "P2", "Minor", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Safe area", "Notch + home-indicator safe-area respected (iPhone X+)",
     "1. D-414N\n2. Verify #app top padding ≥ 47px (env(safe-area-inset-top))\n3. Verify footer bottom-margin ≥ 34px",
     "Safe-area padding applied",
     "P2", "Minor", "Manual", "Selenium cannot set safe-area directly; use real iOS device or BrowserStack",
     "30 min", "Mobile UI engineer", "BrowserStack iOS"),
    ("Safe area", "Pixel notch / hole-punch safe area on Android",
     "1. Real Pixel 6 device\n2. Verify no UI elements clipped under hole-punch",
     "No clipping",
     "P3", "Minor", "Manual", "",
     "30 min", "Mobile UI engineer", "Real device"),
    ("Header / nav", "Primary nav fits in 360px without overflow",
     "1. D-360, load #/home\n2. nav.primary-nav element\n3. Verify scrollWidth <= clientWidth OR has horizontal-scroll affordance",
     "Nav fits OR horizontal-scrolls cleanly",
     "P1", "Major", "Automated (Selenium)", "9 nav links (Grammar / Vocab / Kanji / Reading / Listening / Test / Mock / Missed / Progress)",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Header / nav", "Secondary nav (search + locale + fullscreen + settings) fits at 320px",
     "1. D-320\n2. nav.secondary-nav\n3. Verify all 4 children visible (search-input + locale-toggle + fullscreen-toggle + #/settings)",
     "All 4 secondary controls visible",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Header / nav", "Sticky header stays at top during long-page scroll",
     "1. D-360, load #/learn/grammar (long list)\n2. scrollTo(0, 5000)\n3. Verify .top-bar getBoundingClientRect().top === 0",
     "Header sticky",
     "P1", "Major", "Automated (Selenium)", "Sticky check requires CSS support",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Header / nav", "iOS Safari address-bar collapse does not cause layout jump",
     "1. Real iPhone Safari\n2. Scroll down 50px (triggers address-bar collapse)\n3. Capture screenshot before/after\n4. Diff for layout shift",
     "CLS-equivalent < 0.1",
     "P2", "Minor", "Manual", "Selenium cannot reproduce Safari URL bar collapse",
     "30 min", "Mobile UI engineer", "Real iOS device"),
    ("Footer", "Footer is reachable on every route without keyboard scroll-traps",
     "1. D-360, load each route\n2. JS: scrollTo(0, document.body.scrollHeight)\n3. Verify footer.app-footer in viewport",
     "Footer reachable on every route",
     "P2", "Minor", "Automated (Selenium)", "",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Search input", "Search input does not auto-zoom on iOS (font-size ≥ 16px)",
     "1. D-375 (iOS profile)\n2. Inspect #search-input computed font-size\n3. Verify >= 16px",
     "font-size >= 16px (prevents iOS auto-zoom)",
     "P1", "Major", "Automated (Selenium)", "Common iOS gotcha; <16px triggers zoom-on-focus",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Search input", "Soft keyboard appearance does not occlude search results",
     "1. D-390\n2. tap #search-input, type 'tabe'\n3. Verify visualViewport.height shrinks (keyboard up)\n4. Verify autocomplete results stay above keyboard",
     "Results visible above soft keyboard",
     "P1", "Major", "Manual", "Selenium can't trigger soft keyboard; use Appium or real device",
     "20 min", "Mobile UI engineer", "Appium"),
    ("Locale toggle", "Locale toggle round-trips EN ↔ HI across all routes",
     "1. D-360, load #/home\n2. Tap #locale-toggle\n3. Verify <html lang> changes; [data-locale-label] flips\n4. Reload\n5. Verify locale persists in localStorage",
     "Locale persists across reload + routes",
     "P1", "Major", "Automated (Selenium)", "",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Locale toggle", "Locale toggle is reachable via single-tap (target ≥ 48x48px)",
     "1. D-360\n2. Inspect #locale-toggle bounding rect\n3. Verify width >= 44 AND height >= 44 (Apple HIG)",
     "Tap target meets HIG / Material",
     "P1", "Major", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Fullscreen toggle", "Fullscreen toggle responds to tap on mobile",
     "1. D-390, Chrome\n2. tap #fullscreen-toggle\n3. Verify document.fullscreenElement OR :-webkit-full-screen",
     "Fullscreen enters; toggle exits",
     "P2", "Minor", "Automated (Selenium)", "Fullscreen API not supported on iOS Safari; expected-fail there",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Settings icon", "Settings cog navigates to #/settings",
     "1. D-360\n2. tap .icon-btn[href='#/settings']\n3. Verify hash === '#/settings' AND main shows settings content",
     "Navigation completes",
     "P1", "Major", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Touch targets", "Every nav link is ≥ 44x44 CSS pixels",
     "1. D-360, load #/home\n2. For every a[data-route], measure rect\n3. Assert min(width,height) >= 44",
     "All nav targets ≥ 44px",
     "P1", "Major", "Automated (Selenium)", "Apple HIG minimum",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Touch gestures", "Tap (not long-press) triggers nav-link route change",
     "1. D-360\n2. ActionChains: pointer_down → wait 50ms → pointer_up on nav link\n3. Verify hash change",
     "Short tap routes correctly",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "selenium ActionChains"),
    ("Touch gestures", "Double-tap does not zoom (touch-action: manipulation)",
     "1. D-360, load any route\n2. Inspect body computed touch-action\n3. Verify value includes 'manipulation' OR meta viewport user-scalable=no",
     "No 300ms tap delay; no double-tap zoom on UI chrome",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("PWA install", "PWA installable prompt on Android Chrome",
     "1. Real Android Chrome\n2. Visit app, dismiss any banner\n3. Open menu → 'Install app'\n4. Verify install succeeds",
     "App installs as standalone PWA",
     "P2", "Minor", "Manual", "beforeinstallprompt event automation is brittle",
     "20 min", "Mobile UI engineer", "Real Android device"),
    ("PWA install", "Add-to-Home-Screen on iOS Safari",
     "1. iOS Safari\n2. Share → Add to Home Screen\n3. Open from springboard\n4. Verify standalone display (no Safari chrome)",
     "Standalone PWA launches",
     "P2", "Minor", "Manual", "iOS only supports manual A2HS",
     "15 min", "Mobile UI engineer", "Real iOS device"),
    ("Service worker", "Offline mode: previously-visited routes load offline",
     "1. D-360 online; visit #/home, #/learn/grammar\n2. driver.execute_cdp_cmd('Network.emulateNetworkConditions', {offline:True})\n3. Reload each route\n4. Verify content renders",
     "Routes load from SW cache",
     "P1", "Major", "Automated (Selenium)", "Chrome CDP supports offline emulation",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py + CDP"),
    ("Service worker", "Service worker update on new deploy (skipWaiting + clientsClaim)",
     "1. Visit app; SW activates\n2. Deploy new build with bumped CACHE_VERSION\n3. Reload\n4. Verify new version active within 1 reload",
     "SW updates in ≤ 1 reload",
     "P2", "Minor", "Manual", "Requires actual deploy or local SW simulator",
     "20 min", "Mobile UI engineer", "Manual"),
    ("Network slow", "App functions on 3G throttle without UI freezing",
     "1. D-360, CDP Network.emulateNetworkConditions(downloadThroughput=400kbps, latency=400ms)\n2. Load each route\n3. Verify skeleton appears within 200ms; content within 5s",
     "Skeleton + LCP < 5s on 3G",
     "P2", "Minor", "Automated (Selenium)", "",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py + CDP"),
    ("Network flap", "Network ON → OFF → ON does not break the in-progress flow",
     "1. D-360, start a quiz at #/test\n2. Mid-question, set offline\n3. Submit answer (should fail gracefully OR queue)\n4. Restore network\n5. Verify state recovers",
     "Quiz state preserved across network flap",
     "P2", "Minor", "Automated (Selenium)", "",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Background tab", "Background-tab transition pauses audio + does not crash",
     "1. D-390, play audio on #/listening/story/{id}\n2. Switch tab via driver.switch_to.window\n3. Verify audio pauses (HTMLAudioElement.paused === true)",
     "Audio pauses in background",
     "P2", "Minor", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Browser back", "Hardware/software back-button navigates SPA history",
     "1. D-360, navigate Home → Grammar → vocab → kanji\n2. driver.back() x 3\n3. Verify hash returns to '#/home'",
     "Back walks history correctly",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Browser back", "Back from mid-quiz prompts before discarding answer",
     "1. D-360, start #/test\n2. Answer 2/10 questions\n3. driver.back()\n4. Verify confirm prompt OR auto-save",
     "User warned before destructive nav",
     "P2", "Minor", "Manual", "Depends on app's onbeforeunload/popstate handling",
     "15 min", "Mobile UI engineer", "Manual"),
    ("Pull-to-refresh", "Pull-to-refresh does not break SPA state",
     "1. Real mobile Chrome\n2. From #/learn/grammar mid-scroll, pull-down to refresh\n3. Verify route reloads OR is suppressed (overscroll-behavior: contain)",
     "PTR behavior consistent across routes",
     "P2", "Minor", "Manual", "Pull-to-refresh is touch-gesture; Selenium can't simulate fully",
     "20 min", "Mobile UI engineer", "Real device"),
    ("Audio playback", "First audio play requires user gesture on iOS",
     "1. iOS Safari\n2. Navigate #/listening/story/{any}, attempt programmatic .play()\n3. Verify NotAllowedError\n4. Tap play button\n5. Verify audio plays",
     "Audio requires gesture on iOS; works post-tap",
     "P1", "Major", "Manual", "Selenium can't fire trusted-event gesture; use Appium",
     "20 min", "Mobile UI engineer", "Appium / real iOS"),
    ("Audio playback", "Audio play button has visible loading + playing + paused states",
     "1. D-390, tap play\n2. Capture screenshot during load (spinner)\n3. Capture during play (pause icon)\n4. Capture paused (play icon)",
     "3 distinct states visible",
     "P2", "Minor", "Automated (Selenium)", "",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py + visual"),
    ("IME / typing", "Japanese romaji input works in production-type drills",
     "1. D-390, navigate #/drill (production type)\n2. Tap input\n3. Type 'tabemasu' via ActionChains.send_keys\n4. Verify input value becomes 'たべます' (if romaji-mode on)",
     "Romaji→kana conversion works",
     "P1", "Major", "Manual", "Romaji mode is app-internal; native IME needs Appium",
     "20 min", "Mobile UI engineer", "Manual + Appium"),
    ("IME / typing", "Soft keyboard does not occlude the focused input",
     "1. D-390\n2. Focus input near bottom of viewport\n3. Verify visualViewport API shows shrunk viewport\n4. Verify input remains visible via scrollIntoView",
     "Focused input stays visible above keyboard",
     "P1", "Major", "Manual", "Selenium can't open soft keyboard reliably",
     "20 min", "Mobile UI engineer", "Appium"),
    ("Performance", "First Contentful Paint < 2.5s on D-360 throttled",
     "1. Lighthouse mobile run on Pixel 4 emulation\n2. 4x CPU + 4G network\n3. Read LCP / FCP values",
     "FCP < 2.5s, LCP < 4s",
     "P1", "Major", "Automated (Lighthouse)", "",
     "10 min", "Mobile UI engineer", "Lighthouse CI"),
    ("Performance", "Memory: 1-hour study session does not exceed 150 MB heap",
     "1. D-360, drive through 5 modules over 60 min via test script\n2. window.performance.memory.usedJSHeapSize\n3. Verify peak < 150 MB",
     "Heap stays bounded",
     "P2", "Minor", "Automated (Selenium)", "",
     "70 min", "Mobile UI engineer", "ui_mobile_suite.py"),
]

# Per-screen scenarios. Each screen gets initial-load / primary-interaction /
# navigation-away / locale-switch / overflow-or-empty-state where applicable.
PER_SCREEN = [
    # ---- #/home ----
    ("Screen: Home (#/home)", "Home screen primary CTAs are tappable + above-the-fold on D-360",
     "1. D-360, load /\n2. Wait for skeleton to clear (.skeleton-wrap hidden)\n3. Verify primary CTA buttons exist with rect.top < window.innerHeight",
     "Primary CTAs above the fold",
     "P1", "Major", "Automated (Selenium)", "Home.js renders primary cards",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Home (#/home)", "Home noscript fallback renders text-only outline when JS off",
     "1. Disable JS in driver options\n2. Load /\n3. Verify <noscript> content visible (h2 'JLPT N5 — Free study material')",
     "Noscript content rendered",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Home (#/home)", "Home → each module nav-link tappable + routes correctly",
     "1. D-360, load #/home\n2. For each of Grammar/Vocab/Kanji/Reading/Listening/Test/Mock/Missed/Progress nav links\n3. Tap → verify hash matches expected route",
     "All 9 nav links route correctly",
     "P1", "Major", "Automated (Selenium)", "",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Home (#/home)", "Home trust-strip text wraps cleanly at 320px",
     "1. D-320, load #/home\n2. Inspect .footer-trust-strip\n3. Verify no horizontal overflow + readable line-height",
     "Trust strip wraps cleanly",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/learn ----
    ("Screen: Learn (#/learn)", "Learn overview shows both Grammar + Vocab entry points",
     "1. D-360, load #/learn\n2. Verify links a[href='#/learn/grammar'] AND a[href='#/learn/vocab'] present + tappable",
     "Both entry points visible",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Learn (#/learn)", "Learn overview last-viewed deep-link visible if state exists",
     "1. D-360, prime localStorage with lastViewed pattern\n2. Load #/learn\n3. Verify 'resume' card or link to #/learn/{id}",
     "Resume affordance shown",
     "P2", "Minor", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/learn/grammar ----
    ("Screen: Grammar list (#/learn/grammar)", "Grammar list paginates or virtual-scrolls 178 patterns without freeze on D-360",
     "1. D-360, load #/learn/grammar\n2. Wait for first item\n3. Time-to-interactive < 3s\n4. Scroll to end\n5. Verify all items accessible",
     "List performant + scroll-complete",
     "P1", "Major", "Automated (Selenium)", "178 patterns is a stress test on low-RAM Android",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Grammar list (#/learn/grammar)", "Grammar list per-item tap area ≥ 44px",
     "1. D-360\n2. Sample first 10 grammar list items\n3. Each must have min(width,height) >= 44",
     "All list items meet HIG min",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Grammar list (#/learn/grammar)", "Grammar list filters/search input is mobile-friendly",
     "1. D-360, load #/learn/grammar\n2. If filter input present, tap + type\n3. Verify list filters; no auto-zoom",
     "Filter works without zoom",
     "P2", "Minor", "Automated (Selenium)", "Depends on app implementation",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/learn/{patternId} ----
    ("Screen: Grammar detail (#/learn/{id})", "Grammar detail page renders example list scrollable on D-360",
     "1. D-360, load #/learn/n5-001 (or any pid)\n2. Verify examples render (each example.ja, .en, audio button)\n3. Verify scroll works",
     "Detail page scrolls cleanly",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Grammar detail (#/learn/{id})", "Audio play button on grammar examples is tappable",
     "1. D-360, load grammar detail\n2. Find audio play button per example\n3. Tap, verify HTMLAudioElement plays",
     "Audio plays on tap",
     "P1", "Major", "Automated (Selenium)", "May require user-gesture on iOS",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Grammar detail (#/learn/{id})", "Furigana toggle / ruby-text legibility at D-320",
     "1. D-320, load grammar detail with kanji\n2. Verify rt elements (furigana) render above kanji\n3. Verify font-size of rt ≥ 9pt",
     "Furigana legible at smallest size",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Grammar detail (#/learn/{id})", "Prev/Next pattern navigation buttons are tappable",
     "1. D-360, load pattern detail\n2. Tap 'next' link a[href*='#/learn/']\n3. Verify hash updates",
     "Prev/next navigation works",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Grammar detail (#/learn/{id})", "Wrong-form / corrected-pair section wraps cleanly",
     "1. D-360, load any pattern with wrong_corrected_pair\n2. Verify both 'wrong' + 'correct' lines visible",
     "Both lines visible + readable",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/learn/vocab ----
    ("Screen: Vocab list (#/learn/vocab)", "Vocab list loads 1009 entries on D-360 without freeze",
     "1. D-360, load #/learn/vocab\n2. TTI < 3s\n3. Scroll to end of list",
     "Smooth scroll across all 1009",
     "P1", "Major", "Automated (Selenium)", "Virtual scrolling expected",
     "20 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Vocab list (#/learn/vocab)", "Vocab list section headers sticky-pin on scroll (if implemented)",
     "1. D-360, scroll through alphabetical/category groups\n2. Verify section header stays visible while group is in viewport",
     "Section headers sticky",
     "P3", "Minor", "Automated (Selenium)", "Only test if feature exists",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/learn/vocab/{form} ----
    ("Screen: Vocab detail (#/learn/vocab/{form})", "Vocab detail renders all schema fields on mobile",
     "1. D-360, load #/learn/vocab/食べる (or any form)\n2. Verify reading, gloss, gloss_hi (if HI locale), examples, audio button",
     "All fields render",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Vocab detail (#/learn/vocab/{form})", "Pitch-accent diagram (if present) is touch-sized",
     "1. D-360, load vocab detail\n2. If .pitch-accent-diagram exists, verify min height ≥ 30px (readable)",
     "Pitch diagram legible",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Vocab detail (#/learn/vocab/{form})", "Vocab detail prev/next pagination via tap",
     "1. D-360, load any vocab form\n2. Tap next-link a[href*='#/learn/vocab/']\n3. Verify hash updates",
     "Prev/next works",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/kanji ----
    ("Screen: Kanji list (#/kanji)", "Kanji grid is touch-friendly at D-360",
     "1. D-360, load #/kanji\n2. Inspect kanji grid items\n3. Verify each grid cell ≥ 44x44",
     "Grid items meet touch min",
     "P1", "Major", "Automated (Selenium)", "106 kanji",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Kanji list (#/kanji)", "Kanji glyph renders large enough to read on D-320",
     "1. D-320\n2. Inspect any kanji glyph computed font-size\n3. Verify >= 24px",
     "Glyph readable",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/kanji/{glyph} ----
    ("Screen: Kanji detail (#/kanji/{glyph})", "Kanji detail renders stroke-order SVG on mobile",
     "1. D-360, load #/kanji/食\n2. Verify .stroke-order-svg renders\n3. Verify svg.viewBox or width responsive",
     "Stroke-order SVG renders",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Kanji detail (#/kanji/{glyph})", "Kanji stroke-order animation playable via tap",
     "1. D-360, load kanji detail\n2. Tap play-stroke-order button\n3. Verify animation runs",
     "Animation plays",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Kanji detail (#/kanji/{glyph})", "Kanji popover does not occlude entire viewport on D-360",
     "1. D-360, load any page with kanji popover triggers\n2. Tap a kanji\n3. Verify popover height < viewport.height",
     "Popover fits within viewport",
     "P2", "Minor", "Automated (Selenium)", "kanji-popover.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Kanji detail (#/kanji/{glyph})", "Compound-words list (n5_compounds) is scrollable",
     "1. D-360, load kanji with compound list\n2. Verify each compound is tappable + links to vocab/{form}",
     "Compound links work",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/reading ----
    ("Screen: Reading list (#/reading)", "Reading list shows all 54 passages on D-360",
     "1. D-360, load #/reading\n2. Count list items\n3. Verify >= 54",
     "All 54 passages listed",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Reading list (#/reading)", "Reading list item tap routes to #/reading/{id}",
     "1. D-360\n2. Tap first reading item\n3. Verify hash starts with '#/reading/'",
     "Item tap routes correctly",
     "P1", "Major", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/reading/{id} ----
    ("Screen: Reading detail (#/reading/{id})", "Reading passage line-height + font-size readable at D-320",
     "1. D-320, load #/reading/{any}\n2. Inspect passage paragraph\n3. Verify font-size >= 16px, line-height >= 1.5",
     "Body text readable",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Reading detail (#/reading/{id})", "Vocab-preview inline popover works on tap (not hover)",
     "1. D-360, load reading detail\n2. Tap a vocab_preview underline\n3. Verify popover renders + reading/gloss visible",
     "Popover opens on tap",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Reading detail (#/reading/{id})", "Reading audio play / pause works",
     "1. D-360, load reading with audio\n2. Tap play\n3. Verify audio.play() succeeds (post-gesture)",
     "Audio plays",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Reading detail (#/reading/{id})", "Quiz questions below passage are tappable",
     "1. D-360, load reading detail\n2. Scroll to questions\n3. Tap a choice radio/button\n4. Verify state change",
     "Question UI works",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Reading detail (#/reading/{id})", "Question rationale (explanation_en/hi) wraps cleanly",
     "1. D-360, submit answer\n2. Verify rationale shows + wraps without horizontal scroll",
     "Rationale wraps",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/listening ----
    ("Screen: Listening list (#/listening)", "Listening list shows all 50 drills",
     "1. D-360, load #/listening\n2. Count items >= 50",
     "All drills listed",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Listening list (#/listening)", "Listening list per-item shows mondai type + duration",
     "1. D-360, inspect item meta\n2. Verify mondai + duration text visible",
     "Item meta visible",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/listening/story ----
    ("Screen: Listening story (#/listening/story)", "Listening-story list renders + each story tappable",
     "1. D-360, load #/listening/story\n2. Verify story chapter list renders",
     "Story list visible",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Listening story (#/listening/story/{id})", "Listening-story scrubber/seek-bar usable via touch",
     "1. D-360, load any listening-story chapter\n2. Verify play + scrub controls are touch-sized",
     "Player controls touch-sized",
     "P1", "Major", "Automated (Selenium)", "listening-story.js + audio-player.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Listening story (#/listening/story/{id})", "Listening transcript reveal works without breaking layout",
     "1. D-360, tap 'show transcript'\n2. Verify transcript renders below player; no overflow",
     "Transcript reveals cleanly",
     "P1", "Major", "Automated (Selenium)", "listening-transcript.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/test ----
    ("Screen: Test (#/test)", "Test landing page CTAs are above the fold",
     "1. D-360, load #/test\n2. Verify primary CTA (start test) visible without scroll",
     "CTA above the fold",
     "P1", "Major", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Test (#/test)", "Test in-progress: choice options stack vertically on D-360",
     "1. D-360, start test\n2. Inspect choice container\n3. Verify flex-direction: column OR vertical stacking",
     "Choices stack",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Test (#/test)", "Test submit button doesn't shift on soft-keyboard appearance",
     "1. D-390, start test with text-input question\n2. Focus input (keyboard opens)\n3. Verify submit button still visible OR scrolls into view",
     "Submit reachable",
     "P1", "Major", "Manual", "Soft keyboard test",
     "15 min", "Mobile UI engineer", "Appium"),
    ("Screen: Test (#/test)", "Test progress indicator (X of N) visible during scroll",
     "1. D-360\n2. Verify progress bar/text stays sticky OR re-shows on scroll",
     "Progress visible",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/sitting ----
    ("Screen: Mock sitting (#/sitting)", "Mock-paper sitting flow renders paper-list on D-360",
     "1. D-360, load #/sitting\n2. Verify paper-list cards render",
     "Paper list visible",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Mock sitting (#/sitting/{n}/{idx})", "In-progress sitting timer visible + counts down",
     "1. D-360, start mock-paper sitting\n2. Verify timer DOM element\n3. Wait 5s, verify timer decremented",
     "Timer counts down",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Mock sitting (#/sitting/{n}/{idx})", "Inter-section navigation (next section) works on mobile",
     "1. D-360, complete one section\n2. Tap 'next section' button\n3. Verify hash updates to /sitting/{n}/{idx+1}",
     "Section navigation works",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Mock sitting (#/sitting/{n}/{idx})", "Mock sitting persists state on page reload mid-paper",
     "1. D-360, answer 3 questions\n2. driver.refresh()\n3. Verify answers preserved + position restored",
     "State persists on reload",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/missed ----
    ("Screen: Missed (#/missed)", "Missed-items list shows wrong-answer history",
     "1. D-360, prime localStorage with missed-items state\n2. Load #/missed\n3. Verify list renders",
     "Missed list renders",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Missed (#/missed)", "Missed-items empty state shows helpful message",
     "1. D-360, clear localStorage\n2. Load #/missed\n3. Verify 'no missed items yet' empty state",
     "Helpful empty state",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/summary ----
    ("Screen: Progress / Summary (#/summary)", "Summary charts fit within viewport on D-360",
     "1. D-360, load #/summary\n2. Verify chart/graph elements don't overflow",
     "Charts fit",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Progress / Summary (#/summary)", "Summary heatmap legible at D-320",
     "1. D-320, load #/summary\n2. Verify heatmap cells distinguishable",
     "Heatmap legible",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Progress / Summary (#/summary)", "Per-module progress bars sized appropriately",
     "1. D-360\n2. Verify each progress bar height >= 6px, width responsive",
     "Bars sized correctly",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/settings ----
    ("Screen: Settings (#/settings)", "Settings toggles all reachable + tappable on D-360",
     "1. D-360, load #/settings\n2. For each toggle/select, verify size + tap response",
     "All settings interactive",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Settings (#/settings)", "Locale toggle in settings persists + reflects in toggle button",
     "1. D-360, change locale in settings\n2. Navigate to #/home\n3. Verify locale-toggle label matches new locale",
     "Settings ↔ toggle sync",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Settings (#/settings)", "Reset-progress destructive action requires confirmation",
     "1. D-360, tap 'reset progress'\n2. Verify confirm modal\n3. Verify localStorage not cleared until confirm",
     "Destructive action gated",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/changelog ----
    ("Screen: Changelog (#/changelog)", "Changelog renders markdown with mobile-friendly typography",
     "1. D-360, load #/changelog\n2. Verify markdown rendered (headings/lists/code)\n3. Verify no horizontal scroll",
     "Markdown renders mobile-friendly",
     "P2", "Minor", "Automated (Selenium)", "md-viewer.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/privacy ----
    ("Screen: Privacy (#/privacy)", "Privacy policy text wraps + body font readable",
     "1. D-320, load #/privacy\n2. Verify body font-size >= 16px",
     "Privacy readable",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/notices ----
    ("Screen: Notices (#/notices)", "Notices / licenses page lists all dependencies",
     "1. D-360, load #/notices\n2. Verify kanjium, VOICEVOX, etc. attributions present",
     "Attributions visible",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/feedback ----
    ("Screen: Feedback (#/feedback)", "Feedback form (or link to GitHub Issues) tappable",
     "1. D-360, load #/feedback\n2. Verify feedback CTA + textarea/link",
     "Feedback path clear",
     "P1", "Major", "Automated (Selenium)", "feedback.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Feedback (#/feedback)", "Feedback textarea does not occlude submit when keyboard open",
     "1. D-390, focus feedback textarea\n2. Verify submit button visible OR scrolls into view",
     "Submit reachable",
     "P1", "Major", "Manual", "Soft keyboard test",
     "15 min", "Mobile UI engineer", "Appium"),

    # ---- #/diagnostic ----
    ("Screen: Diagnostic (#/diagnostic)", "Diagnostic placement-test loads + first question rendered",
     "1. D-360, load #/diagnostic\n2. Verify first question renders",
     "Diagnostic starts",
     "P1", "Major", "Automated (Selenium)", "diagnostic.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Diagnostic (#/diagnostic)", "Diagnostic completion result page renders recommendation",
     "1. D-360, complete diagnostic\n2. Verify recommendation card with link to suggested module",
     "Recommendation rendered",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/drill ----
    ("Screen: Drill (#/drill)", "Drill cloze / production type renders correctly on D-360",
     "1. D-360, load #/drill\n2. Verify question + input field render",
     "Drill UI works",
     "P1", "Major", "Automated (Selenium)", "drill.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Drill (#/drill)", "Drill answer submission feedback (correct/wrong) visible",
     "1. D-360, submit answer\n2. Verify correct/wrong indicator + explanation",
     "Feedback shown",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Drill (#/drill)", "Drill 'next' button advances to next question",
     "1. D-360, submit answer, tap next\n2. Verify new question loaded",
     "Next advances",
     "P1", "Major", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/examday ----
    ("Screen: Exam day (#/examday)", "Exam-day countdown / checklist renders on D-360",
     "1. D-360, load #/examday\n2. Verify checklist items + countdown if set",
     "Exam-day content renders",
     "P2", "Minor", "Automated (Selenium)", "exam-day.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Exam day (#/examday)", "Date-input control usable on mobile (native date picker)",
     "1. D-360, focus exam-date input\n2. Verify native date picker opens",
     "Native date picker opens",
     "P2", "Minor", "Manual", "Mobile-emulation may not show native picker",
     "10 min", "Mobile UI engineer", "Real device"),

    # ---- #/papers ----
    ("Screen: Papers list (#/papers)", "Papers category list shows all 4 categories on D-360",
     "1. D-360, load #/papers\n2. Verify {bunpou, goi, moji, dokkai} cards",
     "All 4 categories listed",
     "P1", "Major", "Automated (Selenium)", "papers.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Papers category (#/papers/{cat})", "Per-category paper list renders all 7 papers",
     "1. D-360, load #/papers/dokkai\n2. Verify paper-list >= 7 items",
     "Papers listed",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Paper detail (#/papers/{cat}/{n})", "Paper detail question list scrollable + answerable",
     "1. D-360, load any paper\n2. Verify questions render; answer first question; submit",
     "Paper UI works",
     "P1", "Major", "Automated (Selenium)", "",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Paper detail (#/papers/{cat}/{n})", "Paper print-friendly link visible",
     "1. D-360, load paper\n2. Verify a[href*='#/print/'] present",
     "Print link visible",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/print ----
    ("Screen: Print (#/print)", "Print route renders printable layout",
     "1. D-360, load #/print/{id}\n2. Verify body[data-route='print']\n3. CSS @media print: inspect print-only styles",
     "Print layout active",
     "P2", "Minor", "Automated (Selenium)", "print-paper.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Print (#/print/{id}?key=1)", "Answer-key URL param toggles answer reveal",
     "1. Load #/print/{id}?key=1\n2. Verify answers visible\n3. Reload without ?key=1\n4. Verify answers hidden",
     "Key toggle works",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/review ----
    ("Screen: Review (#/review)", "SRS review queue renders due items on D-360",
     "1. D-360, prime localStorage with SRS state\n2. Load #/review\n3. Verify due items render",
     "Review queue renders",
     "P1", "Major", "Automated (Selenium)", "review.js",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Review (#/review)", "Review again/good/easy buttons sized + spaced for thumb",
     "1. D-360, in review session\n2. Verify 3-4 grading buttons each >= 44px tall\n3. Verify spacing >= 8px between them",
     "Grade buttons thumb-friendly",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Review (#/review)", "Review empty-queue state shows when 0 due",
     "1. D-360, clear due state\n2. Load #/review\n3. Verify 'all caught up' empty state",
     "Empty state shown",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/weakareas ----
    ("Screen: Weak areas (#/weakareas)", "Weak-areas list renders + tap routes to remediation",
     "1. D-360, prime weak-area state\n2. Load #/weakareas\n3. Tap an item\n4. Verify routes to relevant lesson",
     "Weak-area drill-down works",
     "P1", "Major", "Automated (Selenium)", "weak-areas.js",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/strategy ----
    ("Screen: Strategy (#/strategy)", "Strategy page / modal renders without horizontal overflow",
     "1. D-360, load #/strategy\n2. Verify content fits viewport",
     "Strategy fits",
     "P2", "Minor", "Automated (Selenium)", "strategy.js / strategy-modal.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/levels ----
    ("Screen: Levels (#/levels)", "Levels page lists tiers + counts",
     "1. D-360, load #/levels\n2. Verify tier sections + per-tier counts",
     "Levels renders",
     "P2", "Minor", "Automated (Selenium)", "levels.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/mining ----
    ("Screen: Mining (#/mining)", "Mining page renders + import/export controls visible",
     "1. D-360, load #/mining\n2. Verify import/export buttons",
     "Mining UI renders",
     "P2", "Minor", "Automated (Selenium)", "mining.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Mining (#/mining)", "Anki-export download triggers on mobile",
     "1. D-360, tap export to Anki\n2. Verify download dialog OR file blob URL",
     "Export works",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- #/authentic ----
    ("Screen: Authentic (#/authentic)", "Authentic-context items render with images / signs",
     "1. D-360, load #/authentic\n2. Verify images render with proper srcset/responsive sizing\n3. Verify alt text present",
     "Authentic items render",
     "P1", "Major", "Automated (Selenium)", "authentic.js",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Screen: Authentic (#/authentic)", "Authentic-item context popover / modal usable on mobile",
     "1. D-360, tap authentic item\n2. Verify modal opens, scrollable, dismissable",
     "Modal mobile-usable",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- Search ----
    ("Search functionality", "Global search returns results across grammar/vocab/kanji",
     "1. D-360, focus #search-input\n2. Type 'taberu'\n3. Verify autocomplete shows grammar/vocab/kanji matches",
     "Cross-corpus search works",
     "P1", "Major", "Automated (Selenium)", "search.js",
     "15 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Search functionality", "Search results dismiss on outside-tap",
     "1. Open search, type query\n2. Tap outside results\n3. Verify results hide",
     "Dismiss works",
     "P2", "Minor", "Automated (Selenium)", "",
     "5 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Search functionality", "Search result tap routes correctly + closes search overlay",
     "1. Open search, type query\n2. Tap a result\n3. Verify hash updates AND overlay closes",
     "Result nav works",
     "P1", "Major", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- Onboarding ----
    ("Onboarding (first visit)", "First-visit auto-detect-locale toast appears once",
     "1. Clear localStorage, set Accept-Language=hi\n2. Load /\n3. Verify locale-suggestion toast appears\n4. Reload, verify toast does NOT reappear",
     "Toast fires once",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Onboarding (first visit)", "Service-worker registers on first visit without blocking content",
     "1. Clear cache + SW\n2. Load /\n3. Verify content renders before SW activate event",
     "SW registers non-blocking",
     "P2", "Minor", "Automated (Selenium)", "pwa.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),

    # ---- Accessibility cross-cut ----
    ("A11y on mobile", "All interactive elements have accessible names (axe-core mobile run)",
     "1. D-360, load each route\n2. Run axe-core via JS injection\n3. Verify 0 violations",
     "0 axe violations per route",
     "P1", "Major", "Automated (Selenium + axe-core)", "Cross-ref tab E. Accessibility",
     "45 min", "Mobile UI engineer", "ui_mobile_suite.py + axe-core"),
    ("A11y on mobile", "VoiceOver navigation through home → grammar → detail completes",
     "1. Real iPhone with VoiceOver enabled\n2. Navigate home to a grammar detail via SR gestures\n3. Verify each step announces sensibly",
     "VO navigation works",
     "P2", "Minor", "Manual", "Cannot automate VoiceOver",
     "30 min", "Mobile UI engineer", "Real iOS device"),

    # ---- Visual regression cross-cut ----
    ("Visual regression", "Per-route screenshot diff baseline at D-360 portrait",
     "1. Capture golden screenshots for each route\n2. Compare against baseline on each CI run",
     "0 unexpected diffs",
     "P2", "Minor", "Automated (Selenium + visual-diff)", "Cross-ref tab K. QA testing",
     "60 min", "Mobile UI engineer", "ui_mobile_suite.py + Percy/Playwright snapshot"),
    ("Visual regression", "Per-route screenshot at D-360 landscape",
     "1. D-360L golden screenshots per route\n2. Compare baseline",
     "0 unexpected diffs",
     "P3", "Minor", "Automated (Selenium + visual-diff)", "",
     "60 min", "Mobile UI engineer", "ui_mobile_suite.py + visual-diff"),

    # ---- Error states ----
    ("Error states", "404 / invalid-route fallback on D-360",
     "1. D-360, load #/nonexistent-route\n2. Verify 'not found' or redirect to home",
     "404 handled gracefully",
     "P2", "Minor", "Automated (Selenium)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Error states", "Corrupt-localStorage recovery on D-360",
     "1. D-360, set localStorage['n5:progress'] = '{{garbage'\n2. Reload\n3. Verify app loads (graceful degradation)",
     "App recovers from corrupt state",
     "P1", "Major", "Automated (Selenium)", "storage.js",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Error states", "Missing data/*.json (404) shows error UI, not blank page",
     "1. D-360, intercept network to 404 a JSON file\n2. Verify error UI",
     "Data 404 handled",
     "P2", "Minor", "Automated (Selenium + CDP)", "",
     "10 min", "Mobile UI engineer", "ui_mobile_suite.py + CDP"),

    # ---- Locale on every screen ----
    ("Locale on every screen", "HI locale text renders correctly on each route",
     "1. D-360, set locale=hi\n2. Visit each of 37 routes\n3. Verify Devanagari renders + no English fallback",
     "All routes localized",
     "P1", "Major", "Automated (Selenium)", "i18n.js",
     "45 min", "Mobile UI engineer", "ui_mobile_suite.py"),
    ("Locale on every screen", "HI text-expansion does not overflow buttons on D-320",
     "1. D-320, set locale=hi\n2. For each route, check buttons for text-overflow",
     "0 overflowing HI buttons",
     "P1", "Major", "Automated (Selenium)", "Cross-ref tab H. Performance / i18n",
     "30 min", "Mobile UI engineer", "ui_mobile_suite.py"),
]


def main() -> None:
    # Backup the workbook first (project backup policy)
    backup = WORKBOOK.parent / (
        WORKBOOK.name + f".bak_{date.today().strftime('%Y_%m_%d')}_mobile_ui_tab")
    if not backup.exists():
        shutil.copy2(WORKBOOK, backup)
        print(f"Backup written: {backup.name}")
    else:
        # Versioned name per backup policy (never overwrite)
        i = 2
        while True:
            alt = WORKBOOK.parent / (backup.name + f"_v{i}")
            if not alt.exists():
                shutil.copy2(WORKBOOK, alt)
                print(f"Backup written: {alt.name}")
                break
            i += 1

    wb = load_workbook(WORKBOOK)

    # If tab exists, remove and recreate (idempotent re-runs)
    if TAB_NAME in wb.sheetnames:
        del wb[TAB_NAME]

    ws = wb.create_sheet(title=TAB_NAME)

    # Title + perspective row
    ws["A1"] = TAB_NAME
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Perspectives: {PERSPECTIVES}.  "
                f"Every SPA route in index.html + js/app.js route table (as of "
                f"commit cb97df1) is covered.  "
                f"Scenarios are Selenium-runnable where the action is "
                f"automatable; some (soft-keyboard, IME, VoiceOver, iOS "
                f"audio-gesture, safe-area, PTR) are explicitly Manual / "
                f"Appium because Selenium cannot reproduce them faithfully.\n\n"
                + DEVICE_PROFILES)
    ws["A2"].alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    ws.row_dimensions[2].height = 220

    # Column headers at row 4
    for col, (h, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(col)].width = width

    # Compose all rows: cross-cutting first, then per-screen
    all_rows = []
    for scen in CROSS_CUTTING:
        all_rows.append(("X", *scen))  # X = cross-cutting prefix
    for scen in PER_SCREEN:
        all_rows.append(("S", *scen))  # S = per-screen prefix

    today = date.today().isoformat()
    for i, (group, subcat, scenario, steps, expected, priority, severity,
            ttype, notes, effort, owner, tools) in enumerate(all_rows, start=1):
        r = 4 + i
        sid = f"O-{group}-{i:03d}"
        persp_num = ""  # Mobile UI doesn't map to one of the 84 persona numbers cleanly
        row_values = [
            sid, subcat, persp_num, scenario, steps, expected,
            priority, severity, ttype, notes,
            effort, owner, tools,
            "",  # Last run date — not yet run
            "Not run",  # Last run result
            "",  # Tracker link
            "",  # Depends on
            "0%",  # Coverage %
        ]
        for col, val in enumerate(row_values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = WRAP

        # Estimate row height from longest content
        max_len = max(len(str(v)) for v in row_values)
        ws.row_dimensions[r].height = max(30, min(150, max_len // 3))

    ws.freeze_panes = "A5"

    wb.save(WORKBOOK)
    print(f"Wrote tab '{TAB_NAME}' to {WORKBOOK.name}")
    print(f"Scenarios: {len(all_rows)}  "
          f"(cross-cutting: {len(CROSS_CUTTING)}, per-screen: {len(PER_SCREEN)})")
    print(f"Routes covered (per-screen rows that hit a route): "
          f"~{len(set(s[0] for s in PER_SCREEN if s[0].startswith('Screen:')))}")


if __name__ == "__main__":
    main()
