"""File GOI-004 / GOI-005 / GOI-006 in the bug tracker as BUG-133/134/135."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
import shutil, os

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_21_goi_004_006"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
print(f"Current max_row: {ws.max_row}")

# Each entry: (title, description, severity, priority)
BUGS = []

BUGS.append((
    "GOI-004 — Off-by-one rationale_hi misalignment goi-7.6 → goi-7.7 → goi-7.8; same class as GOI-001, sweep should have caught it",
    """FILE: data/papers/goi/paper-7.json questions goi-7.6 (Q96) + goi-7.7 (Q97).

Consecutive content-mismatch bugs. Same shape as GOI-001 (goi-6.11 fix), but the original sweep didn't propagate corpus-wide.

goi-7.6 stem: A: きょうは ゆうがた に かえります。 (today I'll go home in the evening)
goi-7.6 rationale (EN): 夕方 ≈ 夜の前 (before night). [CORRECT TOPIC]
goi-7.6 rationale_hi: 「話すのが じょうず」 = 「上手に 話す」... [WRONG - this is goi-7.7's content]

goi-7.7 stem: A: たろうさんは 日本ごを 話すのが じょうずです。 (Taro speaks Japanese well)
goi-7.7 rationale (EN): 「話すのが じょうず」 = 「上手に 話す」... [CORRECT TOPIC]
goi-7.7 rationale_hi: 「(教師に) しゅくだいを 出す」 ≈ 「先生に...」 [WRONG - this is goi-7.8's content]

Off-by-one shift: goi-7.6 carries goi-7.7's rationale_hi; goi-7.7 carries goi-7.8's rationale_hi. goi-7.8 itself has correct content.

Fix direction:
1. goi-7.6 rationale_hi → rewrite for ゆうがた/夕方/夜の前 paraphrase.
2. goi-7.7 rationale_hi → rewrite for じょうずに 話す paraphrase.
3. Update both to rationale_hi_provenance=native_reviewed_2026_05_21.

Corpus-wide sweep across all 100 goi questions found ONLY 2 shifts (goi-7.6, goi-7.7). No additional shifts in goi-1.x through goi-6.x.

Proposed CI invariant (JA-137): rationale_hi must share at least 1 normalized Japanese content-token with stem_html OR correctAnswer. Catches off-by-one shifts at CI time. Would have caught both GOI-001 and GOI-004.""",
    "Major", "P2",
))

BUGS.append((
    "GOI-005 — 7 rationale fields carry fix-history / meta-commentary / version references; GOI-002 fix only closed goi-6.14 leaving 7 sibling occurrences",
    """FILE: data/papers/goi/paper-1.json, paper-3.json, paper-4.json, paper-5.json, paper-7.json.

Sweep for "replaces" / "prior version" / "previous version" / "Strict-N5:" / "documented at" / "policy applied at" / Hindi "पिछले संस्करण" / "पुराने":

  goi-1.5  (rationale): "(replaces ので which leans N4 in major textbooks)" — policy history
  goi-1.10 (rationale + rationale_hi): "はやい distractor replaces the prior あつい... पुराने あつい की जगह लेता है" — distractor-edit history
  goi-3.15 (rationale + rationale_hi): "パジャマ replaces the previous シャツ... पुराने シャツ" — distractor-edit history
  goi-4.6  (rationale): "...replaces the prior tautological 「父は医者 = 父の仕事は医者」" — replacement history including discarded pair
  goi-5.4  (rationale): "Strict-N5: drops the previous keyed form... per the same policy applied at Q97 in v1.12.13" — version + policy cross-reference
  goi-7.7  (rationale): "Strict-N5: also drops the potential form 話せます (N4) used in a previous version"
  goi-7.8  (rationale): "Strict-N5: replaces the previous keyed verb わたす... no longer appears in the goi corpus. Note: kept in kana since 持 is not in the kanji whitelist"

Fix direction: same as PAPER-003 / GOI-002 — strip fix-history from all 7 rationale (and parallel rationale_hi) fields. Keep only the first sentence (the actual paraphrase pedagogy).

Proposed JA-121 trigger extension: "replaces the prior", "replaces the previous", "previous version", "prior version", "Strict-N5:", "in v1.", "policy applied at", "no longer appears", "पिछले संस्करण", "पुराने ... की जगह".""",
    "Low", "P4",
))

BUGS.append((
    "GOI-006 — goi-7.4 rationale_hi has corrupted Japanese token あमारी ありません (Devanagari मा/ी embedded inside kana stream); should be あまく ありません",
    """FILE: data/papers/goi/paper-7.json question goi-7.4 (Q94).

rationale_hi reads: 「あमारी ありません (औपचारिक विनम्र-नकारात्मक)」.

The token 「あमारी」 is mojibake: kana あ + Devanagari मा + Devanagari ी. Expected: 「あまく」 (pure kana). Search-and-replace or transliteration pass substituted kana まく with Devanagari मा and ी.

Result:
  - invalid Japanese (mixed script in a single word)
  - invalid Hindi (मा and ी don't form a recognizable Hindi word)
  - garbles the pedagogical point — the two equivalent polite-negative forms are 「あまく ないです」 and 「あまく ありません」

Same class as the यहाँre mojibake (PAPER-004 Q72/Q80) and by ट्रेन English-fragment (DOKKAI-004). Different surface, same root: text-transformation pass that did not sanitize script boundaries.

Fix direction:
1. Replace 「あमारी ありません」 with 「あまく ありません」 in goi-7.4 rationale_hi.
2. Update rationale_hi_provenance to native_reviewed_2026_05_21.

Corpus-wide sweep with sharper detector (Devanagari LETTER — excluding danda ।/॥ punctuation — embedded inside JP word, no hyphen separator): exactly 1 hit, confirming goi-7.4 is the only true-mojibake instance.

Proposed CI invariant (JA-139): no rationale_hi token may have a Devanagari letter (excluding sentence-end danda) embedded inside a kana/CJK character run. CI tokenizes and fails on any embedded-Devanagari hit.""",
    "Major", "P2",
))

# Append at rows 136, 137, 138 (BUG-133/134/135)
for i, (title, desc, sev, prio) in enumerate(BUGS):
    r = 136 + i
    ws.cell(row=r, column=1, value='="BUG-"&TEXT(ROW()-3,"000")')
    ws.cell(row=r, column=2, value="2026-05-21")
    ws.cell(row=r, column=3, value="Content audit (goi sweep iteration 2)")
    ws.cell(row=r, column=4, value=title)
    ws.cell(row=r, column=5, value=desc)
    ws.cell(row=r, column=6, value=sev)
    ws.cell(row=r, column=7, value=prio)
    ws.cell(row=r, column=8, value="Open")
    print(f"  Row {r}: BUG-{r-3:03d} = {title[:70]}...")

wb.save(XLSX)
# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
for r in [136, 137, 138]:
    print(f"  Verify R{r}: status={ws2.cell(row=r, column=8).value} sev={ws2.cell(row=r, column=6).value}/{ws2.cell(row=r, column=7).value}")
print("Saved.")
