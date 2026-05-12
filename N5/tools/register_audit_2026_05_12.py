"""
Register the 2026-05-12 N5 richness audit findings into the canonical
issue registry (feedback/n5-audit-2026-05-04.xlsx, Items sheet).

Inputs:  feedback/n5-richness-audit-2026-05-12.md (audit transcript)
Output:  feedback/n5-audit-2026-05-04.xlsx mutated in place
         (15 Issue rows + 7 Improvement rows; ID-gap-safe append)

Per prompts/N5Improvement.txt REGISTRATION section:
- Inspect existing Items sheet (header at row 4) to find max(ISSUE-NNN)
  and max(IMP-NNN). Continue numbering from there; do NOT reuse IDs.
- 16-column schema:
    ID, Type, Severity, Priority, Impact, Effort, Category, Location,
    Title, Current state, Why-it-matters, Suggested direction,
    Dependencies, Decision (Fix/Avoid/Defer), Description, Permission.
- Decision defaults:
    Fix   = P1/P2 unblocked work
    Defer = has a precondition (engine decision, anime-quote license)
    Avoid = explicit non-feature

This is the audit's only allowed write per the prompt.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

# Force UTF-8 stdout so the diagnostic prints don't die on cp932 locale.
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "feedback" / "n5-audit-2026-05-04.xlsx"

# ---------------------------------------------------------------------------
# Findings table — 16 columns each. Numbering must align with the audit
# transcript at feedback/n5-richness-audit-2026-05-12.md (ISSUE-111..124,
# IMP-147..153). The transcript was updated to match the registry's actual
# next-IDs after inspecting the sheet (max ISSUE-110, max IMP-146 at
# audit time, 2026-05-12).
# ---------------------------------------------------------------------------

ISSUES = [
    # ISSUE-111 — Per-example grammar audio at 0/1782 (Section 0 TOP-1)
    [
        "ISSUE-111",
        "Issue",
        "MAJOR",
        "P1",
        "HIGH",
        "MEDIUM",
        "Grammar-audio / Authentic-content layer",
        "data/grammar.json patterns[].examples[].audio; tools/build_audio_voicevox.py; tools/build_audio.py",
        "Per-example grammar audio at 0/1782 examples",
        "0/1782 examples have audio rendered. Schema supports examples[].audio but field is universally null across all 178 patterns.",
        "NO incumbent ships per-example audio on grammar (Tofugu, Bunpro, JLPT Sensei stop at text). Single largest leadership-claim opportunity on grammar surface.",
        "VOICEVOX render-loop on examples[].ja text via tools/build_audio_voicevox.py; emit MP3 per example, populate examples[i].audio field.",
        "Q9 (engine decision: VOICEVOX vs gtts vs edge-tts)",
        "Defer",
        "Section 0 TOP-1 of 2026-05-12 audit. Blocked on audio engine decision (see Q9). Once engine choice is final, 1-2 person-days for render batch + UI wire-up.",
        "Requires engine decision (VOICEVOX vs gtts vs edge-tts)",
    ],
    # ISSUE-112 — Common-mistakes >=3 categorized: 0/178 (Section 0 TOP-2)
    [
        "ISSUE-112",
        "Issue",
        "MAJOR",
        "P1",
        "HIGH",
        "MEDIUM",
        "Grammar-pedagogical-depth",
        "data/grammar.json patterns[].common_mistakes[]",
        "Common-mistakes >=3 categorized into 4 N5 error types: 0/178 patterns",
        "178/178 patterns carry >=1 mistake, but 0/178 have >=3 mistakes EACH categorized into {particle, verb_class, conjugation, register}.",
        "Bunpro: 1 generic mistake per pattern. Genki workbook is the only major source systematically categorizing errors. Hitting >=3 categorized clearly exceeds Bunpro and matches a native-teacher classroom checklist.",
        "Enrich each pattern's common_mistakes to >=3 entries, each with category: 'particle' | 'verb_class' | 'conjugation' | 'register'. Renderer (learn-grammar.js) already iterates the array; add category badge.",
        "-",
        "Fix",
        "Section 0 TOP-2 of 2026-05-12 audit. Purely content-authoring; no engine change. Renderer accepts arbitrary mistake counts already.",
        "No permission required",
    ],
    # ISSUE-113 — Onomatopoeia cluster 7/1009 (Section 0 TOP-4)
    [
        "ISSUE-113",
        "Issue",
        "MAJOR",
        "P1",
        "HIGH",
        "LOW",
        "Vocab-pedagogical-depth",
        "data/vocab.json sections '33. Adverbs' + '36. Greetings & set phrases'",
        "Onomatopoeia (giongo/gitaigo) cluster: only 7/1009 entries flagged",
        "7/1009 with onomatopoeia: true flag. Canonical N5 set (pekopeko/nikoniko/dokidoki/wakuwaku/pikapika/yukkuri/chotto/dandan/motto/maamaa) not all schema-marked.",
        "NO incumbent systematically teaches mimetics at N5. Native real Japanese is mimetic-heavy; defensible richness lever called out in the prompt's largest-leverage list.",
        "Verify 10 canonical entries carry onomatopoeia: true + 2 examples each + at least one audio render.",
        "ISSUE-111 (audio infra would cover example audio here)",
        "Fix",
        "Section 0 TOP-4 of 2026-05-12 audit. 10-entry batch; tiny effort. UI badge already exists in vocab detail page when flag is present.",
        "No permission required",
    ],
    # ISSUE-114 — Voice variety 4 vs >=6 (Section 0 TOP-5)
    [
        "ISSUE-114",
        "Issue",
        "MAJOR",
        "P1",
        "HIGH",
        "MEDIUM",
        "Listening-richness",
        "data/listening.json items[].audio_render_meta.voices_used",
        "Listening voice variety: 4 distinct (edge-TTS) vs target >=6 with age-band coverage",
        "4 distinct voices (Nanami/Keita/Aoi/Daichi) rendered via edge-TTS. voice_planned exists on 50/50 items for VOICEVOX assignments but actual rendered audio is edge-TTS. 0/8 child-band and 0/8 elderly-band cells covered.",
        "4 voices = JapanesePod101 free-tier parity. 6-8 with age x gender variety is leadership level. Real exam audio diversifies speakers across age bands.",
        "Run tools/build_audio_voicevox.py to re-render 50 items with planned speaker map (Kasukabe Tsumugi / Shikoku Metan / Zundamon child / Shirakami Kotaro / Aoyama Ryusei / Suzumatsu Akeshi elderly). Update audio_render_meta.voices_used after render.",
        "Q9 (engine decision)",
        "Defer",
        "Section 0 TOP-5 of 2026-05-12 audit. Blocked on engine decision. Single batch job once engine choice is final.",
        "Requires engine decision",
    ],
    # ISSUE-115 — Vocab register tag 61/1009 (P2)
    [
        "ISSUE-115",
        "Issue",
        "MINOR",
        "P2",
        "MEDIUM",
        "MEDIUM",
        "Vocab-pedagogical-depth",
        "data/vocab.json entries[].register",
        "Vocab register tag: 61/1009 entries (6%)",
        "61/1009 entries have register tag. 43 expression entries closed in 2026-05-11 wave; residual 906 are mostly nouns/verbs where register is meaningful but unauthored.",
        "Native-teacher dimension; register-selection failures are the #1 N5 social-error class.",
        "Heuristic tagging by section (function-word neutral / set-phrase polite / casual-marker forms). Hand-curate ambiguous cases.",
        "-",
        "Fix",
        "Q7 (register-tag policy for neutral default — explicit register:'neutral' vs absent field) should be resolved as part of this fix.",
        "No permission required",
    ],
    # ISSUE-116 — Wago/Kango origin tag 0/1009 (P2)
    [
        "ISSUE-116",
        "Issue",
        "MINOR",
        "P2",
        "MEDIUM",
        "LOW",
        "Vocab-pedagogical-depth",
        "data/vocab.json entries[].register_origin (currently absent)",
        "Vocab Wago / Kango / Gairaigo origin tag: 0/1009",
        "0/1009 entries carry wago/kango/gairaigo split. Detectable algorithmically: katakana -> gairaigo; kanji-compound with Sino-Japanese sound -> kango; native morphology -> wago.",
        "Register-appropriate output depends on the wago/kango split (tabemono wago casual vs shokuryo kango formal). Native-teacher dimension absent from every named competitor.",
        "Write deterministic classifier as tools/tag_wago_kango.py pass over data/vocab.json.",
        "-",
        "Fix",
        "Mostly deterministic; ~566 kanji-compound nouns the easy case. Ambiguous gairaigo (loanwords written in hiragana/kanji) need manual review.",
        "No permission required",
    ],
    # ISSUE-117 — ambient_context_audio 0/50 (P2)
    [
        "ISSUE-117",
        "Issue",
        "MINOR",
        "P2",
        "MEDIUM",
        "HIGH",
        "Listening-realism",
        "data/listening.json items[].ambient_context_audio",
        "Ambient context audio missing on listening: 0/50",
        "0/50 listening items carry ambient-mix audio. Real exam audio has cafe/station/classroom ambience under mondai 1-2; currently plays as clean dialogue.",
        "Real-exam fidelity. JLPT N5 sokki-shori items in real audio often have light ambient context to test inference.",
        "Source CC-0 ambient loops + ffmpeg mixdown pass; OR accept terminal gap with documented rationale in Section 10 anti-items.",
        "Q3 (ambient asset sourcing path)",
        "Defer",
        "Blocked on Q3 (asset sourcing). Effort is HIGH because of mixing pipeline + asset curation, not because of code.",
        "Requires ambient asset sourcing decision",
    ],
    # ISSUE-118 — Contrasts cross-link gap 57/178 (P2)
    [
        "ISSUE-118",
        "Issue",
        "MINOR",
        "P2",
        "MEDIUM",
        "MEDIUM",
        "Grammar-pedagogical-depth",
        "data/grammar.json patterns[].contrasts",
        "Contrasts cross-link gap: 57/178 patterns still without a contrast partner",
        "121/178 (68%) have >=1 contrast. 3 prior waves (2026-05-11 + 2026-05-12) authored 14+6+14 entries across 23 pairs. Remaining 57 are mostly auxiliary-verb chains, time-marker variants, register-flexible patterns lacking obvious N5 partners.",
        "Bunpro has 1-2 contrasts per pattern; this app should match. D9 density-9 already 178/178 for the politeness-ladder dimension; contrasts is the residual gap.",
        "4th wave of ~25 more pairs from the remaining 57 single-coverage patterns (auxiliary-verb chains, time-marker variants).",
        "-",
        "Fix",
        "Some patterns genuinely lack natural N5 partners (e.g., copular ja-arimasen) and will remain terminal; aim for 150/178 not 178/178.",
        "No permission required",
    ],
    # ISSUE-119 — Kanji vocab cross-links 92/106 below >=5 (P2)
    [
        "ISSUE-119",
        "Issue",
        "MINOR",
        "P2",
        "MEDIUM",
        "MEDIUM",
        "Interconnection density",
        "data/kanji.json entries[].n5_compounds",
        "Kanji vocab cross-links: 92/106 entries below the >=5 bar",
        "avg 2.9 per kanji; 14/106 hit >=5; 68/106 hit >=2. WaniKani averages 10 vocab links per kanji.",
        "D3 density gap. Closing it raises the aggregate interconnection metric (~123 -> ~140 estimated).",
        "Deterministic vocab-scan: for each kanji, find all vocab entries containing it as substring; populate up to 8 compounds in n5_compounds.",
        "-",
        "Fix",
        "Existing n5_compounds field is truncated for many kanji. Re-derive from vocab.json with no manual authoring needed.",
        "No permission required",
    ],
    # ISSUE-120 — Vocab frequent_patterns reverse-map 161/1009 (P3)
    [
        "ISSUE-120",
        "Issue",
        "MINOR",
        "P3",
        "MEDIUM",
        "LOW",
        "Density / interconnection",
        "data/vocab.json entries[].frequent_patterns",
        "Vocab frequent_patterns reverse-map: 161/1009 (16%) with >=3 patterns",
        "161/1009 entries have >=3 patterns in frequent_patterns. Easily auto-derivable from grammar.examples[].vocab_ids index.",
        "D2b reverse density target >=3 per high-freq vocab; current 1.1 avg. Doubles cross-link richness in vocab detail page.",
        "Write tools/build_frequent_patterns.py that inverts the grammar.examples[].vocab_ids index and populates vocab.frequent_patterns for every entry referenced by >=1 pattern.",
        "-",
        "Fix",
        "Purely deterministic; can be a CI-rebuild step (re-run after any grammar/vocab edit).",
        "No permission required",
    ],
    # ISSUE-121 — Transitivity pair bidirectional incomplete (P3)
    [
        "ISSUE-121",
        "Issue",
        "MINOR",
        "P3",
        "LOW",
        "LOW",
        "Vocab-pedagogical-depth",
        "data/vocab.json entries[].transitivity_pair",
        "Transitivity pair tagging incomplete (Genki-14 not all bidirectional)",
        "20/1009 entries carry transitivity_pair. All 9 Genki-14 pairs present but only one direction tagged per pair (e.g., hairu has -> ireru, but ireru lacks back-pointer).",
        "Bidirectional density; renderer-side, the pair shows one-way only. Native-teacher dimension (vi/vt pairs are the #1 N5 verb-form confusion).",
        "Bidirectional fill pass: for every pair authored, ensure both members carry the partner reference.",
        "-",
        "Fix",
        "9 pairs x 2 directions = 18 edits. ~30 minutes of work.",
        "No permission required",
    ],
    # ISSUE-122 — Real-world signage refs on kanji 18/106 (P3)
    [
        "ISSUE-122",
        "Issue",
        "MINOR",
        "P3",
        "MEDIUM",
        "MEDIUM",
        "Authentic-content layer",
        "data/kanji.json entries[].authentic_refs",
        "Real-world signage refs on kanji: 18/106 (17%)",
        "18/106 kanji cross-linked to authentic-card content. Prompt suggests 30+ N5 kanji have unambiguous real-Japan signage uses (eki / denwa / iriguchi / deguchi / otoko / onna / dai / chu).",
        "Authentic-content layer expansion; density-3 contribution.",
        "Identify ~20 more kanji with unambiguous signage; add to data/authentic.json (new cards or extend existing card kanji_refs) + back-link.",
        "ISSUE-113 overlap",
        "Fix",
        "Q2 (whether real Japan has signage for the remaining 88 kanji) is partially answered: 20 more are credibly signage-worthy; remaining 68 are above-N5 or rare.",
        "No permission required",
    ],
    # ISSUE-123 — Kanji audio per yomi 0/106 (P3)
    [
        "ISSUE-123",
        "Issue",
        "MINOR",
        "P3",
        "LOW",
        "HIGH",
        "Kanji-richness",
        "data/kanji.json entries[].audio_yomi",
        "Kanji audio per on/kun-yomi: 0/106",
        "0/106 kanji ship per-yomi audio.",
        "Pronunciation reinforcement; NHK pedagogical convention. None of the named incumbents (WaniKani, Jisho, Renshuu) ship per-yomi audio.",
        "Render per-yomi audio via VOICEVOX (single-mora utterances) into kanji[].audio_yomi.{on:'...mp3', kun:'...mp3'}.",
        "Q9 (engine decision)",
        "Defer",
        "Blocked on engine decision. Once engine is chosen, ~2 audio files per kanji x 106 = 212 single-mora renders.",
        "Requires engine decision",
    ],
    # ISSUE-124 — Anime/drama citation 0/178 (P4 elevation to TOP-3)
    [
        "ISSUE-124",
        "Issue",
        "IMPROVEMENT",
        "P2",
        "HIGH",
        "MEDIUM",
        "Authentic-content layer",
        "data/grammar.json patterns[].authentic_refs",
        "Anime / J-drama / manga citation layer on grammar: 0/178 patterns cite specific works",
        "12/178 patterns have authentic_refs (all card-based). 0/178 cite anime/drama/manga directly. Prompt names canonical N5-friendly works: Shirokuma Cafe / Chibi Maruko-chan / Sazae-san / Yotsuba&! / ARIA / Doraemon.",
        "Largest single richness lever per prompt's strategic framing. NONE of Bunpro/Tofugu/WaniKani/Renshuu/JapanesePod101 ship systematic anime/drama citations at N5.",
        "Author 20% x 178 = ~36 patterns with 1 anime/drama citation each. Reuse authentic_refs field; one-session authoring batch.",
        "Q4 (anime-quote licensing path)",
        "Defer",
        "Section 0 TOP-3 elevation. Quoting 5-10 word phrases from copyrighted anime should fall under fair-use / educational-quote in most jurisdictions, but content-policy decision needed before authoring batch.",
        "Requires fair-use / educational-quote licensing decision",
    ],
]

IMPROVEMENTS = [
    # IMP-147 — Anime/drama citation per ~20% of grammar patterns (P2, dup of ISSUE-124 for tracker visibility)
    [
        "IMP-147",
        "Improvement",
        "IMPROVEMENT",
        "P2",
        "HIGH",
        "MEDIUM",
        "Authentic-content layer",
        "data/grammar.json patterns[].authentic_refs",
        "Anime / J-drama / manga citation per ~20% of grammar patterns",
        "Same coverage state as ISSUE-124 (12/178 authentic_refs, 0/178 anime/drama citations).",
        "Mirror of ISSUE-124; filed as IMP for tracker visibility under Section 6 P2 improvements.",
        "See ISSUE-124 suggested direction.",
        "ISSUE-124, Q4",
        "Defer",
        "Cross-reference to ISSUE-124. Same content-policy gating.",
        "Requires fair-use / educational-quote licensing decision",
    ],
    # IMP-148 — Genki/MNN textbook-aligned grammar paths (P2)
    [
        "IMP-148",
        "Improvement",
        "IMPROVEMENT",
        "P2",
        "MEDIUM",
        "LOW",
        "Discoverability + paths",
        "js/learn-grammar.js + new route #/learn/grammar/path/<textbook>",
        "Genki / MNN textbook-aligned grammar paths (route)",
        "grammar.json carries sources per pattern listing Genki / MNN references but no UI surface groups them.",
        "Bunpro's stickiness comes from textbook paths. JLPT Sensei free site is ranked partially because of textbook grouping.",
        "12-page render of patterns grouped by Genki I L1..L12 + MNN Ch.1..25. Single route.",
        "-",
        "Fix",
        "Section 3 parity table item #2 (Bunpro: textbook-aligned grammar paths) currently in partial state. Data is already there; just needs route + view.",
        "No permission required",
    ],
    # IMP-149 — Review forecast 7-day (P2)
    [
        "IMP-149",
        "Improvement",
        "IMPROVEMENT",
        "P2",
        "MEDIUM",
        "LOW",
        "Daily-routine UX",
        "js/storage.js (has scaffold getReviewForecast) + new #/review/forecast route",
        "Review forecast (next-7-days projected load)",
        "js/storage.js#getReviewForecast exists; no UI surface.",
        "Bunpro's daily-routine learner stickiness lever. Helps user pace study time.",
        "Aggregate nextDue timestamps from SRS state; render histogram per day for next 7.",
        "-",
        "Fix",
        "Storage scaffolding done; single page-add. Section 3 parity table item #5.",
        "No permission required",
    ],
    # IMP-150 — SRS gating UI (P2)
    [
        "IMP-150",
        "Improvement",
        "IMPROVEMENT",
        "P2",
        "MEDIUM",
        "MEDIUM",
        "Pedagogical pacing",
        "js/storage.js#srsGatingEnabled + js/learn-vocab.js",
        "SRS gating UI integration",
        "srsGatingEnabled storage flag exists; UI integration partial. Currently no UI exposes the toggle; gated cards are not visually locked.",
        "WaniKani pacing pedagogy (vocab gated until kanji is graduated). Reduces overwhelm for new learners.",
        "Settings toggle + render gated cards as locked until kanji is graduated. Lock badge in card thumbnail.",
        "-",
        "Fix",
        "Section 3 parity table item #7 (WaniKani: SRS gating) partial -> done with UI surface.",
        "No permission required",
    ],
    # IMP-151 — Migaku sentence-mining cross-link route (P3)
    [
        "IMP-151",
        "Improvement",
        "IMPROVEMENT",
        "P3",
        "LOW",
        "MEDIUM",
        "Authentic-content cross-link surface",
        "new route #/mining/index",
        "Migaku-style sentence-mining cross-link route",
        "authentic.json + cross-links provide structured equivalent to Migaku's user-driven mining; not exposed in any single sortable index.",
        "Migaku's signature feature. Structured equivalent already exists in data; just needs a UI route.",
        "Single route showing every vocab/kanji entry's authentic-card cross-links in a sortable table.",
        "-",
        "Fix",
        "Section 3 parity table item #11 (Migaku: sentence mining) currently partial.",
        "No permission required",
    ],
    # IMP-152 — Per-pattern PDF print view (P3)
    [
        "IMP-152",
        "Improvement",
        "IMPROVEMENT",
        "P3",
        "MEDIUM",
        "MEDIUM",
        "Print / offline study",
        "js/print-paper.js",
        "Per-pattern PDF print view (JP101-parity lesson notes)",
        "print-paper.js currently covers mock papers; no per-pattern format.",
        "JapanesePod101 paid-tier signature feature (PDF lesson notes per pattern). Offline study path for school printouts.",
        "Extend print engine with per-pattern format: essay + examples + common-mistakes + print-friendly CSS.",
        "-",
        "Fix",
        "Section 3 parity table item #13 (JP101: lesson notes PDF) currently partial.",
        "No permission required",
    ],
    # IMP-153 — Reverse-map vocab->patterns (P3, dup of ISSUE-120 for tracker visibility)
    [
        "IMP-153",
        "Improvement",
        "IMPROVEMENT",
        "P3",
        "MEDIUM",
        "LOW",
        "Density / interconnection",
        "data/vocab.json entries[].frequent_patterns",
        "Reverse-map vocab->patterns (D2b density)",
        "Mirror of ISSUE-120. 161/1009 (16%) at >=3 patterns; 1.1 avg vs target >=3.",
        "Filed as IMP for tracker visibility under Section 6 P3 improvements (the D2b density gap is the lowest-scoring density dimension).",
        "Auto-derive from grammar.examples[].vocab_ids index; see ISSUE-120.",
        "ISSUE-120",
        "Fix",
        "Cross-reference to ISSUE-120. Deterministic; same fix.",
        "No permission required",
    ],
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, read_only=False, data_only=False)
    ws = wb["Items"]

    # Sanity-check header (row 4).
    header = [ws.cell(row=4, column=c).value for c in range(1, 17)]
    expected = [
        "ID",
        "Type",
        "Severity",
        "Priority",
        "Impact",
        "Effort",
        "Category",
        "Location",
        "Title",
        "Current state",
        "Why this matters / Best-in-class",
        "Suggested direction",
        "Dependencies",
        "Decision (Fix / Avoid / Defer)",
        "Description",
        "Permission decision",
    ]
    if header != expected:
        print("ERROR: header mismatch")
        print("  expected:", expected)
        print("  found:   ", header)
        return 1

    # Pre-scan existing IDs to confirm we won't collide.
    existing_ids = set()
    for r in range(5, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            existing_ids.add(str(v).strip())

    rows_to_append = ISSUES + IMPROVEMENTS
    for row in rows_to_append:
        if row[0] in existing_ids:
            print(f"ERROR: ID {row[0]} already exists in sheet; aborting")
            return 1

    # Append to first all-empty trailing row to keep deterministic layout.
    next_row = ws.max_row + 1
    while ws.cell(row=next_row - 1, column=1).value is None and next_row > 5:
        next_row -= 1

    print(f"Appending {len(rows_to_append)} rows starting at row {next_row}")
    print(f"Max ISSUE before:    ISSUE-{max(int(i[7:]) for i in existing_ids if i.startswith('ISSUE-')):03d}")
    print(f"Max IMP   before:    IMP-{max(int(i[4:]) for i in existing_ids if i.startswith('IMP-')):03d}")

    for offset, row in enumerate(rows_to_append):
        r = next_row + offset
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
        print(f"  row {r:>4}: {row[0]:<11} {row[1]:<12} P{row[3]:<3} {row[8][:60]}")

    wb.save(XLSX)
    print(f"\nWrote {len(rows_to_append)} rows to {XLSX}")
    print(f"Final max_row: {ws.max_row}")

    # Post-verify by re-reading.
    wb2 = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws2 = wb2["Items"]
    new_ids = []
    for r in range(5, ws2.max_row + 1):
        v = ws2.cell(row=r, column=1).value
        if v and (str(v).startswith("ISSUE-1") or str(v).startswith("IMP-1")):
            n = int(str(v).split("-")[1])
            if (str(v).startswith("ISSUE-") and 111 <= n <= 124) or (
                str(v).startswith("IMP-") and 147 <= n <= 153
            ):
                new_ids.append(str(v))
    print(f"\nPost-verify: {len(new_ids)} new IDs present in registry")
    print(f"  IDs: {sorted(new_ids)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
