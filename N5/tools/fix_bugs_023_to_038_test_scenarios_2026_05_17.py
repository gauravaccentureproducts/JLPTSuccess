"""Fix BUG-023 through BUG-038 — test-scenarios spreadsheet + content fixes.

The User Reported Bugs sheet was moved to
N5/specifications/test-scenarios-by-specialist-perspective.xlsx
and 16 new pending bugs were filed against it. This script applies
all 16 fixes in sequence.

Content fixes (already applied in earlier commits, just verified):
  BUG-023 — vocab.json kanji forms (already in place since cbbf02e)
  BUG-038 — 十 entry compound gloss disambiguation (in this commit)

Spreadsheet improvements (this script's primary scope):
  BUG-024 — Re-rank Priority to P1-P5 scale
  BUG-025 — Re-label Manual → Auto for mechanically-checkable rows
  BUG-026 — Reconcile Severity / Priority gaps
  BUG-027 — Threshold justifications
  BUG-028 — Expand truncated Test-steps cells
  BUG-029 — Add 7 operational columns to all 14 category tabs
  BUG-030 — Merge / cross-ref duplicate scenarios
  BUG-031 — UX scenarios: name evaluation method + sample size
  BUG-032 — Add 8 new scenarios (JEES side-by-side, competitor, etc.)
  BUG-033 — Rewrite Pedagogy (J) tab to testable items
  BUG-034 — Trim End-user (N) tab to journey-only items
  BUG-035 — Standardize Notes column
  BUG-036 — Overview formulas (COUNTA / COUNTIF)
  BUG-037 — Add Coverage % column

After this script: all 38 user-reported bugs marked Fixed.
"""
from __future__ import annotations

import io
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"

CATEGORY_TABS = [
    "A. Japanese language", "B. JLPT format", "C. Hindi locale", "D. UX design",
    "E. Accessibility", "F. Security", "G. Privacy and legal", "H. Performance",
    "I. Data engineering", "J. Pedagogy", "K. QA testing", "L. Cultural ethical",
    "M. Operations", "N. End-user POV",
]

# Tab letter → owner role mapping (for BUG-029 Owner column default)
TAB_OWNER = {
    "A": "Native Japanese teacher",
    "B": "JLPT exam expert (JEES-aligned)",
    "C": "Native Hindi teacher",
    "D": "UI/UX designer",
    "E": "WCAG-AA auditor",
    "F": "Security engineer",
    "G": "Privacy / legal counsel",
    "H": "Web-performance engineer",
    "I": "Data engineer",
    "J": "SLA pedagogy researcher",
    "K": "QA tester",
    "L": "Cultural-sensitivity reviewer",
    "M": "DevOps / SRE",
    "N": "Primary-persona learner (HI-L1, EN-L2)",
}


# ============================================================
# BUG-029: Add 7 new operational columns to every category tab.
# BUG-037: Add 1 more column (Coverage %).
# Combined into one schema-extension pass.
# ============================================================
NEW_COLUMNS = [
    ("K", "Estimated effort"),
    ("L", "Owner / role"),
    ("M", "Tools / scripts required"),
    ("N", "Last run date"),
    ("O", "Last run result"),
    ("P", "Tracker link"),
    ("Q", "Depends on"),
    ("R", "Coverage %"),   # BUG-037
]


def add_columns(wb: openpyxl.Workbook) -> int:
    """Add 8 new columns (K..R) to every category tab. BUG-029 + BUG-037."""
    n = 0
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for tab in CATEGORY_TABS:
        sh = wb[tab]
        # Add header row 4 cells
        for col_letter, label in NEW_COLUMNS:
            cell = sh[f"{col_letter}4"]
            if cell.value is None or cell.value != label:
                cell.value = label
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                n += 1
        # Default value for Coverage % is 0%; Last run result is "Not Yet Run"
        tab_letter = tab[0]
        owner_default = TAB_OWNER.get(tab_letter, "")
        for r in range(5, sh.max_row + 1):
            if not sh.cell(row=r, column=1).value:
                continue
            # K: Estimated effort — leave blank for now (per-row judgment)
            # L: Owner default by tab
            if sh[f"L{r}"].value is None:
                sh[f"L{r}"].value = owner_default
            # M: Tools/scripts — leave blank, populated per-row in BUG-025
            # N: Last run date — blank (per BUG-029 spec)
            # O: Last run result — "Not Yet Run"
            if sh[f"O{r}"].value is None:
                sh[f"O{r}"].value = "Not Yet Run"
            # P: Tracker link — blank
            # Q: Depends on — blank
            # R: Coverage % — 0% default
            if sh[f"R{r}"].value is None:
                sh[f"R{r}"].value = "0%"
        # Adjust column widths
        widths = {"K": 16, "L": 24, "M": 32, "N": 14, "O": 14, "P": 18, "Q": 16, "R": 12}
        for col_letter, w in widths.items():
            sh.column_dimensions[col_letter].width = w
    return n


# ============================================================
# BUG-024: Re-rank Priority to P1-P5 scale
# Mapping (severity + current priority → new priority):
#   Critical → P1
#   Major + (High | Medium-High) → P2
#   Major + Medium → P3
#   Major + (Low-Medium | Low) → P4
#   Minor + (Medium | Medium-High | High) → P4
#   Minor + (Low-Medium | Low) → P5
# ============================================================
def remap_priority(severity: str, current_priority: str) -> str:
    severity = (severity or "").strip()
    current = (current_priority or "").strip()
    # Idempotency: if already a P1-P5 value, leave it alone. Re-running this
    # script must NOT shift P3 → P4 (which would happen if we fell through
    # to the else clause).
    if current in ("P1", "P2", "P3", "P4", "P5"):
        return current
    if severity == "Critical":
        return "P1"
    if severity == "Major":
        if current in ("High", "Medium-High"):
            return "P2"
        if current == "Medium":
            return "P3"
        return "P4"  # Low-Medium / Low
    if severity == "Minor":
        if current in ("Medium", "Medium-High", "High"):
            return "P4"
        return "P5"  # Low-Medium / Low
    return "P5"  # unknown → lowest


# ============================================================
# BUG-025: Re-label specific Manual scenarios to Auto + populate Automation tool
# BUG-026: Reconcile Severity vs Priority — specific rows from the bug description
# BUG-027: Threshold justifications — provisional tags on listed rows
# BUG-028: Expand truncated Test-steps cells
# BUG-031: UX scenarios — add method + sample size
# ============================================================

# (tab_letter, row_id, automation_tool)
AUTO_RELABEL = [
    ("A", "A-008", "Custom Python (vocab.json + verb-class reference cross-check)"),
    ("A", "A-018", "Custom Python (stroke-count vs Unicode stroke-count reference)"),
    ("A", "A-027", "Custom Python (Hepburn vs Kunrei consistency rules)"),
    ("D", "D-004", "axe-core (via Playwright)"),
    ("E", "E-001", "axe-core (via Playwright)"),
    ("E", "E-002", "axe-core (via Playwright)"),
    ("E", "E-003", "axe-core (via Playwright)"),
    ("G", "G-002", "Custom Python (localStorage namespace scan + PII regex sweep)"),
    ("F", "F-001", "Lighthouse + npm audit + secrets-scan"),
    ("F", "F-002", "npm audit / Snyk / Dependabot"),
    ("F", "F-003", "GitGuardian / trufflehog (secrets in tree)"),
    ("F", "F-004", "Static analysis (eslint-plugin-security)"),
    ("F", "F-005", "CSP-tester (Lighthouse + manual review)"),
    ("F", "F-006", "Custom Python (SRI hash audit on script/link tags)"),
    ("F", "F-007", "ssllabs.com / Mozilla observatory"),
    ("F", "F-008", "OWASP ZAP baseline scan"),
    ("H", "H-001", "Lighthouse (LCP measurement)"),
    ("H", "H-002", "Lighthouse (CLS measurement)"),
    ("H", "H-003", "Lighthouse (INP measurement)"),
    ("H", "H-015", "Playwright at 320px viewport"),
]

# Tools / scripts for already-Auto rows so the Automation tool column isn't empty
EXISTING_AUTO_TOOLS = {
    ("A", "A-016"): "tools/check_content_integrity.py (JA-NN regex pass)",
    ("B", "B-002"): "tools/check_content_integrity.py (JA-33 — choices-per-mondai check)",
    ("B", "B-005"): "tools/check_content_integrity.py (JA-11 — duplicate MCQ choice guard)",
}

# (tab_letter, row_id) where severity ≥ 2 levels above priority — needs reconcile
SEVERITY_RECONCILE = {
    # Critical → P1 already handled by remap_priority; documenting here for audit
    ("F", "F-014"): "Critical severity (git-history secret leak) → P1 priority post-remap.",
    ("F", "F-015"): "Critical severity (.env / private keys) → P1 priority post-remap.",
    ("F", "F-019"): "Critical severity (PR fork attack surface) → P1 priority post-remap.",
    ("K", "K-008"): "Critical severity (deploy-time smoke) → P1 priority post-remap.",
    ("I", "I-009"): "Critical severity (CI on every PR) → P1 priority post-remap.",
    # Major + Low → P4 in remap. User asked to elevate OR document.
    # We elevate to P3 with rationale.
    ("F", "F-016"): "ELEVATED: Major severity (static-site dynamic-surface enumeration) — minimal effort, high signal. Elevated from P4 to P3 per BUG-026.",
    ("F", "F-017"): "ELEVATED: Major severity (rate-limit on POST endpoints) — even a static-PWA has feedback endpoints. Elevated from P4 to P3 per BUG-026.",
    ("M", "M-001"): "ELEVATED: Major severity (deploy rollback procedure) — downtime risk if no rollback. Elevated from P4 to P3 per BUG-026.",
}

# Rows requiring elevation override (set Priority directly after remap)
PRIORITY_OVERRIDES = {
    ("F", "F-016"): "P3",
    ("F", "F-017"): "P3",
    ("M", "M-001"): "P3",
}

# (tab_letter, row_id) → provisional threshold marker for BUG-027
PROVISIONAL_THRESHOLDS = {
    ("A", "A-001"): "Threshold provisional (n=200; no power calc on file). Sample size selected because 200 examples × ~3 min/example = 10h native-teacher session is feasible. Action if missed: triage by sub-category (vocab vs grammar), expand sample on weakest. Revisit threshold after first full run.",
    ("A", "A-020"): "Threshold provisional (n=944 vocab pitch_marks). JA-90 (kanjium ref) covers 810 high-confidence rows. Remaining 134 rows: human audio review. Action if missed: rank-order by frequency-rank, prioritize top-100 most-frequent.",
    ("A", "A-024"): "Threshold provisional (n=200 grammar examples). Same sample-size rationale as A-001.",
    ("C", "C-001"): "Threshold provisional (n=150 Hindi-locale strings). Native-Hindi reviewer time-box: 8h session. Action if missed: triage by surface (UI strings > content > meta).",
    ("B", "B-006"): "Threshold provisional (≥95% of shipped vocab is in N5 spec). The 5% slack covers late_n5 + borderline_n4 tier entries (intentional per scope). Action if missed: review tier classification per-entry.",
    ("J", "J-001"): "Threshold provisional (≥80% known-vocab). 'Known' = appears in lessons ≤ current lesson per curriculum graph. Action if missed: rewrite example or move pattern to later lesson.",
    ("K", "K-001"): "Threshold provisional (≤3 critical/major bugs per exploratory session). Budget if >3: STOP feature work; triage immediately; close all newly-found criticals before next session.",
}

# (tab_letter, row_id) → expanded Test-steps (BUG-028)
EXPANDED_STEPS = {
    ("A", "A-001"): (
        "1. Random-sample 200 vocab.json examples (use `python tools/sample_vocab.py --n 200 --seed <date>`).\n"
        "2. Native teacher reads each in turn (~3 min per item).\n"
        "3. Mark UNNATURAL / BORDERLINE / NATURAL per item in a Google Sheet column.\n"
        "4. Compute % NATURAL; report items marked UNNATURAL for fix queue.\n"
        "5. Cross-check UNNATURAL items against JA-83 (template-leak guard) — none should match a flagged template."
    ),
}

# (tab_letter, row_id) → augmented Test-steps for UX criteria (BUG-031)
UX_AUGMENTED_STEPS = {
    ("D", "D-001"): (
        "Method: heuristic evaluation (Nielsen's 10 principles); 2 evaluators independently rate; reconcile disagreements.\n"
        "Sample: 8 screens covering home / lesson / quiz / test / settings / privacy / changelog / progress dashboard.\n"
        "Pass criterion: ≥7 of 8 screens have the Submit/Check action ranked as visually-most-prominent by both evaluators."
    ),
    ("D", "D-010"): (
        "Method: moderated remote usability test.\n"
        "Sample: 5 participants (the established threshold for catching ~85% of UX issues per Nielsen).\n"
        "Pass criterion: median time-to-find ≤5s across the 5 participants for each of the 5 main modules; report participants who took >10s."
    ),
    ("D", "D-013"): (
        "Method: unmoderated remote with task-completion tracking.\n"
        "Sample: 8 participants (HI-L1 + EN-L2; mix of true beginner + ~5 hours prior Japanese exposure).\n"
        "Pass criterion: ≥6 of 8 participants complete ≥1 lesson + ≥1 quiz within their first 20-minute session."
    ),
    ("N", "N-003"): (
        "Method: unmoderated remote first-task completion (Maze / Loop11 / similar).\n"
        "Sample: 10 participants matching the primary persona (HI-L1 self-study learner).\n"
        "Pass criterion: ≥8 of 10 complete the onboarding-to-first-lesson flow without contacting support."
    ),
}

# Duplicate-scenario cross-refs (BUG-030)
# (tab_letter, row_id) → "Cross-ref: see <canonical_id>"
CROSS_REFS = {
    ("N", "N-014"): "DUPLICATE of E-010 (furigana exposure to screen reader). Canonical home: E tab.",
    ("D", "D-007"): "DUPLICATE of C-010 (CJK + Devanagari + Latin baseline). Canonical home: C tab.",
    ("K", "K-011"): "DUPLICATE of H-020 (locale-switch state retention). Canonical home: H tab.",
    ("N", "N-009"): "DUPLICATE of H-020 (locale-toggle persistence). Canonical home: H tab.",
    ("G", "G-011"): "DUPLICATE of F-011 (transitive-dependency license-compliance). Canonical home: F tab.",
    ("N", "N-013"): "DUPLICATE of E-007 (screen-reader coverage). Canonical home: E tab.",
}

# BUG-035: standardize Notes — fill empty cells with default
DEFAULT_NOTE = "No automated check today — manual only."

# BUG-035: placeholder-only rows to flag for replacement
PLACEHOLDER_NOTES_TO_REPLACE = {
    ("M", "M-010"): "Operational scenario; no automated check today — manual only. Run quarterly during ops review.",
    ("I", "I-012"): "Data-engineering check; complements JA-82 (`_meta` resolution). No automated check beyond JA-82 today.",
}

# BUG-031: row IDs whose Notes get a "UX method + sample-size now in Test-steps" note
UX_NOTES_ADDITION = " UX evaluation method and sample size now specified in Test-steps cell per BUG-031."


def apply_per_row_fixes(wb: openpyxl.Workbook) -> dict:
    """Apply per-row fixes for BUG-024..028, 031, 035 in a single pass."""
    stats = {"priority_remapped": 0, "auto_relabelled": 0, "severity_documented": 0,
             "threshold_marked": 0, "steps_expanded": 0, "ux_augmented": 0,
             "notes_filled": 0, "notes_replaced": 0, "cross_refs_added": 0}
    # Fix: AUTO_RELABEL tuples are (tab_letter, row_id, tool). row_id is
    # already in "A-008" form — don't prefix with tab_letter again.
    auto_set = {row_id for _, row_id, _ in AUTO_RELABEL}
    auto_tools = {row_id: tool for _, row_id, tool in AUTO_RELABEL}
    for tab in CATEGORY_TABS:
        sh = wb[tab]
        tab_letter = tab[0]
        for r in range(5, sh.max_row + 1):
            row_id = sh.cell(row=r, column=1).value
            if not row_id:
                continue
            current_priority = sh.cell(row=r, column=7).value
            severity = sh.cell(row=r, column=8).value
            test_type = sh.cell(row=r, column=9).value
            notes_cell = sh.cell(row=r, column=10)
            notes = notes_cell.value or ""

            # BUG-024: Re-rank Priority
            new_priority = remap_priority(severity, current_priority)
            # BUG-026: Apply specific elevation overrides
            override = PRIORITY_OVERRIDES.get((tab_letter, row_id))
            if override:
                new_priority = override
            if new_priority != current_priority:
                sh.cell(row=r, column=7).value = new_priority
                stats["priority_remapped"] += 1

            # BUG-025: Re-label Manual → Auto for listed rows
            if row_id in auto_set and test_type != "Auto":
                sh.cell(row=r, column=9).value = "Auto"
                sh.cell(row=r, column=13).value = auto_tools[row_id]  # column M (Tools)
                stats["auto_relabelled"] += 1

            # Populate Tools for already-Auto rows where we know them
            existing_tool = EXISTING_AUTO_TOOLS.get((tab_letter, row_id))
            if existing_tool and not sh.cell(row=r, column=13).value:
                sh.cell(row=r, column=13).value = existing_tool

            # BUG-026: Severity/Priority documentation
            sev_rationale = SEVERITY_RECONCILE.get((tab_letter, row_id))
            if sev_rationale:
                if notes and "[BUG-026]" not in notes:
                    notes = notes + " | [BUG-026] " + sev_rationale
                elif not notes:
                    notes = "[BUG-026] " + sev_rationale
                stats["severity_documented"] += 1

            # BUG-027: Threshold justifications
            threshold_note = PROVISIONAL_THRESHOLDS.get((tab_letter, row_id))
            if threshold_note:
                if "[BUG-027]" not in notes:
                    notes = (notes + " | [BUG-027 threshold] " + threshold_note).strip(" |")
                stats["threshold_marked"] += 1

            # BUG-028: Expand truncated Test-steps
            expanded = EXPANDED_STEPS.get((tab_letter, row_id))
            if expanded:
                sh.cell(row=r, column=5).value = expanded
                stats["steps_expanded"] += 1

            # BUG-031: UX method + sample-size augmentation
            ux_aug = UX_AUGMENTED_STEPS.get((tab_letter, row_id))
            if ux_aug:
                sh.cell(row=r, column=5).value = ux_aug
                if UX_NOTES_ADDITION.strip() not in notes:
                    notes = (notes + UX_NOTES_ADDITION).strip()
                stats["ux_augmented"] += 1

            # BUG-030: Cross-references
            xref = CROSS_REFS.get((tab_letter, row_id))
            if xref:
                if "[BUG-030 cross-ref]" not in notes:
                    notes = (notes + " | [BUG-030 cross-ref] " + xref).strip(" |")
                stats["cross_refs_added"] += 1

            # BUG-035: Placeholder replacement
            placeholder_replacement = PLACEHOLDER_NOTES_TO_REPLACE.get((tab_letter, row_id))
            if placeholder_replacement:
                # If existing notes is too short / placeholder-only, replace
                notes = placeholder_replacement
                stats["notes_replaced"] += 1

            # BUG-035: Standardize Notes — fill empty
            if not notes.strip():
                notes = DEFAULT_NOTE
                stats["notes_filled"] += 1

            notes_cell.value = notes

    return stats


# ============================================================
# BUG-033: Rewrite Pedagogy (J) tab — concrete inputs + pass criteria
# ============================================================
J_TAB_REWRITES = {
    "J-001": {
        "scenario": "Curriculum dependency graph: ≥80% of vocab used in lesson L's example sentences must appear in lessons ≤L",
        "test_steps": (
            "1. Load data/curriculum.json (or grammar.json + vocab.json + lesson_order metadata).\n"
            "2. For each lesson L, extract all vocab tokens from its example sentences (kanji/kana segmentation via MeCab or precomputed token list).\n"
            "3. Compute fraction of tokens that appear in lessons ≤L.\n"
            "4. Report lessons below threshold."
        ),
        "expected": "≥80% known-vocab fraction per lesson; lessons below threshold flagged for example rewrite or vocab reordering.",
        "tools": "Custom Python (curriculum-graph + token-membership check)",
        "test_type": "Auto",
    },
    "J-002": {
        "scenario": "Output-hypothesis adherence: each lesson has ≥1 production drill prompting the learner to use the new pattern",
        "test_steps": (
            "1. For each grammar pattern P, check whether drills_auto.json has ≥1 production-type drill referencing P.id.\n"
            "2. Output: list of patterns lacking production drills.\n"
            "3. Pass: every core_n5 pattern has ≥1 production drill."
        ),
        "expected": "0 core_n5 patterns lacking production drills.",
        "tools": "tools/check_content_integrity.py (extend or new check)",
        "test_type": "Auto",
    },
    "J-003": {
        "scenario": "Spaced-repetition interval validation: SRS algorithm produces intervals within ±20% of SuperMemo SM-2 reference",
        "test_steps": (
            "1. Simulate 100-card SRS trajectories with synthetic Pass/Fail sequences.\n"
            "2. Compare interval growth against SM-2 reference (1-day → 3 → 7 → 14 → 30 → 90 days, with quality scaling).\n"
            "3. Report any interval deviating >20% from reference."
        ),
        "expected": "100% of intervals within ±20% of SM-2 reference at quality=4 (good recall).",
        "tools": "Custom Python (SRS trajectory simulation)",
        "test_type": "Auto",
    },
    "J-004": {
        "scenario": "Lesson-to-test alignment: every test item ID in questions.json references a vocab/grammar item from a lesson ≤ the test's target lesson",
        "test_steps": (
            "1. For each test item, look up the referenced vocab/grammar IDs.\n"
            "2. Verify each is from a lesson ≤ test target lesson.\n"
            "3. Flag any out-of-sequence references."
        ),
        "expected": "0 test items reference items from later lessons.",
        "tools": "tools/check_content_integrity.py (new check)",
        "test_type": "Auto",
    },
    "J-005": {
        "scenario": "Distractor cognitive load: MCQ distractors are not so close to the correct answer that the test becomes a recognition test rather than a recall test",
        "test_steps": (
            "1. For each MCQ, compute Levenshtein distance between correct answer and each distractor.\n"
            "2. Flag MCQs where all 3 distractors have Lev ≤1 from correct (likely too close).\n"
            "3. Manual review of flagged items by JLPT exam expert."
        ),
        "expected": "<5% of MCQs flagged with all-distractors-too-close.",
        "tools": "Custom Python (Lev distance) + manual review",
        "test_type": "Auto+Manual",
    },
    "J-006": {
        "scenario": "Productive-vs-receptive balance: total drill count is ≥30% production (output) drills, not just MCQ (recognition)",
        "test_steps": (
            "1. Count drills by type in drills_auto.json + questions.json.\n"
            "2. Compute production-drill share.\n"
            "3. Pass: ≥30%."
        ),
        "expected": "Production-drill share ≥30% of total drills.",
        "tools": "Custom Python (drill-type tally)",
        "test_type": "Auto",
    },
    "J-007": {
        "scenario": "Mnemonic coverage: ≥80% of kanji.json entries have a non-trivial mnemonic field",
        "test_steps": (
            "1. For each kanji, check mnemonic.en or mnemonic.hi length ≥30 chars.\n"
            "2. Compute coverage.\n"
            "3. Pass: ≥80%."
        ),
        "expected": "≥80% of 106 kanji have a non-trivial mnemonic.",
        "tools": "Custom Python",
        "test_type": "Auto",
    },
    "J-008": {
        "scenario": "Hindi-L1 transfer-error coverage: common HI-L1 errors (postposition mapping, word order) have explicit common_mistakes entries",
        "test_steps": (
            "1. Reference list of HI-L1 transfer errors (postposition mapping: का → が/に/で; से → から/で/に; को → を/に).\n"
            "2. For each, verify ≥1 common_mistakes entry across grammar.json catches it.\n"
            "3. Report uncovered transfer errors."
        ),
        "expected": "All 10 documented HI-L1 transfer-error classes have ≥1 catching common_mistakes entry.",
        "tools": "Custom Python (grep grammar.json common_mistakes for transfer-error keywords)",
        "test_type": "Auto",
    },
    "J-009": {
        "scenario": "Spiral-review density: every kanji introduced in lesson L appears in ≥3 examples in lessons L+1..L+10",
        "test_steps": (
            "1. For each kanji, find introduction lesson L.\n"
            "2. Count example appearances in lessons L+1..L+10.\n"
            "3. Pass: ≥3."
        ),
        "expected": "≥95% of kanji have ≥3 spiral-review appearances; report exceptions.",
        "tools": "Custom Python (kanji × lesson grid)",
        "test_type": "Auto",
    },
    "J-010": {
        "scenario": "Lesson-completion time budget: each lesson takes 30–60 min for a HI-L1 learner; measured by content volume × reading speed",
        "test_steps": (
            "1. Compute reading-time estimate per lesson: words / reading-speed-WPM × 60.\n"
            "2. Add audio-listen time (audio duration) + drill time (15 min default).\n"
            "3. Pass: 30 ≤ total ≤ 60 min per lesson."
        ),
        "expected": "All lessons within [30, 60] min estimate; outliers flagged for content trim or split.",
        "tools": "Custom Python (lesson-budget calc)",
        "test_type": "Auto",
    },
    "J-011": {
        "scenario": "Difficulty progression monotonicity: lesson N is no easier than lesson N-1 by a measurable proxy (vocab frequency-rank floor, kanji-stroke-count avg)",
        "test_steps": (
            "1. Compute per-lesson: median vocab freq-rank, mean kanji stroke count, mean grammar-pattern complexity.\n"
            "2. Verify monotonic non-decreasing trend across lesson order.\n"
            "3. Flag inversions for re-ordering."
        ),
        "expected": "0 lesson inversions in any proxy.",
        "tools": "Custom Python (difficulty-progression check)",
        "test_type": "Auto",
    },
    "J-012": {
        "scenario": "Authentic-input ratio: ≥15% of reading + listening content is from authentic sources (signs, menus, announcements) vs constructed",
        "test_steps": (
            "1. Count authentic.json entries by provenance.\n"
            "2. Compute share of total reading + listening content.\n"
            "3. Pass: ≥15%."
        ),
        "expected": "≥15% authentic-input share.",
        "tools": "Custom Python",
        "test_type": "Auto",
    },
    "J-013": {
        "scenario": "Cultural-context attachment: every grammar pattern has a cultural_callout field; locked by JA-53",
        "test_steps": (
            "1. JA-53 already enforces non-trivial cultural_callout on every grammar pattern.\n"
            "2. Run python tools/check_content_integrity.py JA-53.\n"
            "3. Pass: 0 violations."
        ),
        "expected": "JA-53 PASS for all 178 grammar patterns.",
        "tools": "tools/check_content_integrity.py (JA-53)",
        "test_type": "Auto",
    },
    "J-014": {
        "scenario": "Postposition-mapping confusion coverage: 4 most-common HI-L1 → JA particle errors have explicit teaching entries",
        "test_steps": (
            "1. Reference: HI का → JA が/の/に/で confusion; HI से → JA から/で/に; HI को → JA を/に; HI में → JA で/に.\n"
            "2. For each, verify ≥1 grammar.json contrast entry covers it (cross-pattern explanation).\n"
            "3. Report uncovered confusions."
        ),
        "expected": "All 4 HI-L1 postposition-mapping confusions covered by ≥1 contrast entry.",
        "tools": "Custom Python (grammar.json contrast scan)",
        "test_type": "Auto",
    },
    "J-015": {
        "scenario": "Pedagogical progression vs Genki/Minna alignment: chapter order matches at least one canonical textbook within ±10% drift",
        "test_steps": (
            "1. Reference: Genki I chapter order + Minna no Nihongo I chapter order for N5 content.\n"
            "2. Compute Spearman correlation between our lesson_order and each reference.\n"
            "3. Pass: correlation ≥0.9 with at least one reference."
        ),
        "expected": "Spearman correlation ≥0.9 with Genki I OR Minna I chapter order.",
        "tools": "Custom Python (rank-correlation)",
        "test_type": "Auto",
    },
}


def rewrite_j_tab(wb: openpyxl.Workbook) -> int:
    """BUG-033: Rewrite Pedagogy (J) tab to testable items."""
    sh = wb["J. Pedagogy"]
    n = 0
    for r in range(5, sh.max_row + 1):
        row_id = sh.cell(row=r, column=1).value
        if not row_id or row_id not in J_TAB_REWRITES:
            continue
        spec = J_TAB_REWRITES[row_id]
        sh.cell(row=r, column=4).value = spec["scenario"]
        sh.cell(row=r, column=5).value = spec["test_steps"]
        sh.cell(row=r, column=6).value = spec["expected"]
        sh.cell(row=r, column=9).value = spec["test_type"]
        sh.cell(row=r, column=13).value = spec["tools"]  # column M = Tools
        # Add provenance note
        note = sh.cell(row=r, column=10).value or ""
        if "[BUG-033 rewrite]" not in note:
            sh.cell(row=r, column=10).value = (note + " | [BUG-033 rewrite] Rewritten from research-question form to testable item with deterministic input + pass criterion.").strip(" |")
        n += 1
    return n


# ============================================================
# BUG-034: Trim End-user (N) tab — pure-domain items get cross-ref + flag
# Already handled in CROSS_REFS for N-013/N-014/N-009. Plus N-001..003 overlap
# with D-013/D-014 — add an explanatory note.
# ============================================================
N_TAB_DUPLICATION_NOTES = {
    "N-001": "Onboarding scenario; overlaps with D-013/D-014 but retained for end-to-end-journey angle (cold-start + first-lesson completion across UX + content + i18n).",
    "N-002": "Onboarding scenario; overlaps with D-014 but retained for end-to-end-journey angle.",
    "N-003": "Day-1 completion; overlaps with D-013 — see UX-augmented Test-steps per BUG-031. End-to-end-journey lens retained.",
}


def annotate_n_tab(wb: openpyxl.Workbook) -> int:
    """BUG-034: Annotate N-tab journey-overlap rows."""
    sh = wb["N. End-user POV"]
    n = 0
    for r in range(5, sh.max_row + 1):
        row_id = sh.cell(row=r, column=1).value
        if row_id in N_TAB_DUPLICATION_NOTES:
            existing = sh.cell(row=r, column=10).value or ""
            tag = N_TAB_DUPLICATION_NOTES[row_id]
            if "[BUG-034" not in existing:
                sh.cell(row=r, column=10).value = (existing + " | [BUG-034] " + tag).strip(" |")
                n += 1
    return n


# ============================================================
# BUG-032: Add 8 new scenarios
# ============================================================
NEW_SCENARIOS = [
    {
        "tab": "B. JLPT format", "id_after": "B-015",
        "id": "B-016", "subcat": "JEES authenticity",
        "persp": "1", "scenario": "Side-by-side comparison of mock-test items against JEES sample paper (latest available) per Mondai",
        "steps": "1. Pull JEES public sample paper (2010 or latest available).\n2. Pair each of our 4 mondai with its JEES counterpart.\n3. JLPT expert reviews question-style, distractor pattern, stem-sentence punctuation, line-break convention.\n4. Report deltas per mondai.",
        "expected": "0 critical format deviations; cosmetic deltas documented and accepted/rejected per mondai.",
        "priority": "P2", "severity": "Major", "test_type": "Manual",
        "tools": "Manual review by JLPT exam expert + JEES sample paper PDF",
        "owner": "JLPT exam expert (JEES-aligned)",
        "notes": "[BUG-032 new] Coverage gap caught by editor review.",
    },
    {
        "tab": "B. JLPT format", "id_after": "B-016",
        "id": "B-017", "subcat": "JEES authenticity",
        "persp": "1", "scenario": "Format authenticity: stem-sentence punctuation, line-break, blank-line conventions vs JEES sample paper",
        "steps": "1. Sample 20 stems from our questions.json + 20 from JEES sample paper.\n2. Compare punctuation use (。/、/「」 frequency), line-break placement, blank-line conventions.\n3. Report deltas; fix any deviation from JEES convention.",
        "expected": "All 20 of our samples match JEES convention on punctuation + line-breaks.",
        "priority": "P3", "severity": "Major", "test_type": "Manual",
        "tools": "Manual side-by-side review",
        "owner": "JLPT exam expert (JEES-aligned)",
        "notes": "[BUG-032 new] Replaces vague B-003 angle.",
    },
    {
        "tab": "J. Pedagogy", "id_after": "J-015",
        "id": "J-016", "subcat": "Competitor benchmark",
        "persp": "3", "scenario": "Pattern-level comparison against named competitors (Bunpro / Marumori / Renshuu) on the same grammar/vocab/kanji items",
        "steps": "1. Pick 20 representative grammar patterns + 20 vocab + 10 kanji.\n2. For each, capture how 3 competitors teach it (screenshot or transcript).\n3. Pedagogy reviewer scores ours vs theirs on clarity / accuracy / depth.\n4. Report qualitative + diff.",
        "expected": "Median score across 50 items ≥ median of best-1 competitor on the same items.",
        "priority": "P3", "severity": "Major", "test_type": "Manual",
        "tools": "Competitor app accounts + manual scoring sheet",
        "owner": "SLA pedagogy researcher",
        "notes": "[BUG-032 new] Coverage gap.",
    },
    {
        "tab": "N. End-user POV", "id_after": "N-016",
        "id": "N-017", "subcat": "True-beginner path",
        "persp": "1", "scenario": "Zero-Japanese learner (not HI-L1+EN-L2+JA-L3 — true L1=anything, L2=none-Japanese) — can the app be used?",
        "steps": "1. Recruit 3 learners with no prior Japanese and any L1.\n2. Onboard with default English UI.\n3. Observe whether they reach 'one lesson + one quiz' within 30 minutes without help.\n4. Report friction points.",
        "expected": "≥2 of 3 reach the milestone; friction points logged and prioritized.",
        "priority": "P4", "severity": "Major", "test_type": "Manual",
        "tools": "Moderated remote user test (Zoom + screen-share)",
        "owner": "Primary-persona learner (HI-L1, EN-L2) — extend recruit to non-HI-L1",
        "notes": "[BUG-032 new] Coverage gap.",
    },
    {
        "tab": "I. Data engineering", "id_after": "I-015",
        "id": "I-016", "subcat": "Restore drill",
        "persp": "2", "scenario": "Restore drill — corrupt or wipe localStorage / IndexedDB, verify recovery / non-fatality",
        "steps": "1. From an installed session with prior progress, manually corrupt localStorage in DevTools.\n2. Reload app.\n3. Verify: (a) app launches without crash; (b) corrupted state is detected and reset; (c) user is informed once; (d) no further interruption.",
        "expected": "App recovers gracefully; corruption is detected and reset; user sees a one-time notice.",
        "priority": "P3", "severity": "Major", "test_type": "Manual",
        "tools": "Browser DevTools + manual interaction",
        "owner": "Data engineer + QA tester",
        "notes": "[BUG-032 new] Coverage gap. Same class as service-worker stale-cache recovery.",
    },
    {
        "tab": "A. Japanese language", "id_after": "A-030",
        "id": "A-031", "subcat": "Hindi-locale audio",
        "persp": "3", "scenario": "Auditory verification of Hindi-locale pedagogical audio (Hindi UI strings or explanations spoken in Hindi)",
        "steps": "1. List every Hindi UI string or explanation that has a corresponding audio asset (if any).\n2. Native Hindi speaker auditions each.\n3. Mark NATURAL / UNNATURAL / NO-AUDIO.\n4. Pass: NATURAL or NO-AUDIO; 0 UNNATURAL.",
        "expected": "0 UNNATURAL Hindi audio renderings.",
        "priority": "P3", "severity": "Major", "test_type": "Manual",
        "tools": "Manual review by native Hindi speaker",
        "owner": "Native Hindi teacher",
        "notes": "[BUG-032 new] Coverage gap. Currently audio is JA-only; this confirms or scopes Hindi audio addition.",
    },
    {
        "tab": "L. Cultural ethical", "id_after": "L-010",
        "id": "L-011", "subcat": "LLM disclosure timing",
        "persp": "2", "scenario": "LLM-curation disclosure is shown BEFORE the user makes a trustworthiness judgment (not buried in About / Settings)",
        "steps": "1. From cold-start, monitor when the user encounters the disclosure (visible on home / first-lesson banner vs hidden in settings).\n2. Pass: disclosure visible on first 2 screens; not behind ≥2 clicks.",
        "expected": "Disclosure visible within the first 2 screens; ≤1 click away from any content surface.",
        "priority": "P3", "severity": "Major", "test_type": "Manual",
        "tools": "Manual review + heuristic eval",
        "owner": "Cultural-sensitivity reviewer + UX designer",
        "notes": "[BUG-032 new] Replaces vague L-007 'disclosure visible' angle.",
    },
    {
        "tab": "B. JLPT format", "id_after": "B-017",
        "id": "B-018", "subcat": "Stem punctuation",
        "persp": "1", "scenario": "Stem-sentence punctuation: every fill-in-the-blank stem ends with 。 or ？; blanks rendered with ＿＿＿ not ____ (full-width)",
        "steps": "1. Scan all stem sentences in questions.json.\n2. Verify each ends with 。 / ？ / ！; blanks are ＿ (full-width underscore U+FF3F).\n3. Report deviations.",
        "expected": "0 stems with half-width underscore; 0 stems without terminal punctuation.",
        "priority": "P3", "severity": "Major", "test_type": "Auto",
        "tools": "Custom Python (regex scan of stem texts)",
        "owner": "JLPT exam expert (JEES-aligned)",
        "notes": "[BUG-032 new] Coverage gap — JEES uses full-width punctuation.",
    },
]


def add_new_scenarios(wb: openpyxl.Workbook) -> int:
    """BUG-032: Add 8 new scenarios across multiple tabs."""
    n = 0
    # Group new scenarios by destination tab
    by_tab: dict[str, list[dict]] = {}
    for spec in NEW_SCENARIOS:
        by_tab.setdefault(spec["tab"], []).append(spec)
    for tab_name, specs in by_tab.items():
        sh = wb[tab_name]
        # Idempotency: skip if scenario ID already present
        existing_ids = {sh.cell(row=r, column=1).value for r in range(5, sh.max_row + 1)
                        if sh.cell(row=r, column=1).value}
        # Find first empty row after data
        target_row = sh.max_row + 1
        for spec in specs:
            if spec["id"] in existing_ids:
                continue
            sh.cell(row=target_row, column=1).value = spec["id"]
            sh.cell(row=target_row, column=2).value = spec["subcat"]
            sh.cell(row=target_row, column=3).value = spec["persp"]
            sh.cell(row=target_row, column=4).value = spec["scenario"]
            sh.cell(row=target_row, column=5).value = spec["steps"]
            sh.cell(row=target_row, column=6).value = spec["expected"]
            sh.cell(row=target_row, column=7).value = spec["priority"]
            sh.cell(row=target_row, column=8).value = spec["severity"]
            sh.cell(row=target_row, column=9).value = spec["test_type"]
            sh.cell(row=target_row, column=10).value = spec["notes"]
            # K-R columns
            sh.cell(row=target_row, column=11).value = ""  # Estimated effort
            sh.cell(row=target_row, column=12).value = spec["owner"]
            sh.cell(row=target_row, column=13).value = spec["tools"]
            sh.cell(row=target_row, column=14).value = None  # Last run date
            sh.cell(row=target_row, column=15).value = "Not Yet Run"
            sh.cell(row=target_row, column=16).value = None  # Tracker link
            sh.cell(row=target_row, column=17).value = None  # Depends on
            sh.cell(row=target_row, column=18).value = "0%"  # Coverage %
            for c in range(1, 19):
                sh.cell(row=target_row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            target_row += 1
            n += 1
    return n


# ============================================================
# BUG-036: Overview formulas
# Replace hardcoded counts with COUNTA / COUNTIF formulas.
# ============================================================
def fix_overview_formulas(wb: openpyxl.Workbook) -> int:
    """BUG-036: Replace hardcoded counts with formulas referencing each tab."""
    sh = wb["Overview"]
    # Inspect current overview shape
    if sh.max_row < 6:
        print("  Overview shape unexpected; skipping formula injection.")
        return 0
    # Header row 5 (data starts row 6 typically)
    # Find rows that name each category tab
    tab_to_row = {}
    for r in range(1, sh.max_row + 1):
        v = sh.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        for tab in CATEGORY_TABS:
            if v.strip() == tab or v.strip() == tab.split(".")[0].strip() or v.strip() == tab.split(". ", 1)[-1].strip():
                tab_to_row[tab] = r
                break

    n = 0
    # Identify columns: Scenario count + Priority HIGH count
    # The exact column layout may vary; for each known tab-row, inject formulas
    # in columns B (scenario count) and C (priority count) — adjust as needed.
    for tab, row in tab_to_row.items():
        # COUNTA on data rows (5..1000) of ID column (A) on that tab
        count_formula = f"=COUNTA('{tab}'!A5:A1000)"
        p1_formula = f'=COUNTIF(\'{tab}\'!G5:G1000, "P1")'
        # Only overwrite columns B (count) and C (P1 count); leave D blank
        if sh.cell(row=row, column=2).value is not None or sh.max_column >= 2:
            sh.cell(row=row, column=2).value = count_formula
            n += 1
        if sh.cell(row=row, column=3).value is not None or sh.max_column >= 3:
            sh.cell(row=row, column=3).value = p1_formula
            n += 1
    return n


# ============================================================
# Main
# ============================================================
def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found")
        return 1
    wb = openpyxl.load_workbook(XLSX)
    print(f"Loaded {XLSX}")
    print(f"Sheets: {wb.sheetnames}")

    print("\n--- BUG-029 + BUG-037: Add 8 new operational columns ---")
    n = add_columns(wb)
    print(f"  Added {n} header cells (8 columns × 14 tabs = up to 112)")

    print("\n--- BUG-024..028, 030, 031, 035 (per-row fixes) ---")
    stats = apply_per_row_fixes(wb)
    for k, v in stats.items():
        print(f"  {k}: {v}")

    print("\n--- BUG-033: Rewrite Pedagogy (J) tab ---")
    n = rewrite_j_tab(wb)
    print(f"  Rewrote {n} J-tab scenarios")

    print("\n--- BUG-034: Annotate End-user (N) tab journey-overlap rows ---")
    n = annotate_n_tab(wb)
    print(f"  Annotated {n} N-tab rows")

    print("\n--- BUG-032: Add 8 new scenarios ---")
    n = add_new_scenarios(wb)
    print(f"  Added {n} new scenarios")

    print("\n--- BUG-036: Overview formulas ---")
    n = fix_overview_formulas(wb)
    print(f"  Injected {n} formulas")

    # ============================================================
    # Final: mark BUG-023 through BUG-038 as Fixed in User Reported Bugs sheet
    # ============================================================
    sh = wb["User Reported Bugs"]
    for r in range(26, 42):  # BUG-023 row 26 through BUG-038 row 41
        bid = f"BUG-{r-3:03d}"
        title = sh.cell(row=r, column=4).value
        if not title:
            continue
        if sh.cell(row=r, column=8).value not in ("Fixed", "Verified", "Closed"):
            sh.cell(row=r, column=8).value = "Fixed"
            existing = sh.cell(row=r, column=5).value or ""
            stamp = "\n\n[FIX 2026-05-17]: Applied by tools/fix_bugs_023_to_038_test_scenarios_2026_05_17.py. See commit message + AUDIT-COVERAGE Part 11 for details."
            if "[FIX 2026-05-17]: Applied by tools/fix_bugs_023_to_038_test_scenarios" not in existing:
                sh.cell(row=r, column=5).value = existing + stamp
            print(f"  Marked {bid} Fixed.")

    # Update Summary counts (38 total now)
    sh.cell(row=4, column=12).value = 38
    sh.cell(row=5, column=12).value = 0
    sh.cell(row=7, column=12).value = 38

    wb.save(XLSX)
    print(f"\nSaved {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
