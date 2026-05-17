"""Add a 'UI Tests' tab to test-scenarios xlsx with the 55 scenarios
covered by tools/ui_test_suite_2026_05_17.py + file NR-UI-001 bug.
"""
from __future__ import annotations

import io
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
RESULTS = ROOT / "tools" / "ui_test_results_2026_05_17.json"
TODAY = datetime(2026, 5, 17)


# Map test_id -> (sub-category, scenario description, spec ref)
UI_SCENARIO_META = {
    "UI-5.1.1": ("Routes / home", "Home page renders N5 card",
                 "spec §5.1 (Home #/home)"),
    "UI-5.1.2": ("Routes / home", "N4 card hidden per BINDING Rule 1",
                 "spec §5.1 + CLAUDE.md Rule 1"),
    "UI-5.3.1": ("Routes / grammar", "Grammar listing page renders",
                 "spec §5.3 (Grammar #/learn/grammar)"),
    "UI-5.3.2": ("Routes / grammar", "Per-pattern detail #/learn/n5-001 renders",
                 "spec §5.3 + n5-001 data"),
    "UI-5.3.3": ("Data integrity", "Grammar pattern count = 178 (matches spec)",
                 "spec §5.3 + §7.1"),
    "UI-5.4.1": ("Routes / vocab", "Vocab listing page renders",
                 "spec §5.4 (Vocab #/learn/vocab)"),
    "UI-5.4.2": ("Data integrity", "Vocab entry count = 995 (matches spec)",
                 "spec §5.4 + §7.1"),
    "UI-5.5.1": ("Routes / kanji", "Kanji listing page renders",
                 "spec §5.5 (Kanji #/kanji)"),
    "UI-5.5.2": ("Data integrity", "Kanji entry count = 106 (matches spec)",
                 "spec §5.5 + §7.1"),
    "UI-5.6.1": ("Routes / reading", "Reading listing page renders",
                 "spec §5.6 (Reading #/reading)"),
    "UI-5.6.2": ("Data integrity", "Reading passage count = 54 (matches spec)",
                 "spec §5.6 + §7.1"),
    "UI-5.7.1": ("Routes / listening", "Listening listing page renders",
                 "spec §5.7 (Listening #/listening)"),
    "UI-5.7.2": ("Data integrity", "Listening item count = 50 (matches spec)",
                 "spec §5.7 + §7.1"),
    "UI-5.8.1": ("Routes / mock-test", "Papers index #/papers renders",
                 "spec §5.8 (Mock Test #/test + Papers #/papers)"),
    "UI-5.8.2": ("Routes / mock-test", "Mock test #/test renders",
                 "spec §5.8"),
    "UI-5.9.1": ("Routes / drill", "Practice / Drill page renders",
                 "spec §5.9 (Practice #/drill)"),
    "UI-5.10.1": ("Routes / review", "Review page renders",
                  "spec §5.10 (Review #/review)"),
    "UI-5.10.a": ("Routes / missed", "Missed-answers page renders",
                  "spec §5.10a (Wrong-answer history #/missed)"),
    "UI-5.11.1": ("Routes / summary", "Progress / summary renders",
                  "spec §5.11 (Progress / Summary #/summary)"),
    "UI-5.12.1": ("Routes / settings", "Settings page renders",
                  "spec §5.12 (Settings #/settings)"),
    "UI-5.13.1": ("Routes / sitting", "Full mock-paper sitting renders",
                  "spec §5.13 (#/sitting)"),
    "UI-5.14.1": ("Routes / daily", "Daily review queue renders",
                  "spec §5.14 (#/today + #/review)"),
    "UI-5.16.1": ("Routes / meta", "In-app PRIVACY viewer renders",
                  "spec §5.16"),
    "UI-5.16.2": ("Routes / meta", "In-app NOTICES viewer renders",
                  "spec §5.16"),
    "UI-MM-1": ("Static mirrors", "/home/ mirror crawlable",
                "spec §5.1 static-mirror + Phase-0 mirror block"),
    "UI-MM-2": ("Static mirrors", "/changelog/ meta-mirror crawlable",
                "Phase-0 meta-mirror block"),
    "UI-MM-3": ("Static mirrors", "/privacy/ meta-mirror crawlable",
                "Phase-0 meta-mirror block"),
    "UI-MM-4": ("Static mirrors", "/notices/ meta-mirror crawlable",
                "Phase-0 meta-mirror block"),
    "UI-MM-5a": ("Static mirrors", "/learn/grammar/n5-001/ canonical mirror crawlable",
                 "F.16 + F.18 + tools/build_static_mirrors.py"),
    "UI-MM-5b": ("Static mirrors", "/lessons/n5-001.html legacy mirror still serves",
                 "BUG-001 + BUG-010 lineage"),
    "UI-MM-6": ("Static mirrors", "/reading/<id>/ mirror crawlable",
                "tools/build_static_mirrors.py"),
    "UI-MM-7": ("Static mirrors", "/listening/<id>/ mirror crawlable",
                "tools/build_static_mirrors.py"),
    "UI-MM-8": ("Static mirrors", "/learn/vocab/<form>/ mirror crawlable (URL-encoded)",
                "tools/build_static_mirrors.py"),
    "UI-MM-9": ("Static mirrors", "/kanji/<glyph>/ mirror crawlable (URL-encoded)",
                "tools/build_static_mirrors.py"),
    "UI-MM-10": ("Static mirrors", "/learn/grammar/ index mirror",
                 "tools/build_static_mirrors.py"),
    "UI-MM-11": ("Static mirrors", "/learn/vocab/ index mirror",
                 "tools/build_static_mirrors.py"),
    "UI-MM-12": ("Static mirrors", "/kanji/ index mirror",
                 "tools/build_static_mirrors.py"),
    "UI-MM-13": ("Static mirrors", "/reading/ index mirror",
                 "tools/build_static_mirrors.py"),
    "UI-MM-14": ("Static mirrors", "/listening/ index mirror",
                 "tools/build_static_mirrors.py"),
    "UI-SEO-1": ("SEO", "sitemap.xml served + valid XML format",
                 "spec §10.3 + JA-113 meta-mirror"),
    "UI-SEO-2": ("SEO", "robots.txt served",
                 "spec §10.3"),
    "UI-A11Y-1": ("Accessibility", "<html lang> attribute present (WCAG 3.1.1)",
                  "spec §14"),
    "UI-A11Y-2": ("Accessibility", "Skip-to-main link present (WCAG 2.4.1 Level A)",
                  "spec §14 + TS-E-006 fix"),
    "UI-A11Y-3": ("Accessibility", "<title> element present + non-empty",
                  "spec §14 + WCAG 2.4.2"),
    "UI-A11Y-4": ("Accessibility", "<main> landmark present (WCAG 1.3.1)",
                  "spec §14"),
    "UI-A11Y-5": ("Accessibility", "ARIA-live region present for screen-reader announcements",
                  "spec §14 + dynamic-content a11y"),
    "UI-SEC-1": ("Security headers", "CSP meta tag comprehensive + frame-ancestors",
                 "spec §15 + NR-SEC-002"),
    "UI-SEC-2": ("Security headers", "X-Frame-Options meta tag present",
                 "spec §15 + NR-SEC-002 (cosmetic — browsers ignore via meta)"),
    "UI-SW-1": ("PWA / Service Worker", "navigator.serviceWorker API available",
                "spec §9"),
    "UI-SW-2": ("PWA / Service Worker", "SW registers on page load",
                "spec §9"),
    "UI-AUD-1": ("Audio", "MP3 audio file reachable via HTTP",
                 "spec §11"),
    "UI-AUD-2": ("Audio", "data/audio_manifest.json present + populated",
                 "spec §11 + audio_manifest schema"),
    "UI-I18N-1": ("Locale / i18n", "Locale-switch UI element present",
                  "spec §5.12 + §12"),
    "UI-I18N-2": ("Locale / i18n", "locales/{en,hi}.json key parity (JA-108)",
                  "spec §12 + JA-108"),
    "UI-LOG-1": ("Console health", "0 SEVERE console errors on home page",
                 "spec §15 / §16 — no runtime errors"),
}


def add_tab_or_extend(wb) -> int:
    if "UI Tests" in wb.sheetnames:
        ws = wb["UI Tests"]
    else:
        ws = wb.create_sheet("UI Tests")
        # Header row 1: tab title
        ws.cell(row=1, column=1).value = "UI Tests (Selenium)"
        ws.cell(row=2, column=1).value = (
            "Perspectives: UI engineer / E2E-test engineer; "
            "tooling: Selenium 4 (auto-driver via Selenium Manager). "
            "Tests served from local HTTP server matching production "
            "deployment URL structure. Auto-runnable via "
            "tools/ui_test_suite_2026_05_17.py."
        )
        # Header row 4
        headers = ["ID", "Sub-category", "Persp #", "Scenario", "Test steps",
                   "Expected result", "Priority", "Severity if fails",
                   "Test type", "Notes", "Estimated effort", "Owner / role",
                   "Tools / scripts required", "Last run date",
                   "Last run result"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=4, column=i).value = h
        # Style
        bold = Font(bold=True)
        for c in ws[4]:
            c.font = bold

    # Load results
    if not RESULTS.exists():
        print("ui_test_results_2026_05_17.json missing; run UI test suite first")
        return 0
    results_doc = json.loads(RESULTS.read_text(encoding="utf-8"))
    res_by_id = {r["id"]: r for r in results_doc["results"]}

    # Find next empty row >= 5
    next_row = 5
    while ws.cell(row=next_row, column=1).value:
        next_row += 1

    n_added = 0
    for test_id, meta in UI_SCENARIO_META.items():
        if not meta:
            continue
        result = res_by_id.get(test_id, {})
        status = result.get("status", "Not Yet Run")
        note = result.get("note", "")
        scen_subcat, scen_desc, spec_ref = meta
        # Priority + Severity heuristics
        priority = "P3"
        severity = "Major"
        if "5.1." in test_id or "MM-" in test_id or "A11Y" in test_id or "SEC" in test_id or "SW" in test_id:
            priority = "P2"
        if "LOG" in test_id or "5.3.3" in test_id or "5.4.2" in test_id:
            priority = "P3"
        ws.cell(row=next_row, column=1).value = test_id
        ws.cell(row=next_row, column=2).value = scen_subcat
        ws.cell(row=next_row, column=4).value = scen_desc
        ws.cell(row=next_row, column=5).value = (
            f"1. python -m http.server 8765 (from JLPTSuccess root).\n"
            f"2. python tools/ui_test_suite_2026_05_17.py.\n"
            f"3. Inspect line {test_id} of the output."
        )
        ws.cell(row=next_row, column=6).value = f"Status PASS per {spec_ref}"
        ws.cell(row=next_row, column=7).value = priority
        ws.cell(row=next_row, column=8).value = severity
        ws.cell(row=next_row, column=9).value = "Automated (Selenium 4)"
        ws.cell(row=next_row, column=10).value = note[:200] if note else ""
        ws.cell(row=next_row, column=11).value = "5 min/scenario"
        ws.cell(row=next_row, column=12).value = "UI engineer / E2E-test engineer"
        ws.cell(row=next_row, column=13).value = (
            f"tools/ui_test_suite_2026_05_17.py (Selenium 4 + headless Chrome)"
        )
        ws.cell(row=next_row, column=14).value = TODAY
        # Map "PASS" / "FAIL" / "SKIP" to our scenario-result conventions
        if status == "PASS":
            ws.cell(row=next_row, column=15).value = "PASS"
        elif status == "FAIL":
            ws.cell(row=next_row, column=15).value = "FAIL"
        elif status.startswith("FAIL"):
            ws.cell(row=next_row, column=15).value = status
        elif status == "SKIP":
            ws.cell(row=next_row, column=15).value = "Skipped — agent-deferred"
        else:
            ws.cell(row=next_row, column=15).value = "Not Yet Run"
        next_row += 1
        n_added += 1

    # Auto-size columns
    for col_idx in range(1, 16):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    return n_added


def file_bug() -> int:
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["User Reported Bugs"]
    last_row = ws.max_row
    while last_row >= 4 and not ws.cell(row=last_row, column=4).value:
        last_row -= 1
    next_row = last_row + 1
    today = datetime(2026, 5, 17)
    ws.cell(row=next_row, column=2).value = today
    ws.cell(row=next_row, column=3).value = (
        "UI engineer Selenium review (2026-05-17)"
    )
    ws.cell(row=next_row, column=4).value = (
        "NR-UI-001 — Defense-in-depth meta tags (frame-ancestors in CSP, "
        "X-Frame-Options) ignored by modern browsers — clickjacking "
        "protection limited on static hosting"
    )
    ws.cell(row=next_row, column=5).value = (
        "Source: UI test suite run 2026-05-17 (Selenium 4 + Chrome 148) "
        "caught 2 SEVERE console errors on every route load:\n\n"
        "  - 'The Content Security Policy directive `frame-ancestors` "
        "    is ignored when delivered via a <meta> element.'\n"
        "  - 'X-Frame-Options may only be set via an HTTP header sent "
        "    along with a document. It may not be set inside <meta>.'\n\n"
        "Both directives were added to N5/index.html in commit 46be3e1 "
        "(NR-SEC-002 fix). Per HTML spec / Fetch spec, both `frame-"
        "ancestors` and `X-Frame-Options` are HTTP-header-only — "
        "browsers explicitly ignore them when delivered via meta tags. "
        "So the previous NR-SEC-002 fix was cosmetic — it added the "
        "directives to the source but they had no runtime effect.\n\n"
        "Static-hosting constraint: GitHub Pages does not expose HTTP-"
        "header configuration. The only clickjacking-defense paths are:\n"
        "  (a) Move to a static host that exposes header config "
        "      (Cloudflare Pages, Netlify, Vercel, etc.)\n"
        "  (b) Use a Cloudflare Worker / similar header-injecting "
        "      proxy in front of GH Pages\n"
        "  (c) Accept the limitation; rely on the remaining honored "
        "      directives (default-src 'self', script-src 'self', "
        "      connect-src 'self', form-action 'none' — these ARE "
        "      honored via meta and provide substantial defense-in-depth)\n\n"
        "Decision: option (c) — accept the limitation; remove the "
        "ineffective meta tags so they don't mislead future maintainers "
        "into thinking clickjacking protection is in place when it "
        "isn't. The other honored directives provide the bulk of the "
        "defense-in-depth.\n\n"
        "[FIX 2026-05-17]:\n"
        "  - Removed `frame-ancestors 'none';` from the CSP meta tag.\n"
        "  - Removed the X-Frame-Options meta tag.\n"
        "  - Added a comment explaining why and pointing at the "
        "    static-hosting limitation.\n"
        "  - Permissions-Policy (camera/mic/geo/etc.) + Referrer-Policy "
        "    remain — both ARE honored via meta tags."
    )
    ws.cell(row=next_row, column=6).value = "Medium"
    ws.cell(row=next_row, column=7).value = "P3"
    ws.cell(row=next_row, column=8).value = "Fixed"
    ws.cell(row=next_row, column=9).value = "(pending — this commit)"
    ws.cell(row=next_row, column=10).value = today
    wb.save(str(XLSX))
    return 1


def main() -> int:
    wb = openpyxl.load_workbook(str(XLSX))
    print("=== Adding 'UI Tests' tab ===")
    n_added = add_tab_or_extend(wb)
    wb.save(str(XLSX))
    print(f"Added {n_added} UI test scenario rows")

    print("\n=== Filing NR-UI-001 ===")
    file_bug()

    return 0


if __name__ == "__main__":
    sys.exit(main())
