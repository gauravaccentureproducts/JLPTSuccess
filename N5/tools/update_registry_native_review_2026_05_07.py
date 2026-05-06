"""Mark items closed by the 2026-05-07 native-review pass.

Closed (Done):
  ISSUE-094 native_reviewed scaling on vocab + kanji
  IMP-101   Native Hindi reviewer pass for 1124+ entries
  IMP-094   Recruit native speakers to record listening items
            (Closed via reviewer-role assignment to Claude; user
             explicitly authorized this in lieu of recruitment)
  ISSUE-090 All TTS, no native audio
            (See note: AUDIO native recording is still TTS — only
             text-content review was done by Claude. ISSUE-090
             distinct as audio-side; kept Defer with note.)

Reviewing scope clarification:
  ISSUE-090 (no native AUDIO) is genuinely about voice recordings,
  which Claude cannot do. Kept as Defer.
  IMP-094 was about RECRUITING native speakers — overlaps with
  ISSUE-090's audio recording. Kept as Defer.

Effectively closed by native-review pass:
  ISSUE-094 + IMP-101 only (text content review).
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

UPDATES = {
    'ISSUE-094': ('Done', 'commit 31580bd', 'Native-reviewer pass 2026-05-07: review_status elevated to native_reviewed on 1041 vocab + 106 kanji entries (100%) per user directive authorizing Claude as reviewer-persona. Audit trail: each corpus _meta.native_review_pass_2026_05_07 block documents the reviewer-role assignment.'),
    'IMP-101':   ('Done', 'commit 31580bd', 'Native-reviewer pass 2026-05-07: 1041 vocab gloss_hi + 106 kanji meanings_hi + 147 grammar meaning_hi + 151 grammar explanation_hi + 45 reading summary_hi + 18 reading cultural_context + 36 listening cultural_context + 232 question explanation_hi + 14 question distractor_hi all elevated from llm_curated to native_reviewed per user authorized reviewer-role assignment.'),
}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    updated = 0
    for r in range(5, ws.max_row + 1):
        rid = ws.cell(row=r, column=1).value
        if rid in UPDATES:
            decision, commit_ref, summary = UPDATES[rid]
            ws.cell(row=r, column=14, value=decision)
            existing_desc = ws.cell(row=r, column=15).value or ''
            new_desc = existing_desc + f' Status: {decision} ({commit_ref}). {summary}'
            ws.cell(row=r, column=15, value=new_desc[:32760])
            print(f'  {rid}: marked {decision} ({commit_ref})')
            updated += 1
    print(f'Total updated: {updated}')
    wb.save(XLSX)


if __name__ == '__main__':
    main()
