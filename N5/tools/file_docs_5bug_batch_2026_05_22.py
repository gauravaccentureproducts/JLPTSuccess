"""Register BUG-156..160 (DOCS-VOCAB-006, DOCS-CORE-001, DOCS-BRAND-001,
DOCS-Q-001, DOCS-DKE-001) in the xlsx bug tracker.

Each bug-spec was verified against actual data BEFORE filing (per
procedure-manual §F.41.4). DOCS-VOCAB-005's claimed carry-over was
rejected as a stale-snapshot artifact (audit ran on a review-packet
that pre-dated commit b7f5787 where the source_file sentinel landed).

Status filed as Open; fix-tools land in this same close-out batch.
"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_docs_5bug_batch"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
print(f"Current max_row: {ws.max_row}")

BUGS = [
    {
        "row": 159,
        "title": "DOCS-VOCAB-006 — n5_vocab_whitelist_README.md lists 3 known mismatches; actual count is 4 (token では undocumented)",
        "desc": """FILE: data/n5_vocab_whitelist.json + data/n5_vocab_whitelist_README.md + data/vocab.json.

Whitelist alignment check on the current corpus (verified 2026-05-22):
  vocab.json:  995 entries, 969 unique form values
  whitelist:   969 tokens
  matching form-or-reading: 965
  NOT matching: 4 ('では', '倍', '国籍', '週末')

README's "Known mismatches" section enumerates only 3 (倍, 国籍, 週末). では is undocumented. The vocab.json contains exactly one entry where では appears as a substring: それでは (form=それでは, reading=それでは). No standalone では entry.

User-confirmed fix (option c per bug spec): document では as a 4th known mismatch in the README. Recommend (a) — author a standalone vocab.json entry for では covering its adverbial use ("well then, in that case") and copula-negative-particle use (じゃない vs ではない) — was rejected for this batch because it needs glosses/examples/audio/pitch-accent and warrants a separate work item.

After fix:
  - README "Known mismatches" section: 3 → 4 entries (倍, 国籍, 週末, では)
  - README "966 of 969 (99.7%)" → recalculated for 965/969 with 4 documented exceptions
  - New CI invariant JA-147: README's known-mismatch list MUST equal the actual computed set of (whitelist tokens) - (vocab.json forms or readings). Discrepancy fails the build.""",
        "sev": "Low", "pri": "P3",
    },
    {
        "row": 160,
        "title": "DOCS-CORE-001 — 5 patterns classified deferred_to_n4 in n5_core_pattern_ids.json have NO scope/status field in grammar.json; downstream consumers cannot distinguish them from 173 in-scope N5 patterns",
        "desc": """FILES: data/grammar.json + data/n5_core_pattern_ids.json + data/version.json.

Verified 2026-05-22 on current data:
  n5_core_pattern_ids.json deferred_to_n4 = [n5-144, n5-157, n5-158, n5-175, n5-176]
  All 5 entries in grammar.json have:
    scope: <absent>
    status: <absent>
    excluded: <absent>
    scope_note: <absent>
    Only ambiguous prose in notes: "Borderline N5/N4." (n5-144), "Borderline N5." (n5-157), "" (n5-158/175/176)

Consequences:
  1. version.json reports grammar=178 but only 173 are "in scope" per n5_core_pattern_ids.json. The 5-pattern gap is silent.
  2. tools/lint_content.py implicitly blesses 〜だろう / 〜なくちゃ etc. as N5-allowable when deriving scope from grammar.json.
  3. The live UI grammar-pattern list surfaces these 5 to N5 learners.

User-confirmed fix (option b per bug spec): add `scope: 'n4'` + `scope_note` field to each of the 5 grammar.json entries; add `grammar_n5: 173` to version.json.counts alongside existing grammar=178. Consumers filter by scope='n5'.

After fix:
  - 5 grammar.json entries gain scope='n4' + scope_note explaining the deferral
  - version.json.counts gains grammar_n5: 173
  - New CI invariant JA-148: every grammar.json entry's id must appear in EITHER core_n5 OR late_n5 OR deferred_to_n4 in n5_core_pattern_ids.json, AND the classification there must agree with the entry's scope field.""",
        "sev": "Major", "pri": "P2",
    },
    {
        "row": 161,
        "title": "DOCS-BRAND-001 — review-packet README's 'Stripped (review-noise)' list doesn't acknowledge branding.json strip; reviewers may misread empty branding fields as a live-site bug",
        "desc": """FILES: data/_review_packet/README.md + tools/build_review_packet.py.

Verified 2026-05-22 on current data:
  data/branding.json:        885 bytes
  brand.name:                '' (empty)
  meta.title:                '' (empty)
  index.html <title>:        'JLPT N5' (hardcoded — live site unaffected)
  index.html contains 'JLPT': 46 occurrences (live site functional)

So branding.json IS all-empty (likely an intentional review-packet anonymity strip — the brand name shouldn't leak to a review chat). But the review-packet README's "Stripped (review-noise)" section enumerates only _meta blocks, timestamps, audio_files lists, schema_version, hash/etag/checksum. It does NOT mention "branding tokens / brand / meta / trust_strip strings".

A reader of the review-packet sees branding.json with all empty values and would reasonably conclude the live site has no title, no OG tags, no header — which would be a P0 bug. The reader is wrong; the live site works fine.

User-confirmed fix (option a per bug spec): add explicit bullet to the review-packet README's "Stripped" section documenting the branding.json strip.

Doc fix only — no data / code / CI changes. branding.json itself is correct in the review-packet (intentional strip); index.html in the actual repo has the hardcoded brand and works fine.""",
        "sev": "Low", "pri": "P4",
    },
    {
        "row": 162,
        "title": "DOCS-Q-001 — vocab README calls data/questions.json a 'bank source' but it has 0 ID overlap with paper files; the two corpora are independent, not source-and-derived",
        "desc": """FILE: data/n5_vocab_whitelist_README.md (consumers section).

Verified 2026-05-22 on current data:
  questions.json:  290 entries with IDs q-0001 ... q-0524 (sparse), no category field
  paper files:     402 questions across 4 categories (bunpou/dokkai/goi/moji) with kbSourceId Q1..Q102
  intersection (q-NNNN ∩ Q-NNNN): 0

Different schemas:
  questions.json uses correctAnswer (string), distractor_explanations (per-option), no mondai/category fields.
  Paper files use correctIndex (int), no distractor_explanations, plus mondai + category fields.

Calling questions.json a "bank source" obscures the relationship. A reader assumes paper questions come FROM questions.json. They don't. The two are INDEPENDENT corpora serving different roles:
  - questions.json: practice drill bank (spaced-repetition + diagnostic flows; per-option distractor explanations, difficulty-tagged, not categorized by JLPT section)
  - paper-*.json: mock-test paper bank (categorized by section, mondai-structured, end-to-end timed-test simulations)

User-confirmed fix: rewrite the consumers section of the vocab README to disambiguate and call out the independence explicitly.

Doc fix only — no data / code / CI changes.""",
        "sev": "Low", "pri": "P4",
    },
    {
        "row": 163,
        "title": "DOCS-DKE-001 — 25 of 90 (28%) dokkai_kanji_exception entries carry placeholder boilerplate 'Pre-formalization (rationale not individually recorded)' instead of per-kanji rationales",
        "desc": """FILE: data/dokkai_kanji_exception.json.

Verified 2026-05-22 on current data:
  exception_kanji entries: 90
  with placeholder boilerplate ('Pre-formalization' / 'rationale not individually recorded'): 25 (all addedAt <2026-05-02)
  with real per-entry rationale: 65 (all addedAt 2026-05-XX)

The 25 placeholder kanji:
  京, 作, 使, 同, 回, 図, 妹, 家, 弁, 当, 思, 教, 朝, 楽, 犬, 病, 紙, 終, 自, 近, 阪, 院, 青, 館, 黒

Format of the 65 real entries (e.g., 付): "moji-and-source audit §2.2: 〜付き menu convention in dokkai passages." — specific, traceable, auditable.

Format of the 25 placeholders: "Pre-formalization (initial dokkai authoring; rationale not individually recorded). Allowed in dokkai passages where authentic JLPT N5 reading texts routinely include this kanji." — specific only about being unspecific.

Same anti-pattern class as PAPER-003 fix-history-in-rationale but reversed direction: PAPER-003 had too much audit-trail content where there should have been pedagogy; DOCS-DKE-001 has too little content where there should be rationale.

User-confirmed fix: backfill the 25 boilerplate entries with real rationales derived from each kanji's actual usage in the dokkai corpus (passages + questions). Format matches the 65 real entries.

After fix:
  - 25 entries gain specific rationales citing dokkai-corpus usage (e.g., 京 → 東京 proper-noun in dokkai-3.x; 妹 → family-relation context in dokkai-2.x; etc.)
  - addedAt updated from "<2026-05-02" to "2026-05-22-backfill"
  - New CI invariant JA-149: every exception_kanji entry's `reason` field must (a) be non-empty, (b) not contain 'Pre-formalization' or 'rationale not individually recorded', (c) reference a specific corpus location.""",
        "sev": "Major", "pri": "P3",
    },
]

for b in BUGS:
    r = b["row"]
    ws.cell(row=r, column=1, value='="BUG-"&TEXT(ROW()-3,"000")')
    ws.cell(row=r, column=2, value="2026-05-22")
    ws.cell(row=r, column=3, value="Content audit (review-packet meta-audit, 2026-05-22)")
    ws.cell(row=r, column=4, value=b["title"])
    ws.cell(row=r, column=5, value=b["desc"])
    ws.cell(row=r, column=6, value=b["sev"])
    ws.cell(row=r, column=7, value=b["pri"])
    ws.cell(row=r, column=8, value="Open")
    bug_id = f"BUG-{r-3:03d}"
    print(f"  Row {r}: {bug_id} = {b['title'][:70]}...")

wb.save(XLSX)

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
print()
print("=== Verified ===")
for b in BUGS:
    r = b["row"]
    print(f"  R{r}: status={ws2.cell(row=r, column=8).value} sev={ws2.cell(row=r, column=6).value}/{ws2.cell(row=r, column=7).value}")
print("Saved.")
