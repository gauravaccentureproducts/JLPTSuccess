"""Flip BUG-161 (NTR-001) to Fixed."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_flip_ntr_001"
if not os.path.exists(BAK): shutil.copy2(XLSX, BAK)
wb = load_workbook(XLSX); ws = wb["User Reported Bugs"]
ws.cell(row=164, column=8, value="Fixed")
ws.cell(row=164, column=9, value="2026-05-22")
ws.cell(row=164, column=10, value="<pending-main-commit>")
ws.cell(row=164, column=11, value=(
    "Rewrote 99 violating vocab examples to use only whitelist + exception kanji "
    "(mechanical kana-substitution of side-words; headwords preserved). "
    "Provenance bumped to native_reviewed_2026_05_22. New CI invariant JA-150 "
    "locks the predicate: every kanji in any vocab example's ja text must be "
    "in n5_kanji_whitelist OR dokkai_kanji_exception. Brings vocab examples "
    "up to parity with grammar / reading / listening (all already gated). "
    "Tool: tools/fix_ntr_001_vocab_examples_2026_05_22.py."
))
wb.save(XLSX)
print("BUG-161 (NTR-001) -> Fixed")
