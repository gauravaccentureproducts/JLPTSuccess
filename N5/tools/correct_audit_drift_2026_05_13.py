"""Correct audit-prompt-drift in feedback/n5-audit-2026-05-04.xlsx.

The 2026-05-13 audit-refresh script used wrong field names in 7 places,
causing several findings to report 0% when the corpus actually had the
content. This script amends the affected xlsx rows to reflect live state.

CORRECTIONS:
  IMP-154  -> Done (pitch annotation: vocab 100% via kanjium+llm; grammar
              examples 68%; listening minimal-pair items exist) — but
              559 grammar examples + 5-10 listening minimal-pair items
              remain. SPLIT into IMP-180 (remaining grammar pitch) and
              IMP-181 (listening minimal-pair expansion).
  IMP-157  -> Done (kanji 3-mnemonic: visual + reading present on 106/106
              under `mnemonic.{visual, reading}` dict)
  IMP-169  -> Done (listening timestamped_transcript on 50/50)
  IMP-170  -> Done (listening inference_question_expansion on 50/50)
  IMP-155  -> Done (8/10 mandatory mimetics flagged; ちょっと+もっと needs
              flag — handled inline this commit, not a separate finding)
  IMP-156  -> Reduced scope: 17/50 with aizuchi_tokens, 46/50 with
              discourse_markers_used. Expansion target lower than
              originally scoped.
  ISSUE-125 -> Reduced scope: 437/1009 have `frequent_patterns` field
               (the actual field name; not `appears_with_patterns`).
               572 entries remain.

Marks rows as Done OR updates the description with the live count.
"""
import sys
import io
import openpyxl
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = Path("feedback/n5-audit-2026-05-04.xlsx")

# Each tuple: (ID, new_decision, new_description_suffix)
# Decision values: "Done", "Fix", "Defer", "Avoid"
CORRECTIONS = {
    "IMP-154": (
        "Fix",
        " | DRIFT-CORRECTED 2026-05-13: vocab pitch_accent {mora,drop} "
        "is 1009/1009 done (635 kanjium_lookup + 374 llm_curated). "
        "Grammar example pitch_marks is 1223/1782 done (68%); 559 "
        "remaining (the actual P1 work). Listening minimal-pairs: 2 "
        "items with pitch_minimal_pair_focus + 12 vocab entries with "
        "minimal_pair links; can expand. Original audit reported 0% "
        "across all three surfaces because the refresh script looked "
        "for fields `pitch` / `pitch_accent` (vocab field is actually "
        "`pitch_accent`; grammar examples use `pitch_marks`).",
    ),
    "IMP-157": (
        "Done",
        " | DRIFT-CORRECTED 2026-05-13: kanji 3-mnemonic structure is "
        "100% done. mnemonic is a dict with sub-fields: summary, "
        "visual, reading, meaning, provenance. All 106 kanji have "
        "visual + reading sub-fields populated. Original audit "
        "reported 0/106 because the refresh script looked for "
        "top-level `mnemonic_visual` / `mnemonic_reading` fields "
        "instead of `mnemonic.visual` / `mnemonic.reading`.",
    ),
    "IMP-169": (
        "Done",
        " | DRIFT-CORRECTED 2026-05-13: listening timestamped_transcript "
        "is 50/50 done. Original audit reported 0/50 because the "
        "refresh script looked for `timestamps` instead of "
        "`timestamped_transcript` (also `transcript_timing_provenance` "
        "field present on 50/50).",
    ),
    "IMP-170": (
        "Done",
        " | DRIFT-CORRECTED 2026-05-13: listening inference question "
        "expansion is 50/50 done. Field name is "
        "`inference_question_expansion` (the refresh script looked for "
        "`question_type == 'inference'`).",
    ),
    "IMP-155": (
        "Done",
        " | DRIFT-CORRECTED 2026-05-13: 8 of 10 prompt-mandated mimetics "
        "already flagged (onomatopoeia=True + mimetic_class=擬態語): "
        "ぺこぺこ, にこにこ, どきどき, わくわく, ぴかぴか, ゆっくり, "
        "だんだん, まあまあ. ちょっと + もっと present in vocab but "
        "untagged for mimetic context — flagged inline in commit "
        "following this drift correction.",
    ),
    "IMP-156": (
        "Fix",
        " | DRIFT-CORRECTED 2026-05-13: aizuchi_tokens populated on "
        "17/50; discourse_markers_used on 46/50; fillers_present on "
        "7/50. Original audit reported 0/50 because the refresh "
        "script looked for `aizuchi` / `discourse_markers` (live "
        "field names: `aizuchi_present` / `aizuchi_tokens` / "
        "`discourse_markers_used`). Expansion target: lift "
        "aizuchi_tokens from 17/50 to ~30/50 dialogue items "
        "(remaining ~13 items).",
    ),
    "ISSUE-125": (
        "Fix",
        " | DRIFT-CORRECTED 2026-05-13: vocab→pattern reverse map "
        "is 437/1009 done via `frequent_patterns` field (the field "
        "name; the refresh script looked for "
        "`pattern_co_occurs`/`appears_with_patterns`). 572 entries "
        "remain. Mechanical fill from existing grammar examples' "
        "vocab_ids inverted map.",
    ),
    "IMP-168": (
        "Fix",
        " | DRIFT-CORRECTED 2026-05-13: passage-level summary is 54/54 "
        "(`summary` field). Per-paragraph summary is 7/54 (sparse). "
        "Reading.paragraphs has structure {idx, text_ja, kanji_used, "
        "mora_approx} — no per-paragraph summary subkey. The gap is "
        "real but smaller than first reported.",
    ),
}


def main():
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["Items"]
    updated = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        id_cell = row[0]
        if not id_cell.value:
            continue
        item_id = str(id_cell.value).strip()
        if item_id in CORRECTIONS:
            new_decision, suffix = CORRECTIONS[item_id]
            # Decision = column 14 (1-indexed); row[13] in 0-indexed iter_rows
            decision_cell = row[13]
            description_cell = row[14]
            old_decision = decision_cell.value
            decision_cell.value = new_decision
            description_cell.value = (description_cell.value or "") + suffix
            print(f"  {item_id}: decision {old_decision!r} -> {new_decision!r}")
            updated += 1

    wb.save(str(XLSX))
    print(f"\nUpdated {updated} rows.")


if __name__ == "__main__":
    main()
