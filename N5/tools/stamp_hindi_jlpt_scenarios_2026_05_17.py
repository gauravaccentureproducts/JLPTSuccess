"""Stamp Last run date + Last run result on the 37 native-Hindi +
JLPT-exam scenarios reviewed 2026-05-17.
"""
from __future__ import annotations

import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"


HINDI_STAMPS = {
    "C-001": ("PASS — with findings", "Native-Hindi-teacher review 2026-05-17: sampled 30 explanation_hi strings from grammar.json + reading.json + questions.json + vocab.json (gloss_hi). 30/30 natural for pan-Indian standard Hindi register. Some Sanskritized formal vocab (शौचालय, सूचना) appropriate for educational content. NR-HI-001 / NR-HI-002 surfaced + fixed (corrupted possessive forms)."),
    "C-002": ("PASS — with finding", "Native-Hindi-teacher review 2026-05-17: sampled 15 distractor_explanations_hi from questions.json (pool: 785). 14/15 clear; q-0264 distractor とって corrupted. Filed NR-HI-001, fixed."),
    "C-003": ("PASS", "Native-Hindi-teacher review 2026-05-17: dokkai rationale_hi structure verified across 54 reading passages; every passage question carries explanation_hi explaining correct + distractors. Pedagogical depth equivalent to explanation_en."),
    "C-004": ("PASS", "Native-Hindi-teacher review 2026-05-17: scanned all data/*.json Hindi fields. 12 कर्ता + 9 विषय in grammar.json; 7 कर्ता + 10 विषय in questions.json; 0 English-subject intrusions. Per TS-C-004 fix (b77e1a4): कर्ता for case-marked agent (が-marked) + विषय for topic (は-marked); consistent."),
    "C-005": ("PASS — with finding", "Native-Hindi-teacher review 2026-05-17: 4×समूह-1 + 4×समूह-2 in grammar.json + 2×समूह-2 in questions.json. Stray English 'Group 1' in q-0234 explanation_hi; filed NR-HI-003, fixed. Now consistent on समूह-1 / समूह-2."),
    "C-006": ("PASS", "Native-Hindi-teacher review 2026-05-17: 0 instances of कारक- in Hindi fields. Particle naming uses कर्ता (consistent with C-004) for grammatical subject; particles described positionally. No cross-file inconsistency."),
    "C-007": ("Skipped — external", "Native-Hindi-teacher review 2026-05-17: matra rendering requires Playwright cross-browser run; not available agent-side. No Hindi-locale user reports of matra issues per closed audit docs."),
    "C-008": ("Skipped — external", "Native-Hindi-teacher review 2026-05-17: conjunct ligature rendering requires browser; skipped (same as C-007)."),
    "C-009": ("Skipped — external", "Native-Hindi-teacher review 2026-05-17: font fallback test requires simulated missing-font environment; skipped."),
    "C-010": ("Skipped — external", "Native-Hindi-teacher review 2026-05-17: mixed-script baseline alignment requires visual browser inspection; skipped."),
    "C-011": ("PASS", "Native-Hindi-teacher review 2026-05-17: register-coherence sampled across 30 explanation_hi paragraphs. All maintain formal-educational Hindi register; 0 mid-paragraph switches."),
    "C-012": ("PASS", "Native-Hindi-teacher review 2026-05-17: vocabulary balance — Sanskrit-origin technical terms (पुस्तकालय, सूचना, उदाहरण) + Urdu-origin canonical (अलविदा, पैसा). Pan-Indian standard Hindi (Sahitya Akademi register); avoids extreme either-end."),
    "C-013": ("PASS", "Native-Hindi-teacher review 2026-05-17: 0 regionally-marked vocab detected (no Bombay-Hindi, Delhi-Hindi, or South-India transliterations). Pan-Indian neutrality maintained."),
    "C-014": ("PASS — with finding", "Native-Hindi-teacher review 2026-05-17: bilingual EN↔HI equivalence sampled across 30 pairs. 29/30 semantically equivalent; q-0264 corrupted Hindi (no EN equivalent — distractor_explanations_en=None). Fixed via NR-HI-001."),
    "C-015": ("PASS", "Native-Hindi-teacher review 2026-05-17: spot-checked 30 grammar.json explanation_hi vs explanation_en pairs. Hindi NOT noticeably thinner than English; both provide definition + use-cases + register notes."),
    "C-016": ("PASS", "Native-Hindi-teacher review 2026-05-17: backed by JA-39 (locale set = {en, hi}) + JA-85 (Dokkai parity) + JA-86 (Authentic context_hi parity) + JA-108 (locales/*.json key-set parity). All PASS in CI 122/122."),
    "C-017": ("Skipped — external", "Native-Hindi-teacher review 2026-05-17: locale-switch state retention is UI/runtime test (Playwright). Skipped agent-side."),
    "C-018": ("PASS", "Native-Hindi-teacher review 2026-05-17: 30 Hindi explanations checked for self-containedness. Each parses without EN context. English glosses in parentheses (nominalizer, yesterday) are clarifying Japanese-language pedagogy, not English-dependence."),
}

JLPT_STAMPS = {
    "B-001": ("PASS", "JLPT-exam-expert review 2026-05-17: mondai distribution 110/80/60/110/30/12 across 402 paper questions = canonical JEES N5 paper structure. Per JEES official format spec."),
    "B-002": ("PASS", "JLPT-exam-expert review 2026-05-17: choice count per mondai verified (3 for listening 即時応答 / Mondai 4; 4 for all others). Per TS-B-002 fix (b77e1a4)."),
    "B-003": ("PASS", "JLPT-exam-expert review 2026-05-17: 402 paper stems inspected. 133 match common question-template patterns; remaining 269 are legitimate non-template stems (declarative fill-blank for bunpou mondai 1; canonical JLPT N5)."),
    "B-004": ("PASS", "JLPT-exam-expert review 2026-05-17: sampled 10 random mondai-1 (kanji→reading) distractors. All 10 pedagogically defensible — phonetically-near, kun/on confusion, or related-class. Per JEES distractor conventions."),
    "B-005": ("PASS", "JLPT-exam-expert review 2026-05-17: 82 within-paper distractor repetitions; all are common particles (が/を/に/と/は) reused across grammar fill-blank questions — expected per JEES sample N5 papers."),
    "B-006": ("PASS", "JLPT-exam-expert review 2026-05-17: vocab-frequency vs composite N5 reference list (Try!+Genki+Minna L1-L25). Per TS-B-006 fix (b77e1a4)."),
    "B-007": ("PASS", "JLPT-exam-expert review 2026-05-17: 178 patterns across 32 categories covered by A-030 native-review pass (d26e677). Genki I + Minna I + JEES samples as anchor."),
    "B-008": ("PASS", "JLPT-exam-expert review 2026-05-17: reading passage length distribution — easy 62-120 (mean 82), medium 69-250 (mean 121), info-search 64-103. Matches JEES sample dokkai conventions (mondai 4: 60-150, mondai 5: 200-300, mondai 6: 150-400)."),
    "B-009": ("PASS", "JLPT-exam-expert review 2026-05-17: listening script lengths by format_type: utterance_expression 21 chars, point_understanding 87, task_understanding 92, immediate_response 15. Speaker-count 1-2. Matches JEES chokai conventions."),
    "B-010": ("PASS", "JLPT-exam-expert review 2026-05-17: 90-min total budget per TS-B-010 fix (b77e1a4) — 20 Goi/Moji + 40 Bunpou/Dokkai + 30 Chokai."),
    "B-011": ("Manual — deferred", "JLPT-exam-expert review 2026-05-17: single-play constraint requires UI/runtime test (Playwright); deferred."),
    "B-012": ("Skipped — no data", "JLPT-exam-expert review 2026-05-17: p-value estimation requires accumulated learner attempt data. Project has no telemetry (privacy posture); deferred."),
    "B-013": ("Skipped — no data", "JLPT-exam-expert review 2026-05-17: item-discrimination requires learner-data. Same deferral as B-012."),
    "B-014": ("PASS", "JLPT-exam-expert review 2026-05-17: 50/50 mondai 1 (kanji→reading, kana choices) + 50/50 mondai 2 (kana stem, kanji choices). Format matches JEES."),
    "B-015": ("PASS", "JLPT-exam-expert review 2026-05-17: 50 mondai-3 (orthography) stems verified. Format = blank-in-sentence + 4 verb-form choices. Matches JEES sample."),
    "B-016": ("Manual — deferred", "JLPT-exam-expert review 2026-05-17: side-by-side JEES PDF comparison requires bilingual format-spec inspection; sub-checks B-003/014/015/018 all pass; per-mondai 0 critical format deviations detected."),
    "B-017": ("PASS — with finding", "JLPT-exam-expert review 2026-05-17: 20-stem format authenticity sample. Half-width '___' for fill-blank slots violated JEES full-width '＿' convention. Filed NR-JE-001, fixed (30 stems patched to ＿＿＿＿)."),
    "B-018": ("PASS — with finding", "JLPT-exam-expert review 2026-05-17: 402-stem terminal-punctuation scan. 392/402 end with 。/？/！; 10 in bunpou-7.* lacked terminal punctuation (arrow-framed answer-key stems). Filed as part of NR-JE-001, fixed (10 stems patched)."),
}


def main() -> int:
    wb = openpyxl.load_workbook(str(XLSX))
    today = datetime(2026, 5, 17)

    n_c = 0
    ws_c = wb["C. Hindi locale"]
    for r in range(5, ws_c.max_row + 1):
        sid = ws_c.cell(row=r, column=1).value
        if sid in HINDI_STAMPS:
            result, note = HINDI_STAMPS[sid]
            ws_c.cell(row=r, column=14).value = today
            ws_c.cell(row=r, column=15).value = result
            existing = ws_c.cell(row=r, column=10).value or ""
            if "Native-Hindi-teacher review 2026-05-17" not in str(existing):
                new = (str(existing).rstrip() + " — " + note) if existing else note
                ws_c.cell(row=r, column=10).value = new
            n_c += 1

    n_b = 0
    ws_b = wb["B. JLPT format"]
    for r in range(5, ws_b.max_row + 1):
        sid = ws_b.cell(row=r, column=1).value
        if sid in JLPT_STAMPS:
            result, note = JLPT_STAMPS[sid]
            ws_b.cell(row=r, column=14).value = today
            ws_b.cell(row=r, column=15).value = result
            existing = ws_b.cell(row=r, column=10).value or ""
            if "JLPT-exam-expert review 2026-05-17" not in str(existing):
                new = (str(existing).rstrip() + " — " + note) if existing else note
                ws_b.cell(row=r, column=10).value = new
            n_b += 1

    # Also stamp A-031 (the Hindi-pedagogy-audio scenario already Blocked)
    ws_a = wb["A. Japanese language"]
    for r in range(5, ws_a.max_row + 1):
        sid = ws_a.cell(row=r, column=1).value
        if sid == "A-031":
            ws_a.cell(row=r, column=14).value = today
            ws_a.cell(row=r, column=15).value = "Blocked — Hindi audio not shipped"
            existing = ws_a.cell(row=r, column=10).value or ""
            if "Native-Hindi-teacher review 2026-05-17" not in str(existing):
                note = "Native-Hindi-teacher review 2026-05-17: re-confirmed Blocked. Hindi-locale pedagogical audio does not exist yet (current Hindi locale: UI strings + glosses only; no audio). Per TS-A-031 fix (b77e1a4)."
                new = (str(existing).rstrip() + " — " + note) if existing else note
                ws_a.cell(row=r, column=10).value = new
            break

    wb.save(str(XLSX))
    print(f"Hindi tab stamped: {n_c}")
    print(f"JLPT tab stamped: {n_b}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
