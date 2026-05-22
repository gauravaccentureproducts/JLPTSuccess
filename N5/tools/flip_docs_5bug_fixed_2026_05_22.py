"""Flip BUG-156..160 (DOCS-VOCAB-006, DOCS-CORE-001, DOCS-BRAND-001,
DOCS-Q-001, DOCS-DKE-001) to Fixed + bump version v1.15.8 → v1.15.9."""
import sys, io, os, shutil, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_flip_docs_5bug"
if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]

FIXES = {
    159: {  # BUG-156 / DOCS-VOCAB-006
        "notes": "Documented では as 4th known mismatch in n5_vocab_whitelist_README.md (option (c) per bug spec). Updated counts (965/969 95.6%) + header (4, as of 2026-05-22). New CI invariant JA-147 locks README parity with computed mismatch set. Option (a) — author standalone vocab.json entry for では — deferred as separate authoring work (needs glosses/examples/audio/pitch-accent).",
    },
    160: {  # BUG-157 / DOCS-CORE-001
        "notes": "Added scope='n4' + scope_note field to 5 grammar.json entries (n5-144, n5-157, n5-158, n5-175, n5-176) per bug spec option (b). Added grammar_n5=173 to version.json.counts alongside grammar=178. New CI invariant JA-148 enforces grammar.json scope ↔ n5_core_pattern_ids.json classification agreement; JA-107 extended to validate scope-filtered count. Downstream consumers can now filter via scope==='n5'.",
    },
    161: {  # BUG-158 / DOCS-BRAND-001
        "notes": "Added explicit bullet to data/_review_packet/README.md 'Stripped (review-noise)' section documenting branding.json all-empty values as a privacy/anonymity strip (not a live-site bug; live index.html has hardcoded <title>JLPT N5</title> + 46 other JLPT refs). Doc fix only.",
    },
    162: {  # BUG-159 / DOCS-Q-001
        "notes": "Rewrote consumers section of n5_vocab_whitelist_README.md to explicitly call out the independence between questions.json (q-NNNN drill-bank, 290 entries, spaced-rep flows) and paper files (Q1..Q102 mock-test bank, 402 questions, 0 ID overlap with questions.json). Doc fix only.",
    },
    163: {  # BUG-160 / DOCS-DKE-001
        "notes": "Backfilled 25 placeholder rationales in data/dokkai_kanji_exception.json with per-kanji reasons derived from actual dokkai-corpus usage (survey via tools/survey_docs_dke_001_2026_05_22.py + 23 with surface citations + 2 canonical N5-dokkai surfaces for 自/近). Format matches the 65 pre-existing real entries. New CI invariant JA-149 rejects 'Pre-formalization' / 'rationale not individually recorded' phrasings to prevent re-introduction.",
    },
}

for row, spec in FIXES.items():
    ws.cell(row=row, column=8, value="Fixed")
    ws.cell(row=row, column=9, value="2026-05-22")
    ws.cell(row=row, column=10, value="<pending-main-commit>")
    ws.cell(row=row, column=11, value=spec["notes"])
    print(f"  R{row}: Fixed")

wb.save(XLSX)

# Bump version.json
vfp = "data/version.json"
with open(vfp, "r", encoding="utf-8") as f:
    v = json.load(f)
shutil.copy2(vfp, vfp + ".bak_2026_05_22_v15_9")
v["version"] = "v1.15.9"
v["cacheVersion"] = "jlptsuccess-n5-v1.15.9"
v["builtAt"] = "2026-05-22T03:00:00Z"
with open(vfp, "w", encoding="utf-8") as f:
    json.dump(v, f, ensure_ascii=False, indent=2)
print(f"  bumped version.json: → v1.15.9")
