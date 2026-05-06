"""Round-4 audit registrar - append ISSUE-025..034 + IMP-045..056 +
Q13..Q18 to feedback/n5-audit-2026-05-04.xlsx.

The audit-only prompt at prompts/N5Improvement.txt was rewritten between
round-3 and round-4 to add a STRATEGIC POSITIONING / SALEABILITY frame
+ niche-fit axis (N1 multilingual / N2 privacy / N3 institutional /
N4 all-in-one). This script preserves the existing 14-column schema
on the Audit findings sheet and adds a niche tag inline within the
Title field as `[N1]`, `[N2]`, etc., since adding a 15th column would
break existing fix scripts. The preceding round-3 audit-prompt text
re-write is at commit 0fa46ca.

Idempotent: rows whose ID already exists in the target sheet are
skipped. After append, runs the formatting fix to extend the dropdown
data-validation sqref + mirror row styles to the new rows so the user
gets a usable spreadsheet view.
"""
from __future__ import annotations
import io, sys
from copy import copy
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# 14-column rows for "Audit findings".
# Cols: ID, Type, Severity, Priority, Impact, Effort, Category, Location,
#       Title, Current state, Why this matters, Suggested direction,
#       Dependencies, Decision (Fix / Avoid / Defer)
# Niche tag is prepended to Title as `[N1]` / `[N2]` / `[N3]` / `[N4]` /
# `[none]` for grep-ability without schema change.
ISSUES = [
    # P1
    ('ISSUE-025', 'Issue', 'BLOCKER', 'P1', 'HIGH', 'LOW',
     'Documentation / legal',
     'repo root + N5/ - no LICENSE / LICENSE.md / COPYING file',
     '[N3] No LICENSE file at repo root - GitHub treats unlicensed repo as "all rights reserved", blocks legal fork',
     'CONTENT-LICENSE.md exists for content; code is unlicensed. Vocational schools / language schools / NGOs cannot legally fork or self-host.',
     'Closes off niche N3 (institutional / self-hosted) entirely. Without an explicit code license, every potential institutional adopter walks away.',
     'Add LICENSE (MIT or Apache-2) at repo root + cross-reference from README and CONTENT-LICENSE.md.',
     '-', ''),
    ('ISSUE-026', 'Issue', 'MAJOR', 'P1', 'HIGH', 'HIGH',
     'i18n', 'locales/{vi,id,ne,zh}.json + data/grammar.json + data/vocab.json + data/kanji.json',
     '[N1] vi/id/ne/zh locales at 44% UI coverage; content body (explanations / glosses / meanings) 0% translated',
     '4 non-English locale files have 33/75 keys each (44%). Content body fields (explanation_en, gloss, meanings, common_mistakes) are 100% English-only - no per-locale fields exist in the data schema.',
     'The recommended primary niche (multilingual non-English-native learners) hinges on translation depth. UI-only translation is the floor; content translation is the differentiator vs Bunpro/Renshuu.',
     'Translate the 42 missing keys per locale (native-reviewed); add explanation_vi/_id/_ne/_zh + gloss_vi/_id/_ne/_zh + meanings_vi/_id/_ne/_zh content fields.',
     'Q13, Q14', ''),
    ('ISSUE-027', 'Issue', 'MAJOR', 'P1', 'HIGH', 'LOW',
     'UX / discoverability', 'js/home.js syllabus header',
     '[N2] privacy claim invisible on home - moves the most-defensible niche claim from invisible to visible',
     'Home renders the syllabus dashboard with no privacy band, no hero copy, no "no login required" trust signal. PRIVACY.md content is reachable only via the footer link.',
     'Niche N2 is currently invisible to first-time visitors. The strongest competitive differentiator (vs Bunpro/Duolingo/Renshuu account requirements) is hidden.',
     'Add a hairline trust band above the syllabus title: "No login · No tracking · Works offline · Open source · 100% on-device".',
     'ISSUE-025', ''),
    # P2
    ('ISSUE-028', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'LOW',
     'i18n / discoverability', 'index.html secondary-nav + js/settings.js (set-locale only)',
     '[N1] locale switcher only in Settings - invisible on first paint',
     'The locale-switcher <select> is buried in Settings. A non-English-native first-time visitor sees the English UI by default and has no signposting that other locales exist.',
     'Discoverability of the multilingual claim. Without a visible language chip on first paint, the multilingual feature might as well not exist for the audience that needs it most.',
     'Add a 5-chip locale-switcher (EN · VI · ID · NE · ZH) to the header, between the search input and fullscreen toggle.',
     '-', ''),
    ('ISSUE-029', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'i18n / UX', 'js/i18n.js:59-60',
     '[N1] Accept-Language detection silent - no user-visible feedback when locale is auto-picked',
     'navigator.language is read once on first init and silently used to pick the default locale. Users with mismatched system + UI language never know the option exists.',
     'Silent locale-detection is a missed opportunity to surface multilingual support.',
     'On first init, if navigator.language matches a non-EN SUPPORTED locale and the user has not explicitly chosen, show a one-time toast "App language: <native name>. Change in Settings."',
     'ISSUE-028', ''),
    ('ISSUE-030', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'MEDIUM',
     'Content / trust',
     'data/grammar.json + data/vocab.json + data/kanji.json + data/reading.json + data/listening.json',
     '[none] no native-review provenance tags on any content (trust signal for N1 + N4)',
     'No item in any corpus carries a review_status / native_reviewed / llm_curated / provenance tag. A learner cannot distinguish native-authored content from auto-generated content.',
     'Trust-erosion vector. Critical for N4 (all-in-one credibility) and unlocks honest content labeling for the multilingual content (N1) when translations land.',
     'Add review_status: "native_reviewed" | "llm_curated" | "auto_generated" to every top-level item; render a small badge on detail pages.',
     'Q16', ''),
    # P3
    ('ISSUE-031', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'MEDIUM',
     'Documentation', 'README.md',
     '[N3] README has no self-host / fork guide - niche N3 unclaimable without it',
     'README is dev-oriented. No "host this yourself" / fork-and-customize guide, no theme-token override docs, no domain-asset audit.',
     'Niche N3 (institutional / self-hosted) cannot be claimed without it.',
     'Add docs/SELF-HOST.md covering fork → brand → deploy + theme-customization seam.',
     'ISSUE-025', ''),
    ('ISSUE-032', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'LOW',
     'Discoverability / SEO', 'index.html head',
     '[none] no og:image / twitter:card / structured data - link previews degrade to "no preview"',
     'Page has description and theme-color meta but no og:image, og:title, og:description, twitter:card, no JSON-LD EducationalApplication schema.',
     'Link previews on social media, Show HN, Reddit, Discord all degrade to "no preview". Affects every niche.',
     'Add og:* / twitter:card meta block + a small JSON-LD schema block + an assets/og-image.png (1200×630).',
     '-', ''),
    ('ISSUE-033', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'Content correctness', 'data/grammar.json (178 patterns)',
     '[N4] core-N5 vs late-N5 not surfaced as count-level distinction',
     'README mentions "178 grammar patterns" but contains tier=late_n5 patterns flagged as "borderline N5/N4". Learner studying for the strict N5 exam gets a slightly inflated count.',
     'Content honesty. Learners may study borderline patterns for an N5 exam where they would not appear.',
     'Add data/n5_core_pattern_ids.json whitelist + invariant; report count as "178 patterns (164 core + 14 late-N5)".',
     '-', ''),
    # P4
    ('ISSUE-034', 'Issue', 'MINOR', 'P4', 'LOW', 'LOW',
     'PWA / discoverability', 'js/pwa.js install banner',
     '[N2] install prompt fires once, no fallback for browsers that do not fire beforeinstallprompt',
     'Install prompt fires once on beforeinstallprompt. No periodic re-prompt, no install-prompt UI for Firefox / iOS Safari, no "you can use this offline" hint for users who have not installed.',
     'N2 (offline-first) is invisible to anyone who dismisses or never sees the install prompt.',
     'Add a small "Install for offline use" link to the home privacy band (when not installed).',
     'ISSUE-027', ''),
]

IMPS = [
    # P1
    ('IMP-048', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'LOW',
     'UX / discoverability', 'home dashboard',
     '[N2] add visible "no login · no tracking · offline" trust band on home',
     'Privacy claim invisible (paired with ISSUE-027).',
     'Best-in-class: DuckDuckGo ships its privacy claim on every search-result page; Signal anchors "private by design" in its hero. JLPTSuccess delivers the contract but does not say so.',
     'Hairline trust band: "No login · No tracking · Works offline · Open source · 100% on-device".',
     'ISSUE-025, ISSUE-027', ''),
    # P2
    ('IMP-045', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'HIGH',
     'Content / i18n', 'data/grammar.json 178 patterns × explanation_en field',
     '[N1] translate grammar explanation_en to vi/id/ne/zh - largest content-translation lift',
     'Explanations English-only.',
     'No competitor does this. Bunpro EN-only. Renshuu has multi-language but auto-translated. Native-reviewed multilingual grammar explanations would be unique to JLPTSuccess.',
     'Add explanation_vi/_id/_ne/_zh per pattern; renderer picks based on current locale; falls back to EN.',
     'ISSUE-026, Q13, Q14', ''),
    ('IMP-046', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'HIGH',
     'Content / i18n', 'data/vocab.json 1041 entries × gloss field',
     '[N1] translate vocab glosses to vi/id/ne/zh',
     'Glosses English-only.',
     'Best-in-class: Jisho.com EN-only; Renshuu auto-translated. Vocabulary glossing in target-language is the most basic translation a multilingual learner needs.',
     'Add gloss_vi/_id/_ne/_zh per entry; consider machine-translation seed + native review pass for budget reasons.',
     'IMP-045, Q14', ''),
    ('IMP-047', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'MEDIUM',
     'Content / i18n', 'data/kanji.json 106 entries × meanings array',
     '[N1] translate kanji meanings to vi/id/ne/zh',
     'Meanings English-only.',
     'Best-in-class: WaniKani EN-only; Kanji Damage EN-only. Multi-language kanji meanings would be unique.',
     'Add meanings_vi/_id/_ne/_zh arrays per entry.',
     'IMP-045, IMP-046, Q14', ''),
    ('IMP-049', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Documentation / legal', 'repo root',
     '[N3] add LICENSE (MIT for code, CC BY-SA for content) + 1-page SELF-HOST.md',
     'No LICENSE file (paired with ISSUE-025), no SELF-HOST guide.',
     'Best-in-class: MDN, freeCodeCamp ship dual licenses (code MIT, content CC BY-SA) + clear self-host docs. JLPTSuccess can match this with a single afternoon of paperwork.',
     'Ship LICENSE (MIT for code) + update CONTENT-LICENSE.md (CC BY-SA 4.0) + new docs/SELF-HOST.md.',
     'ISSUE-025, ISSUE-031', ''),
    # P3
    ('IMP-050', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'HIGH',
     'Content / Kanji', 'data/kanji.json (no radicals/mnemonic fields)',
     '[N4] kanji radical decomposition (KanjiDic2 / Heisig) - partial close of WaniKani depth gap',
     'Kanji surface has stroke order + meanings + readings + sentences but no radical decomposition or mnemonic.',
     'Best-in-class: WaniKani defining moat. Kanji Damage has crude mnemonics (free). Tofugu Kanji Learner Course is paid. Adding minimum-viable radical info closes the visible gap.',
     'Add radicals: [...] + mnemonic: "..." fields, populated from KanjiDic2 + curated mnemonic source. Render below additional_readings block.',
     '-', ''),
    ('IMP-051', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'LOW',
     'SEO / discoverability', 'index.html head + new assets/og-image.png',
     '[none] og:image + twitter:card + JSON-LD EducationalApplication schema for shareability',
     'ISSUE-032 deferred form.',
     'Best-in-class: every modern PWA. JLPTSuccess being shareable on social media is currently a 0/10.',
     'See ISSUE-032.',
     'ISSUE-032', ''),
    ('IMP-052', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'Brand customization', 'css/main.css design tokens',
     '[N3] runtime theme-override seam for institutional forks (no source edits)',
     'Design tokens exist (--color-accent, --color-bg, etc.) but no documented override surface; institutional forks must edit CSS source directly.',
     'Best-in-class: Wikimedia MediaWiki skins allow theme override via a single config file. Discourse has admin-time theme presets.',
     'Load data/theme-overrides.json at startup if present; map keys to CSS custom properties at runtime; document keys in SELF-HOST.md.',
     'ISSUE-031', ''),
    # P4
    ('IMP-053', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'MEDIUM',
     'i18n future-proofing', 'index.html + css/main.css',
     '[N1] RTL support via logical-property CSS (future Arabic/Hebrew/Urdu locales)',
     'No RTL support (<html dir> implicitly LTR; no logical-property CSS).',
     'Irrelevant for currently shipped locales (vi/id/ne/zh all LTR). Adding for Arabic/Hebrew/Urdu would extend N1 reach further.',
     'Replace margin-left/right etc. with margin-inline-start/end; add dir="rtl" per-locale.',
     '-', ''),
    ('IMP-054', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'HIGH',
     'Distribution', 'build pipeline',
     '[N3] Trusted Web Activity / Capacitor wrappers for Play Store / App Store',
     'Only PWA install. No native-app-store path.',
     'Best-in-class: Bunpou web + native via React Native; Renshuu native apps. Institutional users in markets with poor mobile-PWA support cannot easily distribute the PWA.',
     'Investigate Trusted Web Activity (Android Play Store) and Capacitor (iOS App Store) wrappers - neither requires rewriting.',
     'Q17', ''),
    ('IMP-055', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'Tests', 'tests/',
     '[none] Playwright spec coverage for round-3 routes (#/missed, #/sitting, audio-player skin)',
     'No Playwright spec covers the round-3 routes or the daily-goal progress UI.',
     'Best-in-class: every web app of this size has integration tests on critical-path routes.',
     'Extend tests/v1.12.28-features.spec.js (or new spec) with the new routes + audio-player skin button-press flow.',
     '-', ''),
    # P5
    ('IMP-056', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'HIGH',
     'Content / community', 'contributor docs',
     '[N1] translator-contributor on-ramp (CONTRIBUTING / TRANSLATING.md)',
     'No contributor guide for translators. Native speakers willing to fix vi/id/ne/zh strings have no on-ramp.',
     'Best-in-class: Crowdin / Weblate integration; or simpler: a CONTRIBUTING.md with locale-file examples.',
     'Add docs/TRANSLATING.md with a pull-request template.',
     'ISSUE-026', ''),
]

QUESTIONS = [
    ('Q13', 'Primary niche commitment',
     'Round-4 audit recommends N1 (multilingual non-English-native) as primary, N2 (privacy/no-account/offline) as secondary. The audit prompt now explicitly frames around saleability + niche fit; trying to beat Bunpro/WaniKani/Renshuu at their core games is unwinnable.',
     'Does the product owner agree with N1+N2 as primary+secondary, or prefer a different primary (e.g., N4 all-in-one)? Affects every Section-3 IMP tagged with niche-fit.',
     ''),
    ('Q14', 'Translation budget for N1',
     'Native-quality translation of 75 UI keys × 4 locales (300 strings) plus content body (grammar + vocab + kanji = ~1325 unique strings × 4 locales = ~5300 strings). Three paths: (a) commission native translators (USD$1500-5000), (b) machine-translation seed + crowdsourced native review (free but slow), (c) machine-translation only with explicit "auto-translated" labeling.',
     'Pick (a)/(b)/(c) per locale, or a mix. Affects ISSUE-026 + IMP-045/046/047.',
     ''),
    ('Q15', 'License choice',
     'Code license: MIT (permissive) vs Apache-2 (permissive + patent-grant) vs AGPL-3 (copyleft). Content license: confirm CONTENT-LICENSE.md as CC BY-SA 4.0 vs CC BY 4.0 vs CC BY-NC 4.0.',
     'Specific code + content license picks. Affects ISSUE-025 + IMP-049.',
     ''),
    ('Q16', 'Native-review staffing',
     'Native-Japanese review of grammar / vocab / kanji / reading / listening content (ISSUE-030 provenance work) requires a Japanese-speaker pool.',
     'Are we commissioning native reviewers, partnering with a JP language school, or accepting community PRs?',
     ''),
    ('Q17', 'Distribution strategy',
     'PWA-only (current) vs Play Store TWA + App Store Capacitor wrappers (IMP-054).',
     'Decide PWA-only or add native-store wrappers. Affects N3 reach.',
     ''),
    ('Q18', 'Round-2/3 leftover questions',
     'Q4/Q6/Q8/Q11/Q12 from prior rounds remain blank. Some may be obsoleted by Q13 niche commitment; others (Q11 native-audio budget, Q8 localization commitment) are now subsumed by Q14.',
     'Sweep Q4/Q6/Q8/Q11/Q12 - close as obsoleted-by-Q13/Q14 where appropriate, or keep open with explicit rationale.',
     ''),
]


def _copy_style(src_cell, dst_cell) -> None:
    dst_cell.font = copy(src_cell.font)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format


def main() -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)

    # --- Audit findings sheet ---
    ws = wb['Audit findings']
    existing = {str(r[0].value).strip() for r in ws.iter_rows(min_row=5)
                if r and r[0].value}
    new_rows = [r for r in ISSUES + IMPS if r[0] not in existing]
    next_row = ws.max_row + 1
    ref_row = 5  # canonical formatting row
    for r in new_rows:
        for col_offset, val in enumerate(r):
            ws.cell(row=next_row, column=col_offset + 1, value=val)
        # Mirror styles from row 5 to the new row.
        for c in range(1, ws.max_column + 1):
            _copy_style(ws.cell(row=ref_row, column=c),
                        ws.cell(row=next_row, column=c))
        next_row += 1
    last_data_row = ws.max_row
    # Extend the Decision-column data-validation sqref.
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for cr in list(dv.sqref.ranges):
            if str(cr).startswith('N'):
                dv.sqref.add(CellRange(f'N5:N{last_data_row}'))
                break

    # --- Open questions sheet ---
    ws_q = wb['Open questions']
    existing_q = {str(r[0].value).strip() for r in ws_q.iter_rows(min_row=4)
                  if r and r[0].value}
    new_qs = [r for r in QUESTIONS if r[0] not in existing_q]
    next_q_row = ws_q.max_row + 1
    ref_q_row = 4
    for r in new_qs:
        for col_offset, val in enumerate(r):
            ws_q.cell(row=next_q_row, column=col_offset + 1, value=val)
        for c in range(1, ws_q.max_column + 1):
            _copy_style(ws_q.cell(row=ref_q_row, column=c),
                        ws_q.cell(row=next_q_row, column=c))
        next_q_row += 1
    last_q_row = ws_q.max_row
    for dv in ws_q.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        for cr in list(dv.sqref.ranges):
            if str(cr).startswith('E'):
                dv.sqref.add(CellRange(f'E4:E{last_q_row}'))
                break

    wb.save(XLSX)
    print(f'Audit findings:  appended {len(new_rows)} rows '
          f'({sum(1 for r in new_rows if r[0].startswith("ISSUE-"))} issues + '
          f'{sum(1 for r in new_rows if r[0].startswith("IMP-"))} improvements)')
    skipped = (set(r[0] for r in ISSUES + IMPS) & existing)
    if skipped:
        print(f'  skipped (already present): {", ".join(sorted(skipped))}')
    print(f'Open questions:  appended {len(new_qs)} rows')
    skipped_q = (set(r[0] for r in QUESTIONS) & existing_q)
    if skipped_q:
        print(f'  skipped (already present): {", ".join(sorted(skipped_q))}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
