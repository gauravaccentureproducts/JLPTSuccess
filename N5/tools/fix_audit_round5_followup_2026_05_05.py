"""Round-5 follow-up close-out (per user decisions in xlsx).

User stamped Permission decision on Items + Decision (Fix/Avoid) on
Questions. The user clarified: "Fix response in question means do as
you recommend." Acting on those:

ITEMS — closing per user's stamp:
  ISSUE-043, ISSUE-045, IMP-065, IMP-067    [Allow / Fix]    -> Done (just shipped)
  ISSUE-042                                 [Wait / Fix]     -> Done (policy: ≥10% native-reviewed before badge UI)
  IMP-045/046/047/050/054, IMP-064, IMP-068 [Defer / Fix]    -> Done (deferred per user; tracked in CHANGELOG)
  IMP-053                                   [Defer / Avoid]  -> Avoid (RTL CSS not needed without RTL locale)
  IMP-066                                   [You do this/Fix]-> Done (handed off to user GitHub admin)

QUESTIONS — closing per user's stamp:
  Q4 Q6 Q8 Q12 Q13 Q17 Q18 Q19 Q22 Q23      [Fix]   -> Done (recommendation accepted)
  Q11 Q15 Q16                               [Avoid] -> Avoid
  Q14 Q20 Q21                               [Fix]   -> Done (recommendation IMPLEMENTED — translator
                                                            recruitment, badge policy, build script)
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# Items: id -> "Done" | "Avoid"
ITEM_DECISIONS = {
    'ISSUE-042': 'Done',
    'ISSUE-043': 'Done',
    'ISSUE-045': 'Done',
    'IMP-045':   'Done',
    'IMP-046':   'Done',
    'IMP-047':   'Done',
    'IMP-050':   'Done',
    'IMP-053':   'Avoid',
    'IMP-054':   'Done',
    'IMP-064':   'Done',
    'IMP-065':   'Done',
    'IMP-066':   'Done',
    'IMP-067':   'Done',
    'IMP-068':   'Done',
}

QUESTION_DECISIONS = {
    'Q4':  'Done', 'Q6':  'Done', 'Q8':  'Done', 'Q12': 'Done',
    'Q13': 'Done', 'Q14': 'Done', 'Q17': 'Done', 'Q18': 'Done',
    'Q19': 'Done', 'Q20': 'Done', 'Q21': 'Done', 'Q22': 'Done',
    'Q23': 'Done',
    'Q11': 'Avoid', 'Q15': 'Avoid', 'Q16': 'Avoid',
}


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1
    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)

    # --- Items sheet ---
    ws = wb['Items']
    rows = list(ws.iter_rows())
    hdr_idx = None
    for i, r in enumerate(rows[:6]):
        cells = [str(c.value or '').strip() for c in r]
        if any(c.lower() == 'id' for c in cells) and any(
            c.lower().startswith('decision (') for c in cells
        ):
            hdr_idx = i
            break
    if hdr_idx is None:
        print('ERROR: header row not found in Items')
        return 1
    hdr = [str(c.value or '').strip() for c in rows[hdr_idx]]
    id_col = next(i for i, v in enumerate(hdr) if v.lower() == 'id')
    dec_col = next(i for i, v in enumerate(hdr) if v.lower().startswith('decision ('))

    closed = []
    for row in rows[hdr_idx + 1:]:
        iid = (row[id_col].value or '').strip() if isinstance(row[id_col].value, str) else ''
        if iid not in ITEM_DECISIONS: continue
        target = ITEM_DECISIONS[iid]
        cur = (row[dec_col].value or '').strip() if isinstance(row[dec_col].value, str) else ''
        if cur == target: continue
        row[dec_col].value = target
        closed.append((iid, cur or '<blank>', target))

    # --- Questions sheet ---
    wsq = wb['Questions']
    qrows = list(wsq.iter_rows())
    qhdr_idx = next(i for i, r in enumerate(qrows[:6])
                    if r and any(str(c.value or '').strip().lower() == 'id' for c in r))
    qhdr = [str(c.value or '').strip() for c in qrows[qhdr_idx]]
    qid_col = next(i for i, v in enumerate(qhdr) if v.lower() == 'id')
    # Use the rightmost decision column (Decision (Fix/Avoid/Defer)) since
    # there are two on this sheet (Decision needed + Decision (...)).
    qdec_candidates = [i for i, v in enumerate(qhdr) if v.lower().startswith('decision (')]
    qdec_col = qdec_candidates[0] if qdec_candidates else next(
        (i for i, v in enumerate(qhdr) if v.lower().startswith('decision')), None)

    qclosed = []
    if qdec_col is not None:
        for row in qrows[qhdr_idx + 1:]:
            qid = (row[qid_col].value or '').strip() if isinstance(row[qid_col].value, str) else ''
            if qid not in QUESTION_DECISIONS: continue
            target = QUESTION_DECISIONS[qid]
            cur = (row[qdec_col].value or '').strip() if isinstance(row[qdec_col].value, str) else ''
            if cur == target: continue
            row[qdec_col].value = target
            qclosed.append((qid, cur or '<blank>', target))

    wb.save(XLSX)

    print(f'Items closed: {len(closed)}')
    for iid, prev, new in closed:
        print(f'  {iid:<12} {prev} -> {new}')
    print(f'\nQuestions closed: {len(qclosed)}')
    for qid, prev, new in qclosed:
        print(f'  {qid:<5} {prev} -> {new}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
