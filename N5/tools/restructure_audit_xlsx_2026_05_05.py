"""Restructure feedback/n5-audit-2026-05-04.xlsx for clarity:

  1. Drop the 3 reference-only sheets ("Top-10 do these first",
     "Coverage map", "Legend") — they were cumulative reference data,
     not actionable.
  2. Rename "Audit findings" -> "Items" and "Open questions" -> "Questions".
  3. Add a "Plain English" column to "Items" so a non-technical reader
     can understand each row without parsing audit jargon.
  4. Add a "Permission decision" column at the end of "Items" with a
     simpler dropdown (Allow / Don't allow / Defer / Skip) targeted
     at the items where the assistant needs the user's go-ahead.
  5. Add a "Plain English" column to "Questions" so each open product
     decision is described in simple language.

Preserves all existing data, formatting, and the existing
"Decision (Fix / Avoid / Defer)" column. Idempotent — running twice
produces no further change.
"""
from __future__ import annotations
import io, sys
from copy import copy
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).resolve().parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# Plain-English text for each currently-pending item or recently-active
# row. Keys = ID. Format: (plain English, permission-decision-option,
# explanation if pending). For items already Done / Avoid we leave the
# new columns blank — they're closed and don't need re-discussion.
PLAIN_ENGLISH = {
    # Round-5 needs-decision items (still pending)
    'ISSUE-042': (
        "We added a 'review status' label to every lesson (e.g., 'AI-drafted' or 'native-reviewed'). Right now ALL 1405 lessons say 'AI-drafted' which means the label is useless until at least some are upgraded by a native Japanese reviewer.",
        'Wait',
        'Three options: (a) show the badge anyway with a note "all currently AI-drafted, native review in progress", (b) wait until at least 50 items are native-reviewed before showing, (c) skip the badge entirely and use the data only behind the scenes.',
    ),
    'IMP-064': (
        "When a Vietnamese / Indonesian / Nepali / Chinese visitor opens the app for the first time, we currently auto-detect their browser language and show a small toast. Should we add an extra step BEFORE the placement test asking 'Your language is X — keep this or pick another?'",
        'Defer',
        'UX trade-off: extra step = more friction but more clarity. Skipping it = faster but the locale auto-detect might surprise a user who shares a device.',
    ),
    'IMP-066': (
        "Set up the GitHub repo metadata: description, topics (jlpt, n5, japanese, pwa, multilingual, privacy, open-source), homepage URL, and a Releases page listing each version. This makes the repo show up in GitHub searches and look professional.",
        'You do this',
        'Requires your GitHub admin access. I cannot edit repo settings. You can do this from github.com/gauravaccentureproducts/JLPTSuccess > Settings > General.',
    ),
    'IMP-068': (
        "Connect the project to Crowdin or Weblate so non-coding native speakers can translate strings via a web UI instead of editing JSON files in a PR. Free for open-source projects.",
        'Defer',
        'Worth doing only if we get translation contributors. Right now there are zero. Defer until 1-2 native speakers volunteer.',
    ),

    # Round-5 skip-on-error items (still pending)
    'ISSUE-043': (
        "Make the JavaScript files smaller by running them through a minifier (esbuild or terser). Would shrink the 387 KB JS bundle by ~30-40%, making the app load faster on slow phones.",
        'Allow',
        'Needs me to install a new dev tool (esbuild via npm). Failure is silent — if your dev environment blocks npm installs, the build just keeps using the unminified files.',
    ),
    'ISSUE-045': (
        "Add automated screenshot tests for the new pages (Mock Sitting, Wrong-answer history, audio player, trust band). If a future change accidentally breaks the layout, the test catches it.",
        'Allow',
        'Needs Playwright + Chromium browser already installed. If your environment lacks them, the test files still get added but cannot run until the browsers are installed.',
    ),
    'IMP-067': (
        "Add WebP / AVIF versions of the app icons (PNG today). Modern browsers prefer WebP and it is ~30-50% smaller, so install + first-paint is slightly faster.",
        'Allow',
        'Needs Pillow with WebP / AVIF support. May fail silently if the local Pillow lacks the WebP plugin; in that case the PNG icons stay in use.',
    ),
    'IMP-065': (
        "Companion to ISSUE-045: add the actual screenshot snapshots to the visual-regression test suite.",
        'Allow',
        'Same Playwright dependency as ISSUE-045.',
    ),

    # Round-4 deferred Fix items (still pending — content / decision blocked)
    'IMP-045': (
        "Translate the 178 grammar lesson explanations (the long English text explaining each grammar point) into Vietnamese, Indonesian, Nepali, Chinese.",
        'Defer',
        'About 712 sentences to translate. Cannot ship machine-translated grammar explanations — the audit explicitly warned this would damage trust. Needs human translators (paid or volunteer).',
    ),
    'IMP-046': (
        "Translate the 1041 vocabulary glosses (the English meaning of each word like '車 = car') into Vietnamese, Indonesian, Nepali, Chinese.",
        'Defer',
        'About 4164 short strings. Could be machine-translated more safely than grammar (single words have less context to get wrong) but still needs native review for quality.',
    ),
    'IMP-047': (
        "Translate the 106 kanji meanings into Vietnamese, Indonesian, Nepali, Chinese.",
        'Defer',
        'About 424 short strings. Same as IMP-046 but smaller scope.',
    ),
    'IMP-050': (
        "Add radical breakdown + memory hint to each kanji (e.g., 校 = 木 + 交 = tree + cross). This is what WaniKani is famous for.",
        'Defer',
        'Substantial content authoring. Need to import data from KanjiDic2 (free) and write hints. The audit recommends NOT chasing WaniKani depth — minimum-viable only.',
    ),
    'IMP-053': (
        "Make the CSS layout work for right-to-left languages (Arabic, Hebrew, Urdu) in case we add such locales in the future.",
        'Defer',
        'Currently no RTL locale ships, so this is preventive work. Skip until we actually add an RTL language.',
    ),
    'IMP-054': (
        "Wrap the web app as a Trusted Web Activity (Android Play Store) and Capacitor app (iOS App Store) so it can be distributed via app stores.",
        'Defer',
        'Big distribution change. Affects how users install the app. Needs a strategy decision: PWA-only forever, or also app stores?',
    ),
}

# Plain-English wrappers for the open Questions sheet.
QUESTION_PLAIN = {
    'Q4':  "Mock-test paper segmentation: should the 7 papers per section have an explicit difficulty curve (paper 1 easiest -> paper 7 hardest), or stay random?",
    'Q6':  "Service-worker precache strategy: should we cache all audio (~22 MB) on first visit, or lazy-load (current)?",
    'Q8':  "Translation commitment for vi/id/ne/zh: machine-translate everything as a starting point, OR wait for native reviewers, OR commit budget for paid translators?",
    'Q11': "Native-speaker audio recordings: budget USD$300-1500 to replace the synthetic gTTS audio with real Japanese voice talent. Yes / no / defer.",
    'Q12': "Round-2 leftover questions Q2-Q7: most are now answered indirectly by later work. Sweep and close, or keep open with explicit reasoning?",
    'Q13': "Strategic primary niche: I recommended N1 (multilingual non-EN-native learners). Do you agree, or pick a different primary positioning?",
    'Q14': "Translation budget: native-quality reviewers (USD$1500-5000), machine-translation + crowd-sourced review (free but slow), OR machine-only with explicit 'auto-translated' labels?",
    'Q15': "License confirmation: code is MIT (LICENSE shipped), content is CC BY-SA 4.0 (CONTENT-LICENSE.md). Are these final?",
    'Q16': "Native-Japanese review staffing: commission paid reviewers, partner with a Japanese language school, or rely on community PRs?",
    'Q17': "Distribution strategy: PWA-only (current) or also wrap for Android Play Store + iOS App Store via TWA / Capacitor?",
    'Q18': "Round-2/3 leftover questions: resolve as a batch sit-down or continue case-by-case?",
    'Q19': "Build-stamp invariants count: should the build script auto-compute it (now done — ISSUE-035 closed), or accept hand-written placeholder?",
    'Q20': "Native-review staffing recruitment: actively recruit per-locale reviewers via Reddit / Discord / GitHub issue, or wait passively?",
    'Q21': "Provenance badge UI: when do we ship the 'AI-drafted vs native-reviewed' badge? See ISSUE-042 for options a/b/c.",
    'Q22': "SEO marketing: invest time in Show HN / blog posts / SEO improvements, or stay intentionally low-discoverability?",
    'Q23': "Long-deferred questions Q4/Q6/Q8/Q11/Q12/Q13-Q18: resolve as a single batch or case-by-case?",
}

DROP_SHEETS = {'Top-10 do these first', 'Coverage map', 'Legend'}
RENAME = {
    'Audit findings': 'Items',
    'Open questions': 'Questions',
}


def _copy_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.number_format = src.number_format


def main() -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.cell_range import CellRange
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)

    # Step 1: drop reference-only sheets.
    for name in list(wb.sheetnames):
        if name in DROP_SHEETS:
            del wb[name]
            print(f'  dropped sheet: {name!r}')

    # Step 2: rename remaining sheets.
    for old, new in RENAME.items():
        if old in wb.sheetnames and new not in wb.sheetnames:
            wb[old].title = new
            print(f'  renamed sheet: {old!r} -> {new!r}')

    # Step 3: add Plain English + Permission Decision columns to Items.
    ws = wb['Items']
    # Find header row (row 4 in this xlsx).
    header_row_idx = None
    for r in range(1, min(7, ws.max_row + 1)):
        cells = [str(c.value or '').strip() for c in ws[r]]
        if any(c.lower() == 'id' for c in cells) and any(
            c.lower().startswith('decision') for c in cells
        ):
            header_row_idx = r
            break
    if header_row_idx is None:
        print('ERROR: header row not found in Items')
        return 1

    header = [c.value for c in ws[header_row_idx]]
    # Ensure new columns exist; create them at the end if not.
    if 'Plain English' not in header:
        col = ws.max_column + 1
        ws.cell(row=header_row_idx, column=col, value='Plain English')
        # Mirror header style.
        _copy_style(ws.cell(row=header_row_idx, column=1),
                    ws.cell(row=header_row_idx, column=col))
        plain_col = col
        print(f'  Items: added column "Plain English" at col {col}')
    else:
        plain_col = header.index('Plain English') + 1
    header = [c.value for c in ws[header_row_idx]]
    if 'Permission decision' not in header:
        col = ws.max_column + 1
        ws.cell(row=header_row_idx, column=col, value='Permission decision')
        _copy_style(ws.cell(row=header_row_idx, column=1),
                    ws.cell(row=header_row_idx, column=col))
        perm_col = col
        print(f'  Items: added column "Permission decision" at col {col}')
    else:
        perm_col = header.index('Permission decision') + 1

    # Find ID column.
    id_col = next(i for i, v in enumerate(header, start=1)
                  if v and str(v).strip().lower() == 'id')

    # Populate Plain English + Permission decision per row.
    populated_plain = 0
    populated_perm = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        id_cell = row[id_col - 1]
        iid = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
        if iid in PLAIN_ENGLISH:
            plain, perm, _explanation = PLAIN_ENGLISH[iid]
            cell = ws.cell(row=id_cell.row, column=plain_col)
            if not cell.value:
                cell.value = plain
                _copy_style(ws.cell(row=header_row_idx + 1, column=1), cell)
                populated_plain += 1
            cell_p = ws.cell(row=id_cell.row, column=perm_col)
            if not cell_p.value:
                cell_p.value = perm
                _copy_style(ws.cell(row=header_row_idx + 1, column=1), cell_p)
                populated_perm += 1

    # Step 4: data validation on the new Permission decision column —
    # simpler dropdown than the legacy Fix/Avoid/Defer one.
    last_data_row = ws.max_row
    perm_letter = ws.cell(row=header_row_idx, column=perm_col).column_letter
    dv = DataValidation(
        type='list',
        formula1='"Allow,Don\'t allow,Defer,You do this,Wait,Skip"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(CellRange(f'{perm_letter}{header_row_idx + 1}:{perm_letter}{last_data_row}'))
    print(f'  Items: Permission decision dropdown applied on '
          f'{perm_letter}{header_row_idx + 1}:{perm_letter}{last_data_row}')

    print(f'  Items: populated {populated_plain} plain-English / '
          f'{populated_perm} permission rows')

    # Step 5: same treatment on Questions sheet.
    if 'Questions' in wb.sheetnames:
        wsq = wb['Questions']
        q_header_idx = None
        for r in range(1, min(7, wsq.max_row + 1)):
            cells = [str(c.value or '').strip() for c in wsq[r]]
            if any(c.lower() == 'id' for c in cells):
                q_header_idx = r
                break
        if q_header_idx is None:
            print('  Questions: header row not found, skipping')
        else:
            qheader = [c.value for c in wsq[q_header_idx]]
            if 'Plain English' not in qheader:
                col = wsq.max_column + 1
                wsq.cell(row=q_header_idx, column=col, value='Plain English')
                _copy_style(wsq.cell(row=q_header_idx, column=1),
                            wsq.cell(row=q_header_idx, column=col))
                qplain_col = col
                print(f'  Questions: added column "Plain English" at col {col}')
            else:
                qplain_col = qheader.index('Plain English') + 1

            qid_col = next(i for i, v in enumerate(qheader, start=1)
                           if v and str(v).strip().lower() == 'id')
            n = 0
            for row in wsq.iter_rows(min_row=q_header_idx + 1):
                id_cell = row[qid_col - 1]
                qid = (id_cell.value or '').strip() if isinstance(id_cell.value, str) else ''
                if qid in QUESTION_PLAIN:
                    cell = wsq.cell(row=id_cell.row, column=qplain_col)
                    if not cell.value:
                        cell.value = QUESTION_PLAIN[qid]
                        _copy_style(wsq.cell(row=q_header_idx + 1, column=1), cell)
                        n += 1
            print(f'  Questions: populated {n} plain-English rows')

    wb.save(XLSX)
    print('\nDone. xlsx now has only 2 sheets: Items, Questions.')
    print('Plain English column populated for currently-pending items only.')
    print('Permission decision dropdown ready for items needing your call.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
