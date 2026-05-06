"""Round-8 (depth-first) audit registration — 2026-05-06.

Registers findings from the depth-first audit run (post-locale-transition):
- 14 issues (ISSUE-077..090)
- 13 improvements (IMP-102..114)
-  6 questions (Q33..Q38)

Modeled on tools/_register_round7_findings.py. Idempotent: skips existing IDs.
"""
from __future__ import annotations
import io, sys
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# (ID, Type, Severity, Priority, Impact, Effort, Category, Location, Title,
#  Current state, Why this matters / Best-in-class, Suggested direction,
#  Dependencies, Decision (blank), Description, Permission)
ITEMS = [
    # ============================================================
    # ISSUES — round 8 depth-first
    # ============================================================
    ('ISSUE-077', 'Issue', 'BLOCKER', 'P1', 'HIGH', 'HIGH',
     'L1-interference notes (Hindi-specific)',
     'data/grammar.json — 178 patterns, l1_notes.hi field',
     '[N1] [DEPTH] l1_notes.hi 0/178 — niche-N1 unique-claim lever empty',
     '0/178 patterns carry l1_notes.hi. The 8 mandatory Hindi-specific contrast areas (postposition→particle mapping से→から/で, को→を/に, में→に/で; verb-agreement transfer; tense over-marking; politeness mismatch; negative-formation placement; question-particle position; plural marking; counter-system overlap) are absent everywhere.',
     'No competitor delivers L1-targeted notes for Hindi-speaking JLPT learners. Niche-N1 cannot become "credible" without these. Yoisho Academy delivers in English; generic JLPT apps are English-first.',
     'Author top-30 grammar patterns first (n5-001..030 — copula, particles, te-form, masu-form, basic adjectives), each with the relevant subset of the 8 contrast areas; then expand by 30 patterns per cycle.',
     '—', '', '', 'No permission required'),
    ('ISSUE-078', 'Issue', 'MAJOR', 'P1', 'HIGH', 'HIGH',
     'Hindi long-form explanation',
     'data/grammar.json — 178 patterns, explanation_hi field',
     '[N1] [DEPTH] grammar explanation_hi 0/178',
     '0/178 patterns have populated explanation_hi (Devanagari long-form pedagogical explanation). meaning_hi is at 100% but as the short label only.',
     'A Hindi-medium learner clicking into any pattern sees the long-form pedagogical reasoning in English only. Niche-N1 credibility ceiling.',
     'Translate explanation_en → explanation_hi on top-30 patterns first; mark explanation_provenance: machine_translated until native review.',
     '—', '', '', 'No permission required'),
    ('ISSUE-079', 'Issue', 'MAJOR', 'P2', 'HIGH', 'HIGH',
     'Native-Hindi-speaker review (provenance + trust)',
     'locales/hi.json _meta.review_status + every Hindi-translation field across content data',
     '[N1, N2] [META] Native-review = 0% on every hi-locale field',
     '0% native-reviewed across every hi-locale field. Round-6 provenance-badge UI scaffolds (js/provenance-badge.js) sit dormant.',
     'Without native-Hindi-speaker review, the niche-N1 trust ceiling is bounded. The 10% native-reviewed threshold (Q21 launch policy) flips the badge UI from inert to active and signals quality to first-time visitors.',
     'Recruit 1 Hindi-speaking JA teacher (commission or via Indian university JA program) for a single corpus pass — the top-30 grammar patterns are highest-leverage.',
     'Q33', '', '', 'Requires native-reviewer engagement (Q33)'),
    ('ISSUE-080', 'Issue', 'MAJOR', 'P1', 'HIGH', 'HIGH',
     'Vocab depth — collocations',
     'data/vocab.json — 1041 entries, collocations array',
     '[none] [DEPTH] vocab collocations 0/1041 — largest absolute deficit',
     '0/1041 entries carry a collocations array. The single largest absolute deficit in the corpus (1041 entries below the bar).',
     'N5 vocab learning at depth means collocations (雨 → 雨が降る、雨が止む、雨に濡れる、雨宿り). Without them the vocab page is dictionary-shallow. Tofugu / Tobira surface collocations explicitly.',
     'Author collocations on the top-100 frequency-ranked entries first; LLM-curated with spot-check; not native-blocking.',
     '—', '', '', 'No permission required'),
    ('ISSUE-081', 'Issue', 'MAJOR', 'P1', 'HIGH', 'HIGH',
     'Vocab examples floor',
     'data/vocab.json — 1031/1041 entries below the ≥2-examples floor',
     '[N1, N4] [DEPTH] vocab examples-≥-2 only 10/1041 (1%)',
     'Only 10/1041 entries have ≥2 example sentences. The audit-prompt floor is ≥2.',
     'A single example does not show breadth — learners see one usage and miss the spread of contexts a word fits. Niche-N4 all-in-one gap.',
     'Add 1-2 additional examples per entry, drawn from existing grammar.json examples that already cite the vocab; tooling can auto-cross-link.',
     '—', '', '', 'No permission required'),
    ('ISSUE-082', 'Issue', 'MAJOR', 'P2', 'HIGH', 'MEDIUM',
     'Kanji compound cross-links',
     'data/kanji.json — 0/106 kanji at the ≥5-compound floor',
     '[N4] [DEPTH] kanji examples-≥-5 0/106',
     'Every kanji currently has only 2 compound examples. Audit floor: ≥5 compound vocab cross-links per kanji.',
     'Niche-N4 depth gap — WaniKani shows 5-10 compounds per kanji minimum. Learners using only 2 compounds need a separate kanji app.',
     'Auto-derive compounds by reverse-mapping vocab.json: for each kanji glyph, find all vocab entries containing it; cap at 8.',
     '—', '', '', 'No permission required'),
    ('ISSUE-083', 'Issue', 'MINOR', 'P1', 'LOW', 'LOW',
     'Documentation / namespace stragglers',
     'js/kanji.js line 10; js/learn-grammar.js line 14',
     '[none] [META] Stale meanings_vi/_id/_ne/_zh + explanation_vi/_id/_ne/_zh comments',
     'Comments still reference deprecated locale-suffixed fields post-Phase-3 narrowing. Not a runtime bug — comments only — but a doc-vs-code drift signal.',
     'Niche-N2 (privacy / honesty) demands docs match reality; future contributors reading these comments get a stale model.',
     'Edit the two comments to reference meanings_hi / explanation_hi only.',
     '—', '', '', 'No permission required'),
    ('ISSUE-084', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'MEDIUM',
     'Vocab pitch-accent coverage',
     'data/vocab.json — 44/1041 entries (4.2%)',
     '[N1] [DEPTH] vocab pitch_accent only 44/1041 (4.2%)',
     'Round-7 IMP-087 added pitch_accent on the top-44 highest-frequency entries. Remaining 997 absent.',
     'Niche-N1 trust signal for serious learners (NHK/wadoku-sourced data is publicly available). Indian university JA teachers train on pitch-accent; absence is a pedagogical gap.',
     'Extend pitch_accent authoring to top-200 entries via NHK 日本語発音アクセント新辞典 lookup tooling.',
     '—', '', '', 'No permission required'),
    ('ISSUE-085', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'HIGH',
     'Vocab register tags',
     'data/vocab.json — 4/1041 entries',
     '[N1] [DEPTH] vocab register tags 4/1041 (0.4%)',
     'Only the 4 keigo-chain entries from round-7 have register tags. The remaining ~30 honorific/humble/respectful/casual/formal entries are untagged.',
     'Hindi-L1 learners have 3-tier pronoun politeness but no JA-style verb morphology — explicit register tags help the L1→L2 mapping. Niche-N1 unique-claim lever.',
     'Tag the ~30 known keigo-chain entries (いる⇄いらっしゃる⇄おる, 食べる⇄召し上がる⇄いただく etc.) plus the casual-form catalog.',
     '—', '', '', 'No permission required'),
    ('ISSUE-086', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'MEDIUM',
     'Grammar register tagging',
     'data/grammar.json — 178 patterns',
     '[N1] [DEPTH] grammar register tag 0/178',
     'Patterns are not tagged casual / polite / humble / respectful.',
     'Hindi-medium pedagogy benefits from explicit register marking — Hindi 3-tier pronoun system maps to JA register morphology, but learners only get the mapping if patterns are tagged.',
     'Tag each of 178 patterns with register from {casual, polite, humble, respectful, neutral}. Heuristic: mostly polite + ~10 casual + ~5 keigo.',
     'ISSUE-077', '', '', 'No permission required'),
    ('ISSUE-087', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'MEDIUM',
     'Grammar source citations',
     'data/grammar.json — 27/178 patterns',
     '[N4] [DEPTH] grammar sources 27/178 (151 patterns missing)',
     'Patterns n5-031..n5-188 (151 patterns) lack source citations. Round-7 ISSUE-069 covered top-30 only.',
     'Trust signal for institutional adopters (niche-N3) and serious self-study learners (niche-N4). Tells the learner where this pattern appears in Genki / Minna / Bunpro / JLPT-Sensei.',
     'Author sources arrays on the remaining 151 patterns via cross-reference with KnowledgeBank/sources.md.',
     '—', '', '', 'No permission required'),
    ('ISSUE-088', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'MEDIUM',
     'Grammar contrasts gap',
     'data/grammar.json — 88/178 patterns missing contrasts',
     '[N1] [DEPTH] grammar contrasts only 90/178 (50%)',
     'Only 90/178 carry contrasts non-empty. The mandatory N5 contrast set (は/が, から/ので, も/と, で/に etc.) needs to be present per audit-prompt requirements.',
     'Adjacent-pattern disambiguation is high-leverage for Hindi-L1 learners (postposition→particle mapping is exactly contrast-driven).',
     'Author contrasts on the remaining 88 patterns; cross-validate against the 11-item mandatory contrast list.',
     '—', '', '', 'No permission required'),
    ('ISSUE-089', 'Issue', 'MINOR', 'P4', 'LOW', 'HIGH',
     'Listening voice variety',
     'data/listening.json — single voicevox voice on 18 items; 22 carry no voice metadata',
     '[N1] [DEPTH] listening voice variety = 1 voice (carry-over from round-7 ISSUE-062)',
     'Single-voice listening. Audit floor: ≥4 distinct voices.',
     'Listening realism for niche-N4; not blocking but a polish lever.',
     'Add 3 voicevox speakers (tsumugi, takehiro, kasukabe-tsumugi) and re-render 30+ items at build time.',
     '—', '', '', 'No permission required'),
    ('ISSUE-090', 'Issue', 'MINOR', 'P5', 'LOW', 'HIGH',
     'Native audio recordings',
     'audio/listening/* + audio/reading/*',
     '[N1] [DEPTH] All TTS, no native audio',
     '711 MP3s, all voicevox synthesis. No native-speaker audio.',
     'Best-in-class realism; Renshuu / JapanesePod101 ship native voice talent.',
     'Recruit native speakers via TRANSLATING.md or AUDIO-CONTRIBUTORS.md; record 5-10 listening items as native audio.',
     'Q33', '', '', 'Requires native-review/audio budget (Q33)'),

    # ============================================================
    # IMPROVEMENTS — round 8 depth-first
    # ============================================================
    ('IMP-102', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'MEDIUM',
     'Reading question Hindi explanations',
     'data/reading.json — 94 questions across 45 passages',
     '[N1] [DEPTH] Reading question explanation_hi 0/94',
     'All explanations are English-only. Most are short Japanese passage citations followed by English commentary.',
     'No competitor offers Devanagari Hindi rationales for JA reading questions. Niche-N1 unique-claim.',
     'Author explanation_hi on top-20 question rationales; intro phrase + Devanagari summary; preserve Japanese citations as-is.',
     '—', '', '', 'No permission required'),
    ('IMP-103', 'Improvement', 'IMPROVEMENT', 'P1', 'HIGH', 'MEDIUM',
     'Listening Hindi explanations',
     'data/listening.json — 47 items',
     '[N1] [DEPTH] Listening explanation_hi 0/47',
     'All English-only.',
     'Same as IMP-102 — no competitor offers Hindi rationales for listening items.',
     'Author explanation_hi on the top-20 items; preserve Japanese script_ja as-is.',
     '—', '', '', 'No permission required'),
    ('IMP-104', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Kanji confusable_with extensions',
     'data/kanji.json — 13/106 carry confusable_with',
     '[N4] [DEPTH] Kanji confusable_with 13/106 (12%)',
     '93 kanji not in any confusable cluster. Some have valid additional pairs (言/話/語, 学/字, 来/米).',
     'WaniKani surfaces visually-similar kanji on every entry. The app could match this with 5-10 additional clusters.',
     'Add confusable_with cross-links for ~10 additional clusters; render Compare callout per entry.',
     '—', '', '', 'No permission required'),
    ('IMP-105', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'MEDIUM',
     'Listening transcript line-timing',
     'data/listening.json — 0/47 carry lines:[{text_ja, startMs}]',
     '[N4] [DEPTH] Listening transcript lines[] 0/47 (renderer exists from round-6)',
     'Renderer scaffolded round-6 (IMP-070); data field empty everywhere.',
     'JapanesePod101 / NHK Easy News have line-by-line transcripts with audio sync; one of the most-requested learner features.',
     'Extend tools/build_audio.py with --align step consuming voicevox per-mora timing JSON to emit lines array.',
     'IMP-070 (round-6 carry-over)', '', '', 'No permission required'),
    ('IMP-106', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Reading paragraph summary',
     'data/reading.json — 0/45 carry summary',
     '[N4] [DEPTH] Reading summary 0/45',
     'No paragraph-level summary on any passage.',
     'Tofugu / NHK Easy News show summaries on long passages; useful for revision (the user-requested active-recall list-tile pattern).',
     'Author one-sentence summary per passage; rendered above the passage on detail page; collapsed by default.',
     '—', '', '', 'No permission required'),
    ('IMP-107', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'MEDIUM',
     'Reading vocab preview',
     'data/reading.json — 0/45 carry vocab_preview',
     '[N4] [DEPTH] Reading vocab_preview 0/45',
     'No pre-reading vocab list.',
     'Most reading apps surface a vocab preview before the passage to lower the entry-friction.',
     'Auto-derive from vocab_used field already on each passage (top-5 unfamiliar words); render as click-to-expand chip strip.',
     '—', '', '', 'No permission required'),
    ('IMP-108', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'LOW',
     'Kanji recognition_priority',
     'data/kanji.json — 0/106 carry recognition_priority',
     '[N4] [DEPTH] Kanji recognition_priority 0/106',
     'No prioritization signal — learners process kanji in arbitrary order.',
     'WaniKani has explicit recognition tiers + lesson order.',
     'Add recognition_priority: 1|2|3 per kanji based on first-Genki-appearance + JLPT.jp frequency.',
     '—', '', '', 'No permission required'),
    ('IMP-109', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'Kanji stroke-order common mistakes',
     'data/kanji.json — 0/106 carry stroke_order_mistakes',
     '[N4] [DEPTH] Kanji stroke_order_mistakes 0/106',
     'No stroke-order trap notes (e.g., 田 horizontal vs vertical order, 力 direction, 必 order).',
     'WaniKani has these on a small subset; few apps surface them explicitly.',
     'Author for the 15-20 N5 kanji with known textbook stroke-order traps.',
     '—', '', '', 'No permission required'),
    ('IMP-110', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'Indian numeral grouping',
     'js/home.js fmt(n) uses Intl.NumberFormat(en-US)',
     '[N1] [DEPTH] Number formatting Western-only',
     'Counts render as Western 1,003 / 1,041 everywhere. Hindi locale still uses US grouping.',
     'Wikipedia / Indian e-commerce apps use Indian grouping (१,०४१ in Devanagari numerals or 1,041 with comma at lakh boundary).',
     'Use Intl.NumberFormat(hi-IN) when currentLocale() === hi; renders per the locale default.',
     '—', '', '', 'No permission required'),
    ('IMP-111', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'localStorage namespace doc-vs-code',
     'README.md (no namespace mentioned), PRIVACY.md (mentioned), js/storage.js (uses jlpt-n5-tutor:*)',
     '[none] [META] README.md does not mention storage namespace',
     'README does not mention the namespace at all. PRIVACY.md is correct. JA-37 invariant currently passes by checking only PRIVACY.md.',
     'Open-source app docs mention storage namespace explicitly; helps niche-N3 institutional adopters audit data residency.',
     'Add a Storage subsection to README.md naming the namespace. Extend JA-37 to also check README.md if mentioned.',
     '—', '', '', 'No permission required'),
    ('IMP-112', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'Listening cultural context note',
     'data/listening.json — 0/47 carry cultural_context',
     '[N1] [DEPTH] Listening cultural_context 0/47',
     'No cultural-context callouts (e.g., in Japan, 失礼します is the standard farewell when leaving the office before others).',
     'Tofugu / WaniKani sprinkle cultural notes; few JLPT-specific apps do.',
     'Author 1-paragraph cultural context for the 15-20 items where Japan-specific assumptions are load-bearing.',
     '—', '', '', 'No permission required'),
    ('IMP-113', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'Indian-discoverability strategy stub',
     'Repo-level docs / GitHub repo description / topics',
     '[N4] [META] No india / hindi / bharat topics',
     'Repo description + 14 topics include multilingual but not india / hindi / bharat. No Hindi YouTube / Telegram presence.',
     'Established Indian-market study apps have Hindi YouTube presence; Yoisho Academy uses Telegram. This is a Q35-blocked decision.',
     'Add india / hindi / bharat topics to repo. Q35 in Section 6 is the strategy gate.',
     'Q35', '', '', 'Requires Indian-discoverability strategy decision (Q35)'),
    ('IMP-114', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'HIGH',
     'Per-locale README',
     'Repo root + N5/',
     '[N3] [META] README.hi.md absent',
     'README is English-only. No README.hi.md.',
     'Wikipedia per-language portals; OSS projects often ship README.<lc>.md.',
     'Author N5/README.hi.md once the L1 notes + explanation_hi work has momentum.',
     'ISSUE-077, ISSUE-078', '', '', 'No permission required'),
]

# (Q-id, Title, Context, What we need decided, Decision blank)
QUESTIONS = [
    ('Q33', 'Native-Hindi-review budget',
     'Hindi content is 100% llm_curated. The 10% threshold for the round-6 provenance badge to fire requires native review. Two paths: commission a Hindi-speaking JA teacher (cost), or recruit volunteers via Indian university JA programs (slow). This blocks ISSUE-079 + IMP-114 + every Section-3 item with native-review dependency.',
     'Which path, and what cycle-budget? Affects ISSUE-079 + ISSUE-090 (native audio) + IMP-114.',
     ''),
    ('Q34', 'Niche-N1 commitment depth (go-to-market)',
     'The locale transition is shipped (Phase 1-10 complete). Open question: does the product owner double down on niche N1 — committing to native-Hindi-speaker review, India-targeted discoverability (Hindi YouTube / Telegram / Quora), Tier-2/3-city Indian learner UX research — or treat Hindi as a quality side-bet while leaning on niche N2 for global English-speaking reach?',
     'Shapes prioritization of ISSUE-077..079 vs IMP-104..109. Concrete commitment level.',
     ''),
    ('Q35', 'Indian discoverability channel',
     'GitHub Pages alone will not reach Tier-2/3 Indian learners. Hindi YouTube? Telegram groups? Indian Japanese-language teacher partnerships? Show HN is unlikely to reach this audience.',
     'Which 1-2 channels to invest in for the first wave?',
     ''),
    ('Q36', 'Vocab depth-batch order',
     'With 1041 entries and multiple deficit dimensions (collocations 0%, examples-≥-2 1%, pitch-accent 4%, counter 0.4%, register 0.4%), the highest-leverage batch order is non-obvious.',
     'Author one dimension across all 1041 entries (full collocations pass), or all dimensions across the top-100 frequency-ranked entries (depth-per-entry)?',
     ''),
    ('Q37', 'Reading/listening sentence-by-sentence footnote scope',
     'IMP-106/107 (summary + vocab_preview) are cheap. Sentence-by-sentence grammar footnote layer (per the audit-prompt depth dimension) is a substantial UI + content lift.',
     'In scope for next cycle, or defer behind ISSUE-077..079 native-review work?',
     ''),
    ('Q38', 'Mock-paper restructure phase 2/3 reactivation',
     'Round-7 ISSUE-059 phase 1 (mondai backfill) shipped; phase 2 (chokai papers from listening.json) + phase 3 (per-section build_papers.py reweighting + 25/50/30-min timing) sit deferred.',
     'Re-prioritize for next cycle, or defer until Q33/Q34 unlock?',
     ''),
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb['Items']
    existing_ids = set()
    for row in ws.iter_rows(min_row=5, values_only=False):
        if row[0].value:
            existing_ids.add(row[0].value)

    n_added = 0
    for entry in ITEMS:
        rid = entry[0]
        if rid in existing_ids:
            print(f'  skip {rid} (exists)')
            continue
        new_row = ws.max_row + 1
        for ci, val in enumerate(entry, start=1):
            ws.cell(row=new_row, column=ci, value=val)
        n_added += 1

    new_max = ws.max_row
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and 'Fix' in dv.formula1:
                dv.sqref = f'N5:N{new_max}'
            elif dv.formula1 and 'Allow' in dv.formula1:
                dv.sqref = f'P5:P{new_max}'

    print(f'\nItems: added {n_added}; sheet now {ws.max_row} rows.')

    ws2 = wb['Questions']
    existing_qids = set()
    for row in ws2.iter_rows(min_row=1, values_only=False):
        if row and row[0].value:
            existing_qids.add(row[0].value)

    n_q_added = 0
    for q in QUESTIONS:
        if q[0] in existing_qids:
            print(f'  skip {q[0]} (exists)')
            continue
        new_row = ws2.max_row + 1
        for ci, val in enumerate(q, start=1):
            ws2.cell(row=new_row, column=ci, value=val)
        n_q_added += 1

    print(f'Questions: added {n_q_added}; sheet now {ws2.max_row} rows.')

    wb.save(XLSX)
    print('\nSaved.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
