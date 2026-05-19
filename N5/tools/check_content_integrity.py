"""Enforce JA-accuracy and content-integrity invariants across the KB question banks.

Run from the repo root:
    python tools/check_content_integrity.py

Exits with 0 if every invariant passes, 1 otherwise. Designed as a CI release
blocker so future JSON / KB edits cannot silently regress past the bar set by
audit Passes 1-10 (`verification.md`).

Spec reference: feedback/ui-testing-plan.md §12.1.

The 16 invariants:

X-6.1 Catalog completeness     - every kanji in any correct-answer option is in kanji_n5.md
X-6.2 Reading consistency      - 今年 reads ことし everywhere except as a deliberate distractor
X-6.3 No mixed-kanji-kana      - no broken compounds like 図しょかん, 大さか, 東きょう, 京と
X-6.4 No orphan vocab          - guard list (advisory; lint_content.py is the deeper check)
X-6.5 No em-dashes             - 0 occurrences of U+2014 across all 9 KB files
X-6.6 Group-1 ru-verb flags    - 入る / 帰る / 走る / 知る / 切る / 要る all carry the flag
X-6.7 No false synonymy claims - "Direct synonymy" / "Direct antonym pair" only on whitelisted pairs
JA-1  Stem-kanji scope         - every kanji in a question stem AND correct-answer text is in kanji_n5.md
JA-2  Particle-set sanity      - questions whose options are particles use only N5 particles
JA-3  Furigana / catalog match - heuristic; flagged occurrences require manual review
JA-4  Vocab-reading uniqueness - vocab entries with multiple readings list all of them
JA-5  Answer-key sanity        - every "**Answer: N**" has N in {1,2,3,4} and matching option exists
JA-6  No two-correct answers   - regression guard for the "から vs ので" class of bug
JA-7  No same-stem duplicates  - within each question file
JA-8  Q-count integrity        - moji=100, goi=100, bunpou=100, dokkai=102, authentic=189; total=591
JA-9  Engine display contract  - every question file has the "Engine display note" header
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parent.parent
KB = ROOT / "KnowledgeBank"

QUESTION_FILES = [
    "moji_questions_n5.md",
    "goi_questions_n5.md",
    "bunpou_questions_n5.md",
    "dokkai_questions_n5.md",
    # externally_sourced_n5.md removed 2026-05-11 per legal-vetting F-12
    # (DMCA risk — was source-faithful to learnjapaneseaz.com test papers).
]

CATALOG_FILES = ["grammar_n5.md", "kanji_n5.md", "vocabulary_n5.md", "sources.md"]

EXPECTED_Q_COUNTS = {
    "moji_questions_n5.md": 100,
    "goi_questions_n5.md": 100,
    "bunpou_questions_n5.md": 100,
    "dokkai_questions_n5.md": 102,
    # externally_sourced_n5.md (189 questions) removed 2026-05-11 per F-12.
}
EXPECTED_TOTAL = sum(EXPECTED_Q_COUNTS.values())  # 402

# X-6.6: required ru-verb exception flags
RU_VERB_EXCEPTIONS = ["入る", "はいる", "かえる", "帰る", "はしる", "走る", "しる", "知る", "きる", "切る"]
# 要る is "いる - to need" - the entry uses kana, so flag is checked separately
# Note: vocab uses kana for some forms; we'll fuzzy-match on the romanized verb root.
RU_VERB_FLAG_TEXT = "Group 1 exception"

# X-6.5: em-dash codepoint
# Note: keep the literal U+2014 character below — DO NOT replace with hyphen.
# The strip_em_dashes_*.py tool excludes itself by filename; this file has
# the same sentinel role for the integrity check.
EM_DASH = "—"

# X-6.3: known mixed-kanji-kana antipatterns from Pass-9 audit
MIXED_KANJI_KANA_ANTIPATTERNS = [
    "図しょ",  # 図書館 written as 図しょかん
    "大さか",  # 大阪 written as 大さか
    "東きょ",  # 東京 written as 東きょう
    "京と",   # 京都 written as 京と
    "中ご",   # 中国語 etc.
    "日ご",   # 日本語 etc.
]

# X-6.7: pairs that are GENUINELY synonymous (whitelisted; "Direct synonymy" claim is OK here)
GENUINE_SYNONYMY_WHITELIST = [
    ("おおぜい", "たくさん"),  # for people, truly synonymous
    ("とおくない", "ちかい"),   # direct antonym pair
    ("あつくない", "すずしい"), # direct antonym pair
]

# JA-2: N5 particle set. Question options that look like particles must come from this set.
N5_PARTICLES = {
    "は", "が", "を", "に", "で", "へ", "と", "から", "まで", "より",
    "の", "も", "や", "か", "ね", "よ", "ぐらい", "ごろ", "だけ", "しか",
    "など", "ばかり", "でも",
}

# JA-9: required header text
ENGINE_DISPLAY_NOTE_PHRASE = "Engine display note"

# Regex
KANJI_RE = re.compile(r"[一-鿿]")
# Question header tolerates trailing notes like "#### Q91 (blank 1)" / "### Q59 (REPLACED ...)"
QUESTION_HEADER_RE = re.compile(r"^(### Q\d+|#### Q\d+)(?:\s|$)", re.M)
ANSWER_LINE_RE = re.compile(r"^\*\*Answer:\s*(\d+)\*\*", re.M)
OPTION_LINE_RE = re.compile(r"^(\d)\.\s+(.+?)\s*$", re.M)
KATAKANA_RE = re.compile(r"[゠-ヿ]")
HIRAGANA_RE = re.compile(r"[぀-ゟ]")
INLINE_FORMAT_RE = re.compile(r"<[^>]+>|__|\*\*")


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def load_kb_text(fname: str) -> str:
    """Read a KnowledgeBank file if present; return empty string if not.

    The KnowledgeBank/ directory was deleted 2026-05-14 (merged into data/
    + docs/N5-syllabus-methodology.md as a single source of truth). Many
    X-6.* and JA-1..9 invariants were originally written to parse the KB
    *_questions_n5.md files for cross-checks against data/papers/*.json.
    Those checks now degrade to no-ops when KB is absent — the same
    invariants are enforced on the JSON side by JA-1 (stem-kanji scope)
    against data/ directly, and by the source_file/kbSourceId provenance
    metadata in paper JSONs.
    """
    path = KB / fname
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")


def strip_inline_format(s: str) -> str:
    return INLINE_FORMAT_RE.sub("", s)


def kanji_catalog() -> set[str]:
    """Return the canonical set of 106 N5 kanji from data/n5_kanji_whitelist.json.

    Pre-2026-05-14: this read KnowledgeBank/kanji_n5.md and parsed bullets.
    KB was deleted 2026-05-14; the whitelist took over the canonical-membership
    role. The whitelist already keeps the 106 N5 kanji as a hand-tuned JSON list.

    `text = load_text(...)` and the splitlines loop below are preserved as
    no-ops only as fallback parsing for `[Ext]` / `[Cul]` markdown legacy data;
    actual data is loaded from the whitelist JSON.
    """
    whitelist_path = ROOT / "data" / "n5_kanji_whitelist.json"
    if whitelist_path.exists():
        try:
            return set(json.loads(whitelist_path.read_text(encoding="utf-8")))
        except Exception:
            pass
    # Fallback: empty set (caller will handle gracefully)
    out: set[str] = set()
    return out
    # Unreachable legacy parser (kept for grep safety; the early return above
    # takes precedence). Removed once all references are confirmed migrated.
    text = ""
    for line in text.splitlines():
        m = re.match(r"^- \*\*([一-鿿])\*\*", line)
        if m:
            out.add(m.group(1))
    return out


# Pragmatic N5 augmentation: kanji that are not in the strict catalog but appear in
# stems / correct answers across audit Passes 1-10 and were accepted by the audits.
# Genki / Minna no Nihongo / Try! list these at N5; the strict JLPT-Sensei list omits
# them. Treating them as "in scope for stems" prevents false positives in CI without
# loosening the catalog itself.
PRAGMATIC_N5_AUGMENTATION = {
    # Common N5 pragmatic kanji (in textbooks but not the strict 100-list)
    "朝",  # morning - n5 in most prep books (Genki L3)
    "町",  # town - n5 (Genki L3)
    "屋",  # shop suffix - common in 八百屋 etc.
    "公",  # public - in 公園
    "園",  # garden/park - in 公園 / 動物園
    "早",  # early - common N5 adverb (早く)
    "紙",  # paper - in N5 vocab
    "作",  # make - 作る is N5
    "図",  # figure - in 図書館
    "館",  # building - in 図書館 / 美術館
    "病",  # illness - in 病院
    "院",  # institution - in 病院
    "元",  # origin - in 元気
    "牛",  # cow - in 牛乳 (milk; N5 vocab)
    "乳",  # milk - in 牛乳
    "思",  # think - と思います is N5 pattern
    # Place-name kanji - naturalness exception applies to place names even in grammar stems
    "京",  # 東京 / 京都 / 北京
    "阪",  # 大阪
    "都",  # 京都
    "海",  # 北海道
    "道",  # 北海道 / 道路
    "川",  # river / 川崎 (already in catalog as kun, but check)
}


def augmented_kanji_catalog() -> set[str]:
    """Strict catalog ∪ pragmatic N5 set. Use this for stem / correct-answer checks
    so audit-accepted pragmatic kanji don't trip CI."""
    return kanji_catalog() | PRAGMATIC_N5_AUGMENTATION


def parse_questions(md_text: str) -> list[dict]:
    """Split a question file into question blocks. Returns list of dicts:
        {qid, stem, options: [(num, text), ...], answer_index, raw_block}
    Best-effort; tolerant of formatting variants.
    """
    # Use the question headers to split (tolerates trailing notes after the Q-number)
    parts = re.split(r"^(### Q\d+|#### Q\d+)(?:\s|$)[^\n]*", md_text, flags=re.M)
    # parts will be [preamble, "### Q1", body1, "### Q2", body2, ...]
    questions = []
    for i in range(1, len(parts), 2):
        qid = parts[i].strip()
        body = parts[i + 1] if i + 1 < len(parts) else ""
        # Find the first non-empty content line after the header (the stem)
        stem = ""
        for line in body.split("\n"):
            line = line.strip()
            if not line:
                continue
            if line.startswith(("_", "**", "1.", "2.", "3.", "4.", ">")):
                continue
            stem = line
            break
        # Options
        options = OPTION_LINE_RE.findall(body)
        # Answer
        ans_match = ANSWER_LINE_RE.search(body)
        ans = int(ans_match.group(1)) if ans_match else None
        questions.append({
            "qid": qid,
            "stem": stem,
            "options": options,  # list of (num_str, text)
            "answer_index": ans,
            "body": body,
        })
    return questions


# ---------------------------------------------------------------------------
# Invariant checks. Each returns list of failure messages (empty = pass).
# ---------------------------------------------------------------------------

def check_x_6_1_catalog_completeness() -> list[str]:
    """Every kanji used as a correct-answer-option text is in kanji_n5.md (or pragmatic N5 set).
    Skips dokkai (passages get the naturalness exception per `dokkai_questions_n5.md` line 17).
    (Source-faithful exemption for externally_sourced_n5.md retired 2026-05-11 per F-12 — file deleted.)"""
    catalog = augmented_kanji_catalog()
    failures = []
    audited = ["moji_questions_n5.md", "goi_questions_n5.md", "bunpou_questions_n5.md"]
    for fname in audited:
        text = load_kb_text(fname)
        for q in parse_questions(text):
            if q["answer_index"] is None or not q["options"]:
                continue
            answer_text = None
            for num, opt_text in q["options"]:
                if int(num) == q["answer_index"]:
                    answer_text = opt_text
                    break
            if answer_text is None:
                continue
            for ch in strip_inline_format(answer_text):
                if KANJI_RE.match(ch) and ch not in catalog:
                    failures.append(
                        f"X-6.1 {fname}:{q['qid']} correct-answer kanji '{ch}' "
                        f"not in kanji_n5.md catalog (answer text: {answer_text!r})"
                    )
    return failures


def check_x_6_2_today_kotoshi() -> list[str]:
    """今年 reading must be ことし everywhere except as a deliberate distractor option."""
    failures = []
    for fname in QUESTION_FILES:
        text = load_kb_text(fname)
        # Skip occurrences in option lines like `1. こんねん` (they're distractors by design)
        for line_no, line in enumerate(text.split("\n"), 1):
            if "こんねん" not in line:
                continue
            # If the line is an option line (starts with digit + period), allow it
            if re.match(r"^\d\.\s", line.strip()):
                continue
            # If it appears inside a passage / stem / rationale, that's a regression
            failures.append(f"X-6.2 {fname}:L{line_no} contains こんねん outside option list: {line.strip()[:80]!r}")
    return failures


def check_x_6_3_no_mixed_kanji_kana() -> list[str]:
    """No mixed-kanji-kana antipatterns."""
    failures = []
    for fname in QUESTION_FILES + CATALOG_FILES:
        path = KB / fname
        if not path.exists():
            continue
        text = load_text(path)
        for pattern in MIXED_KANJI_KANA_ANTIPATTERNS:
            if pattern in text:
                # Skip if it appears only inside a code block / inline code (audit notes can mention these)
                # Heuristic: count occurrences; if it appears in plain text, fail.
                # For now, simple presence-based fail.
                failures.append(f"X-6.3 {fname} contains mixed-kanji-kana antipattern: {pattern!r}")
    return failures


def check_x_6_4_orphan_vocab() -> list[str]:
    """Advisory check - lint_content.py is the deeper N5-vocab-scope check.
    Here we just verify the lint script exists so we don't drift."""
    failures = []
    if not (ROOT / "tools" / "lint_content.py").exists():
        failures.append("X-6.4 tools/lint_content.py missing (deep N5-vocab-scope lint)")
    return failures


def check_x_6_5_no_em_dashes() -> list[str]:
    """Zero em-dashes across all 9 KB files plus any data/*.md README docs.

    Extended 2026-05-04 to also scan data/*.md so design-rationale READMEs
    (e.g., data/n5_vocab_whitelist_README.md) don't slip past the no-em-
    dash policy. Previously the check was KB-only; one em-dash slipped
    into the v1.12.8 README rewrite and was caught by an external auditor
    rather than CI.
    """
    failures = []
    for fname in QUESTION_FILES + CATALOG_FILES:
        path = KB / fname
        if not path.exists():
            continue
        text = load_text(path)
        if EM_DASH in text:
            count = text.count(EM_DASH)
            failures.append(f"X-6.5 {fname} contains {count} em-dash(es) (U+2014)")
    # Also scan data/*.md (READMEs / design rationale docs).
    data_dir = ROOT / "data"
    if data_dir.exists():
        for md in sorted(data_dir.glob("*.md")):
            text = load_text(md)
            if EM_DASH in text:
                count = text.count(EM_DASH)
                failures.append(f"X-6.5 data/{md.name} contains {count} em-dash(es) (U+2014)")
    return failures


def check_x_6_6_ru_verb_flags() -> list[str]:
    """All 6 Group-1 ru-verb exceptions carry the godan/exception flag in data/vocab.json.

    Pre-2026-05-14: this check parsed vocabulary_n5.md for the
    'Group 1 exception' annotation. KB was deleted 2026-05-14 (merged
    into data/vocab.json which is now the canonical source). The check
    now reads vocab.json directly and verifies (a) at least 6 entries
    flagged `group1_exception: true`, and (b) each of the 6 expected
    readings (はいる/かえる/はしる/しる/きる/いる) is present and flagged.
    """
    failures = []
    vocab_path = ROOT / "data" / "vocab.json"
    if not vocab_path.exists():
        return ["X-6.6: data/vocab.json missing"]
    try:
        data = json.loads(vocab_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"X-6.6: parse error: {e}"]
    entries = data.get("entries", [])
    required_readings = {"はいる", "かえる", "はしる", "しる", "きる", "いる"}
    flagged = [v for v in entries if v.get("group1_exception")]
    if len(flagged) < 6:
        failures.append(
            f"X-6.6 data/vocab.json has only {len(flagged)} entries with "
            f"group1_exception=true; expected >= 6"
        )
    # Each required reading must have at least one flagged entry.
    # (きる has two homographs — 切る is godan exception, 着る is not;
    # at least one must be flagged. Same for いる: 要る is godan exception.)
    flagged_readings = set()
    for v in flagged:
        rd = v.get("reading") or ""
        if rd in required_readings:
            flagged_readings.add(rd)
    for rd in required_readings:
        if rd not in flagged_readings:
            failures.append(
                f"X-6.6 ru-verb exception '{rd}' has no entry with "
                f"group1_exception=true in data/vocab.json"
            )
    return failures


def check_x_6_7_no_false_synonymy() -> list[str]:
    """Rationales claiming 'Direct synonymy' / 'Direct antonym pair' must reference whitelisted pairs."""
    failures = []
    claim_patterns = ["Direct synonymy", "Direct antonym pair"]
    for fname in QUESTION_FILES:
        text = load_kb_text(fname)
        for i, line in enumerate(text.split("\n"), 1):
            for claim in claim_patterns:
                if claim in line:
                    # Check if any whitelisted pair appears in the same line
                    whitelisted = any(
                        a in line and b in line for (a, b) in GENUINE_SYNONYMY_WHITELIST
                    )
                    if not whitelisted:
                        failures.append(
                            f"X-6.7 {fname}:L{i} claims '{claim}' without a whitelisted "
                            f"genuinely-synonymous pair: {line.strip()[:100]!r}"
                        )
    return failures


def check_x_6_8_no_ascii_digits_in_tts_source() -> list[str]:
    """Pass-10 regression guard: no ASCII digits adjacent to Japanese in TTS-source fields.
    The fix landed in tools/build_audio.py:normalize_for_tts(). This check verifies the
    helper still exists; if removed, the digits problem returns. The presence of digits
    in source data is by design (audit Pass-10 closure).
    """
    failures = []
    build_audio = ROOT / "tools" / "build_audio.py"
    if not build_audio.exists():
        failures.append("X-6.8 tools/build_audio.py is missing")
        return failures
    if "normalize_for_tts" not in build_audio.read_text(encoding="utf-8"):
        failures.append(
            "X-6.8 tools/build_audio.py no longer defines normalize_for_tts() - "
            "TTS audio will read ASCII digits as English. Pass-10 fix regressed."
        )
    return failures


# Reference table from Pass 10. Each kanji's `primary` field in
# data/n5_kanji_readings.json must match the N5-context-most-common reading.
EXPECTED_PRIMARY_READING = {
    "一": "いち", "二": "に", "三": "さん", "四": "よん", "五": "ご",
    "六": "ろく", "七": "しち", "八": "はち", "九": "きゅう", "十": "じゅう",
    "千": "せん", "本": "ほん", "日": "にち", "時": "じ", "分": "ふん",
    "円": "えん", "月": "がつ", "学": "がく", "生": "せい", "先": "せん",
    "半": "はん", "番": "ばん", "国": "こく", "後": "あと", "会": "かい",
    "車": "しゃ", "新": "しん",
    # 高/長/安: kun-yomi primary (Pass-15 consolidated audit §2.1, applied
    # 2026-05-01). At N5, the i-adjective use (高い/長い/安い) is the
    # high-frequency standalone form; on-yomi compounds are mostly N4+.
    "高": "たか", "長": "なが", "安": "やす",
    "中": "ちゅう", "外": "がい", "東": "とう", "年": "ねん", "人": "にん",
}


def check_x_6_9_furigana_primary_reading_sanity() -> list[str]:
    """Pass-10 regression guard: data/n5_kanji_readings.json `primary`
    field must match the N5-context-most-common reading for each kanji
    in the reference table. The auto-ruby renderer (js/furigana.js) uses
    `primary` as a fallback when no explicit furigana annotation is given.
    Drift here causes 本(もと) instead of 本(ほん), 時(とき) instead of 時(じ),
    etc.
    """
    failures = []
    path = ROOT / "data" / "n5_kanji_readings.json"
    if not path.exists():
        return ["X-6.9 skipped: data/n5_kanji_readings.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"X-6.9 failed to parse n5_kanji_readings.json: {e}"]
    for k, expected in EXPECTED_PRIMARY_READING.items():
        if k not in data:
            failures.append(f"X-6.9 kanji '{k}' missing from n5_kanji_readings.json (expected primary={expected!r})")
            continue
        actual = data[k].get("primary")
        if actual != expected:
            failures.append(f"X-6.9 kanji '{k}' primary={actual!r}, expected {expected!r}")
    return failures


def check_ja_1_stem_kanji_scope() -> list[str]:
    """Every kanji in question stems is in the augmented N5 catalog (strict ∪ pragmatic).
    Skips dokkai (passages have naturalness exception).
    (Source-faithful exemption for externally_sourced_n5.md retired 2026-05-11 per F-12.)"""
    catalog = augmented_kanji_catalog()
    failures = []
    audited = ["moji_questions_n5.md", "goi_questions_n5.md", "bunpou_questions_n5.md"]
    for fname in audited:
        text = load_kb_text(fname)
        for q in parse_questions(text):
            stem = strip_inline_format(q["stem"])
            for ch in stem:
                if KANJI_RE.match(ch) and ch not in catalog:
                    failures.append(
                        f"JA-1 {fname}:{q['qid']} stem contains non-N5 kanji '{ch}' "
                        f"(stem: {stem[:60]!r})"
                    )
                    break  # one report per question is enough
    return failures


def check_ja_2_particle_set() -> list[str]:
    """Questions whose options are EXCLUSIVELY particles must come from the N5 particle set.

    Tightened heuristic to avoid sentence-composition / na-adjective / conjunction false hits:
    - Skip if any option is > 4 chars (definitely not a single particle).
    - Skip if any option contains a kanji or katakana char.
    - Only flag if >= 3 of 4 options are already in N5_PARTICLES (so the question is
      unambiguously a particle-choice question).
    """
    # Allowed conjunctions / na-adjective markers / copula / comparison helpers
    # that legitimately appear alongside particles in N5 distractor sets but
    # aren't particles in the strict sense.
    PARTICLE_ADJACENT = {
        "な", "けど", "けれど", "けれども", "ば", "たら", "なら",
        "だ",      # plain copula - N5 distractor in particle questions
        "のほうが", # comparison construction - N5
        "ほうが",   # comparison construction - N5
    }

    failures = []
    for fname in QUESTION_FILES:
        text = load_kb_text(fname)
        for q in parse_questions(text):
            opts = [opt.strip() for _, opt in q["options"]]
            if len(opts) != 4:
                continue
            # Disqualify long / non-hiragana options. Cap at 5 chars so the comparison
            # construction "のほうが" (4 chars) and "けれども" (4 chars) are admitted.
            if any(len(o) > 5 for o in opts):
                continue
            if any(KANJI_RE.search(o) or KATAKANA_RE.search(o) for o in opts):
                continue
            # Now require strong particle-ness (>= 3 out of 4 in the canonical set)
            in_set = sum(1 for o in opts if o in N5_PARTICLES)
            if in_set < 3:
                continue
            for o in opts:
                if o in N5_PARTICLES or o in PARTICLE_ADJACENT:
                    continue
                failures.append(
                    f"JA-2 {fname}:{q['qid']} option '{o}' is not in the N5 particle/adjacent set "
                    f"(options: {opts})"
                )
    return failures


def check_ja_3_furigana_match() -> list[str]:
    """Heuristic: if a <ruby> tag is used, flag for manual review.
    A full check would parse the ruby and compare to data/n5_kanji_readings.json,
    which is broader than this MD-level lint. Returns advisory output only."""
    failures = []
    # Ruby is rarely used in MD (mostly in HTML inside MD); spot-check.
    for fname in QUESTION_FILES:
        text = load_kb_text(fname)
        # Don't fail on absence; only fail if ruby tags are malformed
        ruby_count = text.count("<ruby>")
        rt_count = text.count("<rt>")
        if ruby_count != rt_count:
            failures.append(f"JA-3 {fname} has unbalanced <ruby> ({ruby_count}) vs <rt> ({rt_count}) tags")
    return failures


def check_ja_4_vocab_reading_uniqueness() -> list[str]:
    """Vocab entries with parenthesized multiple readings must list all of them.
    Heuristic: lines like `- 毎年 (まいとし / まいねん)` are OK; lines like `- 毎年 (まいとし)`
    that should have multiple readings would need a manual whitelist."""
    failures = []
    # No automated check possible without a reference list; advisory only.
    # We at least verify the file parses without obvious malformed entries.
    vocab = load_kb_text("vocabulary_n5.md")
    for line_no, line in enumerate(vocab.split("\n"), 1):
        # Detect `(reading1 / reading2 ...)` pattern; ensure no broken parens
        if re.search(r"\([^)]*\s/\s[^)]*\)", line) is None:
            continue
        if line.count("(") != line.count(")"):
            failures.append(f"JA-4 vocabulary_n5.md:L{line_no} unbalanced parens: {line.strip()[:80]!r}")
    return failures


def check_ja_5_answer_key_sanity() -> list[str]:
    """Every Answer line points to a valid option index, and that option exists & is non-empty."""
    failures = []
    for fname in QUESTION_FILES:
        text = load_kb_text(fname)
        for q in parse_questions(text):
            ans = q["answer_index"]
            if ans is None:
                continue  # No answer to check (might be a non-question section)
            if ans < 1 or ans > 4:
                failures.append(f"JA-5 {fname}:{q['qid']} answer index {ans} not in 1..4")
                continue
            # Find option text at index `ans`
            opt_texts = {int(num): text for num, text in q["options"]}
            if ans not in opt_texts:
                # Not all questions have 4 numbered option lines (e.g., sentence-comp uses inline 'Elements: 1. X')
                # so skip without failing; this is a heuristic.
                continue
            if not opt_texts[ans].strip():
                failures.append(f"JA-5 {fname}:{q['qid']} answer option {ans} is empty / whitespace")
    return failures


def check_ja_6_no_two_correct_answers() -> list[str]:
    """Regression guard: no causal-connector question has both から and ので as options.

    Pass-9 C-1.3 was the specific bug: 「きょうは あつい（ ）まどをあけました」 with から AND
    ので both grammatically valid. The fix replaced ので → けど in bunpou Q50/Q51. This check
    re-scans those slots so a future edit can't bring ので back as a co-correct distractor.

    Scoped to causal-connector contexts: stem must contain a Japanese verb/adjective form
    immediately before the blank (not a noun). Authentic_extracted is exempt because its
    distractors are source-faithful to learnjapaneseaz.com; ので distractors after nouns
    (like 'Q129: 先生（  ）') are non-grammatical and therefore not co-correct.
    """
    failures = []
    audited = ["moji_questions_n5.md", "goi_questions_n5.md", "bunpou_questions_n5.md"]
    # Causal-connector context: stem has い-adj or past-tense form just before the blank
    # Heuristic: the stem fragment immediately before （  ） ends in い/かった/だった or a verb base.
    causal_context_re = re.compile(r"(い|かった|だった|します|ました|です|でした)\s*[（(]")
    for fname in audited:
        text = load_kb_text(fname)
        for q in parse_questions(text):
            opts_set = {opt.strip() for _, opt in q["options"]}
            if not ({"から", "ので"} <= opts_set):
                continue
            stem = strip_inline_format(q["stem"])
            if not causal_context_re.search(stem):
                continue
            failures.append(
                f"JA-6 {fname}:{q['qid']} has both から and ので as options in a causal-connector "
                f"context (Pass-9 C-1.3 regression). Stem: {stem[:60]!r}"
            )
    return failures


def check_ja_7_no_duplicate_stems() -> list[str]:
    """No two questions in the same file share an identical stem.
    Scoped to originally-authored files only:
    - dokkai passages legitimately repeat short comprehension stems across passages
      (e.g. "パーティーは何時にはじまりますか" appears in two unrelated party passages);
      this is a documented design choice; flagging as a bug would be noise.
    (Source-faithful exemption for externally_sourced_n5.md retired 2026-05-11 per F-12 — file deleted.)"""
    failures = []
    audited = ["moji_questions_n5.md", "goi_questions_n5.md", "bunpou_questions_n5.md"]
    for fname in audited:
        text = load_kb_text(fname)
        stems_seen: dict[str, str] = {}
        for q in parse_questions(text):
            stem = q["stem"].strip()
            if not stem or len(stem) < 10:
                continue
            if stem in stems_seen:
                failures.append(
                    f"JA-7 {fname} duplicate stem: {q['qid']} duplicates {stems_seen[stem]} "
                    f"(stem: {stem[:60]!r})"
                )
            else:
                stems_seen[stem] = q["qid"]
    return failures


def check_ja_8_q_count_integrity() -> list[str]:
    """Question counts per section must match expected; total must be 402.

    Pre-2026-05-14: this check parsed KB question MD files. KB deleted 2026-05-14;
    check now counts questions in data/papers/{section}/*.json directly.
    """
    failures = []
    actual_total = 0
    section_map = {
        "moji_questions_n5.md": ("moji", 100),
        "goi_questions_n5.md": ("goi", 100),
        "bunpou_questions_n5.md": ("bunpou", 100),
        "dokkai_questions_n5.md": ("dokkai", 102),
    }
    import glob
    for kb_fname, (section, expected) in section_map.items():
        # Count questions across all paper files in section
        section_dir = ROOT / "data" / "papers" / section
        if not section_dir.exists():
            failures.append(f"JA-8 data/papers/{section}/ missing")
            continue
        actual = 0
        for p in section_dir.glob("paper-*.json"):
            try:
                d = json.loads(p.read_text(encoding="utf-8"))
                actual += len(d.get("questions") or [])
            except Exception:
                continue
        actual_total += actual
        if actual != expected:
            failures.append(f"JA-8 data/papers/{section}/ has {actual} questions; expected {expected}")
    if actual_total != EXPECTED_TOTAL:
        failures.append(f"JA-8 total Q-count = {actual_total}; expected {EXPECTED_TOTAL}")
    return failures


def check_ja_9_engine_display_contract() -> list[str]:
    """Engine display contract is now enforced by paper-rendering UI, not by MD headers.

    Pre-2026-05-14: this check verified the 'Engine display note' header
    existed in every KB question MD file. KB deleted 2026-05-14; the
    contract now lives in docs/N5-syllabus-methodology.md §F.3 and is
    asserted by the test-mode UI itself (Playwright test suite covers
    answer-line hiding before commit).
    """
    return []


def check_ja_10_no_stub_redirect_text_in_data() -> list[str]:
    """No learner-facing string field in data/*.json contains '(see n5-' redirect text.

    Pass-12 finding: 40 questions in data/questions.json had leftover '(see n5-XXX)' text in
    question_ja, residual from the stub-pattern era. Pass-11 inlined examples in grammar.json
    but missed the parallel cleanup in questions.json. This invariant prevents recurrence.

    The 'notes' field is exempt because it intentionally records cross-references (audit trail).
    """
    LEARNER_FACING_FIELDS = {
        'ja', 'question_ja', 'prompt_ja', 'meaning_ja', 'example',
        'script_ja', 'translation_ja', 'translation_en', 'wrong', 'right',
        'gloss', 'meanings', 'pattern', 'meaning_en', 'explanation_en',
        'title_en', 'title_ja',
    }
    failures = []
    for fname in ['data/grammar.json', 'data/reading.json', 'data/listening.json',
                  'data/questions.json', 'data/vocab.json', 'data/kanji.json']:
        path = ROOT / fname
        if not path.exists():
            continue
        try:
            d = json.loads(path.read_text(encoding='utf-8'))
        except Exception:
            continue

        def walk(obj, path_str=''):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if k == 'notes':
                        continue  # notes are exempt
                    yield from walk(v, f'{path_str}.{k}' if path_str else k)
            elif isinstance(obj, list):
                for i, v in enumerate(obj):
                    yield from walk(v, f'{path_str}[{i}]')
            elif isinstance(obj, str):
                # Get the last segment of the path to check field name
                last_field = path_str.split('.')[-1].split('[')[0] if path_str else ''
                if last_field in LEARNER_FACING_FIELDS or path_str.endswith(']'):
                    if '(see n5-' in obj or '（see n5-' in obj:
                        yield path_str, obj[:80]

        for path_str, snippet in walk(d):
            failures.append(f"JA-10 {fname}:{path_str} contains '(see n5-' text: {snippet!r}")

    return failures


def check_ja_12_kanji_kb_data_consistency() -> list[str]:
    """data/kanji.json must agree with the canonical whitelist on every kanji.

    Pass-13 finding: data/kanji.json had silently-corrupted entries because
    a build tool had a regex bug that swallowed `[Ext]`-tagged entries.
    Specifically: 番 had on=['ごう'] (= 号's reading) and 会 had on=['いん']
    (= 員's reading). Plus 円 had a stale kun=['まる'] that was supposed to
    have been removed.

    This invariant compares data/kanji.json glyphs against the canonical
    scope whitelist (data/n5_kanji_whitelist.json) and reports any drift.

    History note: pre-2026-05-14 this check used KnowledgeBank/kanji_n5.md
    as the canonical glyph list. KB was deleted 2026-05-14 (merged into
    data/ and docs/N5-syllabus-methodology.md as a single source of truth);
    the whitelist took over the canonical-membership role.
    """
    failures = []
    whitelist_path = ROOT / "data" / "n5_kanji_whitelist.json"
    json_path = ROOT / "data" / "kanji.json"
    if not whitelist_path.exists() or not json_path.exists():
        return failures

    # Read whitelist (canonical N5 kanji set)
    try:
        whitelist = json.loads(whitelist_path.read_text(encoding="utf-8"))
    except Exception as e:
        failures.append(f"JA-12 data/n5_kanji_whitelist.json failed to parse: {e}")
        return failures
    canonical_glyphs = set(whitelist)

    # Read kanji.json
    try:
        d = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        failures.append(f"JA-12 data/kanji.json failed to parse: {e}")
        return failures
    json_glyphs = {e["glyph"] for e in d.get("entries", [])}

    missing_in_json = canonical_glyphs - json_glyphs
    extra_in_json = json_glyphs - canonical_glyphs
    for g in sorted(missing_in_json):
        failures.append(
            f"JA-12 whitelist has kanji '{g}' but data/kanji.json does not. "
            f"Add it to data/kanji.json with full schema."
        )
    for g in sorted(extra_in_json):
        failures.append(
            f"JA-12 data/kanji.json has kanji '{g}' but data/n5_kanji_whitelist.json does not. "
            f"Either add to whitelist or remove from kanji.json."
        )
    return failures


def check_ja_11_no_duplicate_choices() -> list[str]:
    """No MCQ question's `choices` array contains duplicates.

    Pass-12 finding: 3 questions (q-0220, q-0223, q-0280) had a duplicate option in the
    choices array (e.g., 'ません' appearing twice). Auto-grading is meaningful only when
    options are distinct. This invariant prevents recurrence in data/questions.json.
    """
    failures = []
    qpath = ROOT / 'data' / 'questions.json'
    if not qpath.exists():
        return failures
    try:
        d = json.loads(qpath.read_text(encoding='utf-8'))
    except Exception:
        return failures
    for q in d.get('questions', []):
        choices = q.get('choices')
        if not isinstance(choices, list):
            continue
        if len(choices) != len(set(choices)):
            from collections import Counter
            dups = [c for c, n in Counter(choices).items() if n > 1]
            failures.append(
                f"JA-11 data/questions.json {q.get('id', '?')} has duplicate choice(s): {dups} "
                f"(full choices: {choices})"
            )
    return failures


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

CHECKS: list[tuple[str, str, callable]] = [
    ("X-6.1", "Catalog completeness",       check_x_6_1_catalog_completeness),
    ("X-6.2", "今年 / こんねん consistency",  check_x_6_2_today_kotoshi),
    ("X-6.3", "No mixed kanji+kana",        check_x_6_3_no_mixed_kanji_kana),
    ("X-6.4", "Lint script present",        check_x_6_4_orphan_vocab),
    ("X-6.5", "No em-dashes",               check_x_6_5_no_em_dashes),
    ("X-6.6", "Ru-verb exception flags",    check_x_6_6_ru_verb_flags),
    ("X-6.7", "No false synonymy claims",   check_x_6_7_no_false_synonymy),
    ("X-6.8", "No ASCII digits in TTS src", check_x_6_8_no_ascii_digits_in_tts_source),
    ("X-6.9", "Primary-reading sanity",     check_x_6_9_furigana_primary_reading_sanity),
    ("JA-1",  "Stem-kanji scope",           check_ja_1_stem_kanji_scope),
    ("JA-2",  "Particle-set sanity",        check_ja_2_particle_set),
    ("JA-3",  "Furigana / catalog match",   check_ja_3_furigana_match),
    ("JA-4",  "Vocab reading uniqueness",   check_ja_4_vocab_reading_uniqueness),
    ("JA-5",  "Answer-key sanity",          check_ja_5_answer_key_sanity),
    ("JA-6",  "No two-correct-answers",     check_ja_6_no_two_correct_answers),
    ("JA-7",  "No duplicate stems in file", check_ja_7_no_duplicate_stems),
    ("JA-8",  "Q-count integrity",          check_ja_8_q_count_integrity),
    ("JA-9",  "Engine display contract",    check_ja_9_engine_display_contract),
    ("JA-10", "No (see n5-) redirect text", check_ja_10_no_stub_redirect_text_in_data),
    ("JA-11", "No duplicate MCQ choices",   check_ja_11_no_duplicate_choices),
    ("JA-12", "Kanji KB / JSON consistency", check_ja_12_kanji_kb_data_consistency),
    ("JA-13", "No out-of-scope kanji in user-facing data", lambda: _check_ja_13_no_out_of_scope_kanji_in_data()),
    ("JA-14", "No auto-ruby code in renderer",  lambda: _check_ja_14_no_auto_ruby_in_renderer()),
    ("JA-15", "Audio refs resolve to files on disk", lambda: _check_ja_15_audio_refs_on_disk()),
    ("JA-16", "Kanji examples use only target-or-whitelist kanji", lambda: _check_ja_16_kanji_examples_in_scope()),
    ("JA-17", "Grammar examples have vocab_ids (homograph guard)", lambda: _check_ja_17_examples_have_vocab_ids()),
    ("JA-18", "Reading explanation kanji subset of passage", lambda: _check_ja_18_reading_explanation_kanji()),
    ("JA-19", "Reading info-search has format_type", lambda: _check_ja_19_reading_info_search_format()),
    ("JA-20", "Reading choices kanji subset of passage", lambda: _check_ja_20_reading_choices_kanji()),
    ("JA-21", "N4-grammar markers require tier=late_n5", lambda: _check_ja_21_n4_grammar_tier_flag()),
    ("JA-22", "Kanji kun readings deduplicated", lambda: _check_ja_22_kun_dedup()),
    ("JA-23", "Listening script choices match choices array", lambda: _check_ja_23_listening_script_choices_match()),
    ("JA-24", "i-adj kanji primary reading is kun-yomi", lambda: _check_ja_24_iadj_kanji_primary_kun()),
    ("JA-25", "Whitelist exceptions documented (Pass-22 F-22.4)", lambda: _check_ja_25_whitelist_exceptions_documented()),
    ("JA-26", "No duplicate question IDs (Pass-23 2026-05-02)", lambda: _check_ja_26_no_duplicate_question_ids()),
    ("JA-27", "No English-translation/title fields in reading/listening (2026-05-02)", lambda: _check_ja_27_no_english_in_japanese_modules()),
    ("JA-28", "Dokkai-paper kanji bounded by N5 + exception list (2026-05-02)", lambda: _check_ja_28_dokkai_kanji_bounded()),
    ("JA-29", "Question subtype taxonomy is closed (paraphrase / kanji_writing only) (2026-05-02)", lambda: _check_ja_29_subtype_taxonomy()),
    ("JA-30", "No past-paper provenance signatures in question text (2026-05-02)", lambda: _check_ja_30_provenance()),
    ("JA-31", "Vocab PoS tags in vocabulary_n5.md agree with data/vocab.json (2026-05-02)", lambda: _check_ja_31_vocab_pos_parity()),
    ("JA-32", "Paper-JSON rationales appear verbatim in source MD (2026-05-04)", lambda: _check_ja_32_paper_rationale_md_parity()),
    ("JA-33", "Listening items carry mondai (1-4) + closed format_type enum (2026-05-05)", lambda: _check_ja_33_listening_mondai_taxonomy()),
    ("JA-34", "Core-N5 vs late-N5 split agrees with grammar.json tier field (2026-05-05)", lambda: _check_ja_34_core_late_split()),
    ("JA-35", "Every content item carries review_status from closed enum (2026-05-05)", lambda: _check_ja_35_review_status()),
    ("JA-36", "Answer-position distribution per corpus within +/-10pp of even (2026-05-06)", lambda: _check_ja_36_answer_position_balance()),
    ("JA-37", "localStorage namespace in code matches PRIVACY.md verbatim (2026-05-06)", lambda: _check_ja_37_namespace_doc_parity()),
    ("JA-38", "Every grammar pattern has >=1 common_mistakes entry (2026-05-06)", lambda: _check_ja_38_common_mistakes_floor()),
    ("JA-39", "Locale set in content data is exactly {en, hi} (2026-05-06)", lambda: _check_ja_39_locale_set_en_hi()),
    ("JA-40", "moji/goi/bunpou paper distractors bounded by N5 + paper-distractor exception (2026-05-08)", lambda: _check_ja_40_paper_distractor_kanji_bounded()),
    ("JA-41", "Hindi prose: Japanese grammatical particles attached to Hindi terms must be in kana (R-1.1, 2026-05-07)", lambda: _check_ja_41_kana_prefix_convention()),
    ("JA-47", "CONTENT-LICENSE.md corpus counts agree with live data/*.json counts (legal-vetting F-3, 2026-05-11)", lambda: _check_ja_47_content_license_counts()),
    ("JA-48", "KanjiVG SVG copyright headers preserved (Ulrich Apel attribution per CC-BY-SA 3.0) (legal-vetting F-8, 2026-05-11)", lambda: _check_ja_48_kanjivg_svg_headers()),
    # JA-49..53 added 2026-05-12 to lock in the audit-cycle gains
    # (ISSUE-115 register tag, ISSUE-116 wago/kango, ISSUE-112
    # common_mistakes categorized, ISSUE-118 contrasts, P2-12 cultural_callout).
    ("JA-49", "Every vocab has register in {neutral, polite, humble, respectful, casual} (ISSUE-115, 2026-05-12)", lambda: _check_ja_49_vocab_register_coverage()),
    ("JA-50", "Every vocab has register_origin in {wago, kango, gairaigo} (ISSUE-116, 2026-05-12)", lambda: _check_ja_50_vocab_register_origin_coverage()),
    ("JA-51", "Every grammar pattern has >=3 categorized common_mistakes (ISSUE-112, 2026-05-12)", lambda: _check_ja_51_common_mistakes_categorized()),
    ("JA-52", "Every grammar pattern has >=1 contrasts with valid with_pattern_id (ISSUE-118, 2026-05-12)", lambda: _check_ja_52_grammar_contrasts_floor()),
    ("JA-53", "Every grammar pattern has cultural_callout with non-trivial content (P2-12, 2026-05-12)", lambda: _check_ja_53_grammar_cultural_callout()),
    # JA-54..61: anti-item enforcement batch (Section-10 anti-items
    # from the 2026-05-12 richness audit). Programmatically lock the
    # frozen contracts so future edits cannot silently violate them.
    ("JA-54", "Every grammar pattern essay totals >=500 chars (Tofugu-bar; 2026-05-12)", lambda: _check_ja_54_essay_min_length()),
    ("JA-55", "Every grammar pattern essay has all 6 sub-fields (intro/why/pitfalls/contrasts/practice/cultural_context; 2026-05-12)", lambda: _check_ja_55_essay_schema()),
    ("JA-56", "Corpus sizes locked at 178/995/106/54/50 (Section-10 anti-items #1-4; updated 2026-05-16: BUG-018 −10 + BUG-017 collision −1 + BUG-019 −3 = 1009→995)", lambda: _check_ja_56_corpus_size_locks()),
    ("JA-57", "No LH/HL pitch notation in vocab.json (use {mora, drop} integer; anti-item #11; 2026-05-12)", lambda: _check_ja_57_no_lh_pitch_notation()),
    ("JA-58", "No 'JLPT.jp official' citations (JEES 出題基準 discontinued 2010; anti-item #12; 2026-05-12)", lambda: _check_ja_58_no_jlpt_jp_current_citation()),
    ("JA-59", "No competitive gamification (no XP/leaderboard/badge/achievement keys or files; anti-item #5 refined; 2026-05-12)", lambda: _check_ja_59_no_gamification_state()),
    ("JA-60", "No account / cloud-sync (no fetch to non-local URLs; anti-item #6; 2026-05-12)", lambda: _check_ja_60_no_account_or_cloud_sync()),
    ("JA-61", "No per-content discussion / comments route (anti-item #7; 2026-05-12)", lambda: _check_ja_61_no_discussion_route()),
    # JA-62..65: round 2 of anti-item + shape-contract enforcement.
    # 2026-05-12. Locks romaji-in-display rule + 3 shape contracts.
    ("JA-62", "No romaji in user-facing Japanese display fields (anti-item #9; 2026-05-12)", lambda: _check_ja_62_no_romaji_in_japanese_fields()),
    ("JA-63", "Authentic kanji_refs lists all N5 kanji in ja text (shape contract; 2026-05-12)", lambda: _check_ja_63_authentic_kanji_refs_complete()),
    ("JA-64", "Common_mistakes have wrong+right+why fields populated (renderer contract; 2026-05-12)", lambda: _check_ja_64_common_mistakes_shape()),
    ("JA-65", "Contrasts notes >=30 chars (substantive explanation bar; 2026-05-12)", lambda: _check_ja_65_contrast_notes_min_length()),
    # JA-66 (2026-05-13): hardening from v1.15.1 Phase 7 drift. JA-13
    # leaf-skips explanation_en + pattern_role as "English commentary",
    # but Phase 7 inadvertently embedded 4 above-N5 kanji (好/嫌/広/方)
    # in those fields. This check programmatically enforces the
    # corpus-wide convention: English commentary fields use only
    # N5-whitelist kanji + kana.
    ("JA-66", "explanation_en + pd_refs.pattern_role: kanji bounded by N5 whitelist (2026-05-13)", lambda: _check_ja_66_explanation_en_kanji_in_scope()),
    # JA-67 (2026-05-13): Density-3 floor regression guard. After the
    # D2/D3 fix (commit d7eaf43) the below-floor count is 24/106
    # (the genuinely-singleton N5 kanji: 力 + 23 with exactly 1 vocab).
    # Locking the count prevents accidental vocab deletions from
    # silently dropping more kanji below the floor.
    ("JA-67", "Density-3 below-floor count locked at 25 (kanji->vocab union method; 2026-05-13, updated 2026-05-16 +1 from BUG-018 dedup losing one 道 vocab usage)", lambda: _check_ja_67_density3_floor_lock()),
    # JA-68 (2026-05-13): cache-version 3-place sync. index.html CSS
    # `?v=`, index.html JS `?v=`, and sw.js CACHE_VERSION must all
    # match. Forgotten bumps in any one breaks the cache-bust on
    # release.
    ("JA-68", "Cache version sync: index.html css + js + sw.js CACHE_VERSION (2026-05-13)", lambda: _check_ja_68_cache_version_sync()),
    # JA-69 (2026-05-13): public_domain_refs legal-status guard. Added
    # after the legal-vetting run-2 audit caught 4 entries citing
    # in-copyright works (Kawabata d.1972, Nakamura ×2 d.1972, Nishijō
    # d.1970) whose pd_status text itself admitted "PD pending until
    # 2043". The fix is mechanical (replace with verified-PD authors);
    # this invariant prevents recurrence by failing on any future ref
    # whose pd_status contains red-flag legal language or whose
    # author_death_year is too recent under Japan's life+70 rule.
    ("JA-69", "public_domain_refs legal-status guard: no 'pending'/'protected' pd_status; author_death_year <= 1955 buffer (2026-05-13)", lambda: _check_ja_69_pd_refs_legal_status()),
    # JA-70 (2026-05-13): pitch_accent.mora must equal count_mora(reading)
    # for every vocab entry with a single-reading form. Added after the
    # native-teacher review caught 110/1009 entries (10.9%) with wrong
    # mora counts — all in the LLM-curated subset. Re-derived via
    # kanjium where possible (88 entries), mechanical mora-count
    # correction for the rest (5 entries). 17 entries skipped (multi-
    # reading slash-separated forms like "なに / なん" — a measurement
    # artifact, not a real error).
    ("JA-70", "Vocab pitch_accent.mora == count_mora(reading) for single-reading entries (2026-05-13)", lambda: _check_ja_70_vocab_pitch_mora_count()),
    # JA-71 (2026-05-13): grammar meaning_ja cross-contamination guard.
    # Added after the native-teacher review caught 13 patterns whose
    # meaning_ja text described a DIFFERENT grammar point (systematic
    # off-by-N misalignment from a translation pass). This check
    # asserts the first 「marker」 in meaning_ja has at least one
    # substring overlap with the pattern field, catching the case
    # where the marker is for a completely different rule.
    ("JA-71", "Grammar meaning_ja first 「marker」 must overlap with pattern field (2026-05-13)", lambda: _check_ja_71_meaning_ja_pattern_alignment()),
    # JA-72 (2026-05-13): gairaigo (loanword) vocab entries should use
    # katakana in their `form` field. Hiragana / kanji loanword forms
    # are a typography mismatch — natives write カタカナ for foreign-
    # origin words. Catches any future loanword authoring that
    # accidentally uses hiragana.
    ("JA-72", "Gairaigo vocab entries use katakana in their form field (2026-05-13)", lambda: _check_ja_72_gairaigo_katakana()),
    # JA-73 (2026-05-13 run-2): reading questions must use `prompt_ja`
    # as the canonical stem-field key (not the legacy `question_ja`).
    # Added after run-2 review caught 9 questions using the legacy
    # key. Schema parity across all reading questions.
    ("JA-73", "Reading questions use canonical prompt_ja key (not question_ja) (2026-05-13)", lambda: _check_ja_73_reading_prompt_ja_canonical()),
    # JA-74 (2026-05-13 opt-outs): vocab entries with multiple kana
    # readings must use the `readings: [...]` list schema (NOT a
    # slash-separated `reading: "なに / なん"` string). The migration
    # ran 2026-05-13; this invariant locks the schema forward.
    ("JA-74", "Vocab `reading` field is a single kana form; multi-reading uses `readings: [...]` list (2026-05-13)", lambda: _check_ja_74_vocab_readings_schema()),
    # JA-75 (2026-05-13): per-pattern meaning_ja marker dictionary.
    # JA-71's character-overlap heuristic catches egregious cross-
    # contamination but misses subtle cases (1 incidental char
    # overlap). JA-75 snapshots the current (verified-correct after
    # 3 native-teacher audits) meaning_ja's distinctive vocabulary
    # as `_meaning_ja_markers` on each pattern. meaning_ja must
    # contain at least one marker. Catches any future drift away
    # from the verified state.
    ("JA-75", "Grammar meaning_ja must contain at least one of its pattern's _meaning_ja_markers (2026-05-13)", lambda: _check_ja_75_meaning_ja_markers()),
    # JA-76 (2026-05-13): kanji on-yomi convention lock. Standard
    # Japanese pedagogical convention writes on-yomi in katakana
    # (ニチ, ジツ) and kun-yomi in hiragana (ひ, き). Genki I, Minna
    # no Nihongo, and JLPT dictionaries all follow this. The corpus
    # converted from hiragana to katakana on-yomi 2026-05-13; this
    # invariant locks the convention forward.
    ("JA-76", "Kanji on-yomi entries are written in katakana (pedagogical convention) (2026-05-13)", lambda: _check_ja_76_on_yomi_katakana()),
    # JA-77 (2026-05-13 run-2): no placeholder text leakage in user-
    # facing fields. Caught 2 entries in n5-098.wrong_corrected_pair
    # that shipped "(unspecified — keep prior)" strings — author meta-
    # commentary that should never have landed in production data.
    ("JA-77", "No placeholder/TODO/unspecified text in shipped data fields (2026-05-13)", lambda: _check_ja_77_no_placeholder_leakage()),
    # JA-78 (2026-05-13 run-2): example sentences should not contain
    # 3+ consecutive same hiragana characters. Catches the "ははは"
    # class — grammatically valid (母 + は particle) but reads as a
    # typo to N5 learners. Use kanji form for clarity.
    ("JA-78", "Grammar example sentences avoid 3+ consecutive same hiragana (clarity) (2026-05-13)", lambda: _check_ja_78_no_repeated_kana_examples()),
    # JA-79 (2026-05-13 run-3): grammar example `form` field must be
    # consistently populated across all examples in a pattern, OR
    # consistently absent. Mixed populations cause UI badges to render
    # partially (broken UX). Per anti-pattern §3.2.34.
    ("JA-79", "Grammar example form-field is all-populated or all-empty within a pattern (2026-05-13)", lambda: _check_ja_79_form_field_consistency()),
    # JA-81 (2026-05-14): boilerplate-leak detection. Discovered in the
    # Phase-1/2 grammar-example content audit that a small set of canned
    # sentences had been copy-pasted into example slots across many
    # patterns without checking pattern-relevance (e.g. "あなたは がくせいですか。"
    # appeared in 21 patterns, "あなたは どなたですか。" in 18). Phase 2 cleanup
    # dropped max occurrences below 10. This invariant locks the gain.
    ("JA-81", "No grammar example sentence repeated in 10+ patterns (boilerplate-leak guard, 2026-05-14)", lambda: _check_ja_81_no_boilerplate_leak()),
    # JA-82 (2026-05-14, IMP-006): every path mentioned in any _meta.see_also
    # or _meta.consumers entry across data/*.json must resolve to an existing
    # file/directory. Cross-reference rot detection.
    ("JA-82", "_meta.see_also / _meta.consumers path references resolve (IMP-006, 2026-05-14)", lambda: _check_ja_82_meta_paths_resolve()),
    # JA-83 (2026-05-15): vocab template-leak regression guard. The
    # 2026-05-15 vocab audit removed 539 templated example sentences
    # across 4 anti-patterns: 'Xを 見ました', 'あの Xは どこですか' (non-
    # location), 'これは Xです' on 3+ex entries (non-demonstrative), and
    # '「X」と あいさつしました' for non-greeting X. JA-83 locks the cleanup
    # so the templates can't drift back via future authoring.
    ("JA-83", "No vocab example uses removed templates (Xを見ました/あのXはどこ/これはXです/Xとあいさつしました) (2026-05-15)", lambda: _check_ja_83_no_vocab_template_leak()),
    # JA-84 (2026-05-15): regression guard from the kanji audit. Covers
    # (a) sentences with non-empty ja but empty translation_en (55
    # entries fixed in round-1) and (b) <u>...</u> HTML markup leaking
    # into ja fields that the renderer at js/kanji.js:347 HTML-escapes
    # (6 entries fixed in round-2).
    ("JA-84", "kanji.json sentences have populated translation_en + no <u> markup leakage (2026-05-15)", lambda: _check_ja_84_kanji_sentence_translations()),
    # JA-85 (2026-05-15): dokkai locale-parity regression guard. The
    # 2026-05-15 dokkai audit filled 83 missing explanation_hi entries
    # (Hindi coverage 19% -> 100%) and set format_role='primary' on 9
    # questions that needed it. JA-85 locks both gains: explanation_hi
    # must be populated whenever explanation_en is; format_role must
    # be present on every question; same rationale_hi rule for the
    # papers/dokkai mock-exam questions.
    ("JA-85", "Dokkai locale + format_role parity (2026-05-15)", lambda: _check_ja_85_dokkai_locale_parity()),
    # JA-86 (2026-05-15): mega-audit regression guard. The 2026-05-15
    # mega-audit filled 88 authentic.json Hindi context entries
    # (round-1) and 375 questions.json distractor explanations + 375
    # parallel Hindi distractor explanations (round-2, across 127
    # MCQ items that had completely-empty distractor maps). JA-86
    # locks both gains so future authoring batches can't leave the
    # parallel-locale fields empty.
    ("JA-86", "authentic context_hi + questions.json distractor en/hi coverage (2026-05-15)", lambda: _check_ja_86_mega_audit_locale_coverage()),
    # JA-87 (2026-05-15 wave-1): cross-corpus reading/gloss consistency.
    # Locks the 2026-05-15 cross-corpus audit's findings — any place
    # that quotes a vocab term must agree with vocab.json canonical
    # reading + gloss. Caught 1 real mismatch (まがる gloss) which was
    # reconciled by widening vocab.json's gloss to match.
    ("JA-87", "Cross-corpus reading/gloss consistency vs vocab.json canonical (2026-05-15 wave-1)", lambda: _check_ja_87_cross_corpus_consistency()),
    # JA-88 (2026-05-15 wave-2): particle-precision L2-error scan
    # across all 12 content corpora (6,309 sentences). Locks zero-
    # finding floor on 10 known L2-error patterns (Xを 好き, double
    # particle, na-adj as i-adj, double-mashita, etc.).
    ("JA-88", "Particle-precision L2-error scan across all corpora (2026-05-15 wave-2)", lambda: _check_ja_88_particle_precision()),
    # JA-89 (2026-05-15 native-teacher pass): five locks from the
    # final native-teacher audit covering counter-noun pedagogy,
    # bare-article EN translations, animacy violations, the あの-doko
    # grammar template, and 毎日-ことができます / あした-つもりです
    # verb template leaks. 159 fixes applied in waves A+B+D combined.
    ("JA-89", "Native-teacher audit locks: counter+article+animacy+templates (2026-05-15 phases A/B/D)", lambda: _check_ja_89_native_teacher_phase_b_d_locks()),
    # JA-90 (2026-05-15 pitch-accent reconciliation): every vocab
    # pitch_accent.drop is validated against the kanjium reference
    # (CC-BY-SA 4.0, pinned commit). Entries with confidence='high'
    # MUST match a drop in the reference; entries with 'medium'/'low'/
    # 'unverified' are accepted as-is (kept from LLM authoring with
    # transparency about uncertainty). The 2026-05-15 reconciliation
    # pass auto-fixed 22 disagreements, validated 810 matches, marked
    # 177 not-in-reference as 'unverified'.
    ("JA-90", "Vocab pitch_accent.drop validated vs kanjium reference (2026-05-15)", lambda: _check_ja_90_pitch_accent_reference_agreement()),
    # JA-96 (2026-05-16): BUG-014 close-out. Catches the bare
    # `<form>が あります。` / `<form>が います。` template-nonsense
    # pattern in vocab.json examples (e.g., 「一月が あります。」 —
    # grammatically possible but semantically nonsensical Japanese).
    # The corpus had 19 such cases; replacement landed in commit ?
    # (see fix_bug_014_vocab_template_nonsense_2026_05_16.py).
    # Numbers JA-91..95 are RESERVED for previously-documented but
    # not-yet-wired invariants (per the BUG-003..009 audit-coverage
    # addendum); JA-96 is the next available slot.
    ("JA-96", "No bare '<form> が あります/います' template examples in vocab.json (BUG-014 guard, 2026-05-16)", lambda: _check_ja_96_no_bare_arimasu_template()),
    # JA-97 (2026-05-16): BUG-015 close-out. Locks the canonical counter
    # schema after the 2026-05-16 normalization migrated 111 legacy
    # string-pair entries + isolated 16 counter-word metadata cases.
    ("JA-97", "vocab.json counter is null OR {kanji,reading} dict; counter_register null; counter_word_metadata isolated to counter-word entries (BUG-015 guard, 2026-05-16)", lambda: _check_ja_97_counter_schema_canonical()),
    # JA-98 (2026-05-16): BUG-016 close-out. Every verb POS in vocab
    # must carry a transitivity classification from a closed enum.
    ("JA-98", "Every vocab verb has transitivity in {transitive, intransitive, contact} (BUG-016 guard, 2026-05-16)", lambda: _check_ja_98_verb_transitivity_coverage()),
    # JA-99 (2026-05-16): BUG-017 close-out. Every kanji that appears
    # in any vocab `form` must be in n5_kanji_whitelist.json or
    # explicitly listed in dokkai_kanji_exception.json. Build fails
    # otherwise. JA-13 catches OOS kanji in grammar examples; this
    # extends the coverage to vocab forms (which JA-13 doesn't
    # currently check).
    ("JA-99", "All kanji in vocab.json `form` fields are in N5 whitelist OR exception list (BUG-017 guard, 2026-05-16)", lambda: _check_ja_99_vocab_form_kanji_in_scope()),
    # JA-100 (2026-05-17): BUG-020 + BUG-023 close-out. Every
    # kanji.json compound/example with a `vocab_id` must have its
    # `form` equal the linked vocab.form EXACTLY. Catches both
    # directions of drift:
    #   - BUG-020 case: kanji.json kanji form ↔ vocab.json kana
    #     form (BUG-017 propagation failure; OOS kanji exposure)
    #   - BUG-023 case: kanji.json kanji form ↔ vocab.json kana
    #     form when the kanji IS in scope (vocab.json was the
    #     laggard; native textbooks all teach the kanji form)
    # First wired as narrow OOS-only check; tightened to strict
    # equality after the BUG-023 audit demonstrated the looser
    # check missed a real bug class.
    ("JA-100", "kanji.json compound/example.form == linked vocab.form STRICT (BUG-020 + BUG-023 guard, 2026-05-17)", lambda: _check_ja_100_kanji_vocab_form_consistency()),
    # JA-101 (2026-05-17): BUG-022 close-out. kanji.json examples
    # use `form` field name only — no `lemma`. Normalized to match
    # vocab.json and the majority of existing entries.
    ("JA-101", "kanji.json example objects use `form` not `lemma` (BUG-022 guard, 2026-05-17)", lambda: _check_ja_101_kanji_examples_form_field_only()),
    # JA-102 (2026-05-17): BUG-021 close-out. The 6 standalone-kun
    # kanji must keep their kun-yomi primary_reading. Specifically
    # locks the post-fix state so a future auto-pipeline can't
    # silently flip them back to on-yomi.
    ("JA-102", "Standalone-kun primary_reading lock for 人/中/外/東/車/国 (BUG-021 guard, 2026-05-17)", lambda: _check_ja_102_standalone_kun_primary_reading()),
    # JA-103 (2026-05-17): BUG-024 close-out. Within any single kanji
    # entry's n5_compounds array, no two rows may have the same
    # (form, reading) tuple. Catches the duplicate-subset-gloss class
    # that BUG-024 reported. Different readings (e.g., 一日 with ついたち
    # vs いちにち) are legitimate polysemy and pass the check.
    ("JA-103", "kanji.json n5_compounds: (form, reading) tuple unique within each entry (BUG-024 guard, 2026-05-17)", lambda: _check_ja_103_kanji_compound_form_reading_unique()),
    # JA-104 (2026-05-17): BUG-041 close-out. reading.json passages use
    # `difficulty` field (not legacy `level`). Closed enum {easy,
    # medium, hard} or absent (info-search passages).
    ("JA-104", "reading.json passages use `difficulty` field (not legacy `level`); closed enum {easy,medium,hard} (BUG-041 guard, 2026-05-17)", lambda: _check_ja_104_reading_difficulty()),
    # JA-105 (2026-05-17): BUG-045 close-out. vocab_preview is a list
    # of vocab_id strings, never a list of dicts.
    ("JA-105", "reading.json vocab_preview is list of vocab_id strings (BUG-045 guard, 2026-05-17)", lambda: _check_ja_105_vocab_preview_shape()),
    # JA-106 (2026-05-17): BUG-044 close-out. format_type is null on
    # non-info-search passages; on info-search passages it's in
    # {schedule_table, menu_list, notice}. Never "comprehension".
    ("JA-106", "reading.json format_type ∈ {null, schedule_table, menu_list, notice} (BUG-044 guard, 2026-05-17)", lambda: _check_ja_106_format_type_enum()),
    # JA-107 (2026-05-17): Cross-Artifact Sync Protocol INV-4 — every
    # count in data/version.json.counts equals the actual length of
    # the referenced corpus file. Catches release-stamp drift where
    # a dedup/migration reduces a corpus but version.json (consumed
    # by app footer + sw.js CACHE_VERSION) isn't updated. Companion
    # to JA-47 (CONTENT-LICENSE.md side).
    ("JA-107", "version.json.counts equal live data array lengths (INV-4 guard, 2026-05-17)", lambda: _check_ja_107_version_json_counts()),
    # JA-108 (2026-05-17): Cross-Artifact Sync Protocol INV-5 — every
    # UI string key in locales/en.json must exist in every other
    # locale file (currently hi.json). Catches UI translation drift
    # where a new surface ships with EN copy and the HI/JA pass is
    # forgotten.
    ("JA-108", "locales/*.json key-set parity across all locales (INV-5 guard, 2026-05-17)", lambda: _check_ja_108_locale_key_parity()),
    # JA-109 (2026-05-17): Cross-Artifact Sync Protocol INV-10 —
    # every `tools/<name>.py` script reference in the governance
    # docs (procedure manual, accuracy prompt, N5Improvement,
    # AUDIT-COVERAGE) must resolve to an actual file on disk.
    # Catches script-rename / -removal drift where the manual
    # still tells builders to run a vanished file.
    ("JA-109", "procedure manual + prompts → script references resolve (INV-10 guard, 2026-05-17)", lambda: _check_ja_109_script_references_resolve()),
    # JA-110 (2026-05-17): BUG-047 close-out. listening.json items
    # must NOT carry the legacy `voice_planned` field. The
    # audio_render_meta block is the canonical source for engine +
    # speaker info post the 2026-05-12 VOICEVOX migration. Pre-fix
    # all 50 items declared voice_planned.engine="edge-tts" while
    # audio_render_meta.voice_provider="voicevox" — same dual-field
    # drift class as BUG-044 (reading.json format_type vs
    # format_role) and BUG-051 (listening.json format vs format_type).
    ("JA-110", "listening.json items deprecate legacy `voice_planned` (BUG-047 guard, 2026-05-17)", lambda: _check_ja_110_no_voice_planned()),
    # JA-111 (2026-05-17): BUG-051 close-out. listening.json items
    # must NOT carry the legacy `format` field; `format_type` is the
    # canonical descriptive field with closed enum
    # {task_understanding, point_understanding, utterance_expression,
    # immediate_response}. Same dual-field drift class as BUG-044
    # and BUG-047.
    ("JA-111", "listening.json drops legacy `format`; `format_type` ∈ closed enum (BUG-051 guard, 2026-05-17)", lambda: _check_ja_111_format_type_only()),
    # JA-112 (2026-05-17): BUG-050 close-out (charitable interpretation).
    # AUDIO.md count claims ("N listening items use M distinct VOICEVOX
    # speakers") must match the live data — N == len(listening.json.items),
    # M == distinct count in audio_render_meta.voices_used across items.
    # Same INV-4 (cross-artifact sync protocol) drift class as JA-47
    # (CONTENT-LICENSE.md) and JA-107 (version.json), extended to the
    # AUDIO.md user-facing doc surface.
    ("JA-112", "AUDIO.md 'N listening items use M distinct VOICEVOX speakers' matches live data (BUG-050 charitable guard, 2026-05-17)", lambda: _check_ja_112_audio_md_listening_counts()),
    # JA-114 (2026-05-17): listening.json pacing_status closed-enum
    # check. After the BUG-048/049 close-out (commit 47d1edc), every
    # item has a measured pacing_morae_per_min + a status reflecting
    # its position in the JLPT N5 target band. Lock the field's
    # value-domain so future regressions (null / new ad-hoc strings)
    # are blocked at CI.
    ("JA-114", "listening.json pacing_status ∈ {in_range, too_slow, too_fast, no_audio, unmeasured} (BUG-048/049 follow-up, 2026-05-17)", lambda: _check_ja_114_pacing_status_enum()),
    # JA-115 (2026-05-17): README.md "Content" section count claims
    # must match live corpus sizes. JA-112's prose-pattern lock
    # extended to the README — same Cross-Artifact Sync Protocol
    # INV-4 class, fourth instance alongside JA-47 (CONTENT-LICENSE.md),
    # JA-107 (version.json), JA-112 (AUDIO.md). Prior README state
    # carried "1041 vocab / 40 reading / 40 listening" against live
    # 995 / 54 / 50 — drift caught and corrected in this same commit.
    ("JA-115", "README.md 'Content' section counts match live data (extends JA-47/107/112 to README surface, 2026-05-17)", lambda: _check_ja_115_readme_counts()),
    # JA-116 (2026-05-17): Cross-Artifact Sync Protocol INV-6 promotion
    # Partial → Wired. Every A-NN audit category in the accuracy prompt,
    # every Phase-0 regression block in N5Improvement, and every FP-NN
    # false-positive class in the accuracy prompt must have ≥1 matching
    # xlsx scenario row (named explicitly in any scenario's
    # Scenario / Notes / Tools column). Prevents prompts↔xlsx drift
    # that the b466293 bulk sync had to close manually.
    ("JA-116", "every A-NN / Phase-0 / FP-NN in prompts has ≥1 matching xlsx scenario row (INV-6 promotion to Wired, 2026-05-17)", lambda: _check_ja_116_prompt_xlsx_coverage()),
    # JA-117 (2026-05-17): Cross-Artifact Sync Protocol INV-7
    # extension. Every cross-corpus passage_id reference (kanji.json
    # `.entries[*].reading_passages[*].passage_id`) resolves to a
    # passage in reading.json; every pattern_id reference (reading.
    # json `.passages[*].grammar_footnotes[*].pattern_id` + nested
    # `.patterns[*].pattern_id`) resolves to a pattern in
    # grammar.json. Promotes the passage_id / pattern_id sub-case of
    # INV-7 (cross-file references resolve) from Partial → Wired.
    ("JA-117", "cross-corpus passage_id (kanji↔reading) + pattern_id (reading↔grammar) refs resolve (INV-7 extension, 2026-05-17)", lambda: _check_ja_117_cross_corpus_id_refs()),
    # JA-118 (2026-05-17): Cross-Artifact Sync Protocol INV-9
    # promotion Partial → Wired. Every Fixed-status row in the User
    # Reported Bugs sheet must have a non-empty Fix Commit cell
    # (commit SHA or explicit "(unattributable)" annotation).
    # Companion populated via tools/populate_bug_fix_commits_2026_05_17.py
    # which scans git log for commit subjects mentioning the BUG-NNN
    # (including range patterns like "BUG-041 through BUG-046").
    ("JA-118", "every Fixed row in xlsx User Reported Bugs has a non-empty Fix Commit cell (INV-9 promotion to Wired, 2026-05-17)", lambda: _check_ja_118_bug_fix_commit_link()),
    # JA-91 (2026-05-17 final unblock + Phase B resolution): cross-pattern
    # explanation_en similarity ≥0.85 Levenshtein. Initial wire-up surfaced
    # 43 pairs (DUPLICATE_PATTERN ×8, CROSS_REFERENCE ×21,
    # ALTERNATIVE_VARIANT ×12, SUBSET ×2); Phase B (2026-05-17) addressed
    # each by rewriting the deferring side's explanation_en so it diverges
    # from the canonical (or both sides, for ALTERNATIVE_VARIANT pairs).
    # data/_ja91_baseline.json now carries an EMPTY baseline_pairs array
    # — JA-91 enforces the threshold unconditionally on the current
    # corpus. The empty-baseline file is retained as a RESOLVED snapshot
    # documenting the prior pair set and its classification legend.
    ("JA-91", "no cross-pattern explanation_en similarity ≥0.85; data/_ja91_baseline.json carries empty baseline (BUG-003 contamination guard, 2026-05-17 Phase B resolution)", lambda: _check_ja_91_explanation_similarity()),
    # JA-92: no EN sentence repeated in 10+ grammar examples (parallel
    # to JA-81 which catches the JA side; BUG-003/005 lineage).
    ("JA-92", "no EN translation_en repeated in 10+ grammar examples (BUG-003/005 boilerplate guard, 2026-05-17 wire-up)", lambda: _check_ja_92_en_boilerplate()),
    # JA-93: pitch_marks total mora == count_mora(form) per vocab
    # entry (BUG-004 lineage).
    ("JA-93", "vocab.json pitch_marks total mora == count_mora(form) (BUG-004 algorithmic mora-count guard, 2026-05-17 wire-up)", lambda: _check_ja_93_pitch_marks_mora()),
    # JA-94 (2026-05-17 final unblock + Phase A resolution): BUG-006
    # pattern-instance contamination guard. Every grammar example's
    # `ja` must contain ≥1 structural marker from its parent pattern's
    # marker list in data/pattern_markers.json (authored via
    # tools/author_pattern_markers_2026_05_17.py — 178-pattern catalog).
    # Initial wire-up surfaced 14 BUG-006-CANDIDATE wrong-example
    # failures; Phase A (2026-05-17) replaced each example with a
    # parent-pattern-demonstrating one. data/_ja94_baseline.json now
    # carries an EMPTY baseline_failing_examples array — JA-94 enforces
    # marker-presence unconditionally on every example (1782 of 1782).
    # The empty-baseline file is retained as a RESOLVED snapshot.
    ("JA-94", "every grammar example contains ≥1 structural marker from data/pattern_markers.json; data/_ja94_baseline.json carries empty baseline (BUG-006 pattern-instance contamination guard, 2026-05-17 Phase A resolution)", lambda: _check_ja_94_pattern_marker_per_example()),
    # JA-95: particle-pattern alignment for particle-category patterns
    # (BUG-009 lineage). Every example of a Particles-category pattern
    # must use the pattern's named particle. First-run wire-up surfaced
    # n5-028 ex[5] missing の (fixed inline 2026-05-17 — replaced ja
    # with possessive-marked variant).
    ("JA-95", "particle-pattern alignment for Particles-category patterns (BUG-009 alignment guard, 2026-05-17 wire-up)", lambda: _check_ja_95_particle_pattern_alignment()),
    # JA-119 (2026-05-17): BUG-050 round-3 close-out. The spec's §7.3
    # example block (a sample version.json shown to developers) must
    # match the live data/version.json content for the count fields.
    # Previously stale at v1.12.50-era values (vocab 1041, reading 45,
    # listening 47, etc.) which caused multiple confused re-reports
    # of BUG-050 against the actual file. Same Cross-Artifact Sync
    # Protocol INV-4 class as JA-47 / JA-107 / JA-112 / JA-115 —
    # fifth surface where corpus counts surface in prose.
    ("JA-119", "spec §7.3 sample version.json matches live data/version.json counts (BUG-050 round-3 guard, 2026-05-17)", lambda: _check_ja_119_spec_version_sample()),
    # JA-113 (2026-05-17): meta-route static-mirror freshness check.
    # For each markdown-sourced meta route (home / changelog / privacy /
    # notices), the static-mirror HTML at /N5/<slug>/index.html must
    # contain the latest H1/H2 heading from the source markdown. Catches
    # the recurring drift class where a maintainer edits the source
    # markdown (CHANGELOG.md / PRIVACY.md / etc.) but forgets to re-run
    # `tools/build_static_mirrors.py --stages meta` in the same commit.
    # Three instances of this drift class were observed in this session
    # (cdef185 / 47d1edc CHANGELOG.md edits without mirror regen);
    # JA-113 prevents recurrence. Same Rule-5 INV-7 (cross-file
    # references resolve) class, extended from data-references to
    # derived-mirror freshness.
    ("JA-113", "meta-route static mirrors (home/changelog/privacy/notices) reflect latest markdown heading (mirror-freshness guard, 2026-05-17)", lambda: _check_ja_113_meta_mirror_freshness()),
    # JA-120 / JA-121 / JA-122 (2026-05-18): PAPER-001..004 close-out.
    # JA-120 — paper bunpou Mondai 1 questions must have a grammarPatternId
    # that matches the canonical particle pattern for their correct answer.
    # PAPER-001 caught 60+ questions where an early auto-tag pass set every
    # ambiguous question to n5-013 (も) regardless of the actual particle.
    # JA-121 — paper rationale / rationale_hi fields must not contain
    # commit-message-style meta-fix phrases (e.g., "auto_inferred",
    # "Stem now anchored", "(was X)", "replaces ... per policy"). PAPER-003
    # caught 6 questions where the learner-facing rationale was contaminated
    # with audit-trail text intended for commit messages.
    # JA-122 — paper rationale_hi fields must not contain English-pattern
    # fragments (apostrophe-s possessive, English contractions, "yet" / "lot"
    # leakage, mojibake artifacts like "यहाँre" / "o'घड़ी"). PAPER-004 caught
    # 30+ questions in Mondai 2 sentence-ordering whose rationale_hi was a
    # word-by-word literal translation of the English rationale rather than
    # natural Hindi.
    ("JA-120", "paper bunpou Mondai-1 grammarPatternId matches canonical particle pattern (PAPER-001 drift guard, 2026-05-18)", lambda: _check_ja_120_paper_pattern_id_matches_correct_particle()),
    ("JA-121", "paper rationale / rationale_hi free of commit-message-style meta-fix history (PAPER-003 drift guard, 2026-05-18)", lambda: _check_ja_121_no_meta_fix_history_in_rationale()),
    ("JA-122", "paper rationale_hi free of English-pattern fragments (apostrophe-s / contractions / mojibake; PAPER-004 drift guard, 2026-05-18)", lambda: _check_ja_122_no_english_pattern_fragments_in_rationale_hi()),
    # JA-123..127 (2026-05-18): LLM-001..005 + REG-001 close-out invariants.
    # JA-123 — every paper-pack JSON has a static HTML mirror at /papers/<id>/index.html (LLM-001 drift guard).
    # JA-124 — sitemap.xml has ≥1000 <loc> entries (LLM-002 regression floor; was 10 pre-fix).
    # JA-125 — data/index.json size_bytes match actual on-disk file sizes (LLM-003 / INV-LLM-3 corpus-count drift class).
    # JA-126 — the 7 LLM-005 summary pages + llms.txt (root + N5) all exist.
    # JA-127 — no wrong_corrected_pair entry with error_category=register may carry a self-contradicting parenthetical like "(formal)" / "(polite)" / etc. on the wrong-field (REG-001 D6 class).
    ("JA-123", "papers static-mirror exists for every data/papers/*/*.json (LLM-001 drift guard, 2026-05-18)", lambda: _check_ja_123_papers_static_mirrors_exist()),
    ("JA-124", "sitemap.xml has ≥1000 URL entries (LLM-002 regression floor, 2026-05-18)", lambda: _check_ja_124_sitemap_includes_all_mirrors()),
    ("JA-125", "data/index.json size_bytes match actual file sizes (LLM-003 / INV-LLM-3 guard, 2026-05-18)", lambda: _check_ja_125_data_index_corpus_sizes_match()),
    ("JA-126", "7 LLM-005 summary pages + llms.txt (root + N5) all exist (LLM-005 close-out, 2026-05-18)", lambda: _check_ja_126_summary_pages_exist()),
    ("JA-127", "no wrong_corrected_pair w/ error_category=register has self-contradicting wrong-field parenthetical (REG-001 D6 guard, 2026-05-18)", lambda: _check_ja_127_register_variant_no_wrong_right_with_register_why()),
    # JA-128..130 (2026-05-18): DOKKAI-001..003 close-out invariants.
    # JA-128 — paper questions must NOT carry passage_text (single-source-of-truth via passages[label].text).
    # JA-129 — paper rationale_hi free of " ago" / " yet" / " lot" untranslated English temporal/quantity markers.
    # JA-130 — every paper question has grammarPatternId as a key (may be null with documented "not_applicable_<reason>" provenance).
    ("JA-128", "no `passage_text` on paper questions; canonical text lives in passages[label].text (DOKKAI-001 drift guard, 2026-05-18)", lambda: _check_ja_128_no_passage_text_on_paper_questions()),
    ("JA-129", "paper rationale_hi free of untranslated English temporal/quantity fragments (\" ago\" / \" yet\" / \" lot\") (DOKKAI-002 drift guard, 2026-05-18)", lambda: _check_ja_129_no_untranslated_english_temporal_in_rationale_hi()),
    ("JA-130", "every paper question has grammarPatternId key; null requires not_applicable_* provenance (DOKKAI-003 schema-shape guard, 2026-05-18)", lambda: _check_ja_130_paper_questions_have_grammarPatternId_key()),
    # JA-131..134 (2026-05-19): MOB-001..019 mobile-UI close-out invariants.
    # JA-131 — locale parity for the nav.all_levels key used by home-up link.
    # JA-132 — MOB-* CSS compliance batch (touch-target rules) remains in main.css + main.min.css.
    # JA-133 — form input font-size >=16px (iOS Safari auto-zoom guard).
    # JA-134 — no dead-end hash-route references in js/home.js or js/listening-story.js.
    ("JA-131", "locales/{en,hi}.json carry nav.all_levels key (MOB-007 drift guard, 2026-05-19)", lambda: _check_ja_131_locale_all_levels_key_parity()),
    ("JA-132", "css/main.css + main.min.css carry MOB-001..016 mobile-UI compliance batch (MOB-002/003/004/005/011/012/013/014/015/016 drift guard, 2026-05-19)", lambda: _check_ja_132_touch_target_css_rules_present()),
    ("JA-133", "css/main.css has input/textarea/select font-size>=16px rule (MOB-006 iOS auto-zoom guard, 2026-05-19)", lambda: _check_ja_133_form_input_font_size_rule()),
    ("JA-134", "js/home.js + js/listening-story.js free of dead-end hash routes #/levels and #/listening/story (MOB-008/009 drift guard, 2026-05-19)", lambda: _check_ja_134_no_dead_end_hash_routes()),
    # JA-135 (2026-05-19): general hash-route-resolution guard.
    # Extends JA-134 from "2 specific known patterns" to "every hash-href
    # in js/*.js resolves to a route in app.js's routes dict". Catches
    # future MOB-008/009-class introductions without per-pattern
    # hardcoding.
    ("JA-135", "every #/X hash-route href in js/*.js resolves to a route in app.js routes dict (MOB-008/009 generalization, 2026-05-19)", lambda: _check_ja_135_all_hash_hrefs_resolve_to_routes()),
    # JA-136 (2026-05-19): GOI-001 copy-paste-content-mismatch guard.
    # No two questions in the same paper file may share a byte-
    # identical rationale_hi value longer than 30 chars. Catches the
    # GOI-001-class copy-paste where goi-6.11 rationale_hi was a
    # verbatim copy of goi-6.12's (about a different question topic).
    ("JA-136", "no rationale_hi shared verbatim by 2+ questions within the same paper (GOI-001 copy-paste guard, 2026-05-19)", lambda: _check_ja_136_rationale_hi_unique_within_paper()),
    # JA-80 was attempted (2026-05-13 run-4) and removed: heuristic
    # "meaning_ja must share ≥1 Japanese substring with meaning_en" had
    # 19 false positives on legitimate patterns where meaning_ja
    # paraphrases meaning_en using different words (e.g., n5-068
    # meaning_en="Plain past negative (-なかった)" but meaning_ja
    # ="ふつうの かこ ひてい" — semantically equivalent, no string
    # overlap). The cross-contamination class requires LLM-level
    # semantic comparison and stays in the manual-review domain.
    # The single run-4 catch (n5-166) was caught by the Phase-0
    # CHECK-6 re-implementation, which is a stricter version of JA-71
    # (no fallback pass); 2 false positives there required manual
    # inspection. Documented lesson at A12 (audit prompt line 465).
]


# ---------------------------------------------------------------------------
# JA-13 / JA-14 added in Pass 13 (auto-furigana removal)
# ---------------------------------------------------------------------------

def _check_ja_13_no_out_of_scope_kanji_in_data() -> list[str]:
    """No out-of-scope kanji appears in user-facing fields of grammar.json,
    questions.json, reading.json, listening.json. Enforces the 'kanji only
    if in N5 syllabus, kana otherwise' rule from the Pass-13 redesign."""
    failures = []
    try:
        whitelist = set(json.loads((ROOT / "data" / "n5_kanji_whitelist.json").read_text(encoding="utf-8")))
    except Exception as e:
        return [f"JA-13 could not load n5_kanji_whitelist.json: {e}"]
    KANJI_LOCAL = re.compile(r"[一-鿿]")
    SKIP_FIELDS = {"translation_en", "explanation_en", "meaning_en", "gloss",
                   "title_en", "prompt_en", "distractor_explanations",
                   "common_mistakes", "reading", "furigana"}
    # Subtrees rooted at these keys are skipped entirely (their children too).
    # ISSUE-068 (round-7 2026-05-06): common_mistakes contains Japanese
    # example sentences that may use any kanji to illustrate the wrong/right
    # pair (the whole point of the field is showing native-grade Japanese).
    # The previous SKIP_FIELDS-on-leaf logic missed these because walk()
    # descended into the dict and the leaf's key became 'wrong'/'right'/'why',
    # not 'common_mistakes'. Subtree-skip closes that hole.
    #
    # ISSUE-056 (round-7 2026-05-06): l1_notes is a dict whose values are
    # translations into the learner's L1. These contain whatever script
    # that language uses (Devanagari for hi) and must not be subjected to
    # the N5-kanji-only rule. Subtree-skip.
    # Round-8 (2026-05-06): cultural_context + summary fields contain
    # mixed Japanese (illustration phrases) + learner-language commentary.
    # Like l1_notes, the whole subtree is exempt from N5-only kanji rule.
    # Round-9 (2026-05-11): authentic_citations is bibliographic metadata
    # (source + context note). The context describes WHERE the pattern
    # appears in standard textbooks (Genki/Minna/etc.) and uses normal
    # Japanese vocabulary for that meta-commentary (図書館, 映画, 基本形 etc.).
    # Not primary learner content — exempt from N5-only kanji rule, like
    # cultural_context and summary which are similar pedagogical commentary.
    # Also exempt wrong_corrected_pair (categorized common-learner mistakes —
    # mirrors common_mistakes; whole purpose is showing native Japanese
    # to illustrate error patterns) and politeness_ladder (4-tier register
    # variation — humble/respectful forms use vocab beyond N5 by design).
    SKIP_SUBTREE_FIELDS = {"common_mistakes", "distractor_explanations",
                           "l1_notes", "cultural_context", "summary",
                           "authentic_citations", "wrong_corrected_pair",
                           "politeness_ladder",
                           # P2-12 (2026-05-11): cultural_callout is
                           # flowing-prose pedagogical commentary on
                           # pattern usage (business etiquette, age
                           # cohort, register risk). Like cultural_context,
                           # the value needs to read as natural Japanese
                           # commentary and may reference N4+ vocabulary
                           # by necessity ("商業", "業務", "警告" etc.).
                           "cultural_callout",
                           # P2-13 (2026-05-11): etymology stories
                           # reference historical Chinese / OBI script
                           # names that go well beyond N5 (e.g., 圓,
                           # 萬, 翟). Same rationale.
                           "etymology",
                           # P2-8 (2026-05-12): reflection_prompts are
                           # post-passage critical-thinking questions
                           # authored in English (so the prompt_en
                           # field is fine), but the prompt_ja variants
                           # may include analytical vocabulary (反応 /
                           # 関係 / 計画 / 自分 etc.) that goes beyond
                           # the N5 kanji whitelist. Same rationale.
                           "reflection_prompts",
                           # P2-14 (2026-05-12): inference_question_expansion
                           # mirrors reflection_prompts — open-ended
                           # follow-up questions whose Japanese phrasing
                           # naturally uses analytical N4+ vocabulary.
                           "inference_question_expansion",
                           # ISSUE-118 wave 4 (2026-05-12): contrasts notes
                           # are pedagogical commentary that explains the
                           # distinction between two grammar patterns and
                           # naturally cites illustrative example phrases
                           # ("映画が好きです", "1冊/3個/2人", "大きい家")
                           # whose kanji set may exceed the strict N5
                           # whitelist. Same rationale as cultural_callout
                           # and etymology — explanatory text that needs
                           # natural Japanese, not curriculum-bound input.
                           "contrasts",
                           # ISSUE-114 (2026-05-12): audio_render_meta on
                           # listening items carries VOICEVOX character
                           # names (春日部つむぎ / 玄野武宏 / 四国めたん /
                           # ずんだもん / 雨晴はう / 青山龍星) for legal
                           # attribution. These character names use kanji
                           # well beyond N5 (春/部/玄/野/武/宏/etc.) but
                           # they are RENDERING METADATA, not learner-
                           # facing content. The runtime audio player
                           # displays only the role label, not the
                           # character name. Same rationale as the
                           # other explanatory-text exemptions above.
                           "audio_render_meta",
                           # 2026-05-13 (v1.15.0): public_domain_refs on
                           # grammar patterns cite literary / government /
                           # proverb sources with author names and work
                           # titles that use kanji well beyond N5
                           # (漱石 / 芥川 / 八雲 / 賢治 / 一葉 / 鷗外 etc.).
                           # These are CITATION METADATA, not learner-
                           # facing comprehension content. The renderer
                           # displays them as authentic external references,
                           # not as text the learner is expected to parse.
                           # Same rationale as the other commentary-field
                           # exemptions above.
                           "public_domain_refs",
                           # 2026-05-13 (JA-75 install): _meaning_ja_markers
                           # is a per-pattern internal-metadata snapshot of
                           # the distinctive vocabulary in meaning_ja, used
                           # by JA-75 to detect content drift. The list
                           # naturally includes Japanese function-name terms
                           # like 場所 / 時間 / 場面 (N4+ kanji) since those
                           # appear in the verified-correct meaning_ja text.
                           # Internal metadata; never user-facing.
                           "_meaning_ja_markers",
                           # BUG-052/053 fix 2026-05-17: voice_variety_plan
                           # in listening.json _meta carries the VOICEVOX
                           # speaker_catalog (春日部つむぎ / 玄野武宏 / 雨晴は
                           # う / 青山龍星 / etc.) and explanatory prose
                           # about the 2026-05-12 render. Same rationale as
                           # audio_render_meta (line 1282) and
                           # public_domain_refs (line 1294) — provenance /
                           # rendering metadata with kanji beyond N5, never
                           # learner-facing. The audio player UI displays
                           # only the role label, never the character name.
                           "voice_variety_plan",
                           # BUG-049 fix 2026-05-17 (surface-only): the
                           # pacing_fix_status block carries the queued-
                           # action description for the slow-pacing
                           # re-render, including the human-readable
                           # explanation with example item IDs. Same
                           # rationale as voice_variety_plan.
                           "pacing_fix_status",
                           # BUG-052 fix 2026-05-17: the legacy
                           # voice_variety_plan_2026_05_07 block is
                           # marked superseded but kept for history;
                           # exempt for the same rationale.
                           "voice_variety_plan_2026_05_07"}
    # ISSUE-056 + 2026-05-06 locale narrowing (IMP-096): locale-suffixed
    # translation fields. The values are translations into hi (Hindi).
    # Pattern: <basename>_<locale> for locale ∈ {hi}. The pre-narrowing
    # locale set was {vi,id,ne,zh}; those keys are pruned from data, but
    # the regex still allows them so historical fixtures or migration
    # tooling running against pre-narrowing data don't false-trip.
    LOCALE_SUFFIX_RE = re.compile(r"^(?:meaning|explanation|gloss|title|prompt|description|note|meanings)_(?:hi|vi|id|ne|zh)$")

    def walk(obj, key, path, hits):
        if key in SKIP_SUBTREE_FIELDS:
            return
        if isinstance(key, str) and LOCALE_SUFFIX_RE.match(key):
            return
        if isinstance(obj, str):
            if key in SKIP_FIELDS: return
            for ch in obj:
                if KANJI_LOCAL.match(ch) and ch not in whitelist:
                    hits.append((path, ch, obj[:60]))
                    return
        elif isinstance(obj, dict):
            for k, v in obj.items(): walk(v, k, f"{path}.{k}", hits)
        elif isinstance(obj, list):
            for i, v in enumerate(obj): walk(v, key, f"{path}[{i}]", hits)
    for fname in ["data/grammar.json", "data/questions.json", "data/reading.json", "data/listening.json"]:
        try:
            d = json.loads((ROOT / fname).read_text(encoding="utf-8"))
        except Exception:
            continue
        hits = []
        walk(d, None, fname, hits)
        for path, kanji, text in hits[:5]:
            failures.append(f"JA-13 {path}: out-of-scope kanji '{kanji}' in {text!r}")
        if len(hits) > 5:
            failures.append(f"JA-13 {fname}: ... and {len(hits) - 5} more")
    return failures


def _check_ja_14_no_auto_ruby_in_renderer() -> list[str]:
    """js/furigana.js must not auto-generate ruby for in-scope N5 kanji.
    The Pass-13 redesign removed this feature because the single-primary
    lookup picks wrong context-dependent readings (大学 displays
    だい+がく, but 大[おお] alone). Guard the regression here."""
    src = (ROOT / "js" / "furigana.js").read_text(encoding="utf-8") if (ROOT / "js" / "furigana.js").exists() else ""
    if not src:
        return ["JA-14: js/furigana.js missing"]
    # Look for the bad pattern: a ruby tag that references readings[ch].primary
    if "readings[ch]?.primary" in src or "readings[ch].primary" in src:
        return ["JA-14: js/furigana.js still references readings[ch].primary - auto-furigana not fully removed"]
    # The function should not import primary readings as a render input
    if "n5KanjiReadings" in src and "primary" in src:
        return ["JA-14: js/furigana.js still wires the readings map into the renderer"]
    return []


def _check_ja_16_kanji_examples_in_scope() -> list[str]:
    """K-1 invariant: every kanji entry's `examples[*].form` must contain
    only kanji that are either (a) the target kanji of the card, or (b)
    in the N5 whitelist. Non-kanji characters (kana) are always allowed.

    Out-of-scope kanji should be substituted with their kana reading
    BEFORE landing in the data file. The renderer doesn't perform the
    substitution at display time; the form here is what's shown.
    """
    failures: list[str] = []
    kanji_path = ROOT / "data" / "kanji.json"
    wl_path = ROOT / "data" / "n5_kanji_whitelist.json"
    if not kanji_path.exists() or not wl_path.exists():
        return ["JA-16: data files missing"]
    try:
        whitelist = set(json.loads(wl_path.read_text(encoding="utf-8")))
        data = json.loads(kanji_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-16: parse error: {e}"]
    KANJI_RE = re.compile(r"[一-鿿]")
    for entry in data.get("entries", []):
        target = entry.get("glyph")
        for ex in entry.get("examples", []):
            form = ex.get("form", "")
            for ch in KANJI_RE.findall(form):
                if ch == target or ch in whitelist:
                    continue
                failures.append(
                    f"JA-16 kanji '{target}' has example '{form}' with "
                    f"out-of-scope kanji '{ch}'. Substitute with kana per K-1 rule."
                )
    return failures


def _check_ja_17_examples_have_vocab_ids() -> list[str]:
    """Every non-empty example in data/grammar.json must have a
    `vocab_ids` field (a list, possibly empty). This prevents the
    homograph mismatch class flagged 2026-05-01 - without explicit
    example→vocab links, the renderer falls back to substring matching
    on the form field, which can't disambiguate homographs.

    Auto-population: tools/link_grammar_examples_to_vocab.py walks every
    example and assigns vocab_ids. Run that tool whenever new examples
    are added; this invariant guards the result.
    """
    failures: list[str] = []
    grammar_path = ROOT / "data" / "grammar.json"
    if not grammar_path.exists():
        return ["JA-17: data/grammar.json missing"]
    try:
        data = json.loads(grammar_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-17: parse error: {e}"]
    for p in data.get("patterns", []):
        for i, ex in enumerate(p.get("examples", [])):
            if not ex.get("ja"):
                continue  # placeholder examples without ja text are ignored
            if "vocab_ids" not in ex or not isinstance(ex["vocab_ids"], list):
                failures.append(
                    f"JA-17 {p.get('id', '?')} examples[{i}] is missing the "
                    f"`vocab_ids` field. Run "
                    f"`python tools/link_grammar_examples_to_vocab.py` to populate."
                )
                if len(failures) >= 20:
                    failures.append("JA-17 ... (truncated at 20 failures)")
                    return failures
    return failures


def _check_ja_15_audio_refs_on_disk() -> list[str]:
    """Every entry in data/audio_manifest.json must point to a file that
    exists on disk. Per data-correction brief §4.1: a release-blocker check
    for "no question, grammar pattern, listening item references a missing
    audio file." If this fails, the runtime app would 404 on `<audio src>`.

    The `skipped: true` flag in the manifest is a build-script status (file
    already on disk, skipped re-rendering); it does NOT mean missing.
    """
    failures: list[str] = []
    manifest_path = ROOT / "data" / "audio_manifest.json"
    if not manifest_path.exists():
        return ["JA-15: data/audio_manifest.json missing"]
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-15: could not parse audio_manifest.json: {e}"]
    items = manifest.get("items", [])
    for it in items:
        path = it.get("path", "")
        # Manifest paths use OS-native separators (Windows backslash on the
        # author machine). Normalise so the check works on any OS.
        rel = path.replace("\\", "/")
        full = ROOT / rel
        if not full.exists():
            failures.append(
                f"JA-15 manifest entry {it.get('id', '?')} points to missing file: {rel}"
            )
            if len(failures) >= 20:
                failures.append(f"JA-15 ... and more (truncated at 20)")
                break
    return failures


def _check_ja_18_reading_explanation_kanji() -> list[str]:
    """Every kanji that appears INSIDE a single-quoted phrase in an
    explanation_en field must also appear in the passage's `ja` text.

    The convention in data/reading.json is that explanations quote the
    relevant passage line in single quotes, like `'毎日 30どより 高いです'`.
    These quoted phrases must match the passage rendering exactly - if
    the passage uses うち (kana) but the explanation quotes 家 (kanji),
    the learner scans for the quoted phrase in the passage, fails to
    find it, and trust degrades.

    We deliberately do NOT check prompt_ja or unquoted explanation text,
    because prompts legitimately use question words (何, 人, etc.) that
    the passage doesn't echo. Only the quoted-quote-of-the-passage
    convention is checked.

    Originally Pass-15-reading audit §2.2: 8 explicit occurrences across
    n5.read.001 / 002 / 008 / 013 / 016 / 022. This invariant prevents
    recurrence and would have flagged all 8.
    """
    failures: list[str] = []
    reading_path = ROOT / "data" / "reading.json"
    if not reading_path.exists():
        return ["JA-18: data/reading.json missing"]
    try:
        data = json.loads(reading_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-18: parse error: {e}"]

    def is_kanji(ch: str) -> bool:
        cp = ord(ch)
        return 0x3400 <= cp <= 0x9FFF

    # Match either ASCII '...' or Japanese 「...」 quoted phrases.
    QUOTE_RE = re.compile(r"'([^']+)'|「([^」]+)」")

    for p in data.get("passages", []):
        pid = p.get("id", "?")
        passage_kanji = {ch for ch in p.get("ja", "") if is_kanji(ch)}
        for q in p.get("questions", []):
            qid = q.get("id", "?")
            text = q.get("explanation_en", "")
            for m in QUOTE_RE.finditer(text):
                quoted = m.group(1) or m.group(2) or ""
                for ch in quoted:
                    if is_kanji(ch) and ch not in passage_kanji:
                        failures.append(
                            f"JA-18 {qid} explanation quotes kanji "
                            f"'{ch}' not in passage {pid} "
                            f"(passage uses different rendering - "
                            f"likely kana). Quote: {quoted[:50]!r}"
                        )
                        if len(failures) >= 20:
                            failures.append("JA-18 ... (truncated at 20)")
                            return failures
                        break
    return failures


def _check_ja_20_reading_choices_kanji() -> list[str]:
    """Reading-question MCQ choices must use the same kana/kanji rendering
    as the passage. If a choice contains a kanji that the passage doesn't
    use, the learner cannot match-answer-to-text and answer authority is
    undermined. Pass-15-reading audit §5.4 (3).

    We accept the case where a choice contains a kanji that ALSO doesn't
    appear in the passage but is conceptually allowed (e.g., 何 or 人 in
    a question-style distractor). Specifically: only flag kanji that
    appear in the OTHER choices' rendering as kana - that's the strong
    inconsistency signal (n5.read.010 q1 had bare numbers but the passage
    used 〜こ counter; that's the class to catch).
    """
    failures: list[str] = []
    reading_path = ROOT / "data" / "reading.json"
    if not reading_path.exists():
        return ["JA-20: data/reading.json missing"]
    try:
        data = json.loads(reading_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-20: parse error: {e}"]

    def is_kanji(ch: str) -> bool:
        cp = ord(ch)
        return 0x3400 <= cp <= 0x9FFF

    for p in data.get("passages", []):
        pid = p.get("id", "?")
        passage_text = p.get("ja", "")
        passage_kanji = {ch for ch in passage_text if is_kanji(ch)}
        for q in p.get("questions", []):
            qid = q.get("id", "?")
            choices = q.get("choices", [])
            correct = q.get("correctAnswer", "")
            # The strong consistency check: every kanji in the
            # CORRECT-answer choice must appear in the passage. If the
            # correct answer rendering doesn't match the passage, the
            # question is inconsistent. (Distractors get more leeway -
            # they're foils.)
            for ch in correct:
                if is_kanji(ch) and ch not in passage_kanji:
                    failures.append(
                        f"JA-20 {qid} correctAnswer uses kanji '{ch}' "
                        f"not in passage {pid} (rendering mismatch - "
                        f"the answer should match the passage's form). "
                        f"correctAnswer={correct!r}"
                    )
                    if len(failures) >= 20:
                        failures.append("JA-20 ... (truncated at 20)")
                        return failures
                    break
    return failures


# Heuristic patterns that strongly suggest N4 grammar in a passage. Each
# entry: (regex, description). Used by JA-21. Patterns are deliberately
# conservative - false negatives are acceptable, false positives must be
# rare since they'd block release.
#
# Calibrated against the live corpus 2026-05-01:
# - Original "(?:う|く|...)と、" pattern false-fired on `3つと、` (counter
#   + と for noun-listing). Tightened with a negative-look-behind on
#   numerals + counter kana so only verb dictionary forms match.
N4_GRAMMAR_PATTERNS = [
    # 〜と conditional: verb-dictionary-form + と + comma. The verb-dict
    # ends in u-row hiragana. Exclude counter kana (つ/こ/本/etc.) which
    # are preceded by a digit or kanji numeral.
    # Negative-look-behind: NOT preceded by a digit, fullwidth digit, or
    # kanji numeral. Then a verb-dict-ending hiragana, then と + 、.
    (re.compile(
        r"(?<![0-9０-９一二三四五六七八九十百千万])"
        r"(?:う|く|ぐ|す|ぬ|ぶ|む|る)と、"),
     "〜と conditional (Verb-dict + と + comma)"),
    # Potential form -られ + ます/ません/た/ない (verb-2 potential).
    (re.compile(r"(?:け|げ|せ|て|ね|べ|め|れ)られ(?:ます|ません|た|ない)"),
     "Potential form (-られ + ます/ません/た/ない)"),
]


def _check_ja_21_n4_grammar_tier_flag() -> list[str]:
    """Reading passages that use heuristically-detected N4 grammar markers
    must be flagged with `tier: "late_n5"` (or be rewritten). Pass-15-
    reading audit §5.4 (4).

    Currently checks for: 〜と conditional, potential form. Both were
    audit-flagged on n5.read.030 / n5.read.007 and have been fixed. This
    check guards against re-introduction.
    """
    failures: list[str] = []
    reading_path = ROOT / "data" / "reading.json"
    if not reading_path.exists():
        return ["JA-21: data/reading.json missing"]
    try:
        data = json.loads(reading_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-21: parse error: {e}"]
    for p in data.get("passages", []):
        pid = p.get("id", "?")
        ja = p.get("ja", "")
        tier = p.get("tier", "core_n5")
        for pat, desc in N4_GRAMMAR_PATTERNS:
            m = pat.search(ja)
            if m and tier not in ("late_n5", "info_search"):
                failures.append(
                    f"JA-21 {pid} contains N4 grammar pattern "
                    f"({desc}) at {m.group(0)!r} but tier is "
                    f"{tier!r}. Either rewrite to N5 or set "
                    f"tier=late_n5."
                )
    return failures


def _check_ja_19_reading_info_search_format() -> list[str]:
    """Reading passages with `level: info-search` must declare a
    `format_type` (schedule_table | menu_list | notice | etc.) so the
    renderer can give them the appropriate visual treatment. Originally
    Pass-15-reading audit §3.6.
    """
    failures: list[str] = []
    reading_path = ROOT / "data" / "reading.json"
    if not reading_path.exists():
        return ["JA-19: data/reading.json missing"]
    try:
        data = json.loads(reading_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-19: parse error: {e}"]
    allowed_formats = {"schedule_table", "menu_list", "notice",
                       "signage", "form"}
    for p in data.get("passages", []):
        if p.get("level") != "info-search":
            continue
        pid = p.get("id", "?")
        if "format_type" not in p:
            failures.append(
                f"JA-19 {pid} (level=info-search) is missing the "
                f"`format_type` field. Allowed: {sorted(allowed_formats)}"
            )
            continue
        if p["format_type"] not in allowed_formats:
            failures.append(
                f"JA-19 {pid} format_type {p['format_type']!r} not in "
                f"allowed set {sorted(allowed_formats)}"
            )
    return failures


def _check_ja_22_kun_dedup() -> list[str]:
    """Every kanji's `kun` reading list in n5_kanji_readings.json must
    contain no duplicate entries. Pass-15 consolidated audit §2.2 found
    10 entries with repeats (二/七/分/見/聞/入/立/休/高/白) - artefacts
    of stripping okurigana.
    """
    failures: list[str] = []
    path = ROOT / "data" / "n5_kanji_readings.json"
    if not path.exists():
        return ["JA-22: data/n5_kanji_readings.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-22: parse error: {e}"]
    for k, entry in data.items():
        kun = entry.get("kun", [])
        if not isinstance(kun, list):
            continue
        if len(kun) != len(set(kun)):
            seen = set()
            dups = []
            for r in kun:
                if r in seen and r not in dups:
                    dups.append(r)
                seen.add(r)
            failures.append(
                f"JA-22 kanji '{k}' has duplicate kun reading(s): "
                f"{dups} (full list: {kun})"
            )
    return failures


def _check_ja_23_listening_script_choices_match() -> list[str]:
    """Listening utterance items that embed numbered choices in
    `script_ja` (style: '1. xxx\\n2. yyy\\n3. zzz') must have those exact
    strings in the `choices` array. Pass-15 consolidated audit §1.1
    found n5.listen.011 mismatch (script said 'ありがとう' but choices
    array said 'ありがとうございます') - a direct grading bug.

    We do not require items to embed numbered choices; this check only
    fires when items DO embed them and validates consistency.
    """
    failures: list[str] = []
    path = ROOT / "data" / "listening.json"
    if not path.exists():
        return ["JA-23: data/listening.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-23: parse error: {e}"]
    for item in data.get("items", []):
        script = item.get("script_ja", "")
        choices = item.get("choices", [])
        if not script or not choices:
            continue
        # Extract numbered lines: '1. xxx', '2. yyy' etc.
        embedded: list[str] = []
        for line in script.split("\n"):
            m = re.match(r"^\s*(\d+)\.\s*(.+?)\s*$", line)
            if m:
                embedded.append(m.group(2).strip())
        if not embedded:
            continue  # No embedded choices; nothing to check.
        # Each embedded choice must appear in the choices array.
        for emb in embedded:
            if emb not in choices:
                failures.append(
                    f"JA-23 {item.get('id','?')} embedded script choice "
                    f"{emb!r} not found in choices array {choices!r}"
                )
    return failures


def _check_ja_24_iadj_kanji_primary_kun() -> list[str]:
    """For kanji whose most common N5 use is as an i-adjective, the
    `primary` reading in n5_kanji_readings.json must be the kun-yomi.
    Pass-15 consolidated audit §2.1: 高/長/安 had on-yomi primaries
    that mis-rendered the high-frequency standalone use 高い/長い/安い.

    The list is curated; new i-adjective kanji can be added when N5
    scope shifts.
    """
    IADJ_KANJI_EXPECTED_KUN = {
        "高": "たか",
        "長": "なが",
        "安": "やす",
    }
    failures: list[str] = []
    path = ROOT / "data" / "n5_kanji_readings.json"
    if not path.exists():
        return ["JA-24: data/n5_kanji_readings.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-24: parse error: {e}"]
    for k, expected_kun in IADJ_KANJI_EXPECTED_KUN.items():
        if k not in data:
            continue  # JA-12 already catches missing entries.
        actual = data[k].get("primary")
        if actual != expected_kun:
            failures.append(
                f"JA-24 kanji '{k}' primary={actual!r} but most-common "
                f"N5 use is i-adjective {k}い (kun {expected_kun!r}). "
                f"Set primary={expected_kun!r}."
            )
    return failures


def _check_ja_25_whitelist_exceptions_documented() -> list[str]:
    """Pass-22 F-22.4: every kanji in n5_kanji_whitelist.json that is NOT in
    the official JLPT N5 scope MUST appear in
    data/n5_kanji_whitelist.exceptions.md with a `WHY:` justification.

    Without this guard, an agent or contributor could silently add an
    out-of-scope kanji to the whitelist to silence a JA-13 violation. With
    this guard, every exception is a deliberate, reviewable, justified
    addition.

    Spec: specifications/procedure-manual-appendix-c-pass22-polish.md C.4.

    The exceptions doc is OPTIONAL - if absent, the check passes (allows
    bootstrapping). Once present, every project-whitelist entry that is
    not in the official scope must be justified there.

    Bootstrapping the official-scope list: the N5 official scope is
    canonically 103 kanji per JLPT.jp (the project whitelist is 106 - the
    extra 3 are documented exceptions). For now we accept the project
    whitelist verbatim if the exceptions doc is absent; once the
    exceptions doc lands, validation kicks in.
    """
    failures: list[str] = []
    project_wl_path = ROOT / "data" / "n5_kanji_whitelist.json"
    exceptions_md_path = ROOT / "data" / "n5_kanji_whitelist.exceptions.md"
    official_scope_path = ROOT / "data" / "n5_official_kanji_scope.json"

    if not project_wl_path.exists():
        return ["JA-25: data/n5_kanji_whitelist.json missing"]
    if not exceptions_md_path.exists():
        # Bootstrapping mode: skip until the exceptions doc lands.
        return []
    if not official_scope_path.exists():
        # Without an official-scope reference we can't compute exceptions.
        # Skip but emit a one-time hint via a comment in the exceptions doc.
        return []

    try:
        project_wl = set(json.loads(project_wl_path.read_text(encoding="utf-8")))
        official = set(json.loads(official_scope_path.read_text(encoding="utf-8")))
    except Exception as e:
        return [f"JA-25: parse error loading whitelist or official scope: {e}"]

    md_text = exceptions_md_path.read_text(encoding="utf-8")
    # Parse: each line "- {KANJI}  WHY: <reason>" registers one exception.
    documented_with_why: dict[str, str] = {}
    for line in md_text.splitlines():
        line = line.strip()
        if not line.startswith("- "):
            continue
        # Skip non-kanji lines (e.g., docstring text starting with "- ").
        body = line[2:].strip()
        if not body or body[0] not in "一二三四五六七八九十百千万円木金土水火日月年人本中大小上下右左前後内外山川田":
            # Heuristic check: first char is a CJK kanji. Use a broader range.
            if not (0x4E00 <= ord(body[0]) <= 0x9FFF):
                continue
        # Pull off the kanji glyph (handles multi-char or trailing space).
        kanji = body[0]
        if "WHY:" in body:
            why = body.split("WHY:", 1)[1].strip()
            # Strip trailing "REVIEW_DATE:..." if present.
            if "REVIEW_DATE:" in why:
                why = why.split("REVIEW_DATE:", 1)[0].strip()
            documented_with_why[kanji] = why
        else:
            documented_with_why[kanji] = ""  # listed but no WHY

    project_exceptions = project_wl - official
    for kanji in sorted(project_exceptions):
        if kanji not in documented_with_why:
            failures.append(
                f"JA-25 kanji '{kanji}' is in n5_kanji_whitelist.json but not "
                f"in official JLPT N5 scope, AND not documented in "
                f"data/n5_kanji_whitelist.exceptions.md. Add an entry with WHY: <reason>."
            )
        elif not documented_with_why[kanji]:
            failures.append(
                f"JA-25 kanji '{kanji}' is documented in exceptions.md but "
                f"lacks WHY: justification. Add WHY: <reason> on the same line."
            )
    return failures


def _check_ja_26_no_duplicate_question_ids() -> list[str]:
    """Pass-23 (2026-05-02): no two entries in data/questions.json may
    share the same `id` field. JA-7 catches duplicate STEMS but not
    duplicate IDs - and the latter happened twice (Pass-16 ↔ Pass-15-P0
    over q-0454..q-0463; parallel-session ↔ Pass-16 over q-0479..q-0488).

    The runtime uses IDs as primary keys for storage and SRS state, so
    duplicate IDs cause one entry's progress to silently overwrite the
    other's. This invariant prevents the regression class going forward.
    """
    from collections import Counter
    failures: list[str] = []
    qpath = ROOT / "data" / "questions.json"
    if not qpath.exists():
        return ["JA-26: data/questions.json missing"]
    try:
        data = json.loads(qpath.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-26: parse error: {e}"]
    ids = [q.get("id") for q in data.get("questions", [])]
    counts = Counter(ids)
    dups = sorted(qid for qid, n in counts.items() if n > 1)
    for qid in dups:
        failures.append(
            f"JA-26 question id '{qid}' appears {counts[qid]} times in "
            f"data/questions.json. Run a dedup tool to renumber the "
            f"second occurrence."
        )
    return failures


def _check_ja_27_no_english_in_japanese_modules() -> list[str]:
    """Per user direction 2026-05-02: dokkai (data/reading.json) and
    listening (data/listening.json) are Japanese-first learner surfaces
    and must not carry English title or English-passage-translation
    fields, since both are rendered to the learner.

    Specifically banned at the item level:
      - title_en       (replaced by title_ja, used in renderIndex/renderRead/
                        renderQuestions/renderResults for reading and
                        renderListening/renderItem for listening)
      - translation_en (was rendered in a "Show English translation"
                        <details> panel on reading; that panel and field
                        are now removed from data + renderer)

    Allowed because they teach (NOT banned):
      - explanation_en (rationale shown after a wrong answer)
      - prompt_en      (legacy on listening; no items currently carry it)
      - any field on data/grammar.json (grammar pattern teaching genuinely
        needs English glosses - out of scope for this invariant)
      - any field on data/questions.json (question stems are short and
        sometimes carry translation_en for the learner; orthogonal to the
        passage-EN-translation rule we are locking down here)
    """
    BANNED = {"title_en", "translation_en"}
    failures: list[str] = []
    for fname, item_key in [
        ("data/reading.json",   "passages"),
        ("data/listening.json", "items"),
    ]:
        path = ROOT / fname
        if not path.exists():
            continue
        try:
            d = json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-27: {fname} parse error: {e}")
            continue
        items = d.get(item_key, [])
        for it in items:
            for bad in BANNED:
                if bad in it:
                    failures.append(
                        f"JA-27 {fname} item id={it.get('id','?')!r} carries "
                        f"banned field '{bad}'. The learner-facing surface "
                        f"in dokkai/listening must be Japanese only; remove "
                        f"the field (and migrate title_en -> title_ja if "
                        f"applicable). See tools/fix_dokkai_titles_remove_en.py "
                        f"and tools/fix_listening_titles_ja.py for prior "
                        f"migrations."
                    )
    return failures


def _load_kanji_exceptions_by_surface(surface: str) -> tuple[set, str | None]:
    """Read data/dokkai_kanji_exception.json (schema v2 as of 2026-05-08) and
    return the set of kanji allowed on the given `surface` ('dokkai-passage'
    or 'paper-distractor'). Returns (set, error-or-None).

    Schema v2 (current):
      {"_meta": {...}, "exception_kanji": [{"kanji": "x", "surfaces": [...]}, ...]}

    Schema v1 (legacy, still tolerated for transitional reads):
      {"_doc": [...], "exception_kanji": ["x", "y", ...]}  # all dokkai-only
    """
    try:
        doc = json.loads((ROOT / "data" / "dokkai_kanji_exception.json").read_text(encoding="utf-8"))
    except Exception as e:
        return set(), f"could not load dokkai_kanji_exception.json: {e}"
    raw = doc.get("exception_kanji", [])
    if not raw:
        return set(), None
    # Schema v2: list of objects with `kanji` + `surfaces`
    if isinstance(raw[0], dict):
        return ({e["kanji"] for e in raw if surface in (e.get("surfaces") or [])}, None)
    # Schema v1 fallback: flat list of strings, all treated as dokkai-passage
    if surface == "dokkai-passage":
        return (set(raw), None)
    return (set(), None)


def _check_ja_28_dokkai_kanji_bounded() -> list[str]:
    """data/papers/dokkai/*.json passages may contain non-N5 kanji ONLY
    if those kanji are explicitly documented in
    data/dokkai_kanji_exception.json with surface 'dokkai-passage'. This
    formalizes the dokkai naturalness exception so the runtime can't
    silently introduce a new non-N5 kanji into reading passages.

    Bunpou / moji / goi were NOT covered here historically — those are
    now enforced by JA-40 against the 'paper-distractor' surface.
    """
    failures: list[str] = []
    try:
        whitelist = set(json.loads(
            (ROOT / "data" / "n5_kanji_whitelist.json").read_text(encoding="utf-8")))
    except Exception as e:
        return [f"JA-28: could not load n5_kanji_whitelist.json: {e}"]
    exception_kanji, err = _load_kanji_exceptions_by_surface("dokkai-passage")
    if err:
        return [f"JA-28: {err}"]
    allowed = whitelist | exception_kanji
    KANJI_RE = re.compile(r"[一-鿿]")
    dokkai_dir = ROOT / "data" / "papers" / "dokkai"
    if not dokkai_dir.exists():
        return []
    offenders: dict[str, list[str]] = {}  # kanji -> sample passages
    for p in sorted(dokkai_dir.glob("*.json")):
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-28: parse error on {p.name}: {e}")
            continue
        def walk(obj):
            if isinstance(obj, str):
                for ch in obj:
                    if KANJI_RE.match(ch) and ch not in allowed:
                        offenders.setdefault(ch, []).append(
                            f"{p.name}: {obj[:60]!r}")
            elif isinstance(obj, dict):
                for v in obj.values(): walk(v)
            elif isinstance(obj, list):
                for v in obj: walk(v)
        walk(d)
    for ch, samples in sorted(offenders.items()):
        s = samples[0]
        failures.append(
            f"JA-28: kanji '{ch}' is neither in N5 catalog nor in "
            f"dokkai_kanji_exception.json (surface 'dokkai-passage'). "
            f"Either replace with kana, or add to "
            f"data/dokkai_kanji_exception.json with surfaces=['dokkai-passage'] "
            f"and a one-line reason. First seen: {s}")
    return failures


def _check_ja_40_paper_distractor_kanji_bounded() -> list[str]:
    """data/papers/{moji,goi,bunpou}/*.json questions may contain non-N5
    kanji ONLY if those kanji are documented in
    data/dokkai_kanji_exception.json with surface 'paper-distractor'.
    Closes the audit gap (ISSUE-007, 2026-05-08): historically the
    'moji-corpus distractor exception' was only documented in question
    rationales, with no automated enforcement. This invariant makes
    every non-whitelist kanji used as a distractor explicit and
    machine-checkable.

    Note: this walks all four fields commonly carrying displayed text:
    stem_html, choices, correctAnswer, prompt. Rationale prose
    (`rationale`) is intentionally NOT walked — that's where the human
    explanation lives and may freely use non-N5 kanji.
    """
    failures: list[str] = []
    try:
        whitelist = set(json.loads(
            (ROOT / "data" / "n5_kanji_whitelist.json").read_text(encoding="utf-8")))
    except Exception as e:
        return [f"JA-40: could not load n5_kanji_whitelist.json: {e}"]
    exception_kanji, err = _load_kanji_exceptions_by_surface("paper-distractor")
    if err:
        return [f"JA-40: {err}"]
    allowed = whitelist | exception_kanji
    KANJI_RE = re.compile(r"[一-鿿]")
    LEARNER_FIELDS = ("stem_html", "choices", "correctAnswer", "prompt", "title_ja", "prompt_ja")
    offenders: dict[str, list[str]] = {}  # kanji -> sample
    for sub in ("moji", "goi", "bunpou"):
        pdir = ROOT / "data" / "papers" / sub
        if not pdir.exists():
            continue
        for p in sorted(pdir.glob("paper-*.json")):
            try:
                d = json.loads(p.read_text(encoding="utf-8"))
            except Exception as e:
                failures.append(f"JA-40: parse error on {sub}/{p.name}: {e}")
                continue
            for q in d.get("questions", []) or []:
                qid = q.get("id", "?")
                for field in LEARNER_FIELDS:
                    val = q.get(field)
                    if val is None:
                        continue
                    text = " ".join(val) if isinstance(val, list) else str(val)
                    for ch in text:
                        if KANJI_RE.match(ch) and ch not in allowed:
                            offenders.setdefault(ch, []).append(
                                f"{sub}/{p.name}#{qid}.{field}")
    for ch, samples in sorted(offenders.items()):
        # Show first 3 occurrences for context
        sample_list = ", ".join(samples[:3]) + (f" (+{len(samples)-3} more)" if len(samples) > 3 else "")
        failures.append(
            f"JA-40: kanji '{ch}' is neither in N5 catalog nor in "
            f"dokkai_kanji_exception.json (surface 'paper-distractor'). "
            f"Either replace with kana, or add to "
            f"data/dokkai_kanji_exception.json with "
            f"surfaces=['paper-distractor'] and a one-line reason. "
            f"Occurs in: {sample_list}")
    return failures


def _check_ja_30_provenance() -> list[str]:
    """Past-paper provenance scan: no question text may contain markers
    that suggest direct copying from JLPT past papers (JEES citations,
    year-numbered exam markers, "過去問" / "真題" terminology, etc.).

    Mirrors the standalone tool tools/audit_provenance.py - kept inline
    here so the standard CI integrity check (one command) catches a
    leak without needing a separate workflow step. See CONTENT-LICENSE.md
    §3 for the policy this enforces.

    Failure here means: a contributor introduced text that could be
    interpreted as a past-paper reference. Fix by re-authoring the
    stem/rationale/note in the project's own voice, not by exempting
    the rule.
    """
    failures: list[str] = []
    SUSPECT = [
        (re.compile(r"\bJEES\b"), "JEES citation"),
        (re.compile(r"Japan\s*Educational\s*Exchanges?", re.I), "JEES full name"),
        (re.compile(r"(19|20)\d{2}年[\s　]*[1-9七十二]+月.*?(本試験|公開|JLPT)"),
         "year+month past-paper marker"),
        (re.compile(r"本試験[\s　]*第\d+回"), "past-paper round number"),
        (re.compile(r"実試験|実問題|真題"), "past-paper terminology"),
        (re.compile(r"(JLPT|日本語能力試験)\s*N[1-5]\s*(20\d{2}|19\d{2})年"),
         "JLPT year-paper citation"),
        (re.compile(r"過去問"), "kakomon (past-question) self-attestation"),
    ]

    def _scan(text: str, where: str) -> None:
        if not isinstance(text, str) or not text:
            return
        for pat, why in SUSPECT:
            m = pat.search(text)
            if m:
                failures.append(
                    f"JA-30 {where}: {why} - '{m.group(0)[:60]}'"
                )

    qpath = ROOT / "data" / "questions.json"
    if qpath.exists():
        try:
            data = json.loads(qpath.read_text(encoding="utf-8"))
        except Exception as e:
            return [f"JA-30: parse error questions.json: {e}"]
        for qq in data.get("questions", []):
            qid = qq.get("id", "?")
            for f in ("question_ja", "prompt_ja", "explanation_en", "rationale"):
                _scan(qq.get(f, ""), f"{qid}.{f}")
            de = qq.get("distractor_explanations") or {}
            if isinstance(de, dict):
                for k, v in de.items():
                    _scan(v, f"{qid}.distractor_explanations.{k}")

    papers_dir = ROOT / "data" / "papers"
    if papers_dir.exists():
        for f in papers_dir.rglob("paper-*.json"):
            try:
                p = json.loads(f.read_text(encoding="utf-8"))
            except Exception:
                continue
            for qq in p.get("questions", []):
                qid = qq.get("id", "?")
                for fld in ("stem_html", "rationale", "explanation_en",
                            "passage_text", "prompt_ja"):
                    _scan(qq.get(fld, ""), f"{qid}.{fld}")

    return failures


def _check_ja_31_vocab_pos_parity() -> list[str]:
    """The PoS tags inline in KnowledgeBank/vocabulary_n5.md
    (`[n.]` / `[v1]` / etc., added 2026-05-02 per DEFER-5) must agree
    with the `pos` field on the corresponding entry in data/vocab.json.

    Drift class this catches:
      - A contributor edits one file (say, fixes a wrong PoS in the
        markdown) but forgets the JSON, or vice versa. Since both
        files describe the same per-form attribute, drift is silent.
      - The PoS-injection pass (DEFER-5) had two homograph entries
        get the wrong tag (`いる - to need` mistagged `[v2]` instead
        of `[v1]`); JA-31 would have caught it before commit.

    Match strategy is *section-aware* because many forms are
    homographs (e.g., 'はる' is a noun in §14 Weather but a verb-1
    in §27 Verbs; 'いる' is verb-2 in §28 but verb-1 in §30). The
    audit tracks the current `## N. Section title` heading while
    parsing the markdown and looks up `(form, reading, section)` in
    JSON for an exact match; if the section doesn't match exactly,
    word-overlap with the JSON's `section` field picks the closest.
    Form-only fallback (any pos in JSON for that form) is the last
    resort.
    """
    POS_ABBREV = {
        "noun": "n.", "verb-1": "v1", "verb-2": "v2", "verb-3": "v3",
        "i-adj": "i-adj", "na-adj": "na-adj", "adverb": "adv.",
        "particle": "part.", "conjunction": "conj.", "pronoun": "pron.",
        "counter": "count.", "numeral": "num.", "demonstrative": "dem.",
        "question-word": "Q-word", "expression": "exp.",
        "interjection": "interj.",
    }

    failures: list[str] = []
    md_path = ROOT / "KnowledgeBank" / "vocabulary_n5.md"
    json_path = ROOT / "data" / "vocab.json"
    if not json_path.exists():
        return ["JA-31: data/vocab.json missing"]
    if not md_path.exists():
        # KB deleted 2026-05-14 (merged into data/ + docs/N5-syllabus-methodology.md
        # as a single source of truth). The PoS-drift class this check guarded
        # against (KB-MD vs JSON drift) is gone now that PoS lives only in
        # data/vocab.json. The check becomes a no-op rather than failing.
        return []

    try:
        vocab = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-31: parse error vocab.json: {e}"]

    def normalize_section(s: str) -> str:
        s = re.sub(r"^\d+\.\s*", "", s).strip().lower()
        return re.sub(r"\s+", " ", s)

    # Build two indexes:
    #   by_full[(form, reading, section_norm)]      = pos_tag (exact match)
    #   by_form_reading[(form, reading)]            = list of (section_norm, pos_tag)
    #   by_form[form]                               = set of pos_tags (any-of fallback)
    # by_full now stores a SET of valid tags per (form, reading, section) key.
    # Homograph case in section 30: `いる` (exist, v2) and `いる` (to need, v1)
    # share the key, and either tag should pass parity for that key.
    # Without this, the dict-based "last write wins" returned only one tag and
    # JA-31 would flag the other MD line as a mismatch despite both being correct.
    by_full: dict[tuple[str, str, str], set[str]] = {}
    by_form_reading: dict[tuple[str, str], list[tuple[str, str]]] = {}
    by_form: dict[str, set[str]] = {}
    for e in vocab.get("entries", []):
        form = e.get("form")
        reading = e.get("reading", form)
        pos = e.get("pos")
        section = e.get("section", "")
        if not (form and pos):
            continue
        tag = POS_ABBREV.get(pos, pos)
        sec_norm = normalize_section(section)
        by_full.setdefault((form, reading, sec_norm), set()).add(tag)
        by_form_reading.setdefault((form, reading), []).append((sec_norm, tag))
        by_form.setdefault(form, set()).add(tag)

    POS_TAGS_RE = "|".join(re.escape(t) for t in
        ["n.", "v1", "v2", "v3", "i-adj", "na-adj", "adv.", "part.",
         "conj.", "pron.", "count.", "num.", "dem.", "Q-word", "exp.",
         "interj."])
    LINE_RE = re.compile(
        r"^(- )([^\s\(]+)(\s+\(([^)]+)\))?( \[(?:Ext|Cul)\])?(\s*-\s*)"
        rf"\[({POS_TAGS_RE})\]\s+(.+)$"
    )
    SECTION_HEADER_RE = re.compile(r"^##\s+(.+?)\s*$")
    is_jp = re.compile(r"^[ぁ-んァ-ヶー一-鿿]")

    text = md_path.read_text(encoding="utf-8")
    current_section_norm = ""
    line_no = 0
    for raw in text.splitlines():
        line_no += 1
        sh = SECTION_HEADER_RE.match(raw)
        if sh:
            current_section_norm = normalize_section(sh.group(1))
            continue
        m = LINE_RE.match(raw)
        if not m:
            continue
        form = m.group(2)
        reading = m.group(4) or form
        md_tag = m.group(7)
        if not is_jp.match(form):
            continue
        # Lookup priority: exact (form, reading, section) → word-overlap
        # tiebreaker on (form, reading) → form-only any-of fallback.
        # by_full holds a SET (homograph case e.g. section-30 いる); MD tag
        # passes if it's a member of the set for the matched key.
        full_set = by_full.get((form, reading, current_section_norm))
        expected: str | None
        if full_set:
            if md_tag in full_set:
                continue  # MD tag matches one of the valid POS for this key
            # Otherwise emit a comma-list of expected values for the error
            expected = "/".join(sorted(full_set))
        else:
            candidates = by_form_reading.get((form, reading), [])
            if len(candidates) == 1:
                expected = candidates[0][1]
            elif candidates:
                cur_words = set(current_section_norm.split())
                best = None; best_score = -1
                for sec, tag in candidates:
                    score = len(cur_words & set(sec.split()))
                    if score > best_score:
                        best, best_score = tag, score
                expected = best
            else:
                expected = None
        if expected is None:
            allowed = by_form.get(form)
            if allowed is None:
                failures.append(
                    f"JA-31 vocabulary_n5.md:{line_no} entry '{form}' "
                    f"({reading}) [{md_tag}] has no matching record in "
                    f"data/vocab.json"
                )
                continue
            if md_tag in allowed:
                continue
            expected_str = "/".join(sorted(allowed))
            failures.append(
                f"JA-31 vocabulary_n5.md:{line_no} '{form}' ({reading}) "
                f"tagged [{md_tag}] but data/vocab.json says "
                f"'{expected_str}'"
            )
            continue
        if md_tag != expected:
            failures.append(
                f"JA-31 vocabulary_n5.md:{line_no} '{form}' ({reading}) "
                f"tagged [{md_tag}] in section '{current_section_norm}' "
                f"but data/vocab.json says '{expected}'"
            )
    return failures


def _check_ja_32_paper_rationale_md_parity() -> list[str]:
    """Paper-JSON rationale kanji must all appear in the source MD Q-block.

    Drift class this catches:
      - KB markdown rationale is corrected to use kana (e.g., 熱がある ->
        ねつが ある for kanji-scope) but the paper extractor wasn't
        re-run, so the paper JSON still carries the stale kanji. The
        auditor's 2026-05-04 flag of paper-2 Q19 ("熱がある" in JSON,
        but MD says "ねつが ある") is the canonical example.

    Method: for each paper JSON in data/papers/{moji,goi,bunpou,dokkai}/
    paper-*.json with a kbSourceId, locate the corresponding MD Q-block.
    For each kanji in the JSON rationale, verify it also appears
    somewhere in the MD Q-block (stem, choices, answer line, or notes).

    This narrow check avoids false-positives on AUTHORED rationales
    (e.g., bunpou-5/6 sentence-rearrange where the rationale was
    expanded during audit fixes) - authored rationales reuse kanji
    that were already in the MD's stem / choices, so they pass. But a
    stale-extracted rationale that uses a kanji the MD has corrected
    to kana fails immediately.

    A complementary check (kana-only JSON rationale where MD has
    kanji) is intentionally NOT enforced - kana is always permissible
    at N5 level; only stale kanji that contradicts a kana-only MD
    surface as drift.
    """
    failures: list[str] = []
    PAPERS_DIR = ROOT / "data" / "papers"
    if not PAPERS_DIR.exists():
        return failures
    KB_BY_CATEGORY = {
        'moji':   KB / 'moji_questions_n5.md',
        'goi':    KB / 'goi_questions_n5.md',
        'bunpou': KB / 'bunpou_questions_n5.md',
        'dokkai': KB / 'dokkai_questions_n5.md',
    }
    KANJI_RE_LOCAL = re.compile(r'[一-鿿]')

    md_cache: dict[str, str] = {}
    for paper_path in sorted(PAPERS_DIR.rglob('paper-*.json')):
        category = paper_path.parent.name
        kb_path = KB_BY_CATEGORY.get(category)
        if kb_path is None or not kb_path.exists():
            continue
        if category not in md_cache:
            md_cache[category] = kb_path.read_text(encoding='utf-8')
        md = md_cache[category]

        try:
            paper = json.loads(paper_path.read_text(encoding='utf-8'))
        except Exception as e:
            failures.append(f"JA-32 {paper_path}: parse error: {e}")
            continue
        for q in paper.get('questions', []):
            rat = q.get('rationale')
            kb_q = q.get('kbSourceId')
            if not rat or not kb_q:
                continue
            block_re = re.compile(
                rf'### {re.escape(kb_q)}\b([\s\S]+?)(?=\n### Q\d|\Z)'
            )
            m = block_re.search(md)
            if not m:
                failures.append(
                    f"JA-32 {paper_path.name} {q.get('id', '?')}: "
                    f"kbSourceId={kb_q} has no matching block in {kb_path.name}"
                )
                continue
            block_text = m.group(0)
            json_kanji = set(KANJI_RE_LOCAL.findall(rat))
            md_kanji = set(KANJI_RE_LOCAL.findall(block_text))
            stale = json_kanji - md_kanji
            if stale:
                failures.append(
                    f"JA-32 {paper_path.name} {q.get('id', '?')} "
                    f"({kb_q}): rationale uses kanji not in MD source: "
                    f"{sorted(stale)} - possible stale extraction "
                    f"(MD may have corrected to kana)"
                )
    return failures


def _check_ja_33_listening_mondai_taxonomy() -> list[str]:
    """ISSUE-016 (audit round 3, 2026-05-05): every listening item must
    carry `mondai` (1..4) and `format_type` from the closed enum:
      task_understanding   (mondai 1, 課題理解)
      point_understanding  (mondai 2, ポイント理解)
      utterance_expression (mondai 3, 発話表現)
      immediate_response   (mondai 4, 即時応答)

    Without this taxonomy the UI cannot surface the official JLPT
    section structure and the test-mode listening flow cannot mirror the
    real exam's per-mondai weighting.
    """
    failures: list[str] = []
    p = ROOT / "data" / "listening.json"
    if not p.exists():
        return ["JA-33: data/listening.json missing"]
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-33: parse error: {e}"]

    ALLOWED_FT = {
        "task_understanding",
        "point_understanding",
        "utterance_expression",
        "immediate_response",
    }
    FT_TO_MONDAI = {
        "task_understanding":   1,
        "point_understanding":  2,
        "utterance_expression": 3,
        "immediate_response":   4,
    }

    for it in data.get("items", []):
        iid = it.get("id", "?")
        m = it.get("mondai")
        ft = it.get("format_type")
        if not isinstance(m, int) or m not in (1, 2, 3, 4):
            failures.append(f"JA-33 {iid}: mondai is missing or not in {{1,2,3,4}} (got {m!r})")
        if ft not in ALLOWED_FT:
            failures.append(f"JA-33 {iid}: format_type {ft!r} not in closed enum {sorted(ALLOWED_FT)}")
        elif FT_TO_MONDAI.get(ft) != m:
            failures.append(
                f"JA-33 {iid}: mondai={m} inconsistent with format_type={ft!r} "
                f"(expected mondai={FT_TO_MONDAI[ft]})"
            )
    return failures


def _check_ja_35_review_status() -> list[str]:
    """ISSUE-030 (audit round 4, 2026-05-05) + BUG-012 (2026-05-16):
    every content item across all 5 corpora must carry `review_status`
    from a closed enum so the UI can render an honest provenance badge.

    BUG-012 update: the value formerly known as `native_reviewed` is
    renamed to `ai_quality_reviewed`. The label was assigned by Claude
    acting as a native-reviewer persona — not by an actual native
    human Japanese teacher. The rename disambiguates at the point of
    use so downstream consumers don't display a "Native-reviewed"
    badge implying human review. A future human-native review pass
    would introduce a separate `human_native_reviewed` value.

    `native_reviewed` is retained in the enum as a transitional
    accepted value during the migration window (any item carrying
    it is a regression — the fix script should have renamed all
    instances).
    """
    failures: list[str] = []
    ALLOWED = {"ai_quality_reviewed", "llm_curated", "auto_generated", "native_reviewed"}
    targets = [
        ("data/grammar.json",   "patterns"),
        ("data/vocab.json",     "entries"),
        ("data/kanji.json",     "entries"),
        ("data/reading.json",   "passages"),
        ("data/listening.json", "items"),
    ]
    for fname, key in targets:
        p = ROOT / fname
        if not p.exists():
            failures.append(f"JA-35: {fname} missing")
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-35: {fname} parse error: {e}")
            continue
        for item in data.get(key, []):
            iid = item.get("id") or item.get("glyph") or item.get("form") or "?"
            rs = item.get("review_status")
            if not rs:
                failures.append(f"JA-35 {fname} {iid}: missing review_status")
            elif rs not in ALLOWED:
                failures.append(f"JA-35 {fname} {iid}: review_status {rs!r} not in closed enum {sorted(ALLOWED)}")
    return failures


def _check_ja_34_core_late_split() -> list[str]:
    """ISSUE-033 (audit round 4, 2026-05-05) + ISSUE-005 (2026-05-14):
    data/n5_core_pattern_ids.json enumerates the strict-N5 vs late-N5
    (borderline N5/N4) vs deferred-to-N4 pattern IDs. Must agree
    byte-for-byte with the live grammar.json tier field, so the home
    count "178 patterns (153 core + 20 late-N5 + 5 deferred)" stays
    honest as content evolves.

    Schema update 2026-05-14: late_n5 became array-of-objects
    {id, rationale, sources_n5, sources_n4}. Plus new deferred_to_n4
    flat array. This check accepts both old (flat string) and new
    (object) shape for late_n5; deferred_to_n4 is optional (treated
    as empty if absent).
    """
    failures: list[str] = []
    g_path = ROOT / "data" / "grammar.json"
    w_path = ROOT / "data" / "n5_core_pattern_ids.json"
    if not g_path.exists():
        return ["JA-34: data/grammar.json missing"]
    if not w_path.exists():
        return ["JA-34: data/n5_core_pattern_ids.json missing"]
    try:
        grammar = json.loads(g_path.read_text(encoding="utf-8"))
        whitelist = json.loads(w_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-34: parse error: {e}"]

    core_actual = sorted(p["id"] for p in grammar.get("patterns", [])
                         if (p.get("tier", "core_n5") or "core_n5") == "core_n5")
    late_actual = sorted(p["id"] for p in grammar.get("patterns", [])
                         if p.get("tier") == "late_n5")
    deferred_actual = sorted(p["id"] for p in grammar.get("patterns", [])
                             if p.get("tier") == "deferred_to_n4")
    core_listed = sorted(whitelist.get("core_n5", []))
    # late_n5 may be flat strings (legacy) or objects (post 2026-05-14)
    late_raw = whitelist.get("late_n5", [])
    if late_raw and isinstance(late_raw[0], dict):
        late_listed = sorted(item.get("id", "") for item in late_raw)
    else:
        late_listed = sorted(late_raw)
    # deferred_to_n4 may be flat strings (legacy) or objects (post 2026-05-14
    # merge of standalone n5_deferred_to_n4.json into this file).
    deferred_raw = whitelist.get("deferred_to_n4", [])
    if deferred_raw and isinstance(deferred_raw[0], dict):
        deferred_listed = sorted(item.get("id", "") for item in deferred_raw)
    else:
        deferred_listed = sorted(deferred_raw)

    if core_actual != core_listed:
        only_actual = set(core_actual) - set(core_listed)
        only_listed = set(core_listed) - set(core_actual)
        if only_actual:
            failures.append(f"JA-34: core_n5 missing from whitelist: {sorted(only_actual)[:5]}")
        if only_listed:
            failures.append(f"JA-34: core_n5 in whitelist but not in grammar.json (drift): {sorted(only_listed)[:5]}")
    if late_actual != late_listed:
        only_actual = set(late_actual) - set(late_listed)
        only_listed = set(late_listed) - set(late_actual)
        if only_actual:
            failures.append(f"JA-34: late_n5 missing from whitelist: {sorted(only_actual)[:5]}")
        if only_listed:
            failures.append(f"JA-34: late_n5 in whitelist but not in grammar.json: {sorted(only_listed)[:5]}")
    if deferred_actual != deferred_listed:
        only_actual = set(deferred_actual) - set(deferred_listed)
        only_listed = set(deferred_listed) - set(deferred_actual)
        if only_actual:
            failures.append(f"JA-34: deferred_to_n4 missing from whitelist: {sorted(only_actual)[:5]}")
        if only_listed:
            failures.append(f"JA-34: deferred_to_n4 in whitelist but not in grammar.json: {sorted(only_listed)[:5]}")
    expected_total = whitelist.get("totalCount")
    actual_total = len(grammar.get("patterns", []))
    if expected_total != actual_total:
        failures.append(f"JA-34: totalCount {expected_total} != live grammar.json count {actual_total}")
    return failures


def _check_ja_29_subtype_taxonomy() -> list[str]:
    """Question subtypes are a closed taxonomy. New subtypes must be added
    here AND get explicit renderer support before shipping; an unknown
    subtype slipping through means the renderer treats it as a plain mcq
    silently, which masks design intent.

    Closing DEFER-2 (Pass-23 r5, 2026-05-02): rather than promote
    paraphrase to a top-level question type (which would require renderer
    changes for marginal gain), we lock the subtype field as the canonical
    extension point. Any new subtype must be registered here.
    """
    ALLOWED_SUBTYPES = {"paraphrase", "kanji_writing"}
    failures: list[str] = []
    qpath = ROOT / "data" / "questions.json"
    if not qpath.exists():
        return ["JA-29: data/questions.json missing"]
    try:
        data = json.loads(qpath.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-29: parse error: {e}"]
    for q in data.get("questions", []):
        sub = q.get("subtype")
        if sub is None:
            continue
        if sub not in ALLOWED_SUBTYPES:
            failures.append(
                f"JA-29 question id '{q.get('id', '?')}' has unknown "
                f"subtype '{sub}'. Allowed: {sorted(ALLOWED_SUBTYPES)}. "
                f"Register the subtype in tools/check_content_integrity.py "
                f"and confirm renderer support before adding new instances."
            )
    return failures


def _check_ja_36_answer_position_balance() -> list[str]:
    """ISSUE-061 (audit round-7, 2026-05-06): correct-answer position
    distribution must be within ±10 percentage points of an even 25/25/25/25
    split per corpus. Catches authoring drift toward "always answer-1
    syndrome" that runtime shuffle in test/drill/diagnostic masks but
    paper builders don't.

    Tolerance is 10pp (so any single position 15-35% is OK). Tighter
    later if needed. Soft-fails per-corpus, listing the offending
    distribution for transparency.
    """
    from collections import Counter
    failures: list[str] = []
    TOLERANCE_PP = 10.0

    targets: list[tuple[str, "Counter[int]", int]] = []
    # 1) data/questions.json - has 'choices' + 'correctAnswer' (string match)
    qpath = ROOT / "data" / "questions.json"
    if qpath.exists():
        try:
            data = json.loads(qpath.read_text(encoding="utf-8"))
            counts = Counter()
            total = 0
            for q in data.get("questions", []):
                cs = q.get("choices", [])
                ca = q.get("correctAnswer")
                if isinstance(cs, list) and ca in cs:
                    counts[cs.index(ca)] += 1
                    total += 1
            if total > 0:
                targets.append(("data/questions.json", counts, total))
        except Exception as e:
            failures.append(f"JA-36: questions.json parse error: {e}")

    # 2) Per-paper-category aggregation - papers carry 'correctIndex'
    papers_dir = ROOT / "data" / "papers"
    if papers_dir.exists():
        for cat_dir in sorted(papers_dir.iterdir()):
            if not cat_dir.is_dir():
                continue
            counts = Counter()
            total = 0
            for pf in sorted(cat_dir.glob("*.json")):
                try:
                    p = json.loads(pf.read_text(encoding="utf-8"))
                except Exception:
                    continue
                for q in p.get("questions", []):
                    ci = q.get("correctIndex")
                    if isinstance(ci, int):
                        counts[ci] += 1
                        total += 1
            if total > 0:
                targets.append((f"data/papers/{cat_dir.name}/", counts, total))

    for name, counts, total in targets:
        # Compute percentages and check tolerance
        ideal = 25.0
        for pos in (0, 1, 2, 3):
            pct = 100.0 * counts.get(pos, 0) / total
            if abs(pct - ideal) > TOLERANCE_PP:
                pct_str = ", ".join(
                    f"pos{p}={100.0 * counts.get(p, 0) / total:.1f}%"
                    for p in (0, 1, 2, 3)
                )
                failures.append(
                    f"JA-36 {name} (n={total}): position {pos} at {pct:.1f}% "
                    f"violates ±{TOLERANCE_PP}pp from 25%. Distribution: {pct_str}. "
                    f"Either rebalance (per-item correctAnswer rotation) or "
                    f"document the runtime-shuffle compensation explicitly."
                )
                break  # one failure per corpus is enough

    return failures


def _check_ja_37_namespace_doc_parity() -> list[str]:
    """ISSUE-065 (audit round-7, 2026-05-06): the localStorage namespace
    used in js/storage.js must appear verbatim in PRIVACY.md. Closes the
    drift vector that triggered round-3 ISSUE-024 and the niche-N2
    privacy-positioning concern.
    """
    failures: list[str] = []
    storage_js = ROOT / "js" / "storage.js"
    privacy_md = ROOT / "PRIVACY.md"
    if not storage_js.exists():
        return ["JA-37: js/storage.js missing"]
    if not privacy_md.exists():
        return ["JA-37: PRIVACY.md missing"]

    code = storage_js.read_text(encoding="utf-8")
    privacy = privacy_md.read_text(encoding="utf-8")

    # Extract namespace prefixes from localStorage.setItem / getItem calls.
    # Pattern: localStorage.setItem('<namespace>:<key>', ...) or template
    # literals starting with the same prefix.
    namespace_re = re.compile(
        r"""localStorage\.(?:setItem|getItem|removeItem)\(\s*['"`]([^'"`:]+:)"""
    )
    namespaces = set(namespace_re.findall(code))
    if not namespaces:
        # No localStorage usage extracted. Treat as PASS (storage layer might
        # have been refactored to a different mechanism).
        return []

    # Each extracted namespace must appear verbatim in PRIVACY.md.
    for ns in sorted(namespaces):
        if ns not in privacy:
            failures.append(
                f"JA-37 namespace {ns!r} used in js/storage.js but not "
                f"documented verbatim in PRIVACY.md. Niche-N2 privacy-doc "
                f"drift; either update PRIVACY.md or change the namespace."
            )
    return failures


def _check_ja_38_common_mistakes_floor() -> list[str]:
    """ISSUE-068 (audit round-7, 2026-05-06): every grammar pattern must
    carry ≥1 common_mistakes entry. Pedagogical floor - patterns at zero
    drop below the per-pattern depth bar the audit-prompt enforces.

    Soft-fails (lists patterns at zero) so authoring waves can drive the
    count down to zero progressively.
    """
    failures: list[str] = []
    g_path = ROOT / "data" / "grammar.json"
    if not g_path.exists():
        return ["JA-38: data/grammar.json missing"]
    try:
        data = json.loads(g_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-38: parse error: {e}"]

    zero_patterns = []
    for p in data.get("patterns", []):
        cms = p.get("common_mistakes")
        if not isinstance(cms, list) or len(cms) == 0:
            zero_patterns.append(p.get("id", "?"))

    if zero_patterns:
        failures.append(
            f"JA-38: {len(zero_patterns)} grammar patterns have zero "
            f"common_mistakes entries. Author at least one specific "
            f"common-mistake per pattern. Patterns: {zero_patterns[:10]}"
            + ("..." if len(zero_patterns) > 10 else "")
        )
    return failures


def _check_ja_39_locale_set_en_hi() -> list[str]:
    """IMP-096 (locale transition 2026-05-06): the supported locale set
    is exactly {en, hi}. The four deprecated locales (vi, id, ne, zh)
    must not appear as translation-field suffixes anywhere in
    grammar.json / vocab.json / kanji.json / reading.json /
    listening.json / questions.json or in any paper-pack JSON. The
    js/i18n.js SUPPORTED list must also be exactly the literal
    ['en', 'hi'].

    locales/ folder must contain exactly en.json + hi.json - no
    additional files, no missing files.
    """
    failures: list[str] = []

    # 1) locales/ folder check
    loc_dir = ROOT / "locales"
    if loc_dir.exists():
        json_files = sorted(p.name for p in loc_dir.glob("*.json"))
        expected = ["en.json", "hi.json"]
        if json_files != expected:
            failures.append(
                f"JA-39: locales/ contains {json_files}; expected {expected}"
            )

    # 2) Content-data scan - no <basename>_<lc> for lc in deprecated
    DEPRECATED = {"vi", "id", "ne", "zh"}
    locale_field_re = re.compile(
        r"^(?:meaning|meanings|explanation|gloss|title|prompt|description|note)_"
        r"(?P<lc>[a-z]{2})$"
    )

    def walk_check(node, path):
        if isinstance(node, dict):
            for k, v in node.items():
                m = locale_field_re.match(k or "")
                if m and m.group("lc") in DEPRECATED:
                    failures.append(
                        f"JA-39 deprecated-locale field at {path}.{k} "
                        f"(post-IMP-096 narrowing requires en/hi only)"
                    )
                # l1_notes / false_friends container check
                if k in {"l1_notes", "false_friends"} and isinstance(v, dict):
                    for subk in v:
                        if subk in DEPRECATED:
                            failures.append(
                                f"JA-39 deprecated-locale key at {path}.{k}.{subk}"
                            )
                walk_check(v, f"{path}.{k}")
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk_check(v, f"{path}[{i}]")

    targets = [
        "data/grammar.json", "data/vocab.json", "data/kanji.json",
        "data/reading.json", "data/listening.json", "data/questions.json",
    ]
    for fname in targets:
        p = ROOT / fname
        if not p.exists():
            continue
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            walk_check(d, fname)
        except Exception as e:
            failures.append(f"JA-39 parse error on {fname}: {e}")

    # Papers
    papers_dir = ROOT / "data" / "papers"
    if papers_dir.exists():
        for pf in papers_dir.rglob("*.json"):
            try:
                d = json.loads(pf.read_text(encoding="utf-8"))
                walk_check(d, str(pf.relative_to(ROOT)).replace("\\", "/"))
            except Exception as e:
                failures.append(f"JA-39 parse error on {pf.name}: {e}")

    # 3) js/i18n.js SUPPORTED list literal - best-effort regex match
    i18n = ROOT / "js" / "i18n.js"
    if i18n.exists():
        src = i18n.read_text(encoding="utf-8")
        m = re.search(r"const\s+SUPPORTED\s*=\s*\[([^\]]+)\]", src)
        if m:
            tokens = re.findall(r"'([^']+)'", m.group(1))
            if sorted(tokens) != sorted(["en", "hi"]):
                failures.append(
                    f"JA-39 js/i18n.js SUPPORTED is {tokens}; expected ['en', 'hi']"
                )
    return failures[:30]  # cap noise


def _check_ja_41_kana_prefix_convention() -> list[str]:
    """Cycle-2 Phase 2 (2026-05-07): Hindi prose must follow the kana-prefix
    convention. Japanese grammatical particles attached to Hindi grammatical
    terms must be in Japanese script (hiragana), NOT Latin romaji and NOT
    Devanagari transliteration.

    Wrong:   ना-विशेषण  (Devanagari な)
    Wrong:   na-विशेषण  (Latin romaji)
    Right:   な-विशेषण  (hiragana)

    Same applies to い-/て-/た-/ない-/ます-/たい-/ば- prefixes.

    The rule applies SYMMETRICALLY across English and Hindi explanation
    fields - "te-form" in English should be "て-form" too.

    Surfaced via reviewer follow-up during cycle-1 (2026-05-07). Documented
    as audit-finding HI-18 + HI-19. Cycle-1 closed all 67 known instances;
    this invariant locks the convention so future regressions fail CI.
    """
    failures: list[str] = []

    # Hindi grammatical-term suffixes that signal a JLPT pedagogical context.
    # When immediately preceded by a hyphen + a known kana-equivalent token
    # (Latin romaji or Devanagari xliteration), it's a violation.
    HINDI_TERMS = [
        "विशेषण", "क्रिया", "रूप", "संज्ञा", "कण", "वाक्य", "पैटर्न",
        "भूत", "नकार", "प्रश्न", "विधेय",
    ]
    EN_TERMS = ["adjective", "adj", "verb", "form", "particle", "predicate"]

    # Latin-romaji + (Devanagari OR English) grammatical-term
    LATIN_ROMAJI = ["na", "i", "te", "ta", "nai", "masu", "desu", "tai", "ba",
                    "tara", "tari", "kute", "nakute", "ru", "u"]
    latin_pat = re.compile(
        r"\b(" + "|".join(LATIN_ROMAJI) + r")-(" +
        "|".join(HINDI_TERMS + EN_TERMS) + r")\b",
        re.IGNORECASE
    )

    # Devanagari syllable + Devanagari term (the original ना-विशेषण bug).
    # Use negative lookbehind to require the syllable is NOT preceded by
    # other Devanagari (which would mean it's part of a longer Hindi word
    # like "रिश्ता" / "कर्ता" / "हिता" — false positives).
    DEV_ROMAJI = ["ना", "ते", "ता", "नाई", "मासू", "देसू", "सूरू"]
    dev_pat = re.compile(
        r"(?<![ऀ-ॿ])(" + "|".join(re.escape(d) for d in DEV_ROMAJI) +
        r")-(" + "|".join(HINDI_TERMS) + r")\b"
    )

    def has_devanagari(s: str) -> bool:
        return any('ऀ' <= ch <= 'ॿ' for ch in s)

    def scan_value(value, ctx: str):
        if not isinstance(value, str):
            return
        # Latin-romaji checks: scan anywhere (EN or HI prose - the rule is
        # symmetric)
        for m in latin_pat.finditer(value):
            failures.append(
                f"JA-41: kana-prefix violation at {ctx}: '{m.group(0)}' "
                f"should use kana (e.g., な-/い-/て-/た-/ない-/ます-/たい-)"
            )
        # Devanagari-xliteration: only flag inside Hindi prose (the
        # value must contain Devanagari for this to be a Hindi field)
        if has_devanagari(value):
            for m in dev_pat.finditer(value):
                failures.append(
                    f"JA-41: kana-prefix Devanagari xliteration at {ctx}: "
                    f"'{m.group(0)}' should use kana (e.g., '{m.group(0)}' "
                    f"-> use hiragana equivalent)"
                )

    # Walk all *_hi and explanation_en / meaning_en fields
    def walk(node, path):
        if isinstance(node, dict):
            for k, v in node.items():
                new_path = f"{path}.{k}" if path else k
                if isinstance(k, str) and (
                    k == "hi" or k.endswith("_hi") or k == "en"
                    or k.endswith("_en") or k in {"meaning_en", "explanation_en",
                    "rationale", "summary", "summary_en"}
                ):
                    scan_value(v, new_path)
                walk(v, new_path)
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk(v, f"{path}[{i}]")

    targets = [
        "data/grammar.json", "data/vocab.json", "data/kanji.json",
        "data/reading.json", "data/listening.json", "data/questions.json",
    ]
    for fname in targets:
        p = ROOT / fname
        if not p.exists():
            continue
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            walk(d, fname)
        except Exception as e:
            failures.append(f"JA-41 parse error on {fname}: {e}")

    # Papers
    papers_dir = ROOT / "data" / "papers"
    if papers_dir.exists():
        for pf in papers_dir.rglob("*.json"):
            if pf.name == "manifest.json":
                continue
            try:
                d = json.loads(pf.read_text(encoding="utf-8"))
                walk(d, str(pf.relative_to(ROOT)).replace("\\", "/"))
            except Exception as e:
                failures.append(f"JA-41 parse error on {pf.name}: {e}")

    return failures[:30]  # cap noise


# ---------------------------------------------------------------------------
# JA-48 added for legal-vetting F-8 (2026-05-11)
# ---------------------------------------------------------------------------

def _check_ja_48_kanjivg_svg_headers() -> list[str]:
    """Every svg/kanji/*.svg preserves the upstream KanjiVG copyright header
    (Ulrich Apel, CC-BY-SA 3.0). Required by the CC-BY-SA license to maintain
    attribution. Header must appear in the first 1KB of the file."""
    failures = []
    svg_dir = ROOT / "svg" / "kanji"
    if not svg_dir.is_dir():
        return ["JA-48 svg/kanji/ directory not found"]
    for svg in sorted(svg_dir.glob("*.svg")):
        try:
            head = svg.read_text(encoding="utf-8", errors="replace")[:1024]
        except Exception as e:
            failures.append(f"JA-48 {svg.name}: read failed: {e}")
            continue
        if "Ulrich Apel" not in head or "Copyright" not in head:
            failures.append(f"JA-48 {svg.name}: KanjiVG copyright header missing (no 'Ulrich Apel' / 'Copyright' in first 1KB)")
    return failures


# ---------------------------------------------------------------------------
# JA-47 added for legal-vetting F-3 (2026-05-11)
# ---------------------------------------------------------------------------

def _check_ja_47_content_license_counts() -> list[str]:
    """CONTENT-LICENSE.md §1 'What is original' table corpus counts agree
    with live data/*.json counts. Prevents the legal-defensibility document
    drifting out of sync after content batches. Added per F-3 (2026-05-11)."""
    failures = []
    try:
        cl_text = (ROOT / "CONTENT-LICENSE.md").read_text(encoding="utf-8")
    except Exception as e:
        return [f"JA-47 could not read CONTENT-LICENSE.md: {e}"]

    def _live_list_count(filename: str, key: str) -> int | None:
        try:
            d = json.loads((ROOT / filename).read_text(encoding="utf-8"))
            v = d.get(key)
            return len(v) if isinstance(v, list) else None
        except Exception:
            return None

    # (regex-with-int-capture, human-label, live-count-or-None)
    # Reads counts straight out of the §1 table cell, e.g. "(178 patterns)".
    checks = [
        (r"`data/grammar\.json`\s*\((\d+)\s*patterns?\)",  "grammar patterns",   _live_list_count("data/grammar.json",   "patterns")),
        (r"`data/questions\.json`\s*\((\d+)\s*MCQ",        "MCQ questions",      _live_list_count("data/questions.json", "questions")),
        (r"`data/reading\.json`\s*\((\d+)\s*passages?\)",  "reading passages",   _live_list_count("data/reading.json",   "passages")),
        (r"`data/listening\.json`\s*\((\d+)\s*drills?\)",  "listening drills",   _live_list_count("data/listening.json", "items")),
        (r"`data/vocab\.json`\s*\((\d+)\s*entries?\)",     "vocab entries",      _live_list_count("data/vocab.json",     "entries")),
        (r"`data/kanji\.json`\s*\((\d+)\s*entries?\)",     "kanji entries",      _live_list_count("data/kanji.json",     "entries")),
    ]
    for pattern, label, live in checks:
        m = re.search(pattern, cl_text)
        if not m:
            failures.append(f"JA-47 could not locate {label} claim in CONTENT-LICENSE.md")
            continue
        if live is None:
            failures.append(f"JA-47 could not compute live count for {label}")
            continue
        claimed = int(m.group(1))
        if claimed != live:
            failures.append(f"JA-47 {label}: CONTENT-LICENSE.md claims {claimed}, live count is {live}")

    # Papers manifest: paperCount + totalQuestions both cross-referenced.
    try:
        pm = json.loads((ROOT / "data" / "papers" / "manifest.json").read_text(encoding="utf-8"))
        live_papers = pm.get("totalPapers")
        live_paper_qs = pm.get("totalQuestions")
        m_papers = re.search(r"`data/papers/.*?\((\d+)\s*papers?\s*\xd7\s*(\d+)\s*paper\s*questions", cl_text)
        if m_papers is None:
            # `×` may be written as ASCII 'x' in some edits; allow fallback
            m_papers = re.search(r"`data/papers/.*?\((\d+)\s*papers?\s*[x\xd7]\s*(\d+)\s*paper\s*questions", cl_text)
        if m_papers:
            cl_papers, cl_qs = int(m_papers.group(1)), int(m_papers.group(2))
            if isinstance(live_papers, int) and cl_papers != live_papers:
                failures.append(f"JA-47 paper count: CONTENT-LICENSE.md claims {cl_papers}, manifest.json has {live_papers}")
            if isinstance(live_paper_qs, int) and cl_qs != live_paper_qs:
                failures.append(f"JA-47 paper questions: CONTENT-LICENSE.md claims {cl_qs}, manifest.json has {live_paper_qs}")
        else:
            failures.append("JA-47 could not locate papers count claim in CONTENT-LICENSE.md")
    except Exception as e:
        failures.append(f"JA-47 papers cross-check failed: {e}")

    return failures


# ---------------------------------------------------------------------------
# JA-49 through JA-53 added 2026-05-12: lock in the 2026-05-12 audit-cycle
# gains so future edits cannot silently regress past today's bar.
# Each is the CI complement of a closed-as-Done audit item (ISSUE-115, 116,
# 112, 118, plus P2-12). Adding them prevents the "audit drift" failure
# mode catalogued in feedback/audit-drift-findings-2026-05-12.md.
# ---------------------------------------------------------------------------


_VALID_REGISTERS = {"neutral", "polite", "humble", "respectful", "casual"}
_VALID_REGISTER_ORIGINS = {"wago", "kango", "gairaigo"}
_VALID_MISTAKE_CATEGORIES = {"particle", "verb_class", "conjugation", "register"}


def _check_ja_49_vocab_register_coverage() -> list[str]:
    """Every vocab entry must carry a `register` field with value in
    {neutral, polite, humble, respectful, casual}. Locks in ISSUE-115
    closure (1009/1009 explicit-neutral default policy)."""
    failures: list[str] = []
    path = ROOT / "data" / "vocab.json"
    if not path.exists():
        return ["JA-49: data/vocab.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-49: parse error: {e}"]
    for v in data.get("entries", []):
        vid = v.get("id", "?")
        reg = v.get("register")
        if not reg:
            failures.append(f"JA-49 {vid}: missing register field")
        elif reg not in _VALID_REGISTERS:
            failures.append(
                f"JA-49 {vid}: register={reg!r} not in {sorted(_VALID_REGISTERS)}"
            )
    return failures


def _check_ja_50_vocab_register_origin_coverage() -> list[str]:
    """Every vocab entry must carry a `register_origin` field with value
    in {wago, kango, gairaigo}. Locks in ISSUE-116 closure (deterministic
    Wago/Kango/Gairaigo classifier)."""
    failures: list[str] = []
    path = ROOT / "data" / "vocab.json"
    if not path.exists():
        return ["JA-50: data/vocab.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-50: parse error: {e}"]
    for v in data.get("entries", []):
        vid = v.get("id", "?")
        ro = v.get("register_origin")
        if not ro:
            failures.append(f"JA-50 {vid}: missing register_origin field")
        elif ro not in _VALID_REGISTER_ORIGINS:
            failures.append(
                f"JA-50 {vid}: register_origin={ro!r} not in {sorted(_VALID_REGISTER_ORIGINS)}"
            )
    return failures


def _check_ja_51_common_mistakes_categorized() -> list[str]:
    """Every grammar pattern must carry >=3 common_mistakes entries, each
    with a `category` field in {particle, verb_class, conjugation,
    register}. Locks in ISSUE-112 closure (the audit's Section 0 TOP-2
    P1 item — categorized error coverage above Bunpro's 1-generic bar)."""
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-51: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-51: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        cms = p.get("common_mistakes") or []
        if not isinstance(cms, list):
            failures.append(f"JA-51 {pid}: common_mistakes is not a list")
            continue
        categorized = [
            cm for cm in cms
            if isinstance(cm, dict) and cm.get("category") in _VALID_MISTAKE_CATEGORIES
        ]
        if len(categorized) < 3:
            failures.append(
                f"JA-51 {pid}: only {len(categorized)} categorized common_mistakes (need >=3)"
            )
        # Also flag any out-of-set categories
        for cm in cms:
            if isinstance(cm, dict):
                cat = cm.get("category")
                if cat and cat not in _VALID_MISTAKE_CATEGORIES:
                    failures.append(
                        f"JA-51 {pid}: common_mistake category={cat!r} "
                        f"not in {sorted(_VALID_MISTAKE_CATEGORIES)}"
                    )
    return failures


def _check_ja_52_grammar_contrasts_floor() -> list[str]:
    """Every grammar pattern must carry >=1 `contrasts` entry. Locks in
    ISSUE-118 closure (178/178 contrast cross-link coverage). Each
    contrast entry must include with_pattern_id pointing to a real
    pattern id."""
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-52: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-52: parse error: {e}"]
    patterns = data.get("patterns", [])
    pattern_ids = {p.get("id") for p in patterns if p.get("id")}
    for p in patterns:
        pid = p.get("id", "?")
        cs = p.get("contrasts") or []
        if not isinstance(cs, list) or len(cs) < 1:
            failures.append(f"JA-52 {pid}: missing or empty contrasts (need >=1)")
            continue
        for c in cs:
            if not isinstance(c, dict):
                failures.append(f"JA-52 {pid}: contrast entry is not a dict")
                continue
            tid = c.get("with_pattern_id")
            if not tid:
                failures.append(f"JA-52 {pid}: contrast missing with_pattern_id")
            elif tid not in pattern_ids:
                failures.append(
                    f"JA-52 {pid}: contrast with_pattern_id={tid!r} not a real pattern"
                )
    return failures


# ---------------------------------------------------------------------------
# JA-54 through JA-61 added 2026-05-12 (anti-item enforcement batch):
# programmatically lock the Section-10 mandatory anti-items so a future
# careless edit cannot silently violate the audit's frozen contracts.
# ---------------------------------------------------------------------------

_CORPUS_LOCKS = {
    ("data/grammar.json",   "patterns", 178),
    ("data/vocab.json",     "entries",  995),   # was 1009; 2026-05-16: BUG-018 -10 + BUG-017 collision -1 + BUG-019 -3 (月/あつい/きって subset-gloss duplicates missed by BUG-018) = 995
    ("data/reading.json",   "passages", 54),
    ("data/listening.json", "items",    50),
}

# Kanji has separate count locking (X-6.6) but the count is the same anyway.
_KANJI_LOCK = 106

_ESSAY_SUBFIELDS = (
    "intro", "why_it_matters", "common_pitfalls",
    "contrasts", "closing_practice_tip", "cultural_context",
)


def _check_ja_54_essay_min_length() -> list[str]:
    """Every grammar pattern's essay sub-fields combined must total >=500
    characters. Locks in the Tofugu-level depth bar referenced by the
    2026-05-09 audit (P2-13 essay enrichment) and confirmed at 178/178
    on 2026-05-12 after the n5-149 4-char top-up."""
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-54: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-54: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        e = p.get("essay") or {}
        if not isinstance(e, dict):
            failures.append(f"JA-54 {pid}: essay is not a dict")
            continue
        total = sum(len(str(e.get(sf, ""))) for sf in _ESSAY_SUBFIELDS)
        if total < 500:
            failures.append(f"JA-54 {pid}: essay total only {total} chars (need >=500)")
    return failures


def _check_ja_55_essay_schema() -> list[str]:
    """Every grammar pattern's essay dict must contain all 6 sub-fields
    (intro, why_it_matters, common_pitfalls, contrasts,
    closing_practice_tip, cultural_context). The renderer in
    js/learn-grammar.js iterates these fields directly; missing keys
    would silently degrade the rendered output."""
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-55: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-55: parse error: {e}"]
    expected = set(_ESSAY_SUBFIELDS)
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        e = p.get("essay")
        if not isinstance(e, dict):
            failures.append(f"JA-55 {pid}: no essay dict")
            continue
        keys = set(e.keys())
        missing = expected - keys
        if missing:
            failures.append(
                f"JA-55 {pid}: essay missing sub-fields {sorted(missing)}"
            )
    return failures


def _check_ja_56_corpus_size_locks() -> list[str]:
    """Section-10 anti-items #1-4: corpus sizes are FROZEN.
      - 178 grammar patterns (exceeds Bunpro N5)
      - 1009 vocab entries (exceeds Genki/MNN/JLPT Sensei)
      - 106 kanji (matches the canonical N5 set)
      - 54 reading passages
      - 50 listening items
    Adding more violates the audit's anti-items and bloats the bundle."""
    failures: list[str] = []
    for relpath, list_key, expected in _CORPUS_LOCKS:
        path = ROOT / relpath
        if not path.exists():
            failures.append(f"JA-56: {relpath} missing")
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-56: {relpath} parse error: {e}")
            continue
        arr = data.get(list_key) or []
        if len(arr) != expected:
            failures.append(
                f"JA-56 {relpath}.{list_key}: {len(arr)} entries (locked at {expected})"
            )

    # Kanji.json has the entries either as dict or list — handle both
    k_path = ROOT / "data" / "kanji.json"
    if k_path.exists():
        try:
            kdata = json.loads(k_path.read_text(encoding="utf-8"))
            KE = kdata.get("entries", kdata)
            if isinstance(KE, dict):
                KE = list(KE.values())
            if len(KE) != _KANJI_LOCK:
                failures.append(
                    f"JA-56 data/kanji.json.entries: {len(KE)} entries (locked at {_KANJI_LOCK})"
                )
        except Exception as e:
            failures.append(f"JA-56 data/kanji.json parse error: {e}")
    return failures


def _check_ja_57_no_lh_pitch_notation() -> list[str]:
    """Section-10 anti-item #11: pitch_accent must use the NHK
    {mora, drop} integer notation, NOT the LH/HL string notation
    (which only works for words <=3 morae and breaks at N4+ scale)."""
    failures: list[str] = []
    path = ROOT / "data" / "vocab.json"
    if not path.exists():
        return ["JA-57: data/vocab.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-57: parse error: {e}"]
    lh_pattern = re.compile(r"\b[LH]{2,}\b")
    for v in data.get("entries", []):
        vid = v.get("id", "?")
        pa = v.get("pitch_accent")
        if pa is None:
            continue
        # Check stringified pitch_accent for LH/HL contiguous patterns
        as_str = json.dumps(pa, ensure_ascii=False)
        if lh_pattern.search(as_str):
            failures.append(
                f"JA-57 {vid}: pitch_accent contains LH/HL string notation "
                f"(use {{mora, drop}} integer form instead): {as_str[:80]}"
            )
    return failures


def _check_ja_58_no_jlpt_jp_current_citation() -> list[str]:
    """Section-10 anti-item #12: the JEES 出題基準 (JLPT.jp official)
    was discontinued in 2010. Citing it as a CURRENT source for
    grammar/vocab/kanji is incorrect — the legitimate citation is the
    archival 旧出題基準 1994/2002, which the corpus already uses.
    Block any new 'JLPT.jp official' tokens that would imply currency."""
    failures: list[str] = []
    for relpath in ("data/grammar.json", "data/vocab.json", "data/kanji.json"):
        path = ROOT / relpath
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8")
        # Look for the literal phrase indicating CURRENT-source claim
        if "JLPT.jp official" in text:
            count = text.count("JLPT.jp official")
            failures.append(
                f"JA-58 {relpath}: {count} occurrences of 'JLPT.jp official' "
                f"(JEES 出題基準 discontinued 2010; use '旧出題基準 1994/2002' instead)"
            )
    return failures


def _check_ja_59_no_gamification_state() -> list[str]:
    """Section-10 anti-item #5 (refined scope): no COMPETITIVE
    gamification — leaderboards / XP / badges / achievements.

    Note: the pre-existing local habit-formation tracker `streak`
    (consecutive-study-days counter) in js/storage.js is intentionally
    permitted. It carries no social / comparative semantics — purely
    local motivation for adult learners. The anti-item's core spirit
    is to avoid COMPETITIVE / public-comparison gamification, which
    would break the privacy-no-account contract. Local streak counts
    do not violate that contract.

    Enforce blocking on the harder-line entries:
      - leaderboards (any comparative ranking surface)
      - XP (experience-point accumulation systems)
      - badges / achievements (collectibles that imply progress-rewards
        beyond plain SRS state)
    """
    failures: list[str] = []
    # Check js/storage.js for COMPETITIVE gamification keys
    storage_path = ROOT / "js" / "storage.js"
    if storage_path.exists():
        text = storage_path.read_text(encoding="utf-8")
        forbidden_keys = [
            r"\bleaderboard\s*:",
            r"\bxp\s*:",
            r"\bexperience_points\s*:",
            r"\bbadge\s*:",
            r"\bachievement\s*:",
            r"\brank\s*:",
        ]
        for pat in forbidden_keys:
            if re.search(pat, text, re.I):
                failures.append(
                    f"JA-59 js/storage.js: competitive-gamification key matching {pat!r} found"
                )
    # Check no standalone files for competitive-gamification routes exist
    for fname in ("leaderboard.js", "xp.js", "badges.js", "achievements.js", "rankings.js"):
        if (ROOT / "js" / fname).exists():
            failures.append(f"JA-59 js/{fname}: competitive-gamification file exists (forbidden)")
    return failures


def _check_ja_60_no_account_or_cloud_sync() -> list[str]:
    """Section-10 anti-item #6 + #7: no account / cloud-sync / per-content
    discussion threads. Enforce by checking that no JS file fetches a
    non-local URL (privacy posture: everything from ./)."""
    failures: list[str] = []
    js_dir = ROOT / "js"
    if not js_dir.is_dir():
        return failures
    # Look for fetch() / XMLHttpRequest with explicit non-relative URL
    # The exemption: data: URIs (no network), blob:, and ./relative paths.
    bad_url_pat = re.compile(
        r"""\bfetch\s*\(\s*["']\s*(https?:\/\/|ws:\/\/|wss:\/\/)""",
        re.I,
    )
    for jsfile in js_dir.glob("*.js"):
        if jsfile.name.startswith("."):
            continue
        try:
            text = jsfile.read_text(encoding="utf-8")
        except Exception:
            continue
        if bad_url_pat.search(text):
            failures.append(
                f"JA-60 {jsfile.relative_to(ROOT)}: fetch() to non-local URL "
                f"(privacy posture: only ./ relative paths allowed)"
            )
    return failures


def _check_ja_61_no_discussion_route() -> list[str]:
    """Section-10 anti-item #7: no per-content discussion / comments.
    Enforce by checking that no js/discussion.js or js/comments.js
    exists, and the router in app.js does NOT register a discussion
    or comments route."""
    failures: list[str] = []
    for fname in ("discussion.js", "comments.js", "thread.js"):
        if (ROOT / "js" / fname).exists():
            failures.append(f"JA-61 js/{fname}: discussion/comments file exists (forbidden)")
    app_path = ROOT / "js" / "app.js"
    if app_path.exists():
        text = app_path.read_text(encoding="utf-8")
        # Look for route registration matching 'discussion' or 'comments'
        bad_route = re.compile(r"['\"](discussion|comments|thread)['\"]\s*:")
        for m in bad_route.finditer(text):
            failures.append(
                f"JA-61 js/app.js: route '{m.group(1)}' registered (forbidden anti-item)"
            )
    return failures


# ---------------------------------------------------------------------------
# JA-62..65 added 2026-05-12 (round 2 of anti-item enforcement):
# romaji-in-data + shape contracts on authentic_refs, common_mistakes,
# and contrasts.
# ---------------------------------------------------------------------------


def _check_ja_62_no_romaji_in_japanese_fields() -> list[str]:
    """Section-10 anti-item #9: no romaji in user-facing Japanese
    display fields. Romaji INPUT (the typed-answer path through
    js/romaji-kana.js) is intentionally permitted; only DISPLAY
    fields are restricted.

    Checks fields that must be in kana/kanji:
      - vocab.form, vocab.reading
      - vocab.examples[].ja
      - grammar.examples[].ja
      - authentic.ja, authentic.reading

    Detection: 3+ consecutive ASCII letters (a-z / A-Z) inside the
    field value. Stand-alone Latin punctuation, single letters
    (e.g., "P" in pitch notation), and numeric-only strings pass.
    Pattern-label fields like grammar.pattern are exempt (some
    legitimate labels are romanized — te-form, ます-form)."""
    failures: list[str] = []
    romaji_pat = re.compile(r"[a-zA-Z]{3,}")

    # vocab
    v_path = ROOT / "data" / "vocab.json"
    if v_path.exists():
        try:
            vdata = json.loads(v_path.read_text(encoding="utf-8"))
        except Exception as e:
            return [f"JA-62: vocab.json parse error: {e}"]
        for v in vdata.get("entries", []):
            vid = v.get("id", "?")
            for fld in ("form", "reading"):
                val = v.get(fld) or ""
                if isinstance(val, str) and romaji_pat.search(val):
                    failures.append(f"JA-62 vocab {vid}.{fld}: romaji found in {val!r}")
            for ex in v.get("examples") or []:
                if isinstance(ex, dict):
                    ja = ex.get("ja") or ""
                    if isinstance(ja, str) and romaji_pat.search(ja):
                        failures.append(f"JA-62 vocab {vid}.examples[].ja: romaji found in {ja!r}")

    # grammar
    g_path = ROOT / "data" / "grammar.json"
    if g_path.exists():
        try:
            gdata = json.loads(g_path.read_text(encoding="utf-8"))
        except Exception as e:
            return failures + [f"JA-62: grammar.json parse error: {e}"]
        for p in gdata.get("patterns", []):
            pid = p.get("id", "?")
            for ex in p.get("examples") or []:
                if isinstance(ex, dict):
                    ja = ex.get("ja") or ""
                    if isinstance(ja, str) and romaji_pat.search(ja):
                        failures.append(f"JA-62 grammar {pid}.examples[].ja: romaji found in {ja!r}")

    # authentic
    a_path = ROOT / "data" / "authentic.json"
    if a_path.exists():
        try:
            adata = json.loads(a_path.read_text(encoding="utf-8"))
        except Exception as e:
            return failures + [f"JA-62: authentic.json parse error: {e}"]
        for c in adata.get("items", []):
            cid = c.get("id", "?")
            for fld in ("ja", "reading"):
                val = c.get(fld) or ""
                if isinstance(val, str) and romaji_pat.search(val):
                    failures.append(f"JA-62 authentic {cid}.{fld}: romaji found in {val!r}")
    return failures


def _check_ja_63_authentic_kanji_refs_complete() -> list[str]:
    """Shape contract: every authentic card's `kanji_refs` array must
    list ALL N5 kanji that appear in its `ja` text. Surfaced as a
    data-integrity gap during today's wave-2/3 signage authoring —
    16 cards were initially under-populated. JA-63 prevents this
    regression."""
    failures: list[str] = []
    a_path = ROOT / "data" / "authentic.json"
    k_path = ROOT / "data" / "kanji.json"
    if not a_path.exists() or not k_path.exists():
        return ["JA-63: required data files missing"]
    try:
        adata = json.loads(a_path.read_text(encoding="utf-8"))
        kdata = json.loads(k_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-63: parse error: {e}"]
    KE = kdata.get("entries", kdata)
    if isinstance(KE, dict):
        KE = list(KE.values())
    n5_glyphs = {k.get("glyph"): k.get("id") for k in KE if k.get("glyph") and k.get("id")}

    for c in adata.get("items", []):
        cid = c.get("id", "?")
        ja = c.get("ja", "")
        refs = set(c.get("kanji_refs") or [])
        expected = set()
        for ch in ja:
            if ch in n5_glyphs:
                expected.add(n5_glyphs[ch])
        missing = expected - refs
        if missing:
            failures.append(
                f"JA-63 {cid}: ja={ja!r} contains N5 kanji not in kanji_refs: {sorted(missing)}"
            )
    return failures


def _check_ja_64_common_mistakes_shape() -> list[str]:
    """Shape contract: every common_mistakes entry must have a
    well-populated form pair + a `why` field. The renderer iterates
    these fields directly (js/learn-grammar.js#renderMistakes); a
    missing field produces visibly broken output ('undefined' or
    blank cells in the mistakes table).

    Two valid shapes (BUG-013, 2026-05-16):

    1. Legacy / error entries: {`wrong`, `right`, `why`} — the
       red-strike / green-check rendering. `wrong` is an
       ungrammatical form; `right` is the corrected form.

    2. Register-variant entries (`kind: "register_variant"`):
       {`form_a`, `form_b`, `why`} — neutral "Form A / Form B"
       rendering. Both forms are grammatically valid; they differ
       in register / formality / pragmatic context. Optional:
       `label_a` / `label_b` register tags.
    """
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-64: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-64: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        for i, cm in enumerate(p.get("common_mistakes") or []):
            if not isinstance(cm, dict):
                failures.append(f"JA-64 {pid}.common_mistakes[{i}]: not a dict")
                continue
            # Choose required field set based on `kind`
            if cm.get("kind") == "register_variant":
                required = ("form_a", "form_b", "why")
            else:
                required = ("wrong", "right", "why")
            for fld in required:
                val = cm.get(fld)
                if not val or (isinstance(val, str) and not val.strip()):
                    failures.append(
                        f"JA-64 {pid}.common_mistakes[{i}].{fld}: missing or empty"
                    )
            # Defense-in-depth: a register_variant entry MUST NOT carry
            # the legacy wrong/right keys (BUG-013 schema migration).
            if cm.get("kind") == "register_variant":
                for stale in ("wrong", "right"):
                    if stale in cm:
                        failures.append(
                            f"JA-64 {pid}.common_mistakes[{i}].{stale}: "
                            f"stale key on register_variant entry (BUG-013); "
                            f"use form_a/form_b only"
                        )
    return failures


def _check_ja_65_contrast_notes_min_length() -> list[str]:
    """Shape contract: every `contrasts` entry's `note` field must be
    >=30 chars. Pedagogically anchors the bar at a substantive
    one-sentence explanation rather than a trivial gloss ('A vs B')."""
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-65: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-65: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        for i, c in enumerate(p.get("contrasts") or []):
            if not isinstance(c, dict):
                continue
            note = c.get("note") or ""
            if isinstance(note, str) and len(note) < 30:
                failures.append(
                    f"JA-65 {pid}.contrasts[{i}].note: only {len(note)} chars (need >=30): {note!r}"
                )
    return failures


def _check_ja_66_explanation_en_kanji_in_scope() -> list[str]:
    """JA-66 (2026-05-13): every grammar pattern's `explanation_en` field
    must contain only kanji that appear on the N5 whitelist
    (data/n5_kanji_whitelist.json).

    Background: JA-13's leaf-skip set includes `explanation_en` because
    it is nominally English commentary. That assumption broke in the
    Phase 7 polish (v1.15.1) when surgical upgrades to 3 patterns
    introduced 4 above-N5 kanji (好/嫌/広/方) embedded inline in the
    English text — slipped past CI and was caught only by manual review.

    This check enforces the convention used throughout the rest of the
    corpus: when explanation_en cites a Japanese word, write it in kana
    (or kana with N5-kanji where possible). Above-N5 kanji belong in
    the example-sentence field or the contrast-pair surface — not in
    the explanation gloss.

    Reciprocal pattern_role (under public_domain_refs) is also nominal
    English commentary; same convention. Both surfaces are checked.
    """
    failures: list[str] = []
    try:
        whitelist = set(json.loads((ROOT / "data" / "n5_kanji_whitelist.json").read_text(encoding="utf-8")))
    except Exception as e:
        return [f"JA-66 could not load n5_kanji_whitelist.json: {e}"]
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-66: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-66: parse error: {e}"]
    KANJI_RE = re.compile(r"[一-鿿]")
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        expl = p.get("explanation_en") or ""
        if isinstance(expl, str):
            bad = sorted({c for c in KANJI_RE.findall(expl) if c not in whitelist})
            if bad:
                failures.append(
                    f"JA-66 {pid}.explanation_en: above-N5 kanji {bad} — "
                    f"use kana or N5-only kanji in English commentary."
                )
        for i, r in enumerate(p.get("public_domain_refs") or []):
            if not isinstance(r, dict):
                continue
            role = r.get("pattern_role") or ""
            if isinstance(role, str):
                bad = sorted({c for c in KANJI_RE.findall(role) if c not in whitelist})
                if bad:
                    failures.append(
                        f"JA-66 {pid}.public_domain_refs[{i}].pattern_role: "
                        f"above-N5 kanji {bad} — pattern_role is English."
                    )
    return failures


def _check_ja_67_density3_floor_lock() -> list[str]:
    """JA-67 (2026-05-13): Density-3 (kanji->vocab) below-floor count
    regression guard. After the D2/D3 fix (commit d7eaf43), 24 of 106
    kanji have <2 vocab uses (intrinsic singletons at frozen N5 width).
    The count is LOCKED at 24 — a drop further below the floor is a
    regression that this check catches.

    Union method: kanji is "used" if it appears in vocab.form OR is
    linked via kanji.n5_compounds.vocab_id. This matches the canonical
    Density-3 measurement.
    """
    failures: list[str] = []
    LOCKED_COUNT = 25  # below-floor count at the lock point (commit d7eaf43);
                       # +1 on 2026-05-16 from BUG-018 dedup (道 lost one of its
                       # two vocab usages when n5.vocab.23-transport.道 was
                       # dropped in favor of n5.vocab.13-locations-and-places-.道)
    try:
        V = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))["entries"]
        K = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))["entries"]
    except Exception as e:
        return [f"JA-67: load error: {e}"]

    n5_glyphs = {k["glyph"] for k in K}
    used = {g: set() for g in n5_glyphs}
    for v in V:
        for ch in (v.get("form") or ""):
            if ch in n5_glyphs:
                used[ch].add(v.get("id"))
    for k in K:
        for c in (k.get("n5_compounds") or []):
            if isinstance(c, dict) and c.get("vocab_id"):
                used[k["glyph"]].add(c["vocab_id"])

    below_floor = sum(1 for g in n5_glyphs if len(used[g]) < 2)
    if below_floor > LOCKED_COUNT:
        glyphs_below = sorted([g for g in n5_glyphs if len(used[g]) < 2],
                              key=lambda g: len(used[g]))
        failures.append(
            f"JA-67: Density-3 below-floor count regressed to {below_floor} "
            f"(locked at {LOCKED_COUNT}). Excess kanji: {glyphs_below[LOCKED_COUNT:]}"
        )
    return failures


def _check_ja_69_pd_refs_legal_status() -> list[str]:
    """JA-69 (2026-05-13): every entry in grammar.json's `public_domain_refs`
    must cite a verified public-domain work. Failure modes caught:

    1. `pd_status` text contains red-flag legal language ('pending',
       'protected', 'in copyright', 'PD in 20XX' for a future year, or
       '(Fallback ref:)' placeholder substring).
    2. `author_death_year` is more recent than 1955 (conservative buffer
       over Japan's mathematical PD threshold of 1967 — works whose
       authors died in 1968+ get the new life+70 rule and aren't PD
       until 2038+; works with deaths 1956-1967 are borderline edge
       cases the corpus avoids by policy).

    Background: legal-vetting audit 2026-05-13 caught 4 in-copyright
    entries that slipped through the original PD-refs expansion
    (Kawabata d.1972, Nakamura ×2 d.1972, Nishijō d.1970). The data
    file's own pd_status field admitted the copyright status ("PD
    pending until 2043"). This check programmatically catches the same
    class going forward.

    Exemptions:
      - source_type == "proverb"/"folk_song" with author == "(traditional)"
        and author_death_year is None → pre-modern anonymous; pass.
      - source_type == "government" → §13 著作権法 exception; pass.
      - source_type == "nhk_easy" with no quoted text → recommendation-
        only reference; pass (the upstream usage rights are NHK's, not
        ours; we don't redistribute the content).
    """
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-69: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-69: parse error: {e}"]

    RED_FLAG_SUBSTRINGS = (
        "pending",
        "protected",
        "in copyright",
        "(fallback ref:)",
    )
    # Match "PD in 20XX" / "PD until 20XX" / "expires 20XX" with year >= 2026.
    FUTURE_PD_RE = re.compile(r"(?:PD in|PD until|expires)\s+(\d{4})", re.IGNORECASE)
    AUTHOR_DEATH_YEAR_BUFFER = 1955  # conservative; mathematical threshold is 1967

    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        for i, r in enumerate(p.get("public_domain_refs") or []):
            if not isinstance(r, dict):
                continue
            source_type = (r.get("source_type") or "").lower()
            pd_status = r.get("pd_status") or ""
            context = r.get("context") or ""
            adyear = r.get("author_death_year")

            # Red-flag substring check (case-insensitive over pd_status + context).
            blob = (pd_status + " " + context).lower()
            for flag in RED_FLAG_SUBSTRINGS:
                if flag in blob:
                    failures.append(
                        f"JA-69 {pid}.public_domain_refs[{i}]: "
                        f"red-flag substring '{flag}' in pd_status/context. "
                        f"Replace with a verified-PD reference."
                    )
                    break

            # Future-PD-year claim.
            for m in FUTURE_PD_RE.finditer(pd_status):
                year = int(m.group(1))
                if year >= 2026:
                    failures.append(
                        f"JA-69 {pid}.public_domain_refs[{i}]: "
                        f"pd_status admits work enters PD in {year} (future) — "
                        f"work is still in copyright today."
                    )

            # author_death_year buffer (skip for source_types that don't
            # rely on an authored work).
            if isinstance(adyear, int) and adyear > AUTHOR_DEATH_YEAR_BUFFER:
                if source_type not in ("proverb", "folk_song", "government", "nhk_easy"):
                    failures.append(
                        f"JA-69 {pid}.public_domain_refs[{i}]: "
                        f"author_death_year={adyear} is more recent than the "
                        f"safe buffer ({AUTHOR_DEATH_YEAR_BUFFER}). Japan PD "
                        f"under life+70 = {adyear + 70} (still in copyright)."
                    )

    return failures


def _check_ja_77_no_placeholder_leakage() -> list[str]:
    """JA-77 (2026-05-13): no placeholder authoring meta-commentary
    in shipped data fields. Catches text like:
      - "(unspecified — keep prior)"
      - "TODO" / "FIXME" / "TBD" / "XXX"
      - "placeholder" / "(temp)" / "INSERT_"
      - "fallback ref" (from earlier PD-refs cleanup)
      - "Pattern-shape placeholder"
    Scans all user-facing text fields across all data files.
    """
    PLACEHOLDER_TERMS = [
        "(unspecified", "TODO:", "TODO ", "FIXME", "XXX",
        "placeholder", "(temp)", "INSERT_", "TBD",
        "keep prior", "Pattern-shape placeholder",
        "Fallback ref", "(fallback ref:)",
        "Sample text", "lorem ipsum",
    ]
    failures: list[str] = []
    files = ["data/grammar.json", "data/vocab.json", "data/kanji.json",
             "data/reading.json", "data/listening.json"]

    def walk(obj, path):
        if isinstance(obj, dict):
            for k, v in obj.items():
                # Skip internal metadata keys (underscore-prefix)
                if isinstance(k, str) and k.startswith("_"):
                    continue
                yield from walk(v, f"{path}.{k}")
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                yield from walk(item, f"{path}[{i}]")
        elif isinstance(obj, str):
            yield (path, obj)

    for fname in files:
        path = ROOT / fname
        if not path.exists():
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        items = data.get("entries") or data.get("passages") or data.get("items") or data.get("patterns") or []
        for item in items:
            item_id = item.get("id", "?")
            for field_path, val in walk(item, item_id):
                for term in PLACEHOLDER_TERMS:
                    if term.lower() in val.lower():
                        short = field_path.split('.', 1)[1] if '.' in field_path else field_path
                        failures.append(
                            f"JA-77 {fname.split('/')[-1]} {item_id}.{short}: "
                            f"placeholder text '{term}' found: {val[:80]!r}"
                        )
                        break  # one finding per field
    return failures


def _check_ja_79_form_field_consistency() -> list[str]:
    """JA-79 (2026-05-13): within each grammar pattern, all examples must
    EITHER all have `form` populated OR none. Mixed populations cause
    the UI to render partial form-badges (broken UX). Per anti-pattern
    §3.2.34.
    """
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-79: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-79: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        exs = p.get("examples") or []
        if not exs:
            continue
        has = sum(1 for ex in exs if isinstance(ex, dict) and ex.get("form"))
        if 0 < has < len(exs):
            missing = [i for i, ex in enumerate(exs)
                       if not (isinstance(ex, dict) and ex.get("form"))]
            failures.append(
                f"JA-79 {pid}: {has}/{len(exs)} examples have `form`; "
                f"missing at indices {missing}. Either fill all or strip all."
            )
    return failures


# JA-80 implementation removed (2026-05-13 run-4):
# The heuristic "meaning_ja must share Japanese substring with meaning_en"
# had 19 false positives — many legitimate meaning_ja entries paraphrase
# meaning_en using DIFFERENT Japanese vocabulary (e.g., n5-068 meaning_en
# uses 「なかった」 while meaning_ja uses 「ふつうの かこ ひてい」 — same
# concept, different words). Cross-contamination detection of this class
# requires semantic comparison, not string overlap, and is left to
# manual native-teacher review (logged as A12 still-cannot-catch case).


def _check_ja_78_no_repeated_kana_examples() -> list[str]:
    """JA-78 (2026-05-13): grammar example sentences should not have 3+
    consecutive same hiragana characters. Catches typo-looking strings
    like ははは / がが / をを that are grammatically valid but read as
    typos to learners. The fix is usually to use kanji form (母は instead
    of ははは).
    """
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-78: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-78: parse error: {e}"]
    REPEAT_RE = re.compile(r"([぀-ゟ])\1{2,}")  # 3+ consecutive same hiragana
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        for i, ex in enumerate(p.get("examples") or []):
            ja = ex.get("ja") or ""
            m = REPEAT_RE.search(ja)
            if m:
                failures.append(
                    f"JA-78 {pid}.examples[{i}]: 3+ consecutive '{m.group(1)}' "
                    f"in {ja!r}. Consider kanji form for clarity."
                )
    return failures


def _check_ja_76_on_yomi_katakana() -> list[str]:
    """JA-76 (2026-05-13): kanji on-yomi entries must be written in
    katakana per standard pedagogical convention. Hiragana characters
    in the `on` array indicate a regression from the 2026-05-13
    katakana-conversion baseline.

    Checks:
      - kanji.entries[*].on[*] — top-level on-yomi display array
      - kanji.entries[*].audio_yomi.on[*].reading — per-yomi audio
        display label (the `audio` path field itself remains in
        hiragana for filename consistency; only the display label
        is locked to katakana)
    """
    failures: list[str] = []
    path = ROOT / "data" / "kanji.json"
    if not path.exists():
        return ["JA-76: data/kanji.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-76: parse error: {e}"]
    HIRAGANA_RE = re.compile(r"[぀-ゟ]")  # U+3040..U+309F
    for k in data.get("entries", []):
        glyph = k.get("glyph")
        # Check top-level on array
        for o in (k.get("on") or []):
            if isinstance(o, str) and HIRAGANA_RE.search(o):
                failures.append(
                    f"JA-76 {glyph}: on-yomi '{o}' contains hiragana; "
                    f"convert to katakana per pedagogical convention."
                )
        # Check audio_yomi.on[].reading display labels
        ay = k.get("audio_yomi") or {}
        if isinstance(ay, dict):
            for entry in (ay.get("on") or []):
                if isinstance(entry, dict):
                    reading = entry.get("reading") or ""
                    if HIRAGANA_RE.search(reading):
                        failures.append(
                            f"JA-76 {glyph}: audio_yomi.on reading display "
                            f"'{reading}' contains hiragana."
                        )
    return failures


def _check_ja_75_meaning_ja_markers() -> list[str]:
    """JA-75 (2026-05-13): per-pattern meaning_ja marker dictionary.

    Each grammar pattern carries a `_meaning_ja_markers: [...]` list
    (snapshot of distinctive vocabulary in the verified-correct
    meaning_ja, populated 2026-05-13 by tools/install_ja75_marker_dict.py).

    meaning_ja MUST contain at least one of its markers as a substring.
    Catches the cross-contamination class that JA-71 misses (1
    incidental character overlap is enough to pass JA-71, but won't
    contain the conceptual markers).

    Patterns without `_meaning_ja_markers` are skipped (back-compat
    for new entries; install script repopulates).
    """
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-75: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-75: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        markers = p.get("_meaning_ja_markers")
        if not markers:
            continue  # skip patterns without markers
        meaning_ja = p.get("meaning_ja") or ""
        # Tightened rule: meaning_ja must contain EITHER (a) at least
        # 2 distinct markers, OR (b) at least 1 "distinctive" marker
        # (≥3 chars, content-word — not a single common kana like
        # です/ます that would coincidentally match unrelated text).
        matched = [m for m in markers if m in meaning_ja]
        if len(matched) >= 2:
            continue  # >=2 marker hits — clearly references this pattern
        distinctive = [m for m in matched if len(m) >= 3]
        if distinctive:
            continue  # at least 1 distinctive content marker
        # Otherwise: failure (either zero markers matched, or only short
        # generic ones like です which doesn't prove the meaning_ja is
        # actually about THIS pattern)
        failures.append(
            f"JA-75 {pid}: meaning_ja has insufficient marker overlap "
            f"(matched={matched}, expected ≥2 markers OR ≥1 distinctive ≥3-char "
            f"marker from {markers[:5]}). meaning_ja[:60]='{meaning_ja[:60]}'"
        )
    return failures


def _check_ja_74_vocab_readings_schema() -> list[str]:
    """JA-74 (2026-05-13): vocab `reading` field must be a single kana
    form. Multi-reading entries (e.g., 何 = なに/なん) use the
    `readings: [...]` list field with `reading` set to the primary.
    Slash-separated strings in `reading` are a schema regression.
    """
    failures: list[str] = []
    path = ROOT / "data" / "vocab.json"
    if not path.exists():
        return ["JA-74: data/vocab.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-74: parse error: {e}"]
    for v in data.get("entries", []):
        reading = v.get("reading") or ""
        if "/" in reading:
            failures.append(
                f"JA-74 {v.get('id')}: `reading` contains '/'. Use "
                f"`readings: [...]` list for multi-reading entries; "
                f"set `reading` to the primary single form."
            )
        readings = v.get("readings")
        if readings is not None:
            if not isinstance(readings, list):
                failures.append(
                    f"JA-74 {v.get('id')}: `readings` must be a list, got {type(readings).__name__}"
                )
            elif readings and reading and reading not in readings:
                failures.append(
                    f"JA-74 {v.get('id')}: `reading`='{reading}' is not in `readings={readings}`."
                )
    return failures


def _check_ja_73_reading_prompt_ja_canonical() -> list[str]:
    """JA-73 (2026-05-13): reading questions must use `prompt_ja` as
    the stem-field key. The legacy `question_ja` key was used for 9
    questions (n5.read.046-054) but was renamed to prompt_ja in the
    run-2 review T-3 fix. This invariant locks the rename forward.
    """
    failures: list[str] = []
    path = ROOT / "data" / "reading.json"
    if not path.exists():
        return ["JA-73: data/reading.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-73: parse error: {e}"]
    for r in data.get("passages", []):
        for qi, q in enumerate(r.get("questions") or []):
            if "question_ja" in q:
                failures.append(
                    f"JA-73 {r.get('id')} q[{qi}]: uses legacy "
                    f"'question_ja' key; rename to canonical 'prompt_ja'."
                )
            if not q.get("prompt_ja"):
                failures.append(
                    f"JA-73 {r.get('id')} q[{qi}]: missing prompt_ja stem."
                )
    return failures


def _check_ja_72_gairaigo_katakana() -> list[str]:
    """JA-72 (2026-05-13): vocab entries flagged as gairaigo (loanwords
    from non-Japanese languages) must have their `form` field written
    in katakana. Hiragana / kanji forms for loanwords are a typography
    mismatch — natives write カタカナ for foreign-origin words.

    A form is "all-katakana" if every Japanese character in it is in
    the katakana range U+30A0-U+30FF, U+31F0-U+31FF (kana extensions),
    or U+FF65-U+FF9F (half-width katakana). Non-Japanese characters
    (Latin, digits, etc.) are allowed (e.g., "T シャツ" is fine).
    """
    failures: list[str] = []
    path = ROOT / "data" / "vocab.json"
    if not path.exists():
        return ["JA-72: data/vocab.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-72: parse error: {e}"]

    # Hiragana range U+3040-U+309F; CJK kanji U+4E00-U+9FFF.
    HIRAGANA_RE = re.compile(r"[぀-ゟ]")
    KANJI_RE = re.compile(r"[一-鿿]")

    for v in data.get("entries", []):
        if v.get("register_origin") != "gairaigo":
            continue
        form = v.get("form") or ""
        # Skip if form is empty
        if not form:
            continue
        if HIRAGANA_RE.search(form) or KANJI_RE.search(form):
            failures.append(
                f"JA-72 {v.get('id')}: gairaigo entry '{form}' contains "
                f"hiragana or kanji; native convention is katakana only "
                f"for loanwords."
            )
    return failures


def _check_ja_70_vocab_pitch_mora_count() -> list[str]:
    """JA-70 (2026-05-13): every vocab entry's pitch_accent.mora must
    equal the actual mora count of its reading. Tokyo NHK convention:
    each kana = 1 mora; small kana (ゃゅょぁぃぅぇぉ) merge with the
    preceding mora; ー and っ are their own mora.

    Skip entries whose reading contains "/" or " " — those are
    multi-reading aliases (e.g., "なに / なん") which are a schema
    quirk, not a real mora-count issue (V-2 in native-teacher review
    documents this).

    Catches the class of bug fixed in commit d870a2e+1 (110 LLM-curated
    entries had systematically wrong mora counts, all corrected via
    kanjium re-lookup + mechanical fallback).
    """
    failures: list[str] = []
    SMALL_KANA = set("ゃゅょぁぃぅぇぉャュョァィゥェォ")

    def count_mora(s: str) -> int:
        return sum(1 for ch in (s or "") if ch not in SMALL_KANA)

    path = ROOT / "data" / "vocab.json"
    if not path.exists():
        return ["JA-70: data/vocab.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-70: parse error: {e}"]

    for v in data.get("entries", []):
        reading = v.get("reading") or ""
        if "/" in reading or " " in reading:
            continue  # multi-reading alias — skip
        pa = v.get("pitch_accent")
        if not isinstance(pa, dict) or "mora" not in pa:
            continue
        expected = count_mora(reading)
        actual = pa.get("mora")
        if actual != expected:
            failures.append(
                f"JA-70 {v.get('id')}: pitch_accent.mora={actual} but "
                f"count_mora('{reading}')={expected}"
            )
    return failures


def _check_ja_71_meaning_ja_pattern_alignment() -> list[str]:
    """JA-71 (2026-05-13): grammar's meaning_ja text's first 「marker」
    must reference the same grammar point as the pattern field.
    Catches the cross-contamination class (commit d870a2e+1 fixed 13
    patterns whose meaning_ja described a different rule entirely).

    Heuristic: extract the first 「marker」 substring from meaning_ja;
    require that AT LEAST ONE non-trivial character from the marker
    appears in the pattern field. Excludes patterns whose pattern
    field is empty or just a tilde-placeholder (see_also stubs).
    """
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-71: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-71: parse error: {e}"]

    QUOTE_RE = re.compile(r"「([^」]+)」")
    TRIVIAL_CHARS = set("〜～「」、。 　・のな")
    # JA-71 false-positive guard: skip patterns whose `pattern` field
    # is Latin-only abstract notation (e.g., "V-plain + N", "Adj + N",
    # "Verb-stem + たいです"). For those, the meaning_ja markers can't
    # be expected to share characters with Latin tokens.
    LATIN_ONLY_RE = re.compile(r"^[A-Za-z0-9 \-\+\(\)\.\/\,]+$")
    NON_KANA_CHAR_RE = re.compile(r"[ぁ-んァ-ン一-鿿]")

    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        pattern = (p.get("pattern") or "").strip()
        meaning_ja = (p.get("meaning_ja") or "").strip()
        if not pattern or pattern in ("〜", "～", "?"):
            continue  # see_also stub — skip
        if not meaning_ja:
            continue
        # Skip patterns with no Japanese characters at all (Latin-only
        # abstract notation like "Verb-stem + たいです" — actually this
        # has Japanese, ok. But "V-plain + N" or "Adj + N" do not).
        if not NON_KANA_CHAR_RE.search(pattern):
            continue
        first_marker_match = QUOTE_RE.search(meaning_ja)
        if not first_marker_match:
            continue  # no quoted marker — skip (some entries are descriptive)
        marker = first_marker_match.group(1)
        # Extract non-trivial chars from marker; require overlap with pattern
        marker_chars = {c for c in marker if c not in TRIVIAL_CHARS}
        pattern_chars = {c for c in pattern if c not in TRIVIAL_CHARS}
        # Fallback pass: if the marker doesn't overlap, check if the
        # FULL meaning_ja text contains the pattern's Japanese kana.
        # This passes cases like n5-065 where the pattern is
        # "Verb-る / Verb-う" and meaning_ja uses example verbs
        # (のむ, 食べる, する) — the kana る IS in meaning_ja, just not
        # in the first 「marker」.
        # Threshold note (run-2 review): tested ≥2-char overlap to catch
        # n5-151/n5-183 cross-contaminations, but it produced 17 false
        # positives on single-character particle patterns (や/も/か/ね/
        # よ/が/に — common N5 grammar markers). Reverted to ≥1.
        # JA-71 catches the egregious systematic-misalignment class but
        # not subtle cases where 1 character coincidentally overlaps;
        # those require manual native-teacher review.
        if marker_chars and pattern_chars and not (marker_chars & pattern_chars):
            full_overlap = pattern_chars & set(meaning_ja)
            if not full_overlap:
                failures.append(
                    f"JA-71 {pid}: meaning_ja first 「marker」='{marker}' shares "
                    f"no characters with pattern field '{pattern}', and full "
                    f"meaning_ja text also doesn't reference the pattern. "
                    f"Possible cross-contamination."
                )
    return failures


def _check_ja_68_cache_version_sync() -> list[str]:
    """JA-68 (2026-05-13): the cache version must match across 3 places:
    index.html's CSS query string, index.html's JS query string, and
    sw.js's CACHE_VERSION constant. Forgotten bumps in any one breaks
    the cache-bust on release.
    """
    failures: list[str] = []
    try:
        index_html = (ROOT / "index.html").read_text(encoding="utf-8")
        sw_js = (ROOT / "sw.js").read_text(encoding="utf-8")
    except Exception as e:
        return [f"JA-68: load error: {e}"]

    css_m = re.search(r'css/main\.min\.css\?v=(\d+\.\d+\.\d+)', index_html)
    js_m = re.search(r'js/min/app\.js\?v=(\d+\.\d+\.\d+)', index_html)
    sw_m = re.search(r"CACHE_VERSION\s*=\s*'jlptsuccess-n5-v(\d+\.\d+\.\d+)'", sw_js)

    if not css_m:
        failures.append("JA-68: index.html missing css/main.min.css?v=X.Y.Z marker")
    if not js_m:
        failures.append("JA-68: index.html missing js/min/app.js?v=X.Y.Z marker")
    if not sw_m:
        failures.append("JA-68: sw.js missing CACHE_VERSION = 'jlptsuccess-n5-vX.Y.Z' marker")
    if css_m and js_m and sw_m:
        versions = {css_m.group(1), js_m.group(1), sw_m.group(1)}
        if len(versions) > 1:
            failures.append(
                f"JA-68: cache versions out of sync — "
                f"css={css_m.group(1)}, js={js_m.group(1)}, sw={sw_m.group(1)}"
            )
    return failures


def _check_ja_53_grammar_cultural_callout() -> list[str]:
    """Every grammar pattern must carry a `cultural_callout` field with
    non-empty content. Locks in P2-12 closure (cultural_callout 178/178
    coverage from audit round 12)."""
    failures: list[str] = []
    path = ROOT / "data" / "grammar.json"
    if not path.exists():
        return ["JA-53: data/grammar.json missing"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-53: parse error: {e}"]
    for p in data.get("patterns", []):
        pid = p.get("id", "?")
        cc = p.get("cultural_callout")
        if not cc:
            failures.append(f"JA-53 {pid}: missing cultural_callout")
        elif isinstance(cc, str) and len(cc.strip()) < 20:
            # The field exists but contains a trivial placeholder
            failures.append(
                f"JA-53 {pid}: cultural_callout too short ({len(cc.strip())} chars, need >=20)"
            )
    return failures


# ---------------------------------------------------------------------------
# JA-81 added for boilerplate-leak detection (2026-05-14)
# ---------------------------------------------------------------------------

def _check_ja_82_meta_paths_resolve() -> list[str]:
    """IMP-006 (2026-05-14): every path string mentioned in any _meta.see_also
    or _meta.consumers entry across data/*.json must resolve to an existing
    file or directory. Catches the silent-rot bug where docs reference paths
    that no longer exist after refactors / moves / deletions.

    Each path string is the entire entry, possibly with a trailing
    parenthetical comment — e.g. "data/grammar.json (per-pattern source)".
    We split off the comment, take the first whitespace-delimited token, and
    check whether that resolves under ROOT.
    """
    from pathlib import Path
    failures = []
    data_dir = ROOT / "data"
    if not data_dir.is_dir():
        return ["JA-82: data/ directory missing"]

    def extract_path(entry):
        """Take the path-like prefix from an entry. Entries are often
        formatted as 'data/file.json (comment)' or just 'data/file.json'."""
        if not isinstance(entry, str):
            return None
        # Strip parenthetical or hyphen-comment suffix
        s = entry.split("(")[0].split(" - ")[0].split(" — ")[0].strip()
        # First whitespace-delimited token is the path
        s = s.split()[0] if s.split() else s
        # Strip a few common comment prefixes that some entries use
        if s.startswith(("tools/", "data/", "js/", "css/", "svg/", "fonts/",
                         "locales/", "audio/", "KnowledgeBank/", "docs/",
                         "feedback/", "prompts/", "scripts/", "tests/",
                         "specifications/", "not-required/", "build/",
                         "assets/")):
            return s
        # Allow root-level files (CHANGELOG.md, README.md, PRIVACY.md, etc.)
        if s and "/" not in s and "." in s:
            return s
        return None  # Not a recognizable path

    def walk_for_meta(obj, json_file):
        """Find any _meta dict in the tree and check its see_also/consumers."""
        if isinstance(obj, dict):
            if "_meta" in obj and isinstance(obj["_meta"], dict):
                meta = obj["_meta"]
                for field in ("see_also", "consumers", "consumers_planned",
                              "consumers_current"):
                    entries = meta.get(field, [])
                    if isinstance(entries, list):
                        for entry in entries:
                            p = extract_path(entry)
                            if p is None:
                                continue
                            full = ROOT / p
                            if not full.exists():
                                failures.append(
                                    f"JA-82 {json_file}: _meta.{field} references "
                                    f"non-existent path {p!r}"
                                )
            for v in obj.values():
                walk_for_meta(v, json_file)
        elif isinstance(obj, list):
            for item in obj:
                walk_for_meta(item, json_file)

    # Walk every data/*.json + data/*/*.json + a few well-known top-level
    # *.meta.json files in data/
    for jpath in list(data_dir.glob("*.json")) + list(data_dir.glob("*/*.json")):
        try:
            d = json.loads(jpath.read_text(encoding="utf-8"))
        except Exception:
            continue
        walk_for_meta(d, str(jpath.relative_to(ROOT)).replace("\\", "/"))

    return failures[:50]  # cap noise


def _check_ja_90_pitch_accent_reference_agreement() -> list[str]:
    """2026-05-15: pitch-accent values validated against kanjium reference.

    For every vocab entry's `pitch_accent.drop`, enforce ONE of:
      (a) confidence == 'high' AND drop ∈ reference_drops_for_this_vocab_id
      (b) confidence in {'medium', 'low', 'unverified'}
          (no reference match required — kept as-is from LLM authoring;
           future native-human review pass should prioritize these)

    Mora count (`pitch_accent.mora`) is separately enforced by JA-70.
    """
    failures = []
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
        ref = json.loads((ROOT / "data" / "n5_pitch_accent_reference.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-90 could not read inputs: {e}"]
    ref_idx = {r["vocab_id"]: r for r in ref.get("entries", [])}
    for e in vocab.get("entries", []):
        vid = e.get("id", "?")
        pa = e.get("pitch_accent")
        if not isinstance(pa, dict):
            continue  # JA-70 handles structural; we only validate drop here
        confidence = pa.get("confidence", "unverified")
        drop = pa.get("drop")
        if drop is None:
            continue
        if confidence == "high":
            ref_entry = ref_idx.get(vid)
            if ref_entry is None:
                failures.append(
                    f"JA-90 {vid} has confidence='high' but no reference "
                    f"entry exists — must be 'unverified'"
                )
                continue
            ref_drops = ref_entry.get("drops", [])
            if drop not in ref_drops:
                failures.append(
                    f"JA-90 {vid} pitch_accent.drop={drop} disagrees with "
                    f"reference {ref_drops} (confidence='high')"
                )
        elif confidence not in {"medium", "low", "unverified"}:
            failures.append(
                f"JA-90 {vid} has invalid pitch_accent.confidence={confidence!r}"
            )
    return failures


def _check_ja_89_native_teacher_phase_b_d_locks() -> list[str]:
    """2026-05-15 native-teacher audit (waves A+B+D) — three regression
    guards in one check:

    LOCK 1: Counter-noun pairing for the native-counter-series entries
            (三つ..九つ) — the example noun must be canonically
            つ-counted (not books → 冊).

    LOCK 2: No bare-article 'There is X.' / 'I see X.' / 'I eat X.'
            etc. for singular common nouns. Plural and mass nouns
            are allowed (chopsticks, water, sweets).

    LOCK 3: No animate-noun + が あります template (animacy violation).
            Animate nouns must use います.

    LOCK 4: No new instances of 'あの Xは どこですか' grammar template
            leaking into grammar.json.

    LOCK 5: No 「毎日 X ことが できます」 nor 「あした X つもりです」
            templates in vocab examples.
    """
    import re
    failures = []
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
        grammar = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-89 could not read corpora: {e}"]

    # LOCK 1: native-counter-series (三つ..九つ) — example index [1]
    # should not use 「本」 with つ counter (books take 冊)
    bad_counter_entries = {f"n5.vocab.8-native-counters-series.{form}"
                           for form in ("三つ", "四つ", "六つ", "七つ", "八つ", "九つ")}
    for e in vocab["entries"]:
        if e["id"] in bad_counter_entries:
            for i, ex in enumerate(e.get("examples", []) or []):
                if "本が" in (ex.get("ja") or "") and "つ" in (ex.get("ja") or ""):
                    failures.append(
                        f"JA-89 (lock1) {e['id']}[{i}] uses 本+つ template; "
                        f"books take 冊 counter")
    # LOCK 2: bare-article 'There is X' on singular common nouns
    bare_pat = re.compile(r"^(There is|There are|I (?:see|eat|have|drink|use)) ([a-z]+)\.?$")
    allowed_bare = {"water","rice","music","snow","rain","fire","milk","sugar","salt",
                    "tea","coffee","bread","meat","fish","chopsticks","sweets",
                    "vegetables","noodles","soup","curry"}
    for e in vocab["entries"]:
        for i, ex in enumerate(e.get("examples", []) or []):
            en = (ex.get("translation_en") or "").strip()
            m = bare_pat.fullmatch(en)
            if m and m.group(2) not in allowed_bare:
                failures.append(
                    f"JA-89 (lock2) {e['id']}[{i}] bare-article 'X' in EN: {en!r}")
    # LOCK 3: animacy violations
    animate = ['男の子','女の子','子ども','犬','いぬ','ねこ','とり','うま','うし',
               'ぶた','にわとり','ぞう','学生','先生','友だち','家族','かぞく',
               'お父さん','お母さん','父','母','兄','姉','弟','妹','日本人',
               'スペイン人','イギリス人','アメリカ人','フランス人','ドイツ人',
               '中国人','かんこく人','そぼ','そふ','おばあさん','おじいさん',
               'おばさん','おじさん']
    animacy_rx = re.compile(
        r"(?<![ぁ-ゟ一-鿿])(" + "|".join(re.escape(n) for n in animate) +
        r")(?:は|が)\s*あります"
    )
    for e in vocab["entries"]:
        for i, ex in enumerate(e.get("examples", []) or []):
            ja = (ex.get("ja") or "").strip()
            m = animacy_rx.search(ja)
            if m:
                failures.append(
                    f"JA-89 (lock3) {e['id']}[{i}] animacy: '{m.group(1)}が あります' should be います: {ja!r}")
    # LOCK 4: 'あの Xは どこですか' template in grammar.json
    doko_rx = re.compile(r"^あの\s+(.{1,10})は\s*どこですか。?$")
    for p in grammar["patterns"]:
        for i, ex in enumerate(p.get("examples", []) or []):
            ja = (ex.get("ja") or "").strip()
            if doko_rx.fullmatch(ja):
                failures.append(
                    f"JA-89 (lock4) {p['id']}[{i}] 'あの Xは どこですか' template: {ja!r}")
    # LOCK 5: '毎日 X ことが できます' / 'あした X つもりです' templates
    t1_rx = re.compile(r"^毎日\s+.*ことが\s+できます。?$")
    t2_rx = re.compile(r"^あした\s+.*つもりです。?$")
    for e in vocab["entries"]:
        for i, ex in enumerate(e.get("examples", []) or []):
            ja = (ex.get("ja") or "").strip()
            if t1_rx.fullmatch(ja):
                failures.append(
                    f"JA-89 (lock5a) {e['id']}[{i}] '毎日 X ことが できます' template")
            if t2_rx.fullmatch(ja):
                failures.append(
                    f"JA-89 (lock5b) {e['id']}[{i}] 'あした X つもりです' template")
    return failures


def _check_ja_87_cross_corpus_consistency() -> list[str]:
    """Wave-1 (2026-05-15): cross-corpus reading/gloss consistency.

    For every place that quotes a vocab term across corpora, the
    inline reading/gloss must agree with the canonical entry in
    vocab.json. Specifically:
      - reading.json vocab_preview[].reading / .gloss / .gloss_hi
      - listening.json vocab_glossary[].reading / .gloss
      - authentic.json items[].reading when ja == single vocab form
    """
    import re
    failures = []
    try:
        v = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
        vocab_idx = {e["id"]: e for e in v["entries"]}
    except Exception as e:
        return [f"JA-87 could not read vocab.json: {e}"]

    def _ok(it_str, v_read, v_readings):
        if it_str == v_read:
            return True
        all_v = set([v_read] + list(v_readings))
        tokens = [t.strip() for t in it_str.replace("/", " ").split() if t.strip()]
        return tokens and all(tok in all_v for tok in tokens)

    # reading.json
    try:
        r = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
    except Exception as e:
        return failures + [f"JA-87 could not read reading.json: {e}"]
    for p in r.get("passages", []):
        pid = p.get("id", "?")
        for item in p.get("vocab_preview") or []:
            if not isinstance(item, dict):
                continue
            vid = item.get("vocab_id")
            canonical = vocab_idx.get(vid)
            if not canonical:
                continue
            it_read = (item.get("reading") or "").strip()
            v_read = (canonical.get("reading") or "").strip()
            v_readings = canonical.get("readings", []) or []
            if it_read and v_read and not _ok(it_read, v_read, v_readings):
                failures.append(f"JA-87 {pid} vocab_preview reading {it_read!r} != vocab.json {v_read!r} for {vid}")
            it_gloss = (item.get("gloss") or "").strip()
            v_gloss = (canonical.get("gloss") or "").strip()
            if it_gloss and v_gloss and it_gloss != v_gloss:
                failures.append(f"JA-87 {pid} vocab_preview gloss {it_gloss!r} != vocab.json {v_gloss!r} for {vid}")
            it_hi = (item.get("gloss_hi") or "").strip()
            v_hi = (canonical.get("gloss_hi") or "").strip()
            if it_hi and v_hi and it_hi != v_hi:
                failures.append(f"JA-87 {pid} vocab_preview gloss_hi mismatch for {vid}")
    # listening.json
    try:
        ls = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
    except Exception as e:
        return failures + [f"JA-87 could not read listening.json: {e}"]
    for it in ls.get("items", []):
        iid = it.get("id", "?")
        for gx in it.get("vocab_glossary") or []:
            if not isinstance(gx, dict):
                continue
            vid = gx.get("vocab_id")
            canonical = vocab_idx.get(vid)
            if not canonical:
                continue
            r_val = (gx.get("reading") or "").strip()
            v_read = (canonical.get("reading") or "").strip()
            v_readings = canonical.get("readings", []) or []
            if r_val and v_read and not _ok(r_val, v_read, v_readings):
                failures.append(f"JA-87 {iid} vocab_glossary reading {r_val!r} != vocab.json {v_read!r} for {vid}")
            g = (gx.get("gloss") or "").strip()
            vg = (canonical.get("gloss") or "").strip()
            if g and vg and g != vg:
                failures.append(f"JA-87 {iid} vocab_glossary gloss {g!r} != vocab.json {vg!r} for {vid}")
    # authentic.json single-word cards
    try:
        au = json.loads((ROOT / "data" / "authentic.json").read_text(encoding="utf-8"))
    except Exception as e:
        return failures + [f"JA-87 could not read authentic.json: {e}"]
    for it in au.get("items", []):
        iid = it.get("id", "?")
        refs = it.get("vocab_refs") or []
        if len(refs) != 1:
            continue
        vid = refs[0] if isinstance(refs[0], str) else refs[0].get("vocab_id")
        canonical = vocab_idx.get(vid)
        if not canonical:
            continue
        if (it.get("ja") or "").strip() != (canonical.get("form") or "").strip():
            continue
        r_val = (it.get("reading") or "").strip()
        v_read = (canonical.get("reading") or "").strip()
        v_readings = canonical.get("readings", []) or []
        if r_val and v_read and not _ok(r_val, v_read, v_readings):
            failures.append(f"JA-87 {iid} authentic reading {r_val!r} != vocab.json {v_read!r} for {vid}")
    return failures


def _check_ja_88_particle_precision() -> list[str]:
    """Wave-2 (2026-05-15): scan every Japanese sentence across all 12
    content corpora for the top L2-error particle / conjugation traps.

    Locks zero-finding floor caught by the 2026-05-15 mega-audit.
    """
    import re
    failures = []
    # Patterns + tags
    PATTERNS = [
        ("WO-WITH-STATIVE-ADJ",
         re.compile(r"[をｦ]\s*(?:が\s*)?(?:大?好き|だいすき|すき|じょうず|上手|へた|下手|嫌い|きらい|ほしい|欲しい|わか(?:り|る|れ)|分か(?:り|る|れ))")),
        ("WO-WITH-ARU-IRU",
         re.compile(r"[をｦ]\s*(?:が\s*)?(?:あります|います|ある|いる)\b")),
        ("IADJ-NEG-WITH-DA", re.compile(r"(?<![ぁ-ゟ一-鿿])くないだ(?![け])")),
        ("DOUBLE-MASHITA", re.compile(r"ました(?:した|ました)")),
        ("DOUBLE-DESU", re.compile(r"ですです")),
        ("DOUBLE-PARTICLE-DENI", re.compile(r"(?<!ひとり)でに\s")),
        ("DOUBLE-PARTICLE-NIDE", re.compile(r"にで\s")),
        ("NAADJ-AS-IADJ", re.compile(r"きれいい|げんきい|しずかい")),
        ("STUTTER-MASHITA", re.compile(r"きました\s+ました")),
        ("DOUBLE-KA", re.compile(r"(?:です|ます)かか")),
    ]

    def scan(corpus, iid, field, ja):
        for tag, rx in PATTERNS:
            m = rx.search(ja)
            if m:
                failures.append(
                    f"JA-88 {corpus}/{iid}.{field}: {tag} matched {m.group(0)!r} in {ja[:60]!r}"
                )

    try:
        # grammar
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
        for p in g.get("patterns", []):
            for i, ex in enumerate(p.get("examples") or []):
                ja = (ex.get("ja") or "").strip()
                if ja:
                    scan("grammar", p.get("id", "?"), f"examples[{i}].ja", ja)
        # vocab
        v = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
        for e in v.get("entries", []):
            for i, ex in enumerate(e.get("examples") or []):
                ja = (ex.get("ja") or "").strip()
                if ja:
                    scan("vocab", e.get("id", "?"), f"examples[{i}].ja", ja)
        # kanji
        k = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
        for e in k.get("entries", []):
            for i, s in enumerate(e.get("sentences") or []):
                ja = (s.get("ja") or "").strip()
                if ja:
                    scan("kanji", e.get("glyph", "?"), f"sentences[{i}].ja", ja)
        # reading
        r = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
        for p in r.get("passages", []):
            ja = (p.get("ja") or "").strip()
            if ja:
                scan("reading", p.get("id", "?"), "ja", ja)
        # listening
        ls = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
        for it in ls.get("items", []):
            ja = (it.get("script_ja") or "").strip()
            if ja:
                scan("listening", it.get("id", "?"), "script_ja", ja)
        # authentic
        au = json.loads((ROOT / "data" / "authentic.json").read_text(encoding="utf-8"))
        for it in au.get("items", []):
            ja = (it.get("ja") or "").strip()
            if ja:
                scan("authentic", it.get("id", "?"), "ja", ja)
        # questions
        qj = json.loads((ROOT / "data" / "questions.json").read_text(encoding="utf-8"))
        for q in qj.get("questions", []):
            ja = (q.get("question_ja") or "").strip()
            if ja:
                scan("questions", q.get("id", "?"), "question_ja", ja)
        # papers
        for cat in ("dokkai", "bunpou", "goi", "moji"):
            pdir = ROOT / "data" / "papers" / cat
            if not pdir.exists():
                continue
            for pf in sorted(pdir.glob("paper-*.json")):
                paper = json.loads(pf.read_text(encoding="utf-8"))
                for q in paper.get("questions") or []:
                    for fld in ("stem_html", "passage_text"):
                        ja = (q.get(fld) or "").strip()
                        if ja:
                            scan(f"papers/{cat}", q.get("id", "?"), f"{pf.name}.{fld}", ja)
    except Exception as e:
        failures.append(f"JA-88 internal error: {e}")
    return failures


def _check_ja_86_mega_audit_locale_coverage() -> list[str]:
    """Regression guard for the 2026-05-15 mega-audit (commit ceb64fc..):
      - authentic.json: every item with non-empty `context` must have a
        non-empty `context_hi` (88 filled in round-1)
      - questions.json: every MCQ with `distractor_explanations` (EN)
        must have parallel `distractor_explanations_hi` covering all
        non-correct choices (375 EN + 375 HI filled in round-2)
    """
    failures = []
    # authentic.json
    try:
        auth = json.loads((ROOT / "data" / "authentic.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-86 could not read authentic.json: {e}"]
    for it in auth.get("items", []):
        iid = it.get("id", "?")
        if (it.get("context") or "").strip() and not (it.get("context_hi") or "").strip():
            failures.append(f"JA-86 {iid} has context but empty context_hi")
    # questions.json distractor parity (MCQ only)
    try:
        qj = json.loads((ROOT / "data" / "questions.json").read_text(encoding="utf-8"))
    except Exception as e:
        return failures + [f"JA-86 could not read questions.json: {e}"]
    for q in qj.get("questions", []):
        if q.get("type") != "mcq":
            continue
        qid = q.get("id", "?")
        choices = q.get("choices") or []
        ans = q.get("correctAnswer")
        if not choices or not ans:
            continue
        de_en = q.get("distractor_explanations") or {}
        de_hi = q.get("distractor_explanations_hi") or {}
        distractors = [c for c in choices if c != ans]
        for dist in distractors:
            if not (de_en.get(dist) or "").strip():
                failures.append(f"JA-86 {qid}.distractor_explanations missing entry for {dist!r}")
            if not (de_hi.get(dist) or "").strip():
                failures.append(f"JA-86 {qid}.distractor_explanations_hi missing entry for {dist!r}")
    return failures


def _check_ja_85_dokkai_locale_parity() -> list[str]:
    """Regression guard for the 2026-05-15 dokkai audit.

    For every reading.json question with a non-empty explanation_en,
    explanation_hi must also be non-empty (Hindi-locale parity — 83
    questions filled in round-2 from 19% coverage to 100%). Same rule
    for paper-dokkai questions: rationale present implies rationale_hi
    present.

    Also requires format_role to be set on every reading.json question
    (mock-test-mode filter at js/reading.js:60 depends on it).
    """
    failures = []
    try:
        reading = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-85 could not read reading.json: {e}"]
    for p in reading.get("passages", []):
        pid = p.get("id", "?")
        for q in p.get("questions", []) or []:
            qid = q.get("id", f"{pid}.q?")
            if (q.get("explanation_en") or "").strip() and not (q.get("explanation_hi") or "").strip():
                failures.append(
                    f"JA-85 {qid} has explanation_en but empty explanation_hi"
                )
            if not (q.get("format_role") or "").strip():
                failures.append(f"JA-85 {qid} missing format_role")
    # Paper-dokkai locale parity
    papers_dir = ROOT / "data" / "papers" / "dokkai"
    if papers_dir.exists():
        for pf in sorted(papers_dir.glob("paper-*.json")):
            try:
                p = json.loads(pf.read_text(encoding="utf-8"))
            except Exception as e:
                failures.append(f"JA-85 could not read {pf.name}: {e}")
                continue
            for q in p.get("questions", []) or []:
                qid = q.get("id", f"{pf.stem}.q?")
                if (q.get("rationale") or "").strip() and not (q.get("rationale_hi") or "").strip():
                    failures.append(
                        f"JA-85 {qid} has rationale but empty rationale_hi"
                    )
    return failures


def _check_ja_84_kanji_sentence_translations() -> list[str]:
    """Every kanji.json entry's `sentences[*].ja` that is non-empty
    must have a non-empty `translation_en`. The 2026-05-15 kanji
    audit added 55 missing translations + the unstripped <u>...</u>
    markup that the renderer was escaping. This regression guard
    catches both classes.
    """
    failures = []
    try:
        kanji = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-84 could not read kanji.json: {e}"]
    for e in kanji.get("entries", []):
        glyph = e.get("glyph", "?")
        for i, s in enumerate(e.get("sentences", []) or []):
            ja = (s.get("ja") or "").strip()
            en = (s.get("translation_en") or "").strip()
            if ja and not en:
                failures.append(
                    f"JA-84 {glyph}.sentences[{i}] has non-empty ja but empty "
                    f"translation_en: {ja[:50]!r}"
                )
            # HTML markup leak in ja (renderer escapes — displays literally)
            if "<u>" in ja or "</u>" in ja:
                failures.append(
                    f"JA-84 {glyph}.sentences[{i}].ja contains HTML <u> markup "
                    f"that the renderer would display literally: {ja[:50]!r}"
                )
    return failures


def _check_ja_83_no_vocab_template_leak() -> list[str]:
    """No vocab example may use the literal templates that the
    2026-05-15 vocab audit removed across 539 entries:
      - 'Xを 見ました。'              (X is single token)
      - 'あの Xは どこですか。'         (where X equals the headword and
                                       the section is not location-like)
      - 'これは Xです。' / 'あれは Xです。' on entries with 3+ examples,
        excluding the demonstrative-section entries (where it IS the
        canonical demonstration of the headword).
    Regression guard for the cleanup so the templates don't drift back
    in via future authoring batches.
    """
    import re
    failures = []
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-83 could not read vocab.json: {e}"]
    pat_miru = re.compile(r"^[^\s]{1,10}を\s*見ました(?:よ|ね|よね)?。?$")
    pat_doko = re.compile(r"^あの\s+(.{1,10})は\s*どこですか。?$")
    pat_kore = re.compile(r"^(これ|あれ|それ)は\s*(.{1,10})です。?$")
    pat_quote = re.compile(r"^「(.{1,10})」と\s*あいさつしました。?$")
    GREETINGS = {"おはよう","こんにちは","こんばんは","ありがとう","さようなら",
                 "おやすみ","ただいま","おかえり","もしもし","いってきます",
                 "いってらっしゃい","おかげさまで"}
    for e in vocab.get("entries", []):
        eid = e.get("id", "?")
        form = e.get("form", "")
        section = e.get("section", "")
        pos = e.get("pos", "")
        is_locationish = "13-locations" in section or "26-house" in section
        is_demonstrative = pos == "demonstrative" or "demonstrative" in section.lower()
        for i, ex in enumerate(e.get("examples", [])):
            ja = (ex.get("ja") or "").strip()
            if pat_miru.fullmatch(ja):
                failures.append(f"JA-83 {eid}[{i}]: template 'Xを 見ました' leaked back: {ja!r}")
            m = pat_doko.fullmatch(ja)
            if m and not is_locationish and m.group(1).strip() == form.strip():
                failures.append(f"JA-83 {eid}[{i}]: template 'あの Xは どこですか' non-location: {ja!r}")
            m = pat_kore.fullmatch(ja)
            if m and not is_demonstrative and len(e.get("examples", [])) >= 3:
                failures.append(f"JA-83 {eid}[{i}]: bare 'これ/あれは Xです' on 3+ex entry: {ja!r}")
            m = pat_quote.fullmatch(ja)
            if m and m.group(1) not in GREETINGS:
                failures.append(f"JA-83 {eid}[{i}]: 「X」と あいさつしました with non-greeting X: {ja!r}")
    return failures


def _check_ja_100_kanji_vocab_form_consistency() -> list[str]:
    """BUG-020 + BUG-023 (2026-05-17) regression guard — STRICT version.

    For every kanji.json `n5_compounds[i]` and `examples[i]` that
    carries a `vocab_id`, assert TWO things:

    1. The vocab_id resolves (linked vocab entry exists).
    2. The kanji.json `form` equals the linked vocab.json `form`
       EXACTLY.

    First wired as narrow OOS-only check post-BUG-020. The BUG-023
    audit demonstrated the narrow check missed a real bug class:
    kanji.json showed kanji forms (友だち, 手, 上手, 足, 目) while
    vocab.json had kana-only forms (ともだち, て, じょうず, あし, め)
    — all 5 kanji are in scope, so the narrow check accepted this
    as "intentional pedagogy". It wasn't; standard N5 textbooks
    teach those words with kanji.

    Strict equality catches both bug directions:
      BUG-020: kanji.json kanji form → vocab.json kana form (OOS
               kanji in kanji corpus; vocab.json was right)
      BUG-023: kanji.json kanji form → vocab.json kana form (kanji
               IS in scope; vocab.json was wrong)
    """
    failures = []
    try:
        kanji = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-100 could not read inputs: {e}"]
    vocab_form_by_id = {e.get("id"): e.get("form") for e in vocab.get("entries", []) if e.get("id")}
    for k in kanji.get("entries", []):
        glyph = k.get("glyph", "?")
        for label, lst in (("n5_compounds", k.get("n5_compounds") or []),
                           ("examples", k.get("examples") or [])):
            for i, item in enumerate(lst):
                if not isinstance(item, dict):
                    continue
                vid = item.get("vocab_id")
                if not vid:
                    continue  # entries without vocab_id are unlinked, can't cross-check
                form_here = item.get("form") or item.get("lemma") or ""
                expected = vocab_form_by_id.get(vid)
                if expected is None:
                    failures.append(
                        f"JA-100 kanji={glyph} {label}[{i}]: vocab_id {vid!r} not in vocab.json"
                    )
                elif form_here != expected:
                    failures.append(
                        f"JA-100 kanji={glyph} {label}[{i}]: form {form_here!r} != "
                        f"linked vocab.form {expected!r} (vocab_id={vid})"
                    )
    return failures


def _check_ja_101_kanji_examples_form_field_only() -> list[str]:
    """BUG-022 (2026-05-17) regression guard.

    kanji.json `examples[i]` objects must use `form` field name only;
    `lemma` is deprecated by the BUG-022 migration. (The schema lives
    on vocab.json's `form` convention; provenance signal stays on
    `auto_derived: true` + `vocab_id`.)
    """
    failures = []
    try:
        kanji = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-101 could not read kanji.json: {e}"]
    for k in kanji.get("entries", []):
        glyph = k.get("glyph", "?")
        for i, ex in enumerate(k.get("examples") or []):
            if not isinstance(ex, dict):
                continue
            if "lemma" in ex:
                failures.append(
                    f"JA-101 kanji={glyph} examples[{i}]: stale `lemma` field "
                    f"(use `form` only post-BUG-022 migration)"
                )
            if "form" not in ex:
                failures.append(
                    f"JA-101 kanji={glyph} examples[{i}]: missing `form` field"
                )
    return failures


def _check_ja_104_reading_difficulty() -> list[str]:
    """BUG-041 (2026-05-17) regression guard.

    reading.json passages: `level` field is RETIRED. Replaced by
    `difficulty` from closed enum {easy, medium, hard} OR absent
    (for info-search passages whose type is in format_role).
    """
    failures = []
    ALLOWED = {"easy", "medium", "hard"}
    try:
        r = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-104 could not read reading.json: {e}"]
    for p in r.get("passages", []):
        pid = p.get("id", "?")
        if "level" in p:
            failures.append(f"JA-104 {pid}: stale `level` field present (use `difficulty` per BUG-041)")
        d = p.get("difficulty")
        if d is not None and d not in ALLOWED:
            failures.append(f"JA-104 {pid}: difficulty {d!r} not in closed enum {sorted(ALLOWED)}")
    return failures


def _check_ja_105_vocab_preview_shape() -> list[str]:
    """BUG-045 (2026-05-17) regression guard.

    reading.json `vocab_preview` is a list of vocab_id strings only,
    never a list of dicts.
    """
    failures = []
    try:
        r = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-105 could not read reading.json: {e}"]
    for p in r.get("passages", []):
        pid = p.get("id", "?")
        vp = p.get("vocab_preview")
        if vp is None:
            continue
        if not isinstance(vp, list):
            failures.append(f"JA-105 {pid}: vocab_preview must be a list; got {type(vp).__name__}")
            continue
        for i, item in enumerate(vp):
            if not isinstance(item, str):
                failures.append(
                    f"JA-105 {pid}: vocab_preview[{i}] must be a string vocab_id; got {type(item).__name__}"
                )
    return failures


def _check_ja_106_format_type_enum() -> list[str]:
    """BUG-044 (2026-05-17) regression guard.

    reading.json `format_type` field: must be null OR in
    {schedule_table, menu_list, notice}. The legacy "comprehension"
    value is RETIRED — it was a redundant duplicate of
    format_role="comprehension".
    """
    failures = []
    ALLOWED = {"schedule_table", "menu_list", "notice"}
    try:
        r = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-106 could not read reading.json: {e}"]
    for p in r.get("passages", []):
        pid = p.get("id", "?")
        ft = p.get("format_type")
        if ft is None:
            continue
        if ft not in ALLOWED:
            failures.append(
                f"JA-106 {pid}: format_type {ft!r} not in {sorted(ALLOWED)} (legacy "
                f"'comprehension' retired per BUG-044)"
            )
    return failures


def _check_ja_92_en_boilerplate() -> list[str]:
    """BUG-003/005 boilerplate guard on EN side (2026-05-17 wire-up).

    Parallel to JA-81 (JA-side boilerplate guard). If the same
    `translation_en` string appears in ≥10 grammar examples across
    distinct patterns, that's likely a placeholder / copy-paste
    artifact (BUG-005 class: translation_en stuck on wrong content).

    The threshold matches JA-81 (10 patterns).
    """
    failures: list[str] = []
    try:
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-92 could not read grammar.json: {e}"]
    # Map each EN string → set of pattern IDs where it appears
    occurrences: dict[str, list[str]] = {}
    for p in g.get("patterns") or []:
        pid = p.get("id", "?")
        for ex in p.get("examples") or []:
            if not isinstance(ex, dict):
                continue
            en = ex.get("translation_en")
            if not isinstance(en, str) or not en.strip():
                continue
            en_norm = en.strip()
            if len(en_norm) < 10:
                # Skip very-short EN strings (e.g., "Yes." / "No.") —
                # legitimate repetition.
                continue
            occurrences.setdefault(en_norm, [])
            if pid not in occurrences[en_norm]:
                occurrences[en_norm].append(pid)
    THRESHOLD = 10
    for en, pids in occurrences.items():
        if len(pids) >= THRESHOLD:
            failures.append(
                f"JA-92 EN translation {en[:60]!r} appears in "
                f"{len(pids)} patterns (≥ {THRESHOLD} threshold; "
                f"first 8: {','.join(pids[:8])})"
            )
    return failures


def _check_ja_93_pitch_marks_mora() -> list[str]:
    """BUG-004 algorithmic mora-count guard (2026-05-17 wire-up).

    For every vocab entry that carries `pitch_marks`, verify that
    the total mora count (sum of per-syllable mora values, or
    len(pitch_marks) when each entry is one mora) matches the
    algorithmically-computed mora count from the entry's `form`
    field using count_morae() (the canonical algorithm in
    not-required/tools-archive/fix_issue_074_pacing_audit_2026_05_06.py).

    Drift of ≥1 mora flags the entry. BUG-004 caught 911 entries
    with stale pitch_marks; this invariant prevents re-introduction.
    """
    # Canonical morae-count algorithm (matches the round-9 baseline
    # used by ISSUE-074 and BUG-004's fix).
    SMALL_KANA = set("ゃゅょぁぃぅぇぉっゎ" + "ャュョァィゥェォッヮ")
    def count_morae(text: str) -> int:
        morae = 0
        for ch in text or "":
            cp = ord(ch)
            if 0x3041 <= cp <= 0x3096:
                if ch not in SMALL_KANA:
                    morae += 1
            elif 0x30A1 <= cp <= 0x30FA:
                if ch not in SMALL_KANA:
                    morae += 1
            elif ch == "ー":
                morae += 1
            elif 0xFF66 <= cp <= 0xFF9D:
                morae += 1
        return morae

    failures: list[str] = []
    try:
        v = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-93 could not read vocab.json: {e}"]
    for entry in v.get("entries") or []:
        eid = entry.get("id", entry.get("form", "?"))
        pitch = entry.get("pitch_marks")
        if not pitch:
            continue
        # pitch_marks may be a list of dicts {char, drop} or a list of
        # bracket-markers like ["L","H","H",...]. Either way, len()
        # gives the mora count.
        if not isinstance(pitch, list):
            continue
        n_pitch = len(pitch)
        # Compute expected mora from reading (preferred) or form
        reading = entry.get("reading") or entry.get("form") or ""
        if not isinstance(reading, str):
            continue
        n_expected = count_morae(reading)
        if n_pitch != n_expected:
            failures.append(
                f"JA-93 {eid}: pitch_marks has {n_pitch} entries but "
                f"count_morae({reading!r})={n_expected} (BUG-004 "
                f"algorithmic mismatch)"
            )
    return failures


def _check_ja_94_pattern_marker_per_example() -> list[str]:
    """BUG-006 pattern-marker guard (2026-05-17 final unblock).

    Every grammar example's `ja` field must contain ≥1 marker from
    its parent pattern's structural-markers list, as catalogued in
    `data/pattern_markers.json`. Markers are STRUCTURAL (particles,
    conjugational endings, characteristic tokens) rather than the
    pre-existing meaning-explanation markers — this is the
    invariant that BUG-006 deserved.

    Authoring pipeline:
      - `tools/author_pattern_markers_2026_05_17.py` derives markers
        from each pattern's `pattern` field + category-specific
        conjugational expansions + per-pattern OVERRIDES, reaching
        99.2% coverage of the 1782 grammar examples.
      - The remaining 14 examples (BUG-006-CANDIDATE class — wrong
        example doesn't demonstrate parent pattern) are snapshotted
        in `data/_ja94_baseline.json` as an allowlist, with a per-
        entry note explaining the classification + suggested fix.

    JA-94 PASSes against the current corpus. It trips on:
      (a) a new grammar example added that contains no marker for
          its parent pattern — usually a copy-paste contamination
          to fix in the example.
      (b) a markers-catalog edit that drops a marker still used by
          an example — fix by re-adding the marker.

    To extend the allowlist (new BUG-006-CANDIDATE found post-
    snapshot), add a `{pattern_id, ex_index, ja, note}` entry to
    the baseline; do NOT silently expand pattern_markers.json
    unless the form is genuinely canonical for the pattern.
    """
    failures: list[str] = []
    try:
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-94 could not read grammar.json: {e}"]
    markers_path = ROOT / "data" / "pattern_markers.json"
    if not markers_path.exists():
        return [
            f"JA-94 markers catalog missing: {markers_path}. Run "
            f"`python tools/author_pattern_markers_2026_05_17.py` "
            f"from the N5/ directory to (re-)author it."
        ]
    try:
        markers_doc = json.loads(markers_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-94 could not read pattern_markers.json: {e}"]
    markers_by_pid = markers_doc.get("patterns") or {}
    baseline_path = ROOT / "data" / "_ja94_baseline.json"
    if not baseline_path.exists():
        return [
            f"JA-94 baseline file missing: {baseline_path}. The "
            f"baseline carries the 14 currently-acceptable "
            f"BUG-006-CANDIDATE wrong-example failures."
        ]
    try:
        baseline_doc = json.loads(baseline_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-94 could not read baseline: {e}"]
    baseline_allowlist: set[tuple[str, int]] = set()
    for entry in baseline_doc.get("baseline_failing_examples") or []:
        pid = entry.get("pattern_id")
        idx = entry.get("ex_index")
        if isinstance(pid, str) and isinstance(idx, int):
            baseline_allowlist.add((pid, idx))
    for p in g.get("patterns") or []:
        pid = p.get("id")
        if not isinstance(pid, str):
            continue
        markers = markers_by_pid.get(pid) or []
        markers = [m for m in markers if isinstance(m, str) and m.strip()]
        if not markers:
            continue
        for i, ex in enumerate(p.get("examples") or []):
            if not isinstance(ex, dict):
                continue
            ja = ex.get("ja", "")
            if not isinstance(ja, str) or not ja:
                continue
            if (pid, i) in baseline_allowlist:
                continue
            if not any(m in ja for m in markers):
                failures.append(
                    f"JA-94 NEW pattern-instance contamination "
                    f"{pid} ex[{i}]: ja={ja[:60]!r} contains none "
                    f"of the parent pattern's structural markers "
                    f"(first 6: {markers[:6]!r}). Either (a) the "
                    f"example doesn't demonstrate the parent pattern "
                    f"— rewrite or move it; or (b) the form is "
                    f"genuinely canonical but not yet captured — "
                    f"expand the OVERRIDES table in "
                    f"tools/author_pattern_markers_2026_05_17.py "
                    f"and regenerate. To allowlist as a "
                    f"BUG-006-CANDIDATE, add an entry to "
                    f"data/_ja94_baseline.json (with a note)."
                )
    return failures


def _check_ja_91_explanation_similarity() -> list[str]:
    """BUG-003 cross-pattern contamination guard (2026-05-17 final unblock).

    For every pair of grammar patterns (i, j), compute SequenceMatcher
    ratio() on their explanation_en. Pairs with ratio ≥ 0.85 are
    candidates for contamination.

    The 2026-05-17 audit classified the corpus's 43 currently-occurring
    pairs into:
      DUPLICATE_PATTERN ×8 (same concept, two IDs)
      CROSS_REFERENCE ×21 (one pattern defers to another)
      ALTERNATIVE_VARIANT ×12 (register / dialect / syntactic variants)
      SUBSET ×2 (one pattern subset of another's coverage)

    All 43 are baselined in data/_ja91_baseline.json. JA-91 PASSes
    on the current corpus by allowlisting them. It trips on any
    NEW pair that wasn't in the baseline — typically signaling a
    fresh pattern was added with explanation_en copied/contaminated
    from an existing pattern.

    To add a new pattern with a known-intentional similarity to an
    existing one, add the (pattern_a, pattern_b) tuple to the
    baseline AND classify it (see _meta.classification_legend in
    the baseline file).
    """
    import difflib
    failures: list[str] = []
    try:
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-91 could not read grammar.json: {e}"]
    # Load baseline allowlist
    baseline_path = ROOT / "data" / "_ja91_baseline.json"
    if not baseline_path.exists():
        return [
            f"JA-91 baseline file missing: {baseline_path}. Either "
            f"author it (see _meta in similar baseline files) or "
            f"disable JA-91 registry entry. The baseline carries the "
            f"43 currently-acceptable similarity pairs and their "
            f"classification."
        ]
    try:
        baseline_data = json.loads(baseline_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-91 could not read baseline: {e}"]
    # Normalize each pair to a frozenset of (a, b) for order-independent
    # membership testing
    baseline_pairs = set()
    for entry in baseline_data.get("baseline_pairs") or []:
        pair = entry.get("pair") or []
        if len(pair) == 2:
            baseline_pairs.add(frozenset(pair))
    patterns = [(p.get("id"), p.get("explanation_en", ""))
                for p in g.get("patterns") or []
                if isinstance(p.get("explanation_en"), str) and p.get("explanation_en", "").strip()]
    THRESHOLD = 0.85
    for i in range(len(patterns)):
        id_i, txt_i = patterns[i]
        for j in range(i + 1, len(patterns)):
            id_j, txt_j = patterns[j]
            sm = difflib.SequenceMatcher(None, txt_i, txt_j, autojunk=False)
            if sm.quick_ratio() < THRESHOLD:
                continue
            if sm.real_quick_ratio() < THRESHOLD:
                continue
            r = sm.ratio()
            if r < THRESHOLD:
                continue
            # Check baseline
            if frozenset([id_i, id_j]) in baseline_pairs:
                continue
            failures.append(
                f"JA-91 NEW cross-pattern explanation_en similarity "
                f"{id_i} vs {id_j}: ratio={r:.3f} ≥ {THRESHOLD}. "
                f"Either (a) the new pattern is a legitimate "
                f"cross-ref / variant of an existing one — add the "
                f"pair to data/_ja91_baseline.json with classification "
                f"(DUPLICATE_PATTERN / CROSS_REFERENCE / "
                f"ALTERNATIVE_VARIANT / SUBSET), or (b) the new "
                f"pattern's explanation accidentally duplicates an "
                f"existing one — rewrite the explanation."
            )
    return failures


def _check_ja_119_spec_version_sample() -> list[str]:
    """BUG-050 round-3 close-out (2026-05-17).

    The implementation spec §7.3 ('version.json - build stamp')
    carries a sample JSON block showing the file's shape with
    illustrative count values. The sample must match the live
    data/version.json count fields exactly — stale sample values
    have been observed to confuse maintainers + auditors who read
    the spec and mistake the example for the actual current state
    (BUG-050 was filed 3 times this session against the actual file
    even though the file was clean; the user was reading the spec
    sample which carried stale v1.12.50-era values).

    Implementation: locate the §7.3 fenced JSON block, parse its
    `counts` object, and compare each key to data/version.json.

    Same Cross-Artifact Sync Protocol INV-4 class as JA-47
    (CONTENT-LICENSE.md), JA-107 (version.json itself), JA-112
    (AUDIO.md), JA-115 (README.md). The fifth user-facing
    prose-with-counts surface, now locked.
    """
    failures: list[str] = []
    SPEC = ROOT / "specifications" / "JLPT-N5-Current-Implementation-Spec.md"
    VERSION = ROOT / "data" / "version.json"
    if not SPEC.exists():
        return [f"JA-119 spec file missing: {SPEC}"]
    if not VERSION.exists():
        return [f"JA-119 version.json missing: {VERSION}"]
    try:
        spec_text = SPEC.read_text(encoding="utf-8")
        version_data = json.loads(VERSION.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-119 read failed: {e}"]

    # Locate §7.3 — find the heading then take the next fenced ```json block
    h_match = re.search(r"^###\s+7\.3\s+`version\.json`", spec_text, re.M)
    if not h_match:
        return [
            "JA-119 spec §7.3 'version.json - build stamp' heading not "
            "found. If the section was renumbered, update JA-119's "
            "anchor regex."
        ]
    after_heading = spec_text[h_match.end():]
    fence_match = re.search(r"```json\s*\n(.+?)```", after_heading, re.S)
    if not fence_match:
        return [
            "JA-119 spec §7.3 has no `json fenced code block in the "
            "next ~2000 chars after the heading. Update JA-119 if the "
            "sample was moved to a different code-fence language."
        ]
    sample_text = fence_match.group(1)
    # Parse the sample as JSON (strip JSON5-ish indent quirks if any)
    try:
        sample = json.loads(sample_text)
    except json.JSONDecodeError as e:
        return [
            f"JA-119 spec §7.3 sample is not valid JSON: {e}. "
            f"Sample text:\n{sample_text[:300]}"
        ]
    sample_counts = sample.get("counts") or {}
    live_counts = version_data.get("counts") or {}

    # Every count key in the sample must match the live value exactly.
    # New live counts that aren't yet in the sample → spec is stale.
    for key, live_v in live_counts.items():
        sample_v = sample_counts.get(key)
        if sample_v is None:
            failures.append(
                f"JA-119 spec §7.3 sample missing 'counts.{key}' "
                f"(live = {live_v}). Update the sample."
            )
            continue
        if sample_v != live_v:
            failures.append(
                f"JA-119 spec §7.3 sample 'counts.{key}={sample_v}' "
                f"does not match live 'counts.{key}={live_v}'. "
                f"Update the sample to current values."
            )
    # Extra keys in sample that aren't live → also drift
    for key in sample_counts:
        if key not in live_counts:
            failures.append(
                f"JA-119 spec §7.3 sample has 'counts.{key}' which is "
                f"NOT in live version.json. Remove from sample or add "
                f"to live counts."
            )
    return failures


def _check_ja_95_particle_pattern_alignment() -> list[str]:
    """BUG-009 particle-pattern alignment guard (2026-05-17 wire-up).

    For every grammar pattern in category 'Particles' whose `pattern`
    field is a single-character particle (は, が, を, に, へ, で, と,
    の, も, から, まで, ね, よ, か), every example's `ja` field must
    contain that particle.

    BUG-009 caught n5-003 (が particle) whose example used は. This
    invariant prevents re-introduction.

    Note: only enforced on pattern-field that's CLEARLY a single
    particle (≤ 2 chars). More complex pattern-fields (e.g.,
    "〜と〜どちら" for comparison patterns) skip this check.
    """
    failures: list[str] = []
    try:
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-95 could not read grammar.json: {e}"]
    # Particles where the pattern field IS the particle (and ≤ 2 chars)
    for p in g.get("patterns") or []:
        pid = p.get("id", "?")
        cat = p.get("category", "")
        ptn = p.get("pattern", "")
        if cat != "Particles" or not isinstance(ptn, str):
            continue
        # Allow "は", "が" etc. plus 2-char ones like "から", "まで"
        candidate = ptn.replace("〜", "").replace("～", "").strip()
        if len(candidate) > 2 or len(candidate) < 1:
            continue
        # Exclude particles that double as common content words
        # (rare at N5; but defensive)
        for i, ex in enumerate(p.get("examples") or []):
            if not isinstance(ex, dict):
                continue
            ja = ex.get("ja", "")
            if not isinstance(ja, str) or not ja:
                continue
            if candidate not in ja:
                failures.append(
                    f"JA-95 {pid} (particle={candidate!r}) ex[{i}]: "
                    f"ja={ja[:50]!r} does not contain the pattern's "
                    f"named particle (BUG-009 alignment violation)"
                )
    return failures


def _check_ja_118_bug_fix_commit_link() -> list[str]:
    """Cross-Artifact Sync Protocol INV-9 promotion to Wired (2026-05-17).

    Every Fixed (or Closed) row in the User Reported Bugs sheet must
    have a non-empty Fix Commit cell. Acceptable values:
      - A git commit SHA (short or long form)
      - The explicit annotation "(unattributable)" — used when no
        commit references the BUG-NNN ID (e.g., bug was filed and
        closed in the same trackerless session before commits)
      - Any non-empty string containing "commit", "SHA", or a
        7+-char hex token (defensive — allows informal references)

    Re-populate via `python tools/populate_bug_fix_commits_2026_05_17.py`
    which scans git log for commit subjects mentioning each BUG-NNN
    (including range patterns like "BUG-041 through BUG-046" or
    "BUG-041..046").

    Skips gracefully if openpyxl is unavailable (CI environment
    without openpyxl) — a missing dep is not drift.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return []
    import openpyxl
    failures: list[str] = []
    XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
    if not XLSX.exists():
        return [f"JA-118 xlsx missing: {XLSX}"]
    try:
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    except Exception as e:
        return [f"JA-118 could not open xlsx: {e}"]
    if "User Reported Bugs" not in wb.sheetnames:
        return [f"JA-118 'User Reported Bugs' sheet missing"]
    sh = wb["User Reported Bugs"]

    # Locate columns by header row 3
    headers = {}
    for c in range(1, sh.max_column + 1):
        v = sh.cell(row=3, column=c).value
        if isinstance(v, str):
            headers[v.strip()] = c
    status_col = headers.get("Status")
    fix_commit_col = headers.get("Fix Commit")
    if status_col is None:
        return [f"JA-118 'Status' header not found in xlsx"]
    if fix_commit_col is None:
        return [f"JA-118 'Fix Commit' header not found in xlsx — run "
                f"tools/populate_bug_fix_commits_2026_05_17.py to add it"]

    for r in range(4, sh.max_row + 1):
        bid = sh.cell(row=r, column=1).value
        if not bid:
            continue
        status = sh.cell(row=r, column=status_col).value
        if status not in ("Fixed", "Closed"):
            continue
        commit_cell = sh.cell(row=r, column=fix_commit_col).value
        if not commit_cell or (isinstance(commit_cell, str) and not commit_cell.strip()):
            # Resolve bug ID
            if isinstance(bid, str) and bid.startswith("="):
                bid_resolved = f"BUG-{r-3:03d}"
            else:
                bid_resolved = str(bid)
            failures.append(
                f"JA-118 {bid_resolved}: Status={status!r} but Fix Commit "
                f"is empty. Run `python tools/populate_bug_fix_commits_"
                f"2026_05_17.py` to back-fill, or set explicitly."
            )
    return failures


def _check_ja_117_cross_corpus_id_refs() -> list[str]:
    """Cross-Artifact Sync Protocol INV-7 extension (2026-05-17).

    Two cross-corpus reference classes were previously relying on
    manual checks:

      - kanji.json `.entries[*].reading_passages[*].passage_id` →
        reading.json passage IDs (363 refs)
      - reading.json `.passages[*].grammar_footnotes[*].pattern_id`
        + nested `.patterns[*].pattern_id` → grammar.json pattern IDs
        (319 refs)

    JA-117 verifies every such reference resolves to an actual entry
    in the target corpus. Companion to JA-15 (audio refs), JA-17
    (vocab_id in grammar examples), JA-82 (_meta.see_also /
    consumers), JA-100 (kanji↔vocab form), JA-105 (vocab_preview
    refs), JA-113 (mirror freshness vs source) — six wired
    invariants now under INV-7.

    INV-7 promoted to Wired (fully covered for the canonical
    cross-corpus reference fields).
    """
    failures: list[str] = []
    try:
        k_data = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
        r_data = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
        g_data = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-117 could not read corpus files: {e}"]

    reading_ids = {p.get("id") for p in r_data.get("passages") or [] if p.get("id")}
    pattern_ids = {p.get("id") for p in g_data.get("patterns") or [] if p.get("id")}

    # Pass 1: kanji.json passage_id → reading.json
    for entry in k_data.get("entries") or []:
        glyph = entry.get("glyph", "?")
        for rp in entry.get("reading_passages") or []:
            if not isinstance(rp, dict):
                continue
            pid = rp.get("passage_id")
            if pid is None:
                continue
            if pid not in reading_ids:
                failures.append(
                    f"JA-117 kanji.json entry '{glyph}' references "
                    f"reading passage_id {pid!r} which does not exist "
                    f"in reading.json"
                )

    # Pass 2: reading.json pattern_id → grammar.json
    for passage in r_data.get("passages") or []:
        rid = passage.get("id", "?")
        for fn in passage.get("grammar_footnotes") or []:
            if not isinstance(fn, dict):
                continue
            # Direct .pattern_id
            pid = fn.get("pattern_id")
            if pid is not None and pid not in pattern_ids:
                failures.append(
                    f"JA-117 reading.json passage '{rid}' grammar_footnote "
                    f"references pattern_id {pid!r} which does not exist "
                    f"in grammar.json"
                )
            # Nested .patterns[*].pattern_id
            for sub in fn.get("patterns") or []:
                if not isinstance(sub, dict):
                    continue
                pid2 = sub.get("pattern_id")
                if pid2 is not None and pid2 not in pattern_ids:
                    failures.append(
                        f"JA-117 reading.json passage '{rid}' grammar_"
                        f"footnote.patterns[] references pattern_id "
                        f"{pid2!r} which does not exist in grammar.json"
                    )
    return failures


def _check_ja_116_prompt_xlsx_coverage() -> list[str]:
    """Cross-Artifact Sync Protocol INV-6 promotion to Wired (2026-05-17).

    Every structured item in N5/prompts/ — A-NN audit category in the
    accuracy prompt, Phase-0 regression block in N5Improvement, FP-NN
    false-positive class in the accuracy prompt — must be named in
    ≥1 scenario row of the test-scenarios xlsx (anywhere in any tab's
    Scenario / Notes / Tools columns; case-insensitive substring).

    The b466293 sync (2026-05-17) added 134 scenarios to cover every
    such item; this invariant prevents the coverage from drifting
    when prompts are extended in future audit cycles. New A-NN or
    new Phase-0 block must come with a scenario row added in the
    same commit.

    Tries openpyxl import; if unavailable (CI environment without
    openpyxl) the check skips gracefully — a missing dep is not
    drift.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return []

    import openpyxl
    failures: list[str] = []

    # Extract the structured items from the prompts
    accuracy = (ROOT / "prompts" / "Japanese language Accuracy check.txt")
    improvement = (ROOT / "prompts" / "N5Improvement.txt")
    if not accuracy.exists() or not improvement.exists():
        return [f"JA-116 prompts directory missing one of the required files"]

    try:
        acc_text = accuracy.read_text(encoding="utf-8")
        imp_text = improvement.read_text(encoding="utf-8")
    except Exception as e:
        return [f"JA-116 could not read prompts: {e}"]

    # A-NN audit categories (e.g., "A1. Incorrect grammar")
    a_codes = set(re.findall(r"^(A\d+)\.", acc_text, re.M))
    # FP-NN false-positive classes (e.g., "FP-1. ウ音便 ...")
    fp_codes = set(re.findall(r"^(FP-\d+)\.", acc_text, re.M))
    # Phase-0 blocks (e.g., "### Phase-0 Grammar (run mechanically ...)")
    # Tokenize by the FIRST 30 chars of the block heading after the
    # "Phase-0 " token — that's the distinctive identifier.
    phase0_titles = re.findall(r"^### (Phase-0[^\n]+)$", imp_text, re.M)
    # Extract a stable identifier per block: the words after "Phase-0 "
    # up to the first " (" or end-of-line.
    phase0_ids = set()
    for title in phase0_titles:
        m = re.match(r"Phase-0\s+([^\(]+?)(?:\s*\(|$)", title)
        if m:
            phase0_ids.add(m.group(1).strip())

    # Now load the xlsx and collect every text-cell across all
    # specialist tabs (limited to non-header rows).
    XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
    if not XLSX.exists():
        return [f"JA-116 xlsx missing: {XLSX}"]
    try:
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    except Exception as e:
        return [f"JA-116 could not open xlsx: {e}"]

    SPECIALIST_TABS = [
        "A. Japanese language", "B. JLPT format", "C. Hindi locale",
        "D. UX design", "E. Accessibility", "F. Security",
        "G. Privacy and legal", "H. Performance", "I. Data engineering",
        "J. Pedagogy", "K. QA testing", "L. Cultural ethical",
        "M. Operations", "N. End-user POV",
    ]
    all_text = []
    for tab_name in SPECIALIST_TABS:
        if tab_name not in wb.sheetnames:
            continue
        sh = wb[tab_name]
        for row in sh.iter_rows(min_row=5, values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    all_text.append(cell)
    combined = "\n".join(all_text)

    # Check each code / Phase-0 id appears in the combined xlsx text.
    # A-NN match: use word-boundary regex to avoid false negatives
    # (e.g., "A1" being matched inside "A100").
    missing_a = []
    for code in sorted(a_codes, key=lambda x: int(x[1:])):
        # Match the literal "A<NN>" as a word (not as part of A<NNN>+)
        pat = re.compile(rf"\b{code}\b")
        if not pat.search(combined):
            missing_a.append(code)
    missing_fp = []
    for code in sorted(fp_codes, key=lambda x: int(x[3:])):
        pat = re.compile(rf"\b{code}\b")
        if not pat.search(combined):
            missing_fp.append(code)
    missing_p0 = []
    for pid in sorted(phase0_ids):
        # Match the Phase-0 identifier (case-insensitive substring;
        # the xlsx scenarios use the full block-title phrase)
        if pid.lower() not in combined.lower():
            missing_p0.append(pid)

    for code in missing_a:
        failures.append(
            f"JA-116 accuracy prompt audit category {code} has no "
            f"matching xlsx scenario row (run tools/sync_test_scenarios_"
            f"with_prompts_feedback_2026_05_17.py to add coverage)"
        )
    for code in missing_fp:
        failures.append(
            f"JA-116 accuracy prompt false-positive class {code} has no "
            f"matching xlsx scenario row"
        )
    for pid in missing_p0:
        failures.append(
            f"JA-116 N5Improvement Phase-0 block 'Phase-0 {pid}' has no "
            f"matching xlsx scenario row"
        )
    return failures


def _check_ja_115_readme_counts() -> list[str]:
    """Cross-Artifact Sync Protocol INV-4 enforcement, README surface
    (2026-05-17).

    README.md's "Content" section carries user-facing count claims:
        "N grammar patterns ... M vocabulary entries ... K N5 kanji ...
         R reading passages ... L listening items ... Q mock-test
         questions ... 28 audited papers ... totalling P questions"

    Each numeric claim must match the live data. Companion to JA-47
    (CONTENT-LICENSE.md), JA-107 (version.json), JA-112 (AUDIO.md).
    Coverage of the INV-4 class extends to the four primary
    user-facing surfaces where corpus counts surface in prose.

    Anchored on the canonical phrasing in the README's "Content" line.
    If the line is rephrased, the regex must be updated in lockstep
    (matches JA-112's contract).
    """
    failures: list[str] = []
    try:
        text = (ROOT / "README.md").read_text(encoding="utf-8")
    except Exception as e:
        return [f"JA-115 could not read README.md: {e}"]

    # Live counts
    try:
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
        v = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
        k = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
        r = json.loads((ROOT / "data" / "reading.json").read_text(encoding="utf-8"))
        l = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
        q = json.loads((ROOT / "data" / "questions.json").read_text(encoding="utf-8"))
        papers_m = json.loads(
            (ROOT / "data" / "papers" / "manifest.json").read_text(encoding="utf-8")
        )
    except Exception as e:
        return [f"JA-115 could not read one or more data files: {e}"]

    live = {
        "grammar": len(g.get("patterns") or []),
        "vocab": len(v.get("entries") or []),
        "kanji": len(k.get("entries") or []),
        "reading": len(r.get("passages") or []),
        "listening": len(l.get("items") or []),
        "questions": len(q.get("questions") or []),
        "papers": papers_m.get("totalPapers"),
        "paper_questions": papers_m.get("totalQuestions"),
    }

    # Anchored patterns. Each tuple = (regex, live_key, label).
    PATTERNS = [
        (r"(\d+)\s+grammar patterns", "grammar", "grammar patterns"),
        (r"(\d+)\s+vocabulary entries", "vocab", "vocabulary entries"),
        (r"(\d+)\s+N5 kanji", "kanji", "N5 kanji"),
        (r"(\d+)\s+reading passages", "reading", "reading passages"),
        (r"(\d+)\s+listening items", "listening", "listening items"),
        (r"(\d+)\s+mock-test questions", "questions", "mock-test questions"),
        (r"(\d+)\s+audited papers", "papers", "audited papers"),
        (r"totalling\s+(\d+)\s+questions", "paper_questions", "paper questions"),
    ]

    # Search only the "Content (current as of ..." block. Limit scope to
    # ~2000 chars after "Content (current as of" to avoid matching
    # historical CHANGELOG / TASKS-style passages.
    anchor = re.search(r"Content\s*\(current as of[^)]+\)", text)
    if not anchor:
        return [
            "JA-115 README.md no longer contains the 'Content (current as of "
            "...)' anchor. If the section was intentionally rephrased, update "
            "JA-115's anchor regex."
        ]
    scope_start = anchor.start()
    scope_end = min(len(text), scope_start + 2000)
    scope = text[scope_start:scope_end]

    for pat, key, label in PATTERNS:
        m = re.search(pat, scope, re.IGNORECASE)
        if not m:
            failures.append(
                f"JA-115 README.md 'Content' section missing the "
                f"'{label}' count claim (pattern: {pat!r}). If the prose "
                f"was rephrased, update JA-115's regex."
            )
            continue
        n_claimed = int(m.group(1))
        n_live = live.get(key)
        if n_live is None:
            failures.append(
                f"JA-115 README.md '{label}' lookup failed for key {key!r}"
            )
            continue
        if n_claimed != n_live:
            failures.append(
                f"JA-115 README.md claims {n_claimed} {label} but live "
                f"count is {n_live} (drift)"
            )
    return failures


def _check_ja_114_pacing_status_enum() -> list[str]:
    """BUG-048/049 follow-up (2026-05-17) regression guard.

    Lock listening.json `pacing_status` to a closed enum:
        {in_range, too_slow, too_fast, no_audio, unmeasured}

    After the BUG-048/049 close-out (commit 47d1edc) every item has
    a measured pacing_morae_per_min + a status field reflecting its
    position in the JLPT N5 target band 180-240 mpm. This invariant
    prevents the field from drifting back to null (the BUG-048
    starting state) or accumulating new ad-hoc values from future
    pipeline changes.

    Same class as JA-106 / JA-111 (closed-enum on a corpus field
    where the value-domain is small and stable).
    """
    failures: list[str] = []
    ALLOWED = {"in_range", "too_slow", "too_fast", "no_audio", "unmeasured"}
    try:
        L = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-114 could not read listening.json: {e}"]
    for it in L.get("items", []):
        iid = it.get("id", "?")
        if "pacing_status" not in it:
            failures.append(f"JA-114 {iid}: missing required `pacing_status` field")
            continue
        v = it.get("pacing_status")
        if v not in ALLOWED:
            failures.append(
                f"JA-114 {iid}: pacing_status {v!r} not in "
                f"{sorted(ALLOWED)}"
            )
    return failures


def _check_ja_113_meta_mirror_freshness() -> list[str]:
    """Meta-route static-mirror freshness guard (2026-05-17).

    For each markdown-sourced meta route, verify the latest H1/H2
    heading text from the source markdown appears in the static-mirror
    HTML. Catches the drift class where the source markdown was
    updated but `tools/build_static_mirrors.py` wasn't re-run.

    Routes checked (markdown-sourced; the 6 stub-body routes don't
    apply since they have no source-of-truth markdown):

      - home      → README.md
      - changelog → CHANGELOG.md
      - privacy   → PRIVACY.md
      - notices   → NOTICES.md

    Detection logic: extract the FIRST H1 / H2 from the markdown
    (which for CHANGELOG-style time-ordered docs is the latest entry;
    for static reference docs is the canonical top header).
    HTML-escape it and check whether it appears in the mirror.

    This is a fast-and-lightweight check, not a full HTML-vs-HTML
    diff — false negatives possible (e.g., heading edited in-place
    without text change), but the common case (new entry added or
    heading text changed) is caught.

    Rule-5 INV-7 class extension: cross-file references resolve →
    derived-mirror reflects source.
    """
    failures: list[str] = []
    routes = [
        ("home", "README.md"),
        ("changelog", "CHANGELOG.md"),
        ("privacy", "PRIVACY.md"),
        ("notices", "NOTICES.md"),
    ]
    import html as _html
    for slug, source_md in routes:
        src_path = ROOT / source_md
        mirror_path = ROOT / slug / "index.html"
        if not src_path.exists():
            failures.append(f"JA-113 source markdown missing: {source_md}")
            continue
        if not mirror_path.exists():
            failures.append(f"JA-113 static mirror missing: {slug}/index.html")
            continue
        try:
            src_text = src_path.read_text(encoding="utf-8")
            mirror_text = mirror_path.read_text(encoding="utf-8")
        except Exception as e:
            failures.append(f"JA-113 could not read {slug}: {e}")
            continue
        # Extract the FIRST H2 (## …) from the source markdown.
        # CHANGELOG-style time-ordered: most recent entry is first.
        # If no H2, fall back to first H1.
        h_match = re.search(r"^##\s+(.+?)\s*$", src_text, re.M)
        if not h_match:
            h_match = re.search(r"^#\s+(.+?)\s*$", src_text, re.M)
        if not h_match:
            continue  # Markdown has no headings; skip
        heading_text = h_match.group(1).strip()
        # Try direct match, then HTML-escaped, then with inline-code
        # backticks stripped (the builder renders `foo` as <code>foo</code>).
        candidates = [
            heading_text,
            _html.escape(heading_text),
            re.sub(r"`([^`]+)`", r"\1", heading_text),
            _html.escape(re.sub(r"`([^`]+)`", r"\1", heading_text)),
        ]
        if any(c in mirror_text for c in candidates):
            continue
        failures.append(
            f"JA-113 {slug}/index.html does not contain the latest "
            f"heading from {source_md}: {heading_text!r}. "
            f"Run `python tools/build_static_mirrors.py --stages meta` "
            f"to regenerate the mirror."
        )
    return failures


def _check_ja_112_audio_md_listening_counts() -> list[str]:
    """BUG-050 (charitable interpretation, 2026-05-17) regression guard.

    AUDIO.md carries a user-facing claim "N listening items use M
    distinct VOICEVOX speakers in rotation". Both N and M must match
    the live data:
      - N == len(listening.json.items)
      - M == |distinct values in union of audio_render_meta.voices_used
              across all items|

    BUG-050 was filed as "version.json declares counts.listening=47"
    but every observable instance of version.json (working tree, git
    history HEAD..HEAD~10, live deployed site) has shown 50. JA-107
    locks the version.json side. The "47" the bug report observed was
    AUDIO.md line 52's stale prose claim — this invariant catches that
    class going forward.

    Same INV-4 (cross-artifact sync protocol) drift class as JA-47
    (CONTENT-LICENSE.md counts) and JA-107 (version.json counts);
    extended to the AUDIO.md user-facing doc surface.
    """
    failures: list[str] = []
    try:
        text = (ROOT / "AUDIO.md").read_text(encoding="utf-8")
    except Exception as e:
        return [f"JA-112 could not read AUDIO.md: {e}"]
    try:
        L = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-112 could not read listening.json: {e}"]

    live_item_count = len(L.get("items", []) or [])
    speakers: set[str] = set()
    for it in L.get("items", []) or []:
        for v in (it.get("audio_render_meta") or {}).get("voices_used") or []:
            speakers.add(v)
    live_speaker_count = len(speakers)

    # Match the canonical claim shape. The regex is anchored on the
    # specific prose pattern AUDIO.md uses; it skips historical
    # references inside fenced code blocks or commits/changelog entries.
    # Line 52 shape: "N listening items use M distinct VOICEVOX speakers"
    pattern = re.compile(
        r"(\d+)\s+listening items use\s+(\d+)\s+distinct VOICEVOX speakers",
        re.IGNORECASE,
    )
    matches = pattern.findall(text)
    if not matches:
        # If the canonical claim was rephrased, that's information
        # drift too — surface it.
        failures.append(
            "JA-112 AUDIO.md no longer contains the canonical "
            "'N listening items use M distinct VOICEVOX speakers' claim. "
            "If the claim was intentionally rephrased, update JA-112's "
            "regex to match the new shape."
        )
        return failures

    for n_str, m_str in matches:
        n = int(n_str)
        m = int(m_str)
        if n != live_item_count:
            failures.append(
                f"JA-112 AUDIO.md claims {n} listening items but "
                f"live count is {live_item_count} (drift)"
            )
        if m != live_speaker_count:
            failures.append(
                f"JA-112 AUDIO.md claims {m} distinct VOICEVOX "
                f"speakers but live count is {live_speaker_count} (drift)"
            )
    return failures


def _check_ja_110_no_voice_planned() -> list[str]:
    """BUG-047 (2026-05-17) regression guard.

    listening.json items must NOT carry the legacy `voice_planned`
    field. The audio_render_meta block is the canonical source for
    engine + speaker info post the 2026-05-12 VOICEVOX migration.
    Pre-fix every item declared voice_planned.engine="edge-tts" even
    though audio_render_meta.voice_provider="voicevox" — exactly the
    dual-field drift class that BUG-044 / BUG-051 also instantiate.
    """
    failures: list[str] = []
    try:
        l = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-110 could not read listening.json: {e}"]
    for it in l.get("items", []):
        iid = it.get("id", "?")
        if "voice_planned" in it:
            failures.append(
                f"JA-110 {iid}: legacy `voice_planned` field still present "
                f"(deprecated per BUG-047; read from audio_render_meta.voice_"
                f"planned_for_engine + audio_render_meta.voice_provider instead)"
            )
    return failures


def _check_ja_111_format_type_only() -> list[str]:
    """BUG-051 (2026-05-17) regression guard.

    listening.json items must NOT carry the legacy `format` field;
    `format_type` is the canonical descriptive field. Closed enum:
    {task_understanding, point_understanding, utterance_expression,
    immediate_response}.

    Same dual-field-redundancy class as BUG-044 (reading.json
    format_type vs format_role) and BUG-047 (listening.json
    voice_planned vs audio_render_meta).
    """
    failures: list[str] = []
    ALLOWED = {
        "task_understanding",
        "point_understanding",
        "utterance_expression",
        "immediate_response",
    }
    try:
        l = json.loads((ROOT / "data" / "listening.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-111 could not read listening.json: {e}"]
    for it in l.get("items", []):
        iid = it.get("id", "?")
        if "format" in it:
            failures.append(
                f"JA-111 {iid}: legacy `format` field still present "
                f"(deprecated per BUG-051; use `format_type` only)"
            )
        ft = it.get("format_type")
        if ft is None:
            failures.append(
                f"JA-111 {iid}: missing required `format_type` field"
            )
        elif ft not in ALLOWED:
            failures.append(
                f"JA-111 {iid}: format_type {ft!r} not in {sorted(ALLOWED)}"
            )
    return failures


def _check_ja_107_version_json_counts() -> list[str]:
    """Cross-Artifact Sync Protocol INV-4 enforcement (2026-05-17).

    Every count declared in `data/version.json.counts` must equal the
    actual array length in the corresponding live corpus file.
    Catches the failure class where dedup / migration passes reduce
    a corpus but version.json (read by app footer + sw.js
    CACHE_VERSION derivation) stays at the old number.

    Companion to JA-47 (CONTENT-LICENSE.md counts ↔ live data).
    JA-47 enforces the legal-attribution side; JA-107 enforces the
    public build-stamp side.
    """
    failures: list[str] = []
    try:
        version = json.loads((ROOT / "data" / "version.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-107 could not read version.json: {e}"]
    counts = version.get("counts") or {}
    if not isinstance(counts, dict):
        return [f"JA-107 version.json.counts is not a dict: got {type(counts).__name__}"]

    # Map declared keys to (corpus filename, container key).
    SOURCES = {
        "grammar": ("grammar.json", "patterns"),
        "vocab": ("vocab.json", "entries"),
        "kanji": ("kanji.json", "entries"),
        "reading": ("reading.json", "passages"),
        "listening": ("listening.json", "items"),
        "questions": ("questions.json", "questions"),
    }
    for key, declared in counts.items():
        if key in SOURCES:
            fname, container = SOURCES[key]
            try:
                data = json.loads((ROOT / "data" / fname).read_text(encoding="utf-8"))
            except Exception as e:
                failures.append(f"JA-107 could not read data/{fname} for {key} count: {e}")
                continue
            actual = len(data.get(container) or [])
            if declared != actual:
                failures.append(
                    f"JA-107 version.json.counts.{key}={declared} but "
                    f"data/{fname}.{container} has {actual} entries (drift; "
                    f"INV-4 violation)"
                )
        elif key in {"papers", "paperQuestions"}:
            # Both come from data/papers/manifest.json's totalPapers/totalQuestions.
            try:
                manifest = json.loads(
                    (ROOT / "data" / "papers" / "manifest.json").read_text(encoding="utf-8")
                )
            except Exception as e:
                failures.append(f"JA-107 could not read data/papers/manifest.json for {key}: {e}")
                continue
            manifest_key = "totalPapers" if key == "papers" else "totalQuestions"
            actual = manifest.get(manifest_key)
            if declared != actual:
                failures.append(
                    f"JA-107 version.json.counts.{key}={declared} but "
                    f"data/papers/manifest.json.{manifest_key}={actual} (drift; "
                    f"INV-4 violation)"
                )
        else:
            # Unknown count key — surface as a warning so we maintain the
            # explicit-source-mapping discipline.
            failures.append(
                f"JA-107 version.json.counts.{key} has no source mapping in JA-107; "
                f"add the corpus file to SOURCES dict in _check_ja_107_*"
            )
    return failures


def _check_ja_108_locale_key_parity() -> list[str]:
    """Cross-Artifact Sync Protocol INV-5 enforcement (2026-05-17).

    Every UI string key in `locales/en.json` must exist in every
    other locale file (`locales/hi.json` currently) and vice-versa.
    Catches the failure class where a new UI surface lands with EN
    copy only, and the HI translator pass is "queued" then forgotten.

    Strictness: full key-set equality. The `_meta` key is included
    (per the 2026-05-17 install decision to mirror it to en.json
    rather than exempt it) so locale-internal metadata stays
    parity-balanced too.
    """
    failures: list[str] = []
    locales_dir = ROOT / "locales"
    if not locales_dir.exists():
        return [f"JA-108 locales/ directory missing at {locales_dir}"]

    locale_files = sorted(locales_dir.glob("*.json"))
    if len(locale_files) < 2:
        return []  # Single-locale project; parity is trivially satisfied.

    def _flatten(d: dict, prefix: str = "") -> set[str]:
        keys: set[str] = set()
        for k, v in d.items():
            kp = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                keys |= _flatten(v, kp)
            else:
                keys.add(kp)
        return keys

    loaded: dict[str, set[str]] = {}
    for lf in locale_files:
        try:
            d = json.loads(lf.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-108 could not read {lf.name}: {e}")
            continue
        loaded[lf.stem] = _flatten(d)

    if len(loaded) < 2:
        return failures

    # Compare every pair.
    reference_name, reference_keys = next(iter(loaded.items()))
    for other_name, other_keys in list(loaded.items())[1:]:
        missing_in_other = reference_keys - other_keys
        missing_in_ref = other_keys - reference_keys
        for k in sorted(missing_in_other):
            failures.append(
                f"JA-108 key {k!r} present in {reference_name}.json but missing from "
                f"{other_name}.json (INV-5 violation)"
            )
        for k in sorted(missing_in_ref):
            failures.append(
                f"JA-108 key {k!r} present in {other_name}.json but missing from "
                f"{reference_name}.json (INV-5 violation)"
            )
    return failures


def _check_ja_109_script_references_resolve() -> list[str]:
    """Cross-Artifact Sync Protocol INV-10 enforcement (2026-05-17).

    Every reference to a `tools/<name>.py` script in the **N5-specific
    governance docs** — accuracy prompt, N5Improvement prompt, and
    audit-coverage doc — must resolve to an actual file on disk.
    Catches the failure class where a fix script is renamed / removed
    but the N5 docs still tell future operators to run it.

    Scope decisions for v1 (2026-05-17):

    * **Excluded:** `JLPT Common/procedure-manual-build-next-jlpt-level.md`.
      That doc lives in its own submodule and intentionally references
      abstract Nx-builder script names (e.g., `tools/build_data.py`,
      `tools/heuristic_audit.py`) that the cross-level methodology
      prescribes a future Nx builder *port and adapt* — not files that
      must exist in N5/tools/. The procedure manual has its own
      doc-integrity discipline; tracking the abstract-vs-concrete
      reference policy belongs there, not in N5's CI.

    * **Included:** every `tools/...py` reference in the three N5
      docs above. These reference real N5 fix/audit scripts and
      must stay aligned with what actually exists.

    Other path references (audio files, data files, schemas) are out
    of scope — those are caught by JA-15 / JA-82 / other invariants.
    """
    failures: list[str] = []
    DOCS = [
        ROOT / "prompts" / "Japanese language Accuracy check.txt",
        ROOT / "prompts" / "N5Improvement.txt",
    ]
    # Also scan the most recent AUDIT-COVERAGE-*.md (open-ended date suffix).
    docs_dir = ROOT / "docs"
    if docs_dir.exists():
        for f in sorted(docs_dir.glob("AUDIT-COVERAGE-*.md")):
            DOCS.append(f)

    # Match `tools/<name>.py` and `N5/tools/<name>.py` patterns inside
    # backticks (the project's canonical script-reference syntax).
    # Avoid the wildcard `tools/*.py` (literal asterisk; not a real
    # filename) and avoid bash glob substitutions.
    SCRIPT_RE = re.compile(r"`((?:N5/)?tools/[A-Za-z0-9_\-./]+\.py)`")

    seen: set[tuple[str, str]] = set()  # (doc_relpath, script_path) — dedupe
    for doc in DOCS:
        if not doc.exists():
            failures.append(f"JA-109 governance doc missing: {doc}")
            continue
        try:
            text = doc.read_text(encoding="utf-8")
        except Exception as e:
            failures.append(f"JA-109 could not read {doc}: {e}")
            continue
        for m in SCRIPT_RE.finditer(text):
            path = m.group(1)
            # Skip wildcard / glob refs.
            if "*" in path:
                continue
            doc_label = doc.name
            if (doc_label, path) in seen:
                continue
            seen.add((doc_label, path))
            # Resolve path relative to ROOT (N5/).
            # `N5/tools/...py` → strip N5/ prefix since ROOT is already N5/.
            rel = path
            if rel.startswith("N5/"):
                rel = rel[len("N5/"):]
            candidate = ROOT / rel
            if not candidate.exists():
                failures.append(
                    f"JA-109 {doc_label} references {path!r} but file does not exist "
                    f"at {candidate} (INV-10 violation)"
                )
    return failures


def _check_ja_103_kanji_compound_form_reading_unique() -> list[str]:
    """BUG-024 (2026-05-17) regression guard.

    Within any single kanji entry's `n5_compounds` array, the
    (form, reading) tuple must be unique. Catches duplicate-subset-
    gloss compounds that were auto-derived from vocab.json before
    the VOCAB-005 / VOCAB-006 dedup cleanup propagated.

    Legitimate polysemy with DIFFERENT readings (e.g., 一日 with
    ついたち vs いちにち on the 一 entry) PASSES this check — the
    (form, reading) tuples differ. The check only fails when two
    rows share both form AND reading.
    """
    failures = []
    try:
        kanji = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-103 could not read kanji.json: {e}"]
    for k in kanji.get("entries", []):
        glyph = k.get("glyph", "?")
        seen: dict[tuple, int] = {}
        for i, c in enumerate(k.get("n5_compounds") or []):
            if not isinstance(c, dict):
                continue
            key = (c.get("form"), c.get("reading"))
            if key[0] is None:
                continue  # skip malformed; other invariants catch shape issues
            if key in seen:
                failures.append(
                    f"JA-103 kanji={glyph} n5_compounds: duplicate (form={key[0]!r}, "
                    f"reading={key[1]!r}) at indices {seen[key]} and {i}"
                )
            else:
                seen[key] = i
    return failures


def _check_ja_102_standalone_kun_primary_reading() -> list[str]:
    """BUG-021 (2026-05-17) regression guard.

    Locks the kun-yomi primary_reading for the 6 kanji whose
    standalone-in-sentence form is the kun-yomi (人 ひと, 中 なか,
    外 そと, 東 ひがし, 車 くるま, 国 くに).
    """
    failures = []
    EXPECTED = {
        "人": "ひと", "中": "なか", "外": "そと",
        "東": "ひがし", "車": "くるま", "国": "くに",
    }
    try:
        kanji = json.loads((ROOT / "data" / "kanji.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-102 could not read kanji.json: {e}"]
    by_glyph = {k.get("glyph"): k for k in kanji.get("entries", []) if k.get("glyph")}
    for glyph, expected in EXPECTED.items():
        e = by_glyph.get(glyph)
        if not e:
            failures.append(f"JA-102 kanji {glyph} not found")
            continue
        pr = e.get("primary_reading")
        if pr != expected:
            failures.append(
                f"JA-102 kanji {glyph}: primary_reading {pr!r} != expected {expected!r} "
                f"(standalone-kun lock per BUG-021)"
            )
    return failures


def _check_ja_99_vocab_form_kanji_in_scope() -> list[str]:
    """BUG-017 (2026-05-16) regression guard.

    Every kanji character appearing in any vocab.json `form` field
    must be in `data/n5_kanji_whitelist.json` OR explicitly listed in
    `data/dokkai_kanji_exception.json`. JA-13 already enforces this
    for grammar example sentences, but vocab `form` fields were not
    previously covered — the BUG-017 audit found 3 OOS kanji
    (倍, 籍, 末) had slipped through.

    The fix migrated those 3 entries to kana forms; this check
    prevents regression.
    """
    import re
    failures = []
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
        whitelist = json.loads((ROOT / "data" / "n5_kanji_whitelist.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-99 could not read inputs: {e}"]
    # n5_kanji_whitelist.json may be a list of kanji strings or a dict;
    # handle both.
    if isinstance(whitelist, dict):
        whitelist_set = set(whitelist.get("kanji", whitelist.get("entries", [])))
    elif isinstance(whitelist, list):
        whitelist_set = set(whitelist)
    else:
        return [f"JA-99 unexpected whitelist shape: {type(whitelist).__name__}"]
    # Optional exception list
    exception_set: set[str] = set()
    exc_path = ROOT / "data" / "dokkai_kanji_exception.json"
    if exc_path.exists():
        try:
            exc = json.loads(exc_path.read_text(encoding="utf-8"))
            if isinstance(exc, dict):
                exception_set = set(exc.get("kanji", exc.get("entries", [])))
            elif isinstance(exc, list):
                exception_set = set(exc)
        except Exception:
            pass
    allowed = whitelist_set | exception_set
    kanji_re = re.compile(r"[一-鿿]")
    for e in vocab.get("entries", []):
        eid = e.get("id", "?")
        form = e.get("form") or ""
        for ch in kanji_re.findall(form):
            if ch not in allowed:
                failures.append(
                    f"JA-99 {eid}: kanji {ch!r} in form {form!r} not in N5 whitelist or exception list"
                )
    return failures


def _check_ja_98_verb_transitivity_coverage() -> list[str]:
    """BUG-016 (2026-05-16) regression guard.

    Pre-fix only 22 of 132 verbs declared `transitivity`. Most N5-core
    verbs (食べる, 飲む, 買う, 会う, 行く, …) lacked the classification.
    Without explicit classification, particle assignment (を for
    transitive, に/が/と for intransitive/contact) is silent at the
    data layer — downstream tools have to infer it.

    Post-fix: 132 of 132 verbs declared. Closed enum:
      - "transitive"   (他動詞 — takes を)
      - "intransitive" (自動詞 — does not take を)
      - "contact"      (encounter / mutual-action — takes に;
                        currently 会う only, reserved for the
                        small set per BUG-008 lesson)
    """
    failures = []
    ALLOWED = {"transitive", "intransitive", "contact"}
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-98 could not read vocab.json: {e}"]
    for e in vocab.get("entries", []):
        pos = (e.get("pos") or "")
        if not pos.startswith("verb"):
            continue
        eid = e.get("id", "?")
        form = e.get("form", "?")
        t = e.get("transitivity")
        if t is None:
            failures.append(f"JA-98 {eid} ({form}): missing transitivity field")
        elif t not in ALLOWED:
            failures.append(
                f"JA-98 {eid} ({form}): transitivity {t!r} not in closed enum {sorted(ALLOWED)}"
            )
    return failures


def _check_ja_97_counter_schema_canonical() -> list[str]:
    """BUG-015 (2026-05-16) regression guard.

    Pre-fix the vocab.json corpus had three inconsistent shapes for
    `counter` (string / dict / null) plus a doubly-overloaded
    `counter_register` field (string register hint OR dict counter-word
    metadata). Migration landed in
    tools/fix_bug_015_counter_schema_2026_05_16.py.

    Post-fix canonical schema:
      counter:                   null   OR  dict {kanji, reading}
      counter_register:          null  (deprecated; reading lives in counter)
      counter_word_metadata:     null/absent  OR  dict (counter-word entries only)
    """
    failures = []
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-97 could not read vocab.json: {e}"]
    for e in vocab.get("entries", []):
        eid = e.get("id", "?")
        c = e.get("counter")
        cr = e.get("counter_register")
        # counter: null OR dict {kanji, reading}
        if c is not None:
            if not isinstance(c, dict):
                failures.append(
                    f"JA-97 {eid}: counter must be null or dict {{kanji,reading}}; got {type(c).__name__}"
                )
            elif not (isinstance(c.get("kanji"), str) and isinstance(c.get("reading"), str)):
                failures.append(
                    f"JA-97 {eid}: counter dict missing kanji or reading; got {sorted(c.keys())}"
                )
        # counter_register: must be null (deprecated)
        if cr is not None:
            failures.append(
                f"JA-97 {eid}: counter_register must be null (deprecated by BUG-015 migration); got {type(cr).__name__}"
            )
    return failures


def _check_ja_96_no_bare_arimasu_template() -> list[str]:
    """BUG-014 (2026-05-16) regression guard.

    A naive template "<form>が あります。" / "<form>が います。" was
    applied to 19 vocab entries during corpus authoring, producing
    semantically-nonsense Japanese (e.g., 「一月が あります。」 — "There is
    January.") for time words, abstract nouns, and bare locations.

    Replacement landed in tools/fix_bug_014_vocab_template_nonsense_2026_05_16.py.
    This check is the regression guard: any vocab example whose JA is
    exactly the headword followed by が あります / が います (with
    optional terminator/whitespace) is flagged.

    The check is intentionally STRICT — only the bare 'X が あります。'
    shape is flagged. Examples that wrap the headword in a location
    qualifier ('えきの 前に Xが あります。'), a time anchor
    ('七月に Xが あります。'), or any other surrounding context are
    accepted: the が-あります frame works fine for events at a time/place;
    the bug was specifically the bare standalone template.
    """
    import re
    failures = []
    try:
        vocab = json.loads((ROOT / "data" / "vocab.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-96 could not read vocab.json: {e}"]
    for e in vocab.get("entries", []):
        eid = e.get("id", "?")
        form = (e.get("form") or "").strip()
        if not form:
            continue
        # Strict pattern: optional whitespace + form + optional whitespace
        # + が + optional whitespace + あります|います + optional 。
        pat = re.compile(
            r"^\s*" + re.escape(form) + r"\s*が\s*(?:あります|います)\s*。?\s*$"
        )
        for i, ex in enumerate(e.get("examples", []) or []):
            if not isinstance(ex, dict):
                continue
            ja = (ex.get("ja") or "").strip()
            if pat.fullmatch(ja):
                failures.append(
                    f"JA-96 {eid}[{i}]: bare '<form>が あります' template "
                    f"(BUG-014 class): {ja!r}"
                )
    return failures


def _check_ja_81_no_boilerplate_leak() -> list[str]:
    """No grammar example sentence appears as the `ja` field in 10 or more
    patterns. This caps the boilerplate-leak class of bug — a small set of
    canned sentences copy-pasted across many patterns without verifying
    pattern-relevance (the bug Phase 1/2 of the 2026-05-14 content audit
    cleaned up). Some cross-pattern reuse is legitimate (e.g. 'わたしは がくせいです。'
    in basic copula patterns), so the threshold is 10 — well above normal
    legitimate reuse, well below the original 14-21 occurrences of the
    worst offenders pre-cleanup."""
    from collections import defaultdict
    failures = []
    try:
        grammar = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-81 could not read grammar.json: {e}"]
    occurrences = defaultdict(list)
    for p in grammar.get("patterns", []):
        pid = p.get("id", "?")
        for ex in p.get("examples", []):
            ja = (ex.get("ja") or "").strip()
            if ja:
                occurrences[ja].append(pid)
    THRESHOLD = 10
    for ja, pids in occurrences.items():
        if len(pids) >= THRESHOLD:
            failures.append(
                f"JA-81 sentence appears in {len(pids)} patterns "
                f"(>= {THRESHOLD} threshold, boilerplate-leak suspected): "
                f"{ja[:50]!r} in {','.join(pids[:8])}"
                + (f"... and {len(pids)-8} more" if len(pids) > 8 else "")
            )
    return failures


# ---------------------------------------------------------------------------
# JA-131..134 (2026-05-19): MOB-001..019 mobile-UI close-out invariants
# ---------------------------------------------------------------------------

def _check_ja_131_locale_all_levels_key_parity() -> list[str]:
    """MOB-007 (BUG-116) drift guard. The 'nav.all_levels' i18n key must
    exist in EVERY locale (en + hi) — used by the home-up link
    rendered in js/home.js. Catches the locale-parity gap where the
    EN string is hard-coded and a Hindi-locale user sees untranslated
    English on the home page.

    Same shape as JA-108 (locale-parity for the full key-set); JA-131
    is a narrower invariant that explicitly names this key because
    its absence shipped to production once already (MOB-007).
    """
    failures: list[str] = []
    locales_dir = ROOT / "locales"
    for loc in ("en", "hi"):
        fp = locales_dir / f"{loc}.json"
        if not fp.exists():
            failures.append(f"JA-131 locales/{loc}.json missing")
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-131 locales/{loc}.json parse error: {e}")
            continue
        nav = d.get("nav", {})
        if "all_levels" not in nav or not nav["all_levels"]:
            failures.append(f"JA-131 locales/{loc}.json missing nav.all_levels key (MOB-007 drift)")
    return failures


def _check_ja_132_touch_target_css_rules_present() -> list[str]:
    """MOB-002..016 (BUG-111..125) drift guard. The mobile-UI compliance
    CSS batch added in css/main.css (and css/main.min.css) must remain
    in place. Specifically: the 'MOB-001..016 mobile UI compliance
    batch' comment header AND a min-height:44px rule on the canonical
    set of touch-target classes (.btn-action, .study-order-link,
    .home-up-link a, .toc-expand-all, .brand-link, .skip-link,
    .btn-tiny).

    Trip condition: if any maintainer accidentally removes the
    compliance block (e.g., via a CSS-cleanup pass), every previously-
    closed MOB-* bug would silently re-regress. JA-132 catches this
    structural removal.
    """
    failures: list[str] = []
    css_files = ["css/main.css", "css/main.min.css"]
    REQUIRED_MARKERS = [
        "MOB-001..016",  # comment header
        ".btn-action",
        ".study-order-link",
        ".home-up-link",
        ".toc-expand-all",
        ".brand-link",
        ".skip-link",
        ".btn-tiny",
    ]
    for css_rel in css_files:
        fp = ROOT / css_rel
        if not fp.exists():
            failures.append(f"JA-132 {css_rel} missing")
            continue
        text = fp.read_text(encoding="utf-8")
        for marker in REQUIRED_MARKERS:
            if marker not in text:
                failures.append(f"JA-132 {css_rel} missing marker {marker!r} (MOB-* CSS regression)")
    return failures


def _check_ja_133_form_input_font_size_rule() -> list[str]:
    """MOB-006 (BUG-115) drift guard. css/main.css must contain a CSS
    rule that sets font-size >= 16px on input/textarea/select to
    prevent iOS Safari auto-zoom on focus.

    Trip condition: removal of the rule OR replacement with < 16px.
    """
    failures: list[str] = []
    fp = ROOT / "css" / "main.css"
    if not fp.exists():
        return ["JA-133 css/main.css missing"]
    text = fp.read_text(encoding="utf-8")
    import re as _re
    # Look for a rule that targets all 3 selectors together AND sets font-size >= 16px
    # The compliance batch uses: input,\n textarea,\n select { font-size: max(1rem, 16px); }
    if "max(1rem, 16px)" not in text:
        # Loose fallback: any input,textarea,select rule with font-size:16px
        if not _re.search(r"input\s*,\s*textarea\s*,\s*select.*?font-size\s*:\s*(?:16px|1rem|max\(1rem)", text, _re.DOTALL):
            failures.append(
                "JA-133 css/main.css missing input/textarea/select font-size:16px rule "
                "(MOB-006 iOS auto-zoom guard)"
            )
    return failures


def _check_ja_136_rationale_hi_unique_within_paper() -> list[str]:
    """GOI-001 (BUG-130, 2026-05-19) drift guard. No two questions within
    the same paper file may share a byte-identical `rationale_hi` value
    longer than 30 characters.

    GOI-001 caught goi-6.11's rationale_hi being a verbatim copy-paste
    of goi-6.12's (about age) on a phone-call stem — a hard learner-
    facing breakage. The original bug spec proposed "rationale_hi must
    share at least one Japanese token with stem/correctAnswer" but
    that's too strict in practice (dictionary-form ↔ polite-form
    variation produces ~100 false positives across the existing corpus,
    most legitimate references).

    The narrower-but-defensible invariant: cross-question duplication
    within the same paper file is always suspicious. Token-overlap-
    based mismatch detection is left to manual review (no CI guard).

    Threshold of 30 chars excludes trivially-short shared phrases (e.g.,
    a 1-line "Hindi-mirror-of-formula" rationale that legitimately
    repeats across multiple questions testing the same form).
    """
    failures: list[str] = []
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if fp.name == "manifest.json":
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        seen: dict[str, list[str]] = {}
        for q in d.get("questions", []):
            rh = (q.get("rationale_hi") or "").strip()
            if not rh or len(rh) < 30:
                continue
            qid = q.get("id", "?")
            seen.setdefault(rh, []).append(qid)
        for rh, qids in seen.items():
            if len(qids) > 1:
                failures.append(
                    f"JA-136 {fp.relative_to(ROOT)}: rationale_hi shared verbatim by {len(qids)} "
                    f"questions ({', '.join(qids)}) — copy-paste suspected (GOI-001 class)"
                )
    return failures


def _check_ja_135_all_hash_hrefs_resolve_to_routes() -> list[str]:
    """MOB-008/009 follow-up (2026-05-19) — general hash-route-resolution
    guard. Extends JA-134 from "2 specific known dead-end patterns"
    to "every internal hash-route href in js/*.js resolves to a route
    handler in app.js OR points at an explicit redirect path".

    Parses app.js to extract the routes-dict keys (the canonical set of
    valid route names), then greps every js/*.js file for hash-route
    hrefs (`href="#/X"` or `href='#/X'` or `` `#/X` `` in template
    literals) and asserts the first path segment exists in the routes
    dict.

    Trip cases caught:
      - href="#/listening/story" while router only maps "listeningstory"
        (the MOB-008 pattern; JA-134 covers this specifically but
        JA-135 catches future analogues without per-pattern
        hardcoding).
      - href="#/X" where X is not in routes dict (new dead-end
        introductions).

    Skips:
      - js/app.js itself (route definitions; would match itself)
      - js/min/* (minified output; source-of-truth is js/*.js)
      - Hash-only fragments (href="#" without a path) — non-routing
        hash anchors.
      - href="#/" (root) — handled by parseRoute default-to-home.
      - Anchors with non-route-prefixed hashes (e.g., href="#footnote-1") — anchors within page, not routes.
    """
    import re as _re
    failures: list[str] = []

    app_js = ROOT / "js" / "app.js"
    if not app_js.exists():
        return ["JA-135 js/app.js missing"]

    # Extract routes-dict keys from app.js. Pattern: `routes = { name: renderName, ... }`
    # Look for the routes-dict block — heuristic: 30+ entries with `: render` mapping.
    app_text = app_js.read_text(encoding="utf-8")
    # Match all `key: render<X>,` patterns in app.js
    routes = set(_re.findall(r"^\s*(\w+)\s*:\s*render\w+", app_text, _re.MULTILINE))
    # Also add 'home' and other special-cased routes
    routes.update({"home", "diagnostic"})
    if len(routes) < 10:
        return [f"JA-135 could not extract routes from app.js (found {len(routes)} entries; expected >=30)"]

    # Scan every js/*.js file (skip min/) for hash-route hrefs
    for js_fp in sorted((ROOT / "js").glob("*.js")):
        if js_fp.name == "app.js":
            continue  # routes defined here
        text = js_fp.read_text(encoding="utf-8")
        # Match href="#/X..." or href='#/X...' or `#/X...` in template literals
        # First path segment of the hash route
        for m in _re.finditer(r'''href\s*=\s*['"`]#/(\w+)''', text):
            route_name = m.group(1)
            if route_name not in routes:
                failures.append(
                    f"JA-135 {js_fp.name}: href=\"#/{route_name}/...\" but '{route_name}' "
                    f"is not in app.js routes dict (dead-end hash route — MOB-008/009 class)"
                )
        # Also catch template literals `#/X...`
        for m in _re.finditer(r'`#/(\w+)', text):
            route_name = m.group(1)
            if route_name not in routes:
                failures.append(
                    f"JA-135 {js_fp.name}: template literal `#/{route_name}/...` but '{route_name}' "
                    f"is not in app.js routes dict (dead-end hash route — MOB-008/009 class)"
                )
    # Dedupe
    return list(dict.fromkeys(failures))


def _check_ja_134_no_dead_end_hash_routes() -> list[str]:
    """MOB-008/009 (BUG-117/118) drift guard. Every internal hash-route
    reference in js/home.js + js/listening-story.js must:
      (a) point to a route that exists in js/app.js routes dict, OR
      (b) point to a path-routed URL (../ or /JLPTSuccess/...).

    Specifically catches:
      - js/home.js using '#/levels' (silent redirect to diagnostic)
      - js/listening-story.js using '#/listening/story' (silent
        redirect to listening index)

    Both were the MOB-008/009 bug shapes. Post-fix: home.js uses '../'
    directly; listening-story.js uses canonical '#/listeningstory'.
    """
    failures: list[str] = []
    home_js = ROOT / "js" / "home.js"
    if home_js.exists():
        text = home_js.read_text(encoding="utf-8")
        if 'href="#/levels"' in text:
            failures.append(
                'JA-134 js/home.js uses href="#/levels" (silently redirects to root; '
                'should be href="../" per MOB-009 fix)'
            )
    story_js = ROOT / "js" / "listening-story.js"
    if story_js.exists():
        text = story_js.read_text(encoding="utf-8")
        if '"#/listening/story' in text:
            failures.append(
                'JA-134 js/listening-story.js uses "#/listening/story..." '
                '(silently redirects to listening index; should be "#/listeningstory..." per MOB-008 fix)'
            )
    return failures


# ---------------------------------------------------------------------------
# JA-128 / JA-129 / JA-130 (2026-05-18): DOKKAI-001..003 close-out invariants
# ---------------------------------------------------------------------------

def _check_ja_128_no_passage_text_on_paper_questions() -> list[str]:
    """DOKKAI-001 (BUG-107) drift guard. Paper questions must NOT carry
    a `passage_text` field — the canonical source is `passages[label].text`
    at the paper top-level, referenced via `passage_label` foreign key.

    Catches the data-redundancy + drift class where the same passage is
    stored on every question AND in the passages[] block, with the
    "> " markdown-blockquote prefix occasionally drifting between the
    two copies.
    """
    failures: list[str] = []
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if fp.name == "manifest.json":
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        for q in d.get("questions", []):
            if "passage_text" in q:
                failures.append(
                    f"JA-128 {fp.relative_to(ROOT)} {q.get('id')} carries `passage_text` field — "
                    f"should be removed (canonical text lives in passages[].text via passage_label foreign key)"
                )
    return failures


def _check_ja_129_no_untranslated_english_temporal_in_rationale_hi() -> list[str]:
    """DOKKAI-002 (BUG-108) drift guard — extends the JA-122 PAPER-004
    fragment scan to cover untranslated English temporal markers that
    snuck through (the "ago" class).

    Trigger substrings:
      - " ago " / " ago," / " ago." / " ago)" — ago in non-bracket context
      - " yet " / " yet." / " yet," / " yet)" — yet (in similar class)
      - " before " - covered by Hindi "पहले" naturally
      - " lot " — quantity filler

    Caught dokkai-1.1 (Q1) `भूत-सकारात्मक रूप (आया एक महीना ago)।` and
    goi-7.1 `यहाँ के लिए 1 वर्ष = आया 1 वर्ष ago।` in the 2026-05-18
    sweep — both rewritten with natural Hindi using `पहले`.
    """
    failures: list[str] = []
    TRIGGERS = [
        " ago ", " ago.", " ago,", " ago)", " ago]",
        " yet ", " yet.", " yet,", " yet)",
        " lot ", " lot.", " lot,",
        # DOKKAI-004 (BUG-129, 2026-05-19) extension: " by " in
        # Devanagari context — caught dokkai-4.1 and goi-7.1.
        " by ", " by.", " by,", " by)", " by]",
        # Conservative: still skip " before " and " then " — both can appear
        # legitimately in technical fragments like "ष-form" or in
        # romanized Japanese grammatical terms.
    ]
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if fp.name == "manifest.json":
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        for q in d.get("questions", []):
            rh = str(q.get("rationale_hi", "") or "")
            for trig in TRIGGERS:
                if trig in rh:
                    failures.append(
                        f"JA-129 {fp.relative_to(ROOT)} {q.get('id')}.rationale_hi contains "
                        f"untranslated English temporal fragment {trig!r} (DOKKAI-002 class)"
                    )
                    break
    return failures


def _check_ja_130_paper_questions_have_grammarPatternId_key() -> list[str]:
    """DOKKAI-003 (BUG-109) drift guard. Every paper question must carry
    `grammarPatternId` as a key — value MAY be null when the question
    tests comprehension / vocab / orthography rather than a specific
    grammar pattern. When null, `grammarPatternId_provenance` must name
    the reason ("not_applicable_comprehension" / "not_applicable_vocab"
    / "not_applicable_orthography" / etc.).

    Same schema-shape pattern as VOCAB-002 counter-field (always a key,
    sometimes null). Prevents "missing field" vs "null value" ambiguity
    for downstream consumers.
    """
    failures: list[str] = []
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if fp.name == "manifest.json":
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        for q in d.get("questions", []):
            qid = q.get("id", "?")
            if "grammarPatternId" not in q:
                failures.append(
                    f"JA-130 {fp.relative_to(ROOT)} {qid} is missing `grammarPatternId` key "
                    f"(must be present, may be null with documented provenance)"
                )
                continue
            gpid = q.get("grammarPatternId")
            if gpid is None:
                # Null requires a typed-not-applicable provenance
                prov = q.get("grammarPatternId_provenance", "")
                if not prov.startswith("not_applicable"):
                    failures.append(
                        f"JA-130 {fp.relative_to(ROOT)} {qid} has grammarPatternId=null but "
                        f"provenance={prov!r} does not document the reason "
                        f"(expected 'not_applicable_<reason>')"
                    )
    return failures


# ---------------------------------------------------------------------------
# JA-123..127 (2026-05-18): LLM-001..005 + REG-001 close-out invariants
# ---------------------------------------------------------------------------

def _check_ja_123_papers_static_mirrors_exist() -> list[str]:
    """Every paper-pack JSON file under data/papers/*/*.json must have a
    corresponding static-mirror HTML page at /papers/<id>/index.html.
    Plus the papers/ landing index page must exist.

    LLM-001 (BUG-094) close-out: per-entity static mirrors are the
    crawler-readable surface that lets LLMs and search engines reach
    the corpus. Without these, the hash-routed SPA is invisible to
    web-fetch tools.
    """
    failures: list[str] = []
    papers_root = ROOT / "papers"
    if not (papers_root / "index.html").exists():
        failures.append("JA-123 papers/index.html (landing page) is missing")
    for fp in sorted((ROOT / "data" / "papers").rglob("*.json")):
        if fp.name == "manifest.json":
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        paper_id = d.get("id")
        if not paper_id:
            continue
        mirror = papers_root / paper_id / "index.html"
        if not mirror.exists():
            failures.append(f"JA-123 paper mirror missing for {paper_id}: expected {mirror.relative_to(ROOT)}")
    return failures


def _check_ja_124_sitemap_includes_all_mirrors() -> list[str]:
    """sitemap.xml must list URLs for every static-mirror directory and the
    7 LLM-005 summary pages. We use loose floor checks: count of <loc>
    entries must be at least (grammar + vocab + kanji + reading + listening
    + papers) + 7 summary pages + 11 meta routes. The exact 1416 ideal
    isn't strict — but anything <1000 indicates regression to the 10-URL
    pre-LLM-002 state.

    LLM-002 (BUG-095) close-out.
    """
    failures: list[str] = []
    sm = ROOT / "sitemap.xml"
    if not sm.exists():
        return ["JA-124 sitemap.xml missing"]
    text = sm.read_text(encoding="utf-8")
    n_locs = text.count("<loc>")
    # Floor: 1000 URLs (catches regression to pre-LLM-002 state at 10)
    if n_locs < 1000:
        failures.append(f"JA-124 sitemap.xml has only {n_locs} <loc> entries (expected ≥1000 after LLM-002 fix; regression suspected)")
    return failures


def _check_ja_125_data_index_corpus_sizes_match() -> list[str]:
    """data/index.json byte-sizes match actual on-disk file sizes; item
    counts match version.json counts where applicable.

    LLM-003 (BUG-096) close-out + INV-LLM-3 (same shape as INV-4 /
    JA-107 corpus-count drift class).
    """
    failures: list[str] = []
    idx_path = ROOT / "data" / "index.json"
    if not idx_path.exists():
        return ["JA-125 data/index.json missing"]
    try:
        idx = json.loads(idx_path.read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-125 could not parse data/index.json: {e}"]
    for entry in idx.get("files", []):
        rel = entry.get("path", "")
        if not rel.startswith("data/"):
            continue
        # Strip "data/" prefix and join with ROOT/data
        sub = rel[len("data/"):]
        on_disk = ROOT / "data"
        for part in sub.split("/"):
            on_disk = on_disk / part
        if not on_disk.exists():
            failures.append(f"JA-125 data/index.json lists {rel} but file does not exist")
            continue
        actual_size = on_disk.stat().st_size
        recorded_size = entry.get("size_bytes")
        if recorded_size != actual_size:
            failures.append(
                f"JA-125 data/index.json {rel}: size_bytes={recorded_size} but actual={actual_size}"
            )
    return failures


def _check_ja_126_summary_pages_exist() -> list[str]:
    """The 7 LLM-005 summary pages must exist at N5/*.html.

    LLM-005 (BUG-105) close-out + INV-LLM-5.
    """
    failures: list[str] = []
    expected = ["home.html", "grammar.html", "vocabulary.html", "kanji.html",
                "reading.html", "listening.html", "test.html"]
    for name in expected:
        if not (ROOT / name).exists():
            failures.append(f"JA-126 summary page missing: {name}")
    # Also check llms.txt at both root and N5
    if not (ROOT / "llms.txt").exists():
        failures.append("JA-126 N5/llms.txt missing")
    root_llms = ROOT.parent / "llms.txt"
    if not root_llms.exists():
        failures.append("JA-126 /JLPTSuccess/llms.txt (root) missing")
    return failures


def _check_ja_127_register_variant_no_wrong_right_with_register_why() -> list[str]:
    """REG-001 (BUG-106) close-out + INV-REG-2/3.

    Catches the drift class where a grammatical register-choice is framed
    as wrong_corrected_pair (WRONG/RIGHT). The specific n5-046 entry was
    migrated by this commit; this invariant prevents re-introduction of
    the same pattern AND surfaces any new sweep finding.

    Trip condition (strict — focused, not exhaustive):
      For every wrong_corrected_pair entry in data/grammar.json where:
        - error_category == "register", AND
        - the wrong-field annotation contains a parenthetical that
          explicitly concedes the form is appropriate in some register
          (i.e. "(formal)", "(polite)", "(in casual conversation)",
          "(among friends)", etc.) — D6 self-contradiction class

    Catches the most-egregious wrong_corrected_pair entries — the ones
    that label themselves as both Wrong AND appropriate to the named
    register. The broader sweep (D1..D5) lives in
    docs/REG-001-SWEEP-1-candidates_2026_05_18.md for native-speaker
    triage on a per-entry basis.
    """
    failures: list[str] = []
    try:
        g = json.loads((ROOT / "data" / "grammar.json").read_text(encoding="utf-8"))
    except Exception as e:
        return [f"JA-127 could not read grammar.json: {e}"]
    SELF_CONTRADICT_MARKERS = [
        "(formal)", "(polite)", "(casual)",
        "(in casual conversation)", "(in formal context)",
        "(among friends)", "(acceptable in",
    ]
    for p in g.get("patterns", []):
        pid = p.get("id", "?")
        for i, item in enumerate(p.get("wrong_corrected_pair", []) or []):
            if not isinstance(item, dict):
                continue
            cat = item.get("error_category") or item.get("category", "")
            if cat != "register":
                continue
            wrong = item.get("wrong", "")
            for marker in SELF_CONTRADICT_MARKERS:
                if marker in wrong:
                    failures.append(
                        f"JA-127 {pid}.wrong_corrected_pair[{i}] is error_category=register "
                        f"AND wrong-field contains self-contradiction marker {marker!r} — "
                        f"REG-001 D6 class. Migrate to register_variant kind in common_mistakes."
                    )
                    break
    return failures


# ---------------------------------------------------------------------------
# JA-120 / JA-121 / JA-122 (2026-05-18): PAPER-001..004 close-out invariants
# ---------------------------------------------------------------------------

def _check_ja_120_paper_pattern_id_matches_correct_particle() -> list[str]:
    """For every bunpou Mondai 1 question whose correct-answer cell is a
    single Japanese particle, the question's grammarPatternId must reference
    the canonical n5-NNN pattern for that particle.

    PAPER-001 (2026-05-18) caught 60+ bunpou questions where the
    `grammarPatternId` field had been auto-set to `n5-013` (も) by an early
    bulk-tag pass but the actual correct answer is は / が / を / に / で
    / etc. The mismatch is invisible at quiz time (the question still plays
    correctly) but breaks downstream consumers: study-plan grouping,
    pattern-frequency analytics, "review weak particles" recommendation.
    This invariant prevents re-introduction of the same drift class.
    """
    PARTICLE_TO_PATTERN = {
        "は": "n5-002", "が": "n5-003", "を": "n5-004", "に": "n5-005",
        "へ": "n5-006", "で": "n5-007", "と": "n5-008", "から": "n5-009",
        "まで": "n5-010", "や": "n5-011", "も": "n5-013", "か": "n5-023",
        "ね": "n5-025", "よ": "n5-026", "の": "n5-028", "だけ": "n5-033",
        "ぐらい": "n5-035", "くらい": "n5-035", "ごろ": "n5-036",
        "など": "n5-037", "ずつ": "n5-038", "より": "n5-095",
    }
    import re as _re
    failures: list[str] = []
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if ".bak" in fp.name:
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-120 could not parse {fp.name}: {e}")
            continue
        for q in d.get("questions", []):
            qid = q.get("id", "")
            if not qid.startswith("bunpou-") or q.get("mondai") != 1:
                continue
            correct_idx = q.get("correctIndex", -1)
            choices = q.get("choices", [])
            if not (0 <= correct_idx < len(choices)):
                continue
            correct_text = _re.sub(r"<[^>]+>", "", str(choices[correct_idx])).strip()
            if correct_text not in PARTICLE_TO_PATTERN:
                continue  # non-particle Q — different test scope
            expected = PARTICLE_TO_PATTERN[correct_text]
            actual = q.get("grammarPatternId", "<missing>")
            if actual != expected:
                failures.append(
                    f"JA-120 {fp.name} {qid}: correct={correct_text!r} "
                    f"expects grammarPatternId={expected} but got {actual} "
                    f"(PAPER-001 drift guard)"
                )
    return failures


def _check_ja_121_no_meta_fix_history_in_rationale() -> list[str]:
    """Paper-question `rationale` and `rationale_hi` fields must not contain
    commit-message-style meta-fix history (e.g., "replaces ... from a prior
    version", "Stem now anchored with X", "(was Y)", "auto_inferred",
    "previously tagged"). Learners see the rationale on the post-answer
    screen; commit-trail content is inappropriate for that surface.

    PAPER-003 (2026-05-18) caught 6 questions with this drift class.
    """
    BAD_PHRASES = [
        "auto_inferred",
        "previously tagged",
        "prior version was",
        "Stem now anchored",
        "Stem now includes",
        "replaces colloquial",
        "replaces ので per",
        "the original option",
        "was dropped because",
        "replaced with",
        "patched to",
        "fix:",
        # GOI-002 / GOI-003 (BUG-131/132, 2026-05-19) extensions:
        # commit-trail / meta-doc-pointer phrases caught in goi/paper-6.
        "Hence the rewording",
        "rewording from a prior",
        "from a prior version",
        "documented at vocabulary_n5.md",
        "documented at",
        "does not bear on",
        "test point this question",
    ]
    failures: list[str] = []
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if ".bak" in fp.name:
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-121 could not parse {fp.name}: {e}")
            continue
        for q in d.get("questions", []):
            qid = q.get("id", "")
            for field in ("rationale", "rationale_hi"):
                v = str(q.get(field, "") or "")
                for phrase in BAD_PHRASES:
                    if phrase in v:
                        failures.append(
                            f"JA-121 {fp.name} {qid}.{field}: contains "
                            f"meta-fix phrase {phrase!r} (PAPER-003 drift guard)"
                        )
                        break  # one finding per field is enough
    return failures


def _check_ja_122_no_english_pattern_fragments_in_rationale_hi() -> list[str]:
    """Paper-question `rationale_hi` fields must not contain obvious
    English-pattern fragments: apostrophe-s possessive ("दोस्त's"), English
    contractions inside Hindi ("मैं'm", "है yet"), or mojibake artifacts
    ("यहाँre", "o'घड़ी"). These all indicate word-by-word literal translation
    rather than natural Hindi.

    PAPER-004 (2026-05-18) caught 30+ questions in Mondai 2 sentence-ordering
    with these artifacts. Rationale_hi is a learner-facing Hindi explanation
    surface; it must read as natural Hindi, not transliterated English.
    """
    BAD_PATTERNS = [
        "'s घर", "'s दिन", "दोस्त's", "माता's",      # apostrophe-s possessive
        "मैं'm", "मैं'have", "मैं'll", "मैं't",        # English contractions
        "यहाँre", "वहाँre", "Iहाँ", "Iम ",            # mojibake artifacts
        "o'घड़ी", "o'क्लॉक",                          # English apostrophe in time
        " yet,", " yet.", "है yet",                  # untranslated yet
        " lot ", "I am ", " have जाना", " have रहा", # English-pattern verb fragments
    ]
    failures: list[str] = []
    paper_files = sorted((ROOT / "data" / "papers").rglob("*.json"))
    for fp in paper_files:
        if ".bak" in fp.name:
            continue
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
        except Exception as e:
            failures.append(f"JA-122 could not parse {fp.name}: {e}")
            continue
        for q in d.get("questions", []):
            qid = q.get("id", "")
            rh = str(q.get("rationale_hi", "") or "")
            for bad in BAD_PATTERNS:
                if bad in rh:
                    failures.append(
                        f"JA-122 {fp.name} {qid}.rationale_hi: contains "
                        f"English-pattern fragment {bad!r} (PAPER-004 drift guard)"
                    )
                    break
    return failures


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]
    verbose = "-v" in argv or "--verbose" in argv

    overall_failures: list[str] = []
    print(f"JLPT N5 Content Integrity - {len(CHECKS)} invariants")
    # Note: invariant count grew from 23 to 24 with JA-15 (audio refs on disk).
    print("=" * 60)

    for code, label, check_fn in CHECKS:
        try:
            failures = check_fn()
        except Exception as e:  # noqa: BLE001 - report any check-internal error
            failures = [f"{code} check raised {type(e).__name__}: {e}"]
        status = "PASS" if not failures else f"FAIL ({len(failures)})"
        print(f"  {code:<6} {label:<32} {status}")
        if failures and verbose:
            for f in failures[:10]:
                print(f"           - {f}")
            if len(failures) > 10:
                print(f"           ... and {len(failures) - 10} more")
        overall_failures.extend(failures)

    print("=" * 60)
    if overall_failures:
        print(f"FAIL: {len(overall_failures)} integrity violation(s)")
        if not verbose:
            print("Run with -v / --verbose to see all violations")
        return 1

    print(f"PASS: all {len(CHECKS)} invariants green")
    return 0


if __name__ == "__main__":
    sys.exit(main())
