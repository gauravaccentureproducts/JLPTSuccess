"""Round-9 audit registration — 2026-05-06.

Registers Section-2 issues + Section-3 improvements + Section-6 questions from
`feedback/audit-round9-2026-05-06.md` into `feedback/n5-audit-2026-05-04.xlsx`.

Idempotent: skips IDs already present in the sheet. Modeled on
`tools/register_dev_issue_list_deferrals_2026_05_05.py`.

Schema (Items sheet, header at row 4):
  1=ID  2=Type  3=Severity  4=Priority  5=Impact  6=Effort  7=Category
  8=Location  9=Title  10=Current state  11=Why-it-matters  12=Suggested-direction
  13=Dependencies  14=Decision  15=Description  16=Permission

Run:
  python tools/register_audit_round9_2026_05_06.py
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).parent.parent / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# (id, type, severity, priority, impact, effort, category, location, title,
#  current_state, why_matters, suggested_direction, dependencies, decision,
#  description, permission)
#
# IDs continue from existing max: ISSUE-090 (round-8), IMP-114 (round-8),
# Q38 (round-8). Round-9 starts at ISSUE-091, IMP-115, Q39.

ITEM_ROWS = [
    # --- ISSUES (P1) ---
    ('ISSUE-091', 'Issue', 'BLOCKER', 'P1', 'HIGH', 'MEDIUM',
     'i18n / content depth (questions)',
     'data/questions.json — every question entry',
     '[N1] [DEPTH] explanation_hi missing on all 290 test-mode questions',
     '0/290 questions have explanation_hi; test-mode renderer falls back to explanation_en regardless of UI locale.',
     'Hindi-locale learners who set EN|HI to HI in settings still see English-only rationale on every wrong answer; breaks niche-N1 contract at the highest-engagement surface even on the 27 native-reviewed grammar patterns.',
     'LLM-persona Hindi translation pass on explanation_en -> explanation_hi; gate behind meaning_provenance per-question; ship in a single batch.',
     '—',
     'Fix',
     'Round-9 audit. Niche-N1 critical: test-mode rationale is the largest learner-engagement surface. Without this, even native-reviewed grammar patterns produce English-only test feedback.',
     'No permission required'),

    ('ISSUE-092', 'Issue', 'MAJOR', 'P1', 'MEDIUM', 'MEDIUM',
     'i18n / content depth (questions)',
     'data/questions.json — 137 questions with distractor_explanations',
     '[N1] [DEPTH] distractor_explanations_hi missing on 137 questions',
     '137/290 questions have distractor_explanations (English); 0/137 have distractor_explanations_hi.',
     'When a Hindi-locale learner gets a question wrong, the per-distractor wrongness rationale renders in English. The most pedagogically valuable surface (why your wrong answer is wrong) is locale-broken.',
     'Same LLM-persona Hindi translation pass extended to distractor_explanations key-by-key.',
     'ISSUE-091',
     'Fix',
     'Round-9 audit. Pairs with ISSUE-091; both ship together for end-to-end Hindi test mode.',
     'No permission required'),

    ('ISSUE-093', 'Issue', 'MAJOR', 'P1', 'HIGH', 'MEDIUM',
     'i18n / native review',
     'data/grammar.json — patterns where meaning_provenance != native_reviewed',
     '[N1] [DEPTH] Q33 LLM-persona Hindi review scaling on remaining 151 grammar patterns',
     '27/178 grammar patterns have l1_notes.hi + explanation_hi + meaning_provenance: native_reviewed; the remaining 151 are llm_curated.',
     'Round-8 validated the Q33 LLM-persona model on top-30 patterns. Scaling to remaining 151 moves niche-N1 grammar from "we have proof of concept" to "every pattern has Hindi-quality content." Validated process; just throughput.',
     'Reuse round-8 batch script pattern; review in 50-pattern chunks for traceability; ship in 3 commits.',
     '—',
     'Fix',
     'Round-9 audit. Highest-leverage scaling target — process is proven, only throughput remains.',
     'No permission required'),

    ('ISSUE-094', 'Issue', 'MAJOR', 'P1', 'HIGH', 'HIGH',
     'i18n / native review',
     'data/vocab.json, data/kanji.json',
     '[N1] [DEPTH] Native_reviewed scaling on vocab (1041) + kanji (106)',
     '0/1041 vocab and 0/106 kanji have review_status: native_reviewed; both surfaces sit at 100% llm_curated for gloss_hi / meanings_hi.',
     'Provenance-badge UI (active on grammar after round-8) needs >=10% native_reviewed per surface (Q21). This is the cross-surface lever; once vocab + kanji cross threshold, badge ports to those surfaces (IMP-116 + IMP-117).',
     'Q33 LLM-persona-review approach in batches of ~100 entries; add gloss_hi_provenance + meanings_hi_provenance fields (IMP-117 prerequisite).',
     'IMP-117',
     'Fix',
     'Round-9 audit. Most expensive single item but highest niche-N1 unlock.',
     'No permission required'),

    # --- ISSUES (P2) ---
    ('ISSUE-095', 'Issue', 'MAJOR', 'P2', 'HIGH', 'MEDIUM',
     'Mock-test authenticity',
     'data/papers/manifest.json + js/papers.js + js/sitting.js',
     '[N4] [META] Mock papers do not match real JLPT-N5 paper shape (35/32/24)',
     '28 papers x 15-Q single-mondai sets (paper-1 = M1 only). No combined-section paper. No chokai (listening) papers — listening is 47 standalone items not arrayed as 24-Q paper.',
     'Real JLPT N5 = 35Qx25min (moji+goi) / 32Qx50min (bunpou+dokkai) / 24Qx30min (chokai). Learners cannot run real-exam dress rehearsal inside the app. Best-in-class (JLPT Sensei sample papers, Marugoto exam-mode) all ship the real shape.',
     'Build virtual paper aggregator: combine moji[paper-N] + goi[paper-N] into "言語知識 paper-N" (~30Q with IMP-086 25-min combined timer); same for bunpou+dokkai; add chokai virtual paper from listening 47-pool (M1x7 + M2x6 + M3x5 + M4x6).',
     'IMP-115, IMP-118',
     'Fix',
     'Round-9 audit. Niche-N4 (all-in-one) defining gap. Q40 covers shape vs author-new tradeoff.',
     'No permission required'),

    ('ISSUE-096', 'Issue', 'MAJOR', 'P2', 'HIGH', 'MEDIUM',
     'Content depth (vocab)',
     'data/vocab.json — entries with len(examples) < 2',
     '[N4] [DEPTH] Vocab examples >=2 missing on 927/1041 entries (89%)',
     '114/1041 (11%) have >=2 examples; spec floor is 2. Worst offenders: 私たち, かれ, かのじょ, みなさん, どなた.',
     'A serious vocab page needs >=2 examples in different contexts. 11% coverage means 89% of vocab pages feel anemic vs Bunpro/Renshuu.',
     'Auto-derive second example from grammar.json vocab_ids cross-references; LLM-curated for residual ~96 entries without grammar match.',
     '—',
     'Fix',
     'Round-9 audit. Largest absolute count vocab deficit; mostly auto-derivable.',
     'No permission required'),

    ('ISSUE-097', 'Issue', 'MAJOR', 'P2', 'HIGH', 'MEDIUM',
     'Content depth (grammar)',
     'data/grammar.json — patterns with len(examples) < 5',
     '[N4] [DEPTH] Grammar examples >=5 missing on 152/178 patterns (85%)',
     '26/178 (15%) have >=5 examples; best-in-class Bunpro is 5-7. Worst offenders: n5-024, n5-027, n5-028, n5-030, n5-031.',
     'Bunpro ships 5-7 per pattern; the perceived-depth gap is visible on a single visit. Currently 152 patterns sit at the spec floor (3) instead of best-in-class.',
     '2 new examples per pattern, LLM-drafted with attachment-surface diversity constraint; spot-check JA-13 (kanji subset) and JA-1 (vocab subset).',
     '—',
     'Fix',
     'Round-9 audit. Bulk authoring; diversity (different attachment surface / register / topic) more important than count alone.',
     'No permission required'),

    ('ISSUE-098', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'LOW',
     'i18n / content depth (reading)',
     'data/reading.json — passages',
     '[N1] [DEPTH] Reading summary_hi missing on all 45 passages',
     '0/45 passages have summary_hi.',
     'Hindi-locale learners read English summary on every passage. Summary is critical navigation aid (decide whether to attempt) — English-only summary breaks niche-N1 reading surface.',
     'LLM-persona Hindi translation of summary -> summary_hi.',
     '—',
     'Fix',
     'Round-9 audit. Low-effort niche-N1 win.',
     'No permission required'),

    ('ISSUE-099', 'Issue', 'MAJOR', 'P2', 'HIGH', 'LOW',
     'Schema / depth (vocab)',
     'data/vocab.json — entries where pos.startswith("verb")',
     '[N4] [DEPTH] Vocab verb_class missing on all 134 verbs',
     '0/134 verbs have verb_class field. 6 known Group-1 exception verbs (入る, 帰る, 走る, 知る, 切る, 要る) all unflagged.',
     'Without this flag, conjugation drill cannot programmatically tell ichidan from godan from irregular. Risk: drill produces wrong forms for the 6 exception verbs (X-6.6 invariant in spec).',
     'Map pos -> verb_class deterministically; add group1_exception: true on the 6 known exception verbs.',
     '—',
     'Fix',
     'Round-9 audit. Low effort; programmatic correctness depends on it.',
     'No permission required'),

    # --- ISSUES (P3) ---
    ('ISSUE-100', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'LOW',
     'Schema / depth (vocab)',
     'data/vocab.json',
     '[N4] [DEPTH] Vocab pair_id (transitivity) only on 22/1041 entries',
     '22 entries paired. The 12 canonical N5 transitivity pairs (開ける/開く, 閉める/閉まる, 入れる/入る, 出す/出る, 始める/始まる, 止める/止まる, つける/つく, 消す/消える, 起こす/起きる, 落とす/落ちる, 直す/直る, 切る/切れる) need full coverage = 24 entries minimum.',
     'Transitive/intransitive pair drilling is a defining N5 grammar exam topic (が vs を, automatic vs deliberate action).',
     'Author the 12 pairs in a single batch script; add pair_id cross-references both directions.',
     '—',
     'Fix',
     'Round-9 audit. Easy wins; 24 entries.',
     'No permission required'),

    ('ISSUE-101', 'Issue', 'MAJOR', 'P3', 'MEDIUM', 'LOW',
     'Content depth (kanji)',
     'data/kanji.json — entries with len(examples) < 5',
     '[N4] [DEPTH] Kanji examples >=5 on 93 kanji (88% deficit)',
     '13/106 kanji have >=5 examples. Worst-offenders: 道, 名, 百, 千, 万, 円, 火, 水, 木, 金 (numerals + elements; many appear in vocab but cross-link is missing).',
     'WaniKani ships consistent depth across all kanji. Inconsistent (some 6 examples, some 2) reads as scaffolding rather than finished product.',
     'Auto-script: for each kanji, pull all vocab entries where form contains the glyph; pick top-N most common; cross-link.',
     '—',
     'Fix',
     'Round-9 audit. Almost entirely auto-derivable from vocab corpus.',
     'No permission required'),

    ('ISSUE-102', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'MEDIUM',
     'Content depth (grammar)',
     'data/grammar.json — patterns with empty contrasts',
     '[N4] [DEPTH] Grammar contrasts missing on 83 patterns (47%)',
     '95/178 (53%) have >=1 contrast. The 11 mandatory N5 contrast clusters (は/が, から/ので, も/と, で/に, けど/が, ~たことがある/~た, ~ている progressive/resultative, ~ましょう/~ませんか, あげる/くれる/もらう) need every pattern in those clusters cross-linked.',
     'Confusable-pair disambiguation is the most common learner failure mode at N5; explicit cross-links help users find both sides.',
     'Audit the 11 mandatory clusters; ensure every pattern in each cluster cross-links to its sibling.',
     '—',
     'Fix',
     'Round-9 audit. Cluster-by-cluster manual cross-link.',
     'No permission required'),

    ('ISSUE-103', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'LOW',
     'Content depth (reading)',
     'data/reading.json — passages',
     '[N4] [DEPTH] Reading cultural_context missing on 45/45',
     '0/45 passages have cultural_context.',
     'Passages mentioning Japan-specific concepts (おにぎり, school clubs, 銭湯, 塾) need brief explainer for non-Japan-domiciled learners. Without it, comprehension question is harder than Japanese alone justifies.',
     '1-2 sentence cultural callout per passage where applicable; ~15-20 of 45 will need one.',
     '—',
     'Fix',
     'Round-9 audit. Low-effort niche-N4 polish.',
     'No permission required'),

    ('ISSUE-104', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'Schema (reading)',
     'data/reading.json — passages',
     '[none] Reading format_role missing on all 45',
     '0/45 passages have format_role (primary vs supplementary discipline per JA invariants).',
     'Spec calls for primary/supplementary classification so paper-builder can weight question-type distribution; without it builder is blind to the structural property.',
     'Tag every passage primary/supplementary based on question type; primary if >=2 inference questions, supplementary if all-fact-retrieval.',
     '—',
     'Fix',
     'Round-9 audit. Spec-compliance hygiene.',
     'No permission required'),

    # --- IMPROVEMENTS (P2) ---
    ('IMP-115', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'MEDIUM',
     'Test mode UI / mock-paper engine',
     'js/papers.js, js/sitting.js, data/papers/manifest.json',
     '[N4] [META] Real-exam-shape full mock paper UI',
     'Test mode runs single-mondai 15-Q sets only. No "full mock paper" entry point.',
     'JLPT Sensei + official jlpt.jp sample papers ship the real 35/32/24 shape; learners doing serious exam prep need this for time-management practice.',
     'Add "Full Mock Test" tile on #/test that aggregates moji[paper-N] + goi[paper-N] (35Q, 25min combined timer) + bunpou[paper-N] + dokkai[paper-N] (32Q, 50min) + chokai-virtual (24Q, 30min). Reuse IMP-086 timing infrastructure.',
     'ISSUE-095, IMP-118',
     'Fix',
     'Round-9 audit. Pairs with ISSUE-095. Q40 covers approach.',
     'No permission required'),

    ('IMP-116', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'MEDIUM',
     'Trust signal / UI',
     'js/learn-vocab.js, js/kanji.js, js/provenance-badge.js',
     '[N1] [META] Provenance-badge UI on vocab + kanji surfaces',
     'Provenance badge active on grammar only (round-8 v1.12.41). Vocab/kanji detail pages do not expose review status.',
     'Bunpro shows native-reviewed indicators on every entry. Cross-surface trust signal.',
     'Port provenance-badge.js rendering to vocab + kanji detail templates; gate behind per-surface >=10% native_reviewed (Q21 rule); ship after IMP-117 lands data shape.',
     'IMP-117, ISSUE-094',
     'Fix',
     'Round-9 audit. Cross-surface badge; activates only when ISSUE-094 reaches threshold.',
     'No permission required'),

    ('IMP-117', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Schema / data shape',
     'data/vocab.json, data/kanji.json, tools/check_content_integrity.py',
     '[N1] [META] gloss_hi_provenance + meanings_hi_provenance data shape',
     'Vocab has gloss_provenance; kanji has meanings_provenance. Hindi-specific provenance is implicit (whole-entry review_status). Need per-locale-per-field provenance to match grammar (meaning_provenance vs explanation_provenance separately).',
     'Mirrors round-6 grammar provenance shape; needed for IMP-116 to render badge correctly.',
     'Add gloss_hi_provenance to vocab schema, meanings_hi_provenance to kanji schema; default to entry-level review_status for backwards compat; add JA-40 invariant to enforce closed enum.',
     '—',
     'Fix',
     'Round-9 audit. Prerequisite for IMP-116 + ISSUE-094.',
     'No permission required'),

    ('IMP-118', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Test mode (listening)',
     'js/papers.js, data/papers/manifest.json',
     '[N4] [META] Listening 24-Q chokai virtual paper',
     'Listening is 47 standalone items, not arrayed as a 24-Q paper.',
     'Real JLPT N5 chokai is 24 questions x 30 minutes. Without virtual paper, learners cannot run a chokai dress rehearsal.',
     'Sample 7 M1 + 6 M2 + 5 M3 + 6 M4 from listening pool; surface as virtual paper in manifest.json under new chokai category; reuse IMP-086 30-min timer.',
     'ISSUE-095',
     'Fix',
     'Round-9 audit. Pairs with ISSUE-095 + IMP-115.',
     'No permission required'),

    # --- IMPROVEMENTS (P3) ---
    ('IMP-119', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'LOW',
     'UI / vocab',
     'js/learn-vocab.js',
     '[N4] [META] Vocabulary keigo-chain visualizer',
     '9 entries with register_chain_id (3 trios from v1.12.42) render as independent leaf pages.',
     'Side-by-side trio panel makes the keigo politeness contrast immediately visible — currently learners have to navigate three pages.',
     'When a vocab page has register_chain_id, render side-by-side trio panel (humble | plain | respectful) at top of detail view.',
     '—',
     'Fix',
     'Round-9 audit. Low-effort UI on already-existing data.',
     'No permission required'),

    ('IMP-120', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'Test mode UI',
     'js/papers.js paper-list grid',
     '[N4] [META] Per-paper timing display (~7 min in real exam pace)',
     'Paper cards show question count + difficulty but no expected time.',
     'Helps learners time-budget self-study; common feature in best-in-class exam prep.',
     'On each paper card, show "~X min in real exam pace" derived from sec-per-question table (43 sec moji/goi, 94 sec bunpou/dokkai, 75 sec chokai).',
     '—',
     'Fix',
     'Round-9 audit. Trivial render addition.',
     'No permission required'),

    ('IMP-121', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'Test mode results',
     'js/papers.js results view',
     '[N4] [META] Full-paper score-report breakdown (real JLPT shape)',
     'After paper submit, shows total correct/total. No per-mondai breakdown, no per-section score, no pass/fail per minimum.',
     'Real JLPT N5 score report shows per-section scores + per-minimum pass status (38/19/19); replicating this shape is essential for exam-prep authenticity.',
     'Score-report UI matching real JLPT shape: 言語知識(文字・語彙) X/35, 言語知識(文法)・読解 X/32, 聴解 X/24, Total X/180, with pass-mark display.',
     'IMP-115',
     'Fix',
     'Round-9 audit. Ships after IMP-115 full-paper UI.',
     'No permission required'),
]

# Section 6 questions — Questions sheet has its own schema (data starts row 2).
# Inspect existing rows for column convention; we use:
#   1=ID  2=Title  3=Description  4=Affects/dependencies  5=Status  6=Source
QUESTIONS_ROWS = [
    ('Q39', 'Native-Hindi-review scaling commitment',
     'Round-8 validated Q33 LLM-persona model on 27 grammar patterns. Scaling to remaining 151 grammar + 1041 vocab + 106 kanji + 45 reading + 47 listening ~= 45-60 hours of LLM-persona review. Decision needed: full corpus-wide scaling now (>=10% per surface = badge unlock), or stage by surface (grammar first -> 100%, then vocab -> 10%, then others)?',
     'Affects ISSUE-093, ISSUE-094, ISSUE-098, IMP-116',
     'Open',
     'Round-9 audit'),

    ('Q40', 'Real-exam-shape mock paper approach',
     'ISSUE-095 + IMP-115 propose virtual aggregator over existing papers. But underlying papers are 15-Q single-mondai sets; combining moji[paper-N] + goi[paper-N] gives 30 questions, not the canonical 35. Decision: build virtual aggregator on existing pool (acceptable approximation), OR defer until paper-redesign cycle that authors true 35/32/24-shape papers from KnowledgeBank?',
     'Affects ISSUE-095, IMP-115, IMP-118',
     'Open',
     'Round-9 audit'),

    ('Q41', 'Vocab counter field scope',
     'Counter-on-nouns work needs 30-50 high-frequency nouns for common counters (本->冊, 車->台, 人->人, 紙->枚, 個). Going to full 589 nouns yields diminishing returns (most do not take a specific counter). Decision: top-50 high-frequency only (~2 hr), or full 589 sweep with explicit "no specific counter" flag (~10 hr)?',
     '—',
     'Open',
     'Round-9 audit'),

    ('Q42', 'Listening voice variety budget',
     'Voice variety has been deferred 3 times (round-3, round-7, round-8) for lack of budget decision. Three options: (a) install VOICEVOX multi-voice locally — free but 4-6 hr integration; (b) ElevenLabs free tier — 10K chars/month, may not cover 47 items; (c) recruit 2-4 native speakers via Indian Japanese-language teacher network. Decision: which lane?',
     'Affects ISSUE-062, ISSUE-089, ISSUE-090, IMP-094',
     'Open',
     'Round-9 audit'),

    ('Q43', 'Provenance-badge UI rollout strategy',
     'IMP-116 ports provenance badge to vocab + kanji. But Q21 says >=10% native_reviewed per surface to show badge. Today vocab + kanji = 0%. Decision: ship data shape now (IMP-117) and let badge self-activate when ISSUE-094 reaches threshold, OR bundle them as one batch?',
     'Affects IMP-116, IMP-117, ISSUE-094',
     'Open',
     'Round-9 audit'),

    ('Q44', 'Onboarding "your first 60 seconds" path',
     'Brand-new user lands on home page and sees five surface cards. No guided "start here" path. Audit notes this as potential improvement but cannot decide whether to build tutorial overlay (clutters), recommended-first-pattern page (low effort), or curriculum mode (out-of-scope for this cycle). Decision: which onboarding shape, if any?',
     '—',
     'Open',
     'Round-9 audit'),
]


def main() -> int:
    if not XLSX.exists():
        print(f'ERROR: workbook not found at {XLSX}')
        return 1

    wb = load_workbook(XLSX)
    ws = wb['Items']
    qsheet = wb['Questions']

    # Build set of existing IDs (Items: data starts row 5)
    existing_items = set()
    for r in range(5, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid:
            existing_items.add(str(rid).strip())

    existing_questions = set()
    for r in range(2, qsheet.max_row + 1):
        rid = qsheet.cell(r, 1).value
        if rid:
            existing_questions.add(str(rid).strip())

    # Append items
    appended = 0
    skipped = 0
    first_row = ws.max_row + 1
    for row in ITEM_ROWS:
        rid = row[0]
        if rid in existing_items:
            print(f'SKIP (already present): {rid}')
            skipped += 1
            continue
        target_r = ws.max_row + 1
        for c, val in enumerate(row, start=1):
            ws.cell(target_r, c).value = val
        appended += 1
    last_row = ws.max_row

    # Append questions
    q_appended = 0
    q_skipped = 0
    q_first = qsheet.max_row + 1
    for q in QUESTIONS_ROWS:
        qid = q[0]
        if qid in existing_questions:
            print(f'SKIP question (already present): {qid}')
            q_skipped += 1
            continue
        qr = qsheet.max_row + 1
        for c, val in enumerate(q, start=1):
            qsheet.cell(qr, c).value = val
        q_appended += 1
    q_last = qsheet.max_row

    wb.save(XLSX)

    n_issues_appended = sum(1 for r in ITEM_ROWS if r[1] == 'Issue' and r[0] not in existing_items)
    n_imps_appended = sum(1 for r in ITEM_ROWS if r[1] == 'Improvement' and r[0] not in existing_items)

    print(f'\nItems sheet: appended {appended} (rows {first_row}..{last_row}); skipped {skipped}.')
    print(f'Questions sheet: appended {q_appended} (rows {q_first}..{q_last}); skipped {q_skipped}.')
    print(f'\nRegistration receipt:')
    if appended + q_appended == 0:
        print(f'  Registered 0 findings (no new items detected)')
    else:
        print(f'  Registered {appended + q_appended} findings into n5-audit-2026-05-04.xlsx')
        print(f'  (Items rows {first_row}..{last_row} = {n_issues_appended} issues + {n_imps_appended} improvements; '
              f'Questions rows {q_first}..{q_last} = {q_appended} questions).')
    return 0


if __name__ == '__main__':
    sys.exit(main())
