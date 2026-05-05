"""Register the deferred items from the 2026-05-05 developer-issue-list
audit into feedback/n5-audit-2026-05-04.xlsx.

The developer-issue-list (24 items) was processed in three commits:
  - 1d42e75 (doc parity)
  - 3358838 (skip-link a11y)
  - 9d0e3a6 (niche-N3 polish)
Plus old-repo redirects + archives.

11 items were fixed; 4 closed as already-fixed (#2/#6/#21) or
false-positive (#7/#13/#23/#4); 6 items were deferred. This script
adds the 6 deferred items + 1 partial item to the canonical registry
so they survive the conversation transcript and can be picked up by
future audit rounds.

ID assignments continue from the existing range:
  - last ISSUE was ISSUE-053 → new issues start at ISSUE-054
  - last IMP was IMP-074 → new improvements start at IMP-075

Idempotent: if any of these IDs already exist in the sheet, the script
warns and skips that row. Run once.
"""
from __future__ import annotations
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# (ID, Type, Severity, Priority, Impact, Effort, Category, Location, Title,
#  Current state, Why-it-matters, Suggested direction, Dependencies,
#  Decision, Description, Permission)
NEW_ROWS = [
    (
        'ISSUE-054', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
        'Documentation / runtime parity',
        'N5/sw.js, real browser DevTools',
        'Service-worker scope conflict between top-level and per-level apps unverified',
        'Top-level JLPTSuccess/index.html does NOT register a service worker; each level (/N5/, /N4/) registers its own SW with subdirectory scope. By design, scopes do not overlap. But this has not been verified end-to-end in DevTools (load /N5/, confirm only one SW with scope /JLPTSuccess/N5/; navigate to /, confirm no SW takes over; navigate to /N4/, confirm a separate SW with its own scope).',
        'Self-host / institutional adopters and serious privacy-conscious users will check this. A clean SW scope map is a niche-N2 + niche-N3 trust signal.',
        'Run a manual DevTools verification across /JLPTSuccess/, /JLPTSuccess/N5/, /JLPTSuccess/N4/; document the resulting SW scope map in N5/specifications/JLPT-N5-Current-Implementation-Spec.md §9.',
        'None',
        'Defer',
        'From the 2026-05-05 developer-issue-list audit (item #19). Verification requires a real browser; cannot be done via curl. Low impact (the architecture is correct by construction; this is a "show-our-working" item).',
        'Verify in any browser; no permission needed beyond the live URL.'
    ),
    (
        'ISSUE-055', 'Issue', 'MINOR', 'P3', 'MEDIUM', 'MEDIUM',
        'UX / Documentation surface',
        'N5/PRIVACY.md, N5/NOTICES.md served as raw .md',
        'Privacy and Notices markdown served raw, mobile UX poor on Safari',
        'The footer links <a href="PRIVACY.md"> and <a href="NOTICES.md"> serve files with Content-Type: text/markdown. Modern Chrome/Firefox render them as plain text; mobile Safari downloads them. Result: users see raw markdown syntax (or get a download dialog) rather than a styled page.',
        'Privacy / Notices are user-trust surfaces. A raw-markdown render breaks the "this is a serious project" signal — niche-N2 (privacy posture credibility) and niche-N3 (institutional adoption).',
        'Add a #/privacy and #/notices SPA route in N5 that fetches the .md, runs a minimal markdown renderer, and renders inside the app shell. Or pre-render to .html at build time. Either is medium-effort.',
        'None',
        'Defer',
        'From the 2026-05-05 developer-issue-list audit (item #15). Design decision deferred — needs a choice between (a) build-time .md→.html rendering or (b) runtime in-app route with markdown-it.',
        'No permission required.'
    ),
    (
        'IMP-075', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'HIGH',
        'N4 sub-app — work-blocked',
        'JLPTSuccess/N4/index.html (head metadata block)',
        'N4 sub-app missing OpenGraph / Twitter Card / JSON-LD social-share metadata',
        'N5 has full og:* + twitter:* + EducationalApplication JSON-LD. N4 only has the basic description / theme-color / viewport / canonical block. Social shares of any /N4/ URL get no preview card — falls back to a bare URL link.',
        'Same pattern as ISSUE-032 / IMP-051 in N5; copying that block to N4 is straightforward when N4 unblocks. Niche-N3 (level-parity for institutional self-host).',
        'Copy the og:* + twitter:* + JSON-LD block from N5/index.html into N4/index.html, swapping URLs and the EducationalApplication.educationalLevel value.',
        'Lifting the N4 work-block (per .claude/CLAUDE.md governance Rule 1)',
        'Defer',
        'From the 2026-05-05 developer-issue-list audit (item #5). Work-blocked: any N4 file edit requires explicit user instruction "unblock N4" first.',
        'Cannot be picked up until the N4 work-block is lifted.'
    ),
    (
        'IMP-076', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'HIGH',
        'N4 sub-app — work-blocked',
        'JLPTSuccess/N4/index.html, JLPTSuccess/N4/js/app.js',
        'N4 nav missing Mock and Missed routes that N5 has',
        'N5 primary-nav: Grammar | Vocabulary | Kanji | Reading | Listening | Test | Mock | Missed | Progress. N4 primary-nav: Grammar | Vocabulary | Kanji | Reading | Listening | Test | Progress. Two routes (#/sitting Mock + #/missed Missed) shipped to N5 in round-3 but never ported to N4.',
        'Per-level feature parity is part of the niche-N4 (all-in-one) and niche-N3 (consistent product across levels) claim.',
        'Port the renderSitting and renderMissed modules from N5 to N4, plus the corresponding nav links. OR document the per-level feature delta explicitly in README.',
        'Lifting the N4 work-block; ISSUE-020/IMP-031/IMP-032 (the original Mock/Missed work) are already done in N5.',
        'Defer',
        'From the 2026-05-05 developer-issue-list audit (item #10). Work-blocked. Note: while N4 is hidden as "Coming soon" on the level picker, this nav-parity gap is technically not user-visible today, so impact is theoretical until N4 unblocks.',
        'Cannot be picked up until the N4 work-block is lifted.'
    ),
    (
        'IMP-077', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'MEDIUM',
        'N4 sub-app — work-blocked',
        'JLPTSuccess/N4/index.html (secondary-nav locale switcher)',
        'N4 has no language switcher despite multilingual claim',
        'N5 ships a 5-chip locale switcher (EN/VI/ID/NE/ZH) in the secondary-nav of every page. N4 has no switcher. The repo description and README both claim multilingual, but only N5 surfaces it.',
        'Niche-N1 (multilingual non-English-native learners) is the primary strategic positioning per the round-4 audit. Per-level multilingual parity is essential to that claim.',
        'Port the locale-chip-group component from N5 secondary-nav to N4. Same data attributes, same i18n module, same CSS class. Bring N4 locale files (en/vi/id/ne/zh) to parity afterwards.',
        'Lifting the N4 work-block; round-4 i18n work in N5 (ISSUE-028 / IMP-058 etc.).',
        'Defer',
        'From the 2026-05-05 developer-issue-list audit (item #11). Work-blocked. While N4 is hidden, this is invisible — but the multilingual claim in marketing surfaces (repo description, level-picker card) is technically incomplete until N4 ships locale parity.',
        'Cannot be picked up until the N4 work-block is lifted.'
    ),
    (
        'IMP-078', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'LOW',
        'Top-level brand / level picker',
        'JLPTSuccess/index.html (N4 card description)',
        'N4 level-picker card has no content-count breakdown',
        'N5 card on the level picker says "178 grammar patterns, 1041 vocab (700+ official + supplementary), 106 kanji, 40 reading + 40 listening drills" — concrete trust signal. N4 card just says "Builds on N5 with everyday topics, basic written passages, and lower-frequency kanji" — no counts.',
        'Card-level content-count parity is a small-but-real trust signal for visitors deciding whether to wait for N4 or use a competitor.',
        'When N4 unblocks and ships authored content, mirror the N5 count pattern in the N4 card description on JLPTSuccess/index.html.',
        'Lifting the N4 work-block + N4 reaching some authored content threshold (currently 0 native-reviewed items).',
        'Defer',
        'From the 2026-05-05 developer-issue-list audit (item #17). Adding counts now would either lie ("0 / 0 / 0") or freeze a forward promise that N4 cannot keep while work-blocked. Skip until N4 has real content.',
        'Cannot be picked up until the N4 work-block is lifted.'
    ),
    (
        'IMP-079', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'MEDIUM',
        'CI / integrity invariant coverage',
        'N5/tools/check_content_integrity.py',
        'No locale-completeness CI invariant',
        'Today there is no integrity invariant verifying that all 5 locale files (locales/en.json, vi.json, id.json, ne.json, zh.json) have parity in keys. A vi.json missing a key fails silently to English fallback at runtime — invisible to CI.',
        'Niche-N1 (multilingual) trust degrades silently if locale drift is not caught. Bunpro / WaniKani do not have this problem because they only ship English; for a multilingual claim, locale-parity CI is table-stakes.',
        'Add a JA-NN invariant (next number in the JA-1..JA-33 sequence) that loads all 5 locale JSONs, computes the symmetric-difference of their key sets, and fails if non-empty.',
        'None — pure tooling addition.',
        'Defer',
        'Surfaced during the 2026-05-05 developer-issue-list audit ("what the issue list missed" — item 6). Low impact today (locale files are small and recently audited); becomes high-impact as the locales are translated to native quality.',
        'No permission required.'
    ),
]

def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False)
    ws = wb['Items']
    existing_ids = set()
    for r in range(5, ws.max_row + 1):
        v = ws[r][0].value
        if v:
            existing_ids.add(v)

    appended = 0
    skipped = 0
    next_row = ws.max_row + 1
    for row in NEW_ROWS:
        if row[0] in existing_ids:
            print(f'  skip {row[0]} (already in sheet)')
            skipped += 1
            continue
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)
        appended += 1
        next_row += 1

    if appended:
        wb.save(XLSX)
        print(f'Appended {appended} rows. Skipped {skipped}.')
    else:
        print(f'Nothing to append. Skipped {skipped}.')

if __name__ == '__main__':
    main()
