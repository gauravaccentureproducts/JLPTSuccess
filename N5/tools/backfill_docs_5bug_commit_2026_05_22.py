"""Back-fill Fix Commit hash for BUG-156..160 (DOCS-* 5-bug close-out)."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_backfill_docs_5bug"
if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

FIX_COMMIT = "b63bdcc"

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
for r in [159, 160, 161, 162, 163]:
    ws.cell(row=r, column=10, value=FIX_COMMIT)
    print(f"  R{r}: commit = {FIX_COMMIT}")
wb.save(XLSX)
