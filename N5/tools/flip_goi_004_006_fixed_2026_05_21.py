"""Flip BUG-133/134/135 (GOI-004/005/006) to Fixed in xlsx tracker.

Includes the 2 horizontal-deployment dokkai mojibake fixes inside the
GOI-006 lineage (JA-139 surfaced dokkai-2.11 + dokkai-3.4 as same-class).

Fix Commit cell is left as a placeholder; will be back-filled by
backfill_fix_commit_2026_05_21.py after the main commit hash is known.
"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_21_flip_goi_004_006"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]

# Header columns: A=BugID(formula) B=Date C=Reporter D=Title E=Description F=Sev G=Pri H=Status
# I=Fix Date J=Fix Commit K=Fix Notes
# Auto-detect Fix Commit / Notes column by scanning header row
hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
print("Header detected:")
for k, v in hdr.items():
    print(f"  col {v}: {k}")

status_col = hdr.get("Status", 8)
fix_date_col = hdr.get("Fix Date", 9)
fix_commit_col = hdr.get("Fix Commit", 10)
fix_notes_col = hdr.get("Fix Notes", 11)

FIXES = {
    136: {
        "status": "Fixed",
        "fix_date": "2026-05-21",
        "fix_commit": "<pending-main-commit>",
        "fix_notes": (
            "GOI-004: rewrote goi-7.6 + goi-7.7 rationale_hi to natural Hindi about their actual stems "
            "(ゆうがた/夕方 paraphrase + じょうずに 話す paraphrase). goi-7.8 was already correct. "
            "JA-137 (narrow off-by-one shift detector: 0 own + ≥2 next overlap) added to CI to prevent "
            "re-introduction. Tool: tools/fix_goi_004_006_2026_05_21.py."
        ),
    },
    137: {
        "status": "Fixed",
        "fix_date": "2026-05-21",
        "fix_commit": "<pending-main-commit>",
        "fix_notes": (
            "GOI-005: stripped fix-history / version-references / replacement-history from 7 rationale "
            "fields across goi-1.5 / goi-1.10 / goi-3.15 / goi-4.6 / goi-5.4 / goi-7.7 / goi-7.8 "
            "(+ Hindi mirrors where present). JA-121 trigger set extended with 'replaces the prior', "
            "'replaces the previous', 'previous version', 'prior version', 'Strict-N5:', 'in v1.', "
            "'policy applied at', 'no longer appears', and Hindi 'पिछले संस्करण' / 'पुराने' / 'की जगह लेता'. "
            "Tool: tools/fix_goi_004_006_2026_05_21.py."
        ),
    },
    138: {
        "status": "Fixed",
        "fix_date": "2026-05-21",
        "fix_commit": "<pending-main-commit>",
        "fix_notes": (
            "GOI-006: replaced mojibake 「あमारी ありません」 with 「あまく ありません」 in goi-7.4 rationale_hi. "
            "JA-139 (no Devanagari letter embedded inside JP-character word in rationale_hi, "
            "excluding sentence-end danda ।/॥) added to CI. Horizontal-deployment sweep surfaced "
            "2 same-class mojibake in dokkai (dokkai-2.11 'ぐらि' → 'ぐらい'; dokkai-3.4 'あमारी' → "
            "'あまり') — both fixed in this batch. Tools: tools/fix_goi_004_006_2026_05_21.py + "
            "tools/fix_dokkai_mojibake_2026_05_21.py."
        ),
    },
}

for row, spec in FIXES.items():
    ws.cell(row=row, column=status_col, value=spec["status"])
    ws.cell(row=row, column=fix_date_col, value=spec["fix_date"])
    ws.cell(row=row, column=fix_commit_col, value=spec["fix_commit"])
    ws.cell(row=row, column=fix_notes_col, value=spec["fix_notes"])
    print(f"  Row {row}: status={spec['status']} fix_date={spec['fix_date']}")

wb.save(XLSX)

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
print()
print("=== Verified state ===")
for r in [136, 137, 138]:
    print(
        f"  R{r}: status={ws2.cell(row=r, column=status_col).value} "
        f"date={ws2.cell(row=r, column=fix_date_col).value} "
        f"commit={ws2.cell(row=r, column=fix_commit_col).value}"
    )
print("Saved.")
