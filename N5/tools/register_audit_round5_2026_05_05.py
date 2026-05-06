"""Round-5 audit registrar - append ISSUE-035..047 + IMP-057..068 + Q19..Q23
to feedback/n5-audit-2026-05-04.xlsx.

Round-5 was the second run of the SALEABILITY / NICHE-FIT prompt at
prompts/N5Improvement.txt (same prompt as round-4, run against post-
v1.12.32 state). The strategic verdict shifted because round-4 closed
16 of 22 items and operationalized 3 of 4 niches; round-5 surfaced
follow-through gaps + drift introduced by the round-4 work.

Idempotent: rows whose ID already exists in the target sheet are
skipped. Mirrors row 5's formatting + extends the dropdown sqref so
the Decision column stays usable on the new rows.
"""
from __future__ import annotations
import io, sys
from copy import copy
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

ISSUES = [
    # P1
    ('ISSUE-035', 'Issue', 'MAJOR', 'P1', 'MEDIUM', 'LOW',
     'Documentation / build infra',
     'data/version.json#invariants + tools/build_version_json.py',
     '[N1] data/version.json#invariants stale ("41/41"; actual 44/44)',
     'The invariants string is hand-written in build_version_json.py rather than computed. Round-3 (JA-33) and round-4 (JA-34, JA-35) added 3 invariants without updating the source string.',
     'Same drift class as ISSUE-001 (footer version, round 1) / ISSUE-014 (README counts, round 3) / ISSUE-019 (paper-count CTA, round 3). Build stamp lies about what was verified.',
     'Have build_version_json.py read CHECKS list length from check_content_integrity.py and write the live count.',
     '-', ''),
    ('ISSUE-036', 'Issue', 'MINOR', 'P1', 'MEDIUM', 'LOW',
     'Documentation / OSS hygiene', 'repo root',
     '[N3] no CONTRIBUTING.md at repo root - institutional contributors blocked',
     'GitHub Community Standards page would mark this red. SELF-HOST.md and TRANSLATING.md cover specific paths; the GitHub-recognized "how do I contribute" template is missing.',
     'Niche-N3 institutional adopters check Community Standards before forking. First-time code contributors hit a wall.',
     'Add CONTRIBUTING.md at repo root pointing to SELF-HOST.md / TRANSLATING.md / docs/ + PR + issue conventions.',
     '-', ''),
    ('ISSUE-037', 'Issue', 'MAJOR', 'P1', 'HIGH', 'LOW',
     'UX / discoverability', 'js/home.js .syllabus-trust-band',
     '[N2] "Free / No ads / No paywall" pill missing from trust band',
     '5 pills shipped (No login / No tracking / Works offline / Open source / 100% on-device). Strongest market differentiator (price) missing - Bunpro paid, WaniKani paid, Renshuu has paid tier, Duolingo has ads.',
     'Niche N2 incomplete without the price claim. Adding one pill closes the gap.',
     'Add 6th pill "Free · No ads · No paywall" to the trust band.',
     '-', ''),
    ('ISSUE-038', 'Issue', 'MINOR', 'P1', 'MEDIUM', 'LOW',
     'SEO / discoverability', 'repo root + N5/',
     '[N3] no robots.txt / sitemap.xml - SEO + institutional indexing degrade',
     'Search-engine crawlers default to index-everything but explicit sitemap accelerates indexing + signals canonical URLs.',
     'Niche-N3 institutional adopters often have IT search-engine policies that look for sitemap.xml as a quality signal.',
     'Add N5/sitemap.xml (static, listing route URLs) + robots.txt (allow all, point at sitemap).',
     '-', ''),
    # P2
    ('ISSUE-039', 'Issue', 'MINOR', 'P2', 'MEDIUM', 'LOW',
     'UX / discoverability', 'index.html primary-nav + js/app.js ROUTES',
     '[none] #/sitting and #/missed not discoverable from primary-nav - round-3 surfaces orphaned',
     '#/sitting reachable only from #/test CTA. #/missed reachable only via direct URL or review.js link.',
     'Learners would never discover these without specifically clicking through Test / Review. Features shipped, discoverability did not.',
     'Add "Mock" link to primary-nav between Test and Progress; add "Past misses" to summary or review header.',
     '-', ''),
    ('ISSUE-040', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'Self-host / deploy', 'js/home.js trust-band href ../../LICENSE',
     '[N2] "Open source" pill href breaks on non-canonical deploys / localhost root',
     '../../LICENSE resolves correctly only from /JLPTSuccess/N5/. Localhost (http://localhost:8000/N5/) or fork at root deploy 404s.',
     'Niche N3 (self-host) breaks because institutional fork visits same UI link. N2 trust signal partially fails on localhost dev.',
     'Serve LICENSE from N5/ as well, OR change link to a stable URL (e.g., GitHub repo /blob/master/LICENSE).',
     '-', ''),
    ('ISSUE-041', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'i18n / schema', 'locales/{vi,id,ne,zh}.json _provenance + _note',
     '[N1] non-EN locale files leak _provenance + _note into i18n keyspace',
     'Round-4 added _provenance + _note for honest provenance metadata, but the i18n.t() walker treats them as resolvable keys. count_keys() reports inflated totals.',
     'Minor housekeeping; t() never asks for _provenance in practice but the leak makes the locale file key-count report misleading.',
     'Rename to nested object _meta: {provenance, note} and ignore underscore-prefixed top-level keys in the counter.',
     '-', ''),
    ('ISSUE-042', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'Content provenance', 'every item in 5 corpora',
     '[none] all 1405 items show review_status="llm_curated" - badge UI would deliver zero differentiation',
     'Round-4 ISSUE-030 scaffold means 1405/1405 items are uniformly llm_curated. A learner seeing "LLM-curated" everywhere may distrust the corpus more, not less.',
     'Round-4 trust-signal investment is on hold until at least one native-review pass lands.',
     'Either hold the badge UI render until a native-review batch ships, OR commission a small native-review pass on the top-50 items to produce a meaningful split.',
     'Q14, Q16, Q21', ''),
    ('ISSUE-043', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'Performance / first-paint', 'bundle weight',
     '[none] JS bundle 387 KB on first paint - round-3 minified CSS but not JS',
     '502 KB on first paint excluding fonts. Round-2 IMP-022 split learn.js but app.js / storage.js / furigana.js / home.js / kanji.js / kanji-popover.js are eager-loaded.',
     'Mid-tier mobile slow start. PWA install prompt fires after first paint, so the user feels slowness before installing.',
     'Add JS minification (esbuild or terser, devDep) to npm run build; lazy-load home.js / kanji.js until their routes hit.',
     '-', ''),
    # P3
    ('ISSUE-044', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'PWA / scope', 'manifest.webmanifest + sw.js',
     '[N3] PWA shortcuts use ./#/review etc. - fork to non-N5 root would need manifest path edits',
     'SW scope is ./ correct for /N5/. manifest start_url and scope both ./ but shortcut paths assume install was from /N5/. Fork deploying to root would need to update both.',
     'Niche-N3 self-host story has a quiet trap: forks have to remember to update manifest paths if they re-root.',
     'Document in SELF-HOST.md "Brand customization → manifest paths" subsection.',
     '-', ''),
    ('ISSUE-045', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'Tests', 'tests/visual-regression.spec.js',
     '[none] no visual-regression snapshots for round-3/4 surfaces',
     'Visual-regression spec covers v1.5.0 home + a few core surfaces. No snapshots for round-3 (#/missed, #/sitting, audio-player skin) or round-4 (trust band, locale chips, Accept-Language toast).',
     'Round-3 + round-4 added 5 new visible UI states; none guarded against pixel drift.',
     'Extend visual-regression.spec.js with screenshot tests: home with trust band, header with locale chips, #/missed empty, #/sitting picker, audio player skin.',
     '-', ''),
    ('ISSUE-046', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'i18n / a11y', 'js/i18n.js _flashAutoLocaleToast',
     '[N1] auto-locale toast renders English copy even when detected locale is non-EN',
     'Toast says "App language: <native-name> - change anytime in Settings." in English. Vietnamese-default user sees Vietnamese language label inside English-framed sentence.',
     'Discoverability + politeness. Toast is the very first thing a non-EN learner sees; should speak to them in their detected language.',
     'Render toast in auto-detected locale (e.g., "Ngôn ngữ ứng dụng: Tiếng Việt - thay đổi trong Cài đặt." for vi). ISSUE-026 already added the keys.',
     'ISSUE-026', ''),
    # P4
    ('ISSUE-047', 'Issue', 'MINOR', 'P4', 'LOW', 'LOW',
     'Documentation', 'README.md',
     '[N3] docs/SELF-HOST.md + docs/TRANSLATING.md not linked from main README',
     'Round-4 added the 2 docs but README does not link them. First-time visitor browsing the GitHub repo does not see them.',
     'Discoverability of self-host + contribution paths.',
     'README "## Documentation" section linking the 4 docs + spec + KnowledgeBank.',
     'ISSUE-036', ''),
]

IMPS = [
    # P1
    ('IMP-057', 'Improvement', 'IMPROVEMENT', 'P1', 'MEDIUM', 'LOW',
     'OSS hygiene', 'repo root + .github/',
     '[N3] add CODE_OF_CONDUCT.md + GitHub issue/PR templates',
     'No CODE_OF_CONDUCT.md, no .github/ISSUE_TEMPLATE/ or PULL_REQUEST_TEMPLATE.md.',
     'Best-in-class: every reasonably-mature OSS project has these. Contributor Covenant 2.1 is the de facto CoC. GitHub Community Standards checklist names them.',
     'Add Contributor Covenant 2.1 + 2 issue templates (bug, content correction) + 1 PR template.',
     '-', ''),
    # P2
    ('IMP-058', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'UX / discoverability', 'js/home.js trust band',
     '[N2] "Free / No ads / No paywall" pill (paired with ISSUE-037)',
     '5 pills, no price-claim pill.',
     'Best-in-class: DuckDuckGo anchors "Always free"; Standard Notes ships "Free, forever" badge alongside privacy. gov.uk services prominently say "Free, no account needed".',
     'Add 6th pill "Free · No ads · No paywall".',
     'ISSUE-037', ''),
    ('IMP-059', 'Improvement', 'IMPROVEMENT', 'P2', 'LOW', 'LOW',
     'i18n / schema', 'locale files',
     '[N1] move _provenance + _note under _meta object (paired with ISSUE-041)',
     'Underscore-prefixed top-level keys leak into the i18n keyspace.',
     'Best-in-class: gettext convention reserves #: comments for translator notes; i18next uses _meta keys.',
     'Nest provenance + note under _meta so the i18n key counter excludes them.',
     'ISSUE-041', ''),
    ('IMP-060', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'Security / OSS hygiene', '.github/dependabot.yml + .github/workflows/',
     '[N3] add Dependabot + GitHub Actions security audit',
     'No Dependabot config. devDependencies (Playwright, axe-core) are version-pinned but no automated security PRs.',
     'Best-in-class: every modern OSS repo runs Dependabot. GitHub free tier covers npm + GitHub Actions.',
     '5-line dependabot.yml watching npm weekly.',
     '-', ''),
    ('IMP-061', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'UX', 'primary-nav + index.html',
     '[none] add Mock + Past-misses links to primary-nav (paired with ISSUE-039)',
     'Round-3 surfaces orphaned.',
     'Best-in-class: Bunpro has every major feature in primary-nav. Hidden routes are wasted features.',
     'Add "Mock" (linking to #/sitting) to primary-nav between Test and Progress; add "Past misses" to summary or review header.',
     'ISSUE-039', ''),
    # P3
    ('IMP-062', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'Content authoring tools', 'tools/ + npm scripts',
     '[N4] unified npm run build that runs integrity check + version + min CSS',
     '100+ Python tools/fix_*.py scripts; npm run build runs build:css + build:version but check_content_integrity.py is not wired in.',
     'Best-in-class: MDN content repo has single yarn build that does everything; CI runs it. JLPTSuccess can match.',
     'Extend npm run build to chain integrity + version + min CSS. Fail the build on integrity failures.',
     '-', ''),
    ('IMP-063', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'PWA / install', 'manifest.webmanifest',
     '[N2] add share_target so OS Share sheet sees JLPTSuccess as Japanese-text target',
     'Round-3 IMP-040 added 3 PWA shortcuts. No share_target, display_override, protocol_handlers.',
     'Best-in-class: Pocket uses share_target for "send to read later"; DuckDuckGo PWA uses protocol_handlers.',
     'Add share_target so the OS Share sheet sees JLPTSuccess as a target for Japanese text → search.',
     '-', ''),
    ('IMP-064', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'Onboarding / discoverability', 'js/app.js first-run logic',
     '[N1] insert locale-confirmation step before diagnostic on first run',
     'Round-3 IMP-044 routes fresh users to #/diagnostic. Diagnostic does not acknowledge locale or surface locale chips. Vietnamese-default learner gets routed to placement test before being told they can switch UI language.',
     'Best-in-class: Duolingo asks for native-language preference before any test. Anki is locale-aware on first launch.',
     'Insert 1-step locale-confirmation between auto-detect toast and diagnostic.',
     'ISSUE-046', ''),
    ('IMP-065', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'Tests', 'tests/visual-regression.spec.js',
     '[none] add visual-regression snapshots for round-3/4 surfaces (paired with ISSUE-045)',
     'No snapshots for new round-3 + round-4 UI states.',
     'Best-in-class: GitHub runs visual-regression on every PR. Round-3 + round-4 didn\'t add snapshots.',
     'Extend with 5 new screenshot snapshots.',
     'ISSUE-045', ''),
    # P4
    ('IMP-066', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'MEDIUM',
     'Distribution / discoverability', 'GitHub repo settings',
     '[N3] populate repo description + topics + homepage + releases',
     'Repo description / topics / homepage URL not set; no GitHub releases page populated.',
     'Best-in-class: freeCodeCamp, MDN, Tofugu repos have polished GitHub-side metadata.',
     'Set repo description, add 5-7 topics (jlpt, n5, japanese, pwa, multilingual, privacy, open-source), set homepage to live URL, populate releases per major version.',
     '-', ''),
    ('IMP-067', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'Performance', 'image assets',
     '[none] add WebP / AVIF variants of icon-192 / icon-512 / mark.png',
     'assets/logo/ ships PNG icons. No WebP / AVIF for size savings.',
     'Best-in-class: 2025-era PWAs ship AVIF for icons; ~30-50% smaller than PNG.',
     'Add WebP siblings to PNG icons; manifest entries can list both.',
     '-', ''),
    # P5
    ('IMP-068', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'HIGH',
     'Content / community', 'integration with translation platforms',
     '[N1] investigate Crowdin / Weblate integration for contributor-friendly translation',
     'No Crowdin / Weblate integration. Round-4 docs/TRANSLATING.md asks for direct PRs to locale JSONs.',
     'Best-in-class: Mozilla uses Pontoon; KDE uses Weblate. Free for OSS.',
     'Investigate Weblate for a contributor-friendly translation UI.',
     'scale of contributor activity', ''),
]

QUESTIONS = [
    ('Q19', 'tools/build_version_json.py invariants source',
     'Current data/version.json#invariants field is hand-written ("41/41"). Stale by 3 since round-3 added JA-33 and round-4 added JA-34 + JA-35. Hand-written placeholder doesn\'t scale.',
     'Should build_version_json.py execute check_content_integrity.py and write the actual count, or is a hand-written placeholder acceptable as long as a release-pass updates it?',
     ''),
    ('Q20', 'Native-review staffing for the multilingual stretch',
     'Round-4 close-out machine-translated 4 locales but explicitly punted native review to "Q14 / Q16 staffing." This is the same blocker, different round. No progress without staffing.',
     'Should the project actively recruit per-locale reviewers (GitHub issue + Reddit + Discord shout-out), or wait passively?',
     ''),
    ('Q21', 'Provenance badge UI launch threshold',
     'ISSUE-042 raised: rendering review_status badge on every item when 100% are llm_curated looks bad - trust signal that delivers no differentiation.',
     'Decision: (a) ship badge with explicit "all currently LLM-curated; native review in progress" hero text, (b) wait until ≥X% are native_reviewed, (c) skip badge UI entirely and use the data only for internal filters.',
     ''),
    ('Q22', 'Search/SEO commitment',
     'Round-4 added og: + JSON-LD. Going further (sitemap.xml, robots.txt, Show HN submission, Dev.to blog post) requires actual marketing investment.',
     'Is there interest in active SEO / marketing, or is the project intentionally low-discoverability?',
     ''),
    ('Q23', 'Round-2/3/4 long-deferred questions resolved as a batch',
     'Q4/Q6/Q8/Q11/Q12/Q13-18 still partially open. Several (Q8, Q14, Q11, Q16) are functionally the same question - "are we hiring / commissioning?".',
     'Worth a single product-owner sit-down to resolve all open Qs as a batch?',
     ''),
]


def _copy_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.number_format = src.number_format


def main() -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.cell_range import CellRange
    except ImportError:
        print('ERROR: openpyxl not installed.')
        return 1
    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)
    ws = wb['Audit findings']
    existing = {str(r[0].value).strip() for r in ws.iter_rows(min_row=5)
                if r and r[0].value}
    new_rows = [r for r in ISSUES + IMPS if r[0] not in existing]
    next_row = ws.max_row + 1
    ref = 5
    for r in new_rows:
        for col_offset, val in enumerate(r):
            ws.cell(row=next_row, column=col_offset + 1, value=val)
        for c in range(1, ws.max_column + 1):
            _copy_style(ws.cell(row=ref, column=c),
                        ws.cell(row=next_row, column=c))
        next_row += 1
    last = ws.max_row
    for dv in ws.data_validations.dataValidation:
        if dv.type != 'list': continue
        for cr in list(dv.sqref.ranges):
            if str(cr).startswith('N'):
                dv.sqref.add(CellRange(f'N5:N{last}'))
                break

    ws_q = wb['Open questions']
    existing_q = {str(r[0].value).strip() for r in ws_q.iter_rows(min_row=4)
                  if r and r[0].value}
    new_qs = [r for r in QUESTIONS if r[0] not in existing_q]
    next_q = ws_q.max_row + 1
    refq = 4
    for r in new_qs:
        for col_offset, val in enumerate(r):
            ws_q.cell(row=next_q, column=col_offset + 1, value=val)
        for c in range(1, ws_q.max_column + 1):
            _copy_style(ws_q.cell(row=refq, column=c),
                        ws_q.cell(row=next_q, column=c))
        next_q += 1
    last_q = ws_q.max_row
    for dv in ws_q.data_validations.dataValidation:
        if dv.type != 'list': continue
        for cr in list(dv.sqref.ranges):
            if str(cr).startswith('E'):
                dv.sqref.add(CellRange(f'E4:E{last_q}'))
                break

    wb.save(XLSX)
    print(f'Audit findings:  appended {len(new_rows)} rows '
          f'({sum(1 for r in new_rows if r[0].startswith("ISSUE-"))} issues + '
          f'{sum(1 for r in new_rows if r[0].startswith("IMP-"))} improvements)')
    if existing & {r[0] for r in ISSUES + IMPS}:
        skipped = sorted(existing & {r[0] for r in ISSUES + IMPS})
        print(f'  skipped: {", ".join(skipped)}')
    print(f'Open questions:  appended {len(new_qs)} rows')
    if existing_q & {r[0] for r in QUESTIONS}:
        skipped_q = sorted(existing_q & {r[0] for r in QUESTIONS})
        print(f'  skipped: {", ".join(skipped_q)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
