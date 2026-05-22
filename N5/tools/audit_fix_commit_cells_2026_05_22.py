"""Survey existing Fix Commit cell shapes to design JA-146 correctly.

Surfaces:
 - hash-shaped values (the expected good case)
 - hash + (+ submodule HASH) form (the file-based commit pattern's output)
 - date strings (the DOCS-VOCAB-003 bug we want to catch)
 - datetime objects (Excel auto-coerced!)
 - explicit <pending-X> sentinels (transient OK)
 - other (free text — should generally fail)
 - empty
"""
import sys, io, re, datetime as _dt
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook("specifications/test-scenarios-by-specialist-perspective.xlsx")
ws = wb["User Reported Bugs"]

HASH_RE = re.compile(r"^[a-f0-9]{7,40}(\s+\(\+\s+submodule\s+[a-f0-9]{7,40}\))?$")
DATE_STR_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
SENTINEL_RE = re.compile(r"^<[^>]+>$")

shapes = Counter()
problems = []
for r in range(4, ws.max_row + 1):
    title = ws.cell(row=r, column=4).value
    status = ws.cell(row=r, column=8).value
    val = ws.cell(row=r, column=10).value
    if not title or status != "Fixed":
        continue

    if val is None or (isinstance(val, str) and not val.strip()):
        shapes["empty"] += 1
        problems.append((r, "EMPTY", title[:60] if title else ""))
        continue

    if isinstance(val, (_dt.datetime, _dt.date)):
        shapes["datetime_object"] += 1
        problems.append((r, f"DATETIME_OBJECT: {val!r}", title[:60] if title else ""))
        continue

    s = str(val).strip()
    if HASH_RE.match(s):
        shapes["hash"] += 1
    elif DATE_STR_RE.match(s):
        shapes["date_string"] += 1
        problems.append((r, f"DATE_STRING: {s!r}", title[:60] if title else ""))
    elif SENTINEL_RE.match(s):
        shapes["sentinel"] += 1
    elif re.match(r"^[a-f0-9]{7,40}\b", s):
        shapes["hash_prefix_with_trailing"] += 1
        problems.append((r, f"HASH+TRAILING: {s[:80]!r}", title[:60] if title else ""))
    else:
        shapes["other"] += 1
        problems.append((r, f"OTHER: {s[:80]!r}", title[:60] if title else ""))

print("Shape distribution across Fixed rows:")
for k, v in shapes.most_common():
    print(f"  {k}: {v}")
print()
print(f"Problematic rows ({len(problems)}):")
for r, shape, title in problems[:30]:
    print(f"  R{r} {shape}: {title}")
if len(problems) > 30:
    print(f"  ... and {len(problems)-30} more")
