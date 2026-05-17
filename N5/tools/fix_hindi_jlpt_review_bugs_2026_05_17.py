"""Apply native-Hindi-teacher + JLPT-exam-expert review findings:
file bugs + apply fixes for the 4 actionable findings surfaced
during the 2026-05-17 review of C-* and B-* scenarios.

Findings (sourced from native-Hindi-teacher + JLPT-exam-specialist
review, cross-checked against Hindi Vyakaran grammar references +
JEES official JLPT N5 sample papers):

  NR-HI-001 (Critical / P1) — q-0264 distractor とって corruption
    distractor_explanations_hi[とって] = "जो's てください (please DO)。"
    The 'जो's' is a Hindi+English-possessive corruption with no
    semantic content. Per Genki I L11 + the intended pedagogy
    (て-form + ください = polite request), the canonical Hindi
    should be: "とって — とる का て-रूप; ください के साथ मिलकर
    विनम्र अनुरोध बनाता है।".

  NR-HI-002 (Major / P2) — q-0186 explanation_hi possessive corruption
    "प्राप्तकर्ता's दृष्टिकोण" uses English possessive 's after
    Hindi noun. Correct Hindi possessive: "प्राप्तकर्ता का दृष्टिकोण".

  NR-HI-003 (Medium / P3) — q-0065 explanation_hi English "Group 1"
    Uses "for Group 1" in a Hindi explanation. Other questions use
    "समूह-1 / समूह-2" consistently per the C-005 terminology lock.
    Fix: standardize to "समूह-1".

  NR-JE-001 (Major / P2) — JLPT format violations in bunpou papers
    Per JEES official sample N5 sentence-construction (bunpou
    mondai 2) format:
    (a) 80 stems in bunpou-5.* and bunpou-7.* use half-width
        underscores '___' instead of canonical full-width '＿' for
        the fill-blank slots.
    (b) 10 stems in bunpou-7.* end with "→ [N]番" without terminal
        punctuation (。/？/！). Per JEES samples, even arrow-
        framed answer-key-stems carry terminal 。.

Run from N5/:
    python tools/fix_hindi_jlpt_review_bugs_2026_05_17.py
"""
from __future__ import annotations

import glob
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"
QUESTIONS_JSON = ROOT / "data" / "questions.json"


# ---------- Fix application ----------

def fix_questions_json() -> int:
    """Fix the 3 Hindi-side findings in questions.json."""
    q = json.loads(QUESTIONS_JSON.read_text(encoding="utf-8"))
    n_fixed = 0
    for question in q.get("questions", []):
        qid = question.get("id")
        # NR-HI-001 q-0264
        if qid == "q-0264":
            de = question.get("distractor_explanations_hi") or {}
            if isinstance(de, dict) and "とって" in de:
                old = de["とって"]
                if "जो's" in old or "जो's" in old:
                    de["とって"] = (
                        "とって — とる का て-रूप; ください के साथ मिलकर "
                        "विनम्र अनुरोध बनाता है। यहाँ नकारात्मक रूप "
                        "अपेक्षित है (とらない), अनुरोध नहीं।"
                    )
                    n_fixed += 1
                    print(f"  q-0264 distractor とって: {old!r} -> {de['とって']!r}")
        # NR-HI-002 q-0186
        if qid == "q-0186":
            old = question.get("explanation_hi") or ""
            if "प्राप्तकर्ता's" in old:
                new = old.replace("प्राप्तकर्ता's दृष्टिकोण",
                                  "प्राप्तकर्ता का दृष्टिकोण")
                question["explanation_hi"] = new
                n_fixed += 1
                print(f"  q-0186 explanation_hi: 's possessive -> का "
                      f"({len(old) - len(new)} chars diff)")
        # NR-HI-003 q-0065
        if qid == "q-0065":
            old = question.get("explanation_hi") or ""
            if "Group 1" in old or "for Group 1" in old:
                new = (old
                       .replace("for Group 1", "समूह-1 के लिए")
                       .replace("Group 1", "समूह-1")
                       .replace("Group 2", "समूह-2"))
                question["explanation_hi"] = new
                n_fixed += 1
                print(f"  q-0065 explanation_hi: 'Group 1' -> 'समूह-1'")
    QUESTIONS_JSON.write_text(
        json.dumps(q, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return n_fixed


def fix_jlpt_stems_in_papers() -> tuple[int, int]:
    """Fix the 2 JLPT-format findings in data/papers/**/*.json.

    Returns (n_underscore_fixes, n_terminal_punct_fixes).
    """
    n_us = 0
    n_term = 0
    for path in sorted(glob.glob(str(ROOT / "data" / "papers" / "**" / "*.json"),
                                  recursive=True)):
        d = json.loads(Path(path).read_text(encoding="utf-8"))
        changed = False
        for q in d.get("questions", []):
            stem = q.get("stem_html") or ""
            if not stem:
                continue
            new_stem = stem
            # (a) Half-width ___ -> full-width ＿
            # Specifically replace 3-underscore tokens that are JLPT
            # mondai-2 fill-blank slots
            if "___" in new_stem:
                # The canonical JLPT mondai-2 format is 4 slots with ★
                # marking the question slot. Replace `___` (3 ASCII
                # underscores) with `＿＿＿＿` (4 fullwidth underscores)
                # to match the JEES convention.
                new_stem = new_stem.replace("___", "＿＿＿＿")
                n_us += 1
            # (b) Stems ending with "→ [N]番" without terminal punctuation
            stem_clean_strip = re.sub(r'<[^>]+>', '', new_stem).rstrip()
            if (stem_clean_strip.endswith("番")
                    and re.search(r'\[\d+\]\s*番$', stem_clean_strip)):
                new_stem = new_stem.rstrip() + "。"
                n_term += 1
            if new_stem != stem:
                q["stem_html"] = new_stem
                changed = True
        if changed:
            Path(path).write_text(
                json.dumps(d, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
    return n_us, n_term


# ---------- Bug filing ----------

BUGS = [
    {
        "title": "NR-HI-001 — q-0264 distractor とって Hindi explanation corruption (\"जो's てください\")",
        "description": (
            "Source: native-Hindi-teacher review run 2026-05-17 of "
            "C. Hindi locale tab (C-001 + C-002 + C-018 scenarios).\n\n"
            "q-0264 distractor_explanations_hi['とって'] = "
            "\"जो's てください (please DO)。\" — the 'जो's' is a "
            "Hindi+English-possessive corruption that has no semantic "
            "content. The English possessive 's was emitted after a "
            "Hindi word (जो = 'that/which') with no Hindi grammatical "
            "function.\n\n"
            "Expected meaning per Genki I L11 + the pedagogy this "
            "distractor explains: とって is the て-form of とる; "
            "combined with ください it forms a polite request. The "
            "correct distractor for this question (which asks for the "
            "negative form とらない) should explain that とって is a "
            "request form, not a negation.\n\n"
            "[FIX 2026-05-17]: Replaced with natural Hindi: "
            "\"とって — とる का て-रूप; ください के साथ मिलकर विनम्र "
            "अनुरोध बनाता है। यहाँ नकारात्मक रूप अपेक्षित है "
            "(とらない), अनुरोध नहीं।\""
        ),
        "severity": "Critical",
        "priority": "P1",
    },
    {
        "title": "NR-HI-002 — q-0186 explanation_hi English-possessive 's intrusion (प्राप्तकर्ता's)",
        "description": (
            "Source: native-Hindi-teacher review run 2026-05-17.\n\n"
            "q-0186 explanation_hi contains the phrase "
            "\"प्राप्तकर्ता's दृष्टिकोण\" which is grammatically incorrect "
            "Hindi — the English possessive 's was attached to a Hindi "
            "noun. The correct Hindi possessive uses the postposition "
            "का/के/की: \"प्राप्तकर्ता का दृष्टिकोण\" (recipient's "
            "perspective).\n\n"
            "Per Hindi Vyakaran (Kamta Prasad Guru) + Sahitya Akademi "
            "Hindi style conventions: English possessive 's is never "
            "appropriate in formal Hindi prose. The Hindi case-marker "
            "का (singular masculine) / के (singular oblique or plural) "
            "/ की (singular feminine) is the canonical possessive marker.\n\n"
            "[FIX 2026-05-17]: Replaced \"प्राप्तकर्ता's दृष्टिकोण\" "
            "with \"प्राप्तकर्ता का दृष्टिकोण\"."
        ),
        "severity": "Major",
        "priority": "P2",
    },
    {
        "title": "NR-HI-003 — q-0065 explanation_hi uses English \"Group 1\" instead of canonical समूह-1",
        "description": (
            "Source: native-Hindi-teacher review run 2026-05-17 / "
            "C-005 verb-class naming consistency scenario.\n\n"
            "q-0065 explanation_hi uses English \"for Group 1\" in the "
            "middle of a Hindi explanation about godan verb negation. "
            "Per the corpus-wide terminology lock (4× समूह-1 + 4× समूह-2 "
            "in grammar.json + 2× समूह-2 in other questions.json "
            "entries), the canonical Hindi term for godan/ichidan verb "
            "classes is समूह-1 / समूह-2.\n\n"
            "Mixing English \"Group 1\" mid-Hindi-explanation breaks the "
            "terminology consistency that C-005 enforces.\n\n"
            "[FIX 2026-05-17]: Replaced \"for Group 1\" with "
            "\"समूह-1 के लिए\" and \"Group 1\" → \"समूह-1\" / "
            "\"Group 2\" → \"समूह-2\" globally in the field."
        ),
        "severity": "Medium",
        "priority": "P3",
    },
    {
        "title": "NR-JE-001 — JLPT format violations: bunpou stems use half-width '___' (should be full-width '＿') + missing terminal punctuation",
        "description": (
            "Source: JLPT-exam-expert review run 2026-05-17 of B. JLPT "
            "format tab (B-017 + B-018 scenarios).\n\n"
            "Per JEES official JLPT N5 sample paper format conventions:\n"
            "  (a) Sentence-construction mondai (bunpou mondai 2) uses "
            "      FULL-WIDTH underscore '＿' (U+FF3F) for fill-blank "
            "      slots, NOT half-width '_' (U+005F). The canonical "
            "      JEES format shows four blanks per stem: "
            "      '＿＿＿＿ ＿＿＿＿ ★ ＿＿＿＿ ＿＿＿＿'.\n"
            "  (b) Every stem ends with terminal punctuation "
            "      (。 / ？ / ！) — including answer-key-stems that "
            "      use arrow notation '→ [N]番'.\n\n"
            "Findings:\n"
            "  - 80 stems across bunpou-5.* and bunpou-7.* papers use "
            "    '___' (three ASCII underscores) instead of '＿' "
            "    (full-width).\n"
            "  - 10 stems in bunpou-7.* end with '→ [N]番' without "
            "    terminal 。.\n\n"
            "Reference: JEES official 2010 JLPT N5 sample paper, "
            "bunpou-dokkai section, mondai 2 (sentence construction).\n\n"
            "[FIX 2026-05-17]: Replaced '___' with '＿＿＿＿' (4 "
            "full-width underscores per slot, matching JEES convention) "
            "and added terminal 。 to the 10 arrow-framed stems."
        ),
        "severity": "Major",
        "priority": "P2",
    },
]


def append_bugs_to_xlsx(commit_placeholder: str) -> int:
    """Append the 4 native-Hindi + JLPT-expert bugs to User Reported Bugs sheet."""
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["User Reported Bugs"]
    # Find last data row
    last_row = ws.max_row
    while last_row >= 4 and not ws.cell(row=last_row, column=4).value:
        last_row -= 1
    next_row = last_row + 1
    n_added = 0
    today = datetime(2026, 5, 17)
    for bug in BUGS:
        ws.cell(row=next_row, column=2).value = today
        # Reporter column 3 — pick appropriate
        if bug["title"].startswith("NR-HI"):
            ws.cell(row=next_row, column=3).value = "Native-Hindi-teacher review (2026-05-17)"
        else:
            ws.cell(row=next_row, column=3).value = "JLPT exam-expert review (2026-05-17)"
        ws.cell(row=next_row, column=4).value = bug["title"]
        ws.cell(row=next_row, column=5).value = bug["description"]
        ws.cell(row=next_row, column=6).value = bug["severity"]
        ws.cell(row=next_row, column=7).value = bug["priority"]
        ws.cell(row=next_row, column=8).value = "Fixed"
        # Fix Commit at col 9 (1-based), Fix Date at col 10
        ws.cell(row=next_row, column=9).value = commit_placeholder
        ws.cell(row=next_row, column=10).value = today
        next_row += 1
        n_added += 1
    wb.save(str(XLSX))
    return n_added


def main() -> int:
    print("=== Fixing questions.json (Hindi findings) ===")
    n_q = fix_questions_json()
    print(f"  Total fixes: {n_q}")
    print()
    print("=== Fixing papers/*.json (JLPT format findings) ===")
    n_us, n_term = fix_jlpt_stems_in_papers()
    print(f"  Half-width underscore fixes: {n_us}")
    print(f"  Terminal punctuation fixes: {n_term}")
    print()
    print("=== Appending bugs to User Reported Bugs sheet ===")
    n_bugs = append_bugs_to_xlsx("(pending — this commit)")
    print(f"  Bugs filed: {n_bugs}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
