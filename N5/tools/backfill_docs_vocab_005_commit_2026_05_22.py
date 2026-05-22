"""Back-fill Fix Commit hash for BUG-136 (DOCS-VOCAB-005)."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_backfill_docs_vocab_005"

if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

FIX_COMMIT = "b7f5787"      # main repo
SUB_COMMIT = "426c91f"      # submodule (procedure manual F.41)
COMBINED = f"{FIX_COMMIT} (+ submodule {SUB_COMMIT})"

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
prev = ws.cell(row=139, column=10).value
ws.cell(row=139, column=10, value=COMBINED)
print(f"  R139: {prev} -> {COMBINED}")

wb.save(XLSX)

# Verify
wb2 = load_workbook(XLSX)
ws2 = wb2["User Reported Bugs"]
print()
print("=== Verified ===")
print(f"  R139: status={ws2.cell(row=139, column=8).value} commit={ws2.cell(row=139, column=10).value}")
