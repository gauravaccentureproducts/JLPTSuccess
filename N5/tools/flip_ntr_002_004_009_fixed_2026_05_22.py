"""Flip BUG-162/163/164/169 (NTR-002/003/004/009) to Fixed."""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_flip_ntr_002_004_009"
if not os.path.exists(BAK): shutil.copy2(XLSX, BAK)
wb = load_workbook(XLSX); ws = wb["User Reported Bugs"]

FLIPS = {
    165: ("NTR-002 — Marked n5-045 with deprecated=True + kept _alias_of='n5-017'. "
          "Cleared the reverse _alias_of from n5-017 (canonical entry no longer needs to "
          "point at a deprecated alias). External references (questions.json, audio_manifest, "
          "etc.) still resolve to n5-045 ID for backward compatibility; UI grammar TOC should "
          "filter deprecated entries (JS filter follow-up; current 'Also see' alias-badge "
          "rendering handles cross-link from n5-045 detail to n5-017 canonical). Tool: "
          "tools/fix_ntr_002_003_004_009_2026_05_22.py."),
    166: ("NTR-003 — Flipped かれ/かのじょ glosses to make boyfriend/girlfriend the "
          "PRIMARY sense (matches modern spoken Japanese) and the third-person pronoun "
          "the marked formal/literary sense. New gloss reads: "
          "'boyfriend (primary); he, him (third-person pronoun, more formal/literary)'. "
          "Provenance bumped to native_reviewed_2026_05_22."),
    167: ("NTR-004 — Added usage_note to あなた entry: 'あなた is distant or formal, or "
          "(from a wife to a husband) intimate. It is NOT the normal way to address someone "
          "you know — use their name + さん, or drop the subject. Appropriate contexts: "
          "form-filling, wedding vows, lyrics, addressing strangers respectfully.' Replaced "
          "example[0] from 'あなたは どなたですか。' (survey-form register) to "
          "'田中さんは どこから 来ましたか。' (name-based-address alternative). Gloss now "
          "carries the caveat. Hindi usage_note_hi mirrored. Provenance "
          "native_reviewed_2026_05_22."),
    172: ("NTR-009 — Hindi glosses for かれ/かのじょ synced with the English re-gloss "
          "(NTR-003 same-commit fix). New gloss_hi: "
          "'बॉयफ्रेंड (प्राथमिक अर्थ); वह (पुरुष, औपचारिक/साहित्यिक सर्वनाम के रूप में)' (parallel "
          "for かのじょ). Both languages now carry the boyfriend/girlfriend-as-primary nuance."),
}
for row, note in FLIPS.items():
    ws.cell(row=row, column=8, value="Fixed")
    ws.cell(row=row, column=9, value="2026-05-22")
    ws.cell(row=row, column=10, value="<pending-main-commit>")
    ws.cell(row=row, column=11, value=note)
    print(f"  R{row}: Fixed")
wb.save(XLSX)
