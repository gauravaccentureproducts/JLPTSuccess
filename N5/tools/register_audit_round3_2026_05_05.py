"""Round-3 audit registrar — append new ISSUE-013..024 + IMP-030..044 + Q8..Q12
to feedback/n5-audit-2026-05-04.xlsx.

The user ran the audit-only prompt at prompts/N5Improvement.txt. The audit
produced 12 new Issues + 15 new Improvements + 5 new Open Questions.
This script appends them to the existing tracker so round-3 close-out
follows the same xlsx-driven workflow as round-1 and round-2.

Idempotent: rows whose ID already exists in the target sheet are skipped.

Schema (Audit findings sheet, row 4):
  ID, Type, Severity, Priority, Impact, Effort, Category, Location,
  Title, Current state, Why this matters / Best-in-class, Suggested direction,
  Dependencies, Decision (Fix / Avoid / Defer)

Schema (Open questions sheet, row 3):
  ID, Topic, Context, Decision needed, Decision (Fix / Avoid / Defer)
"""
from __future__ import annotations
import io, sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

# 14-column rows for "Audit findings". Decision left blank so the user
# stamps Fix / Avoid / Defer per item per the round-1+2 workflow.
ISSUES = [
    # P1
    ('ISSUE-013', 'Issue', 'MAJOR', 'P1', 'HIGH', 'HIGH',
     'Content correctness · Kanji',
     'data/kanji.json (entries[*].additional_readings); js/kanji-popover.js:68-70',
     'kanji additional_readings populated on 1/106 entries — popover renders empty "Other readings" block for 105 kanji',
     '1/106 kanji have additional_readings.{on,kun}; popover (ISSUE-008 wiring v1.12.29) shows an empty disclosure for the other 105.',
     'Producer-consumer drift: IMP-015 declared the field, no authoring pass populated it. Learner sees a feature that delivers nothing.',
     'Author non-N5 on/kun readings for the 105 entries from KanjiDic2 / Joyo readings into additional_readings.{on,kun}.',
     '—', ''),
    ('ISSUE-014', 'Issue', 'MAJOR', 'P1', 'MEDIUM', 'LOW',
     'Documentation', 'README.md (top "Content scale" section)',
     'README counts stale (177 grammar / 1003 vocab / 60 reading+listening)',
     'README quotes 177/1003/106/60. Actual: 178/1041/106/40+40=80.',
     'Same drift class as ISSUE-001 footer-version that round-1 closed. Erodes trust in other claims (privacy, no-network).',
     'Parse JSON at doc-build time or wire a check_doc_consistency invariant that fails CI on README/data divergence.',
     '—', ''),
    ('ISSUE-015', 'Issue', 'MINOR', 'P1', 'LOW', 'LOW',
     'Documentation', 'README.md "Deploy to GitHub Pages" section',
     'README GH-Pages URL says "/JLPT/N5/" — actual deploy is "/JLPTSuccess/N5/"',
     'Path segment "JLPT" is from the pre-monorepo era; the canonical deploy moved to gauravaccentureproducts.github.io/JLPTSuccess/N5/.',
     'Misleading for anyone forking the repo trying to reproduce the deploy.',
     'Replace path with the canonical URL pattern; note the JLPTSuccess monorepo layout.',
     '—', ''),
    ('ISSUE-016', 'Issue', 'MAJOR', 'P1', 'HIGH', 'MEDIUM',
     'Content correctness · Listening', 'data/listening.json items[*]',
     'listening.json has no mondai taxonomy — UI cannot surface the official 課題理解/ポイント理解/発話表現/即時応答 sections',
     '40/40 items have a free-form `format` string; 0/40 carry mondai (1/2/3/4) or a closed format_type enum.',
     'Without the mondai tag, mock-test cannot mirror official JLPT N5 listening structure (mondai 1+2+3 are 4-choice; mondai 4 is 3-choice).',
     'Add mondai (1|2|3|4) and format_type (closed enum) per item; add invariant guarding the closed enum and choice-count.',
     '—', ''),
    # P2
    ('ISSUE-017', 'Issue', 'MINOR', 'P2', 'MEDIUM', 'LOW',
     'Corpus structural · Answer-position',
     'data/papers/goi/paper-1.json, data/papers/moji/paper-1.json',
     'goi/paper-1 and moji/paper-1 answer-position skew 8/2/3/2 (spread 6)',
     'Global paper distribution is 25.1/25.1/24.9/24.9% — these two papers concentrate choice-3. Round-2 IMP-026 covered bunpou but missed these.',
     'Learner gaming: "always pick D" gets ~53% on these papers; defeats face-validity of "audited paper".',
     'Rebalance by re-ordering choice arrays on 4-5 items in each skewed paper; correctIndex updates in lock-step.',
     '—', ''),
    ('ISSUE-018', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'Content correctness · Listening',
     'data/listening.json (5 items with 3-choice arrays)',
     '5/40 listening items have 3-choice arrays — could be official mondai 4 or authoring drift; cannot verify without mondai tag',
     'Without ISSUE-016 mondai tag, the 3-choice items cannot be confirmed as the legitimate mondai-4 format.',
     'Format-authenticity claim hinges on this; either valid (mondai 4) or content bug.',
     'Blocked by ISSUE-016. Once mondai is tagged, invariant: mondai==4 ⇒ len(choices)==3 else 4.',
     'ISSUE-016', ''),
    ('ISSUE-019', 'Issue', 'MINOR', 'P2', 'LOW', 'LOW',
     'Documentation', 'js/test.js:117 (UI string)',
     'Test setup says "25 papers across 4 sections" — actual is 28 (7 × 4)',
     'Hardcoded UI string. Same drift class as ISSUE-014.',
     'User-visible UI string drifts from data; same trust-erosion risk.',
     'Replace 25 with live count read from data/papers/manifest.json.',
     '—', ''),
    ('ISSUE-020', 'Issue', 'MAJOR', 'P2', 'MEDIUM', 'MEDIUM',
     'App correctness · Test mode',
     'js/test.js (no full-paper sitting flow)',
     'No full-paper sitting that mimics official Moji-Goi 25min + Bunpou-Dokkai 50min + Choukai 30min',
     'Length-picker samples grammar only. /papers does single-section sittings. The 25/50/30 budgets are mentioned in the test.js header comment but never composed at runtime.',
     'Test-mode realism vs official JLPT format — Section-1 audit criterion. Learner preparing for the actual sitting needs the timing + section transitions.',
     'Add "Full mock paper" mode chaining moji→goi→bunpou→dokkai→listening with per-section timer.',
     'ISSUE-016', ''),
    # P3
    ('ISSUE-021', 'Issue', 'MINOR', 'P3', 'LOW', 'MEDIUM',
     'Content correctness · Kanji',
     'data/kanji.json (15 entries with empty kun, 1 with empty on)',
     '15 kanji have empty kun reading and 1 has empty on — intentional vs missing is undocumented',
     'Mostly correct linguistically (numerals 万/千/百, etc. have no kun) but no flag distinguishing "intentionally absent" from "missing data".',
     'Popover renders blank lines for these entries; learner cannot tell whether field was authored or skipped.',
     'Add kun_intentional_absence boolean (or render "(none in N5 syllabus)" instead of blank).',
     '—', ''),
    ('ISSUE-022', 'Issue', 'MAJOR', 'P3', 'MEDIUM', 'HIGH',
     'i18n / accessibility', 'locales/{en,id,ne,vi,zh}.json (47 keys each)',
     'Locale files cover only ~6% of UI strings — most JS modules hardcode English',
     '5 locale files × ~47 keys; runtime UI has hundreds of literal English strings hardcoded across js/*.js.',
     'Multi-locale footprint is a public-facing claim (5 manifest files). Learner experience for non-English speakers is ~94% English.',
     'Either extract every literal string to keys + translate, or remove the 4 non-English files. See Q8.',
     'Q8', ''),
    ('ISSUE-023', 'Issue', 'MINOR', 'P3', 'LOW', 'LOW',
     'Documentation · KnowledgeBank', 'README.md KnowledgeBank section',
     'README per-file Q-counts (100 Qs each) match data but per-paper layout (15+15+15+15+15+15+10) is undocumented',
     'Last paper in each section has 10Q while others have 15Q. README does not explain why.',
     'Implementer ergonomics — future content authors might "fix" the apparent imbalance.',
     'Short note in README explaining the 7-paper layout (6 full + 1 short).',
     '—', ''),
    # P4
    ('ISSUE-024', 'Issue', 'MINOR', 'P4', 'LOW', 'MEDIUM',
     'PWA / service worker', 'sw.js:21 (CACHE_VERSION constant)',
     'CACHE_VERSION is hardcoded — same regression risk as the v1.10→v1.12 footer drift round-1 closed',
     'sw.js bumps a literal string each release. ISSUE-001 auto-derived footer version from CHANGELOG; sw.js did not get the same treatment.',
     'A release that forgets the bump silently serves a stale shell to returning users.',
     'Parse the most recent ## v heading in CHANGELOG.md at install-event time; use as CACHE_VERSION.',
     '—', ''),
]

IMPS = [
    # P1
    ('IMP-030', 'Improvement', 'IMPROVEMENT', 'P1', 'MEDIUM', 'LOW',
     'Tests', 'js/app.js:219 (CHANGELOG-version footer parser)',
     'no unit test for footer-version regex — silent failure mode',
     'Footer version is parsed via regex /^## (v\\d+\\.\\d+\\.\\d+)/m. A future CHANGELOG line that does not start with a v version would silently break the footer.',
     'Best-in-class: every web app of this size has a unit test for "given fixture CHANGELOG, footer renders v1.X.Y" — MDN docs site does this.',
     'Add a unit test fixture that asserts the regex matches across realistic CHANGELOG variations.',
     '—', ''),
    # P2
    ('IMP-031', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'MEDIUM',
     'UX · Review surfaces', 'js/review.js, js/summary.js',
     'no wrong-answer history view — Bunpro / WaniKani / Anki all have this',
     'When a learner gets a question wrong, FSRS-4 resets the pattern to box 1 but the wrong answer itself is not surfaced. Summary shows aggregate stats; Review shows due items only.',
     'Best-in-class: Bunpro "Past Reviews" log shows every grade per item; WaniKani leech-trainer; Anki Browser filterable by "again-pressed".',
     'Persist wrongHistory: [{qId, ts, wrongAnswer, correctAnswer}, ...] in storage (capped at last 200), expose at #/review/missed.',
     '—', ''),
    ('IMP-032', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'HIGH',
     'Test mode · Realism', 'js/test.js, js/papers.js',
     'no full-mock-paper sitting (chained Moji-Goi 25min + Bunpou-Dokkai 50min + Choukai 30min)',
     'See ISSUE-020. Reframed as improvement: ship a true sitting simulator beyond single-section papers.',
     'Best-in-class: JLPT Sensei free full N5 mock; Migaku paid exam simulator; jlpt.jp official sample workbook all match this format. App sits below jlpt.jp free PDF here.',
     'Chain 4 paper-N papers + 1 listening paper with per-section timer.',
     'ISSUE-016, ISSUE-020', ''),
    ('IMP-033', 'Improvement', 'IMPROVEMENT', 'P2', 'HIGH', 'HIGH',
     'SRS · Vocabulary + Kanji',
     'js/storage.js:299 (getDuePatternIds is grammar-only)',
     'FSRS-4 SRS only schedules grammar — vocab (1041) + kanji (106) have only binary "Mark as known"',
     'Active learner tracking 1041 vocab words has no daily review queue for them.',
     'Best-in-class: WaniKani is gold standard (every kanji + vocab on SRS); Anki is open-source equivalent. Grammar-only SRS sits below the under-$20 alternative bar.',
     'Extend FSRS-4 state to vocab + kanji; surface unified review queue at #/review.',
     'Q9', ''),
    ('IMP-035', 'Improvement', 'IMPROVEMENT', 'P2', 'MEDIUM', 'LOW',
     'PWA / data-version', 'no data/version.json exists',
     'no machine-readable build-stamp — footer + sw + README drift independently',
     'Footer parses CHANGELOG; SW uses hardcoded constant; README hardcodes counts. None can be cross-checked automatically.',
     'Best-in-class: most static sites ship a version.json (Cloudflare Pages, Vercel, GH Pages all generate one); MDN does it explicitly.',
     'Write {version, builtAt, counts:{...}} at release; consume from sw.js + footer + README-checker.',
     '—', ''),
    # P3
    ('IMP-036', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'UX · Review forecast', 'js/summary.js, js/home.js',
     'no 7-day review forecast — Bunpro / WaniKani / Anki Stats all have this',
     'Dashboard shows "N reviews due today" but no horizon view. Learner cannot anticipate "Wednesday I will have 80 reviews".',
     'Best-in-class: Bunpro, WaniKani, Anki Stats all show 7-day or 30-day forecast bar chart. One-screen UX upgrade that meaningfully changes study cadence.',
     'Aggregate FSRS-4 nextReview timestamps into per-day buckets; render hairline bar chart on #/summary.',
     'IMP-033 (if vocab+kanji also in scope)', ''),
    ('IMP-037', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'Content / Search', 'js/search.js (limited to grammar/vocab/kanji indexes)',
     'header search does not index reading passages, listening transcripts, paper rationales',
     'Search hits the 3 reference corpora only. Reading + listening + paper rationales + grammar explanations are not indexed.',
     'Best-in-class: Jisho indexes every example sentence and every kanji; JLPT Sensei has site-wide search.',
     'Extend indexer to include reading.json#passages[*].ja, listening.json#items[*].script_ja, grammar explanations.',
     '—', ''),
    ('IMP-038', 'Improvement', 'IMPROVEMENT', 'P3', 'MEDIUM', 'MEDIUM',
     'UX · Audio', 'js/reading.js, js/listening.js',
     'audio uses default <audio controls> — no 5s skip-back, no per-clip rate, no segment markers',
     'Settings has a global audio rate but no per-clip nudge. Browser-native UI only.',
     'Best-in-class: NHK Easy News sentence-by-sentence replay; Migaku 1.5×/0.5× toggle inline; native podcast apps default to 5s/30s skip.',
     'Replace <audio controls> with a custom JS player exposing ±5s skip and per-clip rate.',
     '—', ''),
    ('IMP-039', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'UX · Daily goal', 'no setting, no display',
     'no daily-review goal — home shows streak + due-today but no target',
     'Learner cannot set "20 reviews per day" target and see progress against it.',
     'Best-in-class: Duolingo built around daily-goal nags (XP/day); WaniKani ships email reminders.',
     'Add dailyGoalReviews to settings; render single ring on home.',
     'Q12 (round-2: daily-goal default unresolved)', ''),
    ('IMP-040', 'Improvement', 'IMPROVEMENT', 'P3', 'LOW', 'LOW',
     'PWA / install', 'manifest.webmanifest',
     'manifest does not declare app-shortcuts (long-press deeplinks)',
     'Manifest is otherwise correct. shortcuts: array would expose Reviews / Test as long-press deep links from the installed PWA icon.',
     'Best-in-class: NHK Easy News PWA, Bunpro PWA both expose 2-3 long-press shortcuts. Zero cost; better install ergonomics.',
     'Add shortcuts: [{name:"Reviews", url:"#/review"}, {name:"Test", url:"#/test"}] to manifest.',
     '—', ''),
    # P4
    ('IMP-041', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'LOW',
     'Locale infra', 'locales/*.json',
     '5 locale files at 47 keys each — commit to localizing or remove',
     'See ISSUE-022. Either fully localize (HIGH effort) or remove the 4 non-English files (LOW effort).',
     'Keeping unused locales gives false impression of multi-language support; Tatoeba (multi-locale done right) localises every string.',
     'Decision Q8.',
     'ISSUE-022, Q8', ''),
    ('IMP-042', 'Improvement', 'IMPROVEMENT', 'P4', 'MEDIUM', 'HIGH',
     'Content · Native audio',
     'data/audio_manifest.json (voice_default: synthetic-gtts)',
     'all audio is gTTS-synthetic — manifest schema supports voice:"native" but no native recordings ship',
     'Manifest already supports native voice (skipped by tools/build_audio.py for native:true) but no native files.',
     'Best-in-class: NHK announcer recordings; Bunpro native audio; Tofugu Tofu podcast. gTTS is recognisable as synthetic; N5 listener trained on gTTS may struggle with real audio.',
     'Scope native-recording pass for 40 listening + 30 reading (grammar examples can stay synthetic).',
     'Q11', ''),
    ('IMP-043', 'Improvement', 'IMPROVEMENT', 'P4', 'LOW', 'MEDIUM',
     'A11y', 'css/main.css, app shell',
     'no font-size scaling control or high-contrast mode toggle',
     'Theme System/Light/Dark works; reduce-motion is honored. No font-size scaling; no high-contrast toggle.',
     'Best-in-class: gov.uk + MDN ship explicit contrast + font-size controls. WCAG AA recommends user-controlled font scaling beyond browser zoom.',
     'Add font-scale chip group (90/100/115/130%) in Settings; map to :root font-size variable.',
     '—', ''),
    # P5
    ('IMP-044', 'Improvement', 'IMPROVEMENT', 'P5', 'LOW', 'HIGH',
     'Content · Onboarding', 'no first-run tutorial',
     'fresh installs land on home with no walkthrough — diagnostic exists but is not promoted',
     'Diagnostic at #/diagnostic exists but no first-run routing.',
     'Best-in-class: Duolingo placement test as on-ramp; Bunpro setup wizard picks lessons by perceived level.',
     'Route fresh installs to a 3-question placement at #/diagnostic; emit recommended starting lesson.',
     '—', ''),
]

# 5-column rows for "Open questions". Decision blank.
QUESTIONS = [
    ('Q8', 'Localization commitment',
     '5 locale files (en/id/ne/vi/zh × 47 keys) cover ~6% of UI strings. Hundreds of literals are hardcoded English in js/*.js. Half-state implies multi-language support that does not exist.',
     'Extract every literal to keys and translate fully (HIGH effort, broad reach), OR remove 4 non-English locale files and document English-only.',
     ''),
    ('Q9', 'Vocab + kanji SRS scope',
     'Adding FSRS-4 to 1041 vocab + 106 kanji means daily-due could spike to 100+ items for an active learner — closer to WaniKani pace than current grammar-only.',
     'Acceptable, or should the app cap daily-due at N items?',
     ''),
    ('Q10', 'Mock-paper sitting timing',
     'Full sitting is 105 minutes (Moji-Goi 25 + Bunpou-Dokkai 50 + Choukai 30). Real-time pressure is part of exam realism but may be punishing for casual practice.',
     'Enforce real-time timing (no pause), allow pauses, or both as a setting?',
     ''),
    ('Q11', 'Native audio investment',
     '40 listening items + 30 reading passages = ~70 short clips = ~2-3 hours recorded audio plus splicing. Approximate cost USD$300-1500 depending on voice talent.',
     'In or out of scope for this project budget?',
     ''),
    ('Q12', 'Round-2 deferred questions',
     'Q2-Q7 from round-2 still unanswered: timer-default policy, daily-goal default, bunpou segmentation rule, auto-furigana risk acceptance, metered-mobile-data default. Block several round-2 follow-ups.',
     'Resolve Q2-Q7 (separate decisions per Q) before round-3 P3+ items can land.',
     ''),
]


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl not installed. Run `pip install openpyxl`.')
        return 1

    if not XLSX.exists():
        print(f'ERROR: {XLSX} not found.')
        return 1

    wb = load_workbook(XLSX)

    # --- Audit findings sheet (rows 5+) ---
    ws = wb['Audit findings']
    existing_ids = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row and row[0]:
            existing_ids.add(str(row[0]).strip())

    rows_to_append = []
    for row in ISSUES + IMPS:
        if row[0] in existing_ids:
            continue
        rows_to_append.append(row)
    next_row = ws.max_row + 1
    for r in rows_to_append:
        for col_offset, val in enumerate(r):
            ws.cell(row=next_row, column=col_offset + 1, value=val)
        next_row += 1

    # --- Open questions sheet (rows 4+) ---
    ws_q = wb['Open questions']
    existing_qids = set()
    for row in ws_q.iter_rows(min_row=4, values_only=True):
        if row and row[0]:
            existing_qids.add(str(row[0]).strip())

    q_to_append = [r for r in QUESTIONS if r[0] not in existing_qids]
    next_q_row = ws_q.max_row + 1
    for r in q_to_append:
        for col_offset, val in enumerate(r):
            ws_q.cell(row=next_q_row, column=col_offset + 1, value=val)
        next_q_row += 1

    wb.save(XLSX)

    print(f'Audit findings:  appended {len(rows_to_append)} rows '
          f'({sum(1 for r in rows_to_append if r[0].startswith("ISSUE-"))} issues + '
          f'{sum(1 for r in rows_to_append if r[0].startswith("IMP-"))} improvements)')
    if existing_ids & {r[0] for r in ISSUES + IMPS}:
        skipped = sorted(existing_ids & {r[0] for r in ISSUES + IMPS})
        print(f'  skipped (already present): {", ".join(skipped)}')
    print(f'Open questions:  appended {len(q_to_append)} rows')
    if existing_qids & {r[0] for r in QUESTIONS}:
        skipped = sorted(existing_qids & {r[0] for r in QUESTIONS})
        print(f'  skipped (already present): {", ".join(skipped)}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
