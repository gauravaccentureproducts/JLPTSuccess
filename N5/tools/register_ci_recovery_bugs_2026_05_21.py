"""
Register the four bugs surfaced + closed by the 2026-05-21 CI-recovery
triage to the User Reported Bugs sheet.

Run once. Idempotent: skips bugs whose Description matches an already-
present row (per the canonical "BUG-Cxxx — ..." prefix convention).

Source bug commits:
  - BUG-A11Y-001 → c1750a3 (primary-nav) + 9a9d827 (icon-btn + footer)
  - BUG-RECO-001 → 68d9241 (R-14 / R-13 RULES dispatch swap)
  - BUG-CI-001   → 9a9d827 (visual-regression test.skip CI gate)
                   + 0e505e4 (workflow_dispatch regen mechanism)
  - BUG-TEST-001 → b3c4bee + 3349c97 + 68d9241 + 4c491b4 + 18f1774
                   (15+ stale Playwright assertions across 4 spec files)
"""

from __future__ import annotations

import datetime as _dt
import sys
from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "specifications" / "test-scenarios-by-specialist-perspective.xlsx"

REPORTED_DATE = _dt.date(2026, 5, 21)
FIX_DATE = _dt.date(2026, 5, 21)
REPORTED_BY = (
    "Playwright CI recovery (2026-05-21) — first complete smoke-suite "
    "run after the workers 1→2 + video off + visual-regression test.skip "
    "infra-fix batch (commits 397a933..0e505e4); 65 pre-existing failures "
    "surfaced across 4 distinct classes — these are the close-out registrations"
)

BUGS = [
    {
        "title": (
            "BUG-A11Y-001 — Color-contrast trio (3 elements) fails WCAG AA: "
            "primary-nav links + locale-toggle + footer-disclaimer"
        ),
        "description": (
            "FILE: N5/css/main.css (--color-text-muted #6F6D66 + --color-text-faint #9A968C).\n"
            "SURFACED BY: axe-core in N5/tests/p0-smoke.spec.js once the suite ran to completion on CI.\n"
            "INSTANCES:\n"
            "  (a) .primary-nav a — #6F6D66 on tea-green header bg #cfd8b5 = 3.48 contrast (needs 4.5).\n"
            "      Every primary-nav link (Grammar, Vocabulary, Kanji, Reading, Listening, Test, Progress)\n"
            "      failed on every route that renders the header. 408 violation instances on home alone.\n"
            "  (b) .app-header .icon-btn (settings cog, fullscreen, locale-toggle 'HI' label) —\n"
            "      same root cause; same contrast 3.48.\n"
            "  (c) .app-footer .footer-disclaimer — #9A968C (faint) on white footer = 2.95 contrast.\n"
            "      All three are serious-impact WCAG AA color-contrast violations.\n\n"
            "FIX: Introduced new --color-text-on-header token (#4A4A47 in light mode, contrast 5.92 on\n"
            "the tea-green header band; aliased to --color-text-muted #A8A59C in dark mode). Wired through\n"
            ".primary-nav a + .app-header .icon-btn. Footer-disclaimer switched from --color-text-faint\n"
            "to --color-text-muted (4.94 contrast on white). main.min.css regenerated.\n\n"
            "DETECTION: axe-core P0 smoke suite. Was masked by the 15-min Playwright job timeout that had\n"
            "been cancelling the suite mid-run since DEFER-6 closure (2026-05-03). Fix verified by CI run\n"
            "26258169250 (first green run; a11y suite clean)."
        ),
        "severity": "High",
        "priority": "P2",
        "fix_commit": "c1750a3 + 9a9d827",
    },
    {
        "title": (
            "BUG-RECO-001 — Pedagogy recommender R-13 catch-all dominated R-14 mock-paper-ready "
            "(R-14 was dead code for every returning user)"
        ),
        "description": (
            "FILE: N5/js/pedagogy-recommender.js (RULES dispatch array).\n"
            "SURFACED BY: tests/recommender.spec.js R-14 test once it ran end-to-end on CI.\n"
            "ROOT CAUSE: R-13 (continueExplore — 'catch-all returning user gets Open Learn') had\n"
            "condition `if (signal.isReturning)` — too permissive — and dispatched BEFORE R-14\n"
            "(mockPaperReady — grammar≥60% AND kanji≥50%) in the RULES array. Any returning user\n"
            "with mock-ready coverage got R-13's 'Open Learn' recommendation instead of R-14's\n"
            "'Take a mock paper' recommendation. R-14 was effectively dead code.\n\n"
            "EVIDENCE: test signal grammar 0.65 + kanji 0.55 returned rule_id='R-13' instead of\n"
            "the expected 'R-14'. The R-13 test signal (grammar 0.10 + kanji 0.10) still correctly\n"
            "fires R-13 because R-14's thresholds aren't met.\n\n"
            "FIX: Swapped RULES dispatch order to place R14_mockPaperReady BEFORE R13_continueExplore.\n"
            "R-13 remains the true catch-all per its inline doc. Inline comment block added documenting\n"
            "why specific rules must precede catch-alls in any priority chain.\n\n"
            "DETECTION: Unit-tested recommender suite (recommender.spec.js R-14 test). Was masked by\n"
            "the same CI timeout that hid the a11y bugs."
        ),
        "severity": "Medium",
        "priority": "P2",
        "fix_commit": "68d9241",
    },
    {
        "title": (
            "BUG-CI-001 — Visual-regression baselines committed as -win32.png only; CI Linux runner "
            "requests -linux.png → 38 'snapshot doesn't exist' failures per run since 2026-05-03"
        ),
        "description": (
            "FILES: N5/tests/visual-regression.spec.js-snapshots/*.png (76 PNGs, all -win32.png).\n"
            "SURFACED BY: Playwright suite first complete run on CI after timeout / parallelisation fix.\n"
            "ROOT CAUSE: Visual-regression baselines were captured on a Windows dev box; Playwright\n"
            "suffixes each baseline with the OS (-win32.png, -linux.png, -darwin.png). On CI ubuntu-latest\n"
            "the test runner requests -linux.png baselines that have never been generated → for every\n"
            "viewport × project × route combination, Playwright reports 'snapshot doesn't exist, writing\n"
            "actual' which counts as a test failure. 38 unique failures per CI run since DEFER-6 closure\n"
            "(2026-05-03 when the visual-regression spec was wired).\n\n"
            "FIX (two-step):\n"
            "  (a) Interim: tests/visual-regression.spec.js gated to local-only via\n"
            "      `test.skip(!!process.env.CI && process.env.PLAYWRIGHT_UPDATE_SNAPSHOTS !== 'true', ...)`.\n"
            "      Local Windows dev still gets diffing; CI no longer reports the 38 failures.\n"
            "  (b) Permanent: workflow_dispatch input `update_snapshots: true` added to\n"
            "      .github/workflows/playwright.yml. Triggers --update-snapshots run that writes new\n"
            "      Linux baselines and uploads them as a 14-day artifact for manual commit. Once Linux\n"
            "      baselines land, the CI gate is removed and visual-regression runs on every push.\n\n"
            "DETECTION: Same Playwright suite. Now properly gated; future cross-platform dev box additions\n"
            "should follow the per-OS-baselines-or-skip rule documented in procedure manual §F.40.6."
        ),
        "severity": "Medium",
        "priority": "P2",
        "fix_commit": "9a9d827 + 0e505e4",
    },
    {
        "title": (
            "BUG-TEST-001 — 15+ stale Playwright assertions against removed/restructured UI "
            "(homepage hero, locale-chip group, trust-band, count drift)"
        ),
        "description": (
            "FILES: N5/tests/branding.spec.js + p0-smoke.spec.js + round3-features.spec.js + "
            "v1.12.28-features.spec.js + recommender.spec.js.\n"
            "SURFACED BY: Playwright suite first complete run on CI.\n"
            "ROOT CAUSE: When UI was restructured (elements removed, text changed, layout collapsed),\n"
            "tests against those affordances became stale but typically didn't fail loudly until the\n"
            "test suite ran end-to-end. The 15-min CI timeout cancellations had been hiding them.\n\n"
            "INSTANCES (representative — full list in commits):\n"
            "  - `.syllabus-title` + `.syllabus-subtitle` removed when home restructured 2026-05-09.\n"
            "  - `.syllabus-trust-band` + `.trust-pill` removed in favor of in-card niche-N2 messaging.\n"
            "  - `.locale-chip` × 2 → single `#locale-toggle` icon-btn (2026-05-09 redesign).\n"
            "  - 'Start sitting' CTA copy → 'Start full mock test →' (2026-05-10).\n"
            "  - Test-length picker: `button.length-btn` → `<select id='test-length'>`.\n"
            "  - Grammar TOC tier chips removed 2026-05-10 per user feedback.\n"
            "  - Hardcoded counts (177 grammar / 30 reading / 30 listening) drifted to (178 / 54 / 50).\n"
            "  - .brand-link visible text was 'JLPT N5' literal; commit a91bb68 (2026-05-04 'unify brand\n"
            "    mark with landing page') changed it to 'N5'-only-plus-SVG.\n"
            "  - Recommender R-07..R-14 tests didn't null out lastLearnId, so R-06 resume-last dominated.\n"
            "  - First-run onboarding redirect (IMP-044, 2026-05-11) sent every fresh-context CI test\n"
            "    to #/diagnostic.\n\n"
            "FIX (7 batches across commits b3c4bee → 18f1774):\n"
            "  - Replaced hardcoded counts with `fetch('data/version.json')` reads at runtime.\n"
            "  - Removed-UI assertions: either test.skip() with git-history comment or rewritten for\n"
            "    the replacement affordance.\n"
            "  - Brand-link: assert level-code 'N5' + aria-label /JLPT/i (the meaningful invariants).\n"
            "  - Recommender: added `lastLearnId: null` to baselines for R-07..R-14 tests.\n"
            "  - Onboarding redirect: global `test.beforeEach` in 3 spec files sets\n"
            "    `localStorage['jlpt-n5-tutor:onboardingSeen']='1'` via addInitScript before goto.\n"
            "  - Strict-mode locator: JSON-LD via page.evaluate walk; .bank-note via .filter; etc.\n\n"
            "DETECTION: Playwright P0 smoke. The horizontal-deployment-sweep rule (procedure manual\n"
            "§F.37.6 / F.38.5 / F.40.2) now extends to test artifacts: when removing/restructuring\n"
            "UI, grep tests/ for the affected selectors + copy as part of the same commit."
        ),
        "severity": "Minor",
        "priority": "P3",
        "fix_commit": "b3c4bee + 3349c97 + 68d9241 + 4c491b4 + 18f1774",
    },
]


def _bug_already_registered(sheet, title_prefix: str) -> bool:
    """Return True if any existing row's Title (col 4) starts with title_prefix."""
    for r in range(4, sheet.max_row + 1):
        existing_title = sheet.cell(r, 4).value
        if isinstance(existing_title, str) and existing_title.startswith(title_prefix):
            return True
    return False


def main() -> int:
    wb = openpyxl.load_workbook(WORKBOOK)
    sheet = wb["User Reported Bugs"]
    next_row = sheet.max_row + 1
    written = 0
    skipped = 0
    for bug in BUGS:
        title_prefix = bug["title"].split(" — ")[0]  # "BUG-A11Y-001"
        if _bug_already_registered(sheet, title_prefix):
            print(f"SKIP {title_prefix}: already registered")
            skipped += 1
            continue
        # Col 1 is the auto-formula Bug ID; we don't write it (leave as
        # default to maintain the row-based formula).
        sheet.cell(next_row, 1, value='="BUG-"&TEXT(ROW()-3,"000")')
        sheet.cell(next_row, 2, value=REPORTED_DATE)
        sheet.cell(next_row, 3, value=REPORTED_BY)
        sheet.cell(next_row, 4, value=bug["title"])
        sheet.cell(next_row, 5, value=bug["description"])
        sheet.cell(next_row, 6, value=bug["severity"])
        sheet.cell(next_row, 7, value=bug["priority"])
        sheet.cell(next_row, 8, value="Fixed")
        sheet.cell(next_row, 9, value=bug["fix_commit"])
        sheet.cell(next_row, 10, value=FIX_DATE)
        print(f"WRITE row {next_row}: {title_prefix}")
        written += 1
        next_row += 1
    wb.save(WORKBOOK)
    print(f"\nDone. Wrote {written}, skipped {skipped}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
