"""Phase 9 of the en+hi locale transition (2026-05-06).

Registers the architectural decision as IMP-096 in
feedback/n5-audit-2026-05-04.xlsx.

Modeled on tools/_register_round7_findings.py from round-7. Appends
exactly one row (Items sheet) carrying the full rationale + before/
after state.

Idempotent: skips if IMP-096 already exists.
"""
from __future__ import annotations
import io, sys
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'feedback' / 'n5-audit-2026-05-04.xlsx'

ROW = (
    'IMP-096',
    'Improvement',
    'IMPROVEMENT',
    'P1',
    'HIGH',
    'HIGH',
    'Strategic positioning / locale architecture',
    'N5/locales/, N5/js/i18n.js, N5/data/*.json, N5/specifications/*, top-level brand surfaces',
    'Strategic narrowing - 5-locale shell -> en+hi only',
    (
        'Transitioned on 2026-05-06 per market-research conversation '
        '2026-05-06. Hindi added (high-demand low-competition niche per '
        'top-5 JLPT country with no dedicated Hindi-medium prep app); '
        'vi/id/ne/zh removed (saturated competitive markets per market '
        'research; 5-locale shell was diluting depth across surfaces with '
        'no native-quality content in any). Locales before: 5 (en/vi/id/'
        'ne/zh). Locales after: 2 (en/hi). Tag pre-locale-transition at '
        'commit a611664 marks the parent state for revert reference.'
    ),
    (
        'Niche-N1 reframed from "multilingual non-English-native" to '
        '"the only privacy-first no-account offline JLPT app with English '
        '+ native Hindi pedagogy." Concentrates depth-first investment '
        'per the N5Improvement.txt depth-first audit cycle.'
    ),
    (
        'IMPLEMENTATION COMPLETE this row. Future passes: native Hindi '
        'review pass on hi.json (UI chrome) + content surfaces (vocab '
        'gloss_hi 116/1041 -> 50%+; grammar meaning_hi/explanation_hi '
        '0/178 -> 30+; reading + listening explanation_hi 0/45 + 0/47 -> '
        '20+); L1-interference notes for Hindi (postposition mapping, '
        'verb agreement, tense over-marking, politeness mismatch, '
        'negative placement, question particle, plural marking, SOV '
        'shared advantage) on top-15 grammar patterns; Tier-2/3-city '
        'Indian learner UX research.'
    ),
    'None - this row records the architectural decision; subsequent Hindi-content depth work depends on it but does not block this entry.',
    'Done',
    (
        'Backstory: 2026-05-06 deep market research found that '
        'Vietnamese / Indonesian / Burmese / Bengali / Sinhala / '
        'Brazilian-Portuguese / Filipino / Thai / Chinese all have '
        'established competing JLPT prep apps in their native language; '
        'the "no competitor" framing for niche-N1 was wrong for those '
        'locales. Nepali is genuinely uncovered, but the absolute market '
        'is too small (~2-5K JLPT/yr) to be the strategic primary. '
        'Hindi is the unique high-demand-low-competition gap: India is '
        'the 5th-largest JLPT country (~50K applicants/year, 73% at N5 '
        'or N4); no dedicated Hindi-native JLPT prep app was found in '
        'app-store or curated-list searches. The Yoisho Academy '
        'delivers in English with optional Hindi tutoring (closest '
        'competitor, but not an app and not Hindi-medium pedagogy). '
        'Implementation: 10-phase rollout per prompts/LocaleTransitionEnHi.txt '
        '(Phase 0 inventory; Phase 1 add hi additively; Phase 2 migrate '
        'persisted vi/id/ne/zh to en silently; Phase 3 remove deprecated '
        'locales from UI + i18n module + sw.js precache; Phase 4 strip '
        '2606 deprecated-locale keys from content data + seed 116 vocab '
        'gloss_hi + 106 kanji meanings_hi; Phase 5 rewrite docs / specs '
        '/ brand surfaces; Phase 6 tighten CI invariants - JA-13 '
        'extended + new JA-39 locale-set guard; Phase 7 update Playwright '
        'specs from 5-chip to 2-chip; Phase 8 smoke test green; Phase 9 '
        'this registration row; Phase 10 push to origin/master). '
        'CI: 48/48 invariants PASS post-transition.'
    ),
    'No further permission needed - user explicitly requested this transition on 2026-05-06.',
)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb['Items']
    # Check if IMP-096 already exists
    for row in ws.iter_rows(min_row=5, values_only=False):
        if row[0].value == 'IMP-096':
            print('IMP-096 already registered; skipping.')
            return 0
    new_row = ws.max_row + 1
    for ci, val in enumerate(ROW, start=1):
        ws.cell(row=new_row, column=ci, value=val)

    # Extend Decision column data validation to cover the new row
    new_max = ws.max_row
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and 'Fix' in dv.formula1:
                dv.sqref = f'N5:N{new_max}'
            elif dv.formula1 and 'Allow' in dv.formula1:
                dv.sqref = f'P5:P{new_max}'

    wb.save(XLSX)
    print(f'IMP-096 appended to row {new_row}; xlsx saved.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
