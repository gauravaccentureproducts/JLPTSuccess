"""Register NTR-001..NTR-013 (Native Teacher Review batch 2026-05-22)
in the xlsx bug tracker. All claims pre-verified via
tools/verify_native_teacher_review_2026_05_22.py (per F.41.4).
"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XLSX = "specifications/test-scenarios-by-specialist-perspective.xlsx"
BAK = XLSX + ".bak_2026_05_22_ntr_batch"
if not os.path.exists(BAK):
    shutil.copy2(XLSX, BAK)

wb = load_workbook(XLSX)
ws = wb["User Reported Bugs"]
print(f"max_row before: {ws.max_row}")

BUGS = [
    # ===== SEVERITY 1 — P1 / P2 =====
    {
        "title": "NTR-001 — 99 vocab examples leak non-N5, non-exception kanji (LLM-curated pass introduced ~3× the human-baseline rate)",
        "desc": (
            "FILE: data/vocab.json — 99 of 3036 vocab examples contain kanji NOT in n5_kanji_whitelist (106) AND NOT in dokkai_kanji_exception (90) = NOT in the 196-kanji allowed set.\n\n"
            "Provenance split (verified 2026-05-22):\n"
            "  llm_curated examples: 56 / 970 (5.8%)\n"
            "  untagged human baseline: 43 / 2066 (2.1%)\n"
            "The LLM-curation pass introduces ~3× the rate of the human baseline.\n\n"
            "Top offending kanji (by frequency across the 99 violations):\n"
            "  夜(14), 夏(10), 待(7), 色(7), 昼(6), 杯(4), 冬(4), 枚(4), 走(3), 様(2), 族(2), 事(2), 都(2), 答(2), 米(2), 予(2), 報(2), 赤(2), 茶(2), 音(2), 借, 様, 京, 旅, 始, 匹, 才, 朝, 帰.\n\n"
            "Concrete cases (verified):\n"
            "  n5.vocab.10-time-general.びょう ex[2]: '十びょう 待って ください' (待) — llm_curated\n"
            "  n5.vocab.7-numbers.一 ex[2]: 'コーヒーを 一杯 ください' (杯) — llm_curated\n"
            "  n5.vocab.7-numbers.万 ex[2]: '一万円 借りました' (借) — llm_curated\n"
            "  n5.vocab.11-time-days-weeks-month.先週 ex[0]: '先週 京都に 行きました' (都) — <untagged>\n"
            "  n5.vocab.14-nature-and-weather.はれ ex[2]: 'あしたは はれの 予報です' (予,報) — llm_curated\n"
            "  n5.vocab.15-animals.ねこ ex[2]: 'ねこが 二匹 います' (匹) — llm_curated\n"
            "  n5.vocab.1-people-pronouns-and-se.どなた ex[2]: 'どなた様ですか' (様) — llm_curated\n"
            "  n5.vocab.4-body-parts.め ex[2]: 'こん夜は めが つかれました' (夜) — llm_curated\n\n"
            "Diagnostic pattern: the original human-curated example for びょう uses kana ('十びょう まって ください'); the llm_curated example 'naturalized' the orthography to kanji ('十びょう 待って ください') without consulting the whitelist.\n\n"
            "Fix direction (option a per native-teacher review): rewrite the 99 violating examples to use only whitelist∪exception kanji (or use kana for non-whitelist words). Add CI invariant JA-150 to lock the predicate. Option b (add a parallel `reading` field to every vocab example as kana-fallback for UI furigana) is the second-line safety net but is a separate 3,036-example authoring task; deferred.\n\n"
            "Same defect class as how grammar examples / reading passages / listening scripts are already gated — bring vocab examples up to parity."
        ),
        "sev": "Major", "pri": "P1",
    },
    {
        "title": "NTR-002 — n5-045 is a duplicate of n5-017 (both teach 何/なに・なん); n5-045 self-identifies as duplicate but cleanup wasn't finished",
        "desc": (
            "FILES: data/grammar.json + data/n5_core_pattern_ids.json.\n\n"
            "Verified 2026-05-22:\n"
            "  n5-017: pattern='何（なに／なん）' meaning_en='What - pronounced なに or なん depending on what follows'\n"
            "  n5-045: pattern='何（なに・なん）' meaning_en='What'\n"
            "  Both in core_n5 section of n5_core_pattern_ids.json.\n\n"
            "n5-045's own contrasts block self-identifies:\n"
            "  {'with_pattern_id': 'n5-017', 'note': 'This is a duplicate entry - see the canonical pattern.'}\n\n"
            "Fix direction: either (a) merge n5-045 into n5-017 (preferring n5-017's examples / common_mistakes) and delete n5-045 from n5_core_pattern_ids.json + grammar.json, or (b) if external IDs reference n5-045, collapse it to a redirect-only record. Same defect class as PAPER-002 / duplicate-entry merges in earlier batches.\n\n"
            "Recommend (a) because no external references to n5-045 exist in the corpus (verified via grep)."
        ),
        "sev": "Major", "pri": "P2",
    },
    {
        "title": "NTR-003 — かれ / かのじょ glosses are inverted vs modern usage; 'boyfriend/girlfriend' is primary, 'he/she' is literary",
        "desc": (
            "FILE: data/vocab.json.\n\n"
            "Verified 2026-05-22 current data:\n"
            "  かれ:    gloss='he, him (boyfriend - more advanced sense)'\n"
            "  かのじょ: gloss='she, her (girlfriend - more advanced sense)'\n\n"
            "This is backwards. In contemporary spoken Japanese, *boyfriend/girlfriend* is the **primary** sense — that's what a 20-something hears when someone says かれ in conversation. The pronoun 'he/she' sense is literary or distancing and is the marked / context-dependent usage.\n\n"
            "Additional concern: the example 'かれは とても しんせつ です' under the boyfriend reading produces an unintentional double-entendre native learners notice.\n\n"
            "Recommended gloss rewrites (per native-teacher review):\n"
            "  かれ:    'boyfriend (primary); he, him (third-person pronoun, more formal/literary)'\n"
            "  かのじょ: 'girlfriend (primary); she, her (third-person pronoun, more formal/literary)'\n\n"
            "Also re-pick the lead example for each to fit the primary sense (or label it explicitly when the example uses the literary sense)."
        ),
        "sev": "Major", "pri": "P2",
    },
    {
        "title": "NTR-004 — あなた needs a usage warning; current example 'あなたは どなたですか' pairs あなた with どなた = survey-form register",
        "desc": (
            "FILE: data/vocab.json.\n\n"
            "Verified 2026-05-22:\n"
            "  あなた: gloss='you'  (no usage_note)\n"
            "  first example: 'あなたは どなたですか。'\n\n"
            "Every native teacher will flag: あなた is NOT a normal way to address someone you know — it's distant, formal, or (from a wife to a husband) intimate. Using it as a generic second-person pronoun is one of the most common red-flag errors learners ship to Japan with. Pairing あなた with the honorific どなた reads like a survey form, not natural speech.\n\n"
            "Recommended fix:\n"
            "  - Gloss: 'you (use with caution — Japanese typically uses the person\\'s name + さん or drops the subject entirely)'\n"
            "  - Replace example #1 with a context where あなた is actually appropriate (form-filling, wedding vows, lyrics) OR drop あなた and pivot the example to name-based address (e.g., 田中さんは どこから 来ましたか)."
        ),
        "sev": "Major", "pri": "P2",
    },

    # ===== SEVERITY 2 — P3 =====
    {
        "title": "NTR-005 — おはし mislabeled section '20. Tableware and Cooking'; section 20 is actually 'Colors', tableware is section 19",
        "desc": (
            "FILE: data/vocab.json.\n\n"
            "Verified 2026-05-22:\n"
            "  form=おはし reading=おはし section='20. Tableware and Cooking'   ← WRONG\n"
            "Section enumeration:\n"
            "  section 20 entries (15 total): いろ / 白 / 白い / くろ / くろい / ... ← all colors\n"
            "  section 19 entries (12 total): tableware-and-cooking — including form=はし (chopsticks bare form, correctly tagged section 19)\n\n"
            "Fix: re-tag おはし from section '20. Tableware and Cooking' to '19. Tableware and Cooking'. One-line edit. The kana form is correct (箸 is non-N5, so kana-only is fine)."
        ),
        "sev": "Minor", "pri": "P3",
    },
    {
        "title": "NTR-006 — えいが (movie) filed under '26. House and Furniture' instead of entertainment/leisure",
        "desc": (
            "FILE: data/vocab.json.\n\n"
            "Verified 2026-05-22:\n"
            "  form=えいが reading=えいが section='26. House and Furniture'  ← WRONG\n\n"
            "Fix: re-tag えいが to entertainment / leisure section, or '37. Common nouns (misc)' if no entertainment section exists. The item also surfaces in `vocab_used` arrays for reading passages, where misfiling makes section-grouped review lists less useful."
        ),
        "sev": "Minor", "pri": "P3",
    },
    {
        "title": "NTR-007 — 三 kanji mnemonic conflates 三-reading with -さん honorific (different etymologies)",
        "desc": (
            "FILE: data/kanji.json.\n\n"
            "Verified 2026-05-22 — 三 mnemonic.reading reads:\n"
            "  'さん — borrowed everywhere from \"Mr/Ms\" (-さん) to 三月 (sangatsu, March). Kun み in 三つ (mittsu).'\n\n"
            "The honorific さん derives from 様 → さま → さん, which is a separate etymological line from the on-yomi 三/さん (Chinese-derived). Conflating them gives the wrong etymology to curious learners who look it up.\n\n"
            "Fix: soften the mnemonic to something like:\n"
            "  'The sound *san* is everywhere in Japanese — 三月 (March), the honorific -さん, etc. — so the reading is easy to hold onto. (Note: the honorific さん comes from a different root [様/さま]; the shared sound is coincidental.)'\n\n"
            "Native-speaker pass over similar mnemonic etymology claims recommended across the kanji cohort (per native-teacher review §2.3)."
        ),
        "sev": "Minor", "pri": "P3",
    },
    {
        "title": "NTR-008 — 4 pitch-accent primary choices warrant NHK 2016 verification (これ, あなた, みなさん, きのう)",
        "desc": (
            "FILE: data/vocab.json (pitch_accent fields).\n\n"
            "Verified 2026-05-22 current values:\n"
            "  これ:    mora=2, drop=1 [①]                              — review: NHK matches; confidence high ✓\n"
            "  あなた:  mora=3, drop=2 [②], alternates=[1]              — review: NHK gives ⓪ generic + ② spousal; ② as primary is unusual\n"
            "  みなさん: mora=4, drop=2 [②]                              — review: NHK lists ③\n"
            "  きのう:  mora=3, drop=1 [①], alternates=[0, 2]            — review: NHK-standard ②; ⓪ is colloquial; ① unusual as primary\n\n"
            "These are defensible (alternates are listed for あなた + きのう) but warrant a native-pronunciation reviewer pass against NHK 日本語発音アクセント新辞典 2016. If audio recordings use the listed drops, the data matches the audio; annotate which drop the audio uses rather than claiming NHK-standard.\n\n"
            "Source: kanjium-8a0cdaa1; provenance shows by-reading lookup (not exact) for みなさん/きのう/あなた — by-reading is more error-prone than exact. NTR-008 deferred for actual native-speaker pass (pending NATIVE-SPEAKER-RE-VERIFICATION.md options); annotated as 'native-speaker-review-pending' in this batch."
        ),
        "sev": "Minor", "pri": "P4",
    },

    # ===== SEVERITY 3 — P4 =====
    {
        "title": "NTR-009 — boyfriend/girlfriend gloss nuance lost in Hindi sync (gloss vs gloss_hi out of sync)",
        "desc": (
            "FILE: data/vocab.json.\n\n"
            "Verified 2026-05-22:\n"
            "  かれ:    gloss='he, him (boyfriend - more advanced sense)'    gloss_hi='वह (पुरुष)'\n"
            "  かのじょ: gloss='she, her (girlfriend - more advanced sense)'  gloss_hi='वह (महिला)'\n\n"
            "Hindi gloss carries only the literary-pronoun reading; English carries the boyfriend qualifier. Both right for half the truth.\n\n"
            "Fix follows the NTR-003 re-gloss decision; this entry resolves automatically when NTR-003 lands. Filed as a separate item to track that the Hindi sync is explicit."
        ),
        "sev": "Minor", "pri": "P4",
    },
    {
        "title": "NTR-010 — Q-0226 has a subtle particle-nuance trap (は after time noun → contrast) not flagged in explanation_en",
        "desc": (
            "FILE: data/questions.json q-0226.\n\n"
            "Verified 2026-05-22:\n"
            "  question_ja: 'きのうは ばんごはんを たべ（　）。'\n"
            "  correctAnswer: 'ませんでした'\n"
            "  explanation_en: 'ませんでした is the polite past negative. I didn\\'t eat dinner yesterday.'\n\n"
            "The verb conjugation tested IS ませんでした, so the answer is fine. But 'きのうは ばんごはんを 食べませんでした' in natural speech implies CONTRAST ('Yesterday in particular, I didn\\'t eat dinner — but other days I do'). Neutral past-negative would be 'きのう ばんごはんを 食べませんでした' (no particle) or 'に'-marking.\n\n"
            "Above N5 this would be a `high_confusion: true` candidate; at N5 leaving it as-is is defensible, but a one-line note protects learners who notice the wrinkle.\n\n"
            "Fix: append to explanation_en (or add as explanation_hi parallel): 'Note: は after a time noun introduces contrast — fine here, but typical neutral past-negative would drop the particle.'"
        ),
        "sev": "Minor", "pri": "P4",
    },
    {
        "title": "NTR-011 — pronoun `collocations` field is mechanically generated particle-template substitution, not real collocations",
        "desc": (
            "FILE: data/vocab.json (pronoun entries: 私, あなた, かれ, かのじょ, わたしたち, etc.).\n\n"
            "Verified 2026-05-22 — every pronoun's `collocations` field is the same shape with the pronoun substituted:\n"
            "  わたしの ともだち / わたしの なまえ / わたしは がくせい /\n"
            "  わたしも いく / わたしを しる / わたしと あう\n"
            "  ...then あなた / かれ / かのじょ / わたしたち variants of the same 6.\n\n"
            "As a teaching aid this is harmless (the templates are illustrative of particle usage), but it isn't real *collocation* data in the corpus-linguistics sense. Real collocations would be 私の場合 / 彼女いない歴 (etc., out of scope for N5) or 私たちの間 — phrasal patterns specific to each pronoun.\n\n"
            "Fix options:\n"
            "  (a) Rename the field to `particle_examples` to reflect what it actually contains (more honest).\n"
            "  (b) Override the template with real collocations where they exist within N5 scope (e.g., 'わたしも', 'あなたも', 'みなさんへ').\n\n"
            "Recommend (a) — terminology fix is cheap and honest; (b) is a larger authoring task that can come later."
        ),
        "sev": "Minor", "pri": "P4",
    },
    {
        "title": "NTR-012 — 七 kanji entry primary=しち is correct prescriptively but reading_rule should flag なな is more common in modern colloquial counting",
        "desc": (
            "FILE: data/kanji.json (七 entry).\n\n"
            "Verified 2026-05-22:\n"
            "  七.primary_reading = 'しち'\n"
            "  七.on, kun fields include both しち and なな per kanji.json structure\n"
            "  reading_rule starts: 'Generic rule: standalone kanji → typically kun-yomi; compound kanji → typically on-yomi. For 七: standalone use of kun (なな)...'\n\n"
            "しち is prescriptively the on-yomi primary. In modern colloquial Japanese, especially when counting things, **なな is more common**:\n"
            "  七人  → either reading\n"
            "  七時  → しちじ exclusively\n"
            "  七月  → しちがつ exclusively\n"
            "  七つ  → ななつ exclusively\n"
            "  Plain '7' read aloud → NHK uses なな (to disambiguate from いち over phone)\n\n"
            "Fix: append to reading_rule:\n"
            "  'なな is heard in ordering/listing contexts (especially to disambiguate from いち over the phone); しち in fixed time/date compounds (七時 しちじ, 七月 しちがつ). NHK uses なな when reading numerals aloud. 七人 takes either reading.'"
        ),
        "sev": "Minor", "pri": "P4",
    },
    {
        "title": "NTR-013 — pronoun entries carry counter={kanji: '人', reading: 'にん'} indiscriminately; みなさん has counter.reading='人' typo",
        "desc": (
            "FILE: data/vocab.json (pronoun cohort).\n\n"
            "Verified 2026-05-22 — 9 pronoun entries carry counter:\n"
            "  私:        counter={'kanji': '人', 'reading': 'にん'}\n"
            "  私たち:    counter={'kanji': '人', 'reading': 'にん'}\n"
            "  あなた:    counter={'kanji': '人', 'reading': 'にん'}\n"
            "  かれ:      counter={'kanji': '人', 'reading': 'にん'}\n"
            "  かのじょ:  counter={'kanji': '人', 'reading': 'にん'}\n"
            "  かた:      counter={'kanji': '人', 'reading': 'にん'}\n"
            "  人:        counter={'kanji': '人', 'reading': 'にん'}\n"
            "  みなさん:  counter={'kanji': '人', 'reading': '人'}   ← reading field has kanji not kana (typo)\n\n"
            "Two issues:\n"
            "1. For 'we' (私たち), counter '人/にん' is misleading because we don't count 'we's. The counter applies to the NOUN the pronoun refers to, not to the pronoun itself. For first-person plural / mass pronouns, either omit the counter field or document as 'the noun this pronoun refers to counts with 人/にん'.\n"
            "2. みなさん.counter.reading = '人' (kanji) is a typo; should be 'にん' (kana) like the others.\n\n"
            "Fix:\n"
            "  - Fix the みなさん typo: reading '人' → 'にん'.\n"
            "  - Decide policy on pronoun-counter applicability — either drop the counter from non-individual pronouns (私たち, みなさん) or annotate it as a noun-of-reference counter.\n\n"
            "Recommend: keep counter on individual-referring pronouns (私, あなた, かれ, かのじょ, かた, 人) since they count 'a person'; drop or annotate counter on collective pronouns (私たち, みなさん) since the count is of the people, not of the pronoun."
        ),
        "sev": "Minor", "pri": "P4",
    },
]

# Append starting at row 164 (next after BUG-160 at row 163)
start_row = ws.max_row + 1  # = 164
for i, b in enumerate(BUGS):
    r = start_row + i
    ws.cell(row=r, column=1, value='="BUG-"&TEXT(ROW()-3,"000")')
    ws.cell(row=r, column=2, value="2026-05-22")
    ws.cell(row=r, column=3, value="Native teacher / JLPT expert review (v1.15.8 build packet)")
    ws.cell(row=r, column=4, value=b["title"])
    ws.cell(row=r, column=5, value=b["desc"])
    ws.cell(row=r, column=6, value=b["sev"])
    ws.cell(row=r, column=7, value=b["pri"])
    ws.cell(row=r, column=8, value="Open")
    print(f"  Row {r}: BUG-{r-3:03d} = {b['title'][:80]}...")

wb.save(XLSX)
print(f"\nmax_row after: {ws.max_row}")
print(f"Filed {len(BUGS)} bugs (NTR-001..NTR-{len(BUGS):03d}) at rows {start_row}..{start_row + len(BUGS) - 1}")
